"""Auto-match 1-a-1 de un xlsx de Pichincha contra scintela.transacciones_bancarias.

Uso (en EC2):
    cd C:\\programa-core
    python scripts/auto_match_xlsx_one_off.py C:\\tmp\\extracto.xlsx [--no-banco 10] [--dry-run]

Para cada fila del extracto busca en transacciones_bancarias:
    no_banco = NO_BANCO
    AND fecha BETWEEN extr_fecha-3d AND extr_fecha+3d
    AND ABS(importe - extr_monto) < 0.005   (matching exacto)
    AND tipo PC compatible con tipo banco (C->DE/TR/XX/NC/IN/AC, D->CH/ND/DB/GS/PA)
    AND NOT EXISTS match activo

Si hay 1 candidato → crea banco_conciliacion_match.
Si hay >1 con fecha exacta → toma fecha exacta. Si sigue ambiguo → skip.
Si hay 0 → skip.

Imprime resumen al final.

TMT 2026-05-28 — dueña pidió "conecta uno con uno me da igual".
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

# Cargar .env del repo (DB_HOST/DB_NAME/DB_USER/DB_PASSWORD).
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError:
    pass

import psycopg2
from psycopg2.extras import RealDictCursor

import openpyxl

DOCS_CRED = ("DE", "TR", "XX", "NC", "IN", "AC")
DOCS_DEB = ("CH", "ND", "DB", "GS", "PA")


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Path al .xlsx del extracto Pichincha")
    ap.add_argument("--no-banco", type=int, default=10)
    ap.add_argument("--dias-tol", type=int, default=3)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--usuario", default="auto-match-one-off")
    return ap.parse_args()


def parse_xlsx(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    movs = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for ci, h in enumerate(headers, 1):
            row[h] = ws.cell(r, ci).value
        if not row.get("Fecha"):
            continue
        movs.append(row)
    return movs


def parse_fecha(s) -> "datetime.date|None":
    if not s:
        return None
    if hasattr(s, "date"):
        return s.date()
    if hasattr(s, "year"):
        return s
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def main() -> int:
    args = parse_args()
    movs = parse_xlsx(args.xlsx)
    print(f"Extracto: {len(movs)} movs ({args.xlsx})")
    if args.dry_run:
        print("[DRY-RUN] no se inserta en banco_conciliacion_match.")

    conn = psycopg2.connect(
        host=os.environ["DB_HOST"], dbname=os.environ["DB_NAME"],
        user=os.environ["DB_USER"], password=os.environ["DB_PASSWORD"],
        port=os.environ.get("DB_PORT", "5432"),
    )
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=RealDictCursor)

    NB = int(args.no_banco)
    TOL_DIAS = int(args.dias_tol)

    n_matched = 0
    n_no_match = 0
    n_multi = 0
    n_dup = 0
    log_no: list[tuple] = []
    log_multi: list[tuple] = []

    for m in movs:
        f_b = parse_fecha(m.get("Fecha"))
        if not f_b:
            continue
        try:
            monto = float(m.get("Monto") or 0)
        except (TypeError, ValueError):
            continue
        tipo = str(m.get("Tipo") or "C").upper()[:1]
        doc_b = str(m.get("Documento") or "").strip()[:40]
        concepto = str(m.get("Concepto") or "")[:500]
        oficina = str(m.get("Oficina") or "")[:50]
        codigo = str(m.get("Codigo") or "")[:10]

        # Dedup: si ya hay match activo para este real, skip.
        cur.execute(
            "SELECT 1 FROM scintela.banco_conciliacion_match "
            "WHERE no_banco=%s AND real_fecha=%s AND real_documento=%s "
            "  AND real_monto=%s AND real_tipo=%s AND deshecho_en IS NULL LIMIT 1",
            (NB, f_b, doc_b, monto, tipo),
        )
        if cur.fetchone():
            n_dup += 1
            continue

        docs_compat = DOCS_CRED if tipo == "C" else DOCS_DEB
        d1 = f_b - timedelta(days=TOL_DIAS)
        d2 = f_b + timedelta(days=TOL_DIAS)

        cur.execute(
            """
            SELECT t.id_transaccion, t.fecha, t.documento, t.importe
              FROM scintela.transacciones_bancarias t
             WHERE t.no_banco = %s
               AND t.fecha BETWEEN %s AND %s
               AND ABS(t.importe - %s) < 0.005
               AND UPPER(TRIM(t.documento)) IN %s
               AND TRIM(COALESCE(t.stat,'')) <> '*'
               AND NOT EXISTS (
                   SELECT 1 FROM scintela.banco_conciliacion_match mc
                    WHERE mc.id_transaccion = t.id_transaccion AND mc.deshecho_en IS NULL
               )
             ORDER BY ABS(t.fecha - %s)
             LIMIT 5
            """,
            (NB, d1, d2, monto, docs_compat, f_b),
        )
        cands = cur.fetchall()

        if not cands:
            n_no_match += 1
            log_no.append((str(f_b), tipo, monto, doc_b, concepto[:30]))
            continue
        if len(cands) > 1:
            same_day = [c for c in cands if c["fecha"] == f_b]
            if len(same_day) == 1:
                cands = same_day
            else:
                n_multi += 1
                log_multi.append((str(f_b), tipo, monto, [c["id_transaccion"] for c in cands]))
                continue

        pc = cands[0]
        if not args.dry_run:
            cur.execute(
                """
                INSERT INTO scintela.banco_conciliacion_match
                    (no_banco, estado, real_fecha, real_concepto, real_documento,
                     real_monto, real_tipo, real_codigo, real_oficina, id_transaccion, usuario)
                VALUES (%s, 'matched', %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT DO NOTHING
                RETURNING id
                """,
                (NB, f_b, concepto, doc_b, monto, tipo, codigo, oficina, pc["id_transaccion"], args.usuario),
            )
            if cur.fetchone():
                n_matched += 1
            else:
                n_dup += 1
        else:
            n_matched += 1  # contamos como matched en dry-run

    if not args.dry_run:
        conn.commit()
    else:
        conn.rollback()

    print(f"\n=== RESUMEN ===")
    print(f"  conciliados:      {n_matched}")
    print(f"  ya conciliados:   {n_dup}")
    print(f"  multi-match skip: {n_multi}")
    print(f"  sin match:        {n_no_match}")
    print(f"  TOTAL extracto:   {len(movs)}")
    if log_no:
        print(f"\n=== sin match (primeros 20 de {len(log_no)}) ===")
        for f_b, t, mt, doc, c in log_no[:20]:
            print(f"  {f_b} {t} ${mt:>10,.2f} doc={doc:<10s} {c}")
    if log_multi:
        print(f"\n=== multi-match (primeros 10 de {len(log_multi)}) ===")
        for f_b, t, mt, ids in log_multi[:10]:
            print(f"  {f_b} {t} ${mt:>10,.2f} candidatos: {ids}")
    conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
