"""Alinea scintela.transacciones_bancarias.stat con un xlsx exportado de PICHINCH.DBF.

Uso (en EC2):
    python scripts/sync_stat_from_xlsx.py \
        --xlsx C:\\programa-core\\data\\dbase_snapshots\\PICHINCH.xlsx \
        --no-banco 10

Lee el .env (dot env) del CWD para conectarse a la DB intela (scintela.*).
NO el DATABASE_URL Machine (eso apunta a formulas_app). Ver
feedback_ec2_db_url_pc_vs_machine.md.

TMT 2026-05-28 dueña: 'make sure these are the only conciliated movements
please. the *'. Idempotente.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from dotenv import load_dotenv

# Carga del .env contiguo (C:\programa-core\.env en prod)
load_dotenv(Path(__file__).resolve().parent.parent / ".env")

import db  # noqa: E402  — usa la conexión de la app
from openpyxl import load_workbook  # noqa: E402


def _key(fecha, doc, importe, saldo):
    """Clave natural de un movimiento bancario en el DBF/PG."""
    f_iso = fecha.date().isoformat() if hasattr(fecha, "date") else str(fecha)
    d = (doc or "").strip().upper() if isinstance(doc, str) else ""
    try:
        imp = round(float(importe or 0), 2)
    except (TypeError, ValueError):
        imp = 0.0
    try:
        sal = round(float(saldo or 0), 2)
    except (TypeError, ValueError):
        sal = 0.0
    return (f_iso, d, imp, sal)


def _parse_xlsx(path: Path) -> dict[tuple, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("xlsx vacío")
    header = [
        (c or "").strip().upper() if isinstance(c, str) else c for c in rows[0]
    ]
    needed = ["FECHA", "DOC", "IMPORTE", "SALDO", "STAT"]
    ix = {k: header.index(k) for k in needed}
    out: dict[tuple, str] = {}
    for r in rows[1:]:
        try:
            k = _key(r[ix["FECHA"]], r[ix["DOC"]], r[ix["IMPORTE"]], r[ix["SALDO"]])
            raw = r[ix["STAT"]]
            s = raw.strip() if isinstance(raw, str) else ("" if raw is None else str(raw).strip())
            out[k] = "*" if s == "*" else ""
        except Exception:
            continue
    return out


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, type=Path)
    ap.add_argument("--no-banco", type=int, required=True)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(argv)

    print(f"[1/4] leyendo {args.xlsx}")
    xlsx_stat = _parse_xlsx(args.xlsx)
    n_star = sum(1 for v in xlsx_stat.values() if v == "*")
    print(f"      filas únicas: {len(xlsx_stat)} · con stat='*': {n_star}")

    print(f"[2/4] traigo PG no_banco={args.no_banco}")
    pg_rows = db.fetch_all(
        "SELECT id_transaccion, fecha, documento, importe, saldo, "
        "       TRIM(COALESCE(stat,'')) AS stat "
        "  FROM scintela.transacciones_bancarias WHERE no_banco = %s",
        (args.no_banco,),
    ) or []
    print(f"      PG tiene {len(pg_rows)} filas para no_banco={args.no_banco}")

    to_star: list[int] = []
    to_null: list[int] = []
    no_match: list[int] = []
    ok = 0
    for r in pg_rows:
        k = _key(r["fecha"], r["documento"], r["importe"], r["saldo"])
        pg = (r["stat"] or "").strip()
        if k not in xlsx_stat:
            no_match.append(r["id_transaccion"])
            continue
        xl = xlsx_stat[k]
        if xl == "*" and pg != "*":
            to_star.append(r["id_transaccion"])
        elif xl != "*" and pg == "*":
            to_null.append(r["id_transaccion"])
        else:
            ok += 1

    print(
        f"[3/4] plan: marcar '*' en {len(to_star)} · limpiar '*' en "
        f"{len(to_null)} · OK sin cambios: {ok} · sin match en xlsx: {len(no_match)}"
    )

    if args.dry_run:
        print("dry-run: no toco la DB.")
        return 0

    def _chunks(seq, n=500):
        for i in range(0, len(seq), n):
            yield seq[i:i + n]

    n_up_star = 0
    for chunk in _chunks(to_star):
        n_up_star += db.execute(
            "UPDATE scintela.transacciones_bancarias SET stat = '*' "
            "WHERE id_transaccion = ANY(%s)",
            (chunk,),
        ) or 0
    n_up_null = 0
    for chunk in _chunks(to_null):
        n_up_null += db.execute(
            "UPDATE scintela.transacciones_bancarias SET stat = NULL "
            "WHERE id_transaccion = ANY(%s)",
            (chunk,),
        ) or 0

    print(f"[4/4] UPDATEd → '*': {n_up_star} · → NULL: {n_up_null}")
    if no_match:
        print(
            f"      OJO: {len(no_match)} fila(s) PC sin contraparte en xlsx "
            f"(creadas en PC y/o no en DBF). Primeros 10: {no_match[:10]}"
        )
    print("OK ✓")
    return 0


if __name__ == "__main__":
    sys.exit(main())
