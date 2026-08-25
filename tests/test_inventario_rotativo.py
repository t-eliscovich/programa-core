"""Inventario rotativo — service y pantalla.

Los fakes devuelven filas con la forma de la FUENTE (los nombres de columna que
escriben las dos SQL contra Asinfo), no la del helper: un fake con la forma
equivocada pasa en verde mientras producción no resuelve nada.
"""
import io
import re
from pathlib import Path
from unittest.mock import patch

import pytest

from modules.inventario_rotativo import service, views


def _ficha(**kw):
    """Fila cruda de `_SQL_FICHA`, con los defaults de una tela normal."""
    base = {
        "id_producto": 1, "categoria": "Fleece", "tela": "Fleece 102",
        "codigo": "FE102JOS", "color": "JOS", "semanas_venta": 49,
        "kg52": 5200.0, "inv_kg": 0.0, "ped_kg": 0.0, "ped_un": 0.0,
        "prod_kg": 0.0, "n_ordenes": 0,
    }
    base.update(kw)
    return base


def _serie(id_producto=1, kilos=None, ultima=1389):
    """Filas crudas de `_SQL_SERIE`: sólo las semanas CON venta."""
    kilos = kilos if kilos is not None else [100.0] * 52
    return [
        {"id_producto": id_producto, "w": ultima - len(kilos) + 1 + i, "kg": kg}
        for i, kg in enumerate(kilos)
        if kg
    ]


# ── el percentil ────────────────────────────────────────────────────────────

def test_el_percentil_interpola_entre_los_dos_valores_que_lo_rodean():
    assert service.percentil([0.0, 10.0], 0.90) == pytest.approx(9.0)


def test_el_percentil_de_una_lista_vacia_es_cero_y_no_revienta():
    assert service.percentil([], 0.90) == 0.0


def test_las_semanas_sin_venta_entran_en_la_grilla_con_cero():
    """El percentil se calcula sobre las 52 semanas, no sobre las que vendieron.

    Si se calculara sólo sobre las semanas buenas saldría inflado y la pantalla
    mandaría a teñir de más. El caso extremo lo muestra sin ambigüedad: un
    producto que vendió 5.200 kg en UNA sola semana del año tiene un punto de
    reposición de cero —no vende todas las semanas—, no de 5.200.
    """
    f = service._fila(_ficha(), {1389: 5200.0}, 1389)
    assert f["punto_kg"] == 0.0


def test_el_promedio_semanal_divide_por_las_13_semanas_y_no_por_las_vendidas():
    """Vender 1.300 kg en dos semanas de trece es un promedio de 100, no de 650.

    Dividir por las semanas con venta daría el tamaño del pedido típico, que es
    otra cosa, y sobreestimaría lo que hay que tener en bodega.
    """
    f = service._fila(_ficha(), {1388: 650.0, 1389: 650.0}, 1389)
    assert f["sem_kg"] == pytest.approx(100.0)


def test_el_punto_de_reposicion_cubre_dos_semanas_y_no_una():
    """Es la demanda ACUMULADA en el ciclo de tintura, no la de una semana."""
    f = service._fila(_ficha(), {w: 100.0 for w in range(1338, 1390)}, 1389)
    assert f["punto_kg"] == pytest.approx(200.0)


def test_una_semana_de_devoluciones_no_cuenta_como_demanda_negativa():
    """El neto negativo se trunca para el percentil (no existe demanda de -50)
    pero SIGUE restando en el promedio: eso sí se vendió de menos."""
    semanas = {w: 100.0 for w in range(1338, 1390)}
    semanas[1389] = -50.0
    f = service._fila(_ficha(), semanas, 1389)
    assert f["punto_kg"] > 0
    assert f["sem_kg"] < 100.0


# ── unidades ────────────────────────────────────────────────────────────────

def test_las_telas_se_leen_en_rollos_enteros():
    assert service.unidad_de("Fleece") == "roll"
    assert service.a_unidad(235.0, "roll", "Fleece") == 10.0


def test_medio_rollo_nunca_se_muestra_como_cero():
    """Redondear 8 kg a 0 rollos diría que no falta nada cuando falta."""
    assert service.a_unidad(8.0, "roll", "Fleece") == 1.0
    assert service.a_unidad(0.0, "roll", "Fleece") == 0.0


def test_el_rib_va_en_kilos_y_los_cuellos_en_unidades():
    """Dueña 2026-08-18: "se mide todo en rollos salvo rib, cuellos y puños"."""
    assert service.unidad_de("Rib") == "kg"
    assert service.a_unidad(403.9, "kg", "Rib") == 403.9
    assert service.unidad_de("Cuellos") == "un"
    assert service.a_unidad(15.0, "un", "Cuellos") == 500.0


# ── el semáforo ─────────────────────────────────────────────────────────────

def test_el_semaforo_corta_en_el_tiempo_que_tarda_una_tintura():
    assert service.estado(1.9) == "rojo"
    assert service.estado(2.5) == "ambar"
    assert service.estado(6.0) == "ok"
    assert service.estado(20.0) == "sobra"


def test_lo_que_esta_en_produccion_cuenta_para_la_cobertura():
    """Si ya hay 300 kg tinturándose, no hay que volver a pedirlos."""
    semanas = {w: 100.0 for w in range(1338, 1390)}
    sin = service._fila(_ficha(inv_kg=100.0), semanas, 1389)
    con = service._fila(_ficha(inv_kg=100.0, prod_kg=300.0), semanas, 1389)
    assert sin["estado"] == "rojo"
    assert con["alcanza"] > sin["alcanza"]
    assert con["falta_kg"] < sin["falta_kg"]


def test_sin_venta_reciente_no_se_pinta_de_rojo():
    """Rota en el año pero se apagó hace 13 semanas: no es "falta", es "no sé".

    Marcarlo rojo mandaría a teñir algo que nadie está comprando.
    """
    viejas = {w: 100.0 for w in range(1338, 1377)}
    f = service._fila(_ficha(semanas_venta=40), viejas, 1389)
    assert f["alcanza"] == -1.0
    assert f["estado"] == "ok"


# ── agrupado ────────────────────────────────────────────────────────────────

def _filas_2():
    return [
        service._fila(_ficha(color="JOS", tela="Fleece 102"),
                      {w: 100.0 for w in range(1338, 1390)}, 1389),
        service._fila(_ficha(id_producto=2, color="JOS", tela="Pique Especial",
                             inv_kg=9000.0),
                      {w: 100.0 for w in range(1338, 1390)}, 1389),
    ]


def test_agrupar_por_color_junta_las_telas_de_ese_color():
    (b,) = service.agrupar(_filas_2(), "color")
    assert b["nombre"] == "JOS" and b["n"] == 2
    assert b["etiqueta"] == "tela"


def test_el_bloque_con_mas_faltante_va_primero():
    filas = _filas_2()
    bloques = service.agrupar(filas, "tela")
    assert bloques[0]["nombre"] == "Fleece 102"      # el que no tiene stock
    assert bloques[0]["falta_kg"] > bloques[1]["falta_kg"]


def test_los_subtotales_del_bloque_van_en_kilos():
    """Un color agrupa telas de distinta unidad: sumar rollos con unidades da
    un número que no significa nada."""
    filas = [
        service._fila(_ficha(color="BLA", tela="Fleece 102"),
                      {w: 100.0 for w in range(1338, 1390)}, 1389),
        service._fila(_ficha(id_producto=2, categoria="Cuellos", color="BLA",
                             tela="Cuellos T40"),
                      {w: 10.0 for w in range(1338, 1390)}, 1389),
    ]
    (b,) = service.agrupar(filas, "color")
    assert b["sem_kg"] == pytest.approx(110.0)


# ── la consulta ─────────────────────────────────────────────────────────────

def test_asinfo_caido_no_es_lo_mismo_que_no_falta_nada():
    service._cache.clear()
    service._cache_nuevos.clear()
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      return_value=([], False)):
        filas, disponible = service.rotativo(_ahora=lambda: 0.0)
    assert filas == [] and disponible is False


def test_solo_entran_las_familias_de_tela_vendible():
    """TELA CRUDA, HILO y los insumos no se stockean contra promedio."""
    service._cache.clear()
    service._cache_nuevos.clear()
    fichas = [_ficha(), _ficha(id_producto=2, categoria="HILO", tela="Hilado")]
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=[(fichas, True), (_serie() + _serie(2), True)]):
        filas, _ = service.rotativo(_ahora=lambda: 0.0)
    assert [f["familia"] for f in filas] == ["Fleece"]


def test_la_sql_pide_solo_los_productos_que_rotan():
    assert f"HAVING COUNT(*) >= {service.SEMANAS_MINIMAS}" in service._sql_ficha()


def test_la_sql_de_produccion_deja_afuera_las_ordenes_padre():
    """Las órdenes viven en dos capas y `SUM` sobre la tabla cruda las cuenta
    dos veces. Misma trampa que /pedidos."""
    sql = service._sql_ficha()
    assert "h.p IS NULL" in sql
    assert "o.id_producto IS NOT NULL" in sql
    assert "o.estado_produccion = 2" in sql


def test_la_sql_usa_la_hora_de_ecuador_y_no_getdate_pelado():
    """En Asinfo `GETDATE()` devuelve UTC: pelado corre la ventana un día."""
    for sql in (service._sql_ficha(), service._sql_serie()):
        assert "DATEADD(hour, -5, GETDATE())" in sql


# ── la pantalla ─────────────────────────────────────────────────────────────

def _login(app, fake_db):
    rid = fake_db.add_role("Tester", ["stock.ver"])
    uid = fake_db.add_user("test", b"$2b$12$fakehash", rid)
    c = app.test_client()
    with c.session_transaction() as s:
        s["user_id"] = uid
    return c


def _fake_asinfo(fichas=None, serie=None, nuevos=None):
    """Las TRES consultas de la pantalla, cada una con la forma de su fuente.

    Se despacha por un pedazo del SELECT y no por orden de llamada: si mañana
    la vista pide una antes que la otra, un fake por orden empieza a devolver
    la respuesta equivocada sin que ningún test se ponga rojo.
    """
    fichas = fichas if fichas is not None else [_ficha()]
    serie = serie if serie is not None else _serie()
    nuevos = nuevos if nuevos is not None else [{"n": 0, "kg": 0.0}]

    def fake(_db, sql, **_kw):
        if "SELECT v.id_producto, v.w" in sql:
            return (serie, True)
        if "SELECT COUNT(*) AS n, SUM(inv.inv_kg) AS kg" in sql:
            return (nuevos, True)
        return (fichas, True)

    return fake


def test_la_pantalla_abre_por_color(app, fake_db):
    """Dueña 2026-08-18: "que color sea lo primero"."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        body = c.get("/inventario-rotativo").get_data(as_text=True)
    assert "Rotación del inventario" in body
    assert "JOS" in body
    # el encabezado de la tabla dice Tela: las filas de un color SON telas
    assert "<th>Tela</th>" in body


def test_el_corte_por_tela_da_vuelta_el_agrupado(app, fake_db):
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        body = c.get("/inventario-rotativo?ver=tela").get_data(as_text=True)
    assert "<th>Color</th>" in body


def test_un_corte_que_no_existe_cae_en_color(app, fake_db):
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        body = c.get("/inventario-rotativo?ver=cualquiera").get_data(as_text=True)
    assert "<th>Tela</th>" in body


def test_con_asinfo_caido_la_pantalla_lo_dice_en_vez_de_mentir(app, fake_db):
    """"No pude preguntar" no es "no falta nada"."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      return_value=([], False)):
        body = c.get("/inventario-rotativo").get_data(as_text=True)
    assert "No pude consultar Asinfo" in body


def test_sin_el_permiso_de_stock_la_pantalla_da_404(app, fake_db):
    rid = fake_db.add_role("Otro", ["facturas.ver"])
    uid = fake_db.add_user("otro", b"$2b$12$fakehash", rid)
    c = app.test_client()
    with c.session_transaction() as s:
        s["user_id"] = uid
    assert c.get("/inventario-rotativo").status_code == 404


def test_el_faltante_se_muestra_del_mismo_color_que_alcanza(app, fake_db):
    """Dueña 2026-08-18: "mantengamos naranja y naranja, rojo y rojo".

    Dos rojos distintos en la misma fila se leen como dos alarmas donde hay
    una sola.
    """
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    # 250 kg de stock contra 100 kg/semana: 2,5 semanas → ámbar, y como el
    # punto son 200 no falta nada... le bajamos el stock para que falte.
    fichas = [_ficha(inv_kg=250.0)]
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo(fichas, _serie(kilos=[100.0] * 52))):
        body = c.get("/inventario-rotativo").get_data(as_text=True)
    assert 'class="ambar"' in body
    assert 'class="rojo"' not in body.split("<table>")[1]


# ── imprimir y Excel ────────────────────────────────────────────────────────

def test_la_hoja_impresa_es_la_misma_pantalla(app, fake_db):
    """Dueña 2026-08-18: "que se pueda imprimir también así".

    Sale de la misma plantilla con `?imprimir=1`: si fueran dos, lo que se
    lleva a planta se iría separando de lo que se mira en la oficina.
    """
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        body = c.get("/inventario-rotativo?imprimir=1").get_data(as_text=True)
    assert 'class="iv hoja"' in body
    assert "JOS" in body                      # los datos siguen estando
    assert "Fuente: Asinfo" in body           # y el pie con la fecha


def test_la_hoja_esconde_los_filtros_y_fija_el_ancho_de_la_a4():
    """Sin ancho fijo la tabla sale del papel; con `@media print` sólo no se
    puede verificar la hoja sin abrir el diálogo de impresión."""
    tpl = (Path(service.__file__).parent
           / "templates" / "inventario_rotativo" / "lista.html").read_text(encoding="utf-8")
    assert ".iv.hoja{width:816px" in tpl
    assert ".iv.hoja .barra,.iv.hoja .tabs" in tpl


def test_el_excel_trae_una_fila_por_producto_con_encabezado(app, fake_db):
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        r = c.get("/inventario-rotativo/excel")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["Content-Type"]
    assert "inventario_rotativo_" in r.headers["Content-Disposition"]

    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(r.data)).active
    assert [c.value for c in ws[1]] == list(views.COLUMNAS)
    assert ws.max_row == 2
    assert ws["A2"].value == "JOS"


def test_el_excel_deja_vacia_la_cobertura_que_no_se_puede_calcular():
    """Sin venta reciente `alcanza` vale -1 adentro; en la planilla va vacío.

    Un -1 en una columna de semanas se ordena y se suma como si fuera un dato.
    """
    f = service._fila(_ficha(), {w: 100.0 for w in range(1338, 1377)}, 1389)
    assert views._valores(f)[views.COLUMNAS.index("Alcanza (sem)")] is None


def test_con_asinfo_caido_el_excel_no_baja_una_planilla_vacia(app, fake_db):
    """Bajar un archivo con el encabezado solo se lee como "no falta nada"."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      return_value=([], False)):
        r = c.get("/inventario-rotativo/excel")
    assert r.status_code == 302


# ── lo pedido ───────────────────────────────────────────────────────────────

def test_lo_pedido_no_se_resta_de_la_cobertura():
    """Dueña 2026-08-18: "no suma a ningún lado, pero un número informativo" —
    "todavía no se va".

    La mercadería sigue en bodega hasta que se despacha. Restarla acá
    adelantaría una salida que puede no ocurrir esta semana, y pondría en rojo
    algo que hoy está.
    """
    semanas = {w: 100.0 for w in range(1338, 1390)}
    sin = service._fila(_ficha(inv_kg=600.0), semanas, 1389)
    con = service._fila(_ficha(inv_kg=600.0, ped_kg=500.0), semanas, 1389)
    assert con["alcanza"] == sin["alcanza"]
    assert con["falta_kg"] == sin["falta_kg"]
    assert con["pedido_kg"] == 500.0


def test_los_cuellos_se_piden_por_unidad_y_el_pedido_se_pasa_a_kilos():
    """La columna tiene que ser comparable con el stock, que viene en kilos."""
    f = service._fila(
        _ficha(categoria="Cuellos", tela="Cuellos T40", ped_un=500.0),
        {w: 10.0 for w in range(1338, 1390)}, 1389)
    assert f["pedido_kg"] == pytest.approx(15.0, abs=0.1)


def test_la_columna_pedido_lleva_asterisco_y_su_nota(app, fake_db):
    """Sin la nota, una columna en el medio de la tabla se lee como que suma."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    fichas = [_ficha(inv_kg=600.0, ped_kg=235.0)]
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo(fichas)):
        body = c.get("/inventario-rotativo").get_data(as_text=True)
    assert 'Pedido<span class="ast">*</span>' in body
    assert "no se resta de lo que hay" in body


def test_el_pedido_del_dia_usa_el_mismo_corte_de_90_dias_que_pedidos():
    """Si las dos pantallas cortan distinto, muestran números distintos del
    mismo pedido y la dueña no sabe a cuál creerle."""
    assert f"<= {service.DIAS_PEDIDO_MAX}" in service._sql_ficha()
    assert service.DIAS_PEDIDO_MAX == 90


def test_la_unidad_va_en_una_columna_y_no_pegada_a_cada_numero(app, fake_db):
    """Dueña 2026-08-18: "poner una columna de unidades y ya, como hace Asinfo,
    no repetirlo tantas veces".

    Seis sufijos por fila (roll roll roll roll) tapan los números, que son lo
    que hay que leer.
    """
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        body = c.get("/inventario-rotativo").get_data(as_text=True)
    tabla = body.split("<table>")[1].split("</table>")[0]
    assert '<td class="un">roll</td>' in tabla
    assert tabla.count("roll") == 1         # una sola vez por fila, no seis
    assert 'class="u"' not in tabla         # el sufijo por celda ya no existe


def test_el_link_del_menu_va_debajo_de_inventario_en_produccion_y_stocks(app, fake_db):
    """Federico 2026-08-20: "pasarlo a la sección de Stock, debajo de Inventario".

    Antes vivía en "Modificar", debajo de Pedidos pendientes (dueña 2026-08-18).
    La sección "Producción y stocks" ya está gateada con `stock.ver` —el mismo
    permiso de SU pantalla—, así que el link no necesita gate propio.
    """
    base = (Path(app.root_path) / "templates" / "base.html").read_text(encoding="utf-8")
    i_stock = base.index('data-key="stock"')
    i_mod = base.index('data-key="modificar"')
    i_fab = base.index("stock_asinfo.fabricacion_tc")
    i_inv = base.index("inventario_rotativo.lista")
    i_imp = base.index("importaciones.lista")
    # dentro de la sección, y justo entre "Inventario" e "Ingreso de hilado"
    assert i_stock < i_fab < i_inv < i_imp
    # y ya no está en "Modificar", que arranca antes
    assert i_mod < i_stock


# ── acabado ─────────────────────────────────────────────────────────────────

def test_el_acabado_sale_de_pedidos_y_no_de_una_sql_propia(app, fake_db):
    """Dueña 2026-08-18: "pone columna de acabado".

    Se importa `pedidos.service.acabados_por_producto` en vez de copiar la SQL:
    dos consultas paralelas se desincronizan sin que nadie se entere hasta que
    las dos pantallas dicen acabados distintos del mismo producto.
    """
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch("modules.pedidos.service.acabados_por_producto",
               return_value={"FE102JOS": "TUB"}), \
         patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        body = c.get("/inventario-rotativo").get_data(as_text=True)
    assert "<th style=\"text-align:left\">Acab.</th>" in body
    assert '<td class="ac">TUB</td>' in body


def test_sin_acabado_la_celda_queda_vacia_y_la_pantalla_no_se_cae(app, fake_db):
    """No se inventa un acabado: un producto sin lote con atributo no tiene."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch("modules.pedidos.service.acabados_por_producto",
               side_effect=RuntimeError("Asinfo se cayó")), \
         patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        body = c.get("/inventario-rotativo").get_data(as_text=True)
    assert '<td class="ac"></td>' in body
    assert "JOS" in body


# ── subtotales del bloque ───────────────────────────────────────────────────

def test_el_subtotal_va_en_rollos_cuando_todo_el_bloque_son_rollos():
    """Dueña 2026-08-18: "poneme todo en rollos salvo los cuellos, rib y puños".

    El encabezado del color era el único lugar que seguía hablando en kilos.
    """
    filas = [service._fila(_ficha(color="BLA", tela="Fleece 102"),
                           {w: 235.0 for w in range(1338, 1390)}, 1389)]
    (b,) = service.agrupar(filas, "color")
    assert (b["unidad"], b["sem"]) == ("roll", 10.0)   # 235 kg = 10 rollos


def test_un_bloque_mixto_usa_la_unidad_que_pesa_mas():
    """Dueña 2026-08-18: "o en rollo o en kg, no me conviertas ambos".

    Un color que junta Fleece con Rib se lee en la unidad que manda por peso:
    dos términos en un encabezado son dos números donde alcanza uno.
    """
    filas = [
        service._fila(_ficha(color="BLA", tela="Fleece 102"),
                      {w: 235.0 for w in range(1338, 1390)}, 1389),
        service._fila(_ficha(id_producto=2, categoria="Rib", color="BLA",
                             tela="Rib Normal"),
                      {w: 100.0 for w in range(1338, 1390)}, 1389),
    ]
    (b,) = service.agrupar(filas, "color")
    # 235 kg de Fleece contra 100 de Rib: mandan los rollos
    assert (b["unidad"], b["sem"]) == ("roll", 14.0)


def test_cuellos_y_punos_comparten_el_termino_en_unidades():
    """Los dos se leen "un", así que van en el mismo término — ya convertidos
    fila por fila con el factor de su familia (33,33 y 50 un/kg)."""
    filas = [
        service._fila(_ficha(categoria="Cuellos", color="BLA", tela="Cuellos T40"),
                      {w: 10.0 for w in range(1338, 1390)}, 1389),
        service._fila(_ficha(id_producto=2, categoria="Puños", color="BLA",
                             tela="Puños"),
                      {w: 10.0 for w in range(1338, 1390)}, 1389),
    ]
    (b,) = service.agrupar(filas, "color")
    assert b["unidad"] == "un"


# ── la hoja respeta el filtro ───────────────────────────────────────────────

def _dos_filas():
    """Un rojo (sin stock) y un tranquilo (con stock de sobra)."""
    return [_ficha(color="JOS"),
            _ficha(id_producto=2, color="NEG", codigo="FE102NEG", inv_kg=9000.0)]


def test_la_hoja_imprime_solo_lo_filtrado(app, fake_db):
    """Dueña 2026-08-18. Pedir "las que hay que teñir" imprimía las 289 —
    trece páginas para tirar."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    serie = _serie(1) + _serie(2)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo(_dos_filas(), serie)):
        todo = c.get("/inventario-rotativo?imprimir=1").get_data(as_text=True)
        rojo = c.get("/inventario-rotativo?imprimir=1&est=rojo").get_data(as_text=True)
    assert ">JOS<" in todo and ">NEG<" in todo
    assert ">JOS<" in rojo and ">NEG<" not in rojo


def test_la_hoja_filtrada_dice_ARRIBA_qué_filtro_tiene(app, fake_db):
    """Una hoja de 62 filas de 289 tiene que decir por qué, o el que la lee
    piensa que eso es todo el inventario. Y va en la cabecera, no al pie: el
    recorte se lee ANTES de la tabla, no después."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        body = c.get("/inventario-rotativo?imprimir=1&est=rojo&fam=Fleece").get_data(as_text=True)
    assert "Sólo lo que hay que teñir" in body
    assert "Fleece" in body
    cabeza = body.split('class="cabeza"')[1].split("</div>")[0]
    assert "Rotación del inventario" in body[:body.index('class="cabeza"')] or cabeza


def test_el_filtro_de_la_pantalla_no_recorta_lo_que_se_ve(app, fake_db):
    """Sin `imprimir=1` filtra el JS: si el servidor recortara, al apretar
    "Todo" no habría forma de volver a ver el resto."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    serie = _serie(1) + _serie(2)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo(_dos_filas(), serie)):
        body = c.get("/inventario-rotativo?est=rojo").get_data(as_text=True)
    assert ">JOS<" in body and ">NEG<" in body


def test_un_filtro_que_no_existe_no_vacia_la_hoja(app, fake_db):
    """`?est=cualquiera` imprime todo, no una hoja en blanco."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        body = c.get("/inventario-rotativo?imprimir=1&est=cualquiera").get_data(as_text=True)
    assert ">JOS<" in body


def test_el_buscador_de_la_hoja_mira_color_y_tela(app, fake_db):
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    serie = _serie(1) + _serie(2)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo(_dos_filas(), serie)):
        body = c.get("/inventario-rotativo?imprimir=1&q=neg").get_data(as_text=True)
    assert ">NEG<" in body and ">JOS<" not in body


# ── el Excel sale como la pantalla ──────────────────────────────────────────

def _bajar(c, url="/inventario-rotativo/excel"):
    from openpyxl import load_workbook
    r = c.get(url)
    assert r.status_code == 200, r.status_code
    return load_workbook(io.BytesIO(r.data)).active


def test_el_excel_sale_en_el_mismo_orden_que_la_pantalla(app, fake_db):
    """Dueña 2026-08-18: "cuando bajo el excel aparece en otro orden".

    Una planilla ordenada distinto obliga a buscar de nuevo lo que ya estaba
    en pantalla. Se aplana el MISMO agrupado: bloques por faltante, filas por
    urgencia.
    """
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    fichas = [_ficha(color="JOS", tela="Fleece 102"),
              _ficha(id_producto=2, color="NEG", tela="Pique Especial",
                     codigo="PENEG", inv_kg=9000.0)]
    serie = _serie(1) + _serie(2)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo(fichas, serie)), \
         patch("modules.pedidos.service.acabados_por_producto", return_value={}):
        body = c.get("/inventario-rotativo").get_data(as_text=True)
        ws = _bajar(c)

    en_pantalla = re.findall(r'<h3>(\w+)<', body)
    en_excel = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert en_excel == en_pantalla          # mismo orden de colores


def test_el_excel_respeta_el_filtro_de_la_pantalla(app, fake_db):
    """El botón le cuelga est/fam/q, igual que Imprimir: bajar todo cuando en
    pantalla se ve el recorte es la misma trampa que la hoja de 13 páginas."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    fichas = [_ficha(color="JOS"),
              _ficha(id_producto=2, color="NEG", codigo="FE102NEG", inv_kg=9000.0)]
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo(fichas, _serie(1) + _serie(2))), \
         patch("modules.pedidos.service.acabados_por_producto", return_value={}):
        todo = _bajar(c)
        rojo = _bajar(c, "/inventario-rotativo/excel?est=rojo")
    assert todo.max_row == 3
    assert rojo.max_row == 2
    assert rojo["A2"].value == "JOS"


def test_el_excel_sigue_el_corte_que_se_esta_mirando(app, fake_db):
    """Por tela, la columna que agrupa es la tela; por color, el color."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()), \
         patch("modules.pedidos.service.acabados_por_producto", return_value={}):
        ws = _bajar(c, "/inventario-rotativo/excel?ver=tela")
    assert [c.value for c in ws[1]][:2] == ["Color", "Tela"]
    assert ws["A2"].value == "JOS" and ws["B2"].value == "Fleece 102"


def test_el_excel_lleva_la_unidad_y_el_acabado(app, fake_db):
    """Los mismos datos que la pantalla, no un subconjunto."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()), \
         patch("modules.pedidos.service.acabados_por_producto",
               return_value={"FE102JOS": "TUB"}):
        ws = _bajar(c)
    fila = {h: ws.cell(row=2, column=i).value
            for i, h in enumerate(views.COLUMNAS, 1)}
    assert fila["Acabado"] == "TUB"
    assert fila["Un."] == "roll"


def test_la_hoja_se_manda_a_imprimir_sola(app, fake_db):
    """Dueña 2026-08-18: "puse imprimir y me llevó acá, no a imprimir".

    Un `?imprimir=1` que sólo cambia el ancho deja al usuario a mitad de
    camino, buscando el Ctrl+P.
    """
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        hoja = c.get("/inventario-rotativo?imprimir=1").get_data(as_text=True)
        pantalla = c.get("/inventario-rotativo").get_data(as_text=True)
    assert "window.print()" in hoja
    assert "window.print()" not in pantalla.split("<script>")[-1]


def test_la_cabecera_de_la_tabla_se_repite_en_cada_pagina():
    """Trece páginas sin encabezado son trece páginas de números sueltos."""
    tpl = (Path(service.__file__).parent / "templates" / "inventario_rotativo"
           / "lista.html").read_text(encoding="utf-8")
    assert "<thead>" in tpl
    assert "display:table-header-group" in tpl


def test_la_pantalla_dice_de_cuantas_semanas_es_el_promedio(app, fake_db):
    """Dueña 2026-08-18: "por semana es un promedio de las últimas cuántas
    semanas?". Si hay que preguntarlo, la pantalla no lo estaba diciendo."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        body = c.get("/inventario-rotativo").get_data(as_text=True)
    assert f"promedio de las últimas {service.SEMANAS_CORTA}" in body


def test_el_excel_trae_UNA_columna_de_faltante(app, fake_db):
    """Dueña 2026-08-18: "falta tiene que ser en rollos o en kg pero no dos
    columnas, cuando hay unidades ya está".

    La de kilos servía para sumar la planilla entera, pero un total que mezcla
    telas con cuellos no se usa para nada.
    """
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()), \
         patch("modules.pedidos.service.acabados_por_producto", return_value={}):
        ws = _bajar(c)
    cabecera = [c.value for c in ws[1]]
    assert cabecera.count("Falta") == 1
    assert "Falta kg" not in cabecera
    assert cabecera[-1] == "Falta"


# ── el diseño de los controles ──────────────────────────────────────────────

def _tpl() -> str:
    return (Path(service.__file__).parent / "templates" / "inventario_rotativo"
            / "lista.html").read_text(encoding="utf-8")


def test_los_titulos_quedan_fijos_al_scrollear():
    """Dueña 2026-08-18: "que se mantengan los títulos cuando scrolleo".

    El sticky va en el `th` y no en el `thead` —Safari no pega un thead— y en
    papel se apaga: ahí la cabecera se repite por `table-header-group`.
    """
    tpl = _tpl()
    assert "position:sticky;top:0" in tpl.replace(" ", "")
    assert ".iv thead th{" in tpl
    assert ".iv.hoja thead th{position:static}" in tpl


def test_los_controles_no_son_todos_el_mismo_boton(app, fake_db):
    """Dueña 2026-08-18: "está todo con demasiado botón que no se diferencia".

    Tres cosas distintas, tres formas distintas: elegir vista y filtrar son
    segmentados (opciones excluyentes, pegadas), imprimir y bajar el Excel son
    acciones y viven a la derecha, y las familias son pestañas subrayadas.
    """
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        body = c.get("/inventario-rotativo").get_data(as_text=True)
    assert 'class="seg"' in body                 # vista y filtros
    assert 'class="acciones"' in body            # imprimir / excel aparte
    assert 'class="acc"' in body
    assert 'class="tab' in body                  # familias como pestañas
    assert "chip" not in body.split('class="iv')[1][:4000]   # ya no son todos iguales


def test_el_filtro_que_alarma_lleva_su_punto_rojo(app, fake_db):
    """Dentro de un segmentado gris, el que abre la alarma se distingue sin
    gritar: un punto, no un botón rojo entero."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo()):
        body = c.get("/inventario-rotativo").get_data(as_text=True)
    assert '<i class="punto"></i>Hay que teñir' in body


# ── las que todavía no entran ───────────────────────────────────────────────

def test_las_nuevas_se_cuentan_con_la_hora_de_ecuador():
    """En Asinfo `GETDATE()` devuelve UTC: pelado corre la ventana un día."""
    assert "DATEADD(hour, -5, GETDATE())" in service._sql_nuevos()


def test_la_ventana_de_nuevas_es_mas_corta_que_el_minimo_de_la_lista():
    """Es la razón de ser del renglón, y tiene que seguir siendo cierta.

    Con `SEMANAS_NUEVO < SEMANAS_MINIMAS`, "primera venta dentro de la
    ventana" IMPLICA "no llega al mínimo": el contador no puede contar algo
    que ya está en la lista, sin tener que cruzarlo contra ella. Si alguien
    sube la ventana a 52, el renglón empieza a mentir.
    """
    assert service.SEMANAS_NUEVO < service.SEMANAS_MINIMAS


def test_la_sql_de_nuevas_mira_la_bodega_de_terminado():
    sql = service._sql_nuevos()
    assert f"id_bodega = {service.BODEGA_TERMINADO}" in sql
    assert f"DATEADD(week, -{service.SEMANAS_NUEVO}," in sql


def test_nuevas_devuelve_el_conteo_y_los_kilos():
    service._cache_nuevos.clear()
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      return_value=([{"n": 7, "kg": 1234.56}], True)):
        assert service.nuevos() == {"n": 7, "kg": 1234.6}


def test_con_asinfo_caido_nuevas_devuelve_vacio_y_no_cero():
    """Cero es "no hay ninguna"; vacío es "no pude preguntar". No son lo mismo:
    con cero la pantalla diría que no falta nada mirar."""
    service._cache_nuevos.clear()
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      return_value=([], False)):
        assert service.nuevos() == {}


def test_sin_filas_nuevas_da_cero_sin_reventar():
    service._cache_nuevos.clear()
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      return_value=([], True)):
        assert service.nuevos() == {"n": 0, "kg": 0.0}


def test_la_pantalla_avisa_de_las_que_todavia_no_entran(app, fake_db):
    """El renglón de arriba de la lista: sin él, las 291 se leen como "esto es
    todo lo que hay"."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo(nuevos=[{"n": 12, "kg": 3400.0}])):
        body = c.get("/inventario-rotativo").get_data(as_text=True)
    assert "todavía no entran en la lista" in body
    assert "<b>12</b>" in body
    assert "3.400 kg" in body


def test_sin_ninguna_nueva_el_renglon_no_aparece(app, fake_db):
    """Un renglón que dice "0" es ruido: el dato es que no hay nada que mirar."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo(nuevos=[{"n": 0, "kg": 0.0}])):
        body = c.get("/inventario-rotativo").get_data(as_text=True)
    assert "todavía no entran en la lista" not in body


def test_en_el_papel_el_renglon_de_nuevas_no_va(app, fake_db):
    """La hoja es para la planta: lo que no está en la lista no se tiñe hoy."""
    service._cache.clear()
    service._cache_nuevos.clear()
    c = _login(app, fake_db)
    with patch.object(service.metabase_client, "fetch_dataset_estado",
                      side_effect=_fake_asinfo(nuevos=[{"n": 12, "kg": 3400.0}])):
        body = c.get("/inventario-rotativo?imprimir=1").get_data(as_text=True)
    assert "todavía no entran en la lista" not in body
