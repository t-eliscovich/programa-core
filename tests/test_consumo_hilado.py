"""Tests de la pestaña Consumo Hilado de Inventario (dueña 27/08/2026).

Réplica del Excel "Consumos de Material - Resumen" (Power Query a Asinfo):
kg de HILO despachados por OSM, por material y mes, con el % del mes, los
dos promedios del Excel y el saldo actual pegado en la misma tabla. Con
export a Excel que se lleva el filtro de año puesto.
"""
from __future__ import annotations

import datetime as _dt
import os
import sys
from unittest.mock import patch

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from modules.asinfo import service  # noqa: E402
from modules.stock_asinfo.views import _consumo_hilado_arma  # noqa: E402


def _login_stock(app, fake_db):
    rid = fake_db.add_role("Tester", ["stock.ver"])
    uid = fake_db.add_user("test", b"$2b$12$fakehash", rid)
    c = app.test_client()
    with c.session_transaction() as s:
        s["user_id"] = uid
    return c


_ROWS = [
    {"mes": 1, "material": "22/1", "kg": 600.0, "productos": 2},
    {"mes": 1, "material": "20/1", "kg": 400.0, "productos": 1},
    {"mes": 2, "material": "22/1", "kg": 300.0, "productos": 1},
]
_SALDOS = [
    {"material": "22/1", "kg": 5000.0},
    {"material": "40 D LYCRA", "kg": 80.0},  # sin consumo en el año
]


def _render(app, fake_db, ruta="/stock/consumo-hilado",
            rows=(_ROWS, True), saldos=(_SALDOS, True)):
    c = _login_stock(app, fake_db)
    with patch.object(service, "consumo_hilado_mensual", return_value=rows), \
         patch.object(service, "consumo_hilado_saldos", return_value=saldos):
        return c.get(ruta)


def test_pantalla_renderiza_los_dos_cuadros(app, fake_db):
    r = _render(app, fake_db)
    assert r.status_code == 200
    assert b"Despachado por mes" in r.data
    assert b"Promedio del a" in r.data  # Promedio del año y saldo actual
    assert b"22/1" in r.data and b"20/1" in r.data
    # el material sin consumo pero con saldo aparece igual (con su saldo)
    assert b"40 D LYCRA" in r.data
    # las pestañas están (Inventario <-> Consumo Hilado)
    assert b"/stock/fabricacion-tc" in r.data
    # el botón de Excel se lleva el año puesto
    assert b"/stock/consumo-hilado/export.xlsx" in r.data


def test_bridge_caido_renderiza_placeholder(app, fake_db):
    r = _render(app, fake_db, rows=([], False), saldos=([], False))
    assert r.status_code == 200
    assert b"Sin datos disponibles" in r.data


def test_arma_pivot_porcentajes_y_promedios():
    hoy = _dt.date(2026, 8, 27)
    d = _consumo_hilado_arma(_ROWS, 2026, hoy)
    assert [n for n, _ in d["meses"]] == [1, 2]
    assert d["total"] == 1300.0
    assert d["n_meses"] == 8  # meses transcurridos del año en curso
    m22 = d["materiales"][0]
    assert m22["material"] == "22/1"  # ordenado por total DESC
    assert m22["total"] == 900.0
    assert m22["por_mes"][1]["pct"] == 60.0  # 600 de 1000 en enero
    assert m22["por_mes"][2]["pct"] == 100.0
    # promedio del pivot del Excel: total / filas producto×mes
    assert m22["prom_hoja"] == 900.0 / 3
    # promedio mensual real: total / meses transcurridos
    assert m22["prom_mes"] == 900.0 / 8


def test_arma_pivot_anio_cerrado_divide_por_12():
    hoy = _dt.date(2026, 8, 27)
    d = _consumo_hilado_arma(_ROWS, 2025, hoy)
    assert d["n_meses"] == 12


def test_export_xlsx_baja_con_el_anio_puesto(app, fake_db):
    r = _render(app, fake_db, ruta="/stock/consumo-hilado/export.xlsx?anio=2025")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["Content-Type"]
    assert "consumo_hilado_2025.xlsx" in r.headers["Content-Disposition"]
    assert r.data[:2] == b"PK"  # xlsx = zip
    # y el contenido es abrible con lo que hay a mano
    import io

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.data))
    ws = wb.active
    assert ws.title == "Consumo Hilado 2025"
    textos = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert any("22/1" in t for t in textos)
    assert any("Promedio" in t for t in textos)


def test_sin_permiso_da_404(app, fake_db):
    rid = fake_db.add_role("SinStock", ["informes.ver"])
    uid = fake_db.add_user("sin", b"$2b$12$fakehash", rid)
    c = app.test_client()
    with c.session_transaction() as s:
        s["user_id"] = uid
    assert c.get("/stock/consumo-hilado").status_code == 404
    assert c.get("/stock/consumo-hilado/export.xlsx").status_code == 404
