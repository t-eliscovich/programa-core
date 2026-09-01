"""Tests de la sección Análisis → Lo parado.

Lo que se prueba es lo que se puede romper sin síntoma:

  · que un ítem VENDIDO siga en la lista (el pedido de la dueña);
  · que la cohorte no se pise (kg_al_marcar y fecha son de la primera vez);
  · que Metabase caído NO vacíe la pantalla (fail-closed);
  · que la sección sea invisible para quien no tiene el permiso.
"""

from __future__ import annotations

import re
from datetime import date

import pytest

from modules.analisis import asinfo_parado, queries, views

# ── El invariante que pidió la dueña ────────────────────────────────────────

def test_un_item_vendido_sigue_en_la_lista(monkeypatch):
    """`items()` sale de la COHORTE, no de la foto: si la foto ya no lo tiene
    (se vendió y dejó de estar parado), la fila tiene que seguir."""
    sql_visto = {}

    def fake_fetch_all(sql, params=None, conn=None):
        sql_visto["sql"] = " ".join(sql.split())
        return []

    monkeypatch.setattr(queries.db, "fetch_all", fake_fetch_all)
    queries.items()
    s = sql_visto["sql"]
    assert "FROM scintela.parado_cohorte c" in s, "la lista tiene que salir de la cohorte"
    assert "LEFT JOIN scintela.parado_foto" in s, (
        "si el JOIN fuera INNER, un ítem vendido desaparecería — que es "
        "exactamente lo que la dueña pidió que NO pase"
    )


def test_estado_marca_lo_que_se_movio(monkeypatch):
    """resuelto / empezó a moverse / sigue parado salen del SQL; acá se fija
    que el resumen los cuente bien."""
    filas = [
        {"stock_kg": 100, "kg_vendidos": 0, "clientes": 3, "kg_segunda": 0},
        {"stock_kg": 50, "kg_vendidos": 20, "clientes": 5, "kg_segunda": 12},
        {"stock_kg": 0, "kg_vendidos": 80, "clientes": 0, "kg_segunda": 0},
    ]
    r = queries.resumen(filas)
    assert r["n_items"] == 3
    assert r["kg"] == 150
    assert r["kg_vendidos"] == 100
    assert r["movidos"] == 2
    assert r["kg_segunda"] == 12
    assert r["n_segunda"] == 1


# ── Fail-closed contra Metabase ─────────────────────────────────────────────

def test_metabase_caido_no_vacia_la_pantalla(monkeypatch):
    """`fetch_dataset` devuelve [] tanto si no hay filas como si Metabase se
    cayó. Tomar el segundo caso por "no hay parados" borraría la pantalla y
    parecería una buena noticia."""
    monkeypatch.setattr(
        asinfo_parado.metabase_client, "fetch_dataset_estado",
        lambda *a, **k: ([], False))
    with pytest.raises(RuntimeError, match="Metabase no contestó"):
        asinfo_parado.parados()


def test_sin_filas_pero_metabase_ok_devuelve_lista_vacia(monkeypatch):
    monkeypatch.setattr(
        asinfo_parado.metabase_client, "fetch_dataset_estado",
        lambda *a, **k: ([], True))
    assert asinfo_parado.parados() == []


# ── El mapeo de vendedores ──────────────────────────────────────────────────

def test_vendedor_con_espacios_igual_mapea(monkeypatch):
    """⚠ Asinfo devuelve "Ramirez Edgar " con espacio al final. Sin `.strip()`
    el mapeo falla en silencio y TODOS quedan como mostrador — que es un
    estado válido, y por eso no se nota."""
    monkeypatch.setattr(
        asinfo_parado.metabase_client, "fetch_dataset_estado",
        lambda *a, **k: ([{"vendedor": "Ramirez Edgar ", "subcategoria": "X",
                           "codigo": "ABC", "anio": 2026}], True))
    assert asinfo_parado.llamados()[0]["vend_pc"] == "EDG"


def test_vendedor_desconocido_queda_sin_codigo(monkeypatch):
    monkeypatch.setattr(
        asinfo_parado.metabase_client, "fetch_dataset_estado",
        lambda *a, **k: ([{"vendedor": "Cía. Ltda. Intela", "subcategoria": "X",
                           "codigo": "ABC", "anio": 2026}], True))
    assert asinfo_parado.llamados()[0]["vend_pc"] is None


# ── Las definiciones que no se pueden aflojar ───────────────────────────────

def test_el_consumidor_final_no_es_candidato():
    """VPM es el mostrador. Sin excluirlo, telas con 1.700 kg parados mostraban
    un único "candidato" que compró 3 kg y no existe."""
    assert "<> 'VPM'" in asinfo_parado.SQL_LLAMADOS


def test_las_notas_de_credito_no_netean():
    """Sólo documentos de venta (7, 251) y cantidad > 0: "vendió 0 kg en 12
    meses" tiene que significar que no salió, no que salió y volvió."""
    for sql in (asinfo_parado.SQL_PARADOS, asinfo_parado.SQL_LLAMADOS):
        assert "fc.id_documento IN (7, 251)" in sql
        assert "20, 451" not in sql
        assert "dfc.cantidad > 0" in sql


def test_los_llamados_son_por_tela_y_no_por_tela_color():
    """El color no entra en la llamada: quien compra Kiana Forro compra Kiana
    Forro y el color se negocia. Agrupar por tela × color deja listas de uno."""
    assert "GROUP BY pr.nombre_subcategoria_producto, YEAR(fc.fecha), fc.id_empresa" \
        in asinfo_parado.SQL_LLAMADOS


def test_la_fecha_de_venta_se_compara_como_fecha():
    """Metabase devuelve las fechas como texto ISO y Postgres como `date`.
    Comparar str contra date explota; comparar str contra str compara
    alfabéticamente y a veces acierta, que es peor."""
    assert queries._fecha("2026-08-13T00:00:00Z") == date(2026, 8, 13)
    assert queries._fecha(date(2026, 8, 13)) == date(2026, 8, 13)


# ── El candado ──────────────────────────────────────────────────────────────

def test_todas_las_rutas_piden_el_permiso():
    from modules.analisis import views
    for regla in ("inicio", "parado", "parado_actualizar"):
        f = getattr(views, regla)
        assert getattr(f, "_permiso", None) == "analisis.ver", (
            f"{regla} tiene que estar gateada: si no, la sección aparece "
            f"para cualquiera que sepa la URL")


def test_el_menu_no_ofrece_pantallas_que_no_existen(app):
    """Los links de esta app son strings hardcodeados: un link a una ruta que no
    existe NO se ve desde el código, sólo como 404 al clickear. Por eso el test
    no compara contra una lista escrita a mano —que se desactualiza igual— sino
    contra el `url_map` real."""
    from modules.analisis import views

    rutas = {r.rule for r in app.url_map.iter_rules()}
    for m in views.MENU:
        if m["listo"]:
            assert m["url"] in rutas, (
                f"el menú ofrece {m['url']} y esa ruta no existe: es un 404 "
                f"que no se ve leyendo el código")


def test_todos_los_links_internos_de_las_plantillas_existen(app):
    """Ídem para los links escritos a mano DENTRO de las pantallas de la
    sección — el botón «Ver por cliente», el «← Ver por tela», el form de
    actualizar."""
    import re
    from pathlib import Path

    rutas = {r.rule for r in app.url_map.iter_rules()}
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    for archivo in carpeta.glob("*.html"):
        html = archivo.read_text(encoding="utf-8")
        for url in re.findall(r'(?:href|action)="(/analisis[^"?{]*)', html):
            assert url in rutas, f"{archivo.name} linkea a {url}, que no existe"


# ── Lo que se rompió en producción y no puede volver a pasar ────────────────

def test_ningun_diccionario_que_va_a_un_template_usa_claves_de_dict():
    """⚠ En Jinja `x.items` devuelve el MÉTODO del diccionario, no el dato: la
    pantalla imprime "<built-in method items of dict object at 0x…>" donde va
    la cifra. No da error — renderiza 200 con un texto absurdo.

    Pasó DOS veces: primero en la tarjeta "Parado hoy" y después, con el test
    ya escrito pero mirando sólo `resumen()`, en el resumen por grupo. Por eso
    ahora el test recorre todas las funciones que arman diccionarios para un
    template, y hay que agregarlas acá al crearlas."""
    filas = [{"stock_kg": 10, "kg_vendidos": 0, "clientes": 1, "kg_segunda": 0,
              "categoria": "Jersey", "subcategoria": "Jersey 3"}]
    candidatos = [queries.resumen(filas)] + queries.por_grupo(filas)
    for d in candidatos:
        for clave in d:
            assert not hasattr({}, clave), (
                f"la clave '{clave}' choca con dict.{clave} y Jinja va a "
                f"resolver el método antes que el dato")


def test_la_plantilla_no_pide_claves_que_el_resumen_no_tiene():
    import re
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")
    pedidas = set(re.findall(r"resumen\.(\w+)", html))
    assert pedidas <= set(queries.resumen([])), (
        f"la plantilla pide {pedidas - set(queries.resumen([]))}, que no existe")


# ── El tercer escalón: improbable ───────────────────────────────────────────

def test_los_candidatos_no_se_cortan_en_dos_anios():
    """Dueña 17/08/2026: "ponemelos como improbable pero la última fecha aunque
    sea de 2024". Un cliente de 2023 no es un buen candidato, pero es más que
    un renglón vacío."""
    assert "YEAR(GETDATE()) - 1" not in asinfo_parado.SQL_LLAMADOS, (
        "el corte de dos años dejaba 7 telas sin una sola pista teniendo el "
        "dato a mano")
    assert "MAX(anio) AS anio FROM compras" in asinfo_parado.SQL_LLAMADOS, (
        "el año que manda es el ÚLTIMO con compras, sea cual sea")


def test_el_anio_de_los_candidatos_sigue_a_la_vista():
    """La columna Estado se sacó el 18/08/2026 ("el estado tampoco importa
    mucho"), pero de qué año son los candidatos SÍ importa: es la diferencia
    entre llamar a alguien que compró en julio y a uno que compró en 2023. Vive
    en el detalle que se abre."""
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")
    assert "f.pista" in html or "anio_pista" in html


# ── La hoja del vendedor ────────────────────────────────────────────────────

def _con_puntos(filas, puntos=None):
    """Un `db.fetch_all` falso que también sabe contestar por los PUNTOS.

    ⚠ El stub de antes devolvía las mismas filas para TODA consulta, así que
    `puntos_por_tela()` recibía filas de clientes y reventaba. Devuelve siempre
    al menos una fila de puntaje para que no se dispare el camino de "tabla
    vacía", que sale a buscar a Asinfo.
    """
    pfilas = [{"subcategoria": sub, "categoria": None, "kg_base": 0,
               "kg_12m": 0, "meses": None, "nivel": 1, "puntos": pk}
              for sub, pk in (puntos or {"—": 1}).items()]

    def fetch(sql, params=None, conn=None):
        if "parado_punto" in " ".join(str(sql).split()):
            return pfilas
        return filas
    return fetch


def _filas_falsas():
    return [
        {"codigo_cli": "AAA", "nombre": "Cliente Grande", "provincia": "GUAYAS",
         "vend_pc": "FL1", "subcategoria": "Kiana", "kg_cliente": 100,
         "ultima_compra": None, "anio": 2026, "kg_parado": 900,
         "colores_parados": "CAR"},
        {"codigo_cli": "BBB", "nombre": "Cliente Chico", "provincia": "AZUAY",
         "vend_pc": "FL1", "subcategoria": "Microfibra", "kg_cliente": 10,
         "ultima_compra": None, "anio": 2026, "kg_parado": 100,
         "colores_parados": "FRE"},
        {"codigo_cli": "CCC", "nombre": "Cliente Viejo", "provincia": "AZUAY",
         "vend_pc": "FL1", "subcategoria": "Jersey Fancy", "kg_cliente": 50,
         "ultima_compra": None, "anio": 2023, "kg_parado": 500,
         "colores_parados": "FNB"},
    ]


def test_la_hoja_ordena_por_oportunidad_por_defecto(monkeypatch):
    monkeypatch.setattr(queries.db, "fetch_all", _con_puntos(_filas_falsas()))
    r = queries.por_cliente()
    assert [c["codigo"] for c in r["clientes"]] == ["AAA", "BBB"], (
        "la hoja existe para decidir por quién empezar: si el vendedor corta a "
        "la mitad, que haya cortado por lo chico")


def test_los_improbables_van_al_final_y_aparte(monkeypatch):
    monkeypatch.setattr(queries.db, "fetch_all", _con_puntos(_filas_falsas()))
    r = queries.por_cliente()
    assert [c["codigo"] for c in r["improbables"]] == ["CCC"]
    assert "CCC" not in [c["codigo"] for c in r["clientes"]], (
        "mezclado ensucia una lista que el vendedor tiene que poder creer")


def test_un_cliente_con_una_tela_vieja_y_una_nueva_no_es_improbable(monkeypatch):
    filas = _filas_falsas()
    filas[2]["codigo_cli"] = "AAA"          # la vieja es del cliente grande
    filas[2]["nombre"] = "Cliente Grande"
    monkeypatch.setattr(queries.db, "fetch_all", _con_puntos(filas))
    r = queries.por_cliente()
    assert [c["codigo"] for c in r["improbables"]] == []
    grande = next(c for c in r["clientes"] if c["codigo"] == "AAA")
    assert len(grande["telas"]) == 2


def test_el_orden_alfabetico_y_el_de_provincia_existen(monkeypatch):
    monkeypatch.setattr(queries.db, "fetch_all", _con_puntos(_filas_falsas()))
    assert [c["codigo"] for c in queries.por_cliente(orden="codigo")["clientes"]] \
        == ["AAA", "BBB"]
    assert [c["provincia"] for c in queries.por_cliente(orden="provincia")["clientes"]] \
        == ["AZUAY", "GUAYAS"]


def test_un_orden_inventado_no_rompe_la_pantalla(monkeypatch):
    monkeypatch.setattr(queries.db, "fetch_all", _con_puntos(_filas_falsas()))
    r = queries.por_cliente(orden="loquesea")
    assert [c["codigo"] for c in r["clientes"]] == ["AAA", "BBB"]


def test_la_hoja_avisa_que_los_kg_no_se_suman():
    """Una misma tela parada aparece en la lista de todos los que la compran:
    sumar la columna entre clientes da mucho más que el stock real."""
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado_clientes.html").read_text(encoding="utf-8")
    assert "no se suman entre clientes" in html


def test_la_hoja_se_imprime_con_la_misma_plantilla():
    """⚠ En /mi-cartera la oficina y el portal imprimieron órdenes distintos
    ocho días sin síntoma, por tener dos caminos para el mismo papel."""
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado_clientes.html").read_text(encoding="utf-8")
    assert "@media print" in html and "window.print()" in html


# ── El % del parado ─────────────────────────────────────────────────────────

def test_el_porcentaje_se_calcula_sobre_las_filas_que_se_muestran(monkeypatch):
    """⭐ El total del % sale de un SUM() OVER () sobre el mismo conjunto de
    filas que se dibuja. Traer el total por separado es cómo dos números del
    mismo cuadro terminan sin sumar 100."""
    visto = {}

    def fake(sql, params=None, conn=None):
        visto["sql"] = " ".join(sql.split())
        return []

    monkeypatch.setattr(queries.db, "fetch_all", fake)
    queries.items()
    s = visto["sql"]
    assert "OVER ()" in s, "el total tiene que ser el de la misma consulta"
    assert "OVER (PARTITION BY c.subcategoria)" in s, (
        "el % por TELA suma los colores de esa tela")
    assert "NULLIF(SUM(COALESCE(f.stock_kg, 0)) OVER (), 0)" in s, (
        "sin el NULLIF, una lista vacía divide por cero")


def test_la_tabla_no_queda_desalineada_al_agregar_la_columna():
    """La fila de detalle usa colspan: si no acompaña a las columnas de arriba,
    la tabla se desarma y no da ningún error."""
    import re
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")
    # ⚠ Apuntado a la tabla PRINCIPAL por su id: la pantalla tiene otra tabla
    # (el resumen por grupo) y agarrar su encabezado hacía fallar el test por
    # el motivo equivocado.
    principal = html[html.index('<table id="tabla">'):]
    encabezado = re.search(r"<thead><tr>(.*?)</tr></thead>", principal, re.S).group(1)
    columnas = len(re.findall(r"<th", encabezado))
    colspan = int(re.search(r'colspan="(\d+)"', principal).group(1))
    assert colspan == columnas, (
        f"el detalle abarca {colspan} columnas y la tabla tiene {columnas}")


# ── Primera y segunda calidad ───────────────────────────────────────────────

def test_la_calidad_sale_del_lote_y_no_del_producto():
    """⭐ En Asinfo la calidad es el ATRIBUTO 2 del LOTE (PRI/SEG), no una
    propiedad del producto. Por eso un mismo tela × color puede tener kilos de
    las dos y hacen falta DOS columnas: 26 de los ítems parados tienen de las
    dos, y una sola columna obligaría a elegir y perdería la mitad del dato."""
    s = asinfo_parado.SQL_PARADOS
    assert "saldo_producto_lote" in s
    assert "l.id_valor_atributo_2" in s
    assert "v.codigo = 'SEG'" in s
    assert "AS kg_primera" in s and "AS kg_segunda" in s


def test_lo_que_no_esta_marcado_cuenta_como_primera():
    """585.011 lotes no tienen el atributo cargado. Tratarlos como "sin
    calidad" dejaría la mayoría del stock en una tercera categoría que no
    existe para nadie; tratarlos como SEGUNDA sería peor. El CASE los manda a
    primera y la pantalla lo dice."""
    assert "CASE WHEN v.codigo = 'SEG' THEN 0 ELSE lu.saldo END" in \
        asinfo_parado.SQL_PARADOS


def test_los_kilos_por_calidad_suman_el_total_de_la_fila(monkeypatch):
    """Si `saldo_producto` y `saldo_producto_lote` se despegan, primera + segunda
    deja de dar el stock de la fila. Hoy cierran (0,006% en la bodega 53)."""
    filas = [{"stock_kg": 100, "kg_primera": 88, "kg_segunda": 12,
              "kg_vendidos": 0, "clientes": 1}]
    r = queries.resumen(filas)
    assert r["kg_segunda"] == 12
    assert filas[0]["kg_primera"] + filas[0]["kg_segunda"] == filas[0]["stock_kg"]


# ── Grupo / subgrupo y Excel ────────────────────────────────────────────────

def test_el_grupo_de_producto_llega_desde_asinfo():
    """En Asinfo el GRUPO es nombre_categoria_producto y el SUBGRUPO es
    nombre_subcategoria_producto — lo que la pantalla llamaba "tela"."""
    assert "MIN(p.nombre_categoria_producto) AS categoria" in \
        asinfo_parado.SQL_PARADOS
    assert "stk.categoria" in asinfo_parado.SQL_PARADOS


def test_la_pantalla_ofrece_los_dos_desplegables():
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")
    assert 'id="grupo"' in html and 'id="subgrupo"' in html
    assert "data-grupo=" in html and "data-sub=" in html


def test_elegir_un_grupo_recorta_los_subgrupos():
    """Sin esto se puede pedir "Fleece" + "Jersey 3.5": la tabla queda vacía y
    parece que no hay datos, cuando lo que hay es una combinación imposible."""
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")
    assert "function grupoCambio()" in html
    assert "o.dataset.g === g" in html


def test_el_excel_lo_arma_el_servidor_y_baja_todo():
    """Dueña 26/08/2026: *"bajar a excel se baja horrible: bajalo a algo con
    formato"*. Lo que bajaba era un CSV pegado en el navegador con punto y
    coma: todo texto, columnas de un caracter y kilos que Excel no podía sumar.

    ⚠ Baja TODO, no lo filtrado, y el botón lo dice. Los filtros son de
    JavaScript: replicarlos en el servidor sería la misma regla escrita dos
    veces en dos lenguajes, y el día que una cambie el archivo diría algo
    distinto de la pantalla. Uno baja a Excel justamente para filtrar allá —por
    eso van Grupo y Tela como columnas propias."""
    html = _html_parado()
    assert "function bajarExcel()" not in html, (
        "volvió el CSV hecho a mano en el navegador")
    # ⚠ El botón elige la ruta según quién mira: la de la oficina para la
    # oficina, y la del vendedor —que cuelga de su allowlist— para el
    # vendedor. Apuntar los dos a /analisis/parado.xlsx le daba 404 al
    # vendedor (reportado 27/08/2026: "no pueden bajar bien el excel").
    assert "/analisis/parado.xlsx" in html
    assert "'/analisis/competencia/telas.xlsx'" in html and "if mia" in html
    assert "Bajar todo a Excel" in html, "el botón tiene que decir que baja todo"
    # ⚠ El Excel de Vendidos se mudó con su tabla (26/08/2026): ahora vive en
    # `vendidos.html`, que es la pantalla que lo ofrece.
    assert 'href="/analisis/vendidos.xlsx"' in _html_vendidos(), (
        "y la tabla de Vendidos también se baja (dueña: «y lo mismo con los "
        "vendidos»)")


def test_el_excel_lleva_grupo_y_tela_como_columnas():
    """Uno baja a Excel para filtrar y pivotear allá: sin Grupo y Tela en
    columnas propias no se puede."""
    import inspect as _i
    fuente = _i.getsource(views.parado_xlsx) + _i.getsource(views._excel_saldos)
    for col in ('"Grupo"', '"Tela"', '"Color"', '"Queda"', '"Vendido"'):
        assert col in fuente, f"falta la columna {col}"
    assert "excel.respuesta(" in fuente and "excel.libro(" in fuente


def test_el_excel_manda_numeros_y_no_texto():
    """El CSV viejo mandaba "1.779" como cadena: Excel lo alineaba a la
    izquierda y no se podía sumar una columna."""
    from modules.analisis.views import _num
    assert _num("1779") == 1779 and isinstance(_num("1779"), int)
    assert _num(20.5) == 20.5
    assert _num(None) is None and _num("") is None
    assert _num("no es un numero") == "no es un numero"


def test_el_resumen_por_grupo_suma_100(monkeypatch):
    filas = [
        {"categoria": "Jersey", "subcategoria": "Jersey 3", "stock_kg": 60,
         "kg_segunda": 0},
        {"categoria": "Jersey", "subcategoria": "Jersey 3.5", "stock_kg": 20,
         "kg_segunda": 5},
        {"categoria": "Fleece", "subcategoria": "Fleece 102", "stock_kg": 20,
         "kg_segunda": 0},
    ]
    g = queries.por_grupo(filas)
    assert [x["grupo"] for x in g] == ["Jersey", "Fleece"], "de mayor a menor"
    assert round(sum(x["pct"] for x in g), 6) == 100
    assert g[0]["n_items"] == 2 and g[0]["subgrupos"] == 2
    assert g[0]["kg_segunda"] == 5


def test_el_resumen_sale_de_las_mismas_filas_que_la_tabla():
    """Si fuera una consulta aparte, el resumen y la tabla podrían dejar de
    coincidir el día que una cambie de criterio, y no habría síntoma."""
    import inspect
    fuente = inspect.getsource(queries.por_grupo)
    assert "db.fetch" not in fuente, "por_grupo recibe las filas, no las consulta"


def test_sin_grupo_no_desaparece_del_resumen():
    """Una fila sin categoría cargada tiene que sumar igual: si se cayera del
    resumen, los porcentajes seguirían dando 100 y faltarían kilos."""
    g = queries.por_grupo([{"categoria": None, "subcategoria": "X",
                            "stock_kg": 10, "kg_segunda": 0}])
    assert g[0]["grupo"] == "(sin grupo)" and g[0]["kg"] == 10


def test_ordenar_mueve_la_fila_con_su_detalle():
    """⚠ Cada tela son DOS <tr>: la fila y su detalle escondido. Un ordenador
    que mueva filas sueltas deja los detalles pegados a la tela equivocada, y
    como están cerrados no se nota hasta que alguien abre uno."""
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")
    assert "pares.push([tr, tr.nextElementSibling])" in html
    assert "frag.appendChild(fila); frag.appendChild(det);" in html


def test_las_columnas_numericas_ordenan_por_el_numero_y_no_por_el_texto():
    """"1.779" y "580" como texto ordenan al revés: "1" < "5"."""
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")
    assert 'parseFloat(td.dataset.v ?? \'0\')' in html
    import re
    principal = html[html.index('<table id="tabla">'):]
    encabezado = re.search(r'<thead><tr>\n(.*?)</tr></thead>', principal, re.S).group(1)
    numericas = re.findall(r'data-i="(\d+)" data-num="1"', encabezado)
    assert numericas, "alguna columna tiene que estar marcada como numérica"
    # toda columna marcada numérica tiene que tener data-v en sus celdas
    assert html.count('data-v="') >= len(numericas), (
        "una columna numérica sin data-v ordena por el texto formateado")


# ── Excel desde la hoja del vendedor ────────────────────────────────────────

def test_el_excel_de_la_hoja_sale_de_la_misma_funcion_que_la_pantalla(monkeypatch):
    """Si fueran dos caminos, el archivo y el papel podrían decir cosas
    distintas del mismo día."""
    monkeypatch.setattr(queries.db, "fetch_all", _con_puntos(_filas_falsas()))
    plano = queries.por_cliente_plano()
    pantalla = queries.por_cliente()
    assert {f["codigo"] for f in plano} == (
        {c["codigo"] for c in pantalla["clientes"]}
        | {c["codigo"] for c in pantalla["improbables"]})


def test_el_excel_de_la_hoja_distingue_candidatos_de_improbables(monkeypatch):
    monkeypatch.setattr(queries.db, "fetch_all", _con_puntos(_filas_falsas()))
    plano = queries.por_cliente_plano()
    tipos = {f["codigo"]: f["tipo"] for f in plano}
    assert tipos["AAA"] == "candidato" and tipos["CCC"] == "improbable", (
        "aplanado a Excel, un improbable mezclado con los buenos es "
        "indistinguible — y el orden de las filas no lo dice")


def test_el_excel_de_la_hoja_respeta_el_vendedor_elegido(monkeypatch):
    """A diferencia del CSV de «Lo parado», acá el filtro NO es de JavaScript:
    viaja en la URL y lo aplica la misma función que dibuja la pantalla, así que
    puede y debe viajar al archivo."""
    import inspect

    from modules.analisis import views
    fuente = inspect.getsource(views.parado_clientes_csv)
    assert 'request.args.get("vend")' in fuente
    assert "queries.por_cliente_plano(vend, orden)" in fuente


def test_las_telas_de_cada_cliente_salen_de_mayor_a_menor(monkeypatch):
    filas = _filas_falsas()
    for f in filas:
        f["codigo_cli"], f["nombre"] = "AAA", "Uno"
    filas[0]["kg_parado"], filas[1]["kg_parado"], filas[2]["kg_parado"] = 10, 900, 50
    monkeypatch.setattr(queries.db, "fetch_all", _con_puntos(filas))
    telas = queries.por_cliente()["clientes"][0]["telas"]
    assert [float(t["kg_parado"]) for t in telas] == [900, 50, 10]


def test_las_dos_pantallas_tienen_su_boton_de_excel():
    """Dueña 17/08/2026: "de todos lados deberia poder bajar a excel"."""
    from pathlib import Path
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    for archivo, texto in (("parado.html", "Bajar todo a Excel"),
                           ("vendidos.html", "Bajar a Excel"),
                           ("parado_clientes.html", "Bajar a Excel")):
        html = (carpeta / archivo).read_text(encoding="utf-8")
        assert texto in html, f"{archivo} no ofrece bajar a Excel"


def test_la_tela_cruda_no_entra_aunque_este_en_la_bodega_de_terminado():
    """Dueña 17/08/2026: "tela cruda no debería estar.. es producto terminado".
    Hay 2 productos de categoría TELA CRUDA con saldo en la bodega 53 (37 kg):
    estar ahí no los convierte en producto terminado, y el filtro por bodega
    solo no los saca."""
    for sql in (asinfo_parado.SQL_PARADOS, asinfo_parado.SQL_LLAMADOS):
        assert "'TELA CRUDA'" in sql
    assert "'TELA CRUDA'" in asinfo_parado.CATS


def test_los_tres_grupos_chicos_van_juntos(monkeypatch):
    """Dueña 17/08/2026: "los ultimos y mas chicos unilos todos en una
    categoria franela cuellos y punos". Entre los tres son 361 kg — menos del
    1% — y ocupaban tres renglones del resumen para decir casi nada."""
    visto = {}
    monkeypatch.setattr(queries.db, "fetch_all",
                        lambda sql, *a, **k: visto.setdefault("sql", " ".join(sql.split())) and [])
    queries.items()
    s = visto["sql"]
    assert "'Franela', 'Cuellos', 'Puños'" in s
    assert "'FCP'" in s, "la sigla, que es lo que entra en la columna"


def test_los_grupos_chicos_se_unen_al_LEER_y_no_al_guardar():
    """Unirlos en el refresh perdería el dato crudo de Asinfo. Uniéndolos en la
    lectura, el día que uno crezca se separa cambiando una línea."""
    import inspect
    assert "Franela" not in inspect.getsource(queries.actualizar)
    assert "Franela" in inspect.getsource(queries.items)


def test_la_pantalla_de_saldos_ordena_por_puntos(monkeypatch):
    """⭐ Dueña 24/08/2026: "idem para la pantalla de saldos". La pregunta que
    se hace el que abre esta lista es a qué tela conviene ir, y 300 kg de una
    tela de 10 puntos valen más que 3.000 de una de 1. Ordenada por kilos ponía
    arriba justo lo que sale solo."""
    monkeypatch.setattr(
        queries, "puntos_por_tela",
        lambda: {"Microfibra": {"puntos": 10, "nivel_nombre": "Difícil"},
                 "Fleece 102": {"puntos": 1, "nivel_nombre": "Fácil"}})
    filas = queries.con_puntos([
        {"subcategoria": "Fleece 102", "categoria": "Fleece", "stock_kg": 2900,
         "kg_segunda": 0},
        {"subcategoria": "Microfibra", "categoria": "Poliester", "stock_kg": 300,
         "kg_segunda": 0},
    ])
    assert [f["subcategoria"] for f in filas] == ["Microfibra", "Fleece 102"], (
        "300 kg a 10 puntos van ARRIBA de 2.900 kg a 1")
    assert filas[0]["puntos_fila"] == 3000 and filas[1]["puntos_fila"] == 2900
    assert filas[0]["puntos"] == 10 and filas[0]["nivel"] == "Difícil"


def test_una_tela_sin_puntaje_vale_uno_y_no_cero(monkeypatch):
    """⚠ Un kilo vendido nunca puede contar cero. Pasa sólo si la cohorte
    creció después de congelar los puntos."""
    monkeypatch.setattr(queries, "puntos_por_tela", dict)
    filas = queries.con_puntos([
        {"subcategoria": "Tela nueva", "categoria": "Jersey", "stock_kg": 100,
         "kg_segunda": 0}])
    assert filas[0]["puntos"] == 1 and filas[0]["puntos_fila"] == 100


def test_la_pantalla_de_saldos_muestra_lo_que_vale_cada_tela():
    import inspect
    from pathlib import Path
    html = ((Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "parado.html").read_text(encoding="utf-8"))
    assert ">Vale<" in html and ">Puntos<" in html
    # ⚠ "Vale" no puede esconderse en el celular: es la columna que decide.
    import re
    col = re.search(r'<th[^>]*>Vale<', " ".join(html.split()))
    assert col and "opt" not in col.group(0)
    for vista in (views.parado, views.mis_telas, views.parado_xlsx):
        assert "con_puntos" in inspect.getsource(vista), (
            f"{vista.__name__} tiene que traer los puntos")


def test_la_hoja_del_vendedor_dice_lo_que_vale_cada_tela(monkeypatch):
    """⭐ Dueña 24/08/2026: "es el único papel que se lleva a la calle y es
    justo donde falta". Sin los puntos, el que sale con la hoja en la mano no
    sabe cuál de las telas de ese cliente vale diez veces más que la de al
    lado."""
    filas = [
        {"codigo_cli": "AAA", "nombre": "Cliente", "provincia": "Guayas",
         "vend_pc": "RMY", "subcategoria": "Fleece 102", "kg_cliente": 10,
         "ultima_compra": None, "anio": date.today().year, "colores": 1,
         "kg_parado": 2900, "colores_parados": "BLA"},
        {"codigo_cli": "AAA", "nombre": "Cliente", "provincia": "Guayas",
         "vend_pc": "RMY", "subcategoria": "Microfibra", "kg_cliente": 10,
         "ultima_compra": None, "anio": date.today().year, "colores": 1,
         "kg_parado": 300, "colores_parados": "NEG"},
    ]
    monkeypatch.setattr(
        queries.db, "fetch_all",
        _con_puntos(filas, {"Fleece 102": 1, "Microfibra": 10}))
    c = queries.por_cliente()["clientes"][0]
    assert [t["subcategoria"] for t in c["telas"]] == ["Microfibra", "Fleece 102"], (
        "primero la que más puntos da, no la que más pesa")
    assert c["telas"][0]["puntos"] == 10
    assert c["puntos_potencial"] == 300 * 10 + 2900 * 1

    from pathlib import Path
    html = ((Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "parado_clientes.html")
            .read_text(encoding="utf-8"))
    import re
    col = re.search(r"<th[^>]*>Vale<", " ".join(html.split()))
    assert col, "la hoja tiene que traer la columna Vale"
    # ⚠ Ni en el celular ni en el papel: es la cifra que decide.
    assert "opt" not in col.group(0) and "col" not in col.group(0)


# ── La competencia ──────────────────────────────────────────────────────────

#: Qué tela representa a cada grupo en los datos falsos. Los tests hablan en
#: grupos (Jersey, Fleece) porque así se lee la regla, pero el puntaje vive en
#: la TELA: el helper traduce.
_TELA_DE = {"Jersey": "Jersey 3", "Fleece": "Fleece 102"}


def _competencia_falsa(monkeypatch, vendido=None,
                       semanas=None, meses=None, base=None, puntos=None):
    filas = [
        {"categoria": "Jersey", "subcategoria": "Jersey 3", "color": "NEG",
         "stock_kg": 6000, "kg_segunda": 0, "kg_vendidos": 0, "clientes": 1},
        {"categoria": "Fleece", "subcategoria": "Fleece 102", "color": "BLA",
         "stock_kg": 4000, "kg_segunda": 0, "kg_vendidos": 0, "clientes": 1},
    ]
    # ⭐ Por defecto TODA tela vale 1 punto: así los tests que hablan de kilos
    # siguen midiendo lo que decían medir (un punto = un kilo) y los que hablan
    # de puntos pasan un `puntos` distinto y sólo cambia eso.
    puntos = puntos or {}
    pfilas = [{"subcategoria": f["subcategoria"], "categoria": f["categoria"],
               "kg_base": f["stock_kg"] + f["kg_vendidos"], "kg_12m": 1000,
               "meses": 1, "nivel": 1,
               "puntos": puntos.get(f["subcategoria"], 1)} for f in filas]

    def _con_tela(filas_v):
        salida = []
        for r in filas_v or []:
            r = dict(r)
            r.setdefault("subcategoria", _TELA_DE.get(r.get("categoria"), ""))
            salida.append(r)
        return salida

    def fake(sql, params=None, conn=None):
        s = " ".join(sql.split())
        # ⚠ La LISTA se reconoce por su tabla raíz y va PRIMERO: `items()`
        # joinea el puntaje (para el grupo de la tela vendida entera) y las
        # ventas (para abrir lo vendido por categoría), así que su consulta
        # nombra las tres tablas y cualquier rama de abajo se la comía.
        if "FROM scintela.parado_cohorte" in s:
            return filas
        if "date_trunc('week'" in s:
            return semanas or []
        if "date_trunc('month'" in s:
            return meses or []
        # ⚠ `FROM` y no sólo el nombre: desde que `items()` joinea el puntaje
        # para sacar el grupo de la tela vendida entera, su consulta también
        # nombra `parado_punto` y esta rama se la comía —devolvía las filas del
        # puntaje y la lista se quedaba sin `stock_kg`.
        if "FROM scintela.parado_punto" in s:
            return pfilas
        # ⚠ `FROM` otra vez: `items()` joinea `parado_venta` para abrir lo
        # vendido por categoría, así que su consulta también lo nombra.
        if "FROM scintela.parado_venta" in s:
            return _con_tela(vendido)
        if "parado_share" in s:
            return []
        if "parado_base" in s:
            # la meta congelada del día de la largada; vacía = todavía no largó
            return [{"categoria": k, "kg": v, "fijada_el": date(2026, 8, 25)}
                    for k, v in (base or {}).items()]
        return filas

    def fake_one(sql, params=None, conn=None):
        s = " ".join(sql.split())
        if "parado_config" in s:
            # ⚠ Devolver lo mismo para toda clave hacía que `cierre` valiera la
            # fecha de largada y el test de la fecha de cierre pasara por el
            # motivo equivocado.
            return {"valor": {"largada": "2026-08-25",
                              "cierre": "2026-12-31"}[(params or ("",))[0]]}
        return None

    monkeypatch.setattr(queries.db, "fetch_all", fake)
    monkeypatch.setattr(queries.db, "fetch_one", fake_one)
    return queries.competencia()


def test_no_hay_meta_por_vendedor(monkeypatch):
    """⭐ Dueña 24/08/2026: "ya que es por puntos, saquemos la meta, ideal es +
    puntos gana".

    ⚠ Y no era sólo simplificar: la meta ya no decidía NADA. Los siete tenían
    la misma (los puntos totales sobre 7), así que ordenar por "% de su meta"
    era dividir a todos por la misma constante — exactamente el mismo orden que
    ordenar por puntos. Era una cuenta de más en la pantalla."""
    c = _competencia_falsa(monkeypatch, puntos={"Jersey 3": 10, "Fleece 102": 1})
    for r in c["ranking"]:
        assert "meta" not in r, "no puede quedar una meta por vendedor"
        assert "pct" not in r, "ni un % contra una meta"
    assert c["puntos_en_juego"] == 64000, (
        "lo que sí queda es cuántos puntos hay en juego, como referencia")


def test_el_ranking_va_por_puntos(monkeypatch):
    c = _competencia_falsa(monkeypatch, vendido=[
        {"vendedor": "Quintero Jose", "categoria": "Jersey", "kg": 100,
         "ultima": None},
        {"vendedor": "Intela", "categoria": "Jersey", "kg": 50, "ultima": None},
    ])
    assert c["ranking"][0]["vendedor"] == "Quintero Jose"
    assert c["ranking"][0]["puesto"] == 1
    assert c["ranking"][0]["pct_lider"] == 100, "el primero es la vara"
    assert c["ranking"][1]["pct_lider"] == 50, "y el resto, contra él"


def test_intela_compite_como_uno_mas(monkeypatch):
    """Decisión de la dueña con el dato a la vista: Intela es el 51,3% de las
    ventas de estas telas. "Compite como uno mas porque hay una vendedora
    dedicada"."""
    c = _competencia_falsa(monkeypatch)
    assert "Intela" in [r["vendedor"] for r in c["ranking"]]
    assert len(c["ranking"]) == 7


def test_un_vendedor_que_ya_no_esta_no_entra_al_ranking(monkeypatch):
    """Un kilo firmado por alguien que no está entre los siete no le suma a
    nadie en el ranking, pero salió de la bodega igual: cuenta para el grupo.

    ⚠ Ya NO puede pasar en producción: `_quien_vendio()` manda a Intela todo lo
    que no firma uno de los siete, y lo hace al ESCRIBIR `parado_venta`. La
    pantalla decía "152 kg los vendió alguien que ya no compite" y esa frase se
    fue el 25/08/2026 con el caso. El `continue` queda como red y, si volviera
    a pasar, lo avisa la alarma `competencia_vendedor_ajeno` del health."""
    c = _competencia_falsa(monkeypatch, vendido=[
        {"vendedor": "Bedon Hector", "categoria": "Jersey", "kg": 80, "ultima": None}])
    assert "kg_fuera_del_ranking" not in c, (
        "volvió el número que sostenía una frase de un caso ya arreglado")
    assert all(r["kg"] == 0 for r in c["ranking"])
    jersey = next(g for g in c["grupos"] if g["grupo"] == "Jersey")
    assert jersey["liquidado"] == 80, (
        "el kilo salió de la bodega igual: tiene que contar para el grupo")


def test_la_competencia_sale_de_las_mismas_filas_que_lo_parado():
    """Si saliera de otra consulta, el termómetro de acá y el total de allá
    podrían no coincidir el mismo día."""
    import inspect
    assert "items()" in inspect.getsource(queries.competencia)


def test_competencia_muestra_el_al_arrancar_vivo_de_saldos(monkeypatch):
    """⭐ DECISIÓN 01/09/2026. Tamara notó que "Al arrancar" de Saldos (el que
    sube solo cuando entra segunda a una tela vieja, ver `resumen()`) no se
    veía en Competencia — sólo la meta fija (`kg_al_largar`). Pidió mostrarlo
    acá también: *"esta bien que suba"* — no hace falta esconderlo, hace
    falta que las dos pantallas digan lo mismo. `kg_al_arrancar_vivo` sale
    de `resumen()` sobre las MISMAS `filas` que arma `competencia()` (no una
    cuenta aparte), así que es matemáticamente el mismo número que pintaría
    Saldos con esos datos — no una aproximación."""
    c = _competencia_falsa(monkeypatch)
    assert "kg_al_arrancar_vivo" in c
    # Mismo fake DB todavía activo: repetir el cálculo de Saldos sobre las
    # MISMAS filas tiene que dar el mismo número, no una aproximación.
    filas_saldos = queries.con_puntos(queries.items())
    r_saldos = queries.resumen(
        filas_saldos, queries.kg_al_marcar_vivo(filas_saldos),
        largada=c["largada"])
    assert c["kg_al_arrancar_vivo"] == r_saldos["kg_inicial"]


def test_el_vendido_de_arriba_se_mide_contra_el_al_arrancar_vivo():
    """⭐⭐ DECISIÓN 01/09/2026 (tercera vuelta sobre este número, misma
    tarde). Después de agregar el renglón chico con "Al arrancar" vivo AL
    LADO de la meta fija, Tamara: *"no entiendo para que sirve el 50 aca"*
    — con las opciones puestas adelante (sacar la meta y dejar sólo el
    vivo / mantener las dos líneas / una sola línea con los dos números),
    eligió sacar la meta fija de la pantalla entera. El headline "Se
    vendió" y el % avance pasan a medirse contra `kg_al_arrancar_vivo`, no
    contra `kg_al_largar` — un solo número en toda la pantalla, aunque eso
    signifique que el % ya no es perfectamente estable (puede bajar un
    poco si entra segunda nueva más rápido de lo que se vende)."""
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "competencia.html").read_text(
                encoding="utf-8")
    assert "100 * liquidado / kg_al_arrancar_vivo" in html
    assert "de {{ kg_al_arrancar_vivo | num_es(0) }} kg" in html
    assert "kg_al_largar" not in html, (
        "la meta fija ya no se muestra en esta pantalla")


def test_la_pantalla_de_competencia_muestra_el_al_arrancar_vivo():
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "competencia.html").read_text(
                encoding="utf-8")
    assert "kg_al_arrancar_vivo" in html


# ── El candado, con una ruta nueva abierta ──────────────────────────────────

def test_la_competencia_esta_abierta_a_todos():
    """Dueña: "aca tienen acceso todos, vendedores sobre todo incluidos"."""
    from modules.analisis import views
    assert getattr(views.competencia, "_permiso", None) is None, (
        "la competencia no lleva gate de permiso a propósito")


def test_el_vendedor_llega_a_la_competencia_desde_su_portal():
    """El link vive en la barra de abajo de /mi-cartera, que es lo único que el
    vendedor ve. Dice "Competencia" y lleva al tablero: el rótulo y el destino
    son la misma cosa (dueña 25/08/2026: "que el boton diga competencia no
    saldos"). Sin el link, la pantalla abierta no la encuentra nadie."""
    from pathlib import Path
    barra = Path("modules/mi_cartera/templates/mi_cartera/base.html").read_text()
    assert 'href="/analisis/competencia' in barra
    assert ">Competencia</a>" in barra


def test_al_vendedor_se_le_abrio_la_competencia_y_nada_mas():
    """⭐ Dueña 25/08/2026, el día de la largada: "abrilo para vendedores". La
    pantalla estuvo hecha y frenada desde el 17/08.

    ⚠ Lo que NO puede pasar nunca es que se abra /analisis a secas o
    /analisis/parado: ahí está la cartera de TODOS los vendedores."""
    import scope_vendedor
    abiertos = [p for p in scope_vendedor.PREFIJOS_PERMITIDOS
                if p.startswith("/analisis")]
    assert abiertos == ["/analisis/competencia"], (
        f"la Competencia y nada más de /analisis: {abiertos}")


def test_si_algun_dia_se_abre_que_sea_solo_la_competencia():
    """El guard del guard: el día que se habilite, que no se cuele /analisis
    entero de un copy-paste."""
    import scope_vendedor
    for p in scope_vendedor.PREFIJOS_PERMITIDOS:
        assert p not in ("/analisis", "/analisis/parado"), (
            f"{p} le abre al vendedor la cartera de los otros cinco")


def test_ningun_porcentaje_sale_con_quince_decimales():
    """⚠ En Jinja `x if y else 0 | round(1)` aplica el filtro SÓLO al 0: el
    termómetro mostraba "17.284974824441832%". Además con punto decimal, que en
    esta app es separador de miles."""
    import re
    from pathlib import Path
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    for archivo in carpeta.glob("*.html"):
        html = archivo.read_text(encoding="utf-8")
        for expr in re.findall(r"\{\{([^}]*%)", html):
            if "if" in expr and "|" in expr:
                antes_del_filtro = expr.split("|")[0]
                assert antes_del_filtro.strip().startswith("("), (
                    f"{archivo.name}: '{expr.strip()}' — el filtro se aplica "
                    f"sólo a la última rama del if; faltan paréntesis")


def test_todos_los_porcentajes_llevan_un_decimal():
    """Dueña 17/08/2026: "un decimal". Con enteros, un grupo del 0,4% se lee
    como 0% y parece que no existe; con dos, la columna deja de leerse de un
    vistazo."""
    import re
    from pathlib import Path
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    for archivo in carpeta.glob("*.html"):
        html = archivo.read_text(encoding="utf-8")
        for decimales in re.findall(r"num_es\((\d)\)\s*\}\}%", html):
            assert decimales == "1", (
                f"{archivo.name}: un porcentaje con {decimales} decimales")


# ── Lo que ve un vendedor ───────────────────────────────────────────────────

def test_la_cartera_del_vendedor_sale_de_programa_core_y_no_de_asinfo():
    """⚠ `cliente.vend` (Programa Core) y el vendedor de la última factura de
    Asinfo no siempre coinciden. Si acá usara el de Asinfo, un cliente podría
    salirle a uno en esta pantalla y a otro en /mi-cartera, y eso no hay forma
    de explicárselo a nadie."""
    import inspect
    fuente = inspect.getsource(queries.mis_clientes_parado)
    assert "scintela.cliente" in fuente and "c.vend" in queries._ES_MI_CLIENTE
    assert "vend_pc" not in fuente, "vend_pc es el de Asinfo, no la cartera de PC"


def test_el_predicado_de_pertenencia_se_importa_del_portal():
    """⭐ No se compara: se IMPORTA. Escrito a mano acá, el día que en
    /mi-cartera cambien el criterio las dos pantallas le mostrarían al vendedor
    carteras distintas y nadie se enteraría."""
    import inspect

    from modules.mi_cartera import queries as mc
    assert mc._ES_MI_CLIENTE.replace(
        "%(vend)s", "%(cartera)s") == queries._ES_MI_CLIENTE
    fuente = inspect.getsource(queries)
    assert "from modules.mi_cartera.queries import" in fuente, (
        "tiene que venir por import, no copiado")


def test_sin_vendedor_no_devuelve_la_lista_entera(monkeypatch):
    """⚠ Un scope que falla ABIERTO no da error: muestra de más y nadie se
    entera. Con `vend` vacío tiene que devolver [], no todo."""
    llamado = {"veces": 0}

    def fake(*a, **k):
        llamado["veces"] += 1
        return [{"codigo_cli": "AAA"}]

    monkeypatch.setattr(queries.db, "fetch_all", fake)
    assert queries.mis_clientes_parado("") == []
    assert queries.mis_clientes_parado(None) == []
    assert llamado["veces"] == 0, "ni siquiera tiene que consultar"


def test_el_vendedor_del_bloque_sale_del_usuario_y_no_de_la_url():
    """Si viniera del querystring, cualquiera vería la cartera de cualquiera
    cambiando tres letras. La vista lo pide por `_vend_actual()`, que sólo mira
    la URL cuando el usuario es wildcard (preview de la dueña)."""
    import inspect

    from modules.analisis import views
    fuente = inspect.getsource(views.competencia)
    assert "_vend_actual()" in fuente
    assert "request.args" not in fuente, (
        "la vista no lee la URL: eso lo decide _vend_actual")


# ── Semana a semana ─────────────────────────────────────────────────────────

def test_el_acumulado_de_la_semana_mas_nueva_es_el_total(monkeypatch):
    from datetime import date as _d
    c = _competencia_falsa(monkeypatch, semanas=[
        {"semana": _d(2026, 8, 17), "vendedor": "Intela", "kg": 100},
        {"semana": _d(2026, 8, 10), "vendedor": "Intela", "kg": 40},
    ])
    filas = c["semanas"]
    assert [f["semana"] for f in filas] == [_d(2026, 8, 17), _d(2026, 8, 10)], \
        "la semana más nueva va arriba"
    assert filas[0]["acumulado"] == 140 and filas[1]["acumulado"] == 40


def test_el_movimiento_del_ranking_se_recalcula_sin_guardar_nada(monkeypatch):
    """El puesto de la semana pasada sale de descontar lo de esta semana. Sin
    esto habría que guardar un ranking por semana para poder decir "subió"."""
    from datetime import date as _d
    c = _competencia_falsa(
        monkeypatch,
        vendido=[{"vendedor": "Quintero Jose", "categoria": "Jersey", "kg": 500,
                  "ultima": None}],
        semanas=[{"semana": _d(2026, 8, 17), "vendedor": "Quintero Jose",
                  "kg": 500, "puntos": 500}])
    quintero = next(r for r in c["ranking"] if r["vendedor"] == "Quintero Jose")
    assert quintero["puesto"] == 1
    assert quintero["kg_semana"] == 500
    assert quintero["movimiento"] > 0, "venía último y pasó a primero"


def test_los_tres_numeros_de_arriba_cierran_entre_ellos(monkeypatch):
    """"Había" no se guarda: se reconstruye como lo que hay más lo que se
    vendió. Así los tres números cierran siempre, que es lo primero que alguien
    va a chequear de un vistazo."""
    c = _competencia_falsa(monkeypatch, vendido=[
        {"vendedor": "Intela", "categoria": "Jersey", "kg": 250, "ultima": None}])
    assert c["kg_al_largar"] == c["kg_parado"] + c["liquidado"]


def test_no_se_cuentan_las_ventas_anteriores_a_la_largada():
    """La cohorte se marcó el 13/08 y la competencia arranca el 17: sin el
    corte, esos cuatro días le regalarían kilos a quien justo vendió algo."""
    import inspect
    fuente = inspect.getsource(queries.competencia)
    assert "WHERE v.fecha >= %s" in fuente
    assert 'config("largada"' in fuente


def test_la_fila_fantasma_sin_grupo_no_se_dibuja(monkeypatch):
    """Restos de la cohorte que ya no están en la foto (la tela cruda que se
    sacó) sumaban un renglón "(sin grupo) · 0 kg" que sólo confunde."""
    import inspect
    assert 'if float(f["stock_kg"]) > 0 or f["categoria"]' in \
        inspect.getsource(queries.competencia)


def test_las_telas_a_sacar_no_llevan_un_solo_cliente_adentro():
    """Es lo que le faltaba al vendedor: la pantalla le decía quién le compró
    qué, pero no cuántos kilos hay ni de qué colores. Sin clientes adentro, la
    lista es información de fábrica y la puede ver cualquiera."""
    t = queries.telas_a_sacar([
        {"subcategoria": "Kiana", "categoria": "Poliester", "color": "CAR",
         "stock_kg": 100},
        {"subcategoria": "Kiana", "categoria": "Poliester", "color": "LIF",
         "stock_kg": 300},
        {"subcategoria": "Jersey 3", "categoria": "Jersey", "color": "NEG",
         "stock_kg": 50},
    ])
    assert [x["tela"] for x in t] == ["Kiana", "Jersey 3"], "de mayor a menor"
    assert t[0]["kg"] == 400 and t[0]["n_colores"] == 2
    assert t[0]["colores"] == "LIF, CAR", "los colores van por kilos"
    assert [c["cod"] for c in t[0]["colores_lista"]] == ["LIF", "CAR"], \
        "la pantalla los recibe uno por uno, para poner el nombre al lado"
    assert all("cliente" not in k for x in t for k in x)


def test_las_telas_sin_stock_no_se_listan():
    assert queries.telas_a_sacar([
        {"subcategoria": "X", "categoria": "Y", "color": "Z", "stock_kg": 0}]) == []


def test_la_hoja_del_vendedor_usa_la_plantilla_de_la_oficina():
    """⚠ Dos plantillas para el mismo papel divergen a la primera corrección
    que se le hace a una sola. Ya pasó en /mi-cartera: ocho días imprimiendo
    órdenes distintos sin síntoma."""
    import inspect

    from modules.analisis import views
    fuente = inspect.getsource(views.mi_hoja)
    assert '"analisis/parado_clientes.html"' in fuente
    assert "cartera_de=vend" in fuente, (
        "la hoja propia se acota por la cartera de Programa Core")


def test_la_hoja_propia_no_ofrece_el_selector_de_otros_vendedores():
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado_clientes.html").read_text(encoding="utf-8")
    assert "{% if not mia %}" in html


def test_un_vendedor_no_puede_pedir_la_cartera_de_otro_por_la_url():
    """`?vend=` sólo lo respeta un usuario wildcard, para previsualizar. Si el
    que lo manda ES vendedor, gana el suyo."""
    import inspect

    from modules.analisis import views
    fuente = inspect.getsource(views._vend_actual)
    assert 'propio = (g.user or {}).get("vend")' in fuente
    assert "if propio:" in fuente and "return propio" in fuente
    assert '"*" in (g.get("permisos")' in fuente


def test_la_hoja_general_no_se_le_abre_nunca_al_vendedor():
    """/analisis/parado/clientes tiene los clientes de TODOS. La hoja del
    vendedor cuelga de /analisis/competencia/ justamente para no tener que
    abrirle esa."""
    import scope_vendedor
    for p in scope_vendedor.PREFIJOS_PERMITIDOS:
        assert not p.startswith("/analisis/parado")


def test_la_hoja_propia_baja_a_excel_por_su_propia_ruta():
    """Dueña: "de todos lados deberia poder bajar a excel". El CSV de la hoja
    propia cuelga del MISMO camino que la pantalla (…/mi-hoja.csv) para que
    quede dentro del mismo permiso del allowlist, y va recortado por cartera."""
    import inspect

    from modules.analisis import views
    fuente = inspect.getsource(views.mi_hoja_csv)
    assert "cartera_de=vend" in fuente
    assert "/analisis/competencia/mi-hoja.csv" in inspect.getsource(views)


def test_el_csv_de_la_hoja_propia_no_lleva_la_columna_del_vendedor():
    """En la hoja propia son todos suyos: una columna "Vendedor" con el mismo
    valor repetido no dice nada y ocupa lugar."""
    import inspect

    from modules.analisis import views
    assert '"Vendedor"' not in inspect.getsource(views.mi_hoja_csv)


# ── El tope por grupo ───────────────────────────────────────────────────────

def test_un_kilo_dificil_vale_mas_que_uno_facil(monkeypatch):
    """⭐ Dueña 24/08/2026, después de medir las 98 telas: "si una tela es
    fácil de vender deberíamos dar menos puntos". Dos vendedores que sacan los
    MISMOS kilos no van empatados si uno sacó lo que sale solo y el otro lo que
    no salió en un año."""
    c = _competencia_falsa(
        monkeypatch,
        puntos={"Jersey 3": 10, "Fleece 102": 1},
        vendido=[
            {"vendedor": "Intela", "categoria": "Jersey", "kg": 500,
             "ultima": None},
            {"vendedor": "Lopez Felipe", "categoria": "Fleece", "kg": 500,
             "ultima": None},
        ])
    intela = next(r for r in c["ranking"] if r["vendedor"] == "Intela")
    lopez = next(r for r in c["ranking"] if r["vendedor"] == "Lopez Felipe")
    assert intela["kg"] == lopez["kg"] == 500, "los mismos kilos"
    assert intela["puntos"] == 5000 and lopez["puntos"] == 500
    assert intela["puesto"] < lopez["puesto"], (
        "el que sacó lo difícil tiene que ir adelante")


def test_la_meta_en_puntos_sale_de_las_telas_y_no_del_grupo(monkeypatch):
    """⚠ La bolsa de un grupo NO se puede sacar de sus kilos: adentro de un
    mismo grupo conviven telas de 1 punto y de 10."""
    c = _competencia_falsa(monkeypatch,
                           puntos={"Jersey 3": 10, "Fleece 102": 1})
    jersey = next(g for g in c["grupos"] if g["grupo"] == "Jersey")
    fleece = next(g for g in c["grupos"] if g["grupo"] == "Fleece")
    assert jersey["puntos_base"] == 60000    # 6.000 kg × 10
    assert fleece["puntos_base"] == 4000     # 4.000 kg × 1
    assert c["puntos_en_juego"] == 64000
    assert jersey["kg"] < fleece["kg"] * 2 < jersey["puntos_base"], (
        "un grupo puede tener pocos kilos y muchos puntos")


def test_ya_no_hay_tope_por_grupo(monkeypatch):
    """⭐ El tope existía para obligar a tocar los ocho grupos. Los puntos hacen
    ese trabajo sin tener que explicar un tope: adentro de un grupo hay telas
    que salen solas y telas que no salió una en un año, y el tope las igualaba
    a todas. Dueña 24/08/2026."""
    c = _competencia_falsa(monkeypatch, vendido=[
        {"vendedor": "Intela", "categoria": "Jersey", "kg": 5000, "ultima": None}])
    intela = next(r for r in c["ranking"] if r["vendedor"] == "Intela")
    jersey = next(d for d in intela["detalle"] if d["grupo"] == "Jersey")
    assert jersey["puntos"] == 5000, "el kilo cuenta entero, no cortado"
    assert "contado" not in intela, "no queda rastro del tope"
    assert "meta" not in jersey, "ni de la meta por grupo"


def test_el_kilo_igual_sale_de_la_bodega(monkeypatch):
    """El termómetro de la fábrica sigue siendo de KILOS: es la bodega, no el
    marcador."""
    c = _competencia_falsa(monkeypatch, vendido=[
        {"vendedor": "Intela", "categoria": "Jersey", "kg": 5000, "ultima": None}])
    jersey = next(g for g in c["grupos"] if g["grupo"] == "Jersey")
    assert jersey["liquidado"] == 5000
    assert c["liquidado"] == 5000


def test_los_puntos_de_un_vendedor_son_la_suma_de_sus_grupos(monkeypatch):
    c = _competencia_falsa(
        monkeypatch,
        puntos={"Jersey 3": 10, "Fleece 102": 1},
        vendido=[
            {"vendedor": "Intela", "categoria": "Jersey", "kg": 100,
             "ultima": None},
            {"vendedor": "Intela", "categoria": "Fleece", "kg": 200,
             "ultima": None},
        ])
    intela = next(r for r in c["ranking"] if r["vendedor"] == "Intela")
    assert intela["kg"] == 300
    assert intela["puntos"] == 1200          # 100×10 + 200×1
    assert sum(d["puntos"] for d in intela["detalle"]) == intela["puntos"]


def test_el_ranking_ordena_por_puntos_y_no_por_kilos(monkeypatch):
    """Uno con 5.000 kg de tela fácil tiene que ir DETRÁS de otro con menos
    kilos pero de la difícil."""
    c = _competencia_falsa(
        monkeypatch,
        puntos={"Jersey 3": 1, "Fleece 102": 10},
        vendido=[
            {"vendedor": "Intela", "categoria": "Jersey", "kg": 5000,
             "ultima": None},
            {"vendedor": "Lopez Felipe", "categoria": "Fleece", "kg": 800,
             "ultima": None},
        ])
    puestos = {r["vendedor"]: r["puesto"] for r in c["ranking"]}
    assert puestos["Lopez Felipe"] < puestos["Intela"]


def test_la_pantalla_explica_las_reglas():
    """Dueña: "escribamos las reglas del juego. facil". Una competencia cuyas
    reglas hay que preguntar no la juega nadie."""
    # ⚠ Dos cuidados para que el test mire lo que se VE:
    #   · se sacan los comentarios Jinja ({# … #}), donde justamente se explica
    #     por qué NO se dice cierta frase — si no, el test se engancha con su
    #     propia explicación;
    #   · se normalizan los espacios, porque el texto viene cortado en varias
    #     líneas y si no falla por el formato del HTML y no por lo que dice.
    import re as _re
    from pathlib import Path
    crudo = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "competencia.html").read_text(encoding="utf-8")
    html = " ".join(_re.sub(r"\{#.*?#\}", " ", crudo, flags=_re.S).split())
    assert "Las reglas" in html
    assert "No todos los kilos valen igual" in html, (
        "la regla del puntaje tiene que estar escrita")
    assert "meses de venta hay" in html, (
        "y con qué se mide difícil, o el puntaje parece arbitrario")
    assert "el que más puntos hace" in html
    assert "No hay meta" in html, (
        "que no haya meta es una regla, no una omisión (dueña 24/08/2026)")
    assert "no por kilos" not in html, (
        "todo se mide EN kilos; decir que no, confunde (dueña 17/08/2026)")


def test_la_fila_del_vendedor_se_abre_y_muestra_sus_grupos():
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "competencia.html").read_text(encoding="utf-8")
    assert "abrirVend" in html and 'class="vdet"' in html
    for col in ("Vendió", "Puntos"):
        assert col in html
    assert "Su meta" not in html, "la meta se fue de la pantalla"


def test_la_lista_de_telas_muestra_lo_que_vale_cada_una():
    """⚠ `telas_a_sacar` recibe los puntos por parámetro: si los buscara sola,
    sería una segunda lectura que puede contradecir a la del tablero. La
    pantalla tiene que pasárselos."""
    import inspect

    from modules.analisis import views
    fuente = inspect.getsource(views.competencia)
    assert "puntos_por_tela()" in fuente, (
        "la pantalla tiene que pasarle los puntos a la lista de telas")
    t = queries.telas_a_sacar(
        [{"subcategoria": "Kiana", "categoria": "Poliester", "color": "CAR",
          "stock_kg": 100},
         {"subcategoria": "Jersey 3", "categoria": "Jersey", "color": "NEG",
          "stock_kg": 300}],
        {"Kiana": {"puntos": 10, "nivel_nombre": "Difícil"},
         "Jersey 3": {"puntos": 1, "nivel_nombre": "Fácil"}})
    assert [x["tela"] for x in t] == ["Kiana", "Jersey 3"], (
        "ordena por puntos: 100 kg × 10 vale más que 300 × 1")
    assert t[0]["puntos"] == 10 and t[0]["puntos_total"] == 1000
    assert t[1]["puntos_total"] == 300


# ── La copia de "Lo parado" para el vendedor ────────────────────────────────

def test_el_vendedor_ve_la_misma_pantalla_con_sus_clientes():
    """Dueña 17/08/2026: "la tab de que hay que sacar copiala para ellos. y que
    vean con sus clientes… estaba linda diseñada". Es el MISMO template: lo
    único que cambia es de dónde salen los candidatos."""
    import inspect

    from modules.analisis import views
    fuente = inspect.getsource(views.mis_telas)
    assert '"analisis/parado.html"' in fuente
    assert "cartera_de=vend" in fuente
    assert 'f["clientes"] = len(llamados.get(f["subcategoria"], []))' in fuente, (
        "la columna tiene que contar SUS clientes: con el total de la fábrica "
        "diría 137 y al abrir la fila aparecerían tres")


def test_el_excel_del_vendedor_cuenta_lo_mismo_que_su_pantalla():
    """⭐ Reportado 27/08/2026: "no pueden bajar bien el excel los vendedores".
    El botón de la pantalla compartida apuntaba a /analisis/parado.xlsx —la
    ruta de la oficina— y el allowlist se la 404eaba. La ruta del vendedor
    cuelga de /analisis/competencia y baja LO MISMO que su pantalla: sus
    clientes, el vendido incluido, el ajuste de bodega afuera."""
    import inspect

    from modules.analisis import views
    fuente = inspect.getsource(views.mis_telas_xlsx)
    # el mismo pipeline que mis_telas — si divergen, el archivo y la pantalla
    # dejan de contar lo mismo
    assert "cartera_de=vend" in fuente
    assert 'f["clientes"] = len(llamados.get(f["subcategoria"], []))' in fuente
    assert "queries.abrir_en_lineas(" in fuente and "con_puntos" in fuente
    assert "_excel_saldos(" in fuente, "el archivo se arma en UN solo lugar"
    assert '"Sus clientes' in fuente, (
        "la columna dice de quién son los clientes: con el rótulo de la "
        "fábrica, 3 parece un error al lado de los 137 de la oficina")
    # ⚠ sin requiere_permiso: el rol Vendedor no tiene analisis.ver — igual
    # que mis_telas, lo cierra el allowlist
    assert "requiere_permiso" not in fuente


def test_el_vend_del_excel_sale_de_la_sesion():
    """El ?vend= del botón es sólo para la preview de la dueña: la ruta usa
    _vend_actual(), que a un vendedor real le ignora el querystring."""
    import inspect

    from modules.analisis import views
    assert "_vend_actual()" in inspect.getsource(views.mis_telas_xlsx)


def test_los_candidatos_se_pueden_acotar_a_una_cartera(monkeypatch):
    visto = {}

    def fake(sql, params=None, conn=None):
        visto["sql"] = " ".join(sql.split())
        visto["params"] = params
        return []

    monkeypatch.setattr(queries.db, "fetch_all", fake)
    queries.llamados_por_tela(cartera_de="FL1")
    assert "scintela.cliente" in visto["sql"] and "c.vend" in visto["sql"]
    assert visto["params"] == {"cartera": "FL1"}

    queries.llamados_por_tela()
    assert "scintela.cliente" not in visto["sql"], (
        "sin cartera no se filtra: es la pantalla de la oficina")


def test_el_menu_del_vendedor_no_tiene_links_que_le_dan_404(app):
    """Sus pantallas son OTRAS rutas, no un subconjunto de las de la oficina:
    mostrarle el menú de la oficina serían tres links a 404."""
    from modules.analisis import views
    rutas = {r.rule for r in app.url_map.iter_rules()}
    for m in views.MENU_VENDEDOR:
        assert m["url"] in rutas
        assert m["url"].startswith("/analisis/competencia"), (
            f"{m['url']} queda fuera del prefijo que se le habilita")


def test_nadie_tiene_el_boton_de_actualizar_desde_asinfo():
    """⭐ Dueña 25/08/2026: "borra tambien el boton de actualizar desde asinfo".

    Ya no hace falta: el hilo de fondo refresca la foto sola cada 3 horas y la
    bajada dice de cuándo son los datos. Un botón que casi nadie tiene que
    apretar, metido entre los filtros, se aprieta sin querer y se lleva 40
    segundos de espera contra Asinfo.

    ⚠ La RUTA sigue viva: es la salida de emergencia si Asinfo estuvo caído y
    no se puede esperar tres horas."""
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")
    assert "<button type=\"submit\">Actualizar desde Asinfo" not in html
    assert "auto_refresco" in html, "la plantilla dice quién refresca ahora"

    # La salida de emergencia queda: la ruta sigue registrada.
    import inspect as _i
    assert '@analisis_bp.route("/analisis/parado/actualizar"' in _i.getsource(views)


def test_la_fila_sin_stock_y_sin_grupo_no_ensucia_el_resumen():
    """Restos que ya no existen sumaban un renglón "(sin grupo) · 0 kg"."""
    g = queries.por_grupo([
        {"categoria": None, "subcategoria": "X", "stock_kg": 0, "kg_segunda": 0},
        {"categoria": "Jersey", "subcategoria": "Y", "stock_kg": 100,
         "kg_segunda": 0},
    ])
    assert [x["grupo"] for x in g] == ["Jersey"]


def test_un_item_vendido_entero_SI_sigue_en_el_resumen():
    """⚠ 0 kg y CON grupo es un ítem que se vendió entero: es justamente lo que
    la dueña pidió que no desaparezca."""
    g = queries.por_grupo([
        {"categoria": "Jersey", "subcategoria": "Y", "stock_kg": 0,
         "kg_segunda": 0}])
    assert [x["grupo"] for x in g] == ["Jersey"] and g[0]["n_items"] == 1


def test_la_competencia_tiene_fecha_de_cierre(monkeypatch):
    """Dueña 17/08/2026: "es una competencia hasta fin de año". La fecha vive en
    `parado_config` y no hardcodeada, porque la presentación que se les da a los
    vendedores dice la misma: si un día se corre, tiene que cambiar en UN lugar
    y no en dos que se olvidan de coincidir."""
    from datetime import date as _d
    c = _competencia_falsa(monkeypatch)
    assert c["cierre"] == _d(2026, 12, 31)
    assert c["dias_para_el_cierre"] == (_d(2026, 12, 31) - c["hoy"]).days


def test_la_pantalla_dice_cuanto_falta_para_el_cierre():
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "competencia.html").read_text(encoding="utf-8")
    assert "dias_para_el_cierre" in html
    assert "Terminó el" in html, "cuando pase la fecha tiene que decirlo"


def test_la_largada_es_el_martes_25(monkeypatch):
    """Dueña 18/08/2026: "la vamos a empezar el martes que viene". Antes decía
    17/08 y habría contado kilos de una semana antes de que los vendedores
    supieran que la competencia existía."""
    from datetime import date as _d
    c = _competencia_falsa(monkeypatch)
    assert c["largada"] == _d(2026, 8, 25)


# ── Que se pueda mirar en el celular ────────────────────────────────────────

def test_las_pantallas_esconden_columnas_en_el_celular():
    """Dueña 18/08/2026: "los vendedores tienen un distinto portal, tmb lo ven
    asi?". No: ellos andan en el teléfono, y esto nació como tablas anchas de
    escritorio. En pantalla chica hay que SACAR columnas, no achicarlas: una
    tabla de diez columnas a 390 px no se lee ni con lupa."""
    from pathlib import Path
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    base = (carpeta / "base.html").read_text(encoding="utf-8")
    assert "@media (max-width: 780px)" in base
    assert ".opt{display:none!important}" in base, (
        "sin !important pierde contra cualquier display del cascarón")
    for archivo in ("competencia.html", "parado.html"):
        assert 'opt' in (carpeta / archivo).read_text(encoding="utf-8")


def test_en_el_celular_no_se_esconde_lo_que_identifica_la_fila():
    """Las columnas que se sacan son las de CONTEXTO. Si desapareciera el
    vendedor, la tela o el %, la tabla dejaría de decir nada."""
    import re
    from pathlib import Path
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    comp = " ".join((carpeta / "competencia.html").read_text(encoding="utf-8").split())
    for imprescindible in (">Vendedor<", ">Puntos<", ">Cliente<"):
        col = re.search(r'<th[^>]*' + re.escape(imprescindible), comp)
        assert col, f"falta la columna {imprescindible}"
        assert "opt" not in col.group(0), (
            f"{imprescindible} no puede esconderse en el celular")


def test_las_columnas_se_esconden_por_clase_y_no_por_posicion():
    """Escondiendo `td:nth-child(5)`, una columna nueva en el medio correría
    todas y el que desaparecería sería otro — sin ningún error.

    ⚠ `nth-child` SÍ se usa para el rayado de filas, que no depende de qué
    columna es: por eso el test mira sólo las reglas que esconden algo."""
    import re
    from pathlib import Path
    base = ((Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "base.html").read_text(encoding="utf-8"))
    for regla in re.findall(r"([^{}]+)\{([^}]*display\s*:\s*none[^}]*)\}", base):
        assert "nth-child" not in regla[0], (
            f"esconde por posición: {regla[0].strip()}")


def test_en_el_celular_la_calidad_viaja_pegada_a_la_tela():
    """Con la columna Categoría la tabla mide 563 px y el teléfono 390: los
    kilos se cortaban afuera. Abajo de 560 la columna se esconde y la píldora
    va al lado del nombre. Y la copia NO puede ensuciar ni el orden ni el
    Excel — la columna Tela decía "Jersey Fancy PRI"."""
    from pathlib import Path
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    base = (carpeta / "base.html").read_text(encoding="utf-8")
    parado = (carpeta / "parado.html").read_text(encoding="utf-8")

    assert ".qm{display:none}" in base, "la copia se ve también en la pantalla grande"
    assert "#tabla .cal{display:none!important}" in base, (
        "la columna Categoría no se esconde en el celular")
    # ⚠ `block`, no `inline`: al lado del nombre la columna medía 244 px y la
    # tabla se salía igual de la pantalla. Debajo, vale lo que el nombre.
    assert ".qm{display:block" in base, "en el celular la píldora no aparece"
    # en el teléfono chico se achica la letra, NO se parte el nombre
    assert "@media (max-width: 380px)" in base
    assert "#tabla{font-size:11px}" in base

    # ⚠ Se chequea la CLASE, no el número de columna: el 25/08/2026 entró la
    # columna Forma antes que Categoría y el data-i corrió. Lo que importa es
    # que el encabezado lleve `cal`, que es lo que lo esconde en el celular
    # junto con sus celdas.
    assert re.search(r'class="ord cal" data-i="\d+"', parado), (
        "el encabezado de Categoría no se esconde con su columna")
    assert re.search(r'class="ord forma" data-i="\d+"', parado), (
        "y la Forma tiene la suya")
    assert '<span class="qm">{{ calidad }}</span></td>' in parado
    # ⚠ Una sola vez se DECIDE PRI/SEG. Antes esto contaba las píldoras; desde
    # el 25/08/2026 el bloque tiene una rama más (la línea abierta por calidad
    # trae la suya), así que lo que se fija es lo de verdad: que la decisión
    # viva en UN `{% set calidad %}` y que los dos lugares donde se dibuja la
    # usen. Dos copias del if se despegan a la primera corrección.
    assert parado.count("{% set calidad %}") == 1, (
        "las píldoras se arman en más de un lugar")
    assert parado.count("{{ calidad }}") == 2, (
        "la píldora se dibuja en la columna y, en el celular, pegada a la tela")
    # ⚠ El ORDEN lee el texto sin la copia. (El Excel ya no: desde el
    # 26/08/2026 lo arma el servidor desde la base, así que ni ve el HTML.)
    assert "c.querySelectorAll('.qm').forEach(e => e.remove())" in parado
    assert "return texto(td).trim().toLowerCase();" in parado


def test_los_kilos_congelados_no_se_mueven_cuando_se_mueve_el_stock(monkeypatch):
    """Dueña 18/08/2026, después de ver el número moverse solo: el mismo día,
    sin una sola venta, «había al arrancar» pasó de 52.407 a 51.654 kg —753 kg
    de ajustes de bodega—. Congelados, el stock de hoy puede hacer lo que
    quiera: contra qué se mide sale de los kilos del día de la largada.

    ⚠ Se llamaba `meta_kg` / `meta_fijada_el`. Desde el 25/08/2026 no hay metas
    y el nombre mentía: es la BASE, no una meta."""
    congelada = {"Jersey": 6000, "Fleece": 4000}
    c = _competencia_falsa(monkeypatch, base=congelada)
    assert c["kg_al_largar"] == 10000
    assert c["fijada_el"] == date(2026, 8, 25)

    # ahora la bodega dice otra cosa —un ajuste, no una venta— y la base NO se
    # entera: sigue valiendo lo mismo
    c2 = _competencia_falsa(monkeypatch, base={"Jersey": 6000, "Fleece": 3000})
    assert c2["kg_al_largar"] == 9000       # sólo cambia si cambia la BASE
    assert c["kg_al_largar"] != c2["kg_al_largar"]
    # …y sin base, la pantalla es una previa que se calcula con lo de hoy
    previa = _competencia_falsa(monkeypatch)
    assert previa["fijada_el"] is None
    assert previa["kg_al_largar"] == 10000


def test_un_grupo_despejado_entero_no_se_cae_de_la_tabla(monkeypatch):
    """Si un grupo se vende del todo deja de estar en la foto. Sin esto se caía
    de la tabla y la base total se achicaba justo cuando alguien había hecho
    bien el trabajo — el que lo despejó perdía el puntaje."""
    c = _competencia_falsa(monkeypatch,
                           base={"Jersey": 6000, "Fleece": 4000, "Lycra": 2000})
    grupos = {g["grupo"]: g for g in c["grupos"]}
    assert "Lycra" in grupos, "el grupo despejado se cayó de la tabla"
    assert grupos["Lycra"]["kg_base"] == 2000
    assert grupos["Lycra"]["kg"] == 0
    assert c["kg_al_largar"] == 12000


def test_la_base_se_fija_una_sola_vez_y_recien_desde_la_largada(monkeypatch):
    """Se escribe en el primer refresco del día de la largada o después. Si se
    escribiera antes, congelaría kilos de una semana en la que nadie estaba
    compitiendo; si se reescribiera, la meta volvería a moverse —que es
    justamente lo que se vino a arreglar."""
    escritos = []

    def fake_execute(sql, params=None, conn=None):
        if "parado_base" in " ".join(sql.split()):
            escritos.append(params)

    monkeypatch.setattr(queries.db, "execute", fake_execute)
    monkeypatch.setattr(queries, "config", lambda k, d=None: "2026-08-25")
    monkeypatch.setattr(queries, "items", lambda conn=None: [
        {"categoria": "Jersey", "stock_kg": 100, "kg_vendidos": 40},
        {"categoria": "Jersey", "stock_kg": 10, "kg_vendidos": 0},
        {"categoria": None, "stock_kg": 5, "kg_vendidos": 0}])

    # antes de la largada no escribe nada
    monkeypatch.setattr(queries, "today_ec", lambda: date(2026, 8, 24))
    monkeypatch.setattr(queries.db, "fetch_all", lambda *a, **k: [])
    assert queries._fijar_base() is None
    assert escritos == []

    # el día de la largada sí, y con stock + lo ya vendido (110 + 40)
    monkeypatch.setattr(queries, "today_ec", lambda: date(2026, 8, 25))
    base = queries._fijar_base()
    assert base == {"Jersey": 150}, base
    assert escritos == [("Jersey", 150.0, date(2026, 8, 25))]

    # ya fijada: no la vuelve a tocar
    escritos.clear()
    monkeypatch.setattr(queries.db, "fetch_all",
                        lambda *a, **k: [{"categoria": "Jersey"}])
    assert queries._fijar_base() is None
    assert escritos == []


def test_el_desplegable_va_por_codigo_y_el_nombre_no_se_pierde():
    """Dueña 19/08/2026: "cliente podemos poner solo codigo, no hace falta el
    nombre?". Con el nombre, una fila del desplegable ocupaba cuatro renglones
    en el celular y la tabla no entraba en 390 px.

    Lo que NO puede pasar es que el nombre desaparezca del todo: sigue en el
    `title` del código y en `data-buscar`, así que buscar "ARELLANO" sigue
    encontrando la fila aunque en pantalla diga WFA."""
    from pathlib import Path
    parado = ((Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis" / "parado.html")
              .read_text(encoding="utf-8"))
    detalle = parado[parado.index("{% if cand %}"):]
    assert "<th>Cliente</th>" not in detalle, "el desplegable volvió a traer el nombre"
    assert "{{ c.nombre }}" not in detalle.replace('title="{{ c.nombre }}"', ""), (
        "el nombre se imprime como columna")
    assert 'title="{{ c.nombre }}"' in detalle, "el nombre se perdió del todo"
    # el buscador lo sigue teniendo: es lo que hace que se pueda buscar por nombre
    assert "map(attribute='nombre')" in parado

    # la hoja del vendedor SÍ lleva el nombre grande: ahí la unidad es el cliente
    hoja = ((Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "parado_clientes.html")
            .read_text(encoding="utf-8"))
    assert "{{ c.nombre }}" in hoja


def test_en_la_app_del_vendedor_no_se_repite_su_propio_codigo():
    """Dueña 19/08/2026: "el vendedor ya sabe quién es". En `/mi-cartera` de
    saldos sólo ve SUS clientes, así que la columna Vend. decía FL1 en todas
    las filas — una columna entera para no decir nada. En la oficina se queda:
    ahí se mezclan los seis y el mostrador."""
    from pathlib import Path
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    parado = (carpeta / "parado.html").read_text(encoding="utf-8")
    i = parado.index("<th>Provincia</th>")
    assert "{% if not mia %}<th>Vend.</th>{% endif %}" in parado[i:i + 120], (
        "la columna Vend. no está condicionada al modo oficina")
    assert "{% if not mia %}<td class=\"g vnd\">" in parado, (
        "el encabezado se esconde pero la celda no: la fila se corre")

    hoja = (carpeta / "parado_clientes.html").read_text(encoding="utf-8")
    j = hoja.index('<p class="meta">')
    assert "{% if not mia %}" in hoja[j:j + 160], (
        "la hoja propia sigue repitiendo el código del vendedor")


def test_la_pantalla_no_explica_la_formula_vieja_de_metas():
    """La bajada de «Por grupo» quedó describiendo la PRIMERA fórmula —«la meta
    de cada grupo es su propio peso en el parado»—, la que la dueña cazó con
    "¿por qué 8k kilos?". Se cambió el cálculo y el texto se quedó: contradecía
    a la columna Meta de la misma tabla, que dice 100% en todas las filas.

    ⚠ La tabla «Por grupo» se sacó entera el 25/08/2026 ("este por grupo
    borrar"), así que el texto tampoco está. Lo que el test sigue cuidando es
    que la frase vieja no vuelva por ningún lado."""
    from pathlib import Path
    html = ((Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "competencia.html")
            .read_text(encoding="utf-8"))
    import re
    texto = re.sub(r"\{#.*?#\}", " ", html, flags=re.S)
    assert "su propio peso en el parado" not in texto
    assert "<h2>Por grupo</h2>" not in texto, (
        "la tabla Por grupo volvió: la dueña la sacó el 25/08/2026")


def test_la_fecha_del_refresco_se_muestra_en_hora_de_ecuador():
    """`actualizado` se guarda con NOW() y el servidor corre en UTC. Sin el
    filtro, a las 19:41 del 18 en la fábrica la pantalla decía "datos al
    19/08/2026 00:42": mañana. Lo vio la dueña el mismo día que se estrenó."""
    from pathlib import Path
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    for nombre in ("parado.html", "parado_clientes.html"):
        html = (carpeta / nombre).read_text(encoding="utf-8")
        assert "estado.actualizado.strftime" not in html, (
            f"{nombre} muestra la hora del servidor, no la de Ecuador")
        assert "estado.actualizado | hora_ec" in html


def test_la_hoja_en_el_celular_deja_solo_la_tela_y_la_fecha():
    """Dueña 19/08/2026, mirando una ficha en el teléfono: "acá solo el cliente
    y último día que compró, más no hace falta" — y "y chiquito". En la mano,
    la hoja sirve para saber a quién llamar y hace cuánto no compra; los
    colores y las dos cifras de kilos son para armar la recorrida en el
    escritorio o en el papel, y ahí siguen estando."""
    import re
    from pathlib import Path
    hoja = ((Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "parado_clientes.html")
            .read_text(encoding="utf-8"))
    movil = re.search(r"@media screen and \(max-width:\s*\d+px\)\s*\{(.+?)\n\}",
                      hoja, re.S)
    assert movil, "la hoja no tiene bloque de celular"
    css = movil.group(1)
    assert ".cli .col,.cli .hay,.cli .compro{display:none}" in css, (
        "en el celular siguen los colores y los kilos")
    assert ".cli h3 .kg{display:none}" in css, "los kilos siguen en el encabezado"
    # las celdas se agarran por clase, no por posición
    for clase in ("col", "hay", "compro", "ult"):
        assert f'{clase}"' in hoja, f"la celda {clase} no tiene clase"
    # en papel NO cambia nada: el bloque es `screen`
    assert "@media screen and (max-width" in hoja
    imp = hoja[hoja.index("@media print{"):]
    for clase in (".col", ".hay", ".compro"):
        assert f"{clase}{{display:none" not in imp, (
            f"el papel también se está comiendo {clase}")


def test_las_pantallas_tratan_de_usted():
    """Dueña 18/08/2026: "en ecuador se trata de usted". El voseo se cuela de a
    una palabra por vez, así que el test lo caza en cualquier plantilla nueva de
    la sección."""
    import re
    from pathlib import Path
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    # ⚠ Los imperativos van SÓLO con tilde: "toca", "baja" y "deja" sin tilde
    # son formas de usted o de tercera persona y son correctas. Sin esa
    # distinción el test marcaba "le toca" y "Bajar a Excel" como voseo.
    voseo = re.compile(
        r"\b(toc[áí]|baj[áí]|entrás|eleg[íi]|volvé|fijate|tenés|pod[ée]s|"
        r"querés|vos|tuyo|tuyos|tus|llevás|ponés|dejá|mirá|andá|hacé|sacá|"
        r"apretá|revisá|escribí|abrí|and[áa]te)\b", re.I)
    malas = []
    for archivo in carpeta.glob("*.html"):
        # sin comentarios Jinja: ahí se explica el porqué y puede nombrarlo
        texto = re.sub(r"\{#.*?#\}", " ", archivo.read_text(encoding="utf-8"),
                       flags=re.S)
        for m in voseo.finditer(texto):
            malas.append(f"{archivo.name}: {m.group(0)}")
    assert not malas, f"voseo en pantallas que ven los vendedores: {malas}"


def test_las_pantallas_hablan_de_saldos_y_no_de_sacar():
    """Dueña 18/08/2026: "que hay que sacar, lo podemos llamar distinto en todos
    lados" → eligió SALDOS. "Sacar" suena a que la fábrica se lo quiere sacar de
    encima, y a ellos se les está pidiendo que lo ofrezcan; "saldos" es además
    la palabra que el cliente ya usa cuando pregunta."""
    import re
    from pathlib import Path
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    prohibido = re.compile(r"(hay que sacar|Lo parado|liquidar lo parado)", re.I)
    malas = []
    for archivo in carpeta.glob("*.html"):
        # sin comentarios: ahí se explica el cambio y puede nombrar lo viejo
        texto = re.sub(r"(\{#.*?#\}|/\*.*?\*/|//[^\n]*)", " ",
                       archivo.read_text(encoding="utf-8"), flags=re.S)
        for m in prohibido.finditer(texto):
            malas.append(f"{archivo.name}: {m.group(0)}")
    assert not malas, f"quedó el nombre viejo: {malas}"


def test_el_menu_dice_saldos():
    from modules.analisis import views
    assert any(m["titulo"] == "Saldos" for m in views.MENU)
    assert any(m["titulo"] == "Saldos" for m in views.MENU_VENDEDOR)


# ── Toda la segunda entra a la competencia ──────────────────────────────────

def test_entra_toda_la_segunda_y_no_solo_la_parada():
    """Dueña 18/08/2026: "agreguemos toda la tela de segunda a la competencia".

    La consulta ya no filtra —contesta por toda la bodega—, así que la regla
    vive en el `stock_kg`: la tela que no entra como `parado` entra con sus
    kilos de SEGUNDA, y con 0 kg de segunda queda en 0 y no entra."""
    s = " ".join(asinfo_parado.SQL_PARADOS.split())
    assert "ELSE ISNULL(cal.kg_segunda, 0) END AS stock_kg" in s


def test_de_una_tela_que_se_vende_entran_SOLO_los_kilos_de_segunda():
    """⚠ Esos ítems tienen 61.272 kg de primera que salen solos. Sumarlos
    habría inflado la competencia con tela que ya se vende y la meta dejaría de
    significar algo."""
    s = asinfo_parado.SQL_PARADOS
    assert "ELSE ISNULL(cal.kg_segunda, 0) END      AS stock_kg" in s
    assert "ELSE 0 END                              AS kg_primera" in s


def test_cada_fila_guarda_por_que_entro():
    """Sin `motivo`, en la pantalla no hay forma de distinguir 300 kg parados de
    300 kg de segunda de una tela que se vende todas las semanas."""
    assert "'parado' ELSE 'segunda' END AS motivo" in asinfo_parado.SQL_PARADOS
    import inspect
    assert "motivo" in inspect.getsource(queries.actualizar)
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")
    # ⚠ Se chequea la IDEA, no la palabra. Antes esto pedía "f.kg_segunda" en
    # el template: era un proxy de la columna "De 2ª", que ya no existe. Hoy el
    # motivo se ve por la CATEGORÍA, y la categoría se calcula en un solo lugar
    # (`queries.categoria_de`) — el template la lee resuelta. Lo que no puede
    # faltar es que la fila la muestre.
    assert "f.cat" in html
    assert queries.categoria_de({"kg_segunda": 300}) == "SEG"
    assert queries.categoria_de({"kg_primera": 300}) == "PRI"


# ── El premio del mes ───────────────────────────────────────────────────────

def test_el_premio_del_mes_va_por_kilos_y_sin_tope(monkeypatch):
    """Dueña 18/08/2026: "un premio mensual por kg totales quizás?". Es a
    propósito la carrera contraria a la del año: aquélla premia repartir entre
    tipos de tela, ésta premia sacar kilos, y le da algo para ganar al que viene
    último en el porcentaje. Sin eso, el que se descuelga en octubre no vuelve
    a mirar la pantalla."""
    from datetime import date as _d
    c = _competencia_falsa(monkeypatch, meses=[
        {"mes": _d(2026, 10, 1), "vendedor": "Quintero Jose", "kg": 400},
        {"mes": _d(2026, 10, 1), "vendedor": "Intela", "kg": 900},
    ])
    oct_ = next(m for m in c["meses"] if m["mes"] == _d(2026, 10, 1))
    assert oct_["ganador"] == "Intela", "gana por kilos, no por % de meta"
    assert oct_["kg"] == 1300
    assert [x["vendedor"] for x in oct_["podio"]] == ["Intela", "Quintero Jose"]


def test_agosto_y_septiembre_cuentan_juntos(monkeypatch):
    """⚠ Del 25 al 31 de agosto hay cinco días hábiles: un "premio del mes" por
    esa semana no es un mes."""
    from datetime import date as _d
    c = _competencia_falsa(monkeypatch, meses=[
        {"mes": _d(2026, 8, 1), "vendedor": "Intela", "kg": 100},
        {"mes": _d(2026, 9, 1), "vendedor": "Intela", "kg": 250},
    ])
    assert [m["mes"] for m in c["meses"]] == [_d(2026, 9, 1)]
    assert c["meses"][0]["kg"] == 350


def test_el_mes_en_curso_se_marca_como_en_juego(monkeypatch):
    from datetime import date as _d
    c = _competencia_falsa(monkeypatch, meses=[
        {"mes": _d(2026, 9, 1), "vendedor": "Intela", "kg": 100}])
    assert c["meses"][0]["cerrado"] is False


def test_la_pantalla_explica_las_dos_carreras():
    import re
    from pathlib import Path
    crudo = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "competencia.html").read_text(encoding="utf-8")
    # ⚠ Sin comentarios: cuentan qué frase se sacó y por qué, así que nombran
    # justo lo que este test prohíbe.
    html = " ".join(re.sub(r"\{#.*?#\}", " ", crudo, flags=re.S).split())
    assert "El premio del mes" in html
    # ⚠ Decía "kilos totales, sin tope". Lo de "sin tope" no era cierto —cuenta
    # los mismos kilos que puntúan— y venía del TOPE POR GRUPO, la regla del
    # 17/08 que se sacó el 24. Lo que la pantalla tiene que decir es lo que
    # DISTINGUE a las dos carreras: una va por kilos y la otra por puntos.
    assert "sin tope" not in html
    assert "kilos</b> en vez de puntos" in html


def test_la_tabla_semana_a_semana_no_se_dibuja():
    """Dueña 18/08/2026: "ok saca semana a semana". Con el premio del mes
    arriba era demasiada tabla para el vendedor.

    ⚠ El CÁLCULO por semana se sigue haciendo: de ahí salen la columna "Esta
    semana" del ranking y las flechas de subió/bajó. Lo que se saca es el
    dibujo, no el dato."""
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "competencia.html").read_text(encoding="utf-8")
    assert "<h2>Semana a semana</h2>" not in html
    assert "Esta semana" in html, "la columna del ranking se queda"


def test_el_nombre_de_la_tela_no_se_parte_en_dos_renglones():
    """Dueña 18/08/2026: "intenta que el nombre no ocupa mas de una fila". Con
    10 columnas la tabla aprieta esa celda y "Jersey Fancy" se parte en dos, lo
    que duplica el alto de la fila y hace que la lista se lea a saltos.

    ⚠ En el celular vuelve a partirse a propósito: a 390 px no entra de otra
    forma, y ahí el mal menor es la fila alta."""
    from pathlib import Path
    css = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
           "templates" / "analisis" / "base.html").read_text(encoding="utf-8")
    assert "#tabla td.tela{white-space:nowrap}" in css
    i = css.index("@media (max-width: 780px)")
    # ⭐ En el celular NO vuelve a partirse: entra porque la columna Estado ya
    # no existe.
    assert "#tabla td.tela{white-space:normal}" not in css[i:]
    from pathlib import Path as _P
    html = (_P(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")
    assert ">Estado</th>" not in html


def test_una_fila_sin_nada_se_va_pero_la_liquidada_se_queda():
    """Dueña 18/08/2026: "más de 0, si no hay nada para qué están?".

    ⚠ La condición lleva las DOS cosas: 0 kg Y 0 vendidos. Una fila con 0 kg
    que SÍ vendió algo es un ítem liquidado entero, y ésa se queda — es la que
    muestra que la competencia funcionó. Confundirlas borraría la buena
    noticia."""
    from pathlib import Path
    sql = (Path(__file__).resolve().parent.parent / "migrations" /
           "0203_saldos_limpieza.sql").read_text(encoding="utf-8")
    assert "stock_kg <= 0 AND kg_vendidos <= 0" in sql
    assert sql.count("kg_vendidos <= 0") == 2, "las dos tablas con la misma regla"


def test_los_kilos_chicos_no_se_muestran_como_cero():
    """Un ítem de 0,4 kg mostrado como "0" parece un error de la pantalla — fue
    literalmente la pregunta de la dueña."""
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")
    assert "f.stock_kg | num_es(1) if f.stock_kg < 1" in html


def test_la_foto_no_pierde_el_grupo_al_liquidarse_un_item():
    """Un ítem vendido entero ya no viene de Asinfo, así que su grupo quedaba
    en NULL y la fila aparecía con "—" justo en las filas "resuelto", que son
    las que uno mira para ver si esto funciona."""
    import inspect
    fuente = inspect.getsource(queries.actualizar)
    assert "grupo_previo" in fuente
    assert '(p or {}).get("categoria") or grupo_previo.get(k)' in fuente


# ── Las mejoras de la tabla ─────────────────────────────────────────────────

def _html_parado():
    from pathlib import Path
    return (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")


def _html_vendidos():
    """La tabla de Vendidos se mudó a su propia pantalla el 26/08/2026."""
    from pathlib import Path
    return (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "vendidos.html").read_text(encoding="utf-8")


def test_los_indices_de_orden_coinciden_con_las_columnas():
    """⚠ Cada `th` ordena por su `data-i`. Al sacar una columna del medio, si
    los índices no se renumeran, tocar "Clientes" ordena por otra cosa — y no
    da ningún error: sólo ordena mal."""
    import re
    html = _html_parado()
    enc = re.search(r"<thead><tr>\n(.*?)</tr></thead>", html, re.S).group(1)
    idx = [int(x) for x in re.findall(r'data-i="(\d+)"', enc)]
    assert idx == list(range(len(idx))), f"índices salteados o repetidos: {idx}"
    assert len(idx) == int(re.search(r'colspan="(\d+)"', html).group(1))


def test_lo_vendido_no_va_apilado_en_la_celda_de_los_kilos():
    """Primero fue una columna vacía ("—" en las 711 filas); después la puse
    debajo del stock como "−43 vendidos" y tampoco se entendía: dos números de
    kilos apilados sin decir qué era cada uno. Dueña 18/08/2026: "que quiere
    decir x vendidos? deja de ponerme cosas raras".

    Se cuenta igual, y desde el 25/08/2026 se muestra en la tabla VENDIDOS del
    pie —renglón por renglón, con su día y su vendedor—: la frase que estaba al
    abrir la fila se fue con el resto de la descripción ("estas descripciones
    tmbn borrar"). Lo que este test sigue cuidando es que no vuelva a la celda
    de los kilos."""
    html = _html_vendidos()
    assert ">Vendidos</th>" not in html
    assert "vendidos</div>" not in html
    assert 'id="vendidos"' in html, "la tabla de Vendidos del pie es la que cuenta ahora"


def test_el_cero_de_clientes_se_marca():
    """Es la señal más accionable de la tabla —"no hay a quién llamar"— y era
    un cero gris como cualquier otro."""
    assert "'sincli' if not f.clientes" in _html_parado()


def test_el_encabezado_queda_fijo_al_scrollear():
    """⚠ Un sticky muere si un ancestro fija la altura: es el mismo error que
    tuvo el appbar de /mi-cartera durante meses, sin dar ningún síntoma."""
    from pathlib import Path
    css = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
           "templates" / "analisis" / "base.html").read_text(encoding="utf-8")
    assert "#tabla thead th{position:sticky;top:var(--alto-barra)" in css
    assert "background:#fff" in css, "sin fondo sólido las filas se ven por atrás"
    # sin comentarios: ahí se explica justamente cuál es el error a evitar
    import re as _re
    reglas = _re.sub(r"/\*.*?\*/", " ", css, flags=_re.S)
    assert "height:100%" not in reglas, "un ancestro con altura fija mataría el sticky"


def test_las_filas_alternadas_van_de_a_cuatro():
    """Cada tela son DOS filas (la visible y su detalle), así que el ciclo del
    rayado es de 4 y no de 2: con nth-child(2n) se pintarían los detalles."""
    from pathlib import Path
    css = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
           "templates" / "analisis" / "base.html").read_text(encoding="utf-8")
    assert "#tabla tbody tr:nth-child(4n+1) td" in css


def test_abrir_una_fila_cierra_la_anterior():
    html = _html_parado()
    assert "#tabla tr.det.abierta" in html and "o.classList.remove('abierta')" in html


def test_el_encabezado_se_pega_debajo_de_la_barra_y_no_atras():
    """🐛 La barra de arriba también es sticky (top:0, z-index 5) y mide 52 px:
    con el encabezado en top:0 quedaba ESCONDIDO detrás de ella. Se veía sólo
    scrolleando en la pantalla real — el sticky "funcionaba", pero abajo de
    otra cosa."""
    from pathlib import Path
    css = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
           "templates" / "analisis" / "base.html").read_text(encoding="utf-8")
    assert "#tabla thead th{position:sticky;top:var(--alto-barra)" in css
    assert "--alto-barra:52px" in css
    i = css.index("@media (max-width: 780px)")
    assert "--alto-barra:84px" in css[i:], (
        "en el celular la barra se parte en dos líneas y tapa el encabezado")


def test_se_puede_filtrar_por_calidad():
    """Dueña 18/08/2026: "porque hay cosas que se vendieron recien?". Porque 305
    de las 307 filas con venta reciente entraron por sus kilos de SEGUNDA: la
    tela se vende bien, lo que no sale es la segunda. Mezcladas la lista se
    contradice sola —dice "hace 12 meses que no se vende" al lado de una venta
    de la semana pasada—, así que se pueden separar."""
    html = _html_parado()
    assert 'id="calidad"' in html
    assert "data-cal=" in html
    assert "tr.dataset.cal.split(' ').includes(cal)" in html, (
        "una fila puede ser PRI y SEG a la vez: el filtro tiene que mirar la "
        "lista, no comparar el string entero")
    assert "Sólo SEG" in html


def test_la_calidad_es_una_columna_y_no_una_pildora_suelta():
    """Dueña 18/08/2026: "no asi. no es eficiente. tiene que haber columna PRI
    SEG y que se pueda filtrar". Una píldora con kilos al costado no se ordena
    ni se filtra; una columna sí, y se lee en vertical."""
    from pathlib import Path
    todo = " ".join(_html_parado().split())
    # sólo la tabla principal: el resumen por grupo SÍ tiene una columna con
    # los kilos de segunda de cada grupo, y ahí está bien
    html = todo[todo.index('<table id="tabla">'):]
    assert ">Categoría</th>" in html
    assert ">De 2ª</th>" not in html
    # ⚠ Desde el 25/08/2026 la píldora se dibuja en el macro compartido —la usan
    # la lista, los vendidos y la hoja impresa—, así que el markup se chequea
    # ahí; en la tabla lo que tiene que estar es la COLUMNA que la muestra.
    # ⚠ La fila vendida entera no tiene lote que mirar, así que no se inventa
    # una categoría: por eso el `set` viene con su `if`.
    # ⚠ La píldora sale del macro compartido — cómo se le pasan los datos es
    # asunto del template. Lo que se chequea es que la use, no la firma.
    assert "{% set calidad %}" in html and "fm.cal(" in html
    assert 'class="cal">{{ calidad }}' in html
    macro = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "_forma.html").read_text(encoding="utf-8")
    assert 'class="q pri">PRI' in macro and 'class="q seg">SEG' in macro
    # ⚠ Se chequea la IDEA, no la frase: los textos se acortaron el 20/08/2026
    # (dueña: "todo muy wordy") y un test pegado a la redacción obliga a elegir
    # entre el test y la copia. Lo que no puede faltar es que la pantalla
    # explique por qué hay filas con una venta reciente: son los kilos SEG.
    # ⚠ El 25/08/2026 la BAJADA se fue entera ("esto ocupa un montón de
    # renglones"), así que la explicación vive en el pie — que es donde va la
    # letra chica. Sigue estando; cambió de lugar.
    pie = todo[todo.index("<footer>"):]
    assert "SEG" in pie and "quieta" in pie, (
        "el pie explica por qué hay filas con venta reciente")


def test_los_filtros_viven_en_la_direccion_y_se_recuerdan():
    """Dueña 18/08/2026: filtrar, mirar una fila, volver y tener que rearmar
    todo. En la URL además se pueden mandar por WhatsApp."""
    html = _html_parado()
    assert "history.replaceState" in html
    assert "localStorage.setItem" in html
    assert "new URLSearchParams(location.search)" in html


def test_la_direccion_le_gana_a_lo_guardado():
    """Si alguien te mandó un link es porque quiere que veas ESO, no lo que vos
    tenías filtrado ayer."""
    import re
    html = " ".join(_html_parado().split())
    i = html.index("function ponerFiltros()")
    cuerpo = html[i:i + 700]
    assert re.search(r"if \(\[\.\.\.url\.keys\(\)\].*?\) \{.*?url\.get", cuerpo), (
        "primero la URL; el localStorage sólo si la URL no trae nada")


def test_el_grupo_se_aplica_antes_que_el_subgrupo():
    """⚠ `grupoCambio()` borra el subgrupo que no pertenece al grupo elegido.
    Si se restauran en orden alfabético, el subgrupo se pone y el grupo lo
    borra — el link compartido abre a medias y nadie sabe por qué."""
    html = " ".join(_html_parado().split())
    assert "if (c === 'grupo') { e.value = v[c]; grupoCambio(); }" in html


def test_todo_se_cuenta_desde_la_largada_y_no_desde_que_entro_cada_fila():
    """Dueña 18/08/2026: "hace todo desde 25/08".

    Antes cada fila medía desde su propia `fecha_marcado` (13/08), así que la
    pantalla de Saldos y la de Competencia daban dos números distintos para "lo
    vendido" y el primero que los comparara no iba a saber cuál creer.
    `fecha_marcado` sigue guardada —es cuándo entró cada tela— pero ya no manda
    sobre la cuenta."""
    import inspect
    fuente = inspect.getsource(queries.actualizar)
    assert 'config("largada"' in fuente
    # el corte sigue estando: el renglón anterior a la largada se saltea al
    # armar la lista de la competencia
    assert "_fecha(v[\"fecha\"]) < desde_f" in fuente
    assert "MIN(fecha_marcado)" not in fuente, (
        "ya no se arranca desde la fila más vieja de la cohorte")


def test_la_pantalla_dice_desde_cuando_cuenta():
    """Un "vendido" sin fecha al lado invita justamente a la comparación que
    generó la confusión (dueña 24/08/2026). Y desde la vuelta atrás del
    31/08/2026 (tarde) esto importa el doble: "Vendido" tiene que ser
    SIEMPRE el mismo número que /analisis/competencia, y las dos pantallas
    comparten la misma ventana ("desde la largada") — decirlo en las dos
    partes es lo que evita la pregunta "¿por qué acá dice otra cosa?"."""
    html = " ".join(_html_parado().split())
    assert "Vendido" in html and "desde el 25/08" in html


# ── Intela como una cartera más ─────────────────────────────────────────────

def test_se_puede_mirar_intela_por_separado(monkeypatch):
    """Dueña 18/08/2026: "me haces uno para intela? osea mostrador y todo lo
    que no sea vendedores". Es el 51,3% de las ventas de estas telas y no se
    podía aislar.

    ⚠ Se resuelve con `vend_pc IS NULL` y no inventándole un código: Intela no
    está en `scintela.vendedor` y no debería estarlo, porque no es una
    persona."""
    visto = {}

    def fake(sql, params=None, conn=None):
        # ⚠ Se guarda SÓLO la primera consulta —la de la hoja—: `por_cliente`
        # después pide los puntos, y si se pisara `visto` el test terminaría
        # mirando otra query y pasando o fallando por el motivo equivocado.
        if "sql" not in visto:
            visto["sql"] = " ".join(sql.split())
            visto["params"] = params
        # Una fila de puntaje para que no salga a buscar a Asinfo.
        if "parado_punto" in " ".join(sql.split()):
            return [{"subcategoria": "—", "categoria": None, "kg_base": 0,
                     "kg_12m": 0, "meses": None, "nivel": 1, "puntos": 1}]
        return []

    monkeypatch.setattr(queries.db, "fetch_all", fake)
    monkeypatch.setattr(queries.db, "fetch_one", lambda *a, **k: None)
    queries.por_cliente("INTELA")
    assert "%(vend)s = 'INTELA' AND l.vend_pc IS NULL" in visto["sql"]
    assert visto["params"]["vend"] == "INTELA"


def test_intela_esta_en_los_dos_desplegables():
    from pathlib import Path
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    for archivo in ("parado.html", "parado_clientes.html"):
        html = (carpeta / archivo).read_text(encoding="utf-8")
        assert "Intela (mostrador)" in html, f"falta en {archivo}"


def test_una_tela_cuyos_clientes_son_del_mostrador_aparece_con_INTELA():
    """El filtro de la pantalla de saldos arma `data-vend` salteando los nulos:
    sin agregar INTELA a mano, una tela que sólo compra el mostrador quedaba
    fuera de todos los filtros por vendedor."""
    html = " ".join(_html_parado().split())
    assert "rejectattr('vend_pc')" in html and "['INTELA']" in html


def _html_hoja():
    from pathlib import Path
    return (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado_clientes.html").read_text(encoding="utf-8")


def test_los_dos_kilos_de_la_hoja_dicen_de_quien_son():
    """Dueña 18/08/2026: "no se que quiere decir le vendimos / kg en saldo".
    Eran dos cifras de kilos una al lado de la otra sin decir de quién era cada
    una: una es lo que HAY en bodega y la otra lo que ese cliente COMPRÓ."""
    html = _html_hoja()
    assert "Hay para ofrecerle" in html and "Él ya compró" in html
    assert "Kg en saldo</th>" not in html and "Le vendimos</th>" not in html
    assert 'class="anio">en {{ t.anio }}' in html, (
        "sin el año, 'ya compró' no dice si fue en julio o en 2023")


def test_las_fichas_encolumnan_todas_igual():
    """⚠ Sin `table-layout:fixed` cada ficha arma sus columnas según el largo
    de SU contenido: un cliente con "Fleece 96 Sin Perchar" y otro con "Naty"
    quedaban con las cifras en lugares distintos y la hoja se lee saltando."""
    html = _html_hoja()
    assert "table-layout:fixed" in html
    assert "<colgroup>" in html
    import re
    anchos = [int(x) for x in re.findall(r"\.cli col\.c\d\{width:(\d+)%\}", html)]
    assert sum(anchos) == 100, f"los anchos suman {sum(anchos)}%"


def test_una_sola_palabra_para_la_categoria():
    """Dueña 18/08/2026: "otra columna que diga categoria y pones PRI SEG, no
    repetir con segunda". Convivían "calidad", "segunda", "de 2ª" y "SEG" para
    la misma cosa: cuatro nombres obligan a traducir mentalmente en cada
    columna."""
    import re
    html = _html_parado()
    visible = re.sub(r"(\{#.*?#\}|<!--.*?-->|//[^\n]*)", " ", html, flags=re.S)
    for palabra in ("De 2ª", "de segunda", "segunda calidad", "Primera y segunda"):
        assert palabra not in visible, f"quedó «{palabra}» conviviendo con SEG"
    assert ">Categoría</th>" in visible


def test_la_hoja_sale_alfabetica_por_codigo():
    """Dueña 18/08/2026: "cuando es 'a quien ofrecerle que' ordena
    alfabeticamente".

    ⭐ Por CÓDIGO y no por nombre, aunque la ficha muestre el nombre grande: es
    el mismo orden con el que salen la hoja de estado de cuenta de la oficina y
    la de /mi-cartera, y el vendedor tiene el papel y el celular delante a la
    vez. Dos alfabéticos distintos para el mismo cliente son peor que uno."""
    import inspect

    from modules.analisis import views
    for vista in (views.parado_clientes, views.mi_hoja, views.mi_hoja_csv,
                  views.parado_clientes_csv):
        fuente = inspect.getsource(vista)
        assert 'request.args.get("orden") or "codigo"' in fuente, (
            f"{vista.__name__} sigue arrancando por oportunidad")


def test_ordenar_por_oportunidad_sigue_disponible():
    """Se cambia el DEFAULT, no se saca la opción: para decidir por quién
    empezar sigue siendo la mejor."""
    assert "oportunidad" in queries.ORDENES


def test_el_titulo_de_saldos_no_promete_lo_mismo_que_la_otra_pantalla():
    """Dueña 20/08/2026: "¿por qué Saldos se llama 'y a quién llamar', si ya hay
    otra página que dice 'A quién ofrecerle qué'?".

    Las dos muestran los mismos clientes; lo que cambia es el EJE (por tela vs.
    por cliente). Con los dos títulos prometiendo clientes, el que entra no sabe
    cuál abrir."""
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")
    titulo = html.split("<h1>")[1].split("</h1>")[0]
    assert titulo == "Saldos, tela por tela"
    assert "llamar" not in titulo and "ofrec" not in titulo


def test_el_premio_del_mes_lista_a_los_siete_ordenados_por_kilos(monkeypatch):
    """Dueña 20/08/2026: "acá me gustaría tener a todos ordenados por kg".

    El podio de tres dejaba afuera a la mitad de la tabla —y en el teléfono se
    veía UN nombre, porque las columnas de la derecha se esconden—. Van los
    siete, también los que todavía no vendieron: un cero en la lista dice más
    que no figurar."""
    from datetime import date as _d
    c = _competencia_falsa(monkeypatch, meses=[
        {"mes": _d(2026, 10, 1), "vendedor": "Quintero Jose", "kg": 400},
        {"mes": _d(2026, 10, 1), "vendedor": "Intela", "kg": 900},
    ])
    r = next(m for m in c["meses"] if m["mes"] == _d(2026, 10, 1))["ranking"]
    assert len(r) == len(queries.COMPETIDORES) == 7
    assert [x["vendedor"] for x in r[:2]] == ["Intela", "Quintero Jose"]
    assert [x["kg"] for x in r[:2]] == [900, 400]
    assert [x["puesto"] for x in r] == [1, 2, 3, 4, 5, 6, 7]
    assert all(x["kg"] == 0 for x in r[2:]), "los que no vendieron van con cero"
    # el orden de los empatados en cero lo fija el nombre, no el diccionario
    assert [x["vendedor"] for x in r[2:]] == sorted(x["vendedor"] for x in r[2:])


def test_la_pantalla_dibuja_la_lista_entera_del_mes():
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "competencia.html").read_text(encoding="utf-8")
    assert "m.ranking" in html, "sin esto la pantalla sigue mostrando sólo el podio"
    assert "Kg del mes" in html


# ──────────────────────────────────────────────────────────────────────────
# Qué kilo puntúa · migración 0212
# ──────────────────────────────────────────────────────────────────────────


@pytest.mark.parametrize(
    ("motivo", "calidad", "cuenta"),
    [
        ("parado", "PRI", True),
        ("parado", "SEG", True),
        ("parado", None, True),
        # ⭐ VUELTA ATRÁS 01/09/2026 (ver el docstring de `cuenta_el_kilo`).
        # `calidad` acá YA es la del LOTE REAL que despachó
        # (`asinfo_parado._JOIN_LOTE_DESPACHO`), no lo que decía la factura
        # — por eso un `segunda` con calidad "SEG" sigue contando: es el caso
        # Kiana Forro (lote real SEGUNDA, factura mal cargada en PRIMERA), y
        # con la calidad resuelta contra el lote real ya no importa qué decía
        # la factura.
        ("segunda", "SEG", True),
        # ⭐ Lo que el 31/08 dejaba pasar y el 01/09 cierra: una tela que se
        # vende sola (verificado en vivo: Fleece 96 Sin Perchar, Rib Normal,
        # Rib Acanalado — 600 a 3.800 kg reales de primera desde la largada)
        # llenaba su tope de segunda (12-130 kg) con ventas de PRIMERA de
        # verdad, sin mover un gramo de la segunda marcada.
        ("segunda", "PRI", False),
        ("segunda", None, False),
        # Cohorte vieja, sin motivo guardado: cuenta todo, como venía.
        (None, "PRI", True),
    ],
)
def test_solo_la_segunda_de_verdad_puntua_para_un_item_segunda(motivo, calidad, cuenta):
    assert queries.cuenta_el_kilo(motivo, calidad) is cuenta


def test_las_devoluciones_restan_en_lo_que_puntua():
    """⭐ Dueña 24/08/2026: "dale contemos devoluciones". La nota de crédito es
    el documento 20 o 451 y hay que RESTARLA: sin esto un vendedor facturaba un
    saldo, sumaba los puntos, el cliente devolvía la tela y los puntos quedaban.

    ⚠ Se imputa al día de la factura MADRE, no al suyo: una devolución de
    octubre de una venta de agosto tiene que caer en agosto."""
    sql = " ".join(asinfo_parado._sql_vendido("2026-08-25").split())
    assert "fc.id_documento IN (7, 251, 20, 451)" in sql
    assert "THEN -dfc.cantidad ELSE dfc.cantidad END" in sql
    assert "COALESCE(pad.fecha, fc.fecha)" in sql
    assert "id_factura_cliente_padre" in sql


def test_la_dificultad_de_la_tela_tambien_es_neta():
    """Con la devolución adentro la tela parece más fácil de colocar de lo que
    es, y se le da MENOS puntaje del que corresponde."""
    sql = " ".join(asinfo_parado.SQL_VENTA_TELA.split())
    assert "fc.id_documento IN (7, 251, 20, 451)" in sql
    assert "THEN -dfc.cantidad ELSE dfc.cantidad END" in sql
    # y trae la SEG aparte, para poder ponerle puntaje propio
    assert "kg_seg_12m" in sql


def test_lo_vendido_trae_la_calidad_del_lote_real_del_despacho():
    """CERRADO 01/09/2026 — reemplaza al test del 24/08 que llevaba este
    nombre. `dfc.id_lote` (la columna de la LÍNEA de factura) sigue en NULL
    siempre (eso no cambió: no se usa ni antes ni ahora). Lo que sí cambió es
    de dónde sale la calidad: antes, del atributo 2 de la línea de factura
    (lo que alguien tipeó); ahora, primero del LOTE REAL que salió por el
    despacho (`detalle_despacho_cliente.codigo_lote` → `lote`), y sólo si no
    hay despacho vinculado cae a la línea de factura — ver
    `test_la_calidad_de_una_venta_sale_del_lote_real_del_despacho`."""
    sql = " ".join(asinfo_parado._sql_vendido("2026-08-25").split())
    assert ("COALESCE(lot_desp.id_valor_atributo_2, dfc.id_valor_atributo_2, "
            "mad.va2) = 4 THEN 'SEG'") in sql
    assert "dfc.id_lote" not in sql, "la línea de factura no tiene lote propio"


def test_la_devolucion_hereda_la_calidad_y_el_vendedor_de_la_factura_madre():
    """⚠ En las líneas de nota de crédito el atributo de calidad viene en NULL
    (17 de 17, verificado el 24/08/2026) y `v_ventas` no tiene la NC.

    Sin heredar los dos de la madre: la devolución de un kilo SEG se etiquetaba
    PRI y en un ítem de segunda quedaba sin contar —la venta sumaba los puntos y
    la devolución no los restaba—, y el kilo negativo se le restaba a 'Intela',
    que también compite, en vez de al vendedor que lo facturó.

    ⚠ TOP 1 y no un JOIN: la madre puede repetir el mismo producto en dos
    renglones y un JOIN duplicaría los kilos de la devolución.
    """
    sql = " ".join(asinfo_parado._sql_vendido("2026-08-25").split())
    assert "OUTER APPLY (SELECT TOP 1 m.id_valor_atributo_2" in sql
    assert ("ON vx.id_factura_cliente = COALESCE(fc.id_factura_cliente_padre, "
            "fc.id_factura_cliente)") in sql


def test_la_fecha_de_la_madre_es_solo_para_la_nota_de_credito():
    """Con un COALESCE pelado, una factura común con padre cargado se mudaría a
    la fecha del padre y podría salirse de la ventana de la competencia."""
    sql = " ".join(asinfo_parado._sql_vendido("2026-08-25").split())
    assert ("CASE WHEN fc.id_documento IN (20, 451) THEN "
            "COALESCE(pad.fecha, fc.fecha) ELSE fc.fecha END") in sql


def test_el_motivo_de_la_cohorte_se_escribe_una_sola_vez():
    """El INSERT de la cohorte es ON CONFLICT DO NOTHING: sin un UPDATE, los 734
    ítems anteriores a la migración 0212 se quedaban con el motivo en NULL —o
    sea contando toda la primera. Y con `motivo IS NULL` en el WHERE, una vez
    escrito no se vuelve a mover: la regla de un ítem no puede cambiar en la
    mitad de la carrera.

    La ÚNICA excepción es la tela reciente: si el ítem entró como `parado` y hoy
    resulta que su tela se hizo hace menos de 6 meses, sus kilos de primera
    nunca debieron contar y el motivo se corrige a `segunda`."""
    import inspect as _i
    fuente = " ".join(_i.getsource(queries.actualizar).split())
    assert "UPDATE scintela.parado_cohorte SET motivo = %s" in fuente
    assert "AND (motivo IS NULL OR (%s AND motivo = 'parado'))" in fuente


def test_las_pantallas_respetan_la_bandera_cuenta():
    """La regla vive en UN lugar (el refresh la escribe en `parado_venta.cuenta`)
    y las pantallas la respetan. Con tres WHERE distintos, tarde o temprano uno
    queda sin actualizar y el ranking y el total dejan de coincidir."""
    import inspect as _i
    fuente = _i.getsource(queries)
    # Cada lectura de parado_venta o FILTRA por la bandera (las tres que suman)
    # o la SELECCIONA para mostrarla (el detalle de qué vendió cada uno). Lo que
    # no puede pasar es que una lectura la ignore y sume kilos que no puntúan.
    # ⚠ Se mira hasta el FINAL de la consulta, no una ventana de N caracteres.
    # La ventana fija ya obligó a agrandarla dos veces el 26/08/2026 —cada
    # LATERAL nuevo empuja el WHERE más lejos del FROM— y agrandarla es aflojar
    # el test: con la ventana muy grande empieza a ver el `v.cuenta` de la
    # consulta de al lado y deja de probar nada.
    for m in re.finditer(r"FROM scintela\.parado_venta v", fuente):
        fin = fuente.find('"""', m.start())
        trozo = fuente[m.start():fin if fin > 0 else len(fuente)]
        assert "AND v.cuenta" in trozo or "v.cuenta," in trozo, trozo[:220]


def test_once_noventa_y_ocho_meses_es_doce():
    """⭐ Dueña 24/08/2026: "11.98 es igual que 12". El corte entre 4 y 10
    puntos está en 12 meses parados, y Jersey Forro Spun daba 11,98: sus
    2.448 kg —el ítem más grande de la lista— valían 4 en vez de 10 por una
    diferencia del 0,2%, menos que el error de medición de la bodega."""
    nivel, meses = queries._nivel(2448.3, 2452.35)
    assert round(meses, 2) == 11.98
    assert nivel == 3 and queries.PUNTOS[nivel] == 10


def test_el_redondeo_no_mueve_a_ninguna_otra_tela():
    """El redondeo es de UN decimal a propósito: arregla el borde sin correr la
    línea. Las cuatro telas que están cerca del otro corte (1 mes) se quedan
    donde estaban."""
    for kg_base, kg_12m in [(659.6, 7673.95),    # Jersey Lycra 3.3 — 1,03
                            (191.25, 2023.4),    # Franela — 1,13
                            (21.25, 223.1),      # Stefi — 1,14
                            (204.0, 1974.0)]:    # WAFFER — 1,24
        nivel, _ = queries._nivel(kg_base, kg_12m)
        assert nivel == 2, (kg_base, kg_12m)
    # y una que sí es fácil de verdad sigue siendo fácil
    assert queries._nivel(100, 12000)[0] == 1


def test_hay_un_solo_numero_de_puntos_en_juego(monkeypatch):
    """⭐ Dueña 25/08/2026: "en juego es distinto que la presentacion, puntos".

    La Competencia mostraba la bolsa CONGELADA (`kg_base × puntos`, los kilos
    del día en que se fijó el puntaje) y Saldos la suma de la foto de HOY, que
    se mueve con la bodega. Dos números con el mismo nombre, y el de Saldos
    nunca vuelve a coincidir con uno impreso.
    """
    monkeypatch.setattr(queries, "puntos_por_tela", lambda: {
        "Jersey 3": {"kg_base": 100, "puntos": 10},
        "Fleece 102": {"kg_base": 50, "puntos": 4},
    })
    assert queries.bolsa_congelada() == 1200

    # y la pantalla la usa, en vez de sumar la foto de hoy
    import inspect as _i
    fuente = _i.getsource(views.parado) + _i.getsource(views.mis_telas)
    assert fuente.count("bolsa=queries.bolsa_congelada()") == 2


def test_la_lista_para_imprimir_va_por_tela_y_color(app, monkeypatch):
    """⭐ Dueña 25/08/2026: "quiero una lista en pdf para imprimir todas las
    telas", tela y color.

    ⚠ En papel va ALFABÉTICA, no por puntos como la pantalla: en una hoja no se
    filtra ni se busca, se busca con el dedo. Y los ítems sin kilos no van: son
    los que ya se vendieron, que en la pantalla se quedan a propósito para ver
    si la competencia funcionó, pero en una lista para salir a ofrecer sobran.
    """
    filas = [
        {"subcategoria": "Zeta", "color": "NEG", "stock_kg": 100, "puntos": 10},
        {"subcategoria": "Alfa", "color": "BLA", "stock_kg": 10, "puntos": 1},
        {"subcategoria": "Alfa", "color": "ROJ", "stock_kg": 50, "puntos": 1},
        {"subcategoria": "Alfa", "color": "VER", "stock_kg": 0, "puntos": 1},
    ]
    for f in filas:
        f["puntos_fila"] = f["stock_kg"] * f["puntos"]
    monkeypatch.setattr(views.queries, "items", lambda: filas)
    monkeypatch.setattr(views.queries, "con_puntos", lambda f: f)
    monkeypatch.setattr(views.queries, "estado", lambda: {})

    with app.test_request_context():
        d = views._hoja_saldos()

    assert [t for t, _ in d["bloques"]] == ["Alfa", "Zeta"], "alfabética"
    assert [f["color"] for f in d["bloques"][0][1]] == ["ROJ", "BLA"], (
        "adentro de una tela, por kilos")
    assert d["kg_total"] == 160 and d["puntos_total"] == 1060
    assert all(float(f["stock_kg"]) > 0 for f in d["filas"]), (
        "lo que ya no está en bodega no se sale a ofrecer")


def test_el_pdf_de_saldos_sale_del_mismo_html_que_la_pantalla():
    """Dos plantillas para el mismo papel divergen a la primera corrección que
    se le hace a una sola. Y si el servidor no tiene navegador, el PDF avisa en
    vez de romper — la pantalla de imprimir sigue andando."""
    import inspect as _i
    fuente = _i.getsource(views.saldos_imprimir_pdf)
    assert 'render_template("analisis/parado_impreso.html"' in fuente
    assert "pdf_motor.desde_html" in fuente
    assert "SinMotor" in fuente and "503" in fuente
    assert 'render_template("analisis/parado_impreso.html"' in _i.getsource(
        views.saldos_imprimir), "la pantalla y el PDF, el mismo template"


def test_la_lista_impresa_la_pueden_ver_los_vendedores():
    """Cuelga de /analisis/competencia, el prefijo que tienen abierto. No lleva
    un solo cliente adentro: son telas, colores y kilos de la fábrica."""
    from scope_vendedor import PREFIJOS_PERMITIDOS, _path_permitido
    for ruta in ("/analisis/competencia/telas/imprimir",
                 "/analisis/competencia/telas/imprimir.pdf"):
        assert _path_permitido(ruta, PREFIJOS_PERMITIDOS), ruta


def test_la_forma_sale_del_lote_y_no_separa_los_kilos():
    """⭐ Dueña 25/08/2026: "sumar tubular y abierta" y, enseguida, "agregar si
    es tubular o abierta". Las dos cosas a la vez: los kilos siguen en UNA fila
    por tela × color, y la fila dice de qué forma son.

    ⚠ En Asinfo la forma es un atributo del LOTE y sus slots no siguen el
    número del atributo: la calidad (atributo 2) está en `id_valor_atributo_2`
    pero la forma (atributo 1) está en el `_3`, y el color en el `_1`. Por eso
    se filtra por el CÓDIGO del valor: si mañana el slot cambia, la consulta da
    0 en vez de confundir un color con una forma.
    """
    sql = " ".join(asinfo_parado.SQL_PARADOS.split())
    assert "t.codigo = 'TUB'" in sql and "t.codigo = 'ABI'" in sql
    assert "valor_atributo t ON t.id_valor_atributo = l.id_valor_atributo_3" in sql
    # los kilos de la fila NO se abren por forma: sigue habiendo un stock_kg
    assert "AS kg_tubular" in sql and "AS kg_abierta" in sql
    # Las dos agrupaciones —la del stock y la de la calidad— van por tela y
    # color y nada más: ni el GROUP BY ni el SELECT abren la fila por forma.
    for g in re.findall(r"GROUP BY [^)]*", sql):
        assert "atributo" not in g and "codigo = 'TUB'" not in g, g


def test_la_forma_se_resuelve_una_sola_vez_en_la_consulta():
    """TUB, ABI o vacío. Se arma en el SQL y no en cada plantilla: la pantalla y
    la hoja impresa tienen que decir lo mismo.

    ⭐ Y NUNCA las dos (dueña 25/08/2026: "telas no pueden ser tub y abi al
    mismo tiempo"). Medido ese día contra la bodega 53: en cada tela × color
    que figuraba con las dos, la forma minoritaria era el 0-4% de los lotes
    —Fleece 96 Perchado NEG, 1.296 abiertos contra 36.878 tubulares—. Son lotes
    mal marcados, no dos formas."""
    import inspect as _i
    fuente = " ".join(_i.getsource(queries.items).split())
    assert "TUB ABI" not in fuente, (
        "volvió el caso mezclado: sostiene que una tela puede ser tubular y "
        "abierta a la vez")
    assert "THEN 'TUB'" in fuente and "THEN 'ABI'" in fuente
    assert "AS forma" in fuente


def _forma(**kg):
    """Renderiza el macro de la forma con los kilos que se le pasen."""
    from jinja2 import Environment, FileSystemLoader

    import filters as F
    env = Environment(loader=FileSystemLoader("modules/analisis/templates"))
    env.filters["num_es"] = F.num_es
    tpl = env.from_string(
        '{% import "analisis/_forma.html" as fm %}{{ fm.forma(f) }}')
    return " ".join(tpl.render(f=kg).split())


def test_la_hoja_abre_el_color_por_forma_Y_por_calidad():
    """⭐ Dueña 25/08/2026: "idem con PRI y SEG como tubular y abierta, no es lo
    mismo". Tubular y abierta se cortan distinto; primera y segunda se venden a
    precios distintos. Un renglón de 171 kg que son 95 tubulares de segunda y 76
    abiertas de primera promete cuatro cosas a la vez."""
    out = queries.abrir_en_lineas([{
        "subcategoria": "Jersey 3", "color": "JME", "stock_kg": 171,
        "puntos": 4, "puntos_fila": 684,
        "kg_tub_pri": 0, "kg_tub_seg": 95, "kg_abi_pri": 76, "kg_abi_seg": 0}])
    assert [(f["forma_fila"], f["cal_fila"], f["stock_kg"]) for f in out] == [
        ("TUB", "SEG", 95), ("ABI", "PRI", 76)]
    assert sum(f["stock_kg"] for f in out) == 171, "las líneas cierran"
    assert [f["puntos_fila"] for f in out] == [380, 304]
    # cada línea lleva su propia calidad, para que la píldora no mienta
    assert [(f["kg_segunda"], f["kg_primera"]) for f in out] == [(95, 0), (0, 76)]


def test_un_color_de_una_sola_manera_no_se_abre():
    """Si viene todo tubular y todo de primera, la fila ya lo dice: abrirla en
    una sola línea sería el mismo renglón con una columna más."""
    f = {"subcategoria": "X", "color": "BAN", "stock_kg": 127, "puntos": 4,
         "puntos_fila": 508, "kg_tub_pri": 0, "kg_tub_seg": 127,
         "kg_abi_pri": 0, "kg_abi_seg": 0}
    out = queries.abrir_en_lineas([f])
    assert out == [f] and "forma_fila" not in out[0]


def test_la_hoja_abre_el_color_en_dos_lineas_cuando_hay_las_dos_formas():
    """⭐ Dueña 25/08/2026: "no, una cantidad por tubular. una cantidad por
    abierta… dos lineas para el color cuando hay ambas telas". Tubular y
    abierta no son la misma tela: se cortan distinto y el cliente pide una o la
    otra."""
    filas = [
        {"subcategoria": "Jersey 3", "color": "ROB", "stock_kg": 129,
         "kg_tub_pri": 90, "kg_tub_seg": 0, "kg_abi_pri": 39, "kg_abi_seg": 0,
         "puntos": 4, "puntos_fila": 516},
        {"subcategoria": "Jersey 3", "color": "BAN", "stock_kg": 127,
         "kg_tub_pri": 127, "kg_tub_seg": 0, "kg_abi_pri": 0, "kg_abi_seg": 0,
         "puntos": 4, "puntos_fila": 508},
    ]
    out = queries.abrir_en_lineas(filas)
    assert len(out) == 3, "la que tiene las dos formas se abre; la otra no"
    dos = [f for f in out if f["color"] == "ROB"]
    assert [f["forma_fila"] for f in dos] == ["TUB", "ABI"]
    assert [f["stock_kg"] for f in dos] == [90, 39]
    assert sum(f["stock_kg"] for f in dos) == 129, "las dos líneas cierran"
    assert [f["puntos_fila"] for f in dos] == [360, 156]
    # la que tiene una sola forma queda intacta, sin `forma_fila`
    una = [f for f in out if f["color"] == "BAN"][0]
    assert "forma_fila" not in una and una["stock_kg"] == 127


def test_lo_que_no_cuadra_entre_las_dos_tablas_va_en_su_propia_linea():
    """Los kilos por forma salen del LOTE y el total de la fila de otra tabla:
    cierran al 0,006% pero no son la misma consulta. Lo que sobra va en una
    línea SIN forma — un kilo inventado en la columna de la izquierda es peor
    que un renglón que dice "no sé de qué forma es"."""
    out = queries.abrir_en_lineas([
        {"subcategoria": "X", "color": "NEG", "stock_kg": 200,
         "kg_tub_pri": 90, "kg_tub_seg": 0, "kg_abi_pri": 39, "kg_abi_seg": 0,
         "puntos": 1, "puntos_fila": 200}])
    assert [f["stock_kg"] for f in out] == [90, 39, 71]
    assert out[-1]["forma_fila"] == ""
    # y una diferencia de menos de un kilo no ensucia la hoja con un renglón
    chico = queries.abrir_en_lineas([
        {"subcategoria": "X", "color": "NEG", "stock_kg": 129.4,
         "kg_tub_pri": 90, "kg_tub_seg": 0, "kg_abi_pri": 39, "kg_abi_seg": 0,
         "puntos": 1, "puntos_fila": 129}])
    assert len(chico) == 2


def test_la_fila_dice_UNA_forma_y_nunca_las_dos():
    """⭐ Dueña 25/08/2026: *"telas no pueden ser tub y abi al mismo tiempo"*.

    Antes la fila decía «TUB ABI» cuando el ítem tenía lotes de las dos, y se
    partía en dos renglones. Medido ese día contra la bodega 53: en cada tela ×
    color que figuraba con las dos, la forma minoritaria era el 0-4% de los
    lotes —Fleece 96 Perchado NEG, 1.296 abiertos contra 36.878 tubulares—.
    Son lotes mal marcados. Ahora `formas()` elige la mayoría y el refresco
    dobla los kilos de la otra sobre ella, así que este caso no llega ni al
    macro; si llegara —datos viejos, sin refrescar— manda el tubular y NO se
    dicen las dos.

    ⚠ La sigla y NADA de kilos: los kilos de la línea están en su columna.
    "TUB 422" al lado de un renglón de 171 kg son dos números que no cierran.
    """
    con_dos = _forma(kg_tubular=120, kg_abierta=51)
    assert con_dos == '<span class="fm">TUB</span>', (
        "la fila volvió a decir las dos formas")

    assert _forma(kg_tubular=200, kg_abierta=0) == '<span class="fm">TUB</span>'
    assert _forma(kg_tubular=0, kg_abierta=200) == '<span class="fm">ABI</span>'
    # el lote que no lo dice no inventa una forma
    assert "—" in _forma(kg_tubular=0, kg_abierta=0)


def test_la_forma_de_la_tela_sale_de_la_mayoria_de_los_lotes():
    """`formas()` devuelve UNA sigla: la que tienen más lotes. Con una bandera
    SÍ/NO, un solo lote mal marcado entre 24.000 alcanzaba para que la fila
    dijera las dos."""
    llamadas = {}

    def falso(sql):
        llamadas["sql"] = sql
        return [
            {"subcategoria": "Fleece 96 Perchado", "color": "NEG",
             "tub": 36878, "abi": 1296},
            {"subcategoria": "Lluvia", "color": "MAR", "tub": 2, "abi": 900},
            {"subcategoria": "Sin lotes", "color": "XXX", "tub": 0, "abi": 0},
        ]

    import modules.analisis.asinfo_parado as ap
    viejo = ap._filas
    ap._filas = falso
    try:
        out = ap.formas()
    finally:
        ap._filas = viejo
    assert out[("Fleece 96 Perchado", "NEG")] == "TUB"
    assert out[("Lluvia", "MAR")] == "ABI"
    assert ("Sin lotes", "XXX") not in out, "sin lotes marcados no se inventa forma"
    assert "SUM(CASE WHEN t.codigo = 'TUB'" in llamadas["sql"], (
        "volvió la bandera SÍ/NO: un lote mal marcado vuelve a mandar")


def test_la_forma_se_dibuja_con_UN_macro_en_las_dos_pantallas():
    """Dos copias del mismo `if` se despegan a la primera corrección — ya pasó
    con la píldora de calidad."""
    from pathlib import Path
    base = Path("modules/analisis/templates/analisis")
    for hoja in ("parado.html", "parado_impreso.html"):
        t = (base / hoja).read_text(encoding="utf-8")
        assert '{% import "analisis/_forma.html" as fm %}' in t, hoja
        assert "fm.forma(f)" in t, hoja
        assert "TUB" not in t, f"{hoja} dibuja la forma por su cuenta"


def test_la_fila_del_vendedor_abre_QUE_vendio_no_solo_el_grupo():
    """⭐ Dueña 25/08/2026: "esto me sigue sin abrir pique". El cuadro que se
    abre decía el GRUPO —Pique, 102 kg— y ahí se terminaba: qué tela y qué
    color no estaba en ninguna pantalla.

    ⚠ Este test existe porque el bloque se perdió una vez: el commit lo
    nombraba, la función de datos estaba, y la plantilla había quedado sin el
    cambio. En vivo no se veía nada y el commit decía que sí."""
    from pathlib import Path
    t = (Path("modules/analisis/templates/analisis/competencia.html")
         .read_text(encoding="utf-8"))
    # ⭐ Desde el 25/08/2026 los renglones van DENTRO de su grupo y no en una
    # tabla aparte más abajo (dueña: "quería ver el pique especial de lo vendido
    # dentro de pique"). Lo que este test cuida sigue siendo lo mismo: que al
    # abrir un vendedor se vea QUÉ vendió y no sólo el grupo.
    assert "{% for v in d.lineas %}" in t
    assert "{% for v in r.vendido %}" not in t, (
        "la tabla aparte volvió: los renglones van colgados de su grupo")
    assert "v.subcategoria" in t and "v.color" in t
    # ⚠ SÓLO lo que puntúa (dueña 25/08/2026: "lo que no cuenta para puntos ni
    # lo muestres"). La consulta ya filtra por la bandera; acá se fija que la
    # plantilla no vuelva a dibujar los que no suman.
    cuerpo = re.sub(r"\{#.*?#\}", "", t, flags=re.S)   # sin los comentarios
    assert "no suma" not in cuerpo and "nocuenta" not in cuerpo


# ── Sólo tela ESTANCADA, no la recién hecha (dueña 25/08/2026) ──────────────

def test_la_tela_recien_hecha_no_entra_como_parada():
    """*"Que la competencia tenga solo tela estancada, no tela que se hizo
    recientemente y no se movió"*.

    "No vendió un kilo en 12 meses" lo cumple sola la tela que salió de
    producción el mes pasado: nunca tuvo la chance de venderse."""
    sql = " ".join(asinfo_parado.SQL_PARADOS.split())
    assert (f"AND s.fecha <= DATEADD(day, -{asinfo_parado.DIAS_QUIETO}, "
            "GETDATE())") in sql
    assert asinfo_parado._QUIETO in asinfo_parado._ES_PARADO
    assert asinfo_parado._QUIETO not in asinfo_parado._SIN_VENTA


def test_la_antiguedad_se_mide_por_el_saldo_y_no_por_los_rollos():
    """⚠ El 11 y el 25/04/2026 un re-loteo de bodega le creó rollos NUEVOS a
    tela vieja, sin una sola orden de fabricación detrás. Midiendo por rollo,
    Rib Spun AMF —última venta 17/11/2022— daba producción fresca y se caía de
    la lista: 12 ítems y 314 kg de la tela más clavada que hay, justo al revés
    de lo que la regla busca.

    Los rollos cambian de número; el saldo del producto no. Y es lo único que
    se puede medir para la tela que ya se vendió entera, que no tiene rollos."""
    sql = " ".join(asinfo_parado.SQL_PARADOS.split())
    assert "FROM a_corte u JOIN producto p" in sql
    assert "kg_antes" in asinfo_parado._QUIETO
    assert "id_lote" not in asinfo_parado._QUIETO, (
        "la antigüedad no puede volver a salir del lote")


def test_los_kilos_viejos_tienen_que_llegar_al_minimo():
    """No alcanza con que hubiera algún kilo: la tela tiene que haber tenido
    parados los mismos 20 kg que se le piden a cualquier otra."""
    assert (f"ISNULL(antes.kg_antes, 0) >= {asinfo_parado.MIN_KG}"
            in asinfo_parado._QUIETO)


def test_sin_foto_vieja_de_la_bodega_no_se_filtra_nada():
    """⚠ El filtro excluye sólo con PRUEBA de que la tela es reciente.

    Si `saldo_producto` no llega hasta el corte (Asinfo purgó historia), ninguna
    tela figuraría con saldo viejo y la lista entera se vaciaría sola — y una
    lista vacía se lee como "ya no queda nada parado", que es la peor manera de
    romperse."""
    q = " ".join(asinfo_parado._QUIETO.split())
    assert (f"hist.desde > DATEADD(day, -{asinfo_parado.DIAS_QUIETO}, "
            "GETDATE())") in q
    sql = " ".join(asinfo_parado.SQL_PARADOS.split())
    assert "hist AS ( SELECT MIN(fecha) AS desde FROM saldo_producto" in sql
    assert "CROSS JOIN hist" in sql


def test_la_tela_pedida_no_compite():
    """Dueña 25/08/2026: *"si la tela se produjo por un pedido, tiene que salir
    de la competencia"*. Esa tela ya tiene dueño: darle puntos a quien la
    facture es pagar por una venta que ya estaba hecha."""
    assert "NOT " + asinfo_parado._PEDIDA in asinfo_parado._ES_PARADO
    sql = " ".join(asinfo_parado.SQL_PARADOS.split())
    assert "FROM v_saldos_comprometidos_detallado v" in sql, (
        "los pedidos salen de la misma vista que la pantalla /pedidos")


def test_un_pedido_viejo_no_salva_a_la_tela():
    """*"Si es hace más de 90 días asumo que quedó estancada"*: ese pedido quedó
    en la nada y la tela se estancó igual, así que vuelve a competir."""
    assert (f"DATEADD(day, -{asinfo_parado.DIAS_PEDIDO}, GETDATE())"
            in asinfo_parado._PEDIDA)
    assert asinfo_parado.DIAS_PEDIDO == 90


def test_la_tela_sin_pedido_sigue_siendo_tela_parada():
    """⚠⚠ El bug que casi entra: `ped.ultimo_pedido` viene en NULL para toda
    tela sin pedido —la enorme mayoría—, y `NOT (NULL >= fecha)` es NULL, no
    TRUE. Sin el ISNULL el CASE se iba al ELSE y la lista se quedaba sin UN SOLO
    ítem `parado`: la competencia entera pasaba a contar sólo segunda."""
    assert "ISNULL(ped.ultimo_pedido, '19000101')" in asinfo_parado._PEDIDA


def test_la_segunda_entra_aunque_la_tela_sea_reciente_o_pedida():
    """Dueña 25/08/2026: *"sí, la segunda siempre entra"*. La antigüedad y el
    pedido deciden el motivo `parado`, no el `segunda`: el pedido es de primera,
    y esos kilos siguen siendo un saldo que alguien tiene que colocar."""
    sql = " ".join(asinfo_parado.SQL_PARADOS.split())
    # la antigüedad y el pedido deciden el MOTIVO; los kilos de segunda entran
    # por el ELSE del mismo CASE, sin mirar ni una cosa ni la otra
    assert "ELSE ISNULL(cal.kg_segunda, 0) END AS stock_kg" in sql
    assert "ELSE 'segunda' END AS motivo" in sql


def _asinfo(filas):
    """Metabase contestando `filas` (y contestando OK, que es lo que mira el
    fail-closed de `_filas`)."""
    return lambda *a, **k: (filas, True)


def test_la_reciente_sin_segunda_sale_marcada_y_no_entra(monkeypatch):
    """`parados()` devuelve TODO —también lo que no entra— con dos banderas.
    Tiene que devolverlo: el refresco necesita ver la tela reciente para poder
    apagar de la lista a la que ya había entrado antes de esta regla."""
    monkeypatch.setattr(
        asinfo_parado.metabase_client, "fetch_dataset_estado",
        _asinfo([
            {"subcategoria": "Jersey Forro", "color": "NEG", "stock_kg": 0,
             "kg_segunda": 0, "motivo": "segunda", "nueva": 1,
             "stock_bodega": 900},
            {"subcategoria": "Fleece 102", "color": "AZU", "stock_kg": 300,
             "kg_segunda": 0, "motivo": "parado", "nueva": 0,
             "stock_bodega": 300},
        ]))
    nueva, vieja = asinfo_parado.parados()
    assert nueva["nueva"] is True and nueva["entra"] is False
    assert nueva["stock_bodega"] == 900, "los kilos que quedan afuera se saben"
    assert vieja["nueva"] is False and vieja["entra"] is True


def test_la_reciente_con_segunda_entra_por_sus_kilos_seg(monkeypatch):
    """Marcada como reciente, pero entra: sus kilos SEG son los que la traen, y
    `stock_kg` ya viene con esos y sólo esos."""
    monkeypatch.setattr(
        asinfo_parado.metabase_client, "fetch_dataset_estado",
        _asinfo([{"subcategoria": "Kiana", "color": "ROJ", "stock_kg": 120,
                  "kg_segunda": 120, "motivo": "segunda", "nueva": 1,
                  "stock_bodega": 940}]))
    f = asinfo_parado.parados()[0]
    assert f["nueva"] is True and f["entra"] is True
    assert f["motivo"] == "segunda" and f["stock_kg"] == 120


def test_la_lista_no_muestra_lo_apagado(monkeypatch):
    """La cohorte no borra nunca: la que nunca debió entrar queda con `fuera` en
    TRUE y la lectura la saltea. Sin el WHERE, la fila seguiría apareciendo (con
    LEFT JOIN a la foto) en 0 kg y nadie entendería qué es."""
    visto = {}

    def fake(sql, params=None, conn=None):
        visto["sql"] = " ".join(sql.split())
        return []

    monkeypatch.setattr(queries.db, "fetch_all", fake)
    queries.items()
    assert "WHERE NOT c.fuera" in visto["sql"]


class _DBFalsa:
    """La base de datos del refresco, apuntando lo que se escribe."""

    def __init__(self, cohorte, clientes=None):
        self.cohorte = cohorte
        # ⭐ `{codigo_cli: vend}` de `scintela.cliente` — desde el 31/08/2026
        # la competencia lo pide para saber a quién atribuir cada venta (ver
        # `queries._vend_por_cliente`). Vacío por default: los tests que no
        # lo necesitan siguen viendo todo como "Intela", igual que antes.
        self.clientes = clientes or []
        self.escrito: list[tuple[str, tuple]] = []
        self.leido: list[str] = []

    def tx(self):
        import contextlib
        return contextlib.nullcontext("CONN")

    def execute(self, sql, params=None, conn=None):
        self.escrito.append((" ".join(sql.split()), params))

    def fetch_all(self, sql, params=None, conn=None):
        s = " ".join(sql.split())
        self.leido.append(s)
        # ⚠ Sólo la lectura del refresco. `items()` también sale de la
        # cohorte, pero devuelve otra cosa (la foto pegada al lado).
        if s.startswith("SELECT subcategoria, color, fecha_marcado"):
            return self.cohorte
        if s.startswith("SELECT subcategoria, color, motivo"):
            return self.cohorte
        if s.startswith("SELECT codigo_cli, vend FROM scintela.cliente"):
            return self.clientes
        return []

    def fetch_one(self, sql, params=None, conn=None):
        return {}

    def sql_con(self, texto):
        return [(s, p) for s, p in self.escrito if texto in s]


def _refresco(monkeypatch, parados, cohorte):
    db = _DBFalsa(cohorte)
    monkeypatch.setattr(queries, "db", db)
    monkeypatch.setattr(queries, "today_ec", lambda: date(2026, 8, 25))
    monkeypatch.setattr(asinfo_parado, "parados", lambda: parados)
    monkeypatch.setattr(asinfo_parado, "llamados", lambda: [])
    monkeypatch.setattr(asinfo_parado, "vendido_desde", lambda d: [])
    monkeypatch.setattr(asinfo_parado, "share_por_grupo", lambda: [])
    monkeypatch.setattr(asinfo_parado, "venta_por_tela", lambda: {})
    monkeypatch.setattr(asinfo_parado, "formas", lambda: {})
    monkeypatch.setattr(asinfo_parado, "ultima_venta_antes", lambda d: {})
    return db, queries.actualizar()


def test_el_refresco_apaga_la_tela_reciente_y_no_la_da_de_alta(monkeypatch):
    """La que no está estancada no entra a la cohorte, y si ya estaba (entró el
    13/08, antes de que existiera esta regla) se APAGA. Apagar y no borrar: la
    cohorte es la única tabla que no se puede recalcular."""
    db, res = _refresco(
        monkeypatch,
        parados=[
            {"subcategoria": "Fleece 102", "color": "AZU", "stock_kg": 300,
             "stock_bodega": 300, "motivo": "parado", "nueva": False,
             "entra": True},
            {"subcategoria": "Jersey Forro", "color": "NEG", "stock_kg": 0,
             "stock_bodega": 900, "motivo": "segunda", "nueva": True,
             "entra": False},
        ],
        cohorte=[{"subcategoria": "Fleece 102", "color": "AZU",
                  "fecha_marcado": date(2026, 8, 13), "motivo": "parado"}])

    altas = db.sql_con("INSERT INTO scintela.parado_cohorte")
    assert [p[0] for _, p in altas] == ["Fleece 102"], (
        "la tela reciente no puede darse de alta en la cohorte")
    apagadas = db.sql_con("SET fuera = TRUE")
    assert [p for _, p in apagadas] == [("Jersey Forro", "NEG")]
    assert res["nuevas"] == 1 and res["nuevas_kg"] == 900


def test_la_apagada_vuelve_sola_cuando_cumple_los_meses(monkeypatch):
    """No hace falta tocar nada: el día que sus rollos pasan los 6 meses, Asinfo
    la devuelve como parada y el refresco la vuelve a encender — con la fecha en
    que se marcó, porque nunca se fue de la cohorte."""
    db, _ = _refresco(
        monkeypatch,
        parados=[{"subcategoria": "Jersey Forro", "color": "NEG",
                  "stock_kg": 900, "stock_bodega": 900, "motivo": "parado",
                  "nueva": False, "entra": True}],
        cohorte=[])
    assert [p for _, p in db.sql_con("SET fuera = FALSE")] == [
        ("Jersey Forro", "NEG")]
    assert not db.sql_con("SET fuera = TRUE")


def test_la_foto_y_la_competencia_no_miran_lo_apagado(monkeypatch):
    """`parado_foto` y `parado_venta` se rehacen desde la cohorte: si la lectura
    no filtrara, la tela apagada seguiría con foto y seguiría dando puntos."""
    db, _ = _refresco(monkeypatch, parados=[], cohorte=[])
    lecturas = [s for s in db.leido
                if s.startswith("SELECT subcategoria, color, fecha_marcado")]
    assert lecturas and all(
        s.endswith("FROM scintela.parado_cohorte WHERE NOT fuera")
        for s in lecturas), lecturas


def test_la_meta_y_los_puntos_se_congelan_sobre_lo_que_se_acaba_de_escribir(
        monkeypatch):
    """⚠ `items()` tiene que leer con la conexión de la transacción. Sin eso, la
    meta y el puntaje se congelan sobre la foto ANTERIOR —la única commiteada— y
    quedan calculados sobre un universo que ya no existe: justo lo que pasaría
    el día que se saca la tela reciente."""
    vistas = []
    monkeypatch.setattr(queries, "items", lambda conn=None: vistas.append(conn) or [])
    _refresco(monkeypatch, parados=[], cohorte=[])
    assert vistas and all(c == "CONN" for c in vistas), vistas


def test_el_refresco_cuenta_aparte_las_recientes_y_las_pedidas(monkeypatch):
    """En la pantalla no significan lo mismo: la tela reciente vuelve sola
    cuando cumpla los días, la pedida cuando el pedido salga. Un solo número
    sumado no dejaría explicar ninguna de las dos."""
    db, res = _refresco(
        monkeypatch,
        parados=[
            {"subcategoria": "Toper", "color": "COA", "stock_kg": 0,
             "stock_bodega": 257, "motivo": "segunda", "nueva": True,
             "pedida": False, "entra": False},
            {"subcategoria": "Pique Nido", "color": "CRO", "stock_kg": 0,
             "stock_bodega": 195, "motivo": "segunda", "nueva": False,
             "pedida": True, "entra": False},
            {"subcategoria": "Fleece 102", "color": "AZU", "stock_kg": 300,
             "stock_bodega": 300, "motivo": "parado", "nueva": False,
             "pedida": False, "entra": True},
        ],
        cohorte=[{"subcategoria": "Fleece 102", "color": "AZU",
                  "fecha_marcado": date(2026, 8, 13), "motivo": "parado"}])

    assert (res["nuevas"], res["nuevas_kg"]) == (1, 257)
    assert (res["pedidas"], res["pedidas_kg"]) == (1, 195)
    apagadas = {p for _, p in db.sql_con("SET fuera = TRUE")}
    assert apagadas == {("Toper", "COA"), ("Pique Nido", "CRO")}


def test_la_pedida_tambien_pierde_el_motivo_congelado(monkeypatch):
    """Un ítem que entró como `parado` el 13/08 y hoy resulta que está pedido no
    puede seguir sumando sus kilos de PRIMERA: nunca debieron contar. Es la
    misma corrección que para la tela reciente."""
    db, _ = _refresco(
        monkeypatch,
        parados=[{"subcategoria": "Jersey 115", "color": "ROJ", "stock_kg": 40,
                  "stock_bodega": 130, "motivo": "segunda", "nueva": False,
                  "pedida": True, "entra": True}],
        cohorte=[])
    motivos = db.sql_con("SET motivo = %s")
    assert motivos and motivos[0][1] == ("segunda", "Jersey 115", "ROJ", True)


def test_la_tela_que_ya_se_vendio_no_pierde_su_lugar(monkeypatch):
    """⭐ El invariante de la dueña: "si empezamos a venderlas, que no se nos
    vayan de la lista". Vender lo parado es lo que la competencia premia, así
    que la tela que se vendió entera —sin stock y sin banderas— se queda."""
    db, _ = _refresco(
        monkeypatch,
        parados=[{"subcategoria": "Fleece 102", "color": "AZU", "stock_kg": 0,
                  "stock_bodega": 0, "motivo": "parado", "nueva": False,
                  "pedida": False, "entra": False}],
        cohorte=[{"subcategoria": "Fleece 102", "color": "AZU",
                  "fecha_marcado": date(2026, 8, 13), "motivo": "parado"}])
    assert not db.sql_con("SET fuera = TRUE"), (
        "sin stock pero sin bandera: se vendió, y ésa es la que hay que premiar")


def test_la_reciente_que_se_vendio_entera_igual_pierde_los_puntos(monkeypatch):
    """El agujero que destapó la largada: Intela arrancó con 640 puntos y 554
    eran UNA venta de Jersey 3 BLA. Apenas se vende, la tela pasa a tener ventas
    en 12 meses y deja de figurar como parada — si el refresco sólo mirara la
    lista de hoy, no la vería nunca más y los puntos quedarían puestos.

    Por eso las banderas se calculan para toda la bodega, vendida o no."""
    db, res = _refresco(
        monkeypatch,
        parados=[{"subcategoria": "Jersey 3", "color": "BLA", "stock_kg": 0,
                  "stock_bodega": 0, "motivo": "segunda", "nueva": True,
                  "pedida": False, "entra": False}],
        cohorte=[{"subcategoria": "Jersey 3", "color": "BLA",
                  "fecha_marcado": date(2026, 8, 13), "motivo": "parado"}])
    assert [p for _, p in db.sql_con("SET fuera = TRUE")] == [("Jersey 3", "BLA")]
    assert res["nuevas"] == 1


def test_item_parado_que_vende_sigue_mostrando_su_stock_completo(monkeypatch):
    """Dueña 31/08/2026: *"si teniamos sin vender desde 2024 y se vendio con la
    competencia, tiene que seguir contando. si no cada vez que vendemos un
    rollo va a dejar de contar"*.

    Fleece Fancy FNB: parada desde 2024, entro a la cohorte como 'parado'.
    Vendio 95 de 448,5 kg por la propia competencia -eso hace que HOY
    `asinfo_parado.parados()` ya no la vea "sin venta en 12 meses" y le
    calcule motivo='segunda'/stock_kg=0 para esta vuelta (medido en vivo el
    31/08/2026: la pantalla mostraba "Queda: vendido" con 353,65 kg de
    primera todavia en bodega 53). La cohorte tiene que seguir mandando: la
    foto debe mostrar esos 353,65 kg, no 0."""
    db, _ = _refresco(
        monkeypatch,
        parados=[{"subcategoria": "Fleece Fancy", "color": "FNB",
                  # Lo que calcularia HOY `_ES_PARADO` (ya vendio, cae a
                  # 'segunda'/stock_kg=0) -- y los crudos SIN el CASE WHEN,
                  # que son los que la cohorte necesita para no perderlos.
                  "stock_kg": 0, "motivo": "segunda", "nueva": False,
                  "pedida": False, "entra": False,
                  "stock_bodega": 353.65, "kg_segunda": 0,
                  "kg_primera_bodega": 353.65, "kg_tub_pri_bodega": 353.65,
                  "kg_tub_seg": 0, "kg_abi_pri_bodega": 0, "kg_abi_seg": 0,
                  "categoria": "Fleece"}],
        cohorte=[{"subcategoria": "Fleece Fancy", "color": "FNB",
                  "fecha_marcado": date(2026, 8, 17), "motivo": "parado"}])
    fotos = [p for sql, p in db.escrito
             if "INSERT INTO scintela.parado_foto" in sql]
    assert len(fotos) == 1
    stock_kg, kg_primera, motivo = fotos[0][2], fotos[0][7], fotos[0][10]
    assert stock_kg == 353.65, "el motivo CONGELADO tiene que mandar, no el de hoy"
    assert kg_primera == 353.65
    assert motivo == "parado"


def test_item_segunda_de_verdad_sigue_mostrando_solo_su_segunda(monkeypatch):
    """El espejo del test de arriba: un color que vende bien (NEG) y solo entro
    a la cohorte por su segunda tiene que seguir mostrando SOLO esa segunda,
    no toda la bodega -- el fix de FNB no puede hacer que la primera que se
    vende sola vuelva a contar."""
    db, _ = _refresco(
        monkeypatch,
        parados=[{"subcategoria": "Fleece 96 Perchado", "color": "NEG",
                  "stock_kg": 22, "motivo": "segunda", "nueva": False,
                  "pedida": False, "entra": True,
                  "stock_bodega": 7112.35, "kg_segunda": 22,
                  "kg_primera_bodega": 7090.35, "kg_tub_pri_bodega": 7090.35,
                  "kg_tub_seg": 22, "kg_abi_pri_bodega": 0, "kg_abi_seg": 0,
                  "categoria": "Fleece"}],
        cohorte=[{"subcategoria": "Fleece 96 Perchado", "color": "NEG",
                  "fecha_marcado": date(2026, 8, 18), "motivo": "segunda"}])
    fotos = [p for sql, p in db.escrito
             if "INSERT INTO scintela.parado_foto" in sql]
    assert len(fotos) == 1
    stock_kg, kg_primera = fotos[0][2], fotos[0][7]
    assert stock_kg == 22, "'segunda' tiene que seguir mostrando SOLO su segunda"
    assert kg_primera == 0


def test_las_tarjetas_se_leen_de_corrido(monkeypatch):
    """Dueña 25/08/2026: *"si teníamos 54k, lo que se vendió no puede ser 1k. o
    deberíamos tener meta inicial, vendido, cuánto queda actualmente"*.

    Al arrancar − vendido = queda. Y el arranque son los kilos CONGELADOS de la
    largada, no una reconstrucción con la foto de hoy: si no, el número de esta
    pantalla y el de la competencia se separan en cuanto la bodega se mueva."""
    filas = [{"stock_kg": 53441, "kg_vendidos": 1281, "kg_segunda": 0}]
    r = queries.resumen(filas, kg_base=54722)
    assert (r["kg_inicial"], r["kg_vendidos"], r["kg"]) == (54722, 1281, 53441)
    assert r["kg_movido"] == 0

    # sin base congelada todavía, el arranque se reconstruye y cierra igual
    r = queries.resumen(filas)
    assert r["kg_inicial"] == 54722 and r["kg_movido"] == 0

    # y lo que la resta no explica queda a la vista
    r = queries.resumen(filas, kg_base=55000)
    assert r["kg_movido"] == 278


def test_salido_antes_de_la_largada_no_es_movido_ni_puntua(monkeypatch):
    """⚠⚠ DECISIÓN 31/08/2026. Un ítem marcado el 17/08 (ocho días antes de
    la largada del 25/08) pudo haber vendido algo en el medio: ese kilo salió
    de la bodega de verdad —"Queda" ya no lo tiene— pero `kg_vendidos` sólo
    cuenta desde la largada, así que antes cerraba como `kg_movido`: tela que
    parecía perdida sin estarlo. Kiana Forro 1.45 LIF, real: 534,65 kg.

    Tamara, con el caso puesto adelante: *"si ya estuvo en el listado,
    permanece… los vendedores no pueden que desaparezca de vez en cuando las
    cosas"* — pero también *"si se vendieron antes del 25 no cuentan para
    nadie"*. Las dos juntas piden un casillero aparte: se explica, no puntúa.

    Un ítem marcado DESPUÉS de la largada no tiene este colchón: ahí un
    residuo sigue siendo `kg_movido` de verdad, la alarma que hay que mirar."""
    filas_antes = [{"stock_kg": 45.5, "kg_vendidos": 0, "kg_segunda": 0,
                     "kg_al_marcar": 580.15,
                     "fecha_marcado": date(2026, 8, 17)}]
    r = queries.resumen(filas_antes, kg_base=580.15, largada=date(2026, 8, 25))
    assert r["kg_salido_antes"] == 534.65
    assert r["kg_movido"] == 0
    # ⭐ SEGUNDA VUELTA, 31/08/2026 (misma tarde): "kg_vendidos" (lo que
    # pinta la tarjeta) sigue siendo PURO — 0 acá, porque este ítem no vendió
    # nada DESDE la largada. El ajuste para que la resta cierre igual vive en
    # "Al arrancar" (`kg_inicial`), que baja retroactivo en vez de que
    # "Vendido" suba: así sigue siendo el MISMO número que
    # /analisis/competencia (Tamara: "basta de numeros distintos por todos
    # lados"). "Al arrancar" queda en 45,5 kg — lo único que esta tela trae
    # a la carrera es lo que tiene hoy, porque lo demás ya se había ido
    # antes de arrancar.
    assert r["kg_vendidos"] == 0
    assert r["kg_inicial"] == 45.5
    assert r["kg_inicial"] - r["kg_vendidos"] - r["kg"] == r["kg_movido"] == 0

    # el mismo residuo, pero SIN pasarle `largada` (compatibilidad hacia
    # atrás: las pantallas que no la mandan ven el comportamiento de siempre)
    r = queries.resumen(filas_antes, kg_base=580.15)
    assert r["kg_salido_antes"] == 0
    assert r["kg_vendidos"] == 0
    assert r["kg_movido"] == 534.65

    # marcado DESPUÉS de la largada: el residuo NO tiene colchón, es
    # kg_movido de verdad
    filas_despues = [{"stock_kg": 45.5, "kg_vendidos": 0, "kg_segunda": 0,
                       "kg_al_marcar": 580.15,
                       "fecha_marcado": date(2026, 8, 26)}]
    r = queries.resumen(filas_despues, kg_base=580.15, largada=date(2026, 8, 25))
    assert r["kg_salido_antes"] == 0
    assert r["kg_movido"] == 534.65


def test_movidos_no_cuenta_lo_que_solo_salio_antes():
    """⚠⚠ REVIERTE a propósito el intento de la primera vuelta (mismo día): el
    conteo de ítems de la tarjeta Vendido vuelve a ser SÓLO los que vendieron
    algo DESDE la largada — igual que `kg_vendidos`, igual que Competencia.
    Un ítem que sólo tiene `kg_salido_antes` (vendió todo ANTES del 25/08) no
    suma acá: ese kilo no está en "Vendido", está restado de "Al arrancar"."""
    filas = [
        {"stock_kg": 0, "kg_vendidos": 5, "kg_segunda": 0},  # vendió después
        {"stock_kg": 45.5, "kg_vendidos": 0, "kg_segunda": 0,  # solo antes
         "kg_al_marcar": 580.15, "fecha_marcado": date(2026, 8, 17)},
        {"stock_kg": 10, "kg_vendidos": 0, "kg_segunda": 0},  # no se movió
    ]
    r = queries.resumen(filas, largada=date(2026, 8, 25))
    assert r["movidos"] == 1


def test_al_arrancar_vivo_absorbe_una_tela_nueva_sin_dejar_seg_nueva():
    """⭐ DECISIÓN 31/08/2026 (tarde). Antes, "Al arrancar" era una foto
    congelada por categoría el día de la largada (`kg_al_arrancar()`): una
    tela que se sumaba a la cohorte DESPUÉS nunca entraba ahí, y su stock
    aparecía en "Queda" sin haber estado nunca en "Al arrancar" — el residuo
    salía como `kg_movido` negativo ("kg de SEG nueva"), un asterisco que
    Tamara no quiere ver: *"yo quiero que esta matematica funcione, habia 50k
    se vendio 6k no quedan 46k. a corregir"*.

    Con `kg_al_marcar_vivo()` (suma de `kg_al_marcar` sobre la cohorte de
    HOY, no una foto de un solo día) la tela nueva trae SU PROPIO punto de
    partida y la resta cierra sola."""
    vieja = {"stock_kg": 400, "kg_vendidos": 600, "kg_segunda": 0,
             "kg_al_marcar": 1000, "fecha_marcado": date(2026, 8, 17)}
    nueva = {"stock_kg": 200, "kg_vendidos": 0, "kg_segunda": 200,
             "kg_al_marcar": 200, "fecha_marcado": date(2026, 9, 1)}
    filas = [vieja, nueva]

    kg_base_vivo = queries.kg_al_marcar_vivo(filas)
    assert kg_base_vivo == 1200  # 1.000 de la vieja + 200 de la nueva

    r = queries.resumen(filas, kg_base_vivo, largada=date(2026, 8, 25))
    assert r["kg_inicial"] == 1200
    assert r["kg_vendidos"] == 600
    assert r["kg"] == 600
    assert r["kg_movido"] == 0, "la resta tiene que cerrar sola, sin SEG nueva"

    # ⚠ El PISO de `resumen()` (ver más abajo) es una segunda red: incluso con
    # la foto VIEJA (congelada antes de que la tela nueva se sumara, el bug
    # que este test reproduce), la resta ya no queda en descubierto — "Al
    # arrancar" sube al piso de lo que hoy se puede explicar. Antes de que
    # existiera el piso esto daba -200 ("kg de SEG nueva"); con las dos
    # defensas juntas (kg_al_marcar_vivo + el piso) da 0 aunque el llamador
    # se equivoque de foto.
    r_viejo = queries.resumen(filas, kg_base=1000, largada=date(2026, 8, 25))
    assert r_viejo["kg_movido"] == 0


def test_el_piso_absorbe_segunda_que_sigue_entrando_a_tela_ya_marcada():
    """El caso REAL, verificado por SQL en vivo el 01/09/2026: no es sólo una
    tela nueva entrando a la cohorte (el test de arriba) — una tela marcada
    hace semanas (18/08/2026, una semana antes de la largada) sigue
    recibiendo SEGUNDA todos los días, y su stock de HOY supera lo que tenía
    anotado al marcarse. Ej. real: Toper AZN, marcado con 18,9 kg, con 59,9
    kg en bodega hoy — 41 kg de SEGUNDA que entraron después, no una venta al
    revés. Sin el piso, `kg_al_marcar_vivo()` solo (el fix anterior) NO
    alcanza para este caso: la tela YA estaba en la cohorte antes de la
    largada, así que no hay "tela nueva" que sumar — el hueco está adentro
    de una tela vieja."""
    filas = [{"stock_kg": 59.9, "kg_vendidos": 18.9, "kg_segunda": 59.9,
              "kg_al_marcar": 18.9, "fecha_marcado": date(2026, 8, 18)}]
    kg_base_vivo = queries.kg_al_marcar_vivo(filas)
    assert kg_base_vivo == 18.9  # lo único que sabíamos el día que se marcó

    r = queries.resumen(filas, kg_base_vivo, largada=date(2026, 8, 25))
    assert r["kg_inicial"] == 78.8  # el piso: kg (59,9) + vendido (18,9)
    assert r["kg_movido"] == 0, "la segunda que entra no es una alarma"


def test_la_pantalla_muestra_las_tres_cifras():
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "parado.html").read_text(encoding="utf-8")
    for k in ("Al arrancar", "Vendido", "Queda"):
        assert f'class="k">{k}</span>' in html
    assert "{{ resumen.kg_inicial | num_es(0) }}" in html
    assert "resumen.kg_movido" in html


def _venta(sub, col, kg, dia=25, calidad="PRI", vend="Intela", codigo_cli=None):
    return {"subcategoria": sub, "color": col, "kg": kg, "calidad": calidad,
            "vendedor": vend, "vend_pc": None, "codigo_cli": codigo_cli,
            "fecha": date(2026, 8, dia)}


def _refresco_con_ventas(monkeypatch, parados, cohorte, ventas, clientes=None,
                          devolver_db=False):
    db = _DBFalsa(cohorte, clientes=clientes)
    monkeypatch.setattr(queries, "db", db)
    monkeypatch.setattr(queries, "today_ec", lambda: date(2026, 8, 25))
    monkeypatch.setattr(asinfo_parado, "parados", lambda: parados)
    monkeypatch.setattr(asinfo_parado, "llamados", lambda: [])
    monkeypatch.setattr(asinfo_parado, "vendido_desde", lambda d: ventas)
    monkeypatch.setattr(asinfo_parado, "share_por_grupo", lambda: [])
    monkeypatch.setattr(asinfo_parado, "venta_por_tela", lambda: {})
    monkeypatch.setattr(asinfo_parado, "formas", lambda: {})
    monkeypatch.setattr(asinfo_parado, "ultima_venta_antes", lambda d: {})
    queries.actualizar()
    filas = [(p[1], p[5], p[7]) for sql, p in db.escrito
             if "INSERT INTO scintela.parado_venta" in sql]
    # ⭐ `devolver_db`: para el test que necesita mirar `vend_pc`/`vendedor`
    # (índices 2 y 3 de los mismos params) sin romper a los que ya usan la
    # forma corta (color, kg, cuenta).
    return (db, filas) if devolver_db else filas


def _item(sub, col, kg_antes, stock=300):
    return {"subcategoria": sub, "color": col, "stock_kg": stock,
            "stock_bodega": stock, "motivo": "parado", "nueva": False,
            "pedida": False, "entra": True, "kg_antes": kg_antes}


def _cohorte(sub, col):
    return [{"subcategoria": sub, "color": col,
             "fecha_marcado": date(2026, 8, 13), "motivo": "parado"}]


def test_no_se_puede_sacar_mas_de_lo_que_habia_parado(monkeypatch):
    """Dueña 25/08/2026: *"está mal que sigue contando una tela que había 0 en
    saldo"*. Jersey 3 BLA tenía unos kilos viejos de blanco y 490 tejidos el
    17/07: la regla de la antigüedad mira el ÍTEM, así que lo dejaba entrar
    entero y los kilos de julio puntuaban como si hubieran estado clavados.

    El tope son los kilos que ya estaban al corte. Lo que se vende por encima se
    vende igual, pero no destraba nada: va como un renglón que no cuenta."""
    filas = _refresco_con_ventas(
        monkeypatch,
        parados=[_item("Jersey 3", "BLA", kg_antes=30)],
        cohorte=_cohorte("Jersey 3", "BLA"),
        ventas=[_venta("Jersey 3", "BLA", 554)])
    assert ("BLA", 30.0, True) in filas
    assert ("BLA", 524.0, False) in filas


def test_la_venta_que_entra_entera_en_el_tope_no_se_parte(monkeypatch):
    filas = _refresco_con_ventas(
        monkeypatch,
        parados=[_item("Fleece 102", "AZU", kg_antes=300)],
        cohorte=_cohorte("Fleece 102", "AZU"),
        ventas=[_venta("Fleece 102", "AZU", 120)])
    assert filas == [("AZU", 120.0, True)]


def test_el_tope_se_gasta_en_orden_de_fecha(monkeypatch):
    """Dos ventas y un tope que sólo alcanza para la primera y media: la que
    llegó antes se lleva los kilos. Si se repartiera en el orden en que Asinfo
    devuelve las filas, el puntaje cambiaría de un refresco a otro sin que nadie
    hubiera vendido nada."""
    filas = _refresco_con_ventas(
        monkeypatch,
        parados=[_item("Toper", "COA", kg_antes=100)],
        cohorte=_cohorte("Toper", "COA"),
        ventas=[_venta("Toper", "COA", 60, dia=26),
                _venta("Toper", "COA", 80, dia=25)])
    assert filas[0] == ("COA", 80.0, True)          # la del 25 entra entera
    assert filas[1] == ("COA", 20.0, True)          # de la del 26 entran 20
    assert filas[2] == ("COA", 40.0, False)


def test_la_devolucion_resta_aunque_el_tope_este_gastado(monkeypatch):
    """Si no, un vendedor factura hasta el tope, el cliente devuelve la tela y
    los puntos le quedan puestos."""
    filas = _refresco_con_ventas(
        monkeypatch,
        parados=[_item("Toper", "COA", kg_antes=50)],
        cohorte=_cohorte("Toper", "COA"),
        ventas=[_venta("Toper", "COA", 50, dia=25),
                _venta("Toper", "COA", -20, dia=26)])
    assert filas == [("COA", 50.0, True), ("COA", -20.0, True)]


def test_el_item_que_asinfo_ya_no_devuelve_conserva_su_tope(monkeypatch):
    """El ítem que se vendió entero —o que salió de la bodega— desaparece de la
    consulta de parados: `todas` es lo que hay HOY. Sin tope propio, contaba
    TODO lo que se vendiera de esa tela × color hasta el cierre, así que el día
    que se teje de nuevo esos kilos nuevos puntúan como si hubieran estado
    clavados. Es el agujero de Jersey 3 BLA por la puerta de atrás.

    El tope existe igual y es nuestro: los kilos que el ítem tenía el día que
    entró a la lista. Medidos el 25/08/2026: 9 ítems, 1.129 kg al marcar."""
    cohorte = [{"subcategoria": "Toper", "color": "COA",
                "fecha_marcado": date(2026, 8, 13), "motivo": "parado",
                "kg_al_marcar": 40}]
    filas = _refresco_con_ventas(
        monkeypatch, parados=[], cohorte=cohorte,
        ventas=[_venta("Toper", "COA", 400)])
    assert ("COA", 40.0, True) in filas, "perdió el tope que le pusimos nosotros"
    assert ("COA", 360.0, False) in filas


def test_segunda_de_lote_real_puntua_aunque_la_factura_diga_primera(monkeypatch):
    """⭐ Kiana Forro 1.45 LIF: el lote de bodega era SEGUNDA (580,15 kg al
    marcar) pero las 17 facturas que se lo llevaron salieron cargadas con el
    atributo de calidad en PRIMERA — verificado contra Asinfo. El 31/08 se
    arregló contando CUALQUIER calidad (reabrió otro agujero, cerrado el
    01/09 — ver `test_una_primera_de_verdad_no_puntua_aunque_llene_el_tope`).
    El arreglo de fondo es más arriba, en `vendido_desde()`: `calidad` YA
    viene resuelta contra el LOTE REAL del despacho, no de la factura. Acá,
    en `cuenta_el_kilo`, ese kilo llega como "SEG" —la verdad del lote—, así
    que puntúa igual sin necesitar la excepción de "cualquier calidad".

    La prueba de que esto NO reabre el agujero del 24/08 ("no darle puntos a
    la tela que se vende sola"): el TOPE es el kg de SEGUNDA que había al
    marcar (50), no el kg_antes (5.000, que es el saldo TOTAL —primera y
    segunda— de 90 días atrás, deliberadamente grande acá). Aunque la venta
    sea de 200 kg, sólo 50 puntúan."""
    cohorte = [{"subcategoria": "Kiana Forro 1.45", "color": "LIF",
                "fecha_marcado": date(2026, 8, 13), "motivo": "segunda",
                "kg_al_marcar": 50}]
    item = {"subcategoria": "Kiana Forro 1.45", "color": "LIF",
            "stock_kg": 10, "stock_bodega": 10, "motivo": "segunda",
            "nueva": False, "pedida": False, "entra": True, "kg_antes": 5000}
    filas = _refresco_con_ventas(
        monkeypatch, parados=[item], cohorte=cohorte,
        ventas=[_venta("Kiana Forro 1.45", "LIF", 200, calidad="SEG")])
    assert ("LIF", 50.0, True) in filas, "el kilo del lote real SEG puntúa"
    assert ("LIF", 150.0, False) in filas, "pero el tope de segunda lo frena"


def test_una_primera_de_verdad_no_puntua_aunque_llene_el_tope(monkeypatch):
    """⭐ CERRADO 01/09/2026. Dueña, mirando /analisis/competencia: "porque
    estan en la competencia?" — Fleece 96 Sin Perchar NEG, Rib Normal BLA y
    ~40 ítems más entraban por un resto chico de SEGUNDA (12-130 kg) pegado a
    una tela que se vende sola en PRIMERA (600-3.800 kg reales desde la
    largada). Como cualquier calidad contaba, el tope se llenaba en 1-2 días
    con ventas normales — puntaje completo sin mover un gramo de la segunda,
    que seguía intacta en bodega. Verificado con Metabase: el lote real de
    esas ventas ERA primera de verdad (no un error de carga como Kiana
    Forro) — `dueña: "era de segunda la tela que salio en esas facturas? si
    era de segunda poner en cal = seg"`. No lo era, así que no cuenta."""
    cohorte = [{"subcategoria": "Fleece 96 Sin Perchar", "color": "NEG",
                "fecha_marcado": date(2026, 8, 18), "motivo": "segunda",
                "kg_al_marcar": 44.55}]
    item = {"subcategoria": "Fleece 96 Sin Perchar", "color": "NEG",
            "stock_kg": 44.55, "stock_bodega": 44.55, "motivo": "segunda",
            "nueva": False, "pedida": False, "entra": True, "kg_antes": 44.55}
    filas = _refresco_con_ventas(
        monkeypatch, parados=[item], cohorte=cohorte,
        ventas=[_venta("Fleece 96 Sin Perchar", "NEG", 679, calidad="PRI")])
    assert filas == [("NEG", 679.0, False)], (
        "toda la venta tiene que caer del lado que NO puntúa — nada de "
        "primera de verdad puede llenar el tope de segunda")


def test_sin_dato_de_asinfo_no_hay_tope(monkeypatch):
    """Fail-open a propósito: un ítem que falte en la consulta dejaría a alguien
    sin sus puntos y nadie sabría por qué."""
    item = _item("Toper", "COA", kg_antes=0)
    del item["kg_antes"]
    filas = _refresco_con_ventas(
        monkeypatch, parados=[item], cohorte=_cohorte("Toper", "COA"),
        ventas=[_venta("Toper", "COA", 400)])
    assert filas == [("COA", 400.0, True)]


def test_la_tela_que_se_sigue_tejiendo_no_es_un_saldo():
    """Dueña 25/08/2026, sobre Jersey 3 BLA —kilos viejos de blanco en bodega y
    una orden del 17/07—: *"pero no es saldo si se seguía produciendo ese color x
    tela"*. Que haya kilos viejos no lo convierte en saldo si la fábrica sigue
    tejiéndolo: es un producto vivo, y pagar puntos por venderlo es pagar por una
    venta normal."""
    assert "NOT " + asinfo_parado._EN_PRODUCCION in asinfo_parado._ES_PARADO
    sql = " ".join(asinfo_parado.SQL_PARADOS.split())
    assert "FROM orden_fabricacion o" in sql
    assert (f"DATEADD(day, -{asinfo_parado.DIAS_PRODUCCION}, GETDATE())"
            in asinfo_parado._EN_PRODUCCION)


def test_la_orden_abandonada_no_cuenta_como_produccion():
    """⚠ `estado_produccion = 0` no es "programada", es ABANDONADA: 894 órdenes
    que cuelgan de un padre también en 0 y promedian 660 días. Contarlas dejaría
    media fábrica marcada como viva."""
    sql = " ".join(asinfo_parado.SQL_PARADOS.split())
    assert "o.estado_produccion <> 0" in sql


def test_la_tela_sin_ninguna_orden_sigue_compitiendo():
    """El mismo NULL que en los pedidos: sin ISNULL, la tela que nunca se tejió
    contesta NULL y arrastra toda la regla."""
    assert "ISNULL(fab.ultima_orden, '19000101')" in asinfo_parado._EN_PRODUCCION


def test_el_refresco_apaga_tambien_la_que_se_sigue_tejiendo(monkeypatch):
    db, res = _refresco(
        monkeypatch,
        parados=[{"subcategoria": "Jersey 3", "color": "BLA", "stock_kg": 0,
                  "stock_bodega": 0, "motivo": "segunda", "nueva": False,
                  "pedida": False, "produciendo": True, "entra": False}],
        cohorte=[{"subcategoria": "Jersey 3", "color": "BLA",
                  "fecha_marcado": date(2026, 8, 13), "motivo": "parado"}])
    assert [p for _, p in db.sql_con("SET fuera = TRUE")] == [("Jersey 3", "BLA")]
    assert res["produciendo"] == 1 and res["nuevas"] == 0


def test_los_tres_motivos_se_cuentan_por_separado_en_el_refresco():
    """No significan lo mismo: la reciente vuelve sola con el tiempo, la que se
    teje cuando la fábrica pare, la pedida cuando salga el pedido. Un solo
    número sumado no dejaría explicar ninguna de las tres.

    ⚠ El 26/08/2026 la línea se fue de la PANTALLA ("no nos hace falta eso"),
    pero la cuenta se sigue haciendo y guardando en `parado_refresh`: el día que
    la lista se desplome hay que poder decir por qué."""
    import inspect as _i
    fuente = _i.getsource(queries.actualizar)
    for cuenta in ('"nuevas"', '"pedidas"', '"produciendo"'):
        assert cuenta in fuente, f"el refresco dejó de contar {cuenta}"
    assert "nuevas_kg" in fuente and "pedidas_kg" in fuente

def test_la_fila_en_cero_que_nunca_se_movio_no_se_muestra(monkeypatch):
    """Dueña 25/08/2026: *"sacar las que están en 0 también"*. Sin kilos en
    bodega y sin un kilo vendido no hay nada que ofrecer ni nada que mostrar: la
    tela se fue por un ajuste, un traslado o un recuento.

    ⚠ Pero la que se VENDIÓ se queda aunque esté en 0: es la que muestra que la
    competencia funcionó.

    ⚠⚠ DECISIÓN 31/08/2026: y la que se vendió ANTES de la largada TAMBIÉN se
    queda, aunque `kg_vendidos` (que sólo cuenta desde la largada) la vea en
    0/0. Kiana Forro 1.45 LIF quedó 0 y 0 con este filtro sin la tercera
    condición, y desaparecía ENTERA de la pantalla — 534,65 kg vendidos de
    verdad, invisibles. La tercera condición la salva: si `kg_al_marcar` no
    cierra contra stock+vendido, queda un residuo real y la fila se queda."""
    visto = {}

    def fake(sql, params=None, conn=None):
        visto["sql"] = " ".join(sql.split())
        return []

    monkeypatch.setattr(queries.db, "fetch_all", fake)
    queries.items()
    assert ("AND (COALESCE(f.stock_kg, 0) > 0 OR COALESCE(f.kg_vendidos, 0) > 0 "
            "OR c.kg_al_marcar > COALESCE(f.stock_kg, 0) + "
            "COALESCE(f.kg_vendidos, 0) + 0.01)" in visto["sql"])


def test_la_tabla_de_vendidos_dice_dia_y_vendedor(monkeypatch):
    """Dueña 25/08/2026: *"abajo de saldos poné una tabla que se llame vendidos,
    y ponés la misma info + día vendido + vendedor"*. El total estaba en una
    tarjeta; qué, cuándo y quién había que ir a buscarlo a la Competencia."""
    visto = {}

    def fake(sql, params=None, conn=None):
        visto["sql"] = " ".join(sql.split())
        visto["params"] = params
        return []

    monkeypatch.setattr(queries.db, "fetch_all", fake)
    queries.vendidos("2026-08-25")
    assert "FROM scintela.parado_venta v" in visto["sql"]
    assert "v.fecha, v.vendedor, v.vend_pc" in visto["sql"]
    assert "WHERE v.fecha >= %s AND v.cuenta" in visto["sql"], (
        "sólo lo que puntúa, igual que en la Competencia")
    assert visto["params"] == ("2026-08-25",)


def test_la_pantalla_dibuja_la_tabla_de_vendidos():
    """Dueña 25/08/2026: *"mantené el orden y formato de las de arriba, podés
    reemplazar clientes por vendedor"*. Dos tablas que muestran lo mismo con
    otro orden se leen como dos cosas distintas: las columnas van iguales y en
    el mismo orden, y sólo cambian las de la derecha, que acá significan otra
    cosa (Clientes → Vendedor, y el Día de la venta).

    ⚠ Arriba se fue «Última venta» el 25/08/2026 ("última venta se va de la
    tabla"), así que la de arriba tiene una columna menos. Lo que este test
    cuida es que las que están se lean juntas: las cinco primeras iguales, y
    «lo que queda | lo que salió» en el mismo par de columnas."""
    import inspect as _i
    import re as _re
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "vendidos.html").read_text(encoding="utf-8")
    assert 'id="vendidos-tabla"' in html and ">Vendidos</h2>" in html

    # ⚠ Desde el 26/08/2026 las dos tablas viven en pantallas DISTINTAS —
    # Vendidos se mudó a su tab—, y por eso este test vale más que antes: nadie
    # las va a ver juntas para notar que se despegaron.
    def cabeceras(donde: str, tabla: str) -> list[str]:
        trozo = donde[donde.index(tabla):]
        cab = trozo[trozo.index("<thead>"):trozo.index("</thead>")]
        # ⚠ `<th(?:\s[^>]*)?>` y no `<th[^>]*>`: el segundo matchea `<thead>`
        # y la primera columna sale con media etiqueta adentro.
        return [" ".join(t.split())
                for t in _re.findall(r"<th(?:\s[^>]*)?>(.*?)</th>", cab, _re.S)]

    arriba = cabeceras(_html_parado(), '<table id="tabla">')
    abajo = cabeceras(html, '<table class="resumen" id="vendidos">')
    assert arriba[:5] == abajo[:5], (
        f"el orden de las cinco primeras no coincide: {arriba[:5]} vs {abajo[:5]}")
    # ⭐ Y las dos del medio significan lo MISMO en las dos tablas (dueña
    # 25/08/2026: "¿qué pasa si se vende 10 kg de 100? … poner para vender 90").
    # Arriba: lo que queda y lo que salió. Abajo, lo mismo, en el mismo lugar.
    # ⚠ «Queda» y no «Kg en saldo»: con la columna Vendido al lado, "Kg en
    # saldo 84 · Vendido 20" invita a restar, y los 84 ya tienen la venta
    # descontada. Es además el mismo rótulo que usa la tabla del pie.
    assert arriba[5:7] == ["Queda", "Vendido"]
    assert abajo[5:8] == ["Queda", "Vendido", "Vale"]
    assert abajo[9:] == ["Vendedor", "Día"], "las dos últimas son las que cambian"
    # ⭐ «Última» volvió el mismo día que se fue: lo que molestaba no era la
    # columna sino que la tabla no entraba y quedaba cortada a la derecha
    # ("antes se veía fuera"). Es el dato que explica por qué una tela entra o
    # sale de la lista.
    assert arriba[-1] == "Última"

    # el día con el mismo formato que la fecha de arriba, y el vendedor
    assert "{{ v.fecha.strftime('%d/%m/%y') }}" in html
    assert "{{ v.vendedor }}" in html
    # forma y calidad salen de los MISMOS macros que la lista de Saldos
    # (una vez en cada pantalla desde que Vendidos se mudó, 26/08/2026)
    assert html.count('<td class="forma">{{ fm.forma(') == 1
    assert _html_parado().count('<td class="forma">{{ fm.forma(') == 1
    assert "{{ fm.cal(v) }}" in html
    # las dos pantallas que dibujan esta plantilla tienen que pasarle la tabla
    assert _i.getsource(views).count("vendidos=queries.vendidos(") == 2


def test_el_renglon_vendido_lleva_a_su_tela_arriba():
    """Dueña 25/08/2026: *"si clickeo en poliester me muestre alemania, no otra
    tabla más abajo"*. Un segundo detalle abajo sería la misma tela contada en
    dos lugares de la misma pantalla; lo que hace falta es ir a la fila que ya
    existe.

    ⚠ Y si un filtro la tiene escondida, se limpian los filtros: llevar a
    alguien a una fila invisible es peor que no llevarlo."""
    from pathlib import Path
    html = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / "vendidos.html").read_text(encoding="utf-8")
    # ⚠⚠ Desde el 26/08/2026 Vendidos es su propia pantalla: acá arriba NO hay
    # tabla de saldos que scrollear, así que la fila LLEVA a esa tela en Saldos,
    # ya filtrada. Antes buscaba la fila en la misma página y le hacía scroll;
    # dejar eso habría sido un click mudo.
    assert "location.href='/analisis/parado?" in html, (
        "la fila de Vendidos no lleva a ningún lado")
    assert "subgrupo={{ v.subcategoria" in html, "no filtra por la tela"
    assert "q={{ v.color" in html, (
        "sin el color caería en el primer color de esa tela")
    # ⚠⚠ La fila de Vendidos NO puede llamarse `item`: cuando las dos tablas
    # convivían, el filtro recorría `tr.item` de toda la pantalla y las escondía
    # (25/08/2026). El nombre se mantiene por si algún día vuelven a convivir.
    assert '<tr class="vfila"' in html
    assert "verTela" not in html, (
        "quedó el click viejo, que buscaba una tabla que ya no está acá")
    # ⚠ El destino sigue siendo Saldos, así que la fila de ALLÁ tiene que
    # poder encontrarse por su tela y su color, y avisar adónde llegó.
    saldos = _html_parado()
    assert 'data-color="{{ f.color }}"' in saldos, (
        "la fila de Saldos tiene que poder encontrarse por su color")
    assert "tr.item.marcada>td{background:" in saldos, (
        "sin la marca, el salto deja buscando cuál de las 700 filas era")


def test_lo_de_segunda_que_se_vendio_entero_no_se_apaga(monkeypatch):
    """⚠⚠ El bug que dejó la tabla de Vendidos VACÍA el 25/08/2026.

    Desde que las banderas se calculan para TODA la bodega —hizo falta para la
    tela ya vendida—, contestan también por telas que nunca fueron un saldo
    parado. Una tela que se vende bien y está en producción tiene `produciendo`
    en 1 y, si vendió toda su segunda, `entra` en falso: quedaba marcada `fuera`
    y con ella se borraban sus ventas y los puntos de quien la vendió.

    Sólo se apaga lo que entró como `parado`. Un ítem de SEGUNDA entró por kilos
    que son un saldo se venda la tela o no."""
    db, res = _refresco(
        monkeypatch,
        parados=[{"subcategoria": "Pique Especial", "color": "BAZ",
                  "stock_kg": 0, "stock_bodega": 900, "motivo": "segunda",
                  "nueva": False, "pedida": False, "produciendo": True,
                  "entra": False}],
        cohorte=[{"subcategoria": "Pique Especial", "color": "BAZ",
                  "fecha_marcado": date(2026, 8, 13), "motivo": "segunda"}])
    assert not db.sql_con("SET fuera = TRUE"), (
        "vendió toda su segunda: eso es lo que la competencia premia")
    assert res["produciendo"] == 0


def test_el_renglon_vendido_cuelga_de_su_grupo():
    """Dueña 25/08/2026: *"quería ver el pique especial de lo vendido dentro de
    pique, no una tabla aparte"*. El cuadro decía "Pique · 38 kg" y la tela
    estaba en otra tabla más abajo: había que adivinar de cuál de los ocho
    grupos salía cada renglón."""
    import inspect as _i
    from pathlib import Path
    fuente = _i.getsource(queries.competencia)
    assert 'g["lineas"] = lineas_por_grupo.pop(g["grupo"], [])' in fuente
    # ⚠ lo que no cae en ningún grupo del cuadro no se pierde: sin esto, un
    # renglón desaparecería y los puntos del ranking no cerrarían con lo que se
    # ve en pantalla
    assert '"grupo": cat,' in fuente
    t = (Path("modules/analisis/templates/analisis/competencia.html")
         .read_text(encoding="utf-8"))
    assert "tr.vlinea>td:first-child{padding-left:26px}" in t, (
        "el renglón se lee colgado del grupo")
    assert "tr.vlinea>td{font-size:12px" in t
    assert "tr.vlinea>td{padding" not in t, (
        "con padding en las tres celdas los kilos dejan de encolumnar con los "
        "del grupo")
    # ⚠⚠ NO puede llamarse `linea`: esa clase ya es el renglón de tarjetas del
    # encabezado (un flex con `b` en 19 px) y se comía la fila entera.
    assert '<tr class="linea">' not in t
    base = (Path("modules/analisis/templates/analisis/base.html")
            .read_text(encoding="utf-8"))
    assert ".linea{display:flex" in base, "si esto cambia, revisar el choque"
    assert "vlinea" not in base, "el nombre nuevo tiene que seguir libre"


def test_el_renglon_colgado_es_una_linea():
    """Dueña 25/08/2026: *"muy finito tiene que ser"*. Con el color en negrita
    en su propia celda, cada venta ocupaba tres renglones de alto.

    ⚠ Los puntos estuvieron un día escondidos cuando eran iguales a los kilos
    ("no me repitas info 64, 64"). Funcionaba mientras todas las telas del grupo
    valieran 1: en un grupo MEZCLADO la columna quedaba con un solo número
    —Lycra mostraba 707 en una fila y nada en las otras cuatro— y parecía que
    sólo un color puntuaba (dueña 26/08/2026). Repetir un número es menos malo
    que una columna que se lee como un error, y con todos a la vista el total
    del grupo se puede verificar sumando."""
    from pathlib import Path
    t = (Path("modules/analisis/templates/analisis/competencia.html")
         .read_text(encoding="utf-8"))
    assert "white-space:nowrap" in t
    assert "{% if v.puntos | round(0) != v.kg | round(0) %}" not in t, (
        "volvieron a esconderse los puntos que son iguales a los kilos")
    assert '<td class="n pts">{{ v.puntos | num_es(0) }}</td>' in t


def test_desde_el_detalle_se_salta_a_esa_tela_en_saldos():
    """Dueña 26/08/2026: *"dejame clickear ahí en el producto y me lleve a los
    saldos de ese producto"*. Los filtros de Saldos viajan en la URL, así que
    el link deja la lista filtrada por grupo, tela y color.

    ⚠ El vendedor va a SU copia de la lista: `/analisis/parado` no está en su
    allowlist y le daría 404."""
    from pathlib import Path
    t = (Path("modules/analisis/templates/analisis/competencia.html")
         .read_text(encoding="utf-8"))
    i = t.index('class="alista"')
    link = " ".join(t[i - 400:i].split())
    assert "/analisis/competencia/telas' if vend" in link
    assert "else '/analisis/parado'" in link
    for campo in ("grupo=", "subgrupo=", "q="):
        assert campo in link, f"el link no filtra por {campo}"
    # los tres campos existen en el buscador de Saldos
    parado = (Path("modules/analisis/templates/analisis/parado.html")
              .read_text(encoding="utf-8"))
    assert "const CAMPOS = ['q', 'grupo', 'subgrupo'" in parado


def test_el_que_no_compite_es_la_casa():
    """Dueña 25/08/2026, viendo "Bedon Hector" en la tabla de vendidos: *"bedón
    no es un vendedor"*. Cambió QUIÉN decide (dueña 31/08/2026: ahora es
    `cliente.vend`, no el vendedor de Asinfo — ver
    `test_la_competencia_atribuye_por_cliente_vend_no_por_asinfo`), pero la
    regla de esta función sigue siendo la misma: un código que no es uno de
    los seis activos, o que no vino, lo hizo la casa."""
    assert queries._quien_vendio({"vend_pc": "XYZ"}) == "Intela"
    assert queries._quien_vendio({"vend_pc": None}) == "Intela"
    assert queries._quien_vendio({"vend_pc": ""}) == "Intela"
    assert queries._quien_vendio({}) == "Intela"
    assert queries._quien_vendio({"vend_pc": "EDG"}) == "Ramirez Edgar"
    assert queries._quien_vendio({"vend_pc": "edg"}) == "Ramirez Edgar", (
        "el código llega en minúsculas de algún lado y no debería importar")
    for codigo, nombre in asinfo_parado.PC_VEND.items():
        assert queries._quien_vendio({"vend_pc": codigo}) == nombre
        assert nombre in queries.COMPETIDORES
    # ⚠ se normaliza al ESCRIBIR, no al mostrar: si no, habría kilos que se ven
    # como de Intela y no suman en su fila del ranking
    import inspect as _i
    assert "_quien_vendio(v)" in _i.getsource(queries.actualizar)


def test_la_competencia_atribuye_por_cliente_vend_no_por_asinfo(monkeypatch):
    """Dueña 31/08/2026, sobre la factura 001-099-000183092: el cliente tiene
    `cliente.vend = 'SEP'` en Programa Core —el mismo vendedor que muestra la
    ficha de la factura, Sebastián Proaño— pero esa factura puntual quedó en
    Asinfo firmada por otra persona ("Bedon Hector", que ni siquiera compite) y
    la competencia la contaba para "Intela". Desde ahora manda `cliente.vend`,
    no quién firmó esa venta en particular."""
    db, filas = _refresco_con_ventas(
        monkeypatch,
        parados=[_item("Toper", "COA", kg_antes=100)],
        cohorte=_cohorte("Toper", "COA"),
        ventas=[_venta("Toper", "COA", 60, vend="Bedon Hector",
                        codigo_cli="AAA")],
        clientes=[{"codigo_cli": "AAA", "vend": "SEP"}],
        devolver_db=True)
    completas = [p for sql, p in db.escrito
                 if "INSERT INTO scintela.parado_venta" in sql]
    assert len(completas) == 1
    assert completas[0][2] == "SEP", "vend_pc: el código de cliente.vend"
    assert completas[0][3] == "Proaño Sebastián", (
        "vendedor: el nombre bonito de ESE código, no el de Asinfo")


def test_cliente_sin_vend_asignado_cae_a_intela(monkeypatch):
    """Un cliente sin vendedor en `cliente.vend` (NULL o vacío, o que no está
    en la lista) no tiene cómo atribuirse: cae a Intela — igual que antes de
    este cambio, y aunque en Asinfo la haya firmado un vendedor real."""
    db, _ = _refresco_con_ventas(
        monkeypatch,
        parados=[_item("Toper", "COA", kg_antes=100)],
        cohorte=_cohorte("Toper", "COA"),
        ventas=[_venta("Toper", "COA", 60, vend="Ramirez Edgar",
                        codigo_cli="ZZZ")],
        # "ZZZ" no está en la lista: cliente sin vend, o que no existe
        clientes=[{"codigo_cli": "AAA", "vend": "SEP"}],
        devolver_db=True)
    completas = [p for sql, p in db.escrito
                 if "INSERT INTO scintela.parado_venta" in sql]
    assert completas[0][2] is None
    assert completas[0][3] == "Intela", (
        "sin cliente.vend no importa quién firmó en Asinfo")


def test_vend_de_un_codigo_que_no_compite_no_se_filtra_crudo(monkeypatch):
    """Bug en vivo el 31/08/2026, mismo día del cambio: la factura de James
    1.2 BLA del 25/08 quedó con `vend_pc='BED'` escrito tal cual en
    `parado_venta` —"BED" es un código real de `scintela.vendedor" (alguien
    que no compite), no un vendedor inventado— y `vendidos.html` pinta
    `v.vend_pc or 'Intela'` (el código corto, SIN pasar por `_quien_vendio`):
    la pantalla de Vendidos mostró "BED" en la columna Vendedor en vez de
    "Intela". `vend_pc` tiene que seguir siendo "uno de los seis que
    compiten, o nada" — el mismo contrato de antes del 31/08 — aunque
    `cliente.vend` traiga cualquier código real."""
    db, _ = _refresco_con_ventas(
        monkeypatch,
        parados=[_item("Toper", "COA", kg_antes=100)],
        cohorte=_cohorte("Toper", "COA"),
        ventas=[_venta("Toper", "COA", 60, vend="Ramirez Edgar",
                        codigo_cli="AAA")],
        clientes=[{"codigo_cli": "AAA", "vend": "BED"}],
        devolver_db=True)
    completas = [p for sql, p in db.escrito
                 if "INSERT INTO scintela.parado_venta" in sql]
    assert completas[0][2] is None, (
        "vend_pc: 'BED' no compite, no se escribe crudo")
    assert completas[0][3] == "Intela", (
        "vendedor: normalizado igual, con o sin vend_pc"
    )


def test_el_grupo_del_vendido_sobrevive_a_la_venta(monkeypatch):
    """La tabla muestra justamente lo que se vendió, que es lo primero que deja
    de tener foto: sin el fallback al puntaje —que guarda el grupo por tela— la
    mayoría de los renglones decía "—"."""
    visto = {}

    def fake(sql, params=None, conn=None):
        visto["sql"] = " ".join(sql.split())
        return []

    monkeypatch.setattr(queries.db, "fetch_all", fake)
    queries.vendidos("2026-08-25")
    assert "COALESCE(f.categoria, p.categoria)" in visto["sql"]
    assert "GROUP BY" in visto["sql"] and "p.categoria" in visto["sql"]


def test_solo_puntuan_los_kilos_que_ya_estaban_en_la_competencia(monkeypatch):
    """Dueña 25/08/2026: *"tiene que contar solo kgs que estaban en la
    competencia para empezar"*.

    Son dos preguntas y hay que pasar las dos: `kg_antes` dice cuántos kilos
    estaban parados hace 90 días y `kg_al_marcar` cuántos había el día que la
    tela entró a la lista. Un ítem con 300 kg viejos que recibió producción
    después de entrar no puede puntuar esos kilos nuevos: nunca estuvieron en
    juego."""
    item = _item("Toper", "COA", kg_antes=300, stock=800)
    filas = _refresco_con_ventas(
        monkeypatch, parados=[item],
        cohorte=[{"subcategoria": "Toper", "color": "COA",
                  "fecha_marcado": date(2026, 8, 13), "kg_al_marcar": 120,
                  "motivo": "parado"}],
        ventas=[_venta("Toper", "COA", 500)])
    assert ("COA", 120.0, True) in filas, "el tope es el más chico de los dos"
    assert ("COA", 380.0, False) in filas


def test_sin_kilos_al_marcar_manda_el_saldo_viejo(monkeypatch):
    """La cohorte anterior a esta regla puede no tener el dato; ahí el tope
    sigue siendo el saldo de hace 90 días y no cero — dejar a alguien sin
    puntos por una columna vacía sería peor que el problema."""
    filas = _refresco_con_ventas(
        monkeypatch, parados=[_item("Toper", "HAB", kg_antes=200, stock=200)],
        cohorte=[{"subcategoria": "Toper", "color": "HAB",
                  "fecha_marcado": date(2026, 8, 13), "kg_al_marcar": None,
                  "motivo": "parado"}],
        ventas=[_venta("Toper", "HAB", 150)])
    assert filas == [("HAB", 150.0, True)]


def test_la_forma_sale_de_la_tela_cuando_el_stock_no_la_puede_decir():
    """Dueña 25/08/2026: *"no hay chance que sea —, es tub o abi"*. La forma se
    sacaba de los lotes CON SALDO, así que la tela vendida entera se quedaba sin
    nada que mostrar. Ahora se mira todo lote que pasó por la bodega."""
    sql = " ".join(asinfo_parado.SQL_FORMA.split())
    assert "FROM (SELECT DISTINCT id_producto, id_lote FROM saldo_producto_lote" in sql
    assert "saldo > 0" not in sql, (
        "si sólo mira los lotes con saldo, la tela vendida vuelve a no tener forma")
    assert "l.id_valor_atributo_3" in sql, "la forma vive en ese slot del lote"

    import inspect as _i
    fuente = _i.getsource(queries.actualizar)
    assert "forma_de = asinfo_parado.formas()" in fuente
    assert "forma_de.get(k)" in fuente, "se guarda en la foto, por ítem"
    # y las dos lecturas caen a esa forma cuando los kilos no la dicen
    assert "ELSE COALESCE(f.forma, '') END     AS forma," in _i.getsource(queries.items)
    assert "ELSE COALESCE(f.forma, '') END)    AS forma_fila," in _i.getsource(queries.vendidos)


def test_la_lista_no_dibuja_los_items_sin_un_kilo():
    """Dueña 25/08/2026: *"los que hay 0 no tienen que estar"*. El ítem que se
    vendió entero se queda en la COHORTE —"si empezamos a venderlas, que no se
    nos vayan de la lista"— y sigue contando en el resumen y en la competencia,
    que salen de `base`. Pero en la lista era un renglón de 0 kg, 0 puntos y
    grupo "—": no hay nada que ofrecer ahí, y lo vendido tiene su propia tabla.

    El filtro va sobre `filas` y NO sobre `base`: si se filtrara la base, el
    encabezado dejaría de contar los kilos vendidos y «Se vendió» bajaría."""
    import inspect as _i
    fuente = _i.getsource(views.parado)
    corte = fuente.index("filas = [f for f in filas")
    assert 'float(f.get("stock_kg") or 0) > 0' in fuente[corte:corte + 120]
    assert fuente.index("base = queries.con_puntos") < corte
    assert "resumen=queries.resumen(base" in fuente, (
        "el resumen tiene que seguir saliendo de la base, no de lo dibujado")


def test_vendidos_lleva_su_total_arriba():
    """Dueña 25/08/2026: *"acá poneme un total arriba de todo: debería coincidir
    con los que dicen que vendió"*. Es la comprobación de la pantalla: si esta
    suma no da lo mismo que el «Se vendió» del encabezado y que el ranking de la
    Competencia, hay un renglón contándose de más o de menos."""
    html = _html_vendidos()
    i = html.index('id="vendidos-tabla"')
    j = html.index('id="vendidos"')
    arriba = html[i:j]
    assert "vendidos | sum(attribute='kg')" in arriba, "el total no está, o está debajo de la tabla"
    assert "vendidos | sum(attribute='puntos_fila')" in arriba


def test_el_detalle_no_lista_los_grupos_en_cero():
    """Dueña 25/08/2026: *"si no vendió nada de las otras telas, no
    mostrarlas"*. Los ocho grupos estaban siempre, así que el que vendió una
    sola tela abría su fila y veía un renglón y siete ceros: la pantalla decía
    siete veces "acá no hizo nada" y escondía lo único que sí hizo.

    Se filtra en la CUENTA y no en el template: el total del grupo lo sigue
    armando `competencia()`, así que si un grupo tuviera puntos sin kilos —una
    devolución que deja el neto en cero— tampoco desaparece."""
    import inspect as _i
    fuente = _i.getsource(queries.competencia)
    assert 'd["detalle"] = [g for g in d["detalle"]' in fuente
    assert 'float(g.get("kg") or 0) or float(g.get("puntos") or 0)' in fuente


def test_la_competencia_no_repite_la_tabla_de_saldos():
    """Dueña 25/08/2026: *"toda esta tabla borrar también de competencia, ya
    existe en saldos"*. Era una copia de las ocho telas que más puntos valen:
    los mismos números en dos pantallas se despegan a la primera corrección.
    Queda el link, que es lo que hacía falta."""
    from pathlib import Path
    html = ((Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "competencia.html")
            .read_text(encoding="utf-8"))
    import re
    texto = re.sub(r"\{#.*?#\}", " ", html, flags=re.S)
    assert "<h2>Los saldos</h2>" not in texto
    assert "telas[:8]" not in texto
    assert "/analisis/competencia/telas" in texto, "se fue también el link"
    # ⚠ La sección del vendedor NO se toca: el primer intento de borrar la
    # tabla se llevó puesto «Sus clientes», que está abajo en el mismo archivo.
    assert "<h2>Sus clientes</h2>" in texto


def test_por_grupo_se_pliega_y_los_filtros_tienen_su_cuadro():
    """Dueña 25/08/2026: *"en saldos por grupo dejame hide así no es tan
    grande. filtros mucho más definidos para filtrar tabla abajo"*.

    Los ocho subtotales se miran una vez; los filtros se usan siempre y estaban
    mezclados con los botones de Excel, Imprimir y PDF, todos del mismo tamaño
    y sin rótulo."""
    html = _html_parado()
    # ⚠ Buscar el `<details>` y no el texto "Por grupo": desde que hay índice
    # al costado, la PRIMERA aparición de esas dos palabras es el link.
    i = html.index('<details class="plegable" id="por-grupo">')
    fin = html.index("</summary>", i)
    assert "Por grupo" in html[i:fin], "el desplegable no es el de Por grupo"
    assert " open" not in html[i:i + 60], "arranca abierta: sigue ocupando lo mismo"
    assert '<div class="filtros">' in html
    for campo in ("q", "grupo", "subgrupo", "calidad", "vend"):
        assert f'<label for="{campo}">' in html, f"el filtro {campo} no tiene rótulo"
    # Los botones se van a su propio renglón: no son filtros.
    # ⚠ Desde el 25/08/2026 Excel, Imprimir y PDF viven al lado del TÍTULO
    # ("esto ocupa un montón de renglones"), o sea ANTES de los filtros. Lo que
    # el test cuida sigue siendo lo mismo: que los botones no estén mezclados
    # adentro del cuadro de filtros.
    import re
    ini = html.index('<div class="filtros">')
    fin = html.index("</div>", html.index('id="cuenta"'))
    assert 'id="excel"' not in html[ini:fin], (
        "los botones volvieron adentro del cuadro de filtros")
    assert '<div class="titulo-fila">' in html
    sin_comentarios = re.sub(r"\{#.*?#\}", " ", html, flags=re.S)
    assert "Ver por cliente" not in sin_comentarios, (
        "ya es una pestaña del menú de arriba (dueña 25/08/2026)")


def test_los_filtros_entran_en_un_telefono():
    """Seis controles a lo ancho no entran en 390 px. Van de a dos por renglón
    y el buscador ocupa el renglón entero: es el que más se usa.

    ⚠ Y los controles miden 16 px en el celular a propósito: abajo de eso iOS
    hace zoom solo al tocar el campo y la pantalla queda corrida."""
    from pathlib import Path
    base = ((Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "base.html").read_text(encoding="utf-8"))
    i = base.index("@media (max-width: 780px)")
    movil = base[i:]
    assert ".filtros .ff{min-width:0;flex:1 1 calc(50% - 4px)}" in movil
    assert ".filtros .ff:first-child{flex:1 1 100%}" in movil
    assert ".filtros input,.filtros select{font-size:16px" in movil


def test_la_barra_del_costado_se_fue_con_la_mudanza():
    """Dueña 26/08/2026: *"saquemos esto"*, cuando Vendidos pasó a su tab.

    El índice existía para no scrollear 700 telas hasta Vendidos. Con Vendidos
    en su propia pantalla no hay nada que saltear — y encima engañaba: los links
    scrolleaban bien, pero el título quedaba tapado por la barra pegajosa de
    arriba y parecía que el click no hacía nada.
    """
    html = _html_parado()
    assert '<nav class="secciones"' not in html
    assert 'href="#vendidos-tabla"' not in html, (
        "quedó un link a una tabla que ya no está en esta pantalla")
    assert 'class="gnav"' not in html, (
        "volvió la columna de grupos al costado, que achicaba la tabla")

def test_vendidos_lleva_el_total_debajo_de_sus_columnas():
    """Dueña 25/08/2026: *"totales para la tabla de vendidos"*. Arriba está la
    cifra sola; acá cae debajo de Kg y de Puntos, que es donde se la busca
    cuando se terminó de recorrer los renglones.

    ⚠ El rótulo «Total» va en la columna Tela y no en Grupo: Grupo es `opt` y
    en el teléfono se esconde."""
    html = _html_vendidos()
    tabla = html[html.index('id="vendidos"'):]
    pie = tabla[tabla.index("<tfoot>"):tabla.index("</tfoot>")]
    assert "vendidos | sum(attribute='kg')" in pie
    assert "vendidos | sum(attribute='puntos_fila')" in pie
    assert pie.count("<td") == 11, "el pie tiene que tener las 11 columnas"
    assert '<td class="opt"></td>\n<td><b>Total</b>' in pie


def test_la_tela_vendida_entera_se_queda_tachada():
    """Dueña 25/08/2026: *"¿esto que se vendió por ejemplo? no hay que ponerlo
    en 0. ¿Tachar la fila y decir vendido?"*.

    Primero se habían sacado los renglones de 0 kg —"los que hay 0 no tienen
    que estar"— y eso rompía dos cosas: la lista dejaba de mostrar justo lo que
    la competencia premia, y el click de la tabla de Vendidos, que lleva a la
    fila de esa tela, no encontraba a dónde ir. Lo que molestaba era el CERO
    pelado, no la fila.

    ⚠ El que quedó en cero SIN vender nada sí se va: es un ajuste de bodega."""
    import inspect as _i
    fuente = _i.getsource(views.parado)
    assert 'float(f.get("kg_vendidos") or 0) > 0' in fuente, (
        "la tela vendida entera volvió a desaparecer de la lista")
    html = _html_parado()
    assert "{% set vendida = not f.stock_kg and f.kg_vendidos %}" in html
    # ⚠ La palabra va en la columna KG, donde iba el cero: al lado del nombre no
    # entraba —la columna Tela tiene ancho fijo y tapaba medio nombre— y además
    # el cero era justo lo que molestaba.
    i = html.index('<td class="n" data-v="{{ f.stock_kg }}">')
    assert '<span\n      class="pill vend">vendido</span>' in html[i:i + 260]
    from pathlib import Path
    base = ((Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "base.html").read_text(encoding="utf-8"))
    assert "tr.vendida .nom{text-decoration:line-through}" in base
    assert "tr.vendida td{text-decoration:line-through}" not in base, (
        "los kilos no se tachan: el 0 es un dato")


def test_vendidos_no_promete_un_click_que_no_anda():
    """Dueña 25/08/2026: *"esto borrar, no funciona tampoco"*. La bajada decía
    "toque una fila para ver esa tela arriba" y el click estaba roto: la fila
    de arriba era justamente la que se había ido por tener 0 kg. Ahora la fila
    existe —tachada— y el click anda; el texto igual sobraba."""
    html = _html_vendidos()
    assert "Toque una fila para ver esa tela arriba" not in html
    assert "Lo que salió de la lista desde que arrancó la competencia" not in html
    assert "location.href='/analisis/parado?" in html, "el click sigue estando"


def test_la_pantalla_no_se_corre_para_el_costado_en_el_telefono():
    """Medido el 25/08/2026 a 390, 360 y 320 px: el documento medía 492 contra
    390 de pantalla, así que TODA la pantalla —encabezado, tarjetas y filtros—
    se corría en diagonal. La culpa eran las tablas `resumen` (Por grupo y
    Vendidos), que viven en una `.caja` con `overflow:hidden`: ni se recortan ni
    se pueden correr, empujan.

    ⚠ `:not(.larga)`: la tabla larga va con `overflow:visible` a propósito —un
    ancestro que recorta mata su encabezado fijo—, así que no puede caer en la
    misma regla."""
    from pathlib import Path
    base = ((Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "base.html").read_text(encoding="utf-8"))
    movil = base[base.index("@media (max-width: 780px)"):]
    assert ".caja{overflow-x:auto}" in movil
    # ⚠ La tabla LARGA era la que empujaba: 477 px asomando de una caja de 362,
    # porque en escritorio va con `overflow:visible` para no matar su encabezado
    # fijo. En el teléfono el encabezado fijo se apaga y la tabla se corre.
    assert "#tabla thead th{position:static}" in movil
    # ⚠ Con las DOS clases: `.caja{...}` sola pierde por especificidad contra
    # `.caja.larga{overflow:visible}` y la regla no entraba. A 320 px la página
    # seguía midiendo 350.
    assert ".caja.larga{overflow-x:auto}" in movil


def test_la_pantalla_de_metas_no_vuelve():
    """Dueña 25/08/2026: *"borrar página de metas, no sirve para nada"*. Desde
    el 24/08 la competencia NO tiene metas —gana el que más puntos hace—, así
    que esa pantalla editaba un número que ya no decide nada.

    ⚠ Lo que se queda es `kg_base`: los kilos congelados del día de la largada.
    Ésos no son una meta, son contra qué se mide."""
    import re as _re

    from modules.analisis import views
    assert not hasattr(views, "competencia_metas")
    assert "analisis/competencia_metas.html" not in _i_src(views)

    # ⚠ Sin comentarios: los comentarios cuentan lo que se borró y nombran
    # justamente lo que este test prohíbe.
    codigo = _re.sub(r"#[^\n]*", "", _i_src(queries))
    for muerto in ("_meta_pct(", "meta_pct", "meta_kg", "meta_pts",
                   "parado_meta", "meta_total_pct"):
        assert muerto not in codigo, (
            f"volvió {muerto}, que era de la época de las metas")


def _i_src(mod):
    import inspect
    return inspect.getsource(mod)


def test_el_item_que_asinfo_ya_no_devuelve_igual_recibe_motivo(monkeypatch):
    """El UPDATE del motivo corre sobre lo que HOY está parado. El ítem que
    dejó de calificar —se movió, o se vendió antes de la largada— no aparece
    más en esa consulta y su motivo se quedaba en NULL para siempre.

    Y sin motivo, `cuenta_el_kilo()` cuenta TODO, primera incluida: el día que
    ese ítem vuelva a tener stock y se venda, paga kilos que nunca fueron un
    saldo. Medidos el 25/08/2026: 9 ítems de la cohorte del 17/08, 1.129 kg.

    Se les escribe `segunda`, la regla más exigente: quedarse en la más
    exigente no le regala puntos a nadie, volver sí."""
    escritos = []
    db = _DBFalsa([])
    real = db.execute

    def espiar(sql, params=None, conn=None):
        escritos.append(" ".join(sql.split()))
        return real(sql, params, conn)

    db.execute = espiar
    monkeypatch.setattr(queries, "db", db)
    monkeypatch.setattr(queries, "today_ec", lambda: date(2026, 8, 25))
    monkeypatch.setattr(asinfo_parado, "parados", lambda: [])
    monkeypatch.setattr(asinfo_parado, "llamados", lambda: [])
    monkeypatch.setattr(asinfo_parado, "vendido_desde", lambda d: [])
    monkeypatch.setattr(asinfo_parado, "share_por_grupo", lambda: [])
    monkeypatch.setattr(asinfo_parado, "venta_por_tela", lambda: {})
    monkeypatch.setattr(asinfo_parado, "formas", lambda: {})
    monkeypatch.setattr(asinfo_parado, "ultima_venta_antes", lambda d: {})
    queries.actualizar()
    barrido = [s for s in escritos
               if "parado_cohorte SET motivo = 'segunda'" in s]
    assert barrido, "el ítem que ya no aparece se queda sin motivo para siempre"
    assert "WHERE motivo IS NULL" in barrido[0], (
        "no puede pisar el motivo de los que ya lo tienen: la regla de un ítem "
        "no cambia en la mitad de la carrera")


def test_el_premio_del_mes_no_revienta_con_una_largada_en_diciembre():
    """`largada.month + 1` con la largada en diciembre da `date(a, 13, 1)`, que
    tira ValueError y se lleva puesta la pantalla entera. Hoy la largada es el
    25/08 y no pasa, pero es una trampa puesta para el próximo que corra una
    competencia."""
    import inspect as _i
    fuente = _i.getsource(queries._meses)
    assert "largada.month == 12" in fuente, (
        "diciembre + 1 no existe: la pantalla se cae")
    assert "largada.year + 1" in fuente


def test_el_premio_del_mes_no_dice_que_no_tiene_tope():
    """La pantalla decía "kilos totales, sin tope" y no es cierto: cuenta los
    mismos kilos que puntúan (`v.cuenta`). La frase venía del TOPE POR GRUPO,
    la regla del 17/08 que se sacó el 24 — quedó describiendo otra cosa."""
    from pathlib import Path
    html = ((Path(__file__).resolve().parent.parent / "modules" / "analisis" /
             "templates" / "analisis" / "competencia.html")
            .read_text(encoding="utf-8"))
    import re
    texto = re.sub(r"\{#.*?#\}", " ", html, flags=re.S)
    assert "sin tope" not in texto
    import inspect as _i
    assert "v.cuenta" in _i.getsource(queries._meses), (
        "si algún día cuenta también lo que no puntúa, la frase vuelve a ser cierta")


def test_la_tabla_larga_entra_en_la_pagina():
    """Once columnas con `table-layout:fixed`: si los anchos suman más de 100,
    la tabla se pasa del ancho de la página y la ÚLTIMA queda cortada a la
    derecha. Es lo que le pasaba a «Última venta» (dueña 25/08/2026: "antes se
    veía fuera") y por lo que la columna parecía sobrar."""
    import re
    html = _html_parado()
    anchos = [int(x) for x in re.findall(r"#tabla th:nth-child\(\d+\)\{width:(\d+)%\}", html)]
    assert len(anchos) == 11, f"hay {len(anchos)} anchos para 11 columnas"
    assert sum(anchos) == 100, f"los anchos suman {sum(anchos)}%: la tabla se desborda"


def test_la_fila_vendida_no_tiene_huecos():
    """Dueña 25/08/2026: *"por favor, completá los datos faltantes"*. La fila
    tachada mostraba la categoría en "—" y los puntos en 0.

    · La CATEGORÍA no puede salir del stock —no queda un lote que mirar— pero sí
      de lo que se VENDIÓ, que es el dato que hay.
    · Los PUNTOS de una fila vendida no son cero: son los que ya hizo. El cero
      salía de multiplicar el stock —que es 0— por el valor del kilo, y en una
      fila tachada leer "0 puntos" es exactamente al revés de lo que pasó."""
    html = _html_parado()
    # ⚠ Antes esto miraba la frase textual del template. La regla se mudó a
    # `queries.categoria_de` —vivía repetida en cuatro lugares y sólo el
    # template tenía el caso de la fila vendida; el Excel y la hoja impresa
    # decían PRI (dueña 26/08/2026)—. Ahora se chequea el COMPORTAMIENTO, que
    # es lo que importaba desde el principio.
    vendida_seg = {"kg_primera": 0, "kg_segunda": 0,
                   "kg_vend_pri": 0, "kg_vend_seg": 42.5}
    assert queries.categoria_de(vendida_seg) == "SEG", (
        "la categoría de la fila vendida sale de lo que se vendió")
    assert "f.cat" in html, "y el template la lee ya resuelta"
    assert "{% set pts = (f.kg_vendidos * f.puntos) if vendida else f.puntos_fila %}" in html
    assert 'data-v="{{ pts }}">{{ pts | num_es(0) }}' in html


def test_la_ultima_venta_se_le_pide_a_asinfo_con_el_corte_adentro():
    """Es la columna que justifica que el ítem sea un saldo: cuánto hace que
    nadie lo pedía. Con "la última venta a secas", el primer kilo que se vende
    en la competencia la pisa con la fecha de HOY y la fila borra sola la prueba
    de que estaba clavada — encima con la venta que se acaba de premiar.

    ⚠ El primer intento fue conservar la que ya tenía la foto. NO alcanza: el
    día de la largada la foto ya se había refrescado con la fecha de hoy, así
    que las ocho telas que se vendieron ese día quedaron con la columna en "—"
    y no había de dónde sacarla (dueña 25/08/2026: "¿por qué cuellos no tiene
    última?"). El dato está en Asinfo: hay que ir con la fecha de corte."""
    sql = asinfo_parado._sql_ultima_antes("2026-08-25")
    plano = " ".join(sql.split())
    assert "MAX(CAST(fc.fecha AS date))" in plano
    assert "CAST(fc.fecha AS date) < '2026-08-25'" in plano, (
        "sin el corte, la venta de la competencia pisa la fecha")
    assert "fc.id_documento IN (7, 251)" in plano
    import inspect as _i
    fuente = _i.getsource(queries.actualizar)
    assert "asinfo_parado.ultima_venta_antes(" in fuente
    assert "ultima_antes_de.get(k)" in fuente
    assert "def ultima_antes(" not in fuente, (
        "volvió el intento de reconstruirla desde la foto anterior")


def test_el_movido_ya_no_puede_ser_negativo():
    """⚠⚠ REVIERTE a propósito la frase del 25/08/2026 ("acá aclarar que los
    de segunda siguen entrando") — y el intento a medias de la MISMA tarde
    del 31/08/2026 ("de más en bodega"). Verificado en vivo el 01/09/2026:
    no era sólo una tela nueva sumándose a la cohorte — telas marcadas hace
    semanas siguen recibiendo SEGUNDA todos los días y su stock de hoy supera
    lo que tenían al marcarse. `resumen()` ahora pone un PISO: "Al arrancar"
    nunca puede ser menor que `kg + vendido`, así que `kg_movido` (el
    residuo, ver el docstring de `resumen()`) queda estructuralmente
    imposible de ser negativo. Ya no hace falta explicar "de dónde sale" un
    número que no puede existir."""
    html = _html_parado()
    i = html.index("resumen.kg_movido | num_es(0)")
    assert "que salieron por bodega" in html[i:i + 120]
    assert "de más en bodega" not in html
    assert "de SEG nueva" not in html
    assert "resumen.kg_movido < 0" not in html
    assert "resumen.kg_movido >= 1" in " ".join(html.split())


def test_la_linea_abierta_se_lleva_su_propia_categoria_vendida():
    """La píldora de categoría de una fila VENDIDA se dibuja mirando
    `kg_vend_pri` / `kg_vend_seg` —del stock no se puede, no queda lote—. Si al
    abrir el color en dos líneas quedaran los del ítem entero, la línea de
    primera de un color que vendió las dos diría "SEG" o se quedaría sin
    píldora: estaría mostrando lo que vendió la OTRA línea."""
    fila = {"subcategoria": "Jersey 3", "color": "BLA", "stock_kg": 100,
            "kg_tub_pri": 60, "kg_tub_seg": 40, "kg_abi_pri": 0, "kg_abi_seg": 0,
            "kg_vend_pri": 7, "kg_vend_seg": 3, "puntos": 4}
    lineas = queries.abrir_en_lineas([fila])
    assert len(lineas) == 2
    pri = next(x for x in lineas if x["cal_fila"] == "PRI")
    seg = next(x for x in lineas if x["cal_fila"] == "SEG")
    assert (pri["kg_vendidos"], pri["kg_vend_pri"], pri["kg_vend_seg"]) == (7, 7, 0)
    assert (seg["kg_vendidos"], seg["kg_vend_pri"], seg["kg_vend_seg"]) == (3, 0, 3)




def test_afuera_de_la_lista_no_vuelve_a_la_pantalla():
    """Dueña 26/08/2026: *"no nos hace falta eso"*. La línea explicaba por qué
    la lista tiene 700 ítems y no 4.200 —una pregunta que se hace UNA vez— y
    ocupaba dos renglones arriba de todo, todos los días.

    ⚠ Antes de sacarla se intentó acortarla y después gatearla sólo para la
    dueña. Ninguna de las dos alcanzó: lo que sobraba era el renglón, no su
    largo.

    El dato se sigue calculando y sigue guardado en `parado_refresh`: si algún
    día la lista se desploma y hay que explicar por qué, está ahí."""
    import inspect as _i
    html = _html_parado()
    sin_comentario = html[html.index("{% block cuerpo %}"):] if "{% block cuerpo %}" in html else html
    for muerto in ("{{ estado.nuevas", "{{ estado.produciendo", "{{ estado.pedidas"):
        assert muerto not in sin_comentario, f"volvió {muerto} a la pantalla"
    assert "dias_quieto" not in _i.getsource(views.parado), (
        "se pasaba al contexto sólo para ese texto")
    fuente = _i.getsource(queries.actualizar)
    assert "nuevas" in fuente and "produciendo" in fuente and "pedidas" in fuente


def test_el_numero_de_factura_se_pela_para_buscarla_en_el_programa():
    """Asinfo la numera `001-099-000182637` y Programa Core la busca por `numf`,
    que es el número pelado. Verificado el 26/08/2026 contra las dos facturas
    del día de James 1.2 BLA: 182637 y 182654 existen en `scintela.factura`.

    ⚠ Si el formato cambiara y no quedara ningún dígito, devuelve None y la
    fila se dibuja sin link — antes que mandar a la dueña a otra factura."""
    assert asinfo_parado._numf("001-099-000182637") == 182637
    assert asinfo_parado._numf(" 001-099-000182654 ") == 182654
    assert asinfo_parado._numf("182637") == 182637
    assert asinfo_parado._numf(None) is None
    assert asinfo_parado._numf("") is None
    assert asinfo_parado._numf("001-099-XXX") is None


def test_el_renglon_vendido_guarda_su_factura():
    """Sin el número no hay a dónde linkear, y con el grain viejo —todo lo
    vendido de esa tela ese día— un renglón podía venir de dos facturas y el
    link no sabría a cuál ir. Por eso el número entra al GROUP BY: cada renglón
    es UNA factura."""
    sql = " ".join(asinfo_parado._sql_vendido("2026-08-25").split())
    assert "RTRIM(fc.numero) AS numero" in sql
    assert sql.count("RTRIM(fc.numero)") == 2, "tiene que estar en el GROUP BY"
    import inspect as _i
    fuente = _i.getsource(queries.actualizar)
    assert "calidad, cuenta, numf" in fuente
    assert 'v.get("numf")' in fuente
    # ⭐ Y el número ENTERO, que es el que desempata (mig 0233).
    assert 'v.get("numero")' in fuente


def test_la_factura_es_solo_para_quien_puede_ver_facturas():
    """Dueña 26/08/2026: *"eso sí, pero el link a la factura no"*. El detalle de
    los siete lo dejó abierto para los vendedores; la factura no.

    ⚠ Y sin el permiso el link igual daría 404: sería un link roto además de
    una filtración."""
    from pathlib import Path
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    for nombre in ("competencia.html", "vendidos.html"):
        html = (carpeta / nombre).read_text(encoding="utf-8")
        i = html.index("/facturas/{{ v.numf }}")
        antes = html[i - 260:i]
        assert "tiene_permiso('facturas.ver')" in antes, (
            f"{nombre}: el link a la factura no está gateado")
        assert "v.numf and" in antes, "sin número no se dibuja el link"


def test_la_hoja_impresa_habla_el_mismo_idioma_que_la_pantalla():
    """Dueña 26/08/2026: *"¿y se ven todas las pantallas iguales?"*. La lista
    del vendedor sí —es el MISMO template—, pero la hoja imprimible se arma con
    otra plantilla y se quedó con «Kg en saldo» cuando la tabla pasó a «Queda».

    El que sale a vender con la hoja en la mano y el que mira la pantalla
    tienen que leer la misma palabra."""
    from pathlib import Path
    carpeta = (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
               "templates" / "analisis")
    for nombre in ("parado.html", "parado_impreso.html"):
        html = (carpeta / nombre).read_text(encoding="utf-8")
        import re
        texto = re.sub(r"\{#.*?#\}", " ", html, flags=re.S)
        assert ">Kg en saldo<" not in texto, (
            f"{nombre} sigue diciendo «Kg en saldo»; la tabla dice «Queda»")
        assert ">Queda<" in texto


def test_el_libro_de_excel_sale_con_formato():
    """Dueña 26/08/2026: *"bajalo a algo con formato"*. Un .xlsx de verdad:
    encabezado fijo, filtro automático, anchos pensados y números que Excel
    pueda sumar. Se prueba abriendo el archivo, no leyendo el código."""
    import io

    from openpyxl import load_workbook

    from modules.analisis import excel
    datos = excel.libro([{
        "titulo": "Saldos",
        "columnas": [("Tela", 24, None), ("Queda", 11, "#,##0.0")],
        "filas": [["Jersey Forro Spun", 946], ["Microfibra", 833.5]],
    }])
    wb = load_workbook(io.BytesIO(datos))
    ws = wb["Saldos"]
    assert ws.freeze_panes == "A2", "sin encabezado fijo, a la fila 200 no se sabe qué es cada columna"
    assert ws.auto_filter.ref == "A1:B3"
    assert ws["A1"].value == "Tela" and ws["A1"].font.bold
    assert ws.column_dimensions["A"].width == 24, "la tela no entra en 8 caracteres"
    # ⚠ Lo que arreglaba el pedido: los kilos son NÚMEROS, no cadenas.
    assert ws["B2"].value == 946 and isinstance(ws["B2"].value, int | float)
    assert ws["B2"].number_format == "#,##0.0"
    assert ws["B2"].alignment.horizontal == "right"


def test_vendidos_muestra_el_codigo_del_vendedor_y_no_el_nombre_entero():
    """Dueña 26/08/2026: *"poner código de vendedor en la columna que arriba
    está cliente, no todo"*. Es la misma columna que en la lista de arriba dice
    «Clientes» y trae un número de tres dígitos: con «Proaño Sebastián» adentro
    se estiraba y las dos tablas dejaban de leerse en vertical.

    ⚠ La casa no tiene código de tres letras: va «Intela», que es como se llama
    en el ranking. Y el nombre completo queda en el `title`, que no ocupa."""
    html = _html_vendidos()
    tabla = html[html.index('id="vendidos"'):]
    assert "{{ v.vend_pc or 'Intela' }}" in tabla
    assert 'title="{{ v.vendedor }}"' in tabla, "el nombre entero no se pierde"
    assert "<td>{{ v.vendedor }}</td>" not in tabla


def test_hay_un_boton_para_borrar_los_filtros():
    """Dueña 26/08/2026: *"poneme un botón para borrar los filtros"*. Con seis
    filtros, volver a «todo» era tocar seis desplegables.

    ⚠ Y tiene que borrar TAMBIÉN lo guardado: los filtros viven en la dirección
    y en el navegador, así que dejarlos en blanco en la pantalla sin borrar el
    recuerdo los traía de vuelta al recargar."""
    html = _html_parado()
    assert 'onclick="limpiarFiltros()"' in html
    js = html[html.index("function limpiarFiltros()"):]
    assert "localStorage.removeItem(LLAVE)" in js[:400], "no borra lo guardado"
    assert "history.replaceState" in js[:400], "no borra los filtros de la dirección"
    assert "grupoCambio();" in js[:400], "no vuelve a dibujar la tabla"


def test_se_puede_filtrar_por_lo_que_ya_se_vendio():
    """Dueña 26/08/2026: *"dejame también filtrar por vendido"*. Son los
    renglones que la competencia premió: sin el filtro había que buscarlos a
    ojo entre 709."""
    html = _html_parado()
    assert '<select id="vendido"' in html
    assert 'data-vendido="{{ \'si\' if f.kg_vendidos else \'no\' }}"' in html
    assert "tr.dataset.vendido === vd" in html, "el filtro no se aplica"
    assert "'calidad', 'vendido', 'vend'" in html, (
        "no se recuerda con los demás")


def test_el_contador_no_ocupa_un_renglon_propio():
    """Dueña 26/08/2026: *"el total abajo molesta, en otro lado que no sea un
    renglón más"*. Abajo del cuadro de filtros se comía un renglón entero para
    decir dos números; ahora viaja con el título."""
    html = _html_parado()
    i = html.index('id="tela-por-tela"')
    assert 'id="cuenta"' in html[i:i + 220], "el contador no está en el título"
    fin = html.index("</div>", html.index('<div class="filtros">'))
    assert 'id="cuenta"' not in html[html.index('<div class="filtros">'):fin], (
        "quedó también abajo de los filtros")


def test_el_excel_baja_las_mismas_lineas_que_la_pantalla():
    """Dueña 26/08/2026, mirando el archivo: *"«Kg de primera», «Kg de
    segunda», ¿por qué tengo esto? si ya tengo la división entre PRI y SEG"*.

    Tenía razón, y la causa era otra: el Excel bajaba la fila SIN abrir, así
    que necesitaba las dos columnas de kilos para decir lo que en la pantalla
    ya son dos renglones. Con las líneas abiertas, cada renglón tiene UNA
    categoría y esas columnas no dicen nada. De paso el archivo y la pantalla
    cuentan lo mismo: eran 700 filas contra 709 líneas."""
    import inspect as _i
    fuente = _i.getsource(views.parado_xlsx) + _i.getsource(views._excel_saldos)
    assert "queries.abrir_en_lineas(" in fuente, (
        "el archivo baja la fila sin abrir y la pantalla la abre")
    assert '"Kg de primera"' not in fuente and '"Kg de segunda"' not in fuente
    assert '"Categoría"' in fuente


def test_la_forma_del_excel_sale_de_la_linea():
    """Con las líneas abiertas, la forma es la de ESA línea y no la de la tela
    entera: si no, la línea abierta diría la sigla de la tubular."""
    from modules.analisis.views import _forma
    assert _forma({"forma_fila": "ABI", "forma": "TUB"}) == "ABI"
    assert _forma({"kg_tubular": 10, "kg_abierta": 0}) == "TUB"
    assert _forma({"kg_tubular": 0, "kg_abierta": 10}) == "ABI"
    assert _forma({"forma": "TUB"}) == "TUB"
    assert _forma({}) == ""

# ---------------------------------------------------------------------------
# La CATEGORÍA de una fila (dueña 26/08/2026, leyendo el Excel: "¿cómo puede
# haber algo parado que se vendió hace un mes?" — eran filas que habían entrado
# por sus kilos de SEGUNDA y el archivo las etiquetaba PRI).
# ---------------------------------------------------------------------------

def test_categoria_de_la_fila_vendida_sale_de_lo_vendido():
    """Sin kilos en bodega no hay lote: la categoría sale de la VENTA.

    Éste es el caso que rompía. Una tela que entró por su segunda, la vendió
    entera y quedó en cero: los cuatro contadores de stock dan 0 y la regla
    vieja caía al `else` y devolvía "PRI".
    """
    fila = {"kg_primera": 0, "kg_segunda": 0,
            "kg_vend_pri": 0, "kg_vend_seg": 87.3}
    assert queries.categoria_de(fila) == "SEG"


def test_categoria_de_la_fila_vendida_de_primera():
    fila = {"kg_primera": 0, "kg_segunda": 0,
            "kg_vend_pri": 96.45, "kg_vend_seg": 0}
    assert queries.categoria_de(fila) == "PRI"


def test_categoria_de_la_fila_que_vendio_las_dos():
    fila = {"kg_primera": 0, "kg_segunda": 0,
            "kg_vend_pri": 10, "kg_vend_seg": 5}
    assert queries.categoria_de(fila) == "PRI SEG"


def test_categoria_sin_stock_y_sin_venta_no_inventa():
    """Vacío, no "PRI". Una etiqueta inventada es peor que un guión."""
    assert queries.categoria_de({}) == ""


def test_categoria_manda_el_stock_por_encima_de_lo_vendido():
    """Si quedan kilos, la categoría es la del LOTE aunque haya vendido otra."""
    fila = {"kg_primera": 0, "kg_segunda": 40,
            "kg_vend_pri": 100, "kg_vend_seg": 0}
    assert queries.categoria_de(fila) == "SEG"


def test_categoria_manda_la_linea_ya_abierta():
    fila = {"cal_fila": "PRI", "kg_primera": 0, "kg_segunda": 99,
            "kg_vend_pri": 0, "kg_vend_seg": 99}
    assert queries.categoria_de(fila) == "PRI"


def test_abrir_en_lineas_le_pone_la_categoria_a_todas():
    """Toda fila que sale para la pantalla lleva `cat`, se haya abierto o no."""
    entera = {"stock_kg": 0, "kg_tub_pri": 0, "kg_tub_seg": 0,
              "kg_abi_pri": 0, "kg_abi_seg": 0, "puntos": 1,
              "kg_primera": 0, "kg_segunda": 0,
              "kg_vend_pri": 0, "kg_vend_seg": 20.8}
    partida = {"stock_kg": 30, "kg_tub_pri": 20, "kg_tub_seg": 10,
               "kg_abi_pri": 0, "kg_abi_seg": 0, "puntos": 1,
               "kg_primera": 20, "kg_segunda": 10,
               "kg_vend_pri": 0, "kg_vend_seg": 0}
    salida = queries.abrir_en_lineas([entera, partida])
    assert all("cat" in g for g in salida)
    assert salida[0]["cat"] == "SEG"
    assert {g["cat"] for g in salida[1:]} == {"PRI", "SEG"}


def test_el_excel_dice_lo_mismo_que_la_pantalla():
    """El Excel no puede tener su propia versión de la regla."""
    from modules.analisis import views
    fila = {"kg_primera": 0, "kg_segunda": 0,
            "kg_vend_pri": 0, "kg_vend_seg": 42.5}
    assert views._categoria(fila) == queries.categoria_de(fila) == "SEG"


# ---------------------------------------------------------------------------
# El link del día abre LA factura, no una lista ni la factura de al lado
# (dueña 26/08/2026: *"clickeo en el día y me lleva a la factura equivocada"*).
#
# `numf` es el número PELADO y no es único: la venta de Jersey Listado VIN de
# ese día era `NTEN-10919` (nota de entrega, VPM, $5,53) y en la base había otra
# con el mismo 10919 — `001-099-000010919`, AGL, una devolución de junio—. La
# pantalla de la factura desempata con `?doc=<número completo>`; lo único que
# faltaba era guardarlo.
# ---------------------------------------------------------------------------

def _html_de(nombre):
    from pathlib import Path
    return (Path(__file__).resolve().parent.parent / "modules" / "analisis" /
            "templates" / "analisis" / nombre).read_text(encoding="utf-8")


def test_el_dia_abre_la_factura_directo_y_no_la_lista():
    """`/facturas/<numf>`, no `/facturas?q=`: la lista filtrada obliga a un
    click más y con un numf repetido muestra dos."""
    for nombre in ("vendidos.html", "competencia.html"):
        html = _html_de(nombre)
        assert "/facturas?q=" not in html, f"{nombre}: sigue linkeando a la lista"
        assert "/facturas/{{ v.numf }}" in html, nombre


def test_el_link_lleva_el_numero_completo_para_desempatar():
    """Sin `?doc=` un numf repetido abre la otra factura."""
    for nombre in ("vendidos.html", "competencia.html"):
        html = _html_de(nombre)
        i = html.index("/facturas/{{ v.numf }}")
        trozo = html[i:i + 200]
        assert "?doc=" in trozo, f"{nombre}: el link no desempata"
        assert "v.numero" in trozo, f"{nombre}: no manda el número completo"


def test_sin_numero_completo_el_link_sigue_andando():
    """Las filas viejas tienen `numero` en NULL hasta el próximo refresco: el
    link cae a `/facturas/<numf>` pelado en vez de romperse."""
    for nombre in ("vendidos.html", "competencia.html"):
        trozo = _html_de(nombre)
        i = trozo.index("/facturas/{{ v.numf }}")
        assert "{% if v.numero %}" in trozo[i:i + 120], nombre


def test_el_click_en_el_dia_no_abre_tambien_la_fila():
    """La fila entera tiene `onclick=abrir(this)`: sin frenar la propagación,
    tocar el día abría el detalle Y navegaba, y la página saltaba sola."""
    html = _html_de("vendidos.html")
    i = html.index("/facturas/{{ v.numf }}")
    assert "event.stopPropagation()" in html[i:i + 260], (
        "el click en el día se lo lleva también la fila")


def test_lo_vendido_guarda_el_numero_completo():
    """De nada sirve el link si la columna no está."""
    import inspect as _i
    fuente = _i.getsource(queries.vendidos)
    assert "v.numero" in fuente, "la consulta no trae el número completo"
    from pathlib import Path
    mig = (Path(__file__).resolve().parent.parent / "migrations" /
           "0233_parado_venta_numero_completo.sql").read_text(encoding="utf-8")
    assert "ADD COLUMN IF NOT EXISTS numero" in mig


def test_el_detalle_de_competencia_trae_la_ultima_venta_de_saldos():
    """Duena 01/09/2026: "pone en esta tabla la ultima fecha de vendido de
    esa tela como esta en saldos" — mirando /analisis/competencia, el detalle
    que se abre por vendedor.

    La columna tiene que ser la MISMA que /analisis/parado (`f.ultima_venta`,
    de `parado_foto`, congelada antes de la largada) para esa tela+color, no
    inventar una cuenta aparte. `vendido_detalle()` ya tenia ese JOIN por
    subcategoria+color para sacar `forma_fila`; esto solo agrega la columna."""
    import inspect
    fuente = inspect.getsource(queries.vendido_detalle)
    assert "f.ultima_venta" in fuente, (
        "no lee la ultima_venta de parado_foto: seria otra fuente de verdad")
    assert "AS ultima_venta" in fuente


def test_la_tabla_de_competencia_muestra_la_ultima_venta_de_la_tela():
    """La misma columna, en el mismo formato que /analisis/parado
    (`dd/mm/yy`), y como columna APARTE de "Dia" — "Dia" es la fecha de ESTA
    venta puntual; "Ultima" es cuando se habia vendido antes."""
    from pathlib import Path
    t = (Path("modules/analisis/templates/analisis/competencia.html")
         .read_text(encoding="utf-8"))
    assert "<th class=\"n opt\">Última</th>" in t
    assert ("v.ultima_venta.strftime('%d/%m/%y') if v.ultima_venta "
            "else '—'") in t
    # ⚠ No es la misma columna que "Dia" (v.fecha, la venta puntual de hoy).
    assert "v.ultima_venta" in t and "v.fecha" in t


def test_la_calidad_de_una_venta_sale_del_lote_real_del_despacho():
    """CERRADO 01/09/2026. La calidad que decide si un kilo `segunda` puntua
    tiene que ser la del LOTE que de verdad salio por el despacho, no la que
    alguien tipeo en la linea de la factura -- esa puede estar mal (Kiana
    Forro) o, mas comun, coincidir siempre pero dejar pasar telas que se
    venden solas (Fleece 96 Sin Perchar) si no se verifica.

    Verificado en vivo con Metabase el 01/09/2026: de 5.537 lineas vendidas
    desde la largada, el lote real y la factura coinciden en TODAS salvo 38
    sin despacho vinculado (0,7%) -- por eso el join es LEFT con fallback a
    la factura, nunca NOT NULL."""
    fuente = asinfo_parado._JOIN_LOTE_DESPACHO
    assert "detalle_despacho_cliente" in fuente
    assert "codigo_lote" in fuente
    assert "LEFT JOIN" in fuente, "sin LEFT, una venta sin despacho desaparecería"

    linea = asinfo_parado._CALIDAD_LINEA
    assert "lot_desp.id_valor_atributo_2" in linea
    assert "dfc.id_valor_atributo_2" in linea, "fallback a la factura si no hay despacho"

    sql = asinfo_parado._sql_vendido("2026-08-25")
    assert "_JOIN_LOTE_DESPACHO" not in sql, "el f-string tiene que estar YA interpolado"
    assert "detalle_despacho_cliente" in sql
