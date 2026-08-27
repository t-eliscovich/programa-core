"""/stock/asinfo — vista de stock por producto desde Asinfo.

Blueprint propio (aislado de modules/stock/) montado bajo url_prefix="/stock".
Lee `v_saldo_producto_vista` del ERP vía Metabase API. Asinfo NO tiene
costos cargados (confirmado 2026-05-22), así que solo se muestra cantidad
y `precio_ultima_venta` como proxy informativo para el ~50% que lo tiene.

Filtros disponibles (query string):
    q       — busca en código, nombre y tejido
    tejido  — filtra exactamente por categoría (Jersey / Fleece / Pique / etc.)
    color   — filtra por color hex (#ffffff, #000000, …)
    min     — cantidad mínima
"""
from __future__ import annotations

from flask import Blueprint, abort, render_template, request, url_for

from auth import requiere_login, requiere_permiso
from exports import csv_response

stock_asinfo_bp = Blueprint(
    "stock_asinfo",
    __name__,
    template_folder="templates",
)


@stock_asinfo_bp.route("/quimicos")
@requiere_login
@requiere_permiso("tintura.ver")
def quimicos():
    """Stock de químicos desde formulas_app — MISMA cuenta que el balance/flujo.

    Dueña 2026-07-21: esta pantalla usaba la réplica vieja (stock_quimicos_al_dia,
    ~412) y valuaba SIN IVA, así que su total no coincidía con el "Stock Quí." del
    balance. Ahora usa quimico_inv_formulas (réplica exacta de formulas, c/JSONB +
    A44/10) valuada c/IVA → el total da los mismos ~416 que balance/flujo.
    """
    from types import SimpleNamespace

    from filters import today_ec
    error = None
    rows = []
    fecha_corte = today_ec()
    try:
        from modules.informes.quimico_inv_formulas import quimico_final_por_tipo
        _det = quimico_final_por_tipo(fecha_corte, detalle=True) or {}
        for f in (_det.get("filas") or []):
            rows.append(SimpleNamespace(
                num=f["num"], num_visible=f["num_visible"],
                familia=f.get("familia") or "", nombre=f.get("nombre") or "",
                unidad=f.get("unidad") or "",
                stock_al_dia_kg=float(f.get("final") or 0),
                precio_us=float(f.get("us") or 0),
                valor_us=float(f.get("monto_iva") or 0),   # c/IVA (= balance)
                fecha_lectura=None,
            ))
    except Exception as e:  # noqa: BLE001
        error = str(e)

    # Filtros UI
    q = (request.args.get("q") or "").strip().upper()
    familia_filtro = (request.args.get("familia") or "").strip()

    familias_universo = sorted({(r.familia or "") for r in rows if r.familia})

    if familia_filtro:
        rows = [r for r in rows if r.familia == familia_filtro]
    if q:
        rows = [
            r for r in rows
            if q in (r.nombre or "").upper()
            or q in (r.familia or "").upper()
        ]

    # Quedarse solo con los que tienen algo de stock (o que se movieron)
    rows_con_stock = [r for r in rows if abs(r.stock_al_dia_kg) > 0.001]

    total_productos = len(rows_con_stock)
    total_kg = sum(r.stock_al_dia_kg for r in rows_con_stock)
    total_us = sum(r.valor_us for r in rows_con_stock)   # c/IVA = balance

    # Distribución por familia
    por_familia: dict[str, dict] = {}
    for r in rows_con_stock:
        f = r.familia or "(s/familia)"
        slot = por_familia.setdefault(f, {"n": 0, "kg": 0.0, "us": 0.0})
        slot["n"] += 1
        slot["kg"] += r.stock_al_dia_kg
        slot["us"] += r.valor_us
    distribucion = sorted(por_familia.items(), key=lambda kv: -kv[1]["us"])

    if request.args.get("export") == "csv":
        export_rows = [
            {
                "familia": r.familia,
                "num_visible": r.num_visible,
                "nombre": r.nombre,
                "unidad": r.unidad,
                "stock_kg": round(r.stock_al_dia_kg, 3),
                "precio_us": round(r.precio_us, 4),
                "valor_us": round(r.valor_us, 2),
                "fecha_lectura": (
                    r.fecha_lectura.isoformat() if r.fecha_lectura else ""
                ),
            }
            for r in rows_con_stock
        ]
        return csv_response(
            export_rows,
            columnas=[
                ("familia", "Familia"),
                ("num_visible", "N°"),
                ("nombre", "Nombre"),
                ("unidad", "Unidad"),
                ("stock_kg", "Stock"),
                ("precio_us", "U$/unidad"),
                ("valor_us", "Valor U$"),
                ("fecha_lectura", "Última lectura"),
            ],
            filename="stock_quimicos.csv",
        )

    return render_template(
        "stock_asinfo/quimicos.html",
        rows=rows_con_stock,
        total_productos=total_productos,
        total_kg=total_kg,
        total_us=total_us,
        distribucion=distribucion,
        familias_universo=familias_universo,
        familia_filtro=familia_filtro,
        q=q,
        fecha_corte=fecha_corte,
        error=error,
    )


@stock_asinfo_bp.route("/asinfo-lote")
@requiere_login
@requiere_permiso("stock.ver")
def lote():
    """Stock por LOTE desde Asinfo — réplica del reporte 'Stock Valorado por Lote'.

    Sin bodega seleccionada: landing con los totales por bodega (resumen +
    ancla de reconciliación). Con `bodega=<id>`: detalle de lotes de esa
    bodega con atributos (calidad, color, acabado, estampado, título hilo,
    proveedor). Solo cantidad (kg) — los dólares de Asinfo no son confiables.

    Filtros (query string):
        bodega    — id de bodega (51 Hilo / 52 Tela Cruda / 53 PT / ...)
        q         — busca en código / nombre de producto (empujado al SQL)
        tejido    — categoría del producto
        titulo    — título de hilo
        proveedor — proveedor del lote
        calidad   — PRI / SEG
    """
    from modules.asinfo import service as asinfo_service

    error = None
    try:
        bodega_raw = (request.args.get("bodega") or "").strip()
        id_bodega = int(bodega_raw) if bodega_raw else None
    except (TypeError, ValueError):
        id_bodega = None

    # Landing: totales por bodega
    totales = []
    try:
        totales = asinfo_service.stock_asinfo_lote_totales()
    except Exception as e:  # noqa: BLE001
        error = str(e)

    if id_bodega is None:
        return render_template(
            "stock_asinfo/lote.html",
            modo="landing",
            totales=totales,
            error=error,
        )

    # Detalle de una bodega — filtros empujados al SQL en el service.
    q = (request.args.get("q") or "").strip().upper()
    tejido_filtro = (request.args.get("tejido") or "").strip()
    titulo_filtro = (request.args.get("titulo") or "").strip()
    proveedor_filtro = (request.args.get("proveedor") or "").strip()
    calidad_filtro = (request.args.get("calidad") or "").strip()
    color_filtro = (request.args.get("color") or "").strip()

    rows = []
    try:
        rows = asinfo_service.stock_asinfo_lote(
            id_bodega, q=q, tejido=tejido_filtro, titulo=titulo_filtro,
            proveedor=proveedor_filtro, calidad=calidad_filtro, color=color_filtro,
        )
    except Exception as e:  # noqa: BLE001
        error = str(e)

    bodega_nombre = next(
        (t["bodega"] for t in totales if t["id_bodega"] == id_bodega),
        f"Bodega {id_bodega}",
    )

    # KPIs: totales del set filtrado COMPLETO (window COUNT/SUM), no del recorte.
    total_lotes = rows[0]["_total_lotes"] if rows else 0
    total_kg = rows[0]["_total_kg"] if rows else 0.0
    mostrando = len(rows)
    hay_mas = total_lotes > mostrando

    # Columnas de atributo visibles = sólo las que APORTAN: tienen ≥2 valores
    # distintos en esta bodega. Hilo no tiene atributos → tabla mínima
    # Producto/Lote/Saldo. Una columna con un único valor repetido (ej.
    # Estampado='SE' en todo PT) es ruido → se oculta. El color sólo cuenta si
    # difiere de la categoría (en crudo viene 'TELA CRUDA').
    def _distintos(key, vs_tejido=False):
        vals = set()
        for r in rows:
            v = (r.get(key) or "").strip()
            if v and (not vs_tejido or v != (r.get("tejido") or "").strip()):
                vals.add(v)
        return vals
    cols = {
        "color": len(_distintos("color", vs_tejido=True)) >= 2,
        "calidad": len(_distintos("calidad")) >= 2,
        "titulo_hilo": len(_distintos("titulo_hilo")) >= 2,
        "proveedor": len(_distintos("proveedor")) >= 2,
        "estampado": len(_distintos("estampado")) >= 2,
    }

    # Universos para los dropdowns (de lo traído).
    tejidos_universo = sorted({(r.get("tejido") or "") for r in rows if r.get("tejido")})
    titulos_universo = sorted({(r.get("titulo_hilo") or "") for r in rows if r.get("titulo_hilo")})
    proveedores_universo = sorted({(r.get("proveedor") or "") for r in rows if r.get("proveedor")})
    calidades_universo = sorted({(r.get("calidad") or "") for r in rows if r.get("calidad")})
    colores_universo = sorted({
        r.get("color") for r in rows
        if r.get("color") and r.get("color") != r.get("tejido")
    })

    if request.args.get("export") == "csv":
        columnas = [("codigo", "Código"), ("producto", "Producto"), ("lote", "Lote")]
        if cols["calidad"]:
            columnas.append(("calidad", "Calidad"))
        if cols["titulo_hilo"]:
            columnas.append(("titulo_hilo", "Título Hilo"))
        if cols["proveedor"]:
            columnas.append(("proveedor", "Proveedor"))
        if cols["color"]:
            columnas.append(("color", "Color"))
        if cols["estampado"]:
            columnas.append(("estampado", "Estampado"))
        columnas += [("unidad", "Unidad"), ("saldo", "Saldo")]
        return csv_response(rows, columnas=columnas, filename=f"stock_lote_{id_bodega}.csv")

    return render_template(
        "stock_asinfo/lote.html",
        modo="detalle",
        rows=rows,
        totales=totales,
        id_bodega=id_bodega,
        bodega_nombre=bodega_nombre,
        total_lotes=total_lotes,
        total_kg=total_kg,
        mostrando=mostrando,
        hay_mas=hay_mas,
        cols=cols,
        tejidos_universo=tejidos_universo,
        titulos_universo=titulos_universo,
        proveedores_universo=proveedores_universo,
        calidades_universo=calidades_universo,
        colores_universo=colores_universo,
        q=q,
        tejido_filtro=tejido_filtro,
        titulo_filtro=titulo_filtro,
        proveedor_filtro=proveedor_filtro,
        calidad_filtro=calidad_filtro,
        color_filtro=color_filtro,
        error=error,
    )


@stock_asinfo_bp.route("/asinfo")
@requiere_login
@requiere_permiso("stock.ver")
def lista():
    """Stock por producto en Asinfo. Read-only, cantidad + precio referencial."""
    try:
        min_saldo = float(request.args.get("min") or 0)
    except (TypeError, ValueError):
        min_saldo = 0.0

    q = (request.args.get("q") or "").strip().upper()
    tejido_filtro = (request.args.get("tejido") or "").strip()
    color_filtro = (request.args.get("color") or "").strip().upper()

    # Tabs por bodega (Hilo / Tela Cruda / Tela Terminada). Sin bodega = todas.
    BODEGAS_TABS = [
        (51, "Hilo"),
        (52, "Tela Cruda"),
        (53, "Tela Terminada"),
    ]
    try:
        bodega_raw = (request.args.get("bodega") or "").strip()
        id_bodega = int(bodega_raw) if bodega_raw else None
    except (TypeError, ValueError):
        id_bodega = None

    error = None
    rows = []
    try:
        from modules.asinfo import service as asinfo_service
        rows = asinfo_service.stock_asinfo(min_saldo=min_saldo, id_bodega=id_bodega)
    except Exception as e:  # noqa: BLE001
        error = str(e)

    # Catálogos para los dropdowns — SE CALCULAN SOBRE TODO EL UNIVERSO,
    # no sobre lo filtrado, para que la dueña vea siempre el mismo set.
    tejidos_universo = sorted({(r.get("tejido") or "") for r in rows if r.get("tejido")})
    # `color` ahora es el código de color extraído del nombre (BLA/NEG/MAR/etc.)
    colores_universo = sorted({(r.get("color") or "") for r in rows if r.get("color")})

    # Aplicar filtros (post-cache, en memoria — son ~3500 filas, irrelevante)
    if tejido_filtro:
        rows = [r for r in rows if r.get("tejido") == tejido_filtro]
    if color_filtro:
        rows = [r for r in rows if (r.get("color") or "").upper() == color_filtro]
    if q:
        rows = [
            r for r in rows
            if q in (r.get("codigo") or "").upper()
            or q in (r.get("nombre") or "").upper()
            or q in (r.get("tejido") or "").upper()
            or q in (r.get("subcategoria") or "").upper()
        ]

    # Stats
    total_productos = len(rows)
    total_unidades = sum(r["cantidad_total"] for r in rows)
    valor_proxy = sum(
        r["cantidad_total"] * r["precio_ultima"]
        for r in rows if r["precio_ultima"] > 0
    )
    productos_con_precio = sum(1 for r in rows if r["precio_ultima"] > 0)

    # Distribución por tejido (siempre sobre lo filtrado, para que ayude
    # a explorar): label → {n, kg}
    por_tejido: dict[str, dict] = {}
    for r in rows:
        t = r.get("tejido") or "(s/categoría)"
        slot = por_tejido.setdefault(t, {"n": 0, "kg": 0.0})
        slot["n"] += 1
        slot["kg"] += r["cantidad_total"]
    distribucion = sorted(por_tejido.items(), key=lambda kv: -kv[1]["kg"])

    if request.args.get("export") == "csv":
        return csv_response(
            rows,
            columnas=[
                ("codigo", "Código"),
                ("nombre", "Nombre"),
                ("tejido", "Tejido"),
                ("subcategoria", "Subcategoría"),
                ("color", "Color (hex)"),
                ("cantidad_total", "Cantidad"),
                ("n_bodegas", "Bodegas"),
                ("precio_ultima", "Precio última venta (US)"),
            ],
            filename="stock_asinfo.csv",
        )

    return render_template(
        "stock_asinfo/lista.html",
        rows=rows,
        total_productos=total_productos,
        total_unidades=total_unidades,
        valor_proxy=valor_proxy,
        productos_con_precio=productos_con_precio,
        q=q,
        min_saldo=min_saldo,
        tejido_filtro=tejido_filtro,
        color_filtro=color_filtro,
        tejidos_universo=tejidos_universo,
        colores_universo=colores_universo,
        distribucion=distribucion,
        bodegas_tabs=BODEGAS_TABS,
        id_bodega=id_bodega,
        error=error,
    )


@stock_asinfo_bp.route("/en-proceso")
@requiere_login
@requiere_permiso("stock.ver")
def en_proceso():
    """Stock EN PROCESO (WIP entre pasos): material despachado a órdenes de
    fabricación abiertas pero todavía no devuelto como el producto siguiente."""
    from modules.asinfo import service as asinfo_service

    error = None
    data = {"pasos": [], "ofts": []}
    try:
        data = asinfo_service.stock_en_proceso()
    except Exception as e:  # noqa: BLE001
        error = str(e)

    pasos = data.get("pasos", [])
    ofts = data.get("ofts", [])
    total_en_proceso = sum(p.get("en_proceso", 0) for p in pasos)

    paso_filtro = (request.args.get("paso") or "").strip()
    if paso_filtro:
        try:
            ofts = [o for o in ofts if o.get("id_bodega") == int(paso_filtro)]
        except (TypeError, ValueError) as _e:
            from modules._lib.silencios import avisar
            avisar(__name__, "en_proceso", _e, nivel="debug")

    if request.args.get("export") == "csv":
        return csv_response(
            ofts,
            columnas=[
                ("paso", "Paso"),
                ("oft", "Orden Fabricación"),
                ("prod_codigo", "Código"),
                ("producto", "Producto"),
                ("planif", "Planificado"),
                ("fab", "Fabricado"),
                ("issued", "Material despachado"),
                ("en_proceso", "En proceso"),
            ],
            filename="stock_en_proceso.csv",
        )

    return render_template(
        "stock_asinfo/en_proceso.html",
        pasos=pasos,
        ofts=ofts,
        total_en_proceso=total_en_proceso,
        paso_filtro=paso_filtro,
        error=error,
    )


# ---------------------------------------------------------------------------
# Fabricación TC / PT — las DOS tabs de Stock (TMT 2026-06-10 dueña)
# ---------------------------------------------------------------------------
# Réplica del Excel "Saldos Inventarios Proceso Produccion Nube": una página
# por proceso (TC = hilo→tela cruda en bodega 52; PT = tela cruda→terminado
# en bodega 53), con los totales de TODO el stock arriba y el resto de las
# vistas históricas (por producto, por lote, importaciones) como secciones
# dentro de la misma página. Las URLs viejas siguen vivas, sólo salen del menú.

_PROCESOS = {
    "tc": {"bodega": 52, "titulo": "Inventario TC",
           "material": "HILO", "produce": "TELA CRUDA",
           "bodega_material": 51, "bodega_produce": 52},
    "pt": {"bodega": 53, "titulo": "Inventario PT",
           "material": "TELA CRUDA", "produce": "PRODUCTO TERMINADO",
           "bodega_material": 52, "bodega_produce": 53},
}


def _fabricacion_page(proceso: str):
    from modules.asinfo import service as asinfo_service

    cfg = _PROCESOS.get((proceso or "").lower())
    if not cfg:
        abort(404)

    error = None
    data = {"resumen": {}, "por_tejido": [], "ofts": []}
    try:
        data = asinfo_service.fabricacion_proceso(cfg["bodega"])
    except Exception as e:  # noqa: BLE001
        error = str(e)

    # Totales de TODO el stock (kg por bodega Asinfo) — arriba de la página.
    # Sólo las 3 bodegas del flujo, en orden de proceso: Hilo → TC → PT.
    totales_bodega = []
    _tot: dict = {}
    try:
        _tot = {t["id_bodega"]: t for t in asinfo_service.stock_asinfo_lote_totales()}
        totales_bodega = [_tot[b] for b in (51, 52, 53) if b in _tot]
    except Exception:  # noqa: BLE001
        totales_bodega = []

    # TMT 2026-06-10 dueña: "falta este saldo para los totales de arriba —
    # hacelo lógico y completo". El stock TOTAL en kg es la cadena del flujo:
    #   Hilo + en proceso TC + Tela Cruda + en proceso PT + Prod. Terminado
    # (los saldos en proceso no están en el saldo de ninguna bodega). Por eso
    # acá se traen LOS DOS procesos, no sólo el de la tab actual.
    otro_bodega = 53 if cfg["bodega"] == 52 else 52
    try:
        data_otro = asinfo_service.fabricacion_proceso(otro_bodega)
    except Exception:  # noqa: BLE001
        data_otro = {"resumen": {}}
    _res_actual = data.get("resumen", {}) or {}
    _res_otro = data_otro.get("resumen", {}) or {}
    saldo_tc = _res_actual if cfg["bodega"] == 52 else _res_otro
    saldo_pt = _res_actual if cfg["bodega"] == 53 else _res_otro

    def _kg(b):
        return float((_tot.get(b) or {}).get("total_kg") or 0)

    def _lotes(b):
        return int((_tot.get(b) or {}).get("lotes") or 0)

    # TMT 2026-08-11 (dueña): *"quizás en hilo hay 9k que ya salieron pero no se
    # descontaron hasta que esté la OFT"*. El material despachado hoy sin orden
    # se sostiene hasta las 18 (ver `hilo_sin_of`), y sin este renglón el
    # "Hilo total" cierra por un número que no está escrito en ningún lado. Va
    # como UN ESLABÓN MÁS de la cadena —no como alarma: la alarma es la
    # campanita— y desaparece solo cuando no hay nada esperando.
    try:
        from modules.asinfo import hilo_sin_of as _hso
        _esperando = _hso.esperando_orden_kg()
    except Exception:  # noqa: BLE001 -- nunca romper la pantalla
        _esperando = {}
    esp_hilo = float(_esperando.get(51) or 0)
    esp_cruda = float(_esperando.get(52) or 0)

    cadena = [
        {"label": "Hilo", "kg": _kg(51), "sub": f"{_lotes(51):,} lotes".replace(",", "."),
         "url": url_for("stock_asinfo.lote", bodega=51), "actual": False},
        {"label": "En proceso TC", "kg": float(saldo_tc.get("saldo") or 0),
         "sub": f"{int(saldo_tc.get('n_ofts') or 0)} órdenes",
         "url": url_for("stock_asinfo.fabricacion_tc"), "actual": proceso == "tc"},
        {"label": "Tela Cruda", "kg": _kg(52), "sub": f"{_lotes(52):,} lotes".replace(",", "."),
         "url": url_for("stock_asinfo.lote", bodega=52), "actual": False},
        {"label": "En proceso PT", "kg": float(saldo_pt.get("saldo") or 0),
         "sub": f"{int(saldo_pt.get('n_ofts') or 0)} órdenes",
         "url": url_for("stock_asinfo.fabricacion_pt"), "actual": proceso == "pt"},
        {"label": "Prod. Terminado", "kg": _kg(53), "sub": f"{_lotes(53):,} lotes".replace(",", "."),
         "url": url_for("stock_asinfo.lote", bodega=53), "actual": False},
    ]
    total_kg = sum(c["kg"] for c in cadena)

    # Métricas combinadas pedidas por la dueña (2026-06-26):
    #   Hilo total  = Hilo (bodega 51) + En proceso TC (hilo despachado a tejer)
    #   Cruda total = Tela Cruda (bodega 52) + En proceso PT (TC despachada a tinturar)
    hilo_total = _kg(51) + float(saldo_tc.get("saldo") or 0) + esp_hilo
    cruda_total = _kg(52) + float(saldo_pt.get("saldo") or 0) + esp_cruda

    # TMT 2026-07-08 (dueña): "encolumná acá como stock: columna por etapa pero
    # no a lo largo, que sea para abajo también." → la cadena horizontal pasa a
    # ser una TABLA vertical (una fila por etapa, en orden de producción). Se
    # reusan EXACTAMENTE los mismos kg de la cadena/combos; no se recalcula nada.
    # TMT 2026-07-08 (dueña): sin nº de bodega en las etiquetas + fila TOTAL al
    # final (total_kg = toda la cadena Hilo→proc→Cruda→proc→Terminado).
    # El renglón "Despachado, esperando orden de fabricación" sólo aparece
    # cuando hay algo esperando. Sin él, "Hilo total" cerraría por un número
    # que no está escrito en ninguna parte de la tabla.
    _espera = {"label": "Despachado, esperando orden de fabricación",
               "bold": False, "espera": True}
    filas = [
        {"label": "Hilo", "kg": _kg(51), "bold": False},
        {"label": "En proceso (Hilo → Tela Cruda)",
         "kg": float(saldo_tc.get("saldo") or 0), "bold": False},
        *([dict(_espera, kg=esp_hilo)] if esp_hilo > 0 else []),
        {"label": "Hilo total", "kg": hilo_total, "bold": True},
        {"label": "Tela Cruda", "kg": _kg(52), "bold": False},
        {"label": "En proceso (Tela Cruda → Terminada)",
         "kg": float(saldo_pt.get("saldo") or 0), "bold": False},
        *([dict(_espera, kg=esp_cruda)] if esp_cruda > 0 else []),
        {"label": "Cruda total", "kg": cruda_total, "bold": True},
        {"label": "Prod. Terminado", "kg": _kg(53), "bold": False},
        {"label": "TOTAL", "kg": total_kg, "bold": True, "grand": True},
    ]

    # TMT 2026-07-06 (dueña): se eliminó el toggle ?pend (pagado/pendiente).
    # TODO lo recibido cuenta en el stock oficial, sin distinción — la
    # "deuda" de importaciones dejó de predecirse (el restante entra por
    # /compras y el pasivo real vive en posdat).

    # Valor del stock del programa (kg + $) — mismo cálculo que el Balance.
    stock_programa = {}
    try:
        from modules.stock import queries as stock_queries

        stock_programa = stock_queries.resumen_stock()
    except Exception:  # noqa: BLE001
        stock_programa = {}

    # La vista "Inventario" ya no muestra las órdenes de fabricación (OFT):
    # se conserva sólo el inventario en kg y el "Inventario en proceso".
    # fabricacion_proceso() sigue trayendo ofts/por_tejido para calcular el
    # resumen (saldo en proceso), pero no se pasan al template.

    return render_template(
        "stock_asinfo/fabricacion.html",
        proceso=proceso,
        cfg=cfg,
        resumen=data.get("resumen", {}),
        saldo_tc=saldo_tc,
        saldo_pt=saldo_pt,
        totales_bodega=totales_bodega,
        cadena=cadena,
        total_kg=total_kg,
        hilo_total=hilo_total,
        cruda_total=cruda_total,
        filas=filas,
        stock_programa=stock_programa,
        error=error,
    )


@stock_asinfo_bp.route("/fabricacion-tc")
@requiere_login
@requiere_permiso("stock.ver")
def fabricacion_tc():
    return _fabricacion_page("tc")


@stock_asinfo_bp.route("/fabricacion-pt")
@requiere_login
@requiere_permiso("stock.ver")
def fabricacion_pt():
    return _fabricacion_page("pt")


# ---------------------------------------------------------------------------
# Consumo Hilado — pestaña de Inventario (dueña 27/08/2026)
# ---------------------------------------------------------------------------
# Réplica del Excel "Consumos de Material - Resumen" (Power Query a Asinfo):
# kg de HILO despachados por orden de salida de material, por material y mes,
# más el promedio mensual y el saldo actual en inventario. Igual al Excel:
# consumo bruto (sin restar devoluciones), sólo categoría HILO.

_MESES_CORTOS = ["ene", "feb", "mar", "abr", "may", "jun",
                 "jul", "ago", "sep", "oct", "nov", "dic"]


def _consumo_hilado_arma(rows, anio, hoy):
    """Pivotea las filas (mes, material, kg, lineas) del servicio.

    Devuelve dict con:
        meses      — [(nro, 'ene'), …] sólo los meses con data, en orden.
        materiales — [{material, por_mes: {mes: {kg, pct}}, total, pct,
                       prom_linea, prom_mes}] orden total DESC.
        tot_mes    — {mes: kg} totales por columna.
        total      — kg del año.
        n_meses    — divisor del promedio mensual (meses transcurridos).
    """
    por_mat: dict = {}
    tot_mes: dict = {}
    for r in rows:
        m = por_mat.setdefault(r["material"], {"por_mes": {}, "total": 0.0, "lineas": 0})
        m["por_mes"][r["mes"]] = m["por_mes"].get(r["mes"], 0.0) + r["kg"]
        m["total"] += r["kg"]
        m["lineas"] += r["lineas"]
        tot_mes[r["mes"]] = tot_mes.get(r["mes"], 0.0) + r["kg"]

    meses = [(n, _MESES_CORTOS[n - 1]) for n in sorted(tot_mes)]
    total = sum(tot_mes.values())

    # Divisor del promedio mensual: igual que la columna "Promedio Mensual
    # Real (Total ÷ 8)" del Excel — meses transcurridos del año en curso,
    # 12 para años cerrados.
    if anio == hoy.year:
        n_meses = max(hoy.month, 1)
    elif anio < hoy.year:
        n_meses = 12
    else:
        n_meses = 1

    materiales = []
    for nombre, m in sorted(por_mat.items(), key=lambda kv: -kv[1]["total"]):
        por_mes = {}
        for n, _ in meses:
            kg = m["por_mes"].get(n, 0.0)
            t = tot_mes.get(n) or 0.0
            por_mes[n] = {"kg": kg, "pct": (kg / t * 100) if t else 0.0}
        materiales.append({
            "material": nombre,
            "por_mes": por_mes,
            "total": m["total"],
            "pct": (m["total"] / total * 100) if total else 0.0,
            # "Promedio" del pivot del Excel: AVG por renglón de despacho.
            "prom_linea": (m["total"] / m["lineas"]) if m["lineas"] else 0.0,
            # "Promedio Mensual Real": total ÷ meses transcurridos.
            "prom_mes": m["total"] / n_meses,
        })
    return {"meses": meses, "materiales": materiales, "tot_mes": tot_mes,
            "total": total, "n_meses": n_meses}


def _consumo_hilado_datos():
    """Trae y arma TODO lo que muestran la pantalla y el export (mismo
    filtro de año: lo que se baja es lo que se ve)."""
    from filters import today_ec
    from modules.asinfo import service as asinfo_service

    hoy = today_ec()
    try:
        anio = int(request.args.get("anio") or hoy.year)
    except (TypeError, ValueError):
        anio = hoy.year
    anio = min(max(anio, 2022), hoy.year)  # Asinfo tiene data desde 2022

    rows, ok = asinfo_service.consumo_hilado_mensual(anio)
    saldos, saldos_ok = asinfo_service.consumo_hilado_saldos()

    data = _consumo_hilado_arma(rows, anio, hoy)
    total_saldos = sum(s["kg"] for s in saldos)

    # Promedio y saldo PEGADOS en una sola tabla (dueña 27/08/2026): mismas
    # filas de material que el cuadro grande, con el saldo al lado. Los
    # materiales que solo tienen saldo (sin consumo en el año) van al final.
    saldo_por_mat = {s["material"]: s["kg"] for s in saldos}
    combinado = [{
        "material": m["material"],
        "prom_linea": m["prom_linea"],
        "prom_mes": m["prom_mes"],
        "saldo": saldo_por_mat.pop(m["material"], None),
    } for m in data["materiales"]]
    combinado += [{"material": mat, "prom_linea": None, "prom_mes": None,
                   "saldo": kg}
                  for mat, kg in sorted(saldo_por_mat.items(),
                                        key=lambda kv: -kv[1])]

    return {
        "anio": anio,
        "anios": list(range(hoy.year, 2021, -1)),
        "ok": ok,
        "saldos_ok": saldos_ok,
        "combinado": combinado,
        "total_saldos": total_saldos,
        **data,
    }


@stock_asinfo_bp.route("/consumo-hilado")
@requiere_login
@requiere_permiso("stock.ver")
def consumo_hilado():
    return render_template("stock_asinfo/consumo_hilado.html",
                           **_consumo_hilado_datos())


@stock_asinfo_bp.route("/consumo-hilado/export.xlsx")
@requiere_login
@requiere_permiso("stock.ver")
def consumo_hilado_export():
    """Excel del Consumo Hilado, con el MISMO año que la pantalla.

    Una hoja con los dos cuadros: despachado por mes (kg y % del mes) y,
    abajo, promedio + saldo por material — formato de la casa (encabezado
    oscuro, kg sin decimales, % con un decimal, panes congelados)."""
    import io

    from flask import Response
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    d = _consumo_hilado_datos()
    meses = d["meses"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Consumo Hilado {d['anio']}"

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="0F172A")
    tot_fill = PatternFill("solid", fgColor="F1F5F9")
    tot_border = Border(top=Side(style="medium", color="94A3B8"))
    kg_fmt = "#,##0"
    pct_fmt = "0.0%"

    def _hdr(row, col, texto):
        c = ws.cell(row=row, column=col, value=texto)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
        return c

    # ── Cuadro 1: despachado por mes (kg + % del mes) ──────────────────
    ws.cell(row=1, column=1,
            value=f"Consumo Hilado — {d['anio']}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1,
            value="kg de hilado despachados por orden de salida de material · fuente Asinfo"
            ).font = Font(size=9, color="64748B")

    fila_hdr = 4
    _hdr(fila_hdr, 1, "Material")
    col = 2
    for _, nombre in meses:
        _hdr(fila_hdr, col, nombre)
        _hdr(fila_hdr, col + 1, "%")
        col += 2
    _hdr(fila_hdr, col, "Total")
    _hdr(fila_hdr, col + 1, "%")
    n_cols = col + 1

    r = fila_hdr + 1
    for m in d["materiales"]:
        ws.cell(row=r, column=1, value=m["material"])
        col = 2
        for n, _ in meses:
            ws.cell(row=r, column=col,
                    value=round(m["por_mes"][n]["kg"])).number_format = kg_fmt
            ws.cell(row=r, column=col + 1,
                    value=m["por_mes"][n]["pct"] / 100).number_format = pct_fmt
            col += 2
        ws.cell(row=r, column=col, value=round(m["total"])).number_format = kg_fmt
        ws.cell(row=r, column=col + 1, value=m["pct"] / 100).number_format = pct_fmt
        r += 1

    if d["materiales"]:
        ws.cell(row=r, column=1, value="Total")
        col = 2
        for n, _ in meses:
            ws.cell(row=r, column=col,
                    value=round(d["tot_mes"][n])).number_format = kg_fmt
            col += 2
        ws.cell(row=r, column=col, value=round(d["total"])).number_format = kg_fmt
        ws.cell(row=r, column=col + 1, value=1).number_format = "0%"
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(bold=True)
            cell.fill = tot_fill
            cell.border = tot_border
        r += 1

    # ── Cuadro 2: promedio y saldo, pegados ────────────────────────────
    r += 2
    ws.cell(row=r, column=1,
            value="Promedio del año y saldo actual, por material"
            ).font = Font(bold=True, size=12)
    r += 1
    _hdr(r, 1, "Material")
    _hdr(r, 2, "Promedio por despacho")
    _hdr(r, 3, f"Promedio mensual (total ÷ {d['n_meses']})")
    _hdr(r, 4, "Saldo en inventario")
    r += 1
    for c_ in d["combinado"]:
        ws.cell(row=r, column=1, value=c_["material"])
        if c_["prom_linea"] is not None:
            ws.cell(row=r, column=2,
                    value=round(c_["prom_linea"])).number_format = kg_fmt
        if c_["prom_mes"] is not None:
            ws.cell(row=r, column=3,
                    value=round(c_["prom_mes"])).number_format = kg_fmt
        if c_["saldo"] is not None:
            ws.cell(row=r, column=4,
                    value=round(c_["saldo"])).number_format = kg_fmt
        r += 1
    if d["combinado"]:
        ws.cell(row=r, column=1, value="Total")
        if d["n_meses"]:
            ws.cell(row=r, column=3,
                    value=round(d["total"] / d["n_meses"])).number_format = kg_fmt
        ws.cell(row=r, column=4,
                value=round(d["total_saldos"])).number_format = kg_fmt
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(bold=True)
            cell.fill = tot_fill
            cell.border = tot_border

    ws.column_dimensions["A"].width = 22
    for i in range(2, max(n_cols, 4) + 1):
        letra = get_column_letter(i)
        ws.column_dimensions[letra].width = 9.5 if i % 2 == 0 else 7
    ws.freeze_panes = "B5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="consumo_hilado_{d["anio"]}.xlsx"'},
    )
