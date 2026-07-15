"""Sync dBase en 1 click — TMT 2026-05-28.

Reemplaza el dance manual de CloudShell + S3 + SSM por una sola página:
    1. Tamara entra a /admin/dbase-sync desde el browser
    2. Adjunta dbf-fresh.tar.gz (el que arma el sandbox desde /Files)
    3. El backend lo recibe, lo extrae a C:\\dbase_fresh y corre
       scripts/sync_dbase_actual.py con I_KNOW_THIS_IS_PROD=1
    4. Stream del stdout/stderr en vivo al browser

POSDAT siempre se excluye del extract (la dueña lo edita a mano por ahora).
"""

from __future__ import annotations

import logging
import os
import shutil
import subprocess
import sys
import tarfile
from pathlib import Path

from flask import Blueprint, Response, render_template, request, stream_with_context

from auth import requiere_login, requiere_permiso

_LOG = logging.getLogger("programa_core.admin_dbase")

bp = Blueprint(
    "admin_dbase",
    __name__,
    url_prefix="/admin/dbase-sync",
    template_folder="templates",
)

# Paths en EC2 Windows. En dev (mac/linux) caen a /tmp para no romper imports.
if sys.platform == "win32":
    TARBALL_PATH = Path(r"C:\dbase_fresh.tar.gz")
    EXTRACT_DIR = Path(r"C:\dbase_fresh")
    PYTHON_EXE = Path(r"C:\Python312\python.exe")
    PC_ROOT = Path(r"C:\programa-core")
else:
    TARBALL_PATH = Path("/tmp/dbf-fresh.tar.gz")
    EXTRACT_DIR = Path("/tmp/dbase_fresh")
    PYTHON_EXE = Path(sys.executable)
    PC_ROOT = Path(__file__).resolve().parent.parent.parent

# Archivos que nunca extraemos al destino, aunque vengan en el tarball.
# POSDAT lo edita la dueña a mano; si el sync lo pisa, se rompe el trabajo manual.
NEVER_EXTRACT = {"POSDAT.DBF"}

# Tamaño máximo aceptado del tarball (10MB es holgado — el real pesa ~900KB).
MAX_TARBALL_BYTES = 10 * 1024 * 1024


def _sesiones_conciliacion_abiertas() -> list:
    """Sesiones de conciliación abiertas (cerrada_en IS NULL). Sincronizar con
    una abierta DESCUADRA la sesión: el sync mueve el ledger (libros +
    conciliados) por debajo, los pendientes-programa colapsan, el lado banco
    queda viejo → la diferencia explota. TMT 2026-06-18: pasó con sesión #40.
    """
    try:
        import db
        return db.fetch_all(
            """
            SELECT id, no_banco, COALESCE(usuario, '') AS usuario, abierta_en
              FROM scintela.banco_conciliacion_sesion
             WHERE cerrada_en IS NULL
             ORDER BY abierta_en DESC
            """
        ) or []
    except Exception:
        return []


@bp.route("/", methods=["GET"])
@requiere_login
@requiere_permiso("usuarios.admin")
def form():
    """Formulario para subir el tarball."""
    return render_template(
        "admin_dbase/sync.html",
        never_extract=sorted(NEVER_EXTRACT),
        target_dir=str(EXTRACT_DIR),
        sesiones_abiertas=_sesiones_conciliacion_abiertas(),
    )


@bp.route("/run", methods=["POST"])
@requiere_login
@requiere_permiso("usuarios.admin")
def run():
    """Recibe tarball, extrae, corre sync_dbase_actual.py, streamea stdout."""
    f = request.files.get("tarball")
    if not f or not f.filename:
        return Response("ERROR: falta el archivo 'tarball'.\n", mimetype="text/plain", status=400)

    # Validamos extensión y tamaño antes de tocar disco.
    if not f.filename.lower().endswith((".tar.gz", ".tgz")):
        return Response(
            f"ERROR: extensión inesperada ({f.filename!r}). Esperaba .tar.gz / .tgz.\n",
            mimetype="text/plain",
            status=400,
        )

    # Persistimos el upload.
    TARBALL_PATH.parent.mkdir(parents=True, exist_ok=True)
    if TARBALL_PATH.exists():
        TARBALL_PATH.unlink()
    f.save(TARBALL_PATH)

    size = TARBALL_PATH.stat().st_size
    if size == 0:
        TARBALL_PATH.unlink(missing_ok=True)
        return Response("ERROR: archivo vacío.\n", mimetype="text/plain", status=400)
    if size > MAX_TARBALL_BYTES:
        TARBALL_PATH.unlink(missing_ok=True)
        return Response(
            f"ERROR: tarball pesa {size:,} bytes (max {MAX_TARBALL_BYTES:,}).\n",
            mimetype="text/plain",
            status=400,
        )

    return Response(
        stream_with_context(_run_pipeline(size, f.filename)),
        mimetype="text/plain",
    )


def _run_pipeline(tarball_bytes: int, original_name: str):
    """Generador que loguea cada paso al cliente y dispara el sync."""

    def line(msg: str) -> str:
        return msg.rstrip("\n") + "\n"

    yield line(f"[1/4] tarball recibido: {original_name} ({tarball_bytes:,} bytes)")

    _abiertas = _sesiones_conciliacion_abiertas()
    if _abiertas:
        yield line("")
        yield line("⚠⚠⚠ OJO: hay una sesión de CONCILIACIÓN ABIERTA ⚠⚠⚠")
        for _s in _abiertas:
            yield line(f"   · sesión #{_s.get('id')} (banco {_s.get('no_banco')}, {_s.get('usuario')})")
        yield line("   Sincronizar MUEVE el ledger por debajo de la sesión y la descuadra")
        yield line("   (los pendientes-programa colapsan, el lado banco queda viejo → la")
        yield line("    diferencia explota). Lo ideal: CERRAR la sesión ANTES de sincronizar.")
        yield line("   (El sync continúa igual — esto es solo un aviso.)")
        yield line("")

    # Limpiar el directorio destino antes de extraer.
    try:
        if EXTRACT_DIR.exists():
            shutil.rmtree(EXTRACT_DIR)
        EXTRACT_DIR.mkdir(parents=True)
    except OSError as exc:
        yield line(f"[ERROR] no pude limpiar {EXTRACT_DIR}: {exc}")
        return

    # Extraer, filtrando archivos vetados (POSDAT.DBF).
    extracted: list[str] = []
    skipped: list[str] = []
    try:
        with tarfile.open(TARBALL_PATH, "r:gz") as tar:
            for member in tar.getmembers():
                if not member.isfile():
                    continue
                # Solo el basename para evitar path traversal.
                name = Path(member.name).name
                if name.upper() in NEVER_EXTRACT:
                    skipped.append(name)
                    continue
                # Forzamos extract plano (sin subdirs).
                member.name = name
                tar.extract(member, EXTRACT_DIR)
                extracted.append(name)
    except (tarfile.TarError, OSError) as exc:
        yield line(f"[ERROR] no pude extraer el tarball: {exc}")
        return

    yield line(f"[2/4] extraídos {len(extracted)} archivos a {EXTRACT_DIR}")
    if skipped:
        yield line(f"       saltados (NEVER_EXTRACT): {', '.join(skipped)}")

    # Validar que haya al menos 1 DBF.
    dbfs = sorted(p.name for p in EXTRACT_DIR.glob("*.DBF"))
    if not dbfs:
        yield line(f"[ERROR] no se encontraron *.DBF en {EXTRACT_DIR}; abortando.")
        return
    yield line(f"       DBFs disponibles ({len(dbfs)}): {', '.join(dbfs[:6])}{' …' if len(dbfs) > 6 else ''}")

    # Disparar el sync. Usamos el mismo script que el sync manual.
    sync_script = PC_ROOT / "scripts" / "sync_dbase_actual.py"
    if not sync_script.exists():
        yield line(f"[ERROR] no existe {sync_script}")
        return

    cmd = [
        str(PYTHON_EXE),
        str(sync_script),
        "--source",
        str(EXTRACT_DIR),
    ]
    yield line(f"[3/4] corriendo: {' '.join(cmd)}")
    yield line("")

    env = os.environ.copy()
    env["I_KNOW_THIS_IS_PROD"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUNBUFFERED"] = "1"

    try:
        proc = subprocess.Popen(  # noqa: S603 — argumentos vienen del propio servidor.
            cmd,
            cwd=str(PC_ROOT),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            env=env,
            text=True,
            bufsize=1,
        )
    except OSError as exc:
        yield line(f"[ERROR] no pude lanzar el subprocess: {exc}")
        return

    assert proc.stdout is not None
    try:
        for raw in proc.stdout:
            yield raw if raw.endswith("\n") else raw + "\n"
    finally:
        proc.stdout.close()
        rc = proc.wait()

    yield line("")
    yield line(f"[4/4] sync_dbase_actual.py terminó con exit={rc}")

    # TMT 2026-06-03: post-sync relink de matches. Cada sync DELETE+INSERT
    # transacciones_bancarias y el SERIAL id_transaccion cambia. Los matches
    # quedaban apuntando a ids muertos → conciliados pasaban a pendientes.
    # La función SQL `relink_matches_post_sync` (mig 0066) recupera el id
    # nuevo via tx_firma.
    if rc == 0:
        # TMT 2026-06-03: el SQL function devolvía relinked=N pero no
        # persistía el UPDATE (pl/pgsql + psycopg2 quirk). Llamamos al helper
        # Python directo.
        try:
            from modules.conciliacion.diag_view import _relink_py
            row = _relink_py(10) or {}
            yield line("")
            yield line("[relink matches post-sync]")
            yield line(f"  matches_total : {row.get('matches_total', 0)}")
            yield line(f"  relinked      : {row.get('relinked', 0)}")
            yield line(f"  sin_firma     : {row.get('sin_firma', 0)} (no se podrán recuperar)")
            yield line(f"  sin_match     : {row.get('sin_match', 0)} (id_transaccion huérfano sin firma match)")
        except Exception as exc:
            yield line(f"[WARN] relink falló (no fatal): {exc}")

    # TMT 2026-07-15 (dueña: "arreglalo para futuro" / red de seguridad): después
    # de cada sync, chequear que TODOS los cruces de conciliación cuadren (que los
    # movimientos del banco atados a cada mov del programa sumen su importe). Si
    # algo quedó "corrido" (banco atado al programa equivocado), lo AVISA acá
    # mismo — así se caza el mismo día, no meses después.
    if rc == 0:
        try:
            from modules.conciliacion.diag_view import _chequeo_cruces
            _chk = _chequeo_cruces(10)
            yield line("")
            yield line("[chequeo cruces conciliación]")
            if _chk["n_mal_atados"] == 0:
                yield line(f"  ✓ {_chk['n_programas_revisados']} cruces OK — todos cuadran")
            else:
                yield line(
                    f"  ⚠ {_chk['n_mal_atados']} grupo(s) mal atados detectados — "
                    f"auto-corrigiendo por monto exacto…")
                for _m in _chk["mal_atados"][:8]:
                    yield line(
                        f"    tx {_m['id_transaccion']} ({_m['concepto']}): "
                        f"banco {_m['suma_banco']:,.2f} vs programa {_m['programa']:,.2f} "
                        f"-> diff {_m['diff']:,.2f}")
                # AUTO-HEAL: re-ata por monto exacto (seguro, nunca nulea).
                from modules.conciliacion.diag_view import _reatar_mal_atados
                _fix = _reatar_mal_atados(10, aplicar=True)
                _chk2 = _chequeo_cruces(10)
                yield line(
                    f"  auto-reatado: {_fix['aplicados_total']} cruce(s) corregidos "
                    f"en {_fix['n_grupos']} grupo(s). Quedan mal atados: "
                    f"{_chk2['n_mal_atados']} (si >0, no había programa real del mismo "
                    f"monto — revisar a mano en /admin/diag-pendientes-banco/chequeo-cruces).")
        except Exception as exc:  # noqa: BLE001
            yield line(f"[WARN] chequeo/auto-reatado cruces falló (no fatal): {exc!r}")

    # TMT 2026-07-07 (dueña): reconcile POSDAT FULL en cada sync. El sync NO
    # extrae POSDAT (NEVER_EXTRACT, para no pisar baselines YY + links
    # mov_doble), pero el archivo SIGUE en el tarball. Acá lo sacamos y
    # corremos el reconcile QUIRÚRGICO COMPLETO (banc 0 Y 9, importes Y fechad):
    # borra los dbf-import/reconcile-dbf SIN link mov_doble e inserta todo el
    # POSDAT.DBF, preservando PC-creados/linkeados. Así el flujo y los pasivos
    # quedan alineados al dBase en CADA sync (antes el reconcile viejo solo
    # matcheaba banc=0 por prov/importe y dejaba las importaciones/fechas viejas
    # → el flujo mostraba -2M). Transaccional; no fatal si falla.
    if rc == 0:
        try:
            from modules.admin_dbase.posdat_reconcile_view import reconcile_posdat_full_desde_dbf
            posdat_tmp = EXTRACT_DIR / "_POSDAT_reconcile.DBF"
            encontrado = False
            with tarfile.open(TARBALL_PATH, "r:gz") as tar:
                for m in tar.getmembers():
                    if m.isfile() and Path(m.name).name.upper() == "POSDAT.DBF":
                        m.name = "_POSDAT_reconcile.DBF"
                        tar.extract(m, EXTRACT_DIR)
                        encontrado = True
                        break
            yield line("")
            yield line("[reconcile POSDAT full post-sync]")
            if not encontrado:
                yield line("  (el tarball no traía POSDAT.DBF — se omite el reconcile)")
            else:
                yield from reconcile_posdat_full_desde_dbf(posdat_tmp, aplicar=True)
                posdat_tmp.unlink(missing_ok=True)
        except Exception as exc:  # noqa: BLE001
            yield line(f"[WARN] reconcile POSDAT full falló (no fatal): {exc!r}")

    # TMT 2026-06-10: import de fichas de clientes en CADA sync. CLIENTES.DBF
    # no tiene mapper en sync_dbase_actual.py, así que las altas nuevas del
    # dBase nunca llegaban a PC → las facturas de Asinfo rebotaban con
    # "cliente no existe en PC" (pasó con 3 facturas el 2026-06-09). El
    # import es rellenar-solo + INSERT-si-falta: no pisa datos de PC.
    if rc == 0:
        try:
            from modules.admin_dbase.clientes_import_view import importar_desde_dbf
            cli_dbf = next(
                (pth for pth in EXTRACT_DIR.iterdir()
                 if pth.is_file() and pth.name.upper() == "CLIENTES.DBF"),
                None,
            )
            yield line("")
            yield line("[clientes-import post-sync]")
            if cli_dbf is None:
                yield line("  (el tarball no traía CLIENTES.DBF — se omite; las altas nuevas de clientes NO llegan)")
            else:
                yield from importar_desde_dbf(cli_dbf, aplicar=True, verbose=False)
        except Exception as exc:  # noqa: BLE001
            yield line(f"[WARN] clientes-import falló (no fatal): {exc!r}")

    # TMT 2026-06-23 (dueña): refrescar el catálogo de tintura (tinto_costos) en
    # CADA sync. El catálogo se sembró una sola vez (mig 0083) desde el mes en
    # curso → códigos viejos (LIF/AZU/MEN/RPA/GRC…) faltaban y la planilla
    # /informes/tinto-carga los rechazaba. Usa COSTOS.DBF si vino en el tarball
    # (maestro, pisa color+costo); si no, reconstruye desde TINTO.DBF completo
    # (costo = Σimporte/Σkg) sin pisar costos editados a mano. "Sync completo
    # directo actualiza los colores", como pidió la dueña.
    if rc == 0:
        try:
            from modules.admin_dbase.tinto_costos_sync import refresh_from_dir
            yield line("")
            yield line("[catálogo tintura post-sync]")
            yield from refresh_from_dir(EXTRACT_DIR, aplicar=True)
        except Exception as exc:  # noqa: BLE001
            yield line(f"[WARN] refresh catálogo tintura falló (no fatal): {exc!r}")

    if rc == 0:
        yield line("OK ✓")
    else:
        yield line("FALLO ✗ — revisar el log de arriba.")


# ---------------------------------------------------------------------------
# Refresco SOLO del catálogo de tintura (sin sync completo) — TMT 2026-06-23.
# La dueña: faltan colores (LIF/AZU/MEN/RPA/GRC…) en /informes/tinto-carga.
# Esta pantalla sube un tarball (o un .DBF) con TINTO.DBF/COSTOS.DBF y refresca
# scintela.tinto_costos SIN tocar bancos/facturas/balance — para arreglar los
# colores sin el riesgo de un sync completo. Reproducible 100% por la UI.
# ---------------------------------------------------------------------------
@bp.route("/tinto-costos", methods=["GET"])
@requiere_login
@requiere_permiso("usuarios.admin")
def tinto_costos_form():
    return render_template("admin_dbase/tinto_costos.html")


@bp.route("/tinto-costos/run", methods=["POST"])
@requiere_login
@requiere_permiso("usuarios.admin")
def tinto_costos_run():
    f = request.files.get("archivo")
    if not f or not f.filename:
        return Response("ERROR: falta el archivo.\n", mimetype="text/plain", status=400)
    aplicar = (request.form.get("aplicar") or "").lower() in ("1", "true", "on", "si", "sí")
    fname = f.filename
    low = fname.lower()
    if not low.endswith((".tar.gz", ".tgz", ".dbf")):
        return Response(
            f"ERROR: extensión inesperada ({fname!r}). Esperaba .tar.gz / .tgz / .DBF.\n",
            mimetype="text/plain", status=400,
        )

    work = EXTRACT_DIR.parent / "_tinto_costos_refresh"
    try:
        if work.exists():
            shutil.rmtree(work)
        work.mkdir(parents=True)
    except OSError as exc:
        return Response(f"ERROR: no pude preparar {work}: {exc}\n", mimetype="text/plain", status=500)

    if low.endswith(".dbf"):
        f.save(work / Path(fname).name)
    else:
        tmp_tar = work / "_upload.tar.gz"
        f.save(tmp_tar)
        try:
            with tarfile.open(tmp_tar, "r:gz") as tar:
                for m in tar.getmembers():
                    if not m.isfile():
                        continue
                    name = Path(m.name).name
                    if name.upper() in ("TINTO.DBF", "COSTOS.DBF"):
                        m.name = name
                        tar.extract(m, work)
        except (tarfile.TarError, OSError) as exc:
            return Response(f"ERROR: no pude extraer el tarball: {exc}\n", mimetype="text/plain", status=400)
        tmp_tar.unlink(missing_ok=True)

    def _stream():
        from modules.admin_dbase.tinto_costos_sync import refresh_from_dir
        yield f"[catálogo tintura] {'APLICAR' if aplicar else 'DRY-RUN'} desde {fname}\n\n"
        try:
            for ln in refresh_from_dir(work, aplicar=aplicar):
                yield ln.rstrip("\n") + "\n"
            yield "\nOK ✓\n" if aplicar else "\nDRY-RUN ✓ (no se escribió nada — tildá 'aplicar' para guardar)\n"
        except Exception as exc:  # noqa: BLE001
            yield f"\n[ERROR] {exc!r}\n"
        finally:
            shutil.rmtree(work, ignore_errors=True)

    return Response(stream_with_context(_stream()), mimetype="text/plain")


# ---------------------------------------------------------------------------
# Importar TODOS los colores desde formulas_app — TMT 2026-06-29 (dueña).
# El catálogo scintela.tinto_costos solo tenía ~118 códigos (COSTOS.DBF) y
# faltaban colores. formulas_app (la app de tintorería, bridge read-only
# formulas_db) es el maestro vivo de fórmulas/colores. Este botón los trae
# TODOS y los upsertea al catálogo (conservador: inserta faltantes con costo 0,
# rellena color vacío, NO pisa costos/colores ya cargados). Sin tarball: lee
# directo de formulas_app por el pool read-only. Reproducible 100% por la UI.
# ---------------------------------------------------------------------------
@bp.route("/tinto-costos/from-formulas", methods=["POST"])
@requiere_login
@requiere_permiso("usuarios.admin")
def tinto_costos_from_formulas():
    aplicar = (request.form.get("aplicar") or "").lower() in ("1", "true", "on", "si", "s\u00ed")

    def _stream():
        from modules.admin_dbase.tinto_costos_sync import refresh_from_formulas_app
        yield f"[cat\u00e1logo tintura \u2190 formulas_app] {'APLICAR' if aplicar else 'DRY-RUN'}\n\n"
        try:
            for ln in refresh_from_formulas_app(aplicar=aplicar):
                yield ln.rstrip("\n") + "\n"
            yield "\nOK \u2713\n" if aplicar else "\nDRY-RUN \u2713 (no se escribi\u00f3 nada \u2014 tild\u00e1 'aplicar' para guardar)\n"
        except Exception as exc:  # noqa: BLE001
            yield f"\n[ERROR] {exc!r}\n"

    return Response(stream_with_context(_stream()), mimetype="text/plain")


# ---------------------------------------------------------------------------
# Sync STAT desde xlsx exportado del DBF — TMT 2026-05-28 (dueña).
# Pedido: 'make sure these are the only conciliated movements — the *'.
# Cuando entre syncs DBF→PC, alguien marcó conciliados en PC (vía
# matcher_banco dual-write antes del fix), hay filas con stat='*' colgado.
# Este endpoint alinea SCINTELA.TRANSACCIONES_BANCARIAS.stat con el archivo
# de dBase (PICHINCH.xlsx) usando (no_banco, fecha, doc, importe, saldo)
# como clave natural — saldo running balance es único por fila.
# ---------------------------------------------------------------------------
@bp.route("/stat-xlsx", methods=["GET"])
@requiere_login
@requiere_permiso("usuarios.admin")
def stat_xlsx_form():
    return render_template("admin_dbase/stat_xlsx.html")


@bp.route("/stat-xlsx/run", methods=["POST"])
@requiere_login
@requiere_permiso("usuarios.admin")
def stat_xlsx_run():

    import db as _db

    f = request.files.get("xlsx")
    if not f or not f.filename:
        return Response("ERROR: falta el archivo 'xlsx'.\n",
                        mimetype="text/plain", status=400)
    if not f.filename.lower().endswith(".xlsx"):
        return Response("ERROR: esperaba un .xlsx.\n",
                        mimetype="text/plain", status=400)

    try:
        no_banco = int(request.form.get("no_banco") or "10")
    except (TypeError, ValueError):
        no_banco = 10

    try:
        from openpyxl import load_workbook
    except ImportError:
        return Response("ERROR: openpyxl no instalado en el server.\n",
                        mimetype="text/plain", status=500)

    def _line(msg: str) -> str:
        return msg.rstrip("\n") + "\n"

    def _stream():
        yield _line(f"[1/4] archivo recibido: {f.filename} · no_banco={no_banco}")
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
        except Exception as exc:
            yield _line(f"[ERROR] no pude abrir el xlsx: {exc}")
            return
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            yield _line("[ERROR] xlsx vacío.")
            return
        # Esperamos header: FECHA DOC PROV CONCEPTO NUM FECHAD IMPORTE SALDO STAT CLAVE
        header = [(c or "").strip().upper() if isinstance(c, str) else c
                  for c in rows[0]]
        needed = {"FECHA": 0, "DOC": 1, "IMPORTE": 6, "SALDO": 7, "STAT": 8}
        try:
            ix = {k: header.index(k) for k in needed}
        except ValueError as exc:
            yield _line(f"[ERROR] header no esperado: {exc} · header={header}")
            return

        # Clave natural: (fecha, doc, importe, saldo). Valor: stat ('*' o '').
        def _key(fecha, doc, importe, saldo):
            f_iso = fecha.date().isoformat() if hasattr(fecha, "date") else str(fecha)
            d = (doc or "").strip().upper()
            try:
                imp = round(float(importe or 0), 2)
            except (TypeError, ValueError):
                imp = 0.0
            try:
                sal = round(float(saldo or 0), 2)
            except (TypeError, ValueError):
                sal = 0.0
            return (f_iso, d, imp, sal)

        xlsx_stat: dict = {}
        for r in rows[1:]:
            try:
                k = _key(r[ix["FECHA"]], r[ix["DOC"]],
                          r[ix["IMPORTE"]], r[ix["SALDO"]])
                s = (r[ix["STAT"]] or "").strip() if isinstance(r[ix["STAT"]], str) \
                    else ("" if r[ix["STAT"]] is None else str(r[ix["STAT"]]).strip())
                xlsx_stat[k] = "*" if s == "*" else ""
            except Exception:
                continue
        n_estrella = sum(1 for v in xlsx_stat.values() if v == "*")
        yield _line(f"[2/4] xlsx parseado: {len(xlsx_stat)} filas únicas, "
                    f"{n_estrella} con stat='*'.")

        # Traer todas las filas PC del banco.
        pg_rows = _db.fetch_all(
            "SELECT id_transaccion, fecha, documento, importe, saldo, "
            "       TRIM(COALESCE(stat,'')) AS stat "
            "  FROM scintela.transacciones_bancarias WHERE no_banco = %s",
            (no_banco,),
        ) or []
        yield _line(f"       PG tiene {len(pg_rows)} filas para no_banco={no_banco}.")

        to_set_estrella: list[int] = []   # PG NULL → '*'
        to_set_null: list[int] = []       # PG '*' → NULL
        no_match: list[int] = []          # no aparece en xlsx
        ok_match: int = 0

        for r in pg_rows:
            k = _key(r["fecha"], r["documento"], r["importe"], r["saldo"])
            pg_stat = (r["stat"] or "").strip()
            if k not in xlsx_stat:
                no_match.append(r["id_transaccion"])
                continue
            xl_stat = xlsx_stat[k]
            if xl_stat == "*" and pg_stat != "*":
                to_set_estrella.append(r["id_transaccion"])
            elif xl_stat != "*" and pg_stat == "*":
                to_set_null.append(r["id_transaccion"])
            else:
                ok_match += 1

        yield _line(
            f"[3/4] plan: marcar '*' en {len(to_set_estrella)} fila(s), "
            f"limpiar '*' en {len(to_set_null)} fila(s). "
            f"OK sin cambios: {ok_match}. Sin match en xlsx: {len(no_match)}."
        )

        # Ejecutar UPDATEs en bloque (chunks para evitar query gigante).
        def _chunks(seq, n=500):
            for i in range(0, len(seq), n):
                yield seq[i:i + n]

        n_up_estrella = 0
        for chunk in _chunks(to_set_estrella):
            n_up_estrella += _db.execute(
                "UPDATE scintela.transacciones_bancarias SET stat = '*' "
                "WHERE id_transaccion = ANY(%s)",
                (chunk,),
            ) or 0
        n_up_null = 0
        for chunk in _chunks(to_set_null):
            n_up_null += _db.execute(
                "UPDATE scintela.transacciones_bancarias SET stat = NULL "
                "WHERE id_transaccion = ANY(%s)",
                (chunk,),
            ) or 0

        yield _line(f"       UPDATEd a '*':  {n_up_estrella}")
        yield _line(f"       UPDATEd a NULL: {n_up_null}")

        if no_match:
            yield _line(f"       OJO: {len(no_match)} fila(s) PC sin "
                        f"contraparte en xlsx (creadas en PC y/o no en DBF).")

        yield _line("")
        yield _line("[4/4] OK. Refrescá /bancos para ver los cambios.")

    return Response(stream_with_context(_stream()), mimetype="text/plain")
