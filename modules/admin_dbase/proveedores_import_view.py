"""Endpoint /admin/proveedores-import — crea/completa proveedores desde FABRICA.

FABRICA.DBF (+ FABRICA.BAK) es el maestro de proveedores del FoxPro: código de
2 letras (FAB) → NOMBRE, EMPRESA, RUC, teléfono, plazo, retenciones, dirección.
Ese maestro NO está en el sync normal, así que los proveedores nunca llegaron a
PC (por eso "el proveedor BP/AC/AQ no existe").

Se sube un .xlsx exportado de FABRICA (header: FAB NOMBRE EMPRESA RUC TELEFONO
TIPO PLAZO RETBASE RETIVA DIRECCION [ACTIVA]). Política:
  - Proveedor que falta en PC  → INSERT con todos los datos.
  - Proveedor que ya existe    → rellena SÓLO los campos vacíos en PC (no pisa).
Idempotente. DRY-RUN por defecto; "Aplicar" escribe en producción.

TMT 2026-06-19 (dueña: "creá todos los que existan en fabrica, no dejes de
importar"). Mirror de clientes_import_view.
"""
from __future__ import annotations

from flask import Blueprint, Response, render_template_string, request, stream_with_context

from auth import requiere_login, requiere_permiso

bp = Blueprint("proveedores_import", __name__, url_prefix="/admin/proveedores-import")

# (col_pc, header_xlsx, maxlen)  — campos string que se rellenan si PC está vacío
FICHA_STR = [
    ("ruc", "RUC", 16),
    ("telefono", "TELEFONO", 30),
    ("representante", "EMPRESA", 100),
    ("tipo", "TIPO", 2),
    ("direccion", "DIRECCION", 200),
]
FICHA_NUM = [("plazo", "PLAZO"), ("retbase", "RETBASE"), ("retiva", "RETIVA")]


def _s(v, maxlen=None) -> str:
    s = ("" if v is None else str(v)).strip()
    return s[:maxlen] if maxlen else s


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _leer_xlsx(file_storage) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [(_s(c).upper()) for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}

    def cell(r, h):
        i = idx.get(h)
        return r[i] if (i is not None and i < len(r)) else None

    out: dict[str, dict] = {}
    for r in rows[1:]:
        cod = _s(cell(r, "FAB"), 5).upper()
        if not cod:
            continue
        nombre = _s(cell(r, "NOMBRE"), 200) or _s(cell(r, "EMPRESA"), 200) or cod
        out[cod] = {
            "codigo_prov": cod,
            "nombre": nombre,
            "ruc": _s(cell(r, "RUC"), 16),
            "telefono": _s(cell(r, "TELEFONO"), 30),
            "representante": _s(cell(r, "EMPRESA"), 100),
            "tipo": _s(cell(r, "TIPO"), 2),
            "direccion": _s(cell(r, "DIRECCION"), 200),
            "plazo": _num(cell(r, "PLAZO")),
            "retbase": _num(cell(r, "RETBASE")),
            "retiva": _num(cell(r, "RETIVA")),
        }
    return list(out.values())


def _leer_pc() -> dict:
    import db
    rows = db.fetch_all(
        "SELECT codigo_prov, nombre, ruc, telefono, representante, tipo, "
        "       direccion, plazo, retbase, retiva FROM scintela.proveedor"
    ) or []
    return {(_s(r["codigo_prov"]).upper()): r for r in rows if _s(r.get("codigo_prov"))}


FORM = """
<!doctype html><meta charset=utf-8><title>Importar proveedores</title>
<div style="max-width:640px;margin:2rem auto;font-family:system-ui">
<h2>Importar proveedores desde FABRICA</h2>
<p>Subí el <b>.xlsx</b> exportado de FABRICA (header: FAB NOMBRE EMPRESA RUC
TELEFONO TIPO PLAZO RETBASE RETIVA DIRECCION). Crea los proveedores que falten
y rellena <b>sólo</b> los campos vacíos de los que ya existen (no pisa datos).
Corre en <b>DRY-RUN</b> salvo que marques Aplicar.</p>
<form method=post action="/admin/proveedores-import/run" enctype="multipart/form-data">
  <input type=hidden name=csrf_token value="{{ csrf_token() }}">
  <input type=file name=xlsx accept=".xlsx" required><br><br>
  <label><input type=checkbox name=apply value=1> Aplicar (escribe en producción)</label><br><br>
  <button type=submit>Correr</button>
</form></div>
"""


@bp.route("/", methods=["GET"])
@requiere_login
@requiere_permiso("usuarios.admin")
def form():
    return render_template_string(FORM)


@bp.route("/run", methods=["POST"])
@requiere_login
@requiere_permiso("usuarios.admin")
def run():
    f = request.files.get("xlsx")
    if not f or not f.filename:
        return Response("ERROR: falta el .xlsx.\n", mimetype="text/plain", status=400)
    if not f.filename.lower().endswith(".xlsx"):
        return Response("ERROR: esperaba un .xlsx.\n", mimetype="text/plain", status=400)
    aplicar = request.form.get("apply") in ("1", "true", "on")
    try:
        filas = _leer_xlsx(f)
    except Exception as exc:  # noqa: BLE001
        return Response(f"ERROR leyendo xlsx: {exc!r}\n", mimetype="text/plain", status=400)
    return Response(stream_with_context(_run(filas, aplicar)), mimetype="text/plain")


def _run(filas: list[dict], aplicar: bool):
    import db
    from modules.proveedores import queries as q

    def line(m=""):
        return m.rstrip("\n") + "\n"

    yield line(f"=== Importar proveedores — {'APLICAR' if aplicar else 'DRY-RUN'} ===")
    if not filas:
        yield line("[ERROR] el xlsx no trajo filas reconocibles (¿header FAB/NOMBRE?).")
        return

    pc = _leer_pc()
    inserts = [d for d in filas if d["codigo_prov"] not in pc]
    existentes = [d for d in filas if d["codigo_prov"] in pc]

    # Updates: rellenar SÓLO campos vacíos en PC.
    updates = []  # (codigo, {col: val})
    for d in existentes:
        cur = pc[d["codigo_prov"]]
        cambios = {}
        for col, _h, _ml in FICHA_STR:
            if d[col] and not _s(cur.get(col)):
                cambios[col] = d[col]
        for col, _h in FICHA_NUM:
            if d[col] is not None and (cur.get(col) is None or float(cur.get(col) or 0) == 0):
                cambios[col] = d[col]
        if cambios:
            updates.append((d["codigo_prov"], cambios))

    yield line(f"xlsx: {len(filas)} proveedores  |  PC ya tiene: {len(pc)}")
    yield line(f"PLAN: {len(inserts)} INSERT (faltan) · {len(updates)} UPDATE (rellenan vacíos)")
    yield line("")
    yield line("--- INSERT (faltan en PC) ---")
    for d in sorted(inserts, key=lambda x: x["codigo_prov"]):
        yield line(f"  {d['codigo_prov']:5} {d['nombre'][:34]:34} ruc={d['ruc'][:14]:14} ret={d['retbase']}/{d['retiva']}")
    yield line("")
    if updates:
        yield line("--- UPDATE (rellenan campos vacíos) ---")
        for cod, cambios in updates[:80]:
            yield line(f"  {cod:5} {', '.join(cambios)}")
        yield line("")

    if not aplicar:
        yield line("DRY-RUN: no se tocó nada. Marcá 'Aplicar' para ejecutar.")
        return

    n_ins = n_up = n_err = 0
    for d in inserts:
        try:
            q.crear(
                codigo_prov=d["codigo_prov"], nombre=d["nombre"],
                ruc=d["ruc"] or None, telefono=d["telefono"] or None,
                representante=d["representante"] or None, tipo=d["tipo"] or None,
                plazo=int(d["plazo"]) if d["plazo"] is not None else None,
                retbase=d["retbase"], retiva=d["retiva"],
                direccion=d["direccion"] or None, usuario="fabrica-import",
            )
            n_ins += 1
        except Exception as exc:  # noqa: BLE001
            n_err += 1
            yield line(f"  [ERR insert {d['codigo_prov']}] {exc}")
    for cod, cambios in updates:
        try:
            kw = {k: v for k, v in cambios.items()}
            if "plazo" in kw and kw["plazo"] is not None:
                kw["plazo"] = int(kw["plazo"])
            q.editar(cod, usuario="fabrica-import", **kw)
            n_up += 1
        except Exception as exc:  # noqa: BLE001
            n_err += 1
            yield line(f"  [ERR update {cod}] {exc}")

    yield line("")
    yield line(f"APLICADO ✓ — {n_ins} creados, {n_up} actualizados, {n_err} errores.")
