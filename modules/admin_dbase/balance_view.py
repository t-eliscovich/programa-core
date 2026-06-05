"""Drift audit: PC vs Banco Pichincha.

Sube el xlsx del extracto y obtiene:
  - Saldo banco (top del xlsx)
  - Saldo PC libros (último mov en transacciones_bancarias)
  - Pendientes históricos
  - Saldo banco esperado = PC + pendientes
  - DRIFT exacto
  - Drift DESAGREGADO por categoría: IVA/COMISION/SENAE/CHEQUES DEV/Transfer/Otros

TMT 2026-05-28 — dueña: "verificar si el banco nos da igual".
"""

from __future__ import annotations

import io
import logging
import re
from collections import defaultdict
from datetime import datetime

import openpyxl
from flask import Blueprint, render_template, request

import db as _db
from auth import requiere_login, requiere_permiso

_LOG = logging.getLogger("programa_core.admin_balance")

bp = Blueprint(
    "admin_balance",
    __name__,
    url_prefix="/admin/balance-banco",
    template_folder="templates",
)


DOCS_CRED_PC = ("DE", "NC", "TR", "XX", "IN", "AC")
DOCS_DEB_PC = ("CH", "ND", "DB", "GS", "PA")


CATS = [
    ("IVA", re.compile(r"\bIVA\b", re.I)),
    ("COMISION", re.compile(r"COMISION", re.I)),
    ("CHEQUE DEVUELTO", re.compile(r"CHEQUE\s+DEVUELTO|COST\s+CHEQUE", re.I)),
    ("PAGO SENAE", re.compile(r"PAGO\s+SENAE", re.I)),
    ("DEPOSITO", re.compile(r"DEPOSITO", re.I)),
    ("TRANSFERENCIA", re.compile(r"TRANSFERENCIA", re.I)),
    ("INTERES", re.compile(r"INTERES", re.I)),
]


def _categorizar(concepto: str) -> str:
    for cat, rx in CATS:
        if rx.search(concepto or ""):
            return cat
    return "OTROS"


def _parse_fecha(s):
    if s is None:
        return None
    if hasattr(s, "date"):
        return s.date()
    if hasattr(s, "year"):
        return s
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_xlsx(content: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    movs = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for ci, h in enumerate(headers, 1):
            row[h] = ws.cell(r, ci).value
        if not row.get("Fecha"):
            continue
        movs.append(row)
    return movs


@bp.route("/", methods=["GET", "POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def balance():
    if request.method == "GET":
        return render_template("admin_dbase/balance.html", resultado=None)

    f = request.files.get("xlsx")
    if not f:
        return render_template("admin_dbase/balance.html",
                               resultado={"error": "Subí el xlsx."})
    no_banco = 10
    try:
        no_banco = int(request.form.get("no_banco") or 10)
    except (TypeError, ValueError):
        pass

    content = f.read()
    if len(content) > 10 * 1024 * 1024:
        return render_template("admin_dbase/balance.html",
                               resultado={"error": "Archivo >10MB."})

    try:
        movs = _parse_xlsx(content)
    except Exception as e:
        _LOG.exception("parse xlsx fail")
        return render_template("admin_dbase/balance.html",
                               resultado={"error": f"Parse: {e}"})

    # Saldo banco top del xlsx (fila con mayor fecha, primer registro)
    movs_con_saldo = [m for m in movs if m.get("Saldo") is not None and m.get("Fecha")]
    if not movs_con_saldo:
        return render_template("admin_dbase/balance.html",
                               resultado={"error": "El xlsx no tiene columna Saldo o Fecha."})
    # Ordenamos por fecha desc, tomamos el primero (running balance más reciente)
    top = sorted(
        movs_con_saldo,
        key=lambda m: (_parse_fecha(m.get("Fecha")) or datetime.min.date(),),
        reverse=True,
    )[0]
    saldo_banco_top = float(top["Saldo"])
    fecha_banco_top = _parse_fecha(top["Fecha"])

    # Saldo PC libros
    row = _db.fetch_one(
        """
        SELECT t.id_transaccion, t.fecha, t.documento, t.importe, t.saldo
        FROM scintela.transacciones_bancarias t
        WHERE t.no_banco=%s ORDER BY t.fecha DESC, t.id_transaccion DESC LIMIT 1
        """,
        (no_banco,),
    )
    saldo_pc = float((row or {}).get("saldo") or 0)
    fecha_pc = (row or {}).get("fecha")

    # Pendientes históricos (banco confirmó, PC no tiene)
    h = _db.fetch_one(
        """
        SELECT COUNT(*) AS n,
          COALESCE(SUM(CASE WHEN COALESCE(tipo,'C')='C' THEN monto ELSE 0 END),0) AS cred,
          COALESCE(SUM(CASE WHEN COALESCE(tipo,'D')='D' THEN monto ELSE 0 END),0) AS deb
        FROM scintela.banco_historicos_pendientes
        WHERE no_banco=%s AND conciliado_en IS NULL
        """,
        (no_banco,),
    ) or {}
    hist_cred = float(h.get("cred") or 0)
    hist_deb = float(h.get("deb") or 0)
    hist_neto = hist_cred - hist_deb

    saldo_esperado = saldo_pc + hist_neto
    drift = round(saldo_banco_top - saldo_esperado, 2)

    # ───────────────────────────────────────────────────────
    # Desglose del drift por "bucket" (pedido dueña 2026-05-28).
    #   Layout:
    #     Sin conciliar (desde última conciliación):
    #       + Ingresos    (DE/NC/TR/XX/IN/AC)
    #       − Egresos     (CH/ND/DB/GS/PA)
    #       = sub-total
    #     + Pendientes históricos PC (mismos campos, fecha vieja)
    #     + Pendientes históricos banco (banco_historicos_pendientes)
    #     = TOTAL drift
    #   Y comparamos con drift_real = saldo_banco − saldo_pc.
    # ───────────────────────────────────────────────────────

    # Última conciliación: max(real_fecha) en banco_conciliacion_match activo.
    ult = _db.fetch_one(
        """
        SELECT MAX(real_fecha) AS f
          FROM scintela.banco_conciliacion_match
         WHERE no_banco=%s AND deshecho_en IS NULL
        """,
        (no_banco,),
    ) or {}
    fecha_ult_conc = ult.get("f")

    # Movs PC sin conciliar, separados en "recientes" (> última conc) vs "viejos" (≤).
    cond_pendiente = (
        " AND TRIM(COALESCE(t.stat,'')) <> '*'"
        " AND NOT EXISTS ("
        "   SELECT 1 FROM scintela.banco_conciliacion_match m"
        "    WHERE m.id_transaccion = t.id_transaccion AND m.deshecho_en IS NULL"
        " )"
    )

    def _sum_pendientes_pc(fecha_op: str, fecha_ref) -> dict:
        """fecha_op = '>' o '<='. fecha_ref puede ser None (=> tomamos todos)."""
        params: list = [no_banco]
        sql = (
            "SELECT "
            "  COUNT(*) AS n, "
            "  COALESCE(SUM(CASE WHEN UPPER(TRIM(t.documento)) IN %s THEN t.importe ELSE 0 END), 0) AS sum_cred, "
            "  COALESCE(SUM(CASE WHEN UPPER(TRIM(t.documento)) IN %s THEN t.importe ELSE 0 END), 0) AS sum_deb "
            "FROM scintela.transacciones_bancarias t "
            "WHERE t.no_banco = %s "
        )
        params = [DOCS_CRED_PC, DOCS_DEB_PC, no_banco]
        if fecha_ref:
            sql += f" AND t.fecha {fecha_op} %s"
            params.append(fecha_ref)
        sql += cond_pendiente
        r = _db.fetch_one(sql, tuple(params)) or {}
        return {
            "n": int(r.get("n") or 0),
            "cred": float(r.get("sum_cred") or 0),
            "deb": float(r.get("sum_deb") or 0),
        }

    if fecha_ult_conc:
        bk_recientes = _sum_pendientes_pc(">", fecha_ult_conc)
        bk_viejos = _sum_pendientes_pc("<=", fecha_ult_conc)
    else:
        # Si NO hay conciliación previa, todo cae como "recientes".
        bk_recientes = _sum_pendientes_pc(">", None)  # cualquier fecha
        bk_viejos = {"n": 0, "cred": 0.0, "deb": 0.0}

    sub_recientes_neto = bk_recientes["cred"] - bk_recientes["deb"]
    sub_viejos_neto = bk_viejos["cred"] - bk_viejos["deb"]

    # Pendientes históricos banco (banco_historicos_pendientes): banco confirmó,
    # PC no tiene. Aportan + al saldo banco (si cred) o − (si deb).
    hist_banco_neto = hist_cred - hist_deb

    # TOTAL drift desde la perspectiva PC:
    # saldo_banco - saldo_pc = + hist_banco_neto - PC_pendientes_neto
    # (PC tiene movs que banco no muestra → bank looks LOWER por PC_neto)
    pc_pend_total_neto = sub_recientes_neto + sub_viejos_neto
    total_drift_desglosado = round(hist_banco_neto - pc_pend_total_neto, 2)

    drift_real = round(saldo_banco_top - saldo_pc, 2)
    no_clasificado = round(drift_real - total_drift_desglosado, 2)

    bucket_resumen = {
        "fecha_ult_conc": fecha_ult_conc,
        "recientes": {
            "n": bk_recientes["n"],
            "ingresos": round(bk_recientes["cred"], 2),
            "egresos": round(bk_recientes["deb"], 2),
            "neto": round(sub_recientes_neto, 2),
        },
        "viejos": {
            "n": bk_viejos["n"],
            "ingresos": round(bk_viejos["cred"], 2),
            "egresos": round(bk_viejos["deb"], 2),
            "neto": round(sub_viejos_neto, 2),
        },
        "hist_banco": {
            "n": int(h.get("n") or 0),
            "cred": round(hist_cred, 2),
            "deb": round(hist_deb, 2),
            "neto": round(hist_banco_neto, 2),
        },
        "drift_desglosado": total_drift_desglosado,
        "drift_real": drift_real,
        "no_clasificado": no_clasificado,
    }

    # Categorización del extracto
    sum_por_cat: dict[str, dict[str, float]] = defaultdict(lambda: {"cred": 0.0, "deb": 0.0, "n": 0})
    for m in movs:
        try:
            monto = float(m.get("Monto") or 0)
        except (TypeError, ValueError):
            continue
        tipo = str(m.get("Tipo") or "C").upper()
        cat = _categorizar(str(m.get("Concepto") or ""))
        sum_por_cat[cat]["n"] += 1
        if tipo == "C":
            sum_por_cat[cat]["cred"] += monto
        else:
            sum_por_cat[cat]["deb"] += monto

    # Para cada fila del xlsx: chequeo si PC tiene el mismo (monto, fecha exacta, tipo).
    # Las que NO tienen contraparte van a la lista "faltan en PC" → drift.
    en_pc_por_cat: dict[str, int] = defaultdict(int)
    falta_en_pc_por_cat_monto: dict[str, float] = defaultdict(float)
    faltantes_banco_en_pc: list[dict] = []  # filas xlsx sin contraparte PC
    for m in movs:
        try:
            monto = float(m.get("Monto") or 0)
        except (TypeError, ValueError):
            continue
        tipo = str(m.get("Tipo") or "C").upper()[:1]
        f_b = _parse_fecha(m.get("Fecha"))
        concepto = str(m.get("Concepto") or "")
        doc_b = str(m.get("Documento") or "").strip()
        oficina = str(m.get("Oficina") or "")
        cat = _categorizar(concepto)
        if not f_b:
            continue
        docs_compat = ("DE", "TR", "XX", "NC", "IN", "AC") if tipo == "C" else ("CH", "ND", "DB", "GS", "PA")
        r = _db.fetch_one(
            """
            SELECT 1 FROM scintela.transacciones_bancarias t
            WHERE t.no_banco=%s AND t.fecha BETWEEN %s AND %s
              AND ABS(t.importe - %s) < 0.005
              AND UPPER(TRIM(t.documento)) IN %s
            LIMIT 1
            """,
            (no_banco, f_b, f_b, monto, docs_compat),
        )
        if r:
            en_pc_por_cat[cat] += 1
        else:
            falta_en_pc_por_cat_monto[cat] += monto
            faltantes_banco_en_pc.append({
                "fecha": f_b,
                "tipo": tipo,
                "documento": doc_b,
                "concepto": concepto[:80],
                "monto": monto,
                "oficina": oficina[:30],
                "categoria": cat,
            })

    # Faltantes_pc_en_banco: movs PC sin conciliar que NO tienen contraparte
    # con monto exacto en el xlsx — son los que PC tiene "de más".
    xlsx_keys = set()
    for m in movs:
        try:
            monto = float(m.get("Monto") or 0)
        except (TypeError, ValueError):
            continue
        f_b = _parse_fecha(m.get("Fecha"))
        if f_b:
            xlsx_keys.add((f_b, round(monto, 2), str(m.get("Tipo") or "C").upper()[:1]))

    pc_sin_conc = _db.fetch_all(
        """
        SELECT t.id_transaccion, t.fecha, t.documento, t.concepto, t.importe, t.prov, t.stat
          FROM scintela.transacciones_bancarias t
         WHERE t.no_banco = %s
           AND TRIM(COALESCE(t.stat,'')) <> '*'
           AND NOT EXISTS (
               SELECT 1 FROM scintela.banco_conciliacion_match m
                WHERE m.id_transaccion = t.id_transaccion AND m.deshecho_en IS NULL
           )
         ORDER BY t.fecha DESC, t.id_transaccion DESC LIMIT 500
        """,
        (no_banco,),
    ) or []
    faltantes_pc_en_banco: list[dict] = []
    for r in pc_sin_conc:
        doc = (r.get("documento") or "").strip().upper()
        tipo_pc = "C" if doc in DOCS_CRED_PC else "D" if doc in DOCS_DEB_PC else "?"
        key = (r.get("fecha"), round(float(r.get("importe") or 0), 2), tipo_pc)
        if key not in xlsx_keys:
            faltantes_pc_en_banco.append({
                "id_transaccion": r.get("id_transaccion"),
                "fecha": r.get("fecha"),
                "documento": doc,
                "concepto": (r.get("concepto") or "")[:80],
                "monto": float(r.get("importe") or 0),
                "tipo": tipo_pc,
                "prov": (r.get("prov") or "").strip(),
            })

    cats_resumen = []
    for cat, vals in sorted(sum_por_cat.items(), key=lambda kv: -abs(kv[1]["cred"] - kv[1]["deb"])):
        n_total = vals["n"]
        n_en_pc = en_pc_por_cat.get(cat, 0)
        n_falta = n_total - n_en_pc
        cats_resumen.append({
            "categoria": cat,
            "n_total": n_total,
            "n_en_pc": n_en_pc,
            "n_falta": n_falta,
            "sum_cred_banco": round(vals["cred"], 2),
            "sum_deb_banco": round(vals["deb"], 2),
            "sum_neto_banco": round(vals["cred"] - vals["deb"], 2),
            "monto_falta_pc": round(falta_en_pc_por_cat_monto.get(cat, 0.0), 2),
        })

    resultado = {
        "saldo_banco_top": round(saldo_banco_top, 2),
        "fecha_banco_top": fecha_banco_top,
        "saldo_pc": round(saldo_pc, 2),
        "fecha_pc": fecha_pc,
        "hist_cred": round(hist_cred, 2),
        "hist_deb": round(hist_deb, 2),
        "hist_neto": round(hist_neto, 2),
        "saldo_esperado": round(saldo_esperado, 2),
        "drift": drift,
        "cuadra": abs(drift) < 1.0,
        "n_movs_xlsx": len(movs),
        "categorias": cats_resumen,
        "no_banco": no_banco,
        "bucket": bucket_resumen,
        "faltantes_banco_en_pc": faltantes_banco_en_pc,
        "faltantes_pc_en_banco": faltantes_pc_en_banco,
        "n_falt_b_en_pc": len(faltantes_banco_en_pc),
        "n_falt_pc_en_b": len(faltantes_pc_en_banco),
        "sum_falt_b_en_pc_signed": round(sum(
            (m["monto"] if m["tipo"] == "C" else -m["monto"]) for m in faltantes_banco_en_pc
        ), 2),
        "sum_falt_pc_en_b_signed": round(sum(
            (m["monto"] if m["tipo"] == "C" else -m["monto"]) for m in faltantes_pc_en_banco
        ), 2),
    }
    return render_template("admin_dbase/balance.html", resultado=resultado)
