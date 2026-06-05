"""Auto-match xlsx → scintela.transacciones_bancarias en 1 click.

Subí el extracto de Pichincha (.xlsx), el backend lo parsea, busca para cada
fila un mov PC con monto EXACTO + fecha ±3d + tipo compatible + no conciliado,
y crea banco_conciliacion_match para los que matchean único. Devuelve un
reporte línea por línea.

TMT 2026-05-28 dueña: "conecta uno con uno me da igual" — bypass SSM/CloudShell.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timedelta

import openpyxl
from flask import Blueprint, Response, render_template, request, stream_with_context

import db as _db
from auth import requiere_login, requiere_permiso

_LOG = logging.getLogger("programa_core.admin_match")

bp = Blueprint(
    "admin_match",
    __name__,
    url_prefix="/admin/conciliar-auto",
    template_folder="templates",
)

DOCS_CRED = ("DE", "TR", "XX", "NC", "IN", "AC")
DOCS_DEB = ("CH", "ND", "DB", "GS", "PA")
TOL_DIAS = 3


@bp.route("/", methods=["GET"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def form():
    return render_template("admin_dbase/auto_match.html")


def _parse_fecha(s):
    if s is None:
        return None
    if hasattr(s, "date"):
        return s.date()
    if hasattr(s, "year"):
        return s
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_xlsx(content: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    movs = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for ci, h in enumerate(headers, 1):
            row[h] = ws.cell(r, ci).value
        if not row.get("Fecha"):
            continue
        movs.append(row)
    return movs


@bp.route("/run", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def run():
    f = request.files.get("xlsx")
    if not f:
        return Response("ERROR: subí el xlsx en el campo 'xlsx'.\n", mimetype="text/plain", status=400)

    no_banco = 10
    try:
        no_banco = int(request.form.get("no_banco") or 10)
    except (TypeError, ValueError):
        pass
    dry_run = request.form.get("dry_run") in ("1", "on", "true")

    content = f.read()
    if len(content) > 10 * 1024 * 1024:
        return Response("ERROR: archivo muy grande (>10MB).\n", mimetype="text/plain", status=400)

    try:
        movs = _parse_xlsx(content)
    except Exception as e:
        _LOG.exception("xlsx parse falló")
        return Response(f"ERROR parseando xlsx: {e}\n", mimetype="text/plain", status=400)

    usuario = "auto-match-web"
    try:
        from flask import g
        usuario = (g.user or {}).get("username", usuario) if g.user else usuario
    except Exception:
        pass

    def generate():
        yield f"Extracto: {len(movs)} movs · banco={no_banco}\n"
        if dry_run:
            yield "[DRY-RUN] no se inserta nada.\n"
        yield "\n"

        n_matched = 0
        n_no_match = 0
        n_multi = 0
        n_dup = 0
        log_no, log_multi = [], []

        for m in movs:
            f_b = _parse_fecha(m.get("Fecha"))
            if not f_b:
                continue
            try:
                monto = float(m.get("Monto") or 0)
            except (TypeError, ValueError):
                continue
            tipo = str(m.get("Tipo") or "C").upper()[:1]
            doc_b = str(m.get("Documento") or "").strip()[:40]
            concepto = str(m.get("Concepto") or "")[:500]
            oficina = str(m.get("Oficina") or "")[:50]
            codigo = str(m.get("Codigo") or "")[:10]

            # Skip ya conciliado
            already = _db.fetch_one(
                "SELECT 1 FROM scintela.banco_conciliacion_match "
                "WHERE no_banco=%s AND real_fecha=%s AND real_documento=%s "
                "AND real_monto=%s AND real_tipo=%s AND deshecho_en IS NULL LIMIT 1",
                (no_banco, f_b, doc_b, monto, tipo),
            )
            if already:
                n_dup += 1
                continue

            docs_compat = DOCS_CRED if tipo == "C" else DOCS_DEB
            d1 = f_b - timedelta(days=TOL_DIAS)
            d2 = f_b + timedelta(days=TOL_DIAS)

            cands = _db.fetch_all(
                """
                SELECT t.id_transaccion, t.fecha
                  FROM scintela.transacciones_bancarias t
                 WHERE t.no_banco = %s
                   AND t.fecha BETWEEN %s AND %s
                   AND ABS(t.importe - %s) < 0.005
                   AND UPPER(TRIM(t.documento)) IN %s
                   AND TRIM(COALESCE(t.stat,'')) <> '*'
                   AND NOT EXISTS (
                       SELECT 1 FROM scintela.banco_conciliacion_match mc
                        WHERE mc.id_transaccion = t.id_transaccion AND mc.deshecho_en IS NULL
                   )
                 LIMIT 5
                """,
                (no_banco, d1, d2, monto, docs_compat),
            ) or []

            if not cands:
                n_no_match += 1
                log_no.append((str(f_b), tipo, monto, doc_b, concepto[:40]))
                continue
            if len(cands) > 1:
                same_day = [c for c in cands if c["fecha"] == f_b]
                if len(same_day) == 1:
                    cands = same_day
                else:
                    n_multi += 1
                    log_multi.append((str(f_b), tipo, monto, [c["id_transaccion"] for c in cands]))
                    continue

            pc = cands[0]
            if dry_run:
                n_matched += 1
                continue
            try:
                row = _db.fetch_one(
                    """
                    INSERT INTO scintela.banco_conciliacion_match
                        (no_banco, estado, real_fecha, real_concepto, real_documento,
                         real_monto, real_tipo, real_codigo, real_oficina,
                         id_transaccion, usuario)
                    VALUES (%s, 'matched', %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT DO NOTHING
                    RETURNING id
                    """,
                    (no_banco, f_b, concepto, doc_b, monto, tipo,
                     codigo, oficina, pc["id_transaccion"], usuario),
                )
                if row:
                    n_matched += 1
                else:
                    n_dup += 1
            except Exception as e:
                _LOG.exception("insert match fail")
                yield f"  [ERR insert] {f_b} ${monto:,.2f}: {e}\n"
                continue

        yield "\n=== RESUMEN ===\n"
        yield f"  conciliados:      {n_matched}\n"
        yield f"  ya conciliados:   {n_dup}\n"
        yield f"  multi-match skip: {n_multi}\n"
        yield f"  sin match:        {n_no_match}\n"
        yield f"  TOTAL extracto:   {len(movs)}\n"
        if log_no:
            yield f"\n=== SIN MATCH (top {min(30, len(log_no))} de {len(log_no)}) ===\n"
            for f_b, t, mt, doc, c in log_no[:30]:
                yield f"  {f_b} {t} ${mt:>10,.2f} doc={doc:<10s} {c}\n"
        if log_multi:
            yield "\n=== MULTI-MATCH (top 15) ===\n"
            for f_b, t, mt, ids in log_multi[:15]:
                yield f"  {f_b} {t} ${mt:>10,.2f} candidatos PC: {ids}\n"

    return Response(stream_with_context(generate()), mimetype="text/plain")
