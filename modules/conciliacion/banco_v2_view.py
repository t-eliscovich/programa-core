"""Endpoints v2 de conciliación bancaria — Reforma Sprint 1 (2026-05-28).

Pantalla post-procesar con 3 tabs (Manual, Impuestos, Transferencias),
balance Pichincha compact arriba sticky, sesión persistente y botón
Terminar y guardar abajo que genera el PDF de pendientes.

Coexiste con el viejo /conciliacion/hub mientras se valida. Una vez que
la dueña confirma, el alias /conciliacion/banco se reapunta a este flujo
y el viejo queda para borrar en sprint 2.
"""
from __future__ import annotations

import logging
from datetime import UTC, date, datetime, timedelta, timezone


def _hora_ec_str(value, fmt: str = "%Y-%m-%d %H:%M") -> str:
    """Convertir datetime UTC → str en hora Ecuador (UTC-5)."""
    if value is None:
        return ""
    if not isinstance(value, datetime):
        return str(value)
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(timezone(timedelta(hours=-5))).strftime(fmt)
from pathlib import Path

from flask import (
    abort,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)

import db as _db
from auth import requiere_login, requiere_permiso
from modules.conciliacion import balance_pichincha as _bp
from modules.conciliacion import sesion as _sesion
from modules.conciliacion.matcher_banco import (
    confirmar_match,
    crear_transaccion_agrupada_desde_reals,
)
from modules.conciliacion.parser_banco import parse_banco_xlsx
from modules.conciliacion.views import (
    _BANCO_PICHINCHA,
    _usuario_actual,
    conciliacion_bp,
)

_LOG = logging.getLogger("programa_core.conciliacion.banco_v2")


# ── Asignación multi-PC (N banco → M programa) ────────────────────────
# TMT 2026-07-14 (dueña): "debería funcionar en cualquier caso". Cuando se
# seleccionan N movs de banco y M de programa, hay que auto-asignar cada
# banco (o grupo de banco) al mov de programa cuyo MONTO firmado le cuadra,
# soportando 1:1, N:1 (depósito partido) y multi-PC a la vez.
#
# REGLA DE ORO (no negociable): NUNCA inventar diferencias. Solo se ata un
# banco (o grupo) a un PC si la suma firmada cuadra con ese PC dentro de
# |Δ|≤tol. Lo que no logra asignarse limpio queda PENDIENTE (se preserva la
# protección 2026-06-18 contra atar al primer PC e inventar diferencias).
#
# Función PURA y testeable (los tests viven en tests/test_asignar_banco.py).
def _subset_para_target(avail, target, tol, cap):
    """Devuelve la lista de índices (en `avail`) cuyo signed sum cuadra con
    `target` dentro de `tol`, o None si no hay subconjunto que cuadre.

    `avail` = list[(banco_id, signed)]. Prefiere el subconjunto MÁS CHICO y,
    a igual tamaño, el de menor |Δ|. Para N chico enumera combinaciones
    exactas; si len(avail) > cap cae a un greedy en la dirección del target.
    Tamaño 1 (par 1:1) lo resuelve el llamador antes de invocar esto, así
    que acá se arranca en tamaño 2.
    """
    from itertools import combinations

    n = len(avail)
    if n < 2:
        return None
    if n > cap:
        # Greedy: acumular de mayor a menor SOLO los que empujan hacia el
        # target (mismo signo). Si nos pasamos sin cuadrar → descartar el PC.
        pos = target >= 0
        order = sorted(
            (i for i in range(n) if (avail[i][1] >= 0) == pos and avail[i][1] != 0),
            key=lambda i: abs(avail[i][1]),
            reverse=True,
        )
        acc = 0.0
        picked = []
        for i in order:
            picked.append(i)
            acc += avail[i][1]
            if abs(acc - target) <= tol:
                return picked
            if (pos and acc > target + tol) or (not pos and acc < target - tol):
                return None
        return None
    # Enumeración exacta por tamaño creciente: primer tamaño con match gana
    # (subconjunto más chico); a igual tamaño, el de menor |Δ|.
    vals = [avail[i][1] for i in range(n)]
    for size in range(2, n + 1):
        best = None
        best_d = None
        for combo in combinations(range(n), size):
            s = sum(vals[i] for i in combo)
            d = abs(s - target)
            if d <= tol and (best_d is None or d < best_d):
                best_d = d
                best = combo
        if best is not None:
            return list(best)
    return None


def asignar_banco_a_programa(banco_firmados, prog_firmados, tol=0.50, cap=18):
    """Asigna movimientos de banco a movimientos de programa por MONTO firmado.

    Args:
      banco_firmados: list[(banco_id, signed)]  — signed = +monto si crédito,
                      −monto si débito. `banco_id` es una clave opaca (índice).
      prog_firmados:  list[(pc_id, signed)]     — signed = delta firmado del PC.
      tol: tolerancia |Δ| para considerar que "cuadra" (default 0.50).
      cap: si un PC enfrenta más de `cap` banco disponibles, se usa greedy.

    Returns:
      (asignaciones, sobrantes) donde
        asignaciones: dict[pc_id -> list[banco_id]]  (grupo por PC; N:1 = lista)
        sobrantes:    list[banco_id]  — banco que NO se asignó limpio → PENDIENTE.

    Garantías:
      - Un banco no se reutiliza en dos PC.
      - Se prefiere el match EXACTO 1:1 (un banco suelto) antes de armar grupos.
      - Un PC se atiende SOLO si algún banco/grupo cuadra con él dentro de tol;
        si no, el PC no recibe nada (y sus banco potenciales quedan sobrantes).
      - PC grandes (|monto| desc) se sirven primero para que se queden con su
        match obvio antes que uno chico se "coma" un banco ambiguo.
    """
    asign: dict = {}
    avail = list(banco_firmados)  # list[(banco_id, signed)]
    pcs = sorted(prog_firmados, key=lambda t: abs(t[1]), reverse=True)

    for pc_id, target in pcs:
        # a) Par 1:1 exacto: un banco suelto que cuadre. El más cercano gana.
        best_i = None
        best_d = None
        for i, (_bid, sval) in enumerate(avail):
            d = abs(sval - target)
            if d <= tol and (best_d is None or d < best_d):
                best_d = d
                best_i = i
        if best_i is not None:
            bid, _sv = avail.pop(best_i)
            asign.setdefault(pc_id, []).append(bid)
            continue
        # b) Grupo N:1: subconjunto de banco cuya suma firmada cuadre.
        chosen = _subset_para_target(avail, target, tol, cap)
        if chosen:
            for idx in sorted(chosen, reverse=True):
                bid, _sv = avail.pop(idx)
                asign.setdefault(pc_id, []).append(bid)
        # c) Si no cuadró nada, el PC queda sin banco (no se inventa nada).

    sobrantes = [bid for bid, _sv in avail]
    return asign, sobrantes


def reconciliacion_completa(asign, sobrantes, prog_ids):
    """Decisión ATÓMICA todo-o-nada (dueña 2026-07-14).

    Una selección se confirma SOLO si se reconcilia COMPLETA:
      - no quedan banco sobrantes (`sobrantes` vacío), Y
      - ningún mov de programa quedó sin contraparte de banco (todo pc_id de
        `prog_ids` está en `asign`).
    Si algo no cierra → incompleta → no se confirma NADA (queda todo pendiente).

    Args:
      asign:     dict[pc_id -> [banco_ids]]  (salida de asignar_banco_a_programa)
      sobrantes: list[banco_id]               (idem)
      prog_ids:  iterable de los pc_id que la dueña seleccionó.

    Returns:
      (completa: bool, pcs_sin_banco: list[pc_id])
    """
    asignados = set(asign.keys())
    pcs_sin_banco = [p for p in prog_ids if p not in asignados]
    completa = (not sobrantes) and (not pcs_sin_banco)
    return completa, pcs_sin_banco


def forzar_asignacion_completa(asign, sobrantes_idx, banco_firmados, prog_firmados):
    """Cierra a la fuerza una asignación que quedó incompleta por diferencia
    de montos (dueña 2026-07-17: "cuando hay diferencia que re pregunte pero
    que deje igual"). Se usa SOLO cuando la dueña ya vio la advertencia de
    diferencia en la vista previa y confirmó igual (aceptar_diferencia=1).

    Empareja por cercanía de monto lo que quedó suelto:
      1) cada PC sin banco toma el banco sobrante de monto más cercano;
      2) los banco sobrantes restantes van al grupo del PC más cercano;
      3) TMT 2026-07-20 (dueña: "deja seleccionar cualquier N de cada lado",
         caso 2 banco ↔ 4 programa): si el banco se agota y quedan PCs
         sueltos, esos PCs van a `pcs_internos` — se concilian como INTERNOS
         (match sin contraparte de banco, mismo mecanismo que
         /banco-v2/descartar-programa), reversibles igual.
    La diferencia queda como gap asumido (visible en /banco-v2/auditar).

    Función PURA. Devuelve (asign, sobrantes, pcs_internos). `sobrantes` solo
    queda no-vacío si no hay ningún PC al que atarlos (selección sin lado
    programa).
    """
    asign = {p: list(b) for p, b in asign.items()}
    banco_val = dict(banco_firmados)
    prog_val = dict(prog_firmados)
    rem = list(sobrantes_idx)
    pcs_internos: list = []
    pcs_sueltos = [p for p, _t in prog_firmados if p not in asign]
    for pc_id in pcs_sueltos:
        if not rem:
            pcs_internos.append(pc_id)
            continue
        target = prog_val.get(pc_id, 0.0)
        best = min(rem, key=lambda i: abs(banco_val.get(i, 0.0) - target))
        rem.remove(best)
        asign[pc_id] = [best]
    for i in list(rem):
        if not asign:
            break
        sval = banco_val.get(i, 0.0)
        pc_best = min(asign.keys(), key=lambda p: abs(prog_val.get(p, 0.0) - sval))
        asign[pc_best].append(i)
        rem.remove(i)
    return asign, rem, pcs_internos


def _conciliar_pc_interno(pc_id: int, batch_id: str, usuario: str,
                          metodo: str = "interno:aceptar_diferencia",
                          conn=None) -> None:
    """Concilia UN mov de programa como INTERNO (sin contraparte de banco):
    match estado='matched' sin real_* + stat='*'. Mismo mecanismo que
    /banco-v2/descartar-programa → auditable y reversible desde deshacer.
    TMT 2026-07-20 (dueña: "deja seleccionar cualquier N de cada lado")."""
    try:
        _db.execute(
            """
            INSERT INTO scintela.banco_conciliacion_match (
                no_banco, estado, metodo,
                id_transaccion, tx_firma, confirm_batch_id, usuario
            ) VALUES (%s, 'matched', %s, %s,
                      scintela.compute_tx_firma(%s), %s, %s)
            """,
            (_BANCO_PICHINCHA, metodo[:40], int(pc_id), int(pc_id),
             batch_id, usuario),
            conn=conn,
        )
    except Exception:
        # Schema viejo sin columna `metodo` (pre-mig 0047).
        _db.execute(
            """
            INSERT INTO scintela.banco_conciliacion_match (
                no_banco, estado, id_transaccion, tx_firma,
                confirm_batch_id, usuario
            ) VALUES (%s, 'matched', %s, scintela.compute_tx_firma(%s), %s, %s)
            """,
            (_BANCO_PICHINCHA, int(pc_id), int(pc_id), batch_id, usuario),
            conn=conn,
        )
    _db.execute(
        """
        UPDATE scintela.transacciones_bancarias
           SET stat = '*'
         WHERE id_transaccion = %s AND no_banco = %s
        """,
        (int(pc_id), _BANCO_PICHINCHA),
        conn=conn,
    )


def _proponer_movimiento_diferencia(
    banco_firmados,
    prog_firmados,
    sobrantes_idx,
    pcs_sin_banco,
    umbral=0.50,
    umbral_confirmar=300.0,
):
    """Propone crear el movimiento de programa que falta para cerrar una
    selección que NO reconcilia completa PERO cuyo único faltante es un
    residuo del lado BANCO (TMT 2026-07-14, dueña: "Agregar movimiento en
    el programa").

    Aplica SOLO si:
      - `pcs_sin_banco` está vacío (todo mov de programa seleccionado matcheó), Y
      - `sobrantes_idx` es NO vacío (quedó banco sin contraparte), Y
      - |dif| > umbral (por debajo ya auto-matchea la tolerancia; nunca se
        inventa una diferencia que no existe).

    `dif = round(Σ signed banco − Σ signed programa, 2)`.
      dif > 0 → banco tiene CRÉDITO de más → crear crédito en programa → 'NC'.
      dif < 0 → banco tiene DÉBITO de más  → crear débito  en programa → 'ND'.
    `importe = abs(dif)`. Si |dif| > umbral_confirmar → requiere_confirmar.

    Función PURA (testeable). Devuelve None cuando NO aplica, o el dict con la
    propuesta.
    """
    if pcs_sin_banco:
        return None
    if not sobrantes_idx:
        return None
    dif = round(
        sum(s for _bid, s in banco_firmados) - sum(s for _pid, s in prog_firmados),
        2,
    )
    if abs(dif) <= umbral:
        return None
    if dif > 0:
        documento = "NC"
        sentido = "crédito"
    else:
        documento = "ND"
        sentido = "débito"
    return {
        "diferencia": dif,
        "importe": round(abs(dif), 2),
        "documento": documento,
        "sentido": sentido,
        "requiere_confirmar": abs(dif) > umbral_confirmar,
        "umbral_confirmar": umbral_confirmar,
    }


# ── Defensas contra corrupción de saldo running ──────────────────────
# TMT 2026-05-29: el "BUG #2" reportado durante el E2E (saldo bajó $43K
# al anular tx de $29) en realidad reveló una cadena previamente corrupta.
# Ahora validamos pre/post cada mutación destructiva. Si detectamos
# descalce, log CRITICAL + opcionalmente recompute desde el inicio.

_SIGNOS_C = ("DE", "TR", "AC", "NC", "IN", "XX")
_SIGNOS_D = ("CH", "ND", "DB", "GS", "PA")


def _signed_delta(documento: str, importe: float, usuario_crea: str = "") -> float:
    """Wrapper que delega en bank_helpers._signed_delta (single source of truth).

    TMT 2026-06-03 audit fix: la versión local antigua ignoraba el signo del
    importe y siempre forzaba sign-by-doc — eso destrozaba NDs reverso (importe
    positivo legítimo del DBF). Ahora respeta convención según usuario_crea:
    legacy DBF importe-signed vs web importe-abs.
    """
    import bank_helpers as _bh_local
    return _bh_local._signed_delta(documento, importe, usuario_crea)


def _verificar_cadena_saldos(no_banco: int, limit: int = 30, conn=None) -> dict:
    """Recorre las últimas N filas y valida saldo = saldo_prev + signed_delta.

    Returns:
        {ok: bool, ultimo_saldo: float, problemas: [...], n_chequeadas: int}
    """
    try:
        rows = _db.fetch_all(
            """
            SELECT id_transaccion, fecha, documento, importe, saldo,
                   COALESCE(usuario_crea, '') AS usuario_crea
              FROM scintela.transacciones_bancarias
             WHERE no_banco = %s AND saldo IS NOT NULL
             ORDER BY fecha DESC, id_transaccion DESC LIMIT %s
            """,
            (int(no_banco), int(limit)),
            conn=conn,
        ) or []
    except Exception as e:
        return {"ok": False, "ultimo_saldo": None, "problemas": [{"error": str(e)}], "n_chequeadas": 0}

    rows = list(reversed(rows))  # ASC
    if not rows:
        return {"ok": True, "ultimo_saldo": 0.0, "problemas": [], "n_chequeadas": 0}

    problemas = []
    saldo_prev = None
    for r in rows:
        s = float(r["saldo"] or 0)
        delta = _signed_delta(
            r.get("documento"),
            float(r.get("importe") or 0),
            r.get("usuario_crea") or "",
        )
        if saldo_prev is not None:
            esperado = round(saldo_prev + delta, 2)
            diff = round(s - esperado, 2)
            if abs(diff) > 0.01:
                problemas.append({
                    "id": r["id_transaccion"], "diff": diff,
                    "esperado": esperado, "grabado": s,
                })
        saldo_prev = s

    return {
        "ok": not problemas,
        "ultimo_saldo": saldo_prev,
        "problemas": problemas[:5],
        "n_chequeadas": len(rows),
    }


def _validar_y_loguear(no_banco: int, contexto: str, delta_esperado: float | None = None) -> bool:
    """Verifica cadena + loguea CRITICAL si está corrupta. Devuelve True si OK."""
    r = _verificar_cadena_saldos(no_banco)
    if not r["ok"]:
        _LOG.critical(
            "CADENA SALDOS CORRUPTA tras %s — banco=%s, problemas=%s",
            contexto, no_banco, r["problemas"],
        )
        return False
    if delta_esperado is not None:
        # Nada que validar acá sin saldo_anterior — se hace en el caller.
        pass
    return True

# Directorio para reportes de sesiones cerradas. data/ está en el repo,
# en el server vive bajo C:\programa-core\data\.
_PDF_DIR = Path("data") / "conciliacion_pdfs"  # legacy nombre, ahora xlsx


def _migracion_lista_o_redirect():
    """Si la tabla banco_conciliacion_sesion no existe, mostrar flash claro
    y redirect en lugar de un 500 críptico.
    """
    if _sesion.tabla_existe():
        return None
    flash(
        "El flujo v2 necesita que se corra la migración 0060 en la DB. "
        "Avisame y la corro en CloudShell.",
        "warn",
    )
    return redirect(url_for("conciliacion.hub"))


# ─── Pantalla principal post-procesar ─────────────────────────────────


@conciliacion_bp.route("/banco-v2", methods=["GET"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_post_procesar():
    """Renderiza la pantalla con los 4 tabs.

    TMT 2026-06-02 dueña: sesión nunca se cierra → siempre operamos contra
    la única sesión abierta del banco. Si no existe ninguna, redirect al
    landing para subir el primer extracto (o usar "Conciliar pendientes").
    """
    r = _migracion_lista_o_redirect()
    if r: return r
    no_banco = _BANCO_PICHINCHA

    sesion = _sesion.sesion_abierta(no_banco)
    if not sesion:
        flash("Subí un extracto para empezar la conciliación.", "info")
        return redirect(url_for("conciliacion.hub"))

    # TMT 2026-06-03: el contador sesion.matches_hechos venía desincronizado
    # con la realidad (decía 14 con 0 matches reales). Lo recomputamos live
    # desde banco_conciliacion_match. Single source of truth.
    try:
        _real_matches = _db.fetch_one(
            """
            SELECT COUNT(*) AS n FROM scintela.banco_conciliacion_match
             WHERE no_banco = %s AND deshecho_en IS NULL
            """,
            (no_banco,),
        ) or {}
        sesion = dict(sesion)
        sesion["matches_hechos"] = int(_real_matches.get("n") or 0)
    except Exception:
        pass

    tab = (request.args.get("tab") or "manual").lower()
    if tab not in ("manual", "impuestos", "transferencias", "conciliados"):
        tab = "manual"

    buckets = _sesion.estado_sesion(sesion, no_banco)
    balance = _bp.calcular(no_banco)

    # TMT 2026-06-21 (dueña: "¿por qué dice 131k si la diferencia es 0?").
    # El "Saldo banco esperado" de la pantalla contaba como Pendientes banco
    # SOLO la hoja (decisión 2026-06-04) e IGNORABA el extracto de la sesión
    # sin cruzar — pero el resumen descargado SÍ lo cuenta (fix 2026-06-16/17).
    # Resultado: el resumen cerraba en 0 y la pantalla mostraba justo el neto
    # del extracto sin cruzar como "diferencia". Acá sumamos ese neto (mismos
    # buckets que _generar_xlsx_pendientes) a los totales `*_total` que el
    # template ya sabe mostrar, para que pantalla == resumen y la diferencia
    # cierre. estado_sesion ya deduplica el extracto contra la hoja.
    _xt_cred = _xt_deb = 0.0
    _xt_n = 0
    for _bucket in ("manual_banco", "impuestos"):
        for _it in (buckets.get(_bucket) or []):
            if _it.get("es_historico") or _it.get("mov") is None:
                continue
            _mv = _it["mov"]
            _tp = (getattr(_mv, "tipo", "C") or "C").upper()
            _mo = float(getattr(_mv, "monto", 0) or 0)
            if _tp == "C":
                _xt_cred += _mo
            else:
                _xt_deb += _mo
            _xt_n += 1
    if _xt_n:
        _xt_cred = round(_xt_cred, 2)
        _xt_deb = round(_xt_deb, 2)
        balance["pendientes_banco_total_creditos"] = round(
            float(balance.get("pendientes_banco_creditos") or 0) + _xt_cred, 2)
        balance["pendientes_banco_total_debitos"] = round(
            float(balance.get("pendientes_banco_debitos") or 0) + _xt_deb, 2)
        balance["neto_pendientes_total"] = round(
            float(balance.get("neto_pendientes") or 0) + _xt_cred - _xt_deb, 2)
        balance["n_pendientes_banco_total"] = int(balance.get("n_pendientes") or 0) + _xt_n
        balance["saldo_banco_esperado"] = round(
            float(balance.get("saldo_si_concilio_todo") or balance.get("saldo") or 0)
            + balance["neto_pendientes_total"], 2)
    # TMT 2026-06-02 dueña: 'si en el extracto tenemos 5k de diferencia'.
    # TMT 2026-06-03: balance_pichincha.calcular() ya incluye el extracto de
    # sesión en neto_pendientes_total. No re-enriquecemos acá — un solo lugar
    # para la math. La dueña pidió "totales que cierren de ambos lados".

    # TMT 2026-06-02 dueña: 'lo deberia implementar el usuario no? porque
    # por el excel no sabemos cual es el ultimo valor'. Prioridad:
    #   1. sesion.saldo_banco_objetivo (manual, escrito por la dueña).
    #   2. sesion.saldo_banco_detectado (saldo real del extracto, capturado al
    #      subir, robusto al dedup). TMT 2026-06-26.
    #   3. Auto-detect = max(fecha).saldo del payload (fallback frágil).
    saldo_manual = sesion.get("saldo_banco_objetivo")
    saldo_detectado = sesion.get("saldo_banco_detectado")
    if saldo_manual is not None:
        try:
            balance["saldo_banco_real"] = float(saldo_manual)
            balance["saldo_banco_real_origen"] = "manual"
        except (TypeError, ValueError):
            balance["saldo_banco_real"] = None
    elif saldo_detectado is not None:
        try:
            balance["saldo_banco_real"] = float(saldo_detectado)
            balance["saldo_banco_real_origen"] = "detectado"
        except (TypeError, ValueError):
            balance["saldo_banco_real"] = None
    else:
        try:
            movs_s = _sesion.cargar_movs(sesion)
            con_fecha = [
                m for m in movs_s
                if getattr(m, "fecha", None) and getattr(m, "saldo", None) is not None
            ]
            if con_fecha:
                ult = max(con_fecha, key=lambda m: m.fecha)
                balance["saldo_banco_real"] = float(ult.saldo)
            elif movs_s:
                v = float(getattr(movs_s[-1], "saldo", None) or 0)
                balance["saldo_banco_real"] = v if v else None
            balance["saldo_banco_real_origen"] = "auto"
        except Exception:
            pass

    # Diferencia esperado vs real (si tenemos ambos).
    if balance.get("saldo_banco_real") is not None and balance.get("saldo_banco_esperado") is not None:
        balance["diferencia_no_clasificada"] = round(
            balance["saldo_banco_real"] - balance["saldo_banco_esperado"], 2
        )

    # TMT 2026-05-29 dueña: 'Hacer un cuarto tab que muestre conciliaciones
    # hasta ahora'. Lista los matches confirmados en esta sesión.
    matches_sesion = _sesion.matches_de_sesion(sesion)

    # TMT 2026-06-02 dueña: 'en esta tabla de conciliados, me tiene que
    # aparecer monto banco, monto programa, y tengo que ver que los totales
    # coincidan'. Enriquezco cada match con monto_prog_signed (el lado PC
    # con signo derivado del documento) y totalizo por lado.
    for m in matches_sesion:
        tipo = (m.get("real_tipo") or "").upper()
        # ¿Tiene lado banco? Sí si vino con real_monto (banco_only y matched).
        # Bancsis_only_ok no tiene lado banco.
        m["has_banco"] = m.get("real_monto") is not None
        try:
            monto_banco = float(m.get("real_monto") or 0)
        except (TypeError, ValueError):
            monto_banco = 0.0
        m["monto_banco_signed"] = monto_banco if tipo == "C" else (-monto_banco if tipo == "D" else 0.0)

        # ¿Tiene lado programa? Sí si el JOIN con transacciones_bancarias
        # devolvió una fila (tb.importe IS NOT NULL). Históricos pueden no
        # tenerlo si la dueña los marcó conciliados sin linkear a una tx PC.
        m["has_programa"] = m.get("tb_importe") is not None
        tb_doc = (m.get("tb_documento") or "").upper()
        try:
            tb_imp = float(m.get("tb_importe") or 0)
        except (TypeError, ValueError):
            tb_imp = 0.0
        # TMT 2026-06-03 audit fix: usar _signed_delta con tb_usuario_crea
        # para que NDs reverso del DBF (importe positivo legítimo) tengan
        # signo correcto. La versión local antigua siempre forzaba sign-by-doc.
        m["monto_prog_signed"] = _signed_delta(tb_doc, tb_imp, m.get("tb_usuario_crea") or "")
        # Diferencia por fila (banco − programa) — solo si AMBOS lados existen.
        if m["has_banco"] and m["has_programa"]:
            m["diff"] = round(m["monto_banco_signed"] - m["monto_prog_signed"], 2)
        else:
            m["diff"] = None  # un lado falta → no aplica diferencia

        # TMT 2026-07-15 (dueña: "volvé a poner cuanto venía de programa"). SOLO
        # DISPLAY: si el vínculo al programa se rompió (el relink viejo lo nuleó)
        # pero el match 'matched' conserva su tx_firma, exponemos el importe del
        # programa que quedó GUARDADO en la firma (fecha|documento|importe|...).
        # NO toca has_programa/monto_prog_signed/diff ni los totales ni el
        # matcher ni el cálculo de pendientes — es puro adorno del historial.
        m["tiene_prog_firma"] = False
        m["prog_firma_signed"] = 0.0
        if (not m["has_programa"]) and (m.get("estado") == "matched"):
            _fp = (m.get("tx_firma") or "").split("|")
            if len(_fp) >= 3:
                try:
                    m["prog_firma_signed"] = _signed_delta(
                        (_fp[1] or "").upper(), float(_fp[2]), "")
                    m["tiene_prog_firma"] = True
                except (TypeError, ValueError):
                    pass

    # Totales agregados por lado.
    # TMT 2026-06-03 dueña: 'mira los totales'. Bug: el agrupado de impuestos
    # crea 12 matches que apuntan a la MISMA tx PC (-$69.01). Sumar
    # monto_prog_signed por cada match contaba la tx 12 veces.
    # Fix: dedupar el lado PC por id_transaccion antes de sumar.
    tot_banco_c = sum(float(m.get("real_monto") or 0) for m in matches_sesion if (m.get("real_tipo") or "").upper() == "C")
    tot_banco_d = sum(float(m.get("real_monto") or 0) for m in matches_sesion if (m.get("real_tipo") or "").upper() == "D")
    _seen_tx = set()
    _seen_firma = set()
    _prog_dedup = []
    for m in matches_sesion:
        tx_id = m.get("id_transaccion")
        if tx_id is not None:
            if tx_id in _seen_tx:
                continue
            _seen_tx.add(tx_id)
            _prog_dedup.append(m["monto_prog_signed"])
        elif m.get("tiene_prog_firma"):
            # TMT 2026-07-15 (dueña: el TOTAL PROGRAMA quedaba raro / DIFERENCIA
            # inflada). Los cruces cuyo vínculo al programa se rompió (relink
            # viejo) no sumaban al total → la diferencia mostraba el hueco como
            # si fuera descuadre. Usamos el valor GUARDADO en la firma (dedup por
            # tx_firma). Solo display del header — no toca datos ni pendientes.
            _f = m.get("tx_firma") or ""
            if _f and _f not in _seen_firma:
                _seen_firma.add(_f)
                _prog_dedup.append(m.get("prog_firma_signed") or 0.0)
    tot_prog_c = sum(v for v in _prog_dedup if v > 0)
    tot_prog_d = sum(-v for v in _prog_dedup if v < 0)
    conciliados_totales = {
        "banco_creditos": round(tot_banco_c, 2),
        "banco_debitos": round(tot_banco_d, 2),
        "banco_neto": round(tot_banco_c - tot_banco_d, 2),
        "prog_creditos": round(tot_prog_c, 2),
        "prog_debitos": round(tot_prog_d, 2),
        "prog_neto": round(tot_prog_c - tot_prog_d, 2),
        "diff_neto": round((tot_banco_c - tot_banco_d) - (tot_prog_c - tot_prog_d), 2),
    }

    return render_template(
        "conciliacion/banco_v2.html",
        sesion=sesion,
        tab_activo=tab,
        buckets=buckets,
        balance=balance,
        saldo_pc_actual=balance,        # alias por si algún include lo busca
        banco_nombre="Pichincha",
        modo="compact",
        matches_sesion=matches_sesion,
        conciliados_totales=conciliados_totales,
    )


# ─── Endpoint: crear sesión a partir del upload ───────────────────────


@conciliacion_bp.route("/banco-v2/crear-sesion", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_crear_sesion():
    """Recibe el xlsx, parsea, mergea en la sesión abierta del banco
    (o crea una si no hay) y redirect al post-procesar.

    TMT 2026-06-02 dueña: 'que compare numero de documento de banco y si
    ya esta en nuestra lista, no agregarse'. El dedupe es row-level por
    documento contra todo lo que ya tenemos (sesión + histos + matches).
    """
    r = _migracion_lista_o_redirect()
    if r: return r
    usuario = _usuario_actual()
    no_banco = _BANCO_PICHINCHA

    f = request.files.get("archivo")
    if not f or not f.filename:
        flash("Falta el archivo.", "error")
        return redirect(url_for("conciliacion.hub"))

    raw = f.read()
    if not raw:
        flash("El archivo vino vacío.", "error")
        return redirect(url_for("conciliacion.hub"))

    try:
        movs = parse_banco_xlsx(raw)
    except Exception as e:
        _LOG.exception("parser falló: %s", e)
        flash(f"No pude parsear el extracto: {e}", "error")
        return redirect(url_for("conciliacion.hub"))
    if not movs:
        flash("El extracto no trajo movimientos.", "warn")
        return redirect(url_for("conciliacion.hub"))

    sesion_id, n_added, n_skipped = _sesion.crear_sesion(
        no_banco=no_banco,
        usuario=usuario,
        movs=movs,
        extracto_hash=None,  # ya no se usa para dedupe
        extracto_nombre=f.filename,
    )

    # TMT 2026-06-26 (dueña: el saldo objetivo tomaba un valor del medio del
    # día / de un día viejo). Capturamos el saldo REAL del extracto = saldo de
    # la fila MÁS NUEVA, calculado del archivo COMPLETO (antes del dedup, que
    # saca de la sesión los movs ya conocidos y por eso perdía el 26/06). Se
    # guarda en la sesión y el auto-detect lo prioriza. El extracto Pichincha
    # viene newest-first, así que entre varios del día máximo tomamos el 1ro.
    try:
        _con_saldo = [m for m in movs
                      if getattr(m, "fecha", None) and getattr(m, "saldo", None) is not None]
        if _con_saldo:
            _max_f = max(m.fecha for m in _con_saldo)
            _detectado = next((float(m.saldo) for m in _con_saldo if m.fecha == _max_f), None)
            if _detectado is not None:
                _db.execute(
                    "UPDATE scintela.banco_conciliacion_sesion "
                    "SET saldo_banco_detectado = %s WHERE id = %s",
                    (_detectado, sesion_id),
                )
    except Exception as e:
        _LOG.warning("no pude detectar saldo del extracto: %s", e)
    if n_added and n_skipped:
        msg = f"Sesión #{sesion_id}: agregadas {n_added} filas nuevas, omitidas {n_skipped} duplicadas."
        cat = "ok"
    elif n_added:
        msg = f"Sesión #{sesion_id}: agregadas {n_added} filas nuevas."
        cat = "ok"
    elif n_skipped:
        msg = f"Sesión #{sesion_id}: las {n_skipped} filas del archivo ya estaban cargadas — nada nuevo para agregar."
        cat = "info"
    else:
        msg = f"Sesión #{sesion_id}: el archivo no trajo movimientos procesables."
        cat = "warn"
    flash(msg, cat)
    return redirect(url_for("conciliacion.banco_post_procesar"))


# ─── Preview antes de confirmar match (tab Manual) ────────────────────


@conciliacion_bp.route("/banco-v2/preview", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_preview():
    """Vista previa de un match antes de confirmarlo.

    TMT 2026-05-29 dueña: 'necesito que haya un final view antes de
    apretar conciliar y display como cambiarian los movimientos'.

    Recibe los IDs/firmas que mandó el tab Manual, resuelve los items,
    valida signos/montos, y muestra ANTES vs DESPUÉS del balance Pichincha
    junto con la lista de side-effects que se aplicarán. La confirmación
    final POSTea al endpoint manual/confirmar con los mismos campos.
    """
    sesion_id = int(request.form.get("sesion_id") or 0)
    sesion = _sesion.sesion_por_id(sesion_id) if sesion_id else None
    if not sesion or sesion.get("cerrada_en"):
        flash("Sesión inválida o cerrada.", "error")
        return redirect(url_for("conciliacion.hub"))
    no_banco = _BANCO_PICHINCHA

    real_ids_csv = (request.form.get("real_ids") or "").strip()
    real_sigs_csv = (request.form.get("real_sigs") or "").strip()
    hist_ids_csv = (request.form.get("hist_ids") or "").strip()
    bancsis_ids_csv = (request.form.get("bancsis_ids") or "").strip()
    try:
        real_idxs = [int(x) for x in real_ids_csv.split(",") if x.strip()]
        hist_ids = [int(x) for x in hist_ids_csv.split(",") if x.strip()]
        bancsis_ids = [int(x) for x in bancsis_ids_csv.split(",") if x.strip()]
    except ValueError:
        flash("IDs inválidos.", "error")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))
    real_sigs = [s for s in real_sigs_csv.split("||") if s.strip()]

    # Resolver reales por firma (mismo método que manual_confirmar).
    movs = _sesion.cargar_movs(sesion)
    real_subset = []
    if real_sigs:
        sig_a_mov: dict[str, object] = {}
        for m in movs:
            key = "|".join([
                m.fecha.isoformat() if m.fecha else "",
                m.documento or "",
                f"{float(m.monto or 0):.2f}",
                m.tipo or "",
            ])
            sig_a_mov[key] = m
        real_subset = [sig_a_mov[s] for s in real_sigs if s in sig_a_mov]

    # Resolver históricos.
    hist_rows = []
    if hist_ids:
        try:
            hist_rows = _db.fetch_all(
                """
                SELECT id, fecha, concepto, documento, monto, tipo
                  FROM scintela.banco_historicos_pendientes
                 WHERE id = ANY(%s)
                """,
                (hist_ids,),
            ) or []
        except Exception:
            hist_rows = []

    # Resolver BANCSIS.
    bancsis_rows = []
    if bancsis_ids:
        try:
            bancsis_rows = _db.fetch_all(
                """
                SELECT tb.id_transaccion, tb.fecha, tb.documento, tb.importe,
                       tb.concepto, tb.prov, tb.numreferencia,
                       COALESCE(tb.usuario_crea, '') AS usuario_crea,
                       COALESCE(
                         (SELECT nombre FROM scintela.cliente
                           WHERE codigo_cli = tb.prov LIMIT 1), ''
                       ) AS prov_nombre
                  FROM scintela.transacciones_bancarias tb
                 WHERE tb.id_transaccion = ANY(%s)
                 ORDER BY tb.fecha, tb.id_transaccion
                """,
                (bancsis_ids,),
            ) or []
        except Exception:
            bancsis_rows = []

    # Validar signos.
    import bank_helpers as _bh
    _DOCS_CRED = ("DE", "TR", "XX", "NC", "IN", "AC")
    def _sign_bancsis(doc: str, imp: float, usuario_crea: str = "") -> int:
        return 1 if _bh._signed_delta(doc, imp, usuario_crea) >= 0 else -1
    def _sign_real(tipo: str) -> int:
        return 1 if (tipo or "").upper() == "C" else -1

    warnings: list[str] = []
    can_confirm = True

    banco_count = len(real_subset) + len(hist_rows)
    programa_count = len(bancsis_rows)

    if banco_count == 0 or programa_count == 0:
        flash("Faltan items de un lado.", "warn")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))

    # Totales por lado.
    banco_total_signed = sum(
        _sign_real(m.tipo) * float(m.monto or 0) for m in real_subset
    ) + sum(
        (1 if (h.get("tipo") or "").upper() == "C" else -1) * float(h.get("monto") or 0)
        for h in hist_rows
    )
    programa_total_signed = sum(
        _bh._signed_delta(
            (b.get("documento") or ""),
            float(b.get("importe") or 0),
            b.get("usuario_crea") or "",
        )
        for b in bancsis_rows
    )

    diff_total = round(banco_total_signed - programa_total_signed, 2)
    if abs(banco_total_signed) > 0.01 and abs(programa_total_signed) > 0.01:
        signo_banco = 1 if banco_total_signed > 0 else -1
        signo_prog = 1 if programa_total_signed > 0 else -1
        if signo_banco != signo_prog:
            warnings.append(
                f"Signos opuestos — banco {'+' if signo_banco > 0 else '−'} vs "
                f"programa {'+' if signo_prog > 0 else '−'}."
            )
            can_confirm = False
    if abs(diff_total) > 0.01:
        warnings.append(
            f"Diferencia ${diff_total:+,.2f}. Si confirmás, queda asumida."
        )

    # Cálculo del balance ANTES y DESPUÉS.
    balance_before = _bp.calcular(no_banco)

    # Deltas a aplicar.
    delta_pc_cred = 0.0   # créditos PC que dejan de ser pendientes
    delta_pc_deb = 0.0    # débitos PC que dejan de ser pendientes
    for b in bancsis_rows:
        doc = (b.get("documento") or "").upper()
        imp = float(b.get("importe") or 0)
        d = _bh._signed_delta(doc, imp, b.get("usuario_crea") or "")
        if d >= 0:
            delta_pc_cred += d
        else:
            delta_pc_deb += -d  # storage convention: debits positivos

    delta_banco_cred = 0.0
    delta_banco_deb = 0.0
    for h in hist_rows:
        m = float(h.get("monto") or 0)
        t = (h.get("tipo") or "").upper()
        if t == "C":
            delta_banco_cred += m
        else:
            delta_banco_deb += m
    # TMT 2026-06-04: el extracto crudo (real_subset) YA NO es pendiente de
    # banco — pendientes de banco = la hoja (históricos). Matchear un mov del
    # extracto contra PC baja pendientes de PROGRAMA (vía bancsis_rows →
    # delta_pc), no de banco. Por eso real_subset no aporta a delta_banco_*.
    # Solo los históricos consumidos (hist_rows) bajan pendientes de banco.
    # (El 'BUG FIX 2026-06-03' que lo restaba acá se revirtió junto con el
    # bloque del extracto en balance_pichincha.calcular().)

    n_after_pc = max(0, int(balance_before.get("n_pendientes_conciliar") or 0) - len(bancsis_rows))

    pc_cred_after = round(float(balance_before.get("pendientes_pc_creditos") or 0) - delta_pc_cred, 2)
    pc_deb_after = round(float(balance_before.get("pendientes_pc_debitos") or 0) - delta_pc_deb, 2)
    pc_neto_after = round(pc_cred_after - pc_deb_after, 2)
    saldo_si_concilio_after = round(float(balance_before.get("saldo") or 0) - pc_neto_after, 2)

    # TMT 2026-06-03 BUG FIX: usar *_total (incluye extracto sesión).
    # Antes: preview usaba pendientes_banco_creditos solo (histos), salto
    # de $475K en "saldo banco esperado".
    banco_cred_after = round(
        float(balance_before.get("pendientes_banco_total_creditos") or balance_before.get("pendientes_banco_creditos") or 0)
        - delta_banco_cred, 2,
    )
    banco_deb_after = round(
        float(balance_before.get("pendientes_banco_total_debitos") or balance_before.get("pendientes_banco_debitos") or 0)
        - delta_banco_deb, 2,
    )
    banco_neto_after = round(banco_cred_after - banco_deb_after, 2)
    saldo_banco_esperado_after = round(saldo_si_concilio_after + banco_neto_after, 2)
    n_after_banco = max(
        0,
        int(balance_before.get("n_pendientes_banco_total") or balance_before.get("n_pendientes") or 0)
        - len(hist_rows),
    )

    balance_after = {
        "saldo": float(balance_before.get("saldo") or 0),
        "pendientes_pc_creditos": pc_cred_after,
        "pendientes_pc_debitos": pc_deb_after,
        "pendientes_conciliar_neto": pc_neto_after,
        "n_pendientes_conciliar": n_after_pc,
        "saldo_si_concilio_todo": saldo_si_concilio_after,
        "pendientes_banco_creditos": banco_cred_after,
        "pendientes_banco_debitos": banco_deb_after,
        "pendientes_banco_total_creditos": banco_cred_after,
        "pendientes_banco_total_debitos": banco_deb_after,
        "neto_pendientes": banco_neto_after,
        "neto_pendientes_total": banco_neto_after,
        "n_pendientes": n_after_banco,
        "n_pendientes_banco_total": n_after_banco,
        "saldo_banco_esperado": saldo_banco_esperado_after,
    }

    return render_template(
        "conciliacion/banco_v2_preview.html",
        sesion=sesion,
        balance_before=balance_before,
        balance_after=balance_after,
        real_subset=real_subset,
        hist_rows=hist_rows,
        bancsis_rows=bancsis_rows,
        banco_total_signed=banco_total_signed,
        programa_total_signed=programa_total_signed,
        diff_total=diff_total,
        warnings=warnings,
        can_confirm=can_confirm,
        # Pass-through para que el submit final repita exactamente.
        real_ids_csv=real_ids_csv,
        real_sigs_csv=real_sigs_csv,
        hist_ids_csv=hist_ids_csv,
        bancsis_ids_csv=bancsis_ids_csv,
        sesion_id=sesion_id,
    )


# ─── Auditar diferencia (root cause de la brecha PC vs banco real) ───


@conciliacion_bp.route("/banco-v2/auditar", methods=["GET"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_auditar():
    """Audit page: muestra los componentes que generan la diferencia entre
    saldo banco esperado (cálculo) y saldo banco real (extracto).

    TMT 2026-05-29 dueña: 'no esta bien que tengamos diferencia' +
    'necesito que vos hagas la auditoria'. Lista TRES diagnósticos:
      1) Drift del saldo running PC (suma de movs vs último saldo).
      2) Matches confirmados con monto distinto entre real y BANCSIS.
      3) Pendientes de banco con monto raro o duplicados de firma.
    """
    no_banco = _BANCO_PICHINCHA
    balance = _bp.calcular(no_banco)

    # ── Check 1: walk-forward saldo running PC ─────────────────────
    # TMT 2026-05-29 dueña: 'esto sigue mal'. Bug en el auditor anterior:
    # asumía que TODAS las importes son positivas y aplicaba ±signo según
    # documento. Pero en la DB hay convención MIXTA — filas legacy del
    # dBase tienen importe ya signado (ej ND con importe=−40,775) y el
    # código nuevo usa importe absoluto. La fórmula simple daba falsos
    # positivos por miles. Fix: usar la misma función _signed_delta de
    # bank_helpers — fuente de verdad usada por trigger e insert.
    import bank_helpers as _bh
    filas_torcidas = []
    last_saldo = None
    try:
        rows_walk = _db.fetch_all(
            """
            SELECT id_transaccion, fecha, documento, importe, saldo, concepto,
                   COALESCE(usuario_crea, '') AS usuario_crea
              FROM scintela.transacciones_bancarias
             WHERE no_banco = %s AND saldo IS NOT NULL
             ORDER BY fecha ASC, id_transaccion ASC
            """,
            (no_banco,),
        ) or []
        saldo_prev = None
        for r in rows_walk:
            s = float(r["saldo"] or 0)
            imp = float(r["importe"] or 0)
            doc = (r.get("documento") or "").upper()
            # TMT 2026-06-03 audit fix: pasamos usuario_crea para que NDs reverso
            # del DBF (importe positivo legítimo) no se reporten como torcidas.
            delta = _bh._signed_delta(doc, imp, r.get("usuario_crea") or "")
            if saldo_prev is not None:
                esperado = round(saldo_prev + delta, 2)
                diff = round(s - esperado, 2)
                if abs(diff) > 0.01:
                    filas_torcidas.append({
                        "id": int(r["id_transaccion"]),
                        "fecha": r["fecha"],
                        "documento": doc,
                        "importe": imp,
                        "delta_aplicado": round(delta, 2),
                        "saldo_grabado": s,
                        "saldo_esperado": esperado,
                        "diferencia": diff,
                        "concepto": (r.get("concepto") or "")[:50],
                    })
            saldo_prev = s
        last_saldo = saldo_prev
        # Top 200 desviaciones por magnitud (las más grandes primero).
        filas_torcidas.sort(key=lambda x: abs(x["diferencia"]), reverse=True)
        filas_torcidas = filas_torcidas[:200]
    except Exception as e:
        _LOG.warning("auditar walk-forward falló: %s", e)
    suma_diff_torcidas = round(sum(x["diferencia"] for x in filas_torcidas), 2)

    # ── Check 2: matches confirmados con drift de monto ────────────
    diff_matches = []
    sum_diff_matches = 0.0
    try:
        diff_matches = _db.fetch_all(
            """
            -- TMT 2026-06-04: agrupar por PC tx. En grupos N:1 (ej. 15 cheques
            -- contra 1 depósito) cada match tiene real_monto chico vs el importe
            -- total del PC → comparar uno a uno daba falso positivo gigante.
            -- Ahora sumamos los real_monto del grupo y comparamos la SUMA vs el
            -- importe del PC. Solo salta si la suma del grupo NO cuadra.
            SELECT MAX(m.creado_en) AS creado_en,
                   MIN(m.real_fecha) AS real_fecha,
                   CASE WHEN COUNT(*) > 1
                        THEN COUNT(*)::text || ' movs (grupo)'
                        ELSE MIN(m.real_documento) END AS real_documento,
                   SUM(m.real_monto) AS real_monto,
                   MIN(m.real_concepto) AS real_concepto,
                   tb.importe   AS tb_importe,
                   tb.documento AS tb_documento,
                   tb.concepto  AS tb_concepto,
                   ROUND(SUM(m.real_monto) - tb.importe, 2) AS diferencia
              FROM scintela.banco_conciliacion_match m
              JOIN scintela.transacciones_bancarias tb
                ON tb.id_transaccion = m.id_transaccion
             WHERE m.no_banco = %s
               AND (m.deshecho_en IS NULL)
               AND m.id_transaccion IS NOT NULL
               AND m.real_monto IS NOT NULL
               AND m.estado = 'matched'
             GROUP BY tb.id_transaccion, tb.importe, tb.documento, tb.concepto
            HAVING ABS(SUM(m.real_monto) - tb.importe) > 0.01
             ORDER BY ABS(SUM(m.real_monto) - tb.importe) DESC
             LIMIT 200
            """,
            (no_banco,),
        ) or []
        sum_diff_matches = sum(float(r.get("diferencia") or 0) for r in diff_matches)
    except Exception as e:
        _LOG.warning("auditar diff_matches falló: %s", e)

    # ── Check 3: pendientes históricos con firmas duplicadas ───────
    duplicados_hist = []
    try:
        duplicados_hist = _db.fetch_all(
            """
            SELECT no_banco, fecha, COALESCE(documento, '') AS documento,
                   monto, tipo, COUNT(*) AS n, SUM(monto) AS suma
              FROM scintela.banco_historicos_pendientes
             WHERE no_banco = %s
               AND conciliado_en IS NULL
             GROUP BY no_banco, fecha, COALESCE(documento, ''), monto, tipo
            HAVING COUNT(*) > 1
             ORDER BY COUNT(*) DESC, ABS(monto) DESC
             LIMIT 100
            """,
            (no_banco,),
        ) or []
    except Exception as e:
        _LOG.warning("auditar duplicados_hist falló: %s", e)

    # ── Diferencia objetivo (lo que la dueña ve en la página) ──────
    # saldo_banco_esperado calculado vs saldo banco real del último extracto
    # de la sesión abierta (si existe).
    sesion = _sesion.sesion_abierta(no_banco, _usuario_actual())
    saldo_banco_real = None
    if sesion:
        try:
            movs_s = _sesion.cargar_movs(sesion)
            con_fecha = [
                m for m in movs_s
                if getattr(m, "fecha", None) and getattr(m, "saldo", None) is not None
            ]
            if con_fecha:
                ult = max(con_fecha, key=lambda m: m.fecha)
                saldo_banco_real = float(ult.saldo)
        except Exception:
            pass
    diferencia_objetivo = None
    if saldo_banco_real is not None and balance.get("saldo_banco_esperado") is not None:
        diferencia_objetivo = round(
            saldo_banco_real - balance["saldo_banco_esperado"], 2
        )

    return render_template(
        "conciliacion/auditar.html",
        balance=balance,
        saldo_banco_real=saldo_banco_real,
        diferencia_objetivo=diferencia_objetivo,
        filas_torcidas=filas_torcidas,
        suma_diff_torcidas=suma_diff_torcidas,
        last_saldo=last_saldo,
        diff_matches=diff_matches,
        sum_diff_matches=round(sum_diff_matches, 2),
        duplicados_hist=duplicados_hist,
    )


# ─── Reabrir sesión cerrada ──────────────────────────────────────────


# ─── Endpoint: setear saldo banco objetivo manual ─────────────────────


@conciliacion_bp.route("/banco-v2/set-saldo-objetivo", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_set_saldo_objetivo():
    """Guarda el saldo banco que la dueña lee del extracto o de la web del
    banco. TMT 2026-06-02 dueña: 'lo deberia implementar el usuario no?
    porque por el excel no sabemos cual es el ultimo valor'.

    Reemplaza el auto-detect por max(fecha) que era frágil cuando había
    múltiples movs el mismo día o varios uploads merged.
    """
    sesion_id = int(request.form.get("sesion_id") or 0)
    sesion = _sesion.sesion_por_id(sesion_id) if sesion_id else None
    if not sesion:
        flash("Sesión no encontrada.", "error")
        return redirect(url_for("conciliacion.hub"))

    raw = (request.form.get("saldo_objetivo") or "").strip()
    if raw == "" or raw.lower() in ("none", "null", "—"):
        # Clear → vuelve al auto-detect.
        try:
            _db.execute(
                """
                UPDATE scintela.banco_conciliacion_sesion
                   SET saldo_banco_objetivo = NULL
                 WHERE id = %s
                """,
                (sesion_id,),
            )
            flash("Saldo banco objetivo borrado — vuelve al auto-detect.", "ok")
        except Exception as e:
            flash(f"Error: {e}", "error")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))

    # Limpia comas/separadores antes de parsear.
    try:
        cleaned = raw.replace(",", "").replace("$", "").strip()
        valor = float(cleaned)
    except (ValueError, TypeError):
        flash(f"Valor inválido: '{raw}'. Usá formato numérico (ej. 2846820.24).", "error")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))

    try:
        _db.execute(
            """
            UPDATE scintela.banco_conciliacion_sesion
               SET saldo_banco_objetivo = %s
             WHERE id = %s
            """,
            (valor, sesion_id),
        )
        flash(f"Saldo banco objetivo: ${valor:,.2f}", "ok")
    except Exception as e:
        flash(f"Error al guardar: {e}", "error")
    return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))


# ─── Endpoint: descartar movs banco (no necesitan conciliarse) ────────


@conciliacion_bp.route("/banco-v2/descartar-banco", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_descartar():
    """Marca movs del lado banco como conciliados SIN contraparte PC.

    TMT 2026-06-02 dueña: 'si hay movimientos que no quiero conciliar,
    porque ya fueron conciliados o porque si del banco como hago? quizas
    tengo entrada y devolucion de un cheque'.

    Casos típicos:
      - Entrada $X + devolución $X del mismo cheque → ambos netean a 0,
        no hay contraparte PC.
      - Movs ya conciliados externamente (ej. desde dBase) que el sistema
        no detecta.
      - Cargos del banco que la dueña decide no trackear.

    Acepta hist_ids[] (histos backfilled) y real_sigs[] (movs del extracto
    de la sesión). Para hist: UPDATE conciliado_en. Para real: los inserta
    como histos con conciliado_en=NOW (queda registro auditable, y como
    están conciliados no aparecen como pendientes).
    """
    sesion_id = int(request.form.get("sesion_id") or 0)
    sesion = _sesion.sesion_por_id(sesion_id) if sesion_id else None
    if not sesion:
        flash("Sesión no encontrada.", "error")
        return redirect(url_for("conciliacion.hub"))

    hist_ids_csv = (request.form.get("hist_ids") or "").strip()
    real_sigs_csv = (request.form.get("real_sigs") or "").strip()
    motivo = (request.form.get("motivo") or "descartado-banco")[:50]

    try:
        hist_ids = [int(x) for x in hist_ids_csv.split(",") if x.strip()]
    except ValueError:
        hist_ids = []
    real_sigs = [s for s in real_sigs_csv.split("||") if s.strip()]

    if not hist_ids and not real_sigs:
        flash("No marcaste ningún mov para descartar.", "warn")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))

    usuario = _usuario_actual()
    no_banco = _BANCO_PICHINCHA
    n_hist = 0
    n_real = 0

    # 1) Marcar históricos directamente — UPDATE conciliado_en.
    if hist_ids:
        try:
            n_hist = _db.execute(
                """
                UPDATE scintela.banco_historicos_pendientes
                   SET conciliado_en = CURRENT_TIMESTAMP,
                       conciliado_por = %s
                 WHERE id = ANY(%s)
                   AND no_banco = %s
                   AND conciliado_en IS NULL
                """,
                (f"descart:{motivo}"[:50], hist_ids, no_banco),
            ) or 0
        except Exception as e:
            _LOG.exception("descartar histos falló: %s", e)
            flash(f"Error al descartar históricos: {e}", "error")
            return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))

    # 2) Para movs del extracto: PRIMERO intentamos marcar como conciliado
    # un histo existente que matchee firma (caso común: el extracto trajo
    # una fila que YA está en banco_historicos_pendientes como pendiente).
    # Si no hay, INSERTAMOS uno nuevo ya conciliado.
    # TMT 2026-06-02 fix: antes hacíamos solo INSERT ... ON CONFLICT DO
    # NOTHING, que silenciosamente saltaba cuando el histo ya estaba y no
    # actualizaba conciliado_en — la dueña veía el mov seguir pendiente.
    if real_sigs:
        movs = _sesion.cargar_movs(sesion)
        sig_a_mov = {}
        for m in movs:
            key = "|".join([
                m.fecha.isoformat() if m.fecha else "",
                m.documento or "",
                f"{float(m.monto or 0):.2f}",
                m.tipo or "",
            ])
            sig_a_mov[key] = m
        for sig in real_sigs:
            mov = sig_a_mov.get(sig)
            if not mov:
                continue
            try:
                # Intento 1: UPDATE del histo existente (pendiente) que matchee firma.
                rc = _db.execute(
                    """
                    UPDATE scintela.banco_historicos_pendientes
                       SET conciliado_en = CURRENT_TIMESTAMP,
                           conciliado_por = %s
                     WHERE no_banco = %s
                       AND fecha = %s
                       AND COALESCE(documento, '') = COALESCE(%s, '')
                       AND monto = %s::numeric
                       AND tipo = %s
                       AND conciliado_en IS NULL
                    """,
                    (
                        f"descart:{motivo}"[:50],
                        no_banco, mov.fecha,
                        (mov.documento or "")[:40],
                        str(mov.monto or 0),
                        (mov.tipo or "C")[:1],
                    ),
                ) or 0
                if rc > 0:
                    n_real += rc
                    continue
                # Intento 2: no había histo existente → INSERT uno ya conciliado.
                _db.execute(
                    """
                    INSERT INTO scintela.banco_historicos_pendientes
                        (no_banco, fecha, concepto, documento, monto, tipo,
                         oficina, detalle, fuente, creado_por,
                         conciliado_en, conciliado_por, codigo)
                    VALUES (%s, %s, %s, %s, %s::numeric, %s, %s, %s, %s, %s,
                            CURRENT_TIMESTAMP, %s, %s)
                    ON CONFLICT DO NOTHING
                    """,
                    (
                        no_banco, mov.fecha,
                        (mov.concepto or "")[:120],
                        (mov.documento or "")[:40],
                        str(mov.monto or 0),
                        (mov.tipo or "C")[:1],
                        (getattr(mov, "oficina", "") or "")[:40],
                        "",  # detalle
                        f"descart:sesion:{sesion_id}",
                        usuario[:50],
                        f"descart:{motivo}"[:50],
                        (getattr(mov, "codigo", "") or "")[:20],
                    ),
                )
                n_real += 1
            except Exception as e:
                _LOG.warning("descartar real_sig %s falló: %s", sig, e)

    total = n_hist + n_real
    flash(
        f"Descartados {total} movs ({n_hist} histos + {n_real} del extracto). "
        f"Marcados como conciliados sin contraparte PC. Motivo: {motivo}.",
        "ok" if total > 0 else "warn",
    )
    return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))


@conciliacion_bp.route("/banco-v2/descartar-programa", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_descartar_programa():
    """Concilia movs del lado PROGRAMA como INTERNOS, sin contraparte de banco.

    Espejo de /descartar-banco para el lado programa (pedido Tamara
    2026-06-19, caso de Alex). Caso típico: la **corrección de un negativo**
    — un asiento OP interno (concepto INOP / 'IN OP') que netea contra su par
    y NO toca el saldo real del banco (BANC=0 en el dBase), así que el
    extracto no tiene una línea para aparearlo. Antes la pantalla NO dejaba
    conciliarlo porque `banco_manual_confirmar` exige un mov de banco; quedaba
    trabado con Δ=importe.

    Marca cada bancsis_id con stat='*' y crea un banco_conciliacion_match
    (estado='matched', metodo='interno') → auditable y REVERSIBLE desde
    'Deshacer conciliados', igual que un match normal.
    """
    sesion_id = int(request.form.get("sesion_id") or 0)
    sesion = _sesion.sesion_por_id(sesion_id) if sesion_id else None
    if not sesion or sesion.get("cerrada_en"):
        flash("Sesión inválida o cerrada.", "error")
        return redirect(url_for("conciliacion.hub"))

    bancsis_ids_csv = (request.form.get("bancsis_ids") or "").strip()
    motivo = (request.form.get("motivo") or "interno")[:50]
    try:
        bancsis_ids = [int(x) for x in bancsis_ids_csv.split(",") if x.strip()]
    except ValueError:
        bancsis_ids = []
    if not bancsis_ids:
        flash("No marcaste ningún mov del programa para conciliar como interno.", "warn")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))

    usuario = _usuario_actual()
    no_banco = _BANCO_PICHINCHA
    import uuid as _uuid
    batch_id = _uuid.uuid4().hex

    tiene_metodo = False
    try:
        tiene_metodo = bool(_db.fetch_one(
            """
            SELECT 1 FROM information_schema.columns
             WHERE table_schema='scintela'
               AND table_name='banco_conciliacion_match'
               AND column_name='metodo'
            """,
        ))
    except Exception:
        pass

    n_ok = 0
    err_msg = None
    metodo = f"interno:{motivo}"[:40]
    for bk_id in bancsis_ids:
        try:
            if tiene_metodo:
                _db.execute(
                    """
                    INSERT INTO scintela.banco_conciliacion_match (
                        no_banco, estado, metodo,
                        id_transaccion, tx_firma, confirm_batch_id, usuario
                    ) VALUES (%s, 'matched', %s, %s,
                              scintela.compute_tx_firma(%s), %s, %s)
                    """,
                    (no_banco, metodo, bk_id, bk_id, batch_id, usuario),
                )
            else:
                _db.execute(
                    """
                    INSERT INTO scintela.banco_conciliacion_match (
                        no_banco, estado, id_transaccion, tx_firma, confirm_batch_id, usuario
                    ) VALUES (%s, 'matched', %s,
                              scintela.compute_tx_firma(%s), %s, %s)
                    """,
                    (no_banco, bk_id, bk_id, batch_id, usuario),
                )
            _db.execute(
                """
                UPDATE scintela.transacciones_bancarias
                   SET stat = '*'
                 WHERE id_transaccion = %s AND no_banco = %s
                """,
                (bk_id, no_banco),
            )
            n_ok += 1
        except Exception as e:
            _LOG.warning("descartar programa %s falló: %s", bk_id, e)
            if err_msg is None:
                err_msg = str(e)

    if n_ok:
        _sesion.incrementar_matches(sesion_id, n_ok)
        flash(
            f"{n_ok} mov(s) de programa conciliado(s) como INTERNO (sin "
            f"contraparte de banco). Motivo: {motivo}. Reversible desde "
            f"'Deshacer conciliados'.",
            "ok",
        )
    else:
        flash(f"No pude conciliar como interno. {err_msg or ''}".strip(), "error")
    return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))


# ─── Endpoint Tab Manual: confirmar pares marcados ────────────────────


def _seleccion_firmada(sesion, form):
    """Núcleo COMPARTIDO de parsing de la selección manual (banco+programa) y
    construcción de los firmados. Lo usan `banco_manual_confirmar` y
    `banco_diferencia_confirmar` para partir del MISMO estado.

    NO hace flashes ni redirects — devuelve un dict con TODO lo que ambos
    endpoints necesitan + `error` (None | "ids_invalidos" | "faltan_movs") para
    que cada endpoint decida su propio flash/redirect. Así el comportamiento
    observable de `banco_manual_confirmar` (flashes/confirms/todo-o-nada) queda
    IDÉNTICO al de antes de esta extracción.

    Devuelve dict con:
      error, real_ids_csv, real_sigs_csv, hist_ids_csv, bancsis_ids_csv,
      real_idxs, hist_ids, bancsis_ids, real_sigs, real_subset, banco_movs,
      banco_firmados, prog_firmados, metodo_resolucion, prog_meta.
    """
    from decimal import Decimal as _D_cf

    import bank_helpers as _bh_confirm
    from modules.conciliacion.parser_banco import MovBanco as _MB_cf

    real_ids_csv = (form.get("real_ids") or "").strip()
    real_sigs_csv = (form.get("real_sigs") or "").strip()
    hist_ids_csv = (form.get("hist_ids") or "").strip()
    bancsis_ids_csv = (form.get("bancsis_ids") or "").strip()

    base = {
        "error": None,
        "real_ids_csv": real_ids_csv,
        "real_sigs_csv": real_sigs_csv,
        "hist_ids_csv": hist_ids_csv,
        "bancsis_ids_csv": bancsis_ids_csv,
        "real_idxs": [],
        "hist_ids": [],
        "bancsis_ids": [],
        "real_sigs": [],
        "real_subset": [],
        "banco_movs": [],
        "banco_firmados": [],
        "prog_firmados": [],
        "metodo_resolucion": "n/a",
        "prog_meta": {},
    }

    try:
        real_idxs = [int(x) for x in real_ids_csv.split(",") if x.strip()]
        hist_ids = [int(x) for x in hist_ids_csv.split(",") if x.strip()]
        bancsis_ids = [int(x) for x in bancsis_ids_csv.split(",") if x.strip()]
    except ValueError:
        base["error"] = "ids_invalidos"
        return base
    real_sigs = [s for s in real_sigs_csv.split("||") if s.strip()]

    base.update(
        real_idxs=real_idxs, hist_ids=hist_ids, bancsis_ids=bancsis_ids,
        real_sigs=real_sigs,
    )

    if (not real_idxs and not hist_ids) or not bancsis_ids:
        base["error"] = "faltan_movs"
        return base

    # TMT 2026-05-29 dueña: 'manual me aparece el mensaje vas a conciliar y
    # cuando pongo ok dice sin cambios. ARREGLALO YA'.
    # Lecciones: el matcher re-corrido a veces NO contiene los movs reales
    # que la dueña seleccionó (PASS 0 los reclasifica a matches[], o el
    # orden cambia y el idx queda stale). Necesitamos un mecanismo DURABLE.
    #
    # Nueva estrategia (sigs como PRIMARIO, idx como fallback solo si el
    # front es viejo):
    #   1. Si hay real_sigs → resolver contra el payload CRUDO de la sesión.
    #      No depende del matcher; las firmas siempre encuentran su mov si
    #      existe en el extracto subido.
    #   2. Si NO hay real_sigs (cliente viejo, antes del deploy de hoy) →
    #      caer al matcher y al idx, como antes.
    movs = _sesion.cargar_movs(sesion)
    real_subset = []
    metodo_resolucion = "n/a"

    if real_sigs:
        sig_a_mov: dict[str, object] = {}
        for m in movs:
            key = "|".join([
                m.fecha.isoformat() if m.fecha else "",
                m.documento or "",
                f"{float(m.monto or 0):.2f}",
                m.tipo or "",
            ])
            sig_a_mov[key] = m
        real_subset = [sig_a_mov[s] for s in real_sigs if s in sig_a_mov]
        metodo_resolucion = f"firma ({len(real_subset)}/{len(real_sigs)})"

    if real_idxs and not real_subset:
        # Fallback: matcher + idx. Solo si las firmas no resolvieron nada.
        from modules.conciliacion.matcher_banco import matchear_extracto_banco
        try:
            res = matchear_extracto_banco(movs, no_banco=_BANCO_PICHINCHA)
            real_only = res.real_only or []
        except Exception as e:
            _LOG.exception("re-match para confirmar manual falló: %s", e)
            real_only = []
        real_subset = [real_only[i] for i in real_idxs if 0 <= i < len(real_only)]
        metodo_resolucion = f"idx ({len(real_subset)}/{len(real_idxs)} en real_only[{len(real_only)}])"

    _LOG.info(
        "manual confirm: real_idxs=%d real_sigs=%d hist=%d bancsis=%d resolución=%s subset=%d",
        len(real_idxs), len(real_sigs), len(hist_ids), len(bancsis_ids),
        metodo_resolucion, len(real_subset),
    )

    banco_movs: list[tuple] = []
    for _m in real_subset:
        banco_movs.append((_m, "matched_manual"))

    hist_rows = []
    if hist_ids:
        try:
            hist_rows = _db.fetch_all(
                """
                SELECT id, no_banco, fecha, concepto, documento, monto, tipo,
                       oficina, detalle
                  FROM scintela.banco_historicos_pendientes
                 WHERE id = ANY(%s)
                """,
                (hist_ids,),
            ) or []
        except Exception as e:
            _LOG.exception("fetch historicos fallo: %s", e)
            hist_rows = []
        for h in hist_rows:
            mov_h = _MB_cf(
                fecha=h.get("fecha"),
                concepto=str(h.get("concepto") or ""),
                documento=str(h.get("documento") or ""),
                monto=_D_cf(str(h.get("monto") or 0)),
                saldo=_D_cf("0"),
                codigo=str(h.get("oficina") or "")[:10],
                tipo=str(h.get("tipo") or "C").upper(),
                oficina=str(h.get("oficina") or ""),
            )
            banco_movs.append((mov_h, "matched_historico"))

    def _signed_banco(mov) -> float:
        s = 1.0 if (getattr(mov, "tipo", "") or "").upper() == "C" else -1.0
        return s * float(getattr(mov, "monto", 0) or 0)

    prog_meta: dict = {}
    if bancsis_ids:
        try:
            _pr = _db.fetch_all(
                """
                SELECT id_transaccion, importe, documento,
                       COALESCE(usuario_crea, '') AS usuario_crea
                  FROM scintela.transacciones_bancarias
                 WHERE id_transaccion = ANY(%s) AND no_banco = %s
                """,
                (bancsis_ids, _BANCO_PICHINCHA),
            ) or []
            for r in _pr:
                prog_meta[int(r["id_transaccion"])] = r
        except Exception as e:
            _LOG.warning("fetch importes programa fallo: %s", e)

    def _signed_prog(bk_id) -> float:
        r = prog_meta.get(int(bk_id))
        if not r:
            return 0.0
        try:
            return float(_bh_confirm._signed_delta(
                (r.get("documento") or ""),
                float(r.get("importe") or 0),
                r.get("usuario_crea") or "",
            ))
        except Exception:
            return float(r.get("importe") or 0)

    banco_firmados = [(i, _signed_banco(m)) for i, (m, _md) in enumerate(banco_movs)]
    prog_firmados = [(int(b), _signed_prog(b)) for b in bancsis_ids]

    base.update(
        real_subset=real_subset,
        banco_movs=banco_movs,
        banco_firmados=banco_firmados,
        prog_firmados=prog_firmados,
        metodo_resolucion=metodo_resolucion,
        prog_meta=prog_meta,
    )
    return base


@conciliacion_bp.route("/banco-v2/manual/confirmar", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_manual_confirmar():
    """Confirma N matches manuales. Por ahora hace N:N por suma — si los
    montos cuadran, mapea 1:1 ordenado por monto; sino acepta el primer
    bancsis como representativo y los demás como sus pares por orden.
    """
    sesion_id = int(request.form.get("sesion_id") or 0)
    sesion = _sesion.sesion_por_id(sesion_id) if sesion_id else None
    if not sesion or sesion.get("cerrada_en"):
        flash("Sesión inválida o cerrada.", "error")
        return redirect(url_for("conciliacion.hub"))

    sel = _seleccion_firmada(sesion, request.form)
    if sel["error"] == "ids_invalidos":
        flash("IDs inválidos.", "error")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))
    if sel["error"] == "faltan_movs":
        flash("Marcá al menos un mov de cada lado.", "warn")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))

    real_idxs = sel["real_idxs"]
    hist_ids = sel["hist_ids"]
    bancsis_ids = sel["bancsis_ids"]
    real_sigs = sel["real_sigs"]
    real_subset = sel["real_subset"]
    metodo_resolucion = sel["metodo_resolucion"]
    banco_movs = sel["banco_movs"]

    n_matches = 0
    err_msg: str | None = None
    usuario = _usuario_actual()
    # TMT 2026-06-03 dueña: 'batch id, cuando selecciono junto'. Cada vez
    # que aprieta Conciliar, los items seleccionados forman un batch nuevo.
    # Si después quiere mover items entre batches, usa "Sacar al grupo nuevo".
    # TMT 2026-06-02 dueña: 'los movimientos no tienen que ser 1:1' / 'puede
    # ser distinto'. Aceptamos cualquier N:M.
    #
    # TMT 2026-07-14 (dueña, TODO-O-NADA): la conciliación de una selección es
    # ATÓMICA. Se calcula la asignación por monto (asignar_banco_a_programa) y
    # SOLO se confirma si la selección se reconcilia COMPLETA (ningún banco
    # sobrante Y ningún PC sin contraparte). Si algo no cierra → no se confirma
    # NADA y queda todo pendiente. Un confirm_batch_id por grupo PC (N:1).
    #
    # TMT 2026-06-03: el lado banco unifica reales del extracto + históricos en
    # UNA lista; el lado programa se resuelve por monto firmado, no por id.

    n_hist = 0
    sobrantes_banco = 0
    pcs_internos: list = []
    incompleto = False
    incompleto_msg: str | None = None
    if banco_movs and bancsis_ids:
        # TMT 2026-07-14 (dueña): "debería funcionar en cualquier caso" +
        # corrección: "deberías dejar TODO pendiente si no se puede matchear,
        # no aprobar la una de las partes". La conciliación de una selección es
        # ATÓMICA (todo-o-nada). Calculamos la asignación ANTES de tocar la DB
        # con `asignar_banco_a_programa` (1:1 exacto, N:1 depósito partido,
        # multi-PC) y SOLO confirmamos si la selección se reconcilia COMPLETA:
        #   - NO quedan banco sobrantes, Y
        #   - NO queda ningún mov de programa sin su contraparte de banco
        #     (nada de stat-only "PC internos" de a una pata).
        # Si algo no cierra → NO se confirma NINGÚN match (ni los pares que sí
        # cerraban): queda TODO pendiente. Mejor pendiente que medio-conciliado.
        # Cada banco se indexa por posición para mapear de vuelta a (mov, metodo).
        banco_firmados = sel["banco_firmados"]
        prog_firmados = sel["prog_firmados"]

        asign, sobrantes_idx = asignar_banco_a_programa(
            banco_firmados, prog_firmados, tol=0.50,
        )
        completa, pcs_sin_banco = reconciliacion_completa(
            asign, sobrantes_idx, bancsis_ids,
        )

        if not completa:
            # TMT 2026-07-14 (dueña, "Agregar movimiento en el programa"): si
            # lo ÚNICO que impide cerrar es un residuo del lado BANCO (todo PC
            # matcheó → pcs_sin_banco vacío, pero hay sobrantes) y la diferencia
            # supera la tolerancia, ofrecemos crear el mov de programa faltante.
            # El preview NO muta nada; la creación va por banco_diferencia_confirmar.
            prop = _proponer_movimiento_diferencia(
                banco_firmados, prog_firmados, sobrantes_idx, pcs_sin_banco,
            )
            if prop:
                return render_template(
                    "conciliacion/banco_v2_diferencia_preview.html",
                    sesion=sesion,
                    prop=prop,
                    real_ids_csv=sel["real_ids_csv"],
                    real_sigs_csv=sel["real_sigs_csv"],
                    hist_ids_csv=sel["hist_ids_csv"],
                    bancsis_ids_csv=sel["bancsis_ids_csv"],
                    n_banco=len(banco_movs),
                    n_prog=len(bancsis_ids),
                )
            # dueña 2026-07-17: "cuando hay diferencia que re pregunte pero
            # que deje igual". La vista previa YA mostró la advertencia de
            # diferencia y la dueña confirmó igual → cerrar la asignación a la
            # fuerza (por cercanía de monto) y conciliar con el gap asumido.
            # Antes esto fallaba todo-o-nada cuando |dif| > tol=0.50 (ej. $1).
            if (request.form.get("aceptar_diferencia") or "") == "1":
                asign, sobrantes_idx, pcs_internos = forzar_asignacion_completa(
                    asign, sobrantes_idx, banco_firmados, prog_firmados,
                )
                # Los PCs que quedaron sin banco se concilian como INTERNOS
                # abajo — para el todo-o-nada ya están resueltos.
                _internos_set = set(pcs_internos)
                completa, pcs_sin_banco = reconciliacion_completa(
                    asign, sobrantes_idx,
                    [b for b in bancsis_ids if b not in _internos_set],
                )
        if not completa:
            # TODO-O-NADA: no confirmar nada. Todo queda pendiente.
            sobrantes_banco = len(sobrantes_idx)
            incompleto = True
            incompleto_msg = (
                f"No se concilió nada: {len(sobrantes_idx)} de banco / "
                f"{len(pcs_sin_banco)} de programa sin match. "
                f"Quedó todo pendiente."
            )
        else:
            # Reconciliación COMPLETA → confirmar cada grupo PC → [banco]. Un
            # confirm_batch_id POR grupo, para que un N:1 (depósito partido)
            # quede como un solo grupo y los distintos PC no se mezclen.
            import uuid as _uuid_grp
            for pc_id, banco_idxs in asign.items():
                grupo_batch = _uuid_grp.uuid4().hex
                for idx in banco_idxs:
                    mov_i, metodo_i = banco_movs[idx]
                    try:
                        confirmar_match(_BANCO_PICHINCHA, mov_i, pc_id,
                                        usuario=usuario, metodo=metodo_i,
                                        confirm_batch_id=grupo_batch)
                        if metodo_i == "matched_historico":
                            n_hist += 1
                        else:
                            n_matches += 1
                    except Exception as e:
                        _LOG.warning("manual confirm PC %s idx %s fallo: %s", pc_id, idx, e)
                        if err_msg is None:
                            err_msg = str(e)
            # PCs sin banco disponible (selección N:M aceptada con diferencia)
            # → conciliados como INTERNOS, un batch por PC (reversibles).
            for pc_id in pcs_internos:
                try:
                    _conciliar_pc_interno(pc_id, _uuid_grp.uuid4().hex, usuario)
                    n_matches += 1
                except Exception as e:
                    _LOG.warning("confirm interno PC %s fallo: %s", pc_id, e)
                    if err_msg is None:
                        err_msg = str(e)
    elif hist_ids and not bancsis_ids:
        flash(
            "Para conciliar historicos hay que seleccionar al menos un "
            "movimiento del programa.",
            "warn",
        )
    total = n_matches + n_hist
    _sesion.incrementar_matches(sesion_id, total)
    if incompleto and incompleto_msg:
        flash("⚠ " + incompleto_msg, "warn")
    parts = []
    if n_matches: parts.append(f"{n_matches} match(es) del extracto")
    if n_hist: parts.append(f"{n_hist} histórico(s)")
    if parts:
        flash(" + ".join(parts) + " conciliado(s).", "ok")
    elif incompleto:
        # Ya flasheamos el mensaje todo-o-nada arriba; no duplicar diagnóstico.
        pass
    else:
        # Diagnóstico corto (dueña 2026-07-17: "saca los mensajes muy wordies").
        # El detalle largo va al log, no al flash.
        diag = (
            f"banco enviado={len(real_idxs)}+{len(hist_ids)}hist, "
            f"programa enviado={len(bancsis_ids)}, "
            f"resolución={metodo_resolucion}"
        )
        _LOG.warning("manual confirm sin cambios: %s err=%s", diag, err_msg)
        if real_subset and bancsis_ids and not n_matches:
            # Llegó al confirm_match pero todos fallaron (raro).
            flash(
                f"Los {len(real_subset)} match(es) fallaron."
                f"{(' ' + err_msg) if err_msg else ''}",
                "error",
            )
        elif real_idxs and not real_subset and not real_sigs:
            # Cliente viejo: idxs mandados sin firmas. Hard-refresh requerido.
            flash("Pantalla desactualizada — recargá con Ctrl+Shift+R.", "error")
        elif (real_idxs or real_sigs) and not real_subset:
            flash(
                "No encontré los movs de banco seleccionados (¿ya conciliados?). "
                "Recargá la página.",
                "error",
            )
        elif hist_ids and not n_hist:
            flash("Históricos no encontrados. Recargá la página.", "error")
        elif err_msg:
            flash(f"No se pudo conciliar: {err_msg}", "error")
        else:
            flash("Sin cambios.", "warn")
    return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id, tab="manual"))


# ─── Endpoint: Agregar movimiento en el programa por la diferencia ────
# TMT 2026-07-14 (dueña): cuando una selección manual NO cierra SOLO porque
# el banco tiene un residuo (todo PC matcheó), se ofrece crear el mov de
# programa faltante (NC si el banco tiene crédito de más, ND si débito) para
# que la diferencia cierre y se concilie todo. Patrón preview→confirm (mirror
# del flujo de impuestos). El preview lo renderiza banco_manual_confirmar; acá
# se re-valida y SOLO se crea si la simulación cierra completa (anti-huérfano).


def _fecha_diferencia(banco_movs, sobrantes_idx):
    """Fecha del mov a crear: la MÁS RECIENTE de las líneas sobrantes del
    banco; si no hay, la máxima fecha de toda la selección; fallback hoy (EC).
    """
    from filters import today_ec

    def _fecha_de(idx):
        try:
            return banco_movs[idx][0].fecha
        except Exception:
            return None

    fechas_sobr = [f for f in (_fecha_de(i) for i in sobrantes_idx) if f]
    if fechas_sobr:
        return max(fechas_sobr)
    fechas_all = [m.fecha for (m, _md) in banco_movs if getattr(m, "fecha", None)]
    if fechas_all:
        return max(fechas_all)
    return today_ec()


@conciliacion_bp.route("/banco-v2/diferencia/confirmar", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_diferencia_confirmar():
    """Crea el movimiento de programa que cierra la diferencia de una selección
    manual con residuo LIMPIO del lado banco, y luego concilia todo.

    Guardas (dinero-crítico):
      - re-valida que la selección siga siendo un residuo limpio (estado pudo
        cambiar entre el preview y el submit) → si no, NO crea nada.
      - gate >$300: exige el checkbox de confirmación.
      - SIMULATE-FIRST: simula agregar el mov y solo crea si la simulación
        reconcilia COMPLETA (evita movimientos huérfanos).
    """
    import bank_helpers as _bh_dif

    sesion_id = int(request.form.get("sesion_id") or 0)
    sesion = _sesion.sesion_por_id(sesion_id) if sesion_id else None
    if not sesion or sesion.get("cerrada_en"):
        flash("Sesión inválida o cerrada.", "error")
        return redirect(url_for("conciliacion.hub"))

    _volver = redirect(url_for("conciliacion.banco_post_procesar",
                               sesion_id=sesion_id, tab="manual"))

    # (b) Re-parsear la selección con el mismo núcleo que el manual.
    sel = _seleccion_firmada(sesion, request.form)
    if sel["error"] or not sel["banco_movs"] or not sel["bancsis_ids"]:
        flash("No pude recuperar la selección — recargá y volvé a intentar.", "error")
        return _volver

    banco_firmados = sel["banco_firmados"]
    prog_firmados = sel["prog_firmados"]
    banco_movs = sel["banco_movs"]
    bancsis_ids = list(sel["bancsis_ids"])

    # (c) Re-correr asignación/reconciliación y re-validar la propuesta.
    asign, sobrantes_idx = asignar_banco_a_programa(
        banco_firmados, prog_firmados, tol=0.50,
    )
    completa, pcs_sin_banco = reconciliacion_completa(
        asign, sobrantes_idx, bancsis_ids,
    )
    if completa:
        # Ya cierra sin agregar nada (estado cambió). No crear.
        flash("La selección ya cierra sola — no hace falta agregar nada. "
              "Volvé a conciliar normalmente.", "warn")
        return _volver
    prop = _proponer_movimiento_diferencia(
        banco_firmados, prog_firmados, sobrantes_idx, pcs_sin_banco,
    )
    if not prop:
        flash("La selección ya no es una diferencia limpia de banco (algo "
              "cambió). No se creó ningún movimiento — revisá la selección.",
              "warn")
        return _volver

    dif = prop["diferencia"]
    documento = prop["documento"]
    importe = prop["importe"]

    # (d) Gate >$300: exige el checkbox.
    if prop["requiere_confirmar"]:
        chk = (request.form.get("confirmar_grande") or "").strip().lower()
        if chk not in ("on", "1", "true", "yes"):
            flash(
                f"Marcá la casilla de confirmación (la diferencia es "
                f"${importe:,.2f} > ${prop['umbral_confirmar']:,.0f}).",
                "warn",
            )
            return _volver

    # (e) Concepto obligatorio.
    concepto = (request.form.get("concepto") or "").strip()
    prov = (request.form.get("prov") or "").strip() or None
    if not concepto:
        flash("Escribí un concepto para el movimiento a crear.", "warn")
        return _volver

    # (f) SIMULATE-FIRST: signed del mov nuevo = dif (NC:+, ND:−). Simulamos
    # agregarlo al lado programa; solo seguimos si la simulación cierra.
    sim_prog = list(prog_firmados) + [("__NEW__", dif)]
    sim_asign, sim_sobr = asignar_banco_a_programa(
        banco_firmados, sim_prog, tol=0.50,
    )
    sim_completa, _sim_pcs = reconciliacion_completa(
        sim_asign, sim_sobr, [pid for pid, _s in sim_prog],
    )
    if not sim_completa:
        flash("No cierra ni agregando el movimiento — revisá la selección. "
              "No se creó nada.", "error")
        return _volver

    # (g) Crear el movimiento por Bancos normal (puede tener side effects).
    from modules.bancos.queries import crear_movimiento_simple

    fecha_mov = _fecha_diferencia(banco_movs, sobrantes_idx)
    try:
        result = crear_movimiento_simple(
            no_banco=_BANCO_PICHINCHA,
            documento=documento,
            importe=importe,
            fecha=fecha_mov,
            concepto=concepto,
            prov=prov,
            usuario=_usuario_actual(),
        )
    except Exception as e:
        _LOG.exception("diferencia: crear_movimiento_simple falló: %s", e)
        flash(f"No pude crear el movimiento: {e}", "error")
        return _volver

    new_id = result.get("id_transaccion")
    side = result.get("side_effect") or result.get("side")
    if not new_id:
        _LOG.critical("diferencia: crear_movimiento_simple sin id_transaccion: %r", result)
        flash("Se intentó crear el movimiento pero no devolvió id — revisá "
              "Bancos antes de reintentar.", "error")
        return _volver

    # (h) Reconstruir con el mov nuevo incluido y conciliar todo.
    try:
        signed_new = float(_bh_dif._signed_delta(documento, importe, _usuario_actual()))
    except Exception:
        signed_new = dif
    bancsis_ids.append(int(new_id))
    prog_firmados_2 = list(prog_firmados) + [(int(new_id), signed_new)]

    asign2, sobr2 = asignar_banco_a_programa(
        banco_firmados, prog_firmados_2, tol=0.50,
    )
    completa2, _pcs2 = reconciliacion_completa(asign2, sobr2, bancsis_ids)

    usuario = _usuario_actual()
    doc_txt = "NC" if documento == "NC" else "ND"
    if completa2:
        import uuid as _uuid_dif
        n_ok = 0
        for pc_id, banco_idxs in asign2.items():
            grupo_batch = _uuid_dif.uuid4().hex
            for idx in banco_idxs:
                mov_i, metodo_i = banco_movs[idx]
                try:
                    confirmar_match(_BANCO_PICHINCHA, mov_i, pc_id,
                                    usuario=usuario, metodo=metodo_i,
                                    confirm_batch_id=grupo_batch)
                    n_ok += 1
                except Exception as e:
                    _LOG.warning("diferencia confirm PC %s idx %s fallo: %s", pc_id, idx, e)
        _sesion.incrementar_matches(sesion_id, n_ok)
        msg = (f"Creé el movimiento de ${importe:,.2f} ({doc_txt}) y concilié "
               f"todo ({n_ok} match(es)).")
        if side:
            msg += f" {side}"
        flash(msg, "ok")
    else:
        # No debería pasar (la simulación cerró) — pero el mov YA se creó.
        _LOG.critical(
            "diferencia: mov %s creado pero la reconciliación NO cerró "
            "(sobr=%d). Requiere revisión manual.", new_id, len(sobr2),
        )
        flash(
            f"⚠ Creé el movimiento de ${importe:,.2f} ({doc_txt}, id {new_id}) "
            f"pero la conciliación NO cerró como se esperaba — revisá el match "
            f"a mano en Manual.",
            "error",
        )
    return _volver


# ─── Endpoint Tab Impuestos ───────────────────────────────────────────


@conciliacion_bp.route("/banco-v2/impuestos/preview", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_impuestos_preview():
    """TMT 2026-06-03 dueña: 'que me de el mismo tipo de confirmacion que
    cuando hago manual, no una confirmacion como ok no ok, que muestre uno
    a uno y los cambios que van a haber'.
    Antes de crear la tx agrupada de impuestos, muestra un preview con
    item-por-item + saldo antes/después + botón confirmar.
    """
    sesion_id = int(request.form.get("sesion_id") or 0)
    sesion = _sesion.sesion_por_id(sesion_id) if sesion_id else None
    if not sesion or sesion.get("cerrada_en"):
        flash("Sesión inválida o cerrada.", "error")
        return redirect(url_for("conciliacion.hub"))

    try:
        real_idxs = [int(x) for x in (request.form.get("real_idxs") or "").split(",") if x.strip()]
    except ValueError:
        real_idxs = []
    if not real_idxs:
        flash("No marcaste ningún movimiento.", "warn")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id, tab="impuestos"))

    from modules.conciliacion.matcher_banco import matchear_extracto_banco
    movs = _sesion.cargar_movs(sesion)
    try:
        res = matchear_extracto_banco(movs, no_banco=_BANCO_PICHINCHA)
        real_only = res.real_only or []
    except Exception as e:
        _LOG.exception("re-match para preview impuestos falló: %s", e)
        real_only = []
    real_subset = [real_only[i] for i in real_idxs if 0 <= i < len(real_only)]
    if not real_subset:
        flash("Los movimientos seleccionados ya no existen.", "error")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id, tab="impuestos"))

    fecha_str = (request.form.get("fecha") or "").strip()
    concepto = (request.form.get("concepto") or "").strip()
    prov = (request.form.get("prov") or "").strip()

    # Calcular monto neto del agrupado (signed).
    monto_neto = sum(
        (1 if (r.tipo or '').upper() == 'C' else -1) * float(r.monto or 0)
        for r in real_subset
    )

    # Balance antes/después.
    balance_before = _bp.calcular(_BANCO_PICHINCHA)
    libros_before = float(balance_before.get("saldo") or 0)
    libros_after = round(libros_before + monto_neto, 2)
    # Pendientes banco: real_subset deja de estar pendiente (entra al match).
    pbtc_before = float(balance_before.get("pendientes_banco_total_creditos")
                        or balance_before.get("pendientes_banco_creditos") or 0)
    pbtd_before = float(balance_before.get("pendientes_banco_total_debitos")
                        or balance_before.get("pendientes_banco_debitos") or 0)
    delta_cred = sum(float(r.monto or 0) for r in real_subset if (r.tipo or '').upper() == 'C')
    delta_deb = sum(float(r.monto or 0) for r in real_subset if (r.tipo or '').upper() == 'D')
    pbtc_after = round(pbtc_before - delta_cred, 2)
    pbtd_after = round(pbtd_before - delta_deb, 2)
    pbtn_after = round(pbtc_after - pbtd_after, 2)
    n_banco_after = max(0, int(balance_before.get("n_pendientes_banco_total")
                                or balance_before.get("n_pendientes") or 0) - len(real_subset))

    # Conciliado_live = libros - pend_pc_neto. Pend_pc no cambia (es banco-only).
    saldo_concilio_after = round(
        libros_after - float(balance_before.get("pendientes_conciliar_neto") or 0), 2
    )
    saldo_banco_esperado_after = round(saldo_concilio_after + pbtn_after, 2)

    balance_after = {
        "saldo": libros_after,
        "pendientes_pc_creditos": balance_before.get("pendientes_pc_creditos"),
        "pendientes_pc_debitos": balance_before.get("pendientes_pc_debitos"),
        "pendientes_conciliar_neto": balance_before.get("pendientes_conciliar_neto"),
        "n_pendientes_conciliar": balance_before.get("n_pendientes_conciliar"),
        "saldo_si_concilio_todo": saldo_concilio_after,
        "pendientes_banco_total_creditos": pbtc_after,
        "pendientes_banco_total_debitos": pbtd_after,
        "neto_pendientes_total": pbtn_after,
        "n_pendientes_banco_total": n_banco_after,
        "saldo_banco_esperado": saldo_banco_esperado_after,
    }

    return render_template(
        "conciliacion/banco_v2_impuestos_preview.html",
        sesion=sesion,
        real_subset=real_subset,
        fecha=fecha_str,
        concepto=concepto,
        prov=prov,
        monto_neto=monto_neto,
        balance_before=balance_before,
        balance_after=balance_after,
        real_idxs_csv=",".join(str(i) for i in real_idxs),
    )


@conciliacion_bp.route("/banco-v2/impuestos/confirmar", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_impuestos_confirmar():
    """Toma los reals de impuestos seleccionados → crea UNA tx BANCSIS
    agrupada y concilia los N reals contra ese mismo id_transaccion.
    """
    sesion_id = int(request.form.get("sesion_id") or 0)
    sesion = _sesion.sesion_por_id(sesion_id) if sesion_id else None
    if not sesion or sesion.get("cerrada_en"):
        flash("Sesión inválida o cerrada.", "error")
        return redirect(url_for("conciliacion.hub"))

    try:
        real_idxs = [int(x) for x in (request.form.get("real_idxs") or "").split(",") if x.strip()]
    except ValueError:
        real_idxs = []
    if not real_idxs:
        flash("No marcaste ningún movimiento.", "warn")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id, tab="impuestos"))

    # CRITICAL FIX 2026-05-29: re-correr el matcher para que los idxs
    # apunten a res.real_only (filtrado), NO a la lista cruda del extracto.
    # Antes movs[i] devolvía un mov COMPLETAMENTE DISTINTO (depósitos por
    # $15K) en lugar del impuesto de \$0.05 → suma 67K en lugar de 14.
    from modules.conciliacion.matcher_banco import matchear_extracto_banco
    movs = _sesion.cargar_movs(sesion)
    try:
        res = matchear_extracto_banco(movs, no_banco=_BANCO_PICHINCHA)
        real_only = res.real_only or []
    except Exception as e:
        _LOG.exception("re-match para confirmar impuestos falló: %s", e)
        real_only = []
    real_subset = [real_only[i] for i in real_idxs if 0 <= i < len(real_only)]
    if not real_subset:
        flash("Los movimientos seleccionados ya no existen en la sesión.", "error")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id, tab="impuestos"))

    # Sanity check: si la suma supera $1000 lo más probable es que la
    # dueña seleccionó cosas grandes por error (default checked roto).
    # Pedimos confirmación adicional en backend levantando warning.
    total_signed = sum(float(r.monto) for r in real_subset if (r.tipo or '').upper()=='C') \
                 - sum(float(r.monto) for r in real_subset if (r.tipo or '').upper()=='D')
    if abs(total_signed) > 1000:
        _LOG.warning("Impuestos confirmar con neto inusual: %.2f n=%d", total_signed, len(real_subset))

    fecha_str = (request.form.get("fecha") or "").strip()
    concepto = (request.form.get("concepto") or "").strip() or None
    prov = (request.form.get("prov") or "").strip() or None
    try:
        fecha = date.fromisoformat(fecha_str) if fecha_str else None
    except ValueError:
        fecha = None

    try:
        result = crear_transaccion_agrupada_desde_reals(
            no_banco=_BANCO_PICHINCHA,
            reals=real_subset,
            fecha=fecha,
            concepto=concepto,
            prov=prov,
            usuario=_usuario_actual(),
        )
        n_matches = int(result.get("n_matches") or len(real_subset))
        _sesion.incrementar_matches(sesion_id, n_matches)
        flash(
            f"Movimiento agrupado creado por ${result.get('monto_neto', 0):,.2f}. "
            f"{n_matches} match(es) conciliados.",
            "ok",
        )
    except Exception as e:
        _LOG.exception("impuestos confirmar falló: %s", e)
        flash(f"Error al crear el movimiento agrupado: {e}", "error")

    return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id, tab="impuestos"))


# ─── Endpoint Tab Transferencias ──────────────────────────────────────


@conciliacion_bp.route("/banco-v2/transferencias/confirmar", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_transferencias_confirmar():
    """Confirma N pares de la PASS 0. Cada par_idx referencia la posición
    en buckets['transferencias']. Re-corremos el matcher (idempotente) y
    aceptamos los pares marcados.
    """
    sesion_id = int(request.form.get("sesion_id") or 0)
    sesion = _sesion.sesion_por_id(sesion_id) if sesion_id else None
    if not sesion or sesion.get("cerrada_en"):
        flash("Sesión inválida o cerrada.", "error")
        return redirect(url_for("conciliacion.hub"))

    try:
        par_idxs = [int(x) for x in (request.form.get("par_idxs") or "").split(",") if x.strip()]
    except ValueError:
        par_idxs = []
    if not par_idxs:
        flash("No marcaste ningún par.", "warn")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id, tab="transferencias"))

    buckets = _sesion.estado_sesion(sesion, _BANCO_PICHINCHA)
    matches = buckets.get("transferencias") or []
    if not matches:
        flash("Los matches ya no están disponibles — algo cambió desde que cargaste la página.", "warn")
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id, tab="transferencias"))

    n_done = 0
    usuario = _usuario_actual()
    for i in par_idxs:
        if i < 0 or i >= len(matches):
            continue
        m = matches[i]
        try:
            confirmar_match(
                no_banco=_BANCO_PICHINCHA,
                real=m.real,
                id_transaccion=m.bancsis.id_transaccion,
                estado="matched",
                usuario=usuario,
                metodo="matched_auto",
            )
            # TMT 2026-06-23: si el par vino de un pendiente histórico (doc-match
            # del backlog), marcarlo conciliado para que salga del backlog.
            _id_hist = getattr(m.real, "id_historico", None)
            if _id_hist:
                _db.execute(
                    """
                    UPDATE scintela.banco_historicos_pendientes
                       SET conciliado_en = CURRENT_TIMESTAMP, conciliado_por = %s
                     WHERE id = %s AND no_banco = %s AND conciliado_en IS NULL
                    """,
                    (usuario[:50], int(_id_hist), _BANCO_PICHINCHA),
                )
            n_done += 1
        except Exception as e:
            _LOG.warning("transferencia confirm falló: %s", e)

    _sesion.incrementar_matches(sesion_id, n_done)
    flash(f"{n_done} transferencia(s) conciliada(s) por documento.", "ok")
    return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id, tab="transferencias"))


# ─── XLSX resumen de pendientes ───────────────────────────────────────
# TMT 2026-06-02 dueña: 'pero si quiero el resumen que imprimiamos!!'
# El XLSX que se generaba al cerrar sesión sigue disponible como descarga
# on-demand vía /banco-v2/resumen-xlsx. Genera el listado actual de
# pendientes (histos + extract no-matcheado) con el resumen contable
# al pie. No cierra ni promueve nada.


def _generar_xlsx_pendientes(sesion: dict, balance: dict) -> str | None:
    """Genera un XLSX formato hoja FEB con los DEPÓSITOS PENDIENTES.

    TMT 2026-05-29 dueña pasó el formato esperado: lista completa de
    banco_historicos_pendientes (todos los movs del banco que NO se
    pudieron conciliar contra el programa), ordenados por fecha asc,
    con columnas FECHA / DETALLE / CODIGO / VALOR / DETALLE-extra.

    Antes el XLSX traía solo los movs del extracto de la sesión actual
    (real_only del bucket Manual) — eso era el subconjunto chico de
    los pendientes del DÍA, no el listado completo del backlog histórico.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        _LOG.warning("openpyxl no instalado — saltando export")
        return None

    no_banco = sesion.get("no_banco") or _BANCO_PICHINCHA

    # Listado completo de pendientes del banco. ORDER BY fecha ASC para
    # que arranquen los más viejos primero (igual al formato Tamara
    # subió: marzo → abril → mayo).
    rows: list[dict] = []
    try:
        rows = _db.fetch_all(
            """
            SELECT fecha, concepto, documento, monto, oficina, detalle, tipo
              FROM scintela.banco_historicos_pendientes
             WHERE no_banco = %s
               AND conciliado_en IS NULL
             ORDER BY fecha ASC, id ASC
            """,
            (no_banco,),
        ) or []
    except Exception as e:
        _LOG.warning("xlsx historicos query falló: %s", e)

    # TMT 2026-07-15 (dueña: "muestra valores que ya fueron conciliados"):
    # excluir los históricos que YA están conciliados por un match ACTIVO
    # (banco_conciliacion_match, deshecho_en IS NULL) aunque su conciliado_en no
    # se haya marcado — pasa cuando una sesión nueva re-incluye un histórico
    # viejo cuyo match sigue vivo. Así el export coincide con lo que la pantalla
    # muestra como conciliado (mismo criterio que balance_pichincha). Clave por
    # (fecha, documento, |monto|) — el documento del banco es único por mov.
    _match_firmas: set = set()
    try:
        _mr = _db.fetch_all(
            """
            SELECT real_fecha, real_documento, real_monto
              FROM scintela.banco_conciliacion_match
             WHERE no_banco = %s AND deshecho_en IS NULL
               AND real_documento IS NOT NULL AND real_documento <> ''
            """,
            (no_banco,),
        ) or []
        for _m in _mr:
            _doc = (_m.get("real_documento") or "").strip()
            if not _doc:
                continue
            try:
                _mo = round(abs(float(_m.get("real_monto") or 0)), 2)
            except (TypeError, ValueError):
                _mo = None
            _match_firmas.add((str(_m.get("real_fecha"))[:10], _doc, _mo))
    except Exception as _e_mf:  # noqa: BLE001
        _LOG.warning("xlsx: no pude cargar match_firmas: %s", _e_mf)

    def _ya_conciliado_por_match(r) -> bool:
        _doc = str(r.get("documento") or "").strip()
        if not _doc:
            return False
        try:
            _mo = round(abs(float(r.get("monto") or 0)), 2)
        except (TypeError, ValueError):
            _mo = None
        return (str(r.get("fecha"))[:10], _doc, _mo) in _match_firmas

    if _match_firmas:
        _antes = len(rows)
        rows = [r for r in rows if not _ya_conciliado_por_match(r)]
        if len(rows) != _antes:
            _LOG.info("xlsx pendientes: excluidos %d histos ya conciliados por match activo",
                      _antes - len(rows))

    # TMT 2026-06-04: el export ya NO incluye el extracto crudo de la sesión
    # (era la misma inflación que sacamos del balance). Pendientes = la hoja
    # (banco_historicos_pendientes); el extracto se usa para cruzar en pantalla.
    rows.sort(key=lambda r: (r.get("fecha") or date.min, str(r.get("documento") or "")))

    # TMT 2026-06-17 (dueña Tamara): se RETIRA la separación "CARGOS DEL BANCO".
    # TODO pendiente del banco (incluidos ISD-PAG, comisiones, SENAE) va en UNA
    # sola lista de pendientes y la diferencia se muestra como un solo número.
    # La separación anterior mezclaba el ISD (impuesto de un pago al exterior,
    # apareado con su principal INTELACI-EXT) con fees y, al restarlo del AJUSTE
    # y re-sumarlo como línea aparte, era NEUTRA en la diferencia: solo ruido.
    rows_reales = list(rows)

    # TMT 2026-06-16 (dueña: "deberia sumarse a pendientes, no en dos secciones
    # ... la diferencia deberia llegar a casi 0"): el extracto nuevo sin cruzar
    # (los "N del extracto a cruzar" del 13-16/06) ES parte de los pendientes
    # del banco. Lo FUSIONAMOS en la misma lista (no sección aparte) y sumamos
    # sus créditos/débitos al AJUSTE para que la DIFERENCIA cierre. estado_sesion
    # ya deduplica el extracto contra la hoja (firma fecha+doc+monto+tipo), así
    # que no se cuenta doble.
    _xt_cred = 0.0
    _xt_deb = 0.0
    # TMT 2026-06-17: el extracto sin cruzar (manual_banco + impuestos) se fusiona
    # COMPLETO en la lista única de pendientes y suma al AJUSTE. Antes los
    # "impuestos" (SENAE, IVA, comisiones, ISD-PAG) iban a una sección CARGOS
    # aparte; ahora van en la misma lista para que la dueña los vea como un
    # pendiente más y la diferencia quede en un solo número.
    try:
        _bk_xt = _sesion.estado_sesion(sesion, no_banco)
        for _bucket in ("manual_banco", "impuestos"):
            for _it in (_bk_xt.get(_bucket) or []):
                if _it.get("es_historico") or _it.get("mov") is None:
                    continue
                _mv = _it["mov"]
                _tp = (getattr(_mv, "tipo", "C") or "C").upper()
                _mo = float(getattr(_mv, "monto", 0) or 0)
                _row_ext = {
                    "fecha": getattr(_mv, "fecha", None),
                    "concepto": getattr(_mv, "concepto", "") or "",
                    "documento": getattr(_mv, "documento", "") or "",
                    "monto": _mo,
                    "tipo": _tp,
                    "oficina": getattr(_mv, "oficina", "") or "",
                    "detalle": getattr(_mv, "oficina", "") or "",
                }
                rows_reales.append(_row_ext)
                if _tp == "C":
                    _xt_cred += _mo
                else:
                    _xt_deb += _mo

        rows_reales.sort(key=lambda rr: (rr.get("fecha") or date.min, str(rr.get("documento") or "")))
    except Exception as _e_xt:
        _LOG.warning("resumen: no pude fusionar extracto sin cruzar: %s", _e_xt)
    _xt_cred = round(_xt_cred, 2)
    _xt_deb = round(_xt_deb, 2)

    # TMT 2026-06-29 dueña: en el resumen descargado, PRIMERO los POSITIVOS
    # (créditos, tipo C) ordenados de fecha más antigua a más actual, y al
    # FINAL todos los NEGATIVOS (débitos, tipo D), también por fecha asc.
    # tipo C → grupo 0 (arriba); cualquier otro (D) → grupo 1 (abajo).
    def _orden_resumen(rr: dict):
        es_credito = (rr.get("tipo") or "C").upper() == "C"
        return (0 if es_credito else 1,
                rr.get("fecha") or date.min,
                str(rr.get("documento") or ""))
    rows_reales.sort(key=_orden_resumen)

    _PDF_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = _PDF_DIR / f"sesion_{sesion['id']}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "MOVIMIENTOS PENDIENTES"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")

    ws["A1"] = "MOVIMIENTOS PENDIENTES"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    ws["A2"] = (
        f"Pichincha · Sesión #{sesion['id']} · "
        f"{_hora_ec_str(datetime.now(), '%Y-%m-%d %H:%M')} · "
        f"{len(rows_reales)} movs"
    )
    ws.merge_cells("A2:E2")

    headers = ["FECHA", "DETALLE", "CODIGO", "VALOR", "DETALLE"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = bold
        c.fill = header_fill

    # BUG #10 fix 2026-05-29: defensa contra CSV-injection.
    # Si un concepto del banco empieza con =, +, -, @, Excel lo evalúa
    # como fórmula al abrir. Prefijamos con apóstrofe para neutralizar.
    def _safe_cell(v: str) -> str:
        s = str(v or "")
        if s and s[0] in ("=", "+", "-", "@", "\t", "\r"):
            return "'" + s
        return s

    # TMT 2026-05-29 dueña: 'i am not seeing some of the expenses' (PAGO
    # SENAE, INTELA C-PAG-*, CHEQUE DEVUELTO). El filtro hard-coded
    # `tipo='C'` saltaba TODOS los débitos. Sacado: ahora se listan
    # créditos Y débitos juntos. Débitos se renderean con paréntesis
    # (convención contable) usando el number_format '#,##0.00;(#,##0.00)'.
    r = 5
    total = 0.0
    n_no_identif = 0
    for row in rows_reales:
        tipo = (row.get("tipo") or "C").upper()
        monto = float(row.get("monto") or 0)
        valor = monto if tipo == "C" else -monto  # negativos para débitos
        total += valor
        fecha = row.get("fecha")
        concepto = row.get("concepto") or ""
        # TMT 2026-06-17: badge "⚠ NO IDENT." en la columna E para filas
        # con "DEPOSITO NO IDENTIFICADO" — Tamara los identificó como
        # parte de la diferencia y quiere distinguirlos a simple vista.
        det_extra = row.get("detalle") or row.get("oficina") or ""
        if "NO IDENTIFICADO" in concepto.upper():
            det_extra = (det_extra + " ⚠ NO IDENT.").strip()
            n_no_identif += 1
        ws.cell(row=r, column=1, value=fecha.strftime("%d/%m/%Y") if fecha else "")
        ws.cell(row=r, column=2, value=_safe_cell(concepto)[:100])
        ws.cell(row=r, column=3, value=_safe_cell(row.get("documento"))[:30])
        ws.cell(row=r, column=4, value=valor).number_format = "+#,##0.00;-#,##0.00;0.00"
        ws.cell(row=r, column=5, value=_safe_cell(det_extra)[:30])
        r += 1

    # ── Resumen contable al pie ───────────────────────────────────────
    # TMT 2026-05-29 dueña — decisión final: layout viejo de 5 filas con
    # DIFERENCIA explícita. Razones:
    #   - Math cierra por construcción (AJUSTE + SISTEMA = TOTAL conciliado)
    #   - SALDO BANCO se imprime aparte como referencia externa
    #   - DIFERENCIA expone honestamente la brecha real-vs-cálculo
    #   - Si DIFERENCIA es chica (centavos) es ruido aceptable
    #   - Si DIFERENCIA es grande (>$1000) → auditar es el next step
    #
    # Layout final:
    #   AJUSTE              = pendientes_banco_creditos − pendientes_banco_debitos
    #   SALDO SISTEMA       = saldo conciliado (libros − pendientes_pc_neto)
    #   TOTAL               = SISTEMA + AJUSTE (cierra)
    #   SALDO BANCO         = último saldo del extracto subido
    #   DIFERENCIA          = SALDO BANCO − TOTAL (cero si todo conciliado)
    saldo_sistema = float(
        balance.get("saldo_si_concilio_todo")
        or balance.get("saldo") or 0
    )
    # TMT 2026-06-16: el AJUSTE ahora incluye también el extracto nuevo sin
    # cruzar (_xt_cred/_xt_deb), no solo la hoja — así la DIFERENCIA cierra.
    pendientes_banco_cred = round(float(balance.get("pendientes_banco_creditos") or 0) + _xt_cred, 2)
    pendientes_banco_deb = round(float(balance.get("pendientes_banco_debitos") or 0) + _xt_deb, 2)

    # Incluir real_only de la sesión actual (movs del extracto sin matchear)
    # para que el AJUSTE refleje TODO lo que cierra contra banco.
    saldo_banco_real = None
    # TMT 2026-06-26 (dueña: "el saldo que cambio a mano no se mantiene al
    # descargar el excel"). El resumen ignoraba el saldo_banco_objetivo manual
    # y siempre auto-detectaba por max(fecha) — distinto de la pantalla, que
    # prioriza el valor que la dueña escribe. Espejamos esa prioridad: 1ro el
    # manual de la sesión, 2do el auto-detect.
    saldo_manual = sesion.get("saldo_banco_objetivo")
    saldo_detectado = sesion.get("saldo_banco_detectado")
    if saldo_manual is not None:
        try:
            saldo_banco_real = float(saldo_manual)
        except (TypeError, ValueError):
            saldo_banco_real = None
    # 2do: saldo real del extracto capturado al subir (robusto al dedup).
    if saldo_banco_real is None and saldo_detectado is not None:
        try:
            saldo_banco_real = float(saldo_detectado)
        except (TypeError, ValueError):
            saldo_banco_real = None
    if saldo_banco_real is None:
        try:
            movs_sesion = _sesion.cargar_movs(sesion)
            if movs_sesion:
                # TMT 2026-06-04: ya no sumamos el extracto a pendientes (AJUSTE =
                # solo histos reales, vía balance). Solo leemos el saldo objetivo.
                con_fecha = [
                    m for m in movs_sesion
                    if getattr(m, "fecha", None) and getattr(m, "saldo", None) is not None
                ]
                if con_fecha:
                    ult = max(con_fecha, key=lambda m: m.fecha)
                    saldo_banco_real = float(ult.saldo)
                else:
                    ult = movs_sesion[-1]
                    saldo_banco_real = float(getattr(ult, "saldo", None) or 0) or None
        except Exception as e:
            _LOG.warning("no pude leer movs sesion: %s", e)

    ajuste = round(pendientes_banco_cred - pendientes_banco_deb, 2)
    # TMT 2026-06-17: TODO pendiente (hoja + extracto sin cruzar, incl. ISD/
    # comisiones/SENAE) está en AJUSTE. TOTAL = SISTEMA + AJUSTE y la
    # DIFERENCIA = SALDO BANCO − TOTAL es el único número a cerrar.
    total_calc = round(saldo_sistema + ajuste, 2)
    diferencia = round(saldo_banco_real - total_calc, 2) if saldo_banco_real is not None else None

    label_col = 3
    val_col = 4

    r += 1  # fila vacía de separación
    # TMT 2026-05-29 dueña: 'puedes separar en ajuste los positivos y los
    # negativos?'. Ajuste = banco_cred − banco_deb; mostramos las dos
    # componentes y el neto antes del SISTEMA.
    rows_resumen = [
        ("+ Pendientes banco créditos", pendientes_banco_cred),
        ("− Pendientes banco débitos", -pendientes_banco_deb),
        ("AJUSTE", ajuste),
        ("SALDO SISTEMA (conciliado)", saldo_sistema),
        ("TOTAL", total_calc),
    ]
    if saldo_banco_real is not None:
        rows_resumen.append(("SALDO BANCO (extracto)", saldo_banco_real))
        rows_resumen.append(("DIFERENCIA", diferencia))

    for label, val in rows_resumen:
        ws.cell(row=r, column=label_col, value=label).font = bold
        if val is not None:
            cell = ws.cell(row=r, column=val_col, value=val)
            cell.font = bold
            # Signo + / − explícito, mismo formato que las filas de movs.
            cell.number_format = "+#,##0.00;-#,##0.00;0.00"
        else:
            ws.cell(row=r, column=val_col, value="—").font = bold
        r += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12

    wb.save(str(xlsx_path))
    return str(xlsx_path)


@conciliacion_bp.route("/banco-v2/resumen-xlsx", methods=["GET"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_resumen_xlsx():
    """Genera y descarga el resumen XLSX de movimientos pendientes.

    TMT 2026-06-02 dueña: 'quiero el resumen que imprimiamos'. Reemplaza
    el viejo flujo "Terminar y guardar" → XLSX. Ahora es un download
    on-demand: NO cierra sesión, NO promueve nada a histos, solo genera
    el listado actual (histos sin conciliar + extracto no matcheado) con
    el resumen contable al pie. Se puede pedir cuantas veces se quiera.
    """
    no_banco = _BANCO_PICHINCHA
    sesion = _sesion.sesion_abierta(no_banco)
    if not sesion:
        flash("No hay sesión abierta — subí un extracto primero.", "info")
        return redirect(url_for("conciliacion.hub"))
    balance = _bp.calcular(no_banco)
    try:
        path = _generar_xlsx_pendientes(sesion, balance)
    except Exception as e:
        _LOG.exception("resumen xlsx falló: %s", e)
        flash(f"No pude generar el resumen: {e}", "error")
        return redirect(url_for("conciliacion.banco_post_procesar"))
    if not path:
        flash("openpyxl no disponible — no se pudo generar.", "error")
        return redirect(url_for("conciliacion.banco_post_procesar"))
    p = Path(path)
    if not p.is_absolute():
        p = Path.cwd() / p
    if not p.exists():
        abort(404)
    fname = f"conciliacion_pendientes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        str(p),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


# ─── Historial ────────────────────────────────────────────────────────


@conciliacion_bp.route("/banco-v2/recompute-saldos", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_recompute_saldos():
    """Recalcula la cadena de saldos completa desde la fila más temprana.

    Útil cuando se detecta descalce (por bugs históricos en
    crear/anular/insert). Walk-forward desde el primer mov del banco
    en Pichincha. Idempotente: si la cadena ya está OK, no cambia nada.
    """
    pre = _verificar_cadena_saldos(_BANCO_PICHINCHA)
    try:
        import bank_helpers
        with _db.tx() as conn:
            # TMT 2026-05-29 dueña: 'busca esos 9k de diferencia'. Bug encontrado:
            # recompute_saldos_desde filtra por id_transaccion >= ancla_id, pero
            # los movs no están en orden estricto por id (imports legacy de dBase
            # crearon filas con fecha tardía e id más bajo). Las filas con
            # fecha > ancla.fecha pero id < ancla.id quedaban FUERA del walk.
            # Fix: pasar ancla_fecha en lugar de ancla_id → filtra por fecha,
            # que sí es chronological y captura todas las filas.
            primera = _db.fetch_one(
                """
                SELECT fecha FROM scintela.transacciones_bancarias
                 WHERE no_banco = %s AND fecha IS NOT NULL
                 ORDER BY fecha ASC LIMIT 1
                """,
                (_BANCO_PICHINCHA,),
                conn=conn,
            )
            if not primera or not primera.get("fecha"):
                flash("Sin movimientos para recalcular.", "info")
                return redirect(url_for("conciliacion.hub"))
            n = bank_helpers.recompute_saldos_desde(
                conn, no_banco=_BANCO_PICHINCHA, no_cta=None,
                ancla_fecha=primera["fecha"],
            )
    except Exception as e:
        _LOG.exception("recompute manual falló: %s", e)
        flash(f"Recalculo falló: {e}", "error")
        return redirect(url_for("conciliacion.hub"))

    post = _verificar_cadena_saldos(_BANCO_PICHINCHA)
    if pre.get("ok") and post.get("ok"):
        flash(
            f"✓ Cadena ya estaba coherente. Último saldo: ${post.get('ultimo_saldo', 0):,.2f}",
            "ok",
        )
    elif post.get("ok"):
        flash(
            f"✓ Cadena recalculada. Encontré {len(pre.get('problemas', []))} "
            f"descalces previos. Último saldo: ${post.get('ultimo_saldo', 0):,.2f}",
            "ok",
        )
    else:
        flash(
            f"⚠ Cadena recalculada pero siguen quedando {len(post.get('problemas', []))} "
            f"discrepancias. Revisar /banco-v2/verificar-saldos.",
            "warn",
        )
    return redirect(url_for("conciliacion.hub"))


@conciliacion_bp.route("/banco-v2/verificar-saldos", methods=["GET"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_verificar_saldos():
    """Devuelve JSON con el estado de la cadena (para auditoría o UI)."""
    from flask import jsonify
    return jsonify(_verificar_cadena_saldos(_BANCO_PICHINCHA, limit=100))


@conciliacion_bp.route("/banco-v2/borrar-sesion", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_borrar_sesion():
    """Borra UNA sesión (la fila) + snapshots + matches + txs creadas
    durante esa sesión.

    Menos destructivo que reset-all: solo afecta UNA sesión a la vez.
    El borrar va por rango temporal (abierta_en..cerrada_en) — todos
    los matches y txs grupales creados dentro de ese rango se borran.
    Los pendientes históricos que se hayan conciliado en esa sesión
    vuelven a aparecer como pendientes.
    """
    try:
        sesion_id = int(request.form.get("sesion_id") or 0)
    except (TypeError, ValueError):
        sesion_id = 0
    if sesion_id <= 0:
        flash("ID de sesión inválido.", "error")
        return redirect(url_for("conciliacion.banco_historial_v2"))

    sesion = _sesion.sesion_por_id(sesion_id)
    if not sesion:
        flash("Sesión no encontrada.", "warn")
        return redirect(url_for("conciliacion.banco_historial_v2"))

    # TMT 2026-06-02 dueña: 'poneme boton anular conciliaciones y borrar'.
    # Sesión abierta también se puede borrar — usamos NOW() como cota
    # superior si no hay cerrada_en. Anula matches y limpia todo.
    usuario = _usuario_actual()
    abierta = sesion.get("abierta_en")
    cerrada = sesion.get("cerrada_en")
    if not cerrada:
        from datetime import datetime
        cerrada = datetime.now(UTC)
    counts = {"matches": 0, "snapshots": 0, "txs_grupales": 0, "historicos_reset": 0}

    try:
        with _db.tx() as conn:
            # 1. Firmas del extracto de ESTA sesión = identidad real de sus
            # movimientos de banco (fecha+doc+monto+tipo). Solo tocaremos los
            # matches cuyo lado real pertenece a ESTA sesión.
            #
            # TMT 2026-07-15 (dueña + Alex — bug importación INV30405): BUG DE
            # RAÍZ. Antes el borrado de matches iba por VENTANA DE TIEMPO
            # (creado_en entre abierta y cerrada). Como los matches NO guardan a
            # qué sesión pertenecen, borrar UNA sesión se llevaba puestos TODOS
            # los matches creados en esa ventana — incluidos los de OTRAS
            # conciliaciones legítimas (ej. un cruce multi-PC hecho por Alex
            # desaparecía entero, sin rastro). Y era HARD DELETE: no quedaba ni
            # el `deshecho` para auditar, por eso los movimientos reaparecían
            # como pendientes sin explicación. Ahora: (a) atamos el borrado por
            # FIRMA del extracto de la sesión (no por tiempo), y (b) usamos
            # SOFT-DELETE (deshecho_en) para que sea REVERSIBLE y quede rastro
            # de qué sesión lo quitó.
            movs_sesion = _sesion.cargar_movs(sesion) or []
            firmas_sesion: set = set()
            for _mv in movs_sesion:
                _doc = (getattr(_mv, "documento", "") or "").strip().upper()
                if not _doc:
                    continue
                try:
                    _mo = f"{abs(float(getattr(_mv, 'monto', 0) or 0)):.2f}"
                except (TypeError, ValueError):
                    continue
                _fe = getattr(_mv, "fecha", None)
                _fe = _fe.isoformat() if hasattr(_fe, "isoformat") else str(_fe or "")
                _ti = (getattr(_mv, "tipo", "") or "").strip().upper()[:1]
                firmas_sesion.add((_fe, _doc, _mo, _ti))
            if not firmas_sesion:
                _LOG.warning(
                    "borrar-sesion #%s: extracto sin firmas recuperables — no se "
                    "tocan matches (conservador, evita borrar de otras sesiones).",
                    sesion_id,
                )

            # 2. Candidatos: matches ACTIVOS creados en la ventana de la sesión.
            # De esos, SOLO los que además tienen firma real EN el extracto de
            # esta sesión son realmente de esta sesión (la ventana sola no basta
            # — ese era el bug). Separamos: matches contra tx PC real (soft-
            # delete, recuperable) vs. matches de tx fabricada por conciliación
            # (hard-delete, porque la tx se borra en el paso 4).
            cand_matches = _db.fetch_all(
                """
                SELECT id, id_transaccion, real_fecha, real_documento,
                       real_monto, real_tipo, COALESCE(metodo, '') AS metodo
                  FROM scintela.banco_conciliacion_match
                 WHERE no_banco = %s
                   AND deshecho_en IS NULL
                   AND creado_en >= %s - interval '5 seconds'
                   AND creado_en <= %s + interval '60 seconds'
                """,
                (_BANCO_PICHINCHA, abierta, cerrada),
                conn=conn,
            ) or []
            match_ids_soft: list = []
            match_ids_hard: list = []
            ids_bancsis_sesion: list = []
            ids_grupales: list = []
            _seen_bancsis: set = set()
            _seen_grup: set = set()
            for _m in cand_matches:
                _doc = (_m.get("real_documento") or "").strip().upper()
                if not _doc:
                    # Sin firma real (p.ej. bancsis_only_ok): no lo podemos atar
                    # a esta sesión con certeza → NO lo tocamos (conservador).
                    continue
                try:
                    _mo = f"{abs(float(_m.get('real_monto') or 0)):.2f}"
                except (TypeError, ValueError):
                    continue
                _fe = str(_m.get("real_fecha"))[:10]
                _ti = (_m.get("real_tipo") or "").strip().upper()[:1]
                if (_fe, _doc, _mo, _ti) not in firmas_sesion:
                    continue  # el mov de banco NO es de esta sesión → no tocar
                _met = _m.get("metodo") or ""
                _idtx = _m.get("id_transaccion")
                if _met in ("created_from_real", "created_from_real_grouped"):
                    match_ids_hard.append(int(_m["id"]))
                    if _idtx and int(_idtx) not in _seen_grup:
                        _seen_grup.add(int(_idtx))
                        ids_grupales.append(int(_idtx))
                else:
                    match_ids_soft.append(int(_m["id"]))
                    if _idtx and int(_idtx) not in _seen_bancsis:
                        _seen_bancsis.add(int(_idtx))
                        ids_bancsis_sesion.append(int(_idtx))
            _match_ids_todos = match_ids_soft + match_ids_hard

            # 3a. Reset históricos conciliados POR ESTOS matches (por
            # conciliado_match_id), no por ventana de tiempo. Vuelven a
            # pendientes solo los históricos que ESTA sesión cruzó.
            counts["historicos_reset"] = 0
            if _match_ids_todos:
                counts["historicos_reset"] = _db.execute(
                    """
                    UPDATE scintela.banco_historicos_pendientes
                       SET conciliado_en = NULL,
                           conciliado_por = NULL,
                           conciliado_match_id = NULL
                     WHERE no_banco = %s
                       AND conciliado_match_id = ANY(%s)
                    """,
                    (_BANCO_PICHINCHA, _match_ids_todos),
                    conn=conn,
                ) or 0

            # 3b. Borrar histos que ESTA sesión auto-promovió (fuente=sesion:N).
            counts["historicos_promovidos_borrados"] = _db.execute(
                """
                DELETE FROM scintela.banco_historicos_pendientes
                 WHERE no_banco = %s
                   AND fuente = %s
                   AND conciliado_en IS NULL
                """,
                (_BANCO_PICHINCHA, f"sesion:{sesion_id}"),
                conn=conn,
            ) or 0

            # 3c. Soft-delete (REVERSIBLE) los matches contra tx PC real de esta
            # sesión; hard-delete solo los de tx fabricada (la tx se borra en el
            # paso 4). `deshecho_por` deja el rastro de qué sesión los quitó, así
            # nunca más desaparece un match sin explicación.
            counts["matches"] = 0
            _des_por = (f"borrar-sesion:{sesion_id}")[:50]
            if match_ids_soft:
                counts["matches"] += _db.execute(
                    """
                    UPDATE scintela.banco_conciliacion_match
                       SET deshecho_en = CURRENT_TIMESTAMP,
                           deshecho_por = %s
                     WHERE id = ANY(%s) AND deshecho_en IS NULL
                    """,
                    (_des_por, match_ids_soft),
                    conn=conn,
                ) or 0
            if match_ids_hard:
                counts["matches"] += _db.execute(
                    "DELETE FROM scintela.banco_conciliacion_match WHERE id = ANY(%s)",
                    (match_ids_hard,),
                    conn=conn,
                ) or 0

            # 3d. Reset stat='*' de los BANCSIS que matcheaban con esta sesión,
            # salvo que tengan OTRO match activo (no romper otra conciliación).
            counts["bancsis_stat_reset"] = 0
            if ids_bancsis_sesion:
                counts["bancsis_stat_reset"] = _db.execute(
                    """
                    UPDATE scintela.transacciones_bancarias
                       SET stat = NULL
                     WHERE no_banco = %s
                       AND id_transaccion = ANY(%s)
                       AND TRIM(COALESCE(stat, '')) = '*'
                       AND NOT EXISTS (
                           SELECT 1 FROM scintela.banco_conciliacion_match m
                            WHERE m.id_transaccion = scintela.transacciones_bancarias.id_transaccion
                              AND m.deshecho_en IS NULL
                       )
                    """,
                    (_BANCO_PICHINCHA, ids_bancsis_sesion),
                    conn=conn,
                ) or 0

            # 4. Borrar txs BANCSIS grupales + recompute LOCAL.
            # TMT 2026-06-03 dueña: 'fijate que borrar sesion nos vuelva a
            # dejar en el mismo punto de partida'. Bug: el recompute partía
            # del PRIMER mov del banco (id_ascendente), re-walking 23k filas
            # y arrastrando cualquier desviación histórica del saldo column.
            # Resultado: libros podía desplazarse $147K después de un borrar.
            # Fix: recompute SOLO desde la fecha más vieja de las txs que
            # acabamos de borrar. El walk se limita a ~30-50 filas, no toca
            # historia previa.
            if ids_grupales:
                # Capturamos la fecha mínima de las txs grupales ANTES de
                # borrarlas, así sabemos desde dónde recalcular.
                # TMT 2026-06-03 FIX drift libros al borrar: anclar el recompute por
                # ID, no por fecha. Las txs creadas por conciliacion tienen los
                # ids mas altos (creadas hoy), asi que anclar por id hace que el
                # walk toque SOLO esas filas y arranque del saldo de la ultima
                # fila DBF (autoritativa). Anclar por fecha re-derivaba toda la
                # jornada sumando importes -> drift de 493K vs el saldo DBF.
                ancla_id_borrar = min(int(x) for x in ids_grupales)
                counts["txs_grupales"] = _db.execute(
                    "DELETE FROM scintela.transacciones_bancarias WHERE id_transaccion = ANY(%s) AND no_banco = %s",
                    (ids_grupales, _BANCO_PICHINCHA),
                    conn=conn,
                ) or 0
                try:
                    import bank_helpers
                    bank_helpers.recompute_saldos_desde(
                        conn, no_banco=_BANCO_PICHINCHA, no_cta=None,
                        ancla_id=ancla_id_borrar,
                    )
                except Exception as e:
                    # TMT 2026-06-03: NO silenciar — antes el except log+pass
                    # escondía drift de saldos cuando recompute fallaba (ej.
                    # por bug de _signed_delta). La dueña veía "borrar OK"
                    # pero libros quedaba wrong. Ahora escalamos al outer
                    # except → flash error visible.
                    _LOG.exception("recompute_saldos en borrar-sesion falló: %s", e)
                    raise RuntimeError(
                        f"Borrar OK pero recompute de saldos falló: {e}. "
                        f"Libros puede estar inconsistente. Hacer sync dBase."
                    ) from e

            # 5. Borrar snapshots de esta sesión.
            # TMT 2026-06-03 audit fix: el OR temporal arrasaba snapshots de
            # OTRAS sesiones que hubieran cerrado en la misma ventana de
            # tiempo (también snapshots reset_total, snapshots de impuestos,
            # etc.). Filtramos solo por evento_ref = sesion_id. Si algún
            # snapshot quedó huérfano del rango pero sin evento_ref, mejor
            # dejarlo — es paper trail, no afecta funcionamiento.
            counts["snapshots"] = _db.execute(
                """
                DELETE FROM scintela.banco_saldo_conc_snapshot
                 WHERE no_banco = %s
                   AND evento_ref = %s
                """,
                (_BANCO_PICHINCHA, str(sesion_id)),
                conn=conn,
            ) or 0

            # 6. Borrar la fila de sesión.
            _db.execute(
                "DELETE FROM scintela.banco_conciliacion_sesion WHERE id = %s AND no_banco = %s",
                (sesion_id, _BANCO_PICHINCHA),
                conn=conn,
            )

            # 7. Reset stat='*' en movs PC linkeados a ESTA sesión.
            # TMT 2026-06-03 audit fix: antes este UPDATE arrasaba TODO mov
            # PC con stat='*' sin match activo, lo que apagaba el flag de
            # movs vinculados a OTRAS sesiones cerradas. Restringimos al
            # set de ids capturados en el paso 4 (matches que se acaban de
            # borrar) — solo afecta lo que esta sesión tocó.
            if ids_bancsis_sesion:
                _db.execute(
                    """
                    UPDATE scintela.transacciones_bancarias
                       SET stat = NULL
                     WHERE no_banco = %s
                       AND id_transaccion = ANY(%s)
                       AND TRIM(COALESCE(stat,'')) = '*'
                       AND COALESCE(usuario_crea,'') NOT IN ('dbf-import','asinfo-backfill')
                       AND NOT EXISTS (
                           SELECT 1 FROM scintela.banco_conciliacion_match m
                            WHERE m.id_transaccion = scintela.transacciones_bancarias.id_transaccion
                              AND m.deshecho_en IS NULL
                       )
                    """,
                    (_BANCO_PICHINCHA, ids_bancsis_sesion),
                    conn=conn,
                )

        # 8. Borrar el archivo XLSX si existe.
        if sesion.get("pdf_path"):
            try:
                path = Path(sesion["pdf_path"])
                if not path.is_absolute():
                    path = Path.cwd() / path
                if path.exists():
                    path.unlink()
            except Exception:
                pass

        _LOG.warning("BORRAR SESIÓN #%s por %s: %s", sesion_id, usuario, counts)
        flash(
            f"✓ Sesión #{sesion_id} borrada. "
            f"{counts['matches']} matches, "
            f"{counts.get('bancsis_stat_reset', 0)} BANCSIS desmarcados, "
            f"{counts['snapshots']} snapshots, "
            f"{counts['txs_grupales']} txs, "
            f"{counts['historicos_reset']} histo reseteados.",
            "ok",
        )
    except Exception as e:
        _LOG.exception("borrar sesión #%s falló: %s", sesion_id, e)
        flash(f"Borrar falló: {e}", "error")

    return redirect(url_for("conciliacion.banco_historial_v2"))


@conciliacion_bp.route("/banco-v2/reset-all", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_reset_all():
    """⚠️ NUCLEAR — borra TODAS las conciliaciones para arrancar de cero.

    TMT 2026-05-29 dueña: 'podes borrar todas las conciliaciones, quizas
    ponele un boton con 5 alertas por las dudas, asi hoy arrancamos de cero'.

    Hace, en una sola transacción atómica:
      1. DELETE de banco_conciliacion_match (matches + deshechos).
      2. DELETE de banco_saldo_conc_snapshot (snapshots).
      3. DELETE de banco_conciliacion_sesion (sesiones).
      4. DELETE de transacciones_bancarias creadas por conciliación
         (created_from_real / created_from_real_grouped).
      5. UPDATE banco_historicos_pendientes SET conciliado_en=NULL
         (vuelven a pendientes).
      6. UPDATE transacciones_bancarias SET stat=NULL WHERE stat='*'
         AND usuario_crea NOT IN ('dbf-import','asinfo-backfill')
         (movs PC vuelven a pendientes; los del DBF original mantienen
         su stat porque ya estaban conciliados antes del programa core).
      7. Snapshot 'reset_total' del nuevo saldo a conciliar.

    Frontend exige 5 confirmaciones (window.confirm) + 1 prompt textual
    'BORRAR TODO' antes de llegar acá.

    Requiere POST con `confirm_text=BORRAR TODO` para evitar accidentes
    si alguien dispara el endpoint sin pasar por el modal.
    """
    # BUG #11 fix 2026-05-29: comparación estricta, sin strip/upper —
    # exige el texto exacto y previene curl con espacios extra.
    confirm_text = request.form.get("confirm_text") or ""
    if confirm_text != "BORRAR TODO":
        flash("Reset cancelado — texto de confirmación incorrecto.", "error")
        return redirect(url_for("conciliacion.hub"))

    usuario = _usuario_actual()
    counts = {"matches": 0, "snapshots": 0, "sesiones": 0,
              "txs_grupales": 0, "historicos_reset": 0, "stat_reset": 0}
    try:
        with _db.tx() as conn:
            # 1. Ids de txs creadas por conciliación (para borrar tx + matches).
            ids_rows = _db.fetch_all(
                """
                SELECT DISTINCT t.id_transaccion
                  FROM scintela.transacciones_bancarias t
                  JOIN scintela.banco_conciliacion_match m
                       ON m.id_transaccion = t.id_transaccion
                 WHERE t.no_banco = %s
                   AND m.metodo IN ('created_from_real','created_from_real_grouped')
                """,
                (_BANCO_PICHINCHA,),
                conn=conn,
            ) or []
            ids_grupales = [r["id_transaccion"] for r in ids_rows if r.get("id_transaccion")]

            # 2a. Borrar todos los histos promovidos por sesiones del programa
            # (fuente LIKE 'sesion:%'). Los originales del banco se conservan.
            counts["historicos_promovidos_borrados"] = _db.execute(
                """
                DELETE FROM scintela.banco_historicos_pendientes
                 WHERE no_banco = %s AND fuente LIKE 'sesion:%%'
                """,
                (_BANCO_PICHINCHA,),
                conn=conn,
            ) or 0

            # 2b. Reset históricos pendientes originales del banco
            # (vuelven a no-conciliados).
            counts["historicos_reset"] = _db.execute(
                """
                UPDATE scintela.banco_historicos_pendientes
                   SET conciliado_en = NULL,
                       conciliado_por = NULL,
                       conciliado_match_id = NULL
                 WHERE no_banco = %s
                   AND conciliado_en IS NOT NULL
                """,
                (_BANCO_PICHINCHA,),
                conn=conn,
            ) or 0

            # 3. Borrar TODOS los matches (activos + deshechos).
            counts["matches"] = _db.execute(
                "DELETE FROM scintela.banco_conciliacion_match WHERE no_banco = %s",
                (_BANCO_PICHINCHA,),
                conn=conn,
            ) or 0

            # 4. Borrar TODAS las sesiones.
            counts["sesiones"] = _db.execute(
                "DELETE FROM scintela.banco_conciliacion_sesion WHERE no_banco = %s",
                (_BANCO_PICHINCHA,),
                conn=conn,
            ) or 0

            # 5. Borrar TODOS los snapshots (perderemos el histórico pero
            # es esperable cuando se hace 'arrancar de cero').
            counts["snapshots"] = _db.execute(
                "DELETE FROM scintela.banco_saldo_conc_snapshot WHERE no_banco = %s",
                (_BANCO_PICHINCHA,),
                conn=conn,
            ) or 0

            # 6. Borrar las txs BANCSIS que fueron creadas por conciliación.
            if ids_grupales:
                counts["txs_grupales"] = _db.execute(
                    "DELETE FROM scintela.transacciones_bancarias WHERE id_transaccion = ANY(%s) AND no_banco = %s",
                    (ids_grupales, _BANCO_PICHINCHA),
                    conn=conn,
                ) or 0
                # Recompute saldos desde el inicio de los borrados.
                try:
                    import bank_helpers
                    siguiente = _db.fetch_one(
                        """
                        SELECT id_transaccion FROM scintela.transacciones_bancarias
                         WHERE no_banco = %s
                         ORDER BY fecha ASC, id_transaccion ASC LIMIT 1
                        """,
                        (_BANCO_PICHINCHA,),
                        conn=conn,
                    )
                    if siguiente and siguiente.get("id_transaccion"):
                        bank_helpers.recompute_saldos_desde(
                            conn, no_banco=_BANCO_PICHINCHA, no_cta=None,
                            ancla_id=int(siguiente["id_transaccion"]),
                        )
                except Exception as e:
                    _LOG.warning("recompute_saldos en reset falló: %s", e)

            # 7. Reset stat='*' en movs PC que fueron marcados conciliados
            # por el flujo PC (NO los del dbf-import original).
            counts["stat_reset"] = _db.execute(
                """
                UPDATE scintela.transacciones_bancarias
                   SET stat = NULL
                 WHERE no_banco = %s
                   AND TRIM(COALESCE(stat,'')) = '*'
                   AND COALESCE(usuario_crea,'') NOT IN ('dbf-import','asinfo-backfill')
                """,
                (_BANCO_PICHINCHA,),
                conn=conn,
            ) or 0

        # Snapshot del nuevo saldo a conciliar (fuera de la tx).
        try:
            from modules.conciliacion import saldo_snapshot as _ss
            _ss.snapshot(
                _BANCO_PICHINCHA, "reset_total",
                evento_ref="reset", usuario=usuario,
                descripcion=f"reset total: {counts}",
            )
        except Exception:
            pass

        _LOG.warning("RESET TOTAL por %s: %s", usuario, counts)
        flash(
            f"✓ Reset completo. {counts['matches']} matches, "
            f"{counts['sesiones']} sesiones, {counts['snapshots']} snapshots, "
            f"{counts['txs_grupales']} txs, {counts['historicos_reset']} histo, "
            f"{counts['stat_reset']} stat reseteados.",
            "ok",
        )
    except Exception as e:
        _LOG.exception("RESET falló: %s", e)
        flash(f"Reset falló: {e}", "error")

    return redirect(url_for("conciliacion.hub"))


@conciliacion_bp.route("/banco-v2/cruzar", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_cruzar_pendientes():
    """Cruza un archivo de DEPOSITOS PENDIENTES (hoja FEB2023 prevalece)
    contra el backlog del sistema (banco_historicos_pendientes), por doc+monto.

    TMT 2026-06-03 duena: 'asegurate que prevalezcan y arranque desde ahi yo'.
    APLICA DIRECTO: el archivo prevalece -> borra del backlog los que sobran
    (no estan en el archivo), agrega los que faltan, corrige montos. Toca solo
    banco_historicos_pendientes con conciliado_en IS NULL. Redirige con resumen.
    """
    f = request.files.get("archivo")
    if not f or not f.filename:
        flash("Falta el archivo de pendientes.", "error")
        return redirect(url_for("conciliacion.hub"))
    raw = f.read()
    if not raw:
        flash("El archivo vino vacio.", "error")
        return redirect(url_for("conciliacion.hub"))
    try:
        from modules.conciliacion.parser_xlsx import parse_pendientes_cruce
        hoja, items, dropped = parse_pendientes_cruce(raw, return_dropped=True)
    except Exception as e:
        _LOG.exception("parse cruce fallo: %s", e)
        flash(f"No pude leer el archivo: {e}", "error")
        return redirect(url_for("conciliacion.hub"))
    if not items:
        flash("El archivo no trajo pendientes reconocibles (revisa la hoja FEB2023).", "warn")
        return redirect(url_for("conciliacion.hub"))
    # 2026-06-19: filas con VALOR que NO se pudo parsear ya NO se tiran en
    # silencio. Las avisamos por fila para que se corrija el archivo (era la
    # causa de "no carga el mov del 17 / la acreditación de 967K").
    if dropped:
        detalle = "; ".join(
            f"fila {d['fila']}: {d['concepto'][:30]} = «{d['valor_raw']}»"
            for d in dropped[:8]
        )
        mas = f" (+{len(dropped) - 8} más)" if len(dropped) > 8 else ""
        flash(
            f"⚠ {len(dropped)} fila(s) del archivo NO se cargaron porque no pude "
            f"leer su VALOR — revisalas en el Excel: {detalle}{mas}",
            "error",
        )

    sis_rows = _db.fetch_all(
        """
        SELECT id, fecha, documento, concepto, monto, tipo
          FROM scintela.banco_historicos_pendientes
         WHERE no_banco = %s AND conciliado_en IS NULL
        """,
        (_BANCO_PICHINCHA,),
    ) or []
    sis_por_doc: dict = {}
    for r in sis_rows:
        doc = str(r.get("documento") or "").strip()
        if not doc:
            continue
        signed = float(r.get("monto") or 0) * (1 if (r.get("tipo") or "").upper() == "C" else -1)
        sis_por_doc.setdefault(doc, []).append({
            "id": r.get("id"), "monto": round(signed, 2),
            "concepto": r.get("concepto") or "", "fecha": r.get("fecha"),
        })

    file_docs = set()
    coinciden, monto_distinto, faltan = [], [], []
    for it in items:
        doc = it["doc"]
        file_docs.add(doc)
        sis = sis_por_doc.get(doc)
        if not sis:
            faltan.append(it)
        elif any(abs(s["monto"] - it["monto"]) <= 0.01 for s in sis):
            coinciden.append(it)
        else:
            monto_distinto.append({**it, "sistema_monto": sis[0]["monto"]})
    sobran = []
    for doc, rows in sis_por_doc.items():
        if doc not in file_docs:
            for s in rows:
                sobran.append({"doc": doc, "monto": s["monto"],
                               "detalle": (s["concepto"] or "")[:60], "id": s["id"]})

# TMT 2026-06-03 duena: 'no me hace falta chequear, asegurate que
    # prevalezcan y arranque desde ahi yo'. Aplicamos directo: el archivo
    # FEB2023 prevalece -> el backlog del sistema queda igual a el.
    n_del = n_add = n_upd = 0
    try:
        with _db.tx() as conn:
            sobran_ids = [s["id"] for s in sobran if s.get("id")]
            if sobran_ids:
                n_del = _db.execute(
                    "DELETE FROM scintela.banco_historicos_pendientes "
                    "WHERE id = ANY(%s) AND no_banco = %s AND conciliado_en IS NULL",
                    (sobran_ids, _BANCO_PICHINCHA), conn=conn,
                ) or 0
            saltados = []
            for it in faltan:
                tipo = "C" if it["monto"] >= 0 else "D"
                # ON CONFLICT DO NOTHING choca contra ux_bhp_firma
                # (no_banco, fecha, documento, monto, tipo). Antes hacíamos
                # n_add+=1 igual → un mov que NO entró se reportaba "agregado"
                # y desaparecía sin aviso (ej. una acreditación cuya firma ya
                # existía). Ahora contamos lo REAL y avisamos lo saltado.
                ins = _db.execute(
                    "INSERT INTO scintela.banco_historicos_pendientes "
                    "(no_banco, fecha, concepto, documento, monto, tipo, fuente, creado_en) "
                    "VALUES (%s, %s, %s, %s, %s, %s, 'cruce-feb2023', CURRENT_TIMESTAMP) "
                    "ON CONFLICT DO NOTHING",
                    (_BANCO_PICHINCHA, it.get("fecha"), (it.get("detalle") or "")[:200],
                     it["doc"], abs(it["monto"]), tipo), conn=conn,
                ) or 0
                if ins:
                    n_add += 1
                else:
                    saltados.append(it)
            for it in monto_distinto:
                tipo = "C" if it["monto"] >= 0 else "D"
                n_upd += _db.execute(
                    "UPDATE scintela.banco_historicos_pendientes SET monto = %s, tipo = %s "
                    "WHERE documento = %s AND no_banco = %s AND conciliado_en IS NULL",
                    (abs(it["monto"]), tipo, it["doc"], _BANCO_PICHINCHA), conn=conn,
                ) or 0
    except Exception as e:
        _LOG.exception("aplicar cruce fallo: %s", e)
        flash(f"Error al aplicar el cruce: {e}", "error")
        return redirect(url_for("conciliacion.hub"))

    flash(
        f"Pendientes actualizados desde {hoja}: {n_del} borrado(s), {n_add} agregado(s), "
        f"{n_upd} corregido(s). El backlog ahora coincide con tu archivo ({len(items)} items).",
        "ok",
    )
    # Avisar lo que el archivo traía como NUEVO pero NO entró (firma ya
    # existente). Para una acreditación que "no refleja como las demás" esto
    # apunta exactamente a la fila culpable.
    if saltados:
        det = "; ".join(
            f"{(s.get('detalle') or '')[:24]} {s['monto']:+,.2f} doc {s['doc'] or '—'}"
            for s in saltados[:6]
        )
        mas = f" (+{len(saltados) - 6} más)" if len(saltados) > 6 else ""
        flash(
            f"⚠ {len(saltados)} mov(s) del archivo NO se agregaron: ya existía una "
            f"fila con la misma firma (fecha+doc+monto+tipo). Si esperabas verlos, "
            f"revisá si están duplicados o ya conciliados: {det}{mas}",
            "warn",
        )
    # Control de cuadre: Σ neto de lo parseado vs lo que el archivo declara.
    # Si no coinciden, algo se dejó de leer (otra forma de "encontrar más").
    neto_parseado = round(sum(it["monto"] for it in items), 2)
    _LOG.info(
        "cruce hoja=%s items=%d neto_parseado=%.2f faltan=%d saltados=%d "
        "monto_distinto=%d sobran=%d",
        hoja, len(items), neto_parseado, len(faltan), len(saltados),
        len(monto_distinto), len(sobran),
    )
    return redirect(url_for("conciliacion.hub"))


@conciliacion_bp.route("/banco-v2/eliminar-pendiente", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_eliminar_pendiente():
    """Borra UNA fila del backlog de pendientes del banco (hard delete).

    TMT 2026-06-19 (dueña): 'no podemos tener filas fantasmas que se
    agreguen'. El "DEPOSITO NO IDENTIFICADO" (322,72) vivía solo en
    banco_historicos_pendientes, sin documento, y "Hacer prevalecer" NO lo
    podía sacar (el cruce ignora filas sin documento). Esto da una forma
    quirúrgica, por pantalla, de eliminar UN pendiente fantasma sin tocar
    el resto. Solo borra filas no conciliadas del banco Pichincha.
    """
    try:
        hist_id = int(request.form.get("hist_id") or 0)
    except (TypeError, ValueError):
        hist_id = 0
    try:
        sesion_id = int(request.form.get("sesion_id") or 0)
    except (TypeError, ValueError):
        sesion_id = 0
    _back = (
        url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id)
        if sesion_id else url_for("conciliacion.banco_post_procesar")
    )
    if hist_id <= 0:
        flash("No identifiqué el pendiente a eliminar.", "error")
        return redirect(_back)

    # Traer la fila primero para el mensaje (y para no borrar a ciegas).
    row = _db.fetch_one(
        """
        SELECT concepto, documento, monto, tipo
          FROM scintela.banco_historicos_pendientes
         WHERE id = %s AND no_banco = %s AND conciliado_en IS NULL
        """,
        (hist_id, _BANCO_PICHINCHA),
    )
    if not row:
        flash("Ese pendiente ya no existe o ya está conciliado.", "warn")
        return redirect(_back)

    n = _db.execute(
        """
        DELETE FROM scintela.banco_historicos_pendientes
         WHERE id = %s AND no_banco = %s AND conciliado_en IS NULL
        """,
        (hist_id, _BANCO_PICHINCHA),
    ) or 0
    if n:
        _LOG.warning(
            "ELIMINAR pendiente #%s por %s: %s %s %.2f",
            hist_id, _usuario_actual(), row.get("concepto"),
            row.get("documento") or "—", float(row.get("monto") or 0),
        )
        flash(
            f"Pendiente eliminado: {(row.get('concepto') or '')[:40]} "
            f"${float(row.get('monto') or 0):,.2f}.",
            "ok",
        )
        try:
            from modules.conciliacion import saldo_snapshot as _ss
            _ss.snapshot(_BANCO_PICHINCHA, "pendiente_eliminado",
                         evento_ref=hist_id, usuario=_usuario_actual(),
                         descripcion=f"eliminar pendiente #{hist_id}")
        except Exception:
            pass
    else:
        flash("No pude eliminar el pendiente.", "error")
    return redirect(_back)


@conciliacion_bp.route("/banco-v2/historial", methods=["GET"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_historial_v2():
    """Listado simple de sesiones para auditoría.

    TMT 2026-06-02: sesión no se cierra → la única fila 'activa' por
    banco es la sesión vigente. El resto son sesiones cerradas (legacy
    pre-mig-0062). Sin columnas de saldo inicial/final/diferencia: eso
    se ve live en la pantalla principal.
    """
    r = _migracion_lista_o_redirect()
    if r: return r
    sesiones = _sesion.listar_sesiones(no_banco=_BANCO_PICHINCHA, limit=200)
    return render_template(
        "conciliacion/banco_v2_historial.html",
        sesiones=sesiones,
    )


# ─── Reabrir sesión cerrada ──────────────────────────────────────────


@conciliacion_bp.route("/banco-v2/reabrir-sesion", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_reabrir_sesion():
    """Reabre una sesión cerrada.

    TMT 2026-06-04 dueña: 'abilita reabrir'. Como el modelo permite UNA sola
    sesión abierta por banco, primero cierra la que esté abierta (sin promover
    nada al backlog — la hoja es la verdad de los pendientes) y después reabre
    la elegida. Redirige a la pantalla activa con la sesión reabierta.
    """
    no_banco = _BANCO_PICHINCHA
    sesion_id = int(request.form.get("sesion_id") or 0)
    if not sesion_id:
        flash("Falta la sesión a reabrir.", "error")
        return redirect(url_for("conciliacion.banco_historial_v2"))
    sesion = _sesion.sesion_por_id(sesion_id)
    if not sesion:
        flash("Sesión no encontrada.", "error")
        return redirect(url_for("conciliacion.banco_historial_v2"))
    if not sesion.get("cerrada_en"):
        return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))
    try:
        with _db.tx() as conn:
            _db.execute(
                """
                UPDATE scintela.banco_conciliacion_sesion
                   SET cerrada_en = CURRENT_TIMESTAMP, cerrada_por = %s
                 WHERE no_banco = %s AND cerrada_en IS NULL AND id <> %s
                """,
                (f"auto-al-reabrir-{sesion_id}"[:50], no_banco, sesion_id),
                conn=conn,
            )
            _db.execute(
                """
                UPDATE scintela.banco_conciliacion_sesion
                   SET cerrada_en = NULL, cerrada_por = NULL
                 WHERE id = %s AND no_banco = %s
                """,
                (sesion_id, no_banco),
                conn=conn,
            )
    except Exception as e:
        _LOG.exception("reabrir sesion %s falló: %s", sesion_id, e)
        flash(f"Error al reabrir la sesión: {e}", "error")
        return redirect(url_for("conciliacion.banco_historial_v2"))
    flash(
        f"Sesión #{sesion_id} reabierta. Cerré la que estaba abierta para dejar "
        f"solo esta (sin tocar los pendientes).",
        "ok",
    )
    return redirect(url_for("conciliacion.banco_post_procesar", sesion_id=sesion_id))


# ─── Cerrar sesión + migrar pendientes a histos (lifecycle mensual) ──


@conciliacion_bp.route("/banco-v2/cerrar-sesion", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_cerrar_sesion():
    """Cierra la sesión activa. NO migra nada.

    TMT 2026-06-04 dueña: 'no puede no estar en extracto y no estar en
    histórico'. Antes el cierre promovía el extracto sin parear a
    banco_historicos_pendientes — eso inflaba pendientes de banco (225 /
    −299k) y, al correr "Hacer prevalecer" con la hoja, esos movs (no
    presentes en la hoja) se borraban y desaparecían. Ahora el cierre solo
    marca cerrada_en; los pendientes de banco son SIEMPRE la hoja
    (banco_cruzar_pendientes). El extracto sin conciliar no se pierde: se
    vuelve a ver subiéndolo en una sesión nueva.
    """
    no_banco = _BANCO_PICHINCHA
    usuario = _usuario_actual()
    sesion = _sesion.sesion_abierta(no_banco)
    if not sesion:
        flash("No hay sesión abierta para cerrar.", "warn")
        return redirect(url_for("conciliacion.banco_post_procesar"))

    sesion_id = int(sesion.get("id") or 0)
    # TMT 2026-06-04 dueña: 'no puede no estar en extracto y no estar en
    # histórico'. El cierre YA NO promueve el extracto sin parear a
    # banco_historicos_pendientes. Esa promoción inflaba pendientes de banco
    # (llegó a 225 / −299k) metiendo el día entero del banco como "pendiente",
    # y al correr "Hacer prevalecer" con la hoja esos movs (que no están en la
    # hoja) se borraban → desaparecían. La ÚNICA verdad de los pendientes de
    # banco es la hoja (banco_cruzar_pendientes). El extracto sin conciliar se
    # vuelve a ver subiéndolo de nuevo en una sesión nueva — no se pierde.
    n_migrados = 0

    # Cerrar la sesión.
    try:
        _db.execute(
            """
            UPDATE scintela.banco_conciliacion_sesion
               SET cerrada_en = CURRENT_TIMESTAMP, cerrada_por = %s
             WHERE id = %s AND cerrada_en IS NULL
            """,
            (usuario[:50], sesion_id),
        )
    except Exception as e:
        _LOG.exception("cerrar sesion %s falló: %s", sesion_id, e)
        flash(f"Error al cerrar la sesión: {e}", "error")
        return redirect(url_for("conciliacion.banco_post_procesar"))

    flash(
        f"Sesión #{sesion_id} cerrada. Los pendientes de banco siguen siendo los "
        f"de la hoja; el extracto sin conciliar lo volvés a ver subiéndolo de nuevo.",
        "ok",
    )
    return redirect(url_for("conciliacion.hub"))


# ─── Sacar items de un grupo (sub-batch) ──────────────────────────────


@conciliacion_bp.route("/banco-v2/sacar-del-grupo", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_sacar_del_grupo():
    """Toma N match_ids y los reasigna a un sub-batch_id nuevo.
    El resto del batch original queda intacto. Útil para extraer un par 1:1
    de un grupo más grande para deshacerlo solo."""
    raw = (request.form.get("match_ids") or "").strip()
    if not raw:
        flash("No seleccionaste items.", "warn")
        return redirect(url_for("conciliacion.banco_deshacer_v2"))
    try:
        ids = [int(x) for x in raw.split(",") if x.strip()]
    except (TypeError, ValueError):
        flash("IDs inválidos.", "error")
        return redirect(url_for("conciliacion.banco_deshacer_v2"))
    if not ids:
        flash("No seleccionaste items.", "warn")
        return redirect(url_for("conciliacion.banco_deshacer_v2"))
    import uuid as _uuid
    new_batch = _uuid.uuid4().hex
    try:
        n = _db.execute(
            """
            UPDATE scintela.banco_conciliacion_match
               SET confirm_batch_id = %s
             WHERE id = ANY(%s) AND deshecho_en IS NULL
            """,
            (new_batch, ids),
        ) or 0
    except Exception as e:
        _LOG.warning("sacar-del-grupo falló: %s", e)
        flash("Error al sacar items.", "error")
        return redirect(url_for("conciliacion.banco_deshacer_v2"))
    if n:
        flash(f"{n} item(s) movidos a un sub-grupo nuevo. Podés deshacerlos aparte.", "ok")
    else:
        flash("No se movió nada (¿ya estaban deshechos?).", "warn")
    return redirect(url_for("conciliacion.banco_deshacer_v2"))


# ─── Deshacer conciliados (pantalla minimalista) ──────────────────────


@conciliacion_bp.route("/banco-v2/deshacer", methods=["GET"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_deshacer_v2():
    """Lista de matches activos + movs BANCSIS creados por conciliación.

    TMT 2026-05-29 dueña: 'no veo las dos entradas en 67k que quiero
    deshacer'. Las txs creadas por crear_transaccion_agrupada_desde_reals
    (impuestos/comisiones) son su propio objeto en transacciones_bancarias
    — anular el match no las borra. Pantalla muestra ambos:

      - Matches individuales (real_only conciliados contra un id_tx)
      - Grupos BANCSIS creados por la propia conciliación (las ND/NC
        del flujo de Impuestos): para borrarlos completamente.
    """
    matches = []
    grupos = []
    batches = []
    try:
        # TMT 2026-06-03: JOIN con transacciones_bancarias para traer el LADO
        # programa (que antes mostraba "—"). Agrupamos por confirm_batch_id
        # para que la N:M se vea como UNA conciliación con varios items.
        matches = _db.fetch_all(
            """
            SELECT m.id, m.creado_en, m.real_fecha, m.real_documento,
                   m.real_concepto, m.real_monto, m.real_tipo,
                   m.usuario, m.id_transaccion, m.metodo,
                   m.confirm_batch_id,
                   t.fecha       AS pc_fecha,
                   t.documento   AS pc_documento,
                   t.importe     AS pc_importe,
                   t.concepto    AS pc_concepto,
                   t.prov        AS pc_prov,
                   COALESCE(
                     (SELECT nombre FROM scintela.cliente
                       WHERE codigo_cli = t.prov LIMIT 1), ''
                   ) AS pc_cliente_nombre
              FROM scintela.banco_conciliacion_match m
              LEFT JOIN scintela.transacciones_bancarias t
                     ON t.id_transaccion = m.id_transaccion
             WHERE m.no_banco = %s
               AND m.deshecho_en IS NULL
             ORDER BY m.creado_en DESC
             LIMIT 500
            """,
            (_BANCO_PICHINCHA,),
        ) or []
    except Exception as e:
        _LOG.warning("listar matches activos falló: %s", e)

    # Agrupar por confirm_batch_id (matches con batch_id NULL = individuales).
    from collections import OrderedDict
    by_batch = OrderedDict()
    for m in matches:
        bid = m.get("confirm_batch_id") or f"_solo_{m['id']}"
        if bid not in by_batch:
            by_batch[bid] = {"batch_id": m.get("confirm_batch_id"), "items": [],
                             "creado_en": m.get("creado_en"), "usuario": m.get("usuario")}
        by_batch[bid]["items"].append(m)
    for bid, b in by_batch.items():
        # Cómputo de totales lado banco / lado programa.
        # TMT 2026-06-03: dedup por id_transaccion en PC. Si N matches
        # apuntan a la misma tx PC (ej. agrupado de impuestos), la tx PC
        # se cuenta UNA sola vez.
        sum_banco = 0.0
        sum_programa = 0.0
        n_banco_lados = 0
        n_pc_lados = 0
        seen_tx = set()
        for it in b["items"]:
            if it.get("real_monto") is not None:
                signo = 1 if (it.get("real_tipo") or "").upper() == "C" else -1
                sum_banco += signo * float(it.get("real_monto") or 0)
                n_banco_lados += 1
            tx_id = it.get("id_transaccion")
            if tx_id is not None and tx_id not in seen_tx:
                seen_tx.add(tx_id)
                doc = (it.get("pc_documento") or "").upper()
                imp = float(it.get("pc_importe") or 0)
                signo_pc = 1 if doc in ("DE","TR","NC","IN","AC","XX") else -1
                sum_programa += signo_pc * imp
                n_pc_lados += 1
        b["sum_banco"] = round(sum_banco, 2)
        b["sum_programa"] = round(sum_programa, 2)
        b["n_banco"] = n_banco_lados
        b["n_pc"] = n_pc_lados
        b["delta"] = round(sum_banco - sum_programa, 2)
        batches.append(b)
    # Sort by creado_en DESC.
    batches.sort(key=lambda b: b.get("creado_en") or "", reverse=True)

    try:
        grupos = _db.fetch_all(
            """
            SELECT t.id_transaccion, t.fecha, t.documento, t.importe,
                   t.concepto, t.saldo, t.usuario_crea,
                   COUNT(m.id) AS n_matches_total,
                   COUNT(m.id) FILTER (WHERE m.deshecho_en IS NULL) AS n_matches_activos
              FROM scintela.transacciones_bancarias t
              JOIN scintela.banco_conciliacion_match m ON m.id_transaccion = t.id_transaccion
             WHERE t.no_banco = %s
               AND m.metodo IN ('created_from_real_grouped','created_from_real')
             GROUP BY t.id_transaccion, t.fecha, t.documento, t.importe,
                      t.concepto, t.saldo, t.usuario_crea
            HAVING COUNT(m.id) FILTER (WHERE m.deshecho_en IS NULL) > 0
             ORDER BY t.fecha DESC, t.id_transaccion DESC
             LIMIT 200
            """,
            (_BANCO_PICHINCHA,),
        ) or []
    except Exception as e:
        _LOG.warning("listar grupos BANCSIS falló: %s", e)

    return render_template(
        "conciliacion/banco_v2_deshacer.html",
        matches=matches,
        grupos=grupos,
        batches=batches,
    )


@conciliacion_bp.route("/banco-v2/anular-grupo", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def banco_anular_grupo():
    """Anula UNA tx BANCSIS creada por conciliación (impuestos / agrupado)
    + todos sus matches activos + revierte stat + recompute saldos.

    TMT 2026-05-29 dueña: 'si creamos por impuestos, si deshaces el
    movimiento también tiene que anular esa carga'.
    """
    try:
        id_tx = int(request.form.get("id_transaccion") or 0)
    except (TypeError, ValueError):
        id_tx = 0
    if id_tx <= 0:
        flash("ID de transacción inválido.", "error")
        return redirect(url_for("conciliacion.banco_deshacer_v2"))

    usuario = _usuario_actual()
    # 1) Validar que es realmente una tx de conciliación (no PC normal).
    tx_row = _db.fetch_one(
        """
        SELECT t.id_transaccion, t.fecha, t.documento, t.importe, t.concepto,
               COALESCE(t.usuario_crea, '') AS usuario_crea,
               (SELECT MIN(m.metodo)
                  FROM scintela.banco_conciliacion_match m
                 WHERE m.id_transaccion = t.id_transaccion) AS metodo
          FROM scintela.transacciones_bancarias t
         WHERE t.id_transaccion = %s AND t.no_banco = %s
        """,
        (id_tx, _BANCO_PICHINCHA),
    )
    if not tx_row:
        flash("No se encontró la transacción.", "error")
        return redirect(url_for("conciliacion.banco_deshacer_v2"))
    if (tx_row.get("metodo") or "") not in ("created_from_real", "created_from_real_grouped"):
        flash("Esta tx NO fue creada por conciliación — no la puedo anular desde acá.", "warn")
        return redirect(url_for("conciliacion.banco_deshacer_v2"))

    # Pre-snapshot del último saldo y validación de cadena.
    pre = _verificar_cadena_saldos(_BANCO_PICHINCHA)
    saldo_pre = pre.get("ultimo_saldo")
    delta_esperado = -_signed_delta(
        tx_row.get("documento"),
        float(tx_row.get("importe") or 0),
        tx_row.get("usuario_crea") or "",
    )

    # 2+3) Hard-delete de matches + DELETE de tx + recompute, todo en una
    # sola transacción atómica.
    import bank_helpers
    try:
        with _db.tx() as conn:
            n_matches = _db.execute(
                """
                DELETE FROM scintela.banco_conciliacion_match
                 WHERE id_transaccion = %s
                """,
                (id_tx,),
                conn=conn,
            ) or 0
            n_del = _db.execute(
                "DELETE FROM scintela.transacciones_bancarias WHERE id_transaccion = %s AND no_banco = %s",
                (id_tx, _BANCO_PICHINCHA),
                conn=conn,
            ) or 0
            # Verificación post-delete dentro de la misma tx.
            still = _db.fetch_one(
                "SELECT 1 FROM scintela.transacciones_bancarias WHERE id_transaccion = %s",
                (id_tx,),
                conn=conn,
            )
            if still:
                raise RuntimeError(f"DELETE no efectivo — tx {id_tx} aún existe")

            # Walk-forward desde la siguiente fila.
            siguiente = _db.fetch_one(
                """
                SELECT id_transaccion FROM scintela.transacciones_bancarias
                 WHERE no_banco = %s
                   AND (fecha > %s OR (fecha = %s AND id_transaccion > %s))
                 ORDER BY fecha ASC, id_transaccion ASC LIMIT 1
                """,
                (_BANCO_PICHINCHA, tx_row["fecha"], tx_row["fecha"], id_tx),
                conn=conn,
            )
            if siguiente and siguiente.get("id_transaccion"):
                try:
                    bank_helpers.recompute_saldos_desde(
                        conn,
                        no_banco=_BANCO_PICHINCHA,
                        no_cta=None,
                        ancla_id=int(siguiente["id_transaccion"]),
                    )
                except Exception as e:
                    _LOG.warning("recompute_saldos falló (sigue): %s", e)
            _LOG.info("anular_grupo tx=%s OK: %s matches borrados, %s tx borrada",
                      id_tx, n_matches, n_del)
    except Exception as e:
        _LOG.exception("anular tx %s falló: %s", id_tx, e)
        flash(f"Error al anular: {e}", "error")
        return redirect(url_for("conciliacion.banco_deshacer_v2"))

    # 4) Decrementar el contador matches_hechos de la sesión activa
    # (si existe) — el counter es running total y se descontrolaba al
    # anular. Lo bajamos por el número de matches que efectivamente
    # se borraron en el grupo.
    if n_matches > 0:
        try:
            _db.execute(
                """
                UPDATE scintela.banco_conciliacion_sesion
                   SET matches_hechos = GREATEST(0, matches_hechos - %s)
                 WHERE no_banco = %s
                   AND usuario = %s
                   AND cerrada_en IS NULL
                """,
                (int(n_matches), _BANCO_PICHINCHA, usuario[:50]),
            )
        except Exception:
            pass

    # 5) Snapshot del nuevo saldo a conciliar.
    try:
        from modules.conciliacion import saldo_snapshot as _ss
        _ss.snapshot(
            _BANCO_PICHINCHA, "grupo_anulado",
            evento_ref=str(id_tx), usuario=usuario,
            descripcion=f"anulada tx agrupada #{id_tx} ({n_matches} matches deshechos)",
        )
    except Exception:
        pass

    # 6) Validación post-anular: el saldo debe haber cambiado en el monto
    # esperado (delta_esperado). Si la diferencia con la realidad es
    # mayor al threshold, log CRITICAL — indica corrupción previa o
    # bug en recompute_saldos_desde.
    post = _verificar_cadena_saldos(_BANCO_PICHINCHA)
    saldo_post = post.get("ultimo_saldo")
    if not post.get("ok"):
        _LOG.critical(
            "ANULAR GRUPO tx=%s: cadena CORRUPTA post-anular. Pre OK=%s. "
            "Problemas=%s",
            id_tx, pre.get("ok"), post.get("problemas"),
        )
        flash(
            "⚠ Cadena de saldos descalibrada — corré recompute desde "
            "el botón 'Verificar y recalcular saldos'.",
            "warn",
        )
    elif saldo_pre is not None and saldo_post is not None:
        diff_real = round(saldo_post - saldo_pre, 2)
        if abs(diff_real - delta_esperado) > 0.5:
            _LOG.critical(
                "ANULAR GRUPO tx=%s: saldo cambió %.2f pero el delta "
                "esperado era %.2f (importe=%s, doc=%s). Pre OK=%s, Post OK=%s.",
                id_tx, diff_real, delta_esperado,
                tx_row.get("importe"), tx_row.get("documento"),
                pre.get("ok"), post.get("ok"),
            )
            flash(
                f"⚠ Saldo cambió ${diff_real:+,.2f} (esperado ${delta_esperado:+,.2f}). "
                f"Indica corrupción previa. Usá 'Verificar y recalcular saldos'.",
                "warn",
            )

    flash(
        f"Movimiento agrupado #{id_tx} anulado: tx borrada, "
        f"{n_matches} match(es) deshechos, saldos recalculados.",
        "ok",
    )
    return redirect(url_for("conciliacion.banco_deshacer_v2"))
