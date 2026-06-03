"""Parser de Excel para "DEPÓSITOS PENDIENTES".

TMT 2026-05-20 — la dueña sube un .xlsx con varias hojas, cada una con un mes
de depósitos cargados en sistema que NO aparecen aún en el extracto del banco
(o que el banco no procesó). Formato típico (con leves variaciones por hoja):

    Fila 1: "DEPÓSITOS PENDIENTES"        (header decorativo)
    Fila 2: FECHA | DETALLE/CONCEPTO | CODIGO | VALOR | DETALLE
    Fila 3+: filas con datos

Variantes vistas:
    * "DETALLE" o "CONCEPTO" — ambos OK
    * "FEC HA" (con espacio) — typo común
    * Columna CODIGO a veces falta (hojas viejas)
    * Hoja con datos vacíos / placeholders — ignorar

Devuelve siempre `list[DepositoPendiente]` con la misma forma.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

_LOG = logging.getLogger("programa_core.conciliacion.parser_xlsx")


@dataclass(frozen=True)
class DepositoPendiente:
    """Un depósito que el sistema cree que entró pero el banco no confirmó.

    fecha     : fecha del depósito (cuando se cargó en sistema o en banco)
    concepto  : descripción que aparece en sistema/banco (texto libre)
    codigo    : código/referencia/no. de transacción (string — puede ser largo)
    valor     : monto en US$ (positivo)
    detalle   : nota libre, opcional (puede traer iniciales del cliente)
    hoja      : nombre de la hoja de origen (para trazabilidad)
    """

    fecha: date | None
    concepto: str
    codigo: str
    valor: Decimal
    detalle: str
    hoja: str


def _parse_fecha_celda(v) -> date | None:
    """Acepta datetime, date, '2026-05-20...' o '20/05/2026' string."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or s.startswith("#"):  # #REF!, #N/A, etc.
        return None
    # ISO YYYY-MM-DD (o datetime stringificado)
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:19] if len(s) > 10 else s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_valor(v) -> Decimal:
    if v is None:
        return Decimal(0)
    if isinstance(v, int | float | Decimal):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return Decimal(0)
    s = str(v).strip().replace(" ", "")
    if not s:
        return Decimal(0)
    # 1.234,56 (es) vs 1234.56 (en) — heurística como en parser.py
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal(0)


def _normalizar(s: str) -> str:
    return (
        (s or "")
        .strip()
        .lower()
        .replace(" ", "")
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )


_HEADER_FECHA = {"fecha", "fech", "fechha"}  # incluye 'FEC HA' normalizado
_HEADER_CONCEPTO = {"concepto", "detalle", "descripcion", "observacion"}
_HEADER_CODIGO = {"codigo", "ref", "referencia", "numero", "ndocumento"}
_HEADER_VALOR = {"valor", "monto", "importe", "credito", "haber"}
_HEADER_NOTA = {"detalle2", "nota", "observ", "obs"}  # segundo "DETALLE" si existe


def _detectar_columnas(headers: list[str]) -> dict:
    """Mapea índice de columna por concepto.

    Si una hoja tiene dos columnas "DETALLE" (como FEB2023), la primera se trata
    como 'concepto' y la segunda como 'nota'. Manejamos eso al detectar.
    """
    mapeo = {"fecha": None, "concepto": None, "codigo": None, "valor": None, "nota": None}
    detalles_vistos: list[int] = []
    for i, h in enumerate(headers):
        n = _normalizar(h)
        if not n:
            continue
        if mapeo["fecha"] is None and n in _HEADER_FECHA:
            mapeo["fecha"] = i
        elif n == "detalle":
            detalles_vistos.append(i)
        elif mapeo["concepto"] is None and n in _HEADER_CONCEPTO:
            mapeo["concepto"] = i
        elif mapeo["codigo"] is None and n in _HEADER_CODIGO:
            mapeo["codigo"] = i
        elif mapeo["valor"] is None and n in _HEADER_VALOR:
            mapeo["valor"] = i
        elif mapeo["nota"] is None and n in _HEADER_NOTA:
            mapeo["nota"] = i
    # Si vimos 2 "DETALLE" y no asignamos concepto/nota, asignamos los
    # detalles a (concepto, nota) en orden.
    if detalles_vistos:
        if mapeo["concepto"] is None and detalles_vistos:
            mapeo["concepto"] = detalles_vistos.pop(0)
        if mapeo["nota"] is None and detalles_vistos:
            mapeo["nota"] = detalles_vistos.pop(0)
    return mapeo


def _es_fila_header(celdas: list) -> bool:
    """¿Esta fila contiene los nombres de columna?"""
    textos = [str(c or "").strip().lower() for c in celdas]
    has_fecha = any("fec" in t for t in textos)
    has_valor = any(t in ("valor", "monto", "importe") for t in textos)
    return has_fecha and has_valor


def _es_fila_decorativa(celdas: list) -> bool:
    """¿Esta fila es solo "DEPÓSITOS PENDIENTES" o similar?"""
    textos = [str(c or "").strip().upper() for c in celdas if c]
    if not textos:
        return False
    return any("PENDIENTE" in t or "DEPOSITO" in t and len(textos) == 1 for t in textos)


def parse_xlsx(raw: bytes) -> list[DepositoPendiente]:
    """Parsea bytes de un .xlsx → lista de depósitos.

    Recorre TODAS las hojas. En cada hoja:
      1. Encuentra la fila de headers (busca columnas con FECHA y VALOR).
      2. Parsea las filas siguientes.
      3. Skippea filas vacías o decorativas.

    Filas con fecha o valor inválido se loggean en DEBUG y se ignoran.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl no instalado — pip install openpyxl") from e

    if not raw:
        return []

    bio = io.BytesIO(raw)
    wb = load_workbook(bio, data_only=True, read_only=True)
    salida: list[DepositoPendiente] = []

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        header_row_idx = None
        col_map: dict = {}

        # Encontrar fila de headers — escaneamos hasta las primeras 10 filas
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i > 10:
                break
            row_list = list(row)
            if _es_fila_header(row_list):
                header_row_idx = i
                col_map = _detectar_columnas([str(c or "") for c in row_list])
                break

        if header_row_idx is None or col_map.get("fecha") is None or col_map.get("valor") is None:
            _LOG.debug("Hoja %s: sin headers reconocibles, skipping", nombre_hoja)
            continue

        # Leer filas siguientes
        for _i, row in enumerate(
            ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1
        ):
            row_list = list(row)
            if not any(c is not None and str(c).strip() for c in row_list):
                continue  # fila totalmente vacía
            try:
                fecha = _parse_fecha_celda(row_list[col_map["fecha"]])
                valor = _parse_valor(row_list[col_map["valor"]])
            except IndexError:
                continue
            if valor <= 0:
                continue
            # Filas sin fecha = ruido (TOTAL, SALDO, etc.). Skip.
            if fecha is None:
                continue
            concepto = ""
            if col_map.get("concepto") is not None:
                try:
                    concepto = str(row_list[col_map["concepto"]] or "").strip()
                except IndexError:
                    pass
            codigo = ""
            if col_map.get("codigo") is not None:
                try:
                    codigo = str(row_list[col_map["codigo"]] or "").strip()
                    if codigo.endswith(".0"):  # openpyxl puede dar floats como '15289222.0'
                        codigo = codigo[:-2]
                except IndexError:
                    pass
            nota = ""
            if col_map.get("nota") is not None:
                try:
                    nota = str(row_list[col_map["nota"]] or "").strip()
                except IndexError:
                    pass

            salida.append(
                DepositoPendiente(
                    fecha=fecha,
                    concepto=concepto,
                    codigo=codigo,
                    valor=valor,
                    detalle=nota,
                    hoja=nombre_hoja,
                )
            )

    wb.close()
    return salida


# ─── Cruce de pendientes (subir hoja "DEPÓSITOS PENDIENTES" para validar) ──
# TMT 2026-06-03 dueña: 'dejame subir un archivo para cruzar por las dudas,
# por si algo cambió desde la última vez'. La hoja FEB2023 prevalece.

_CRUCE_FOOTER = re.compile(r"^\s*(total|saldo|diferencia|sistema|banco)\b", re.I)


def _cruce_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).strip().replace("$", "").replace(" ", "")
    if not s:
        return None
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def _cruce_doc(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def parse_pendientes_cruce(raw: bytes, sheet_pref: str = "FEB2023") -> tuple[str | None, list[dict]]:
    """Parsea una hoja de DEPÓSITOS PENDIENTES para CRUZAR contra el backlog.

    A diferencia de parse_xlsx (que recorre TODAS las hojas y descarta
    negativos), esto:
      - Prioriza la hoja `sheet_pref` (FEB2023) — la dueña dijo que prevalece.
      - INCLUYE negativos (PAGO SENAE, etc. son pendientes válidos).
      - Exige `codigo` numérico (descarta ajustes tipo 'AC97' y footer TOTAL/SALDO).

    Returns: (nombre_hoja_usada, [{doc, monto, detalle}, ...]).
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    names = wb.sheetnames
    order = []
    for n in names:
        if n.strip().upper() == sheet_pref.strip().upper():
            order.insert(0, n)
        else:
            order.append(n)

    for nombre in order:
        ws = wb[nombre]
        hdr = None
        cmap: dict = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i > 12:
                break
            cells = [str(c or "").strip().lower() for c in row]
            ci: dict = {}
            for j, c in enumerate(cells):
                if c in ("fecha", "fec ha") and "fecha" not in ci:
                    ci["fecha"] = j
                elif c == "codigo" and "codigo" not in ci:
                    ci["codigo"] = j
                elif c == "valor" and "valor" not in ci:
                    ci["valor"] = j
                elif c in ("detalle", "concepto") and "detalle" not in ci:
                    ci["detalle"] = j
            if "valor" in ci and ("codigo" in ci or "fecha" in ci):
                hdr = i
                cmap = ci
                break
        if hdr is None or "codigo" not in cmap:
            continue
        items = []
        for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
            doc = _cruce_doc(row[cmap["codigo"]]) if len(row) > cmap["codigo"] else ""
            det = (str(row[cmap["detalle"]] or "").strip()
                   if "detalle" in cmap and len(row) > cmap["detalle"] else "")
            val = _cruce_num(row[cmap["valor"]]) if len(row) > cmap["valor"] else None
            fch = (_parse_fecha_celda(row[cmap["fecha"]])
                   if "fecha" in cmap and len(row) > cmap["fecha"] else None)
            if _CRUCE_FOOTER.match(det):
                continue
            if val is None:
                continue
            if not doc or not doc.isdigit():
                continue
            items.append({"doc": doc, "monto": val, "detalle": det[:60], "fecha": fch})
        return nombre, items
    return None, []
