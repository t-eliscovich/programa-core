"""Parser del xlsx del extracto bancario (Banco Pichincha).

TMT 2026-05-22 — Diferente al parser de "depósitos pendientes":
este xlsx viene directamente del banco con TODOS los movimientos del día
(créditos y débitos), no solo depósitos sistema. Formato fijo:

    Fila 1: Fecha | Concepto | Documento | Monto | Saldo | Codigo | Tipo | Oficina | (extra)
    Fila 2+: una fila por movimiento

Campos:
    Fecha       — DD/MM/YYYY
    Concepto    — descripción libre del banco
    Documento   — número de comprobante / referencia (int)
    Monto       — siempre positivo (el signo lo da Tipo C/D)
    Saldo       — saldo del banco DESPUÉS de aplicar el movimiento
    Codigo      — código de la oficina (001045, 001010, etc.)
    Tipo        — 'C' = crédito (entra plata), 'D' = débito (sale plata)
    Oficina     — nombre legible de la oficina (AG. NORTE, AMERICA, ...)
"""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

_LOG = logging.getLogger("programa_core.conciliacion.parser_banco")


@dataclass(frozen=True)
class MovBanco:
    """Un movimiento del extracto bancario (lado REAL)."""

    fecha: date
    concepto: str
    documento: str       # número de comprobante banco
    monto: Decimal       # siempre positivo
    saldo: Decimal       # saldo banco después del mov
    codigo: str          # código oficina
    tipo: str            # 'C' o 'D'
    oficina: str


def _parse_fecha(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s[:19] if len(s) > 10 else s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(v) -> Decimal:
    if v is None:
        return Decimal(0)
    if isinstance(v, int | float | Decimal):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return Decimal(0)
    s = str(v).strip().replace(" ", "")
    if not s:
        return Decimal(0)
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal(0)


def parse_banco_xlsx(raw: bytes) -> list[MovBanco]:
    """Parsea bytes de un .xlsx del extracto Pichincha → lista de movimientos.

    Reglas:
      - Primera fila = headers (Fecha, Concepto, Documento, Monto, Saldo, Codigo, Tipo, Oficina).
      - Filas con monto = 0 o fecha inválida se ignoran.
      - Filas con Tipo distinto de C/D se ignoran (con warning).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl no instalado — pip install openpyxl") from e

    if not raw:
        return []

    bio = io.BytesIO(raw)
    wb = load_workbook(bio, data_only=True, read_only=True)
    salida: list[MovBanco] = []

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        headers: list[str] = []
        col_idx: dict[str, int] = {}

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_list = list(row)
            if i == 1:
                headers = [str(c or "").strip().lower() for c in row_list]
                for j, h in enumerate(headers):
                    if h in ("fecha", "fec ha", "fecha mov"):
                        col_idx["fecha"] = j
                    elif h == "concepto":
                        col_idx["concepto"] = j
                    elif h == "documento":
                        col_idx["documento"] = j
                    elif h in ("monto", "valor", "importe"):
                        col_idx["monto"] = j
                    elif h == "saldo":
                        col_idx["saldo"] = j
                    elif h in ("codigo", "código"):
                        col_idx["codigo"] = j
                    elif h == "tipo":
                        col_idx["tipo"] = j
                    elif h == "oficina":
                        col_idx["oficina"] = j
                if "fecha" not in col_idx or "monto" not in col_idx or "tipo" not in col_idx:
                    _LOG.warning("Hoja %s sin columnas mínimas (fecha/monto/tipo). Skip.", nombre_hoja)
                    break
                continue

            # Fila de data
            if not any(c is not None and str(c).strip() for c in row_list):
                continue
            try:
                fecha = _parse_fecha(row_list[col_idx["fecha"]])
                monto = _parse_decimal(row_list[col_idx["monto"]])
                # TMT 2026-06-03 audit blindaje: aceptar "CR","DB","CRED","DEB"
                # → primer char C/D. Sin esto, filas con tipo distinto a 1 char
                # exacto se silenciaban y se perdía data del extracto.
                tipo_raw = str(row_list[col_idx["tipo"]] or "").strip().upper()
                tipo = tipo_raw[:1] if tipo_raw else ""
            except IndexError:
                continue
            if fecha is None or monto == 0:
                continue
            if tipo not in ("C", "D"):
                _LOG.debug("Hoja %s fila %d: tipo %r desconocido", nombre_hoja, i, tipo_raw)
                continue

            def _g(key: str, default: str = "") -> str:
                j = col_idx.get(key)
                if j is None:
                    return default
                try:
                    v = row_list[j]
                except IndexError:
                    return default
                if v is None:
                    return default
                s = str(v).strip()
                # openpyxl puede stringificar ints como '38078012.0'
                if s.endswith(".0"):
                    s = s[:-2]
                return s

            saldo = _parse_decimal(row_list[col_idx["saldo"]]) if "saldo" in col_idx else Decimal(0)

            salida.append(
                MovBanco(
                    fecha=fecha,
                    concepto=_g("concepto"),
                    documento=_g("documento"),
                    monto=monto,
                    saldo=saldo,
                    codigo=_g("codigo"),
                    tipo=tipo,
                    oficina=_g("oficina"),
                )
            )

    wb.close()
    return salida
