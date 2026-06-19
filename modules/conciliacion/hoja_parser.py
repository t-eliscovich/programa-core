"""Parser de la "hoja de conciliación" (Excel) → filas de pendientes de banco.

TMT 2026-06-04 dueña: "lo único que se tiene que mantener como pendientes es
lo del archivo (la hoja). si subimos hoja se toma de la hoja."

La hoja es el workbook CONCILIACION_*.xlsx con una pestaña por período (ej.
"FEB2023"). Layout observado:

    fila 1: "DEPÓSITOS PENDIENTES"
    fila 2: FECHA | DETALLE | CODIGO | VALOR | DETALLE
    fila 3..N: datos
        - depósitos pendientes → VALOR positivo (tipo C)
        - cheques/pagos pendientes (SENAE, etc.) → VALOR negativo (tipo D)
        - ajustes tipo "AC97" → sin fecha, VALOR positivo (tipo C)
    final: filas TOTAL / SALDO SISTEMA / SALDO BANCO / DIFERENCIA (se ignoran)

El neto de la hoja (Σ C − Σ D) + SALDO SISTEMA debe dar SALDO BANCO con
diferencia 0 — eso es lo que cuadra la conciliación.

Este módulo SOLO parsea. La carga a banco_historicos_pendientes (DELETE +
INSERT idempotente) vive en `reemplazar_historicos_desde_hoja`.
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime

_LOG = logging.getLogger("programa_core.conciliacion.hoja_parser")

# Filas de cierre que NO son pendientes.
_FILAS_CIERRE = {
    "TOTAL", "SALDO SISTEMA", "SALDO BANCO", "SALDO", "DIFERENCIA",
    "SALDO ANTERIOR", "SALDO FINAL",
}


def _parse_fecha(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_float(v) -> float | None:
    """Parsea un VALOR a float, tolerando formato US y EU.

    BUG histórico (2026-06-19): la versión vieja hacía
    ``str(v).replace(",", "")`` asumiendo SIEMPRE formato US (coma=miles,
    punto=decimal). Con un importe EU escrito como TEXTO — p.ej. el depósito
    de 967.608,22 — eso daba "967.608.22", que ``float()`` rechaza → la fila
    se descartaba en silencio y "no cargaba como las demás". Importes EU
    chicos como "557,58" se convertían en 55758 (inflados x100). Ahora
    detectamos el separador decimal por la posición del último separador y,
    para el caso de una sola coma, por la cantidad de decimales (Ecuador usa
    siempre 2). Si igual no parsea, devolvemos None y el llamador lo SURFACEA
    (no lo tira en silencio).
    """
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "").replace(" ", "")
    if not s:
        return None
    # paréntesis o signo = negativo (formato contable)
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    if s.startswith("-"):
        neg = True
        s = s[1:]
    has_dot, has_comma = "." in s, "," in s
    if has_dot and has_comma:
        # El ÚLTIMO separador es el decimal.
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")   # EU: 967.608,22
        else:
            s = s.replace(",", "")                       # US: 967,608.22
    elif has_comma:
        # Sólo coma: decimal si quedan exactamente 2 dígitos detrás
        # (Ecuador siempre usa 2); si no, es separador de miles.
        partes = s.split(",")
        if len(partes) == 2 and len(partes[1]) == 2:
            s = s.replace(",", ".")                      # 557,58 → 557.58
        else:
            s = s.replace(",", "")                       # 1,068 → 1068
    # sólo punto o sin separador: float() lo maneja tal cual.
    try:
        val = float(s)
    except ValueError:
        return None
    return -val if neg else val


def _elegir_hoja(wb, sheet: str | None):
    if sheet:
        if sheet in wb.sheetnames:
            return wb[sheet]
        # tolerancia case-insensitive / sin espacios
        norm = sheet.strip().lower().replace(" ", "")
        for sn in wb.sheetnames:
            if sn.strip().lower().replace(" ", "") == norm:
                return wb[sn]
        raise ValueError(
            f"La hoja '{sheet}' no existe. Pestañas: {', '.join(wb.sheetnames)}"
        )
    # Auto-detect: primera pestaña con headers FECHA + VALOR.
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            cells = [str(c).strip().upper() for c in row if c is not None]
            if "FECHA" in cells and "VALOR" in cells:
                return ws
    return wb[wb.sheetnames[0]]


def _localizar_columnas(ws) -> tuple[int, dict]:
    """Devuelve (fila_header_1based, {campo: col_idx_0based}).

    Busca la fila con FECHA y VALOR; mapea por nombre. Si no encuentra,
    asume layout A/B/C/D (fecha, concepto, documento, valor).
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        cells = [(j, str(c).strip().upper()) for j, c in enumerate(row) if c is not None]
        upset = {c for _, c in cells}
        if "FECHA" in upset and "VALOR" in upset:
            cols = {}
            detalle_cols = []
            for j, c in cells:
                if c == "FECHA":
                    cols["fecha"] = j
                elif c == "VALOR":
                    cols["valor"] = j
                elif c in ("CODIGO", "CÓDIGO", "DOC", "DOCUMENTO"):
                    cols["documento"] = j
                elif c == "DETALLE":
                    detalle_cols.append(j)
            if detalle_cols:
                cols["concepto"] = detalle_cols[0]
                if len(detalle_cols) > 1:
                    cols["detalle2"] = detalle_cols[1]
            cols.setdefault("fecha", 0)
            cols.setdefault("concepto", 1)
            cols.setdefault("documento", 2)
            cols.setdefault("valor", 3)
            return i, cols
    # Fallback layout fijo.
    return 2, {"fecha": 0, "concepto": 1, "documento": 2, "valor": 3}


def parse_hoja_pendientes(
    source, sheet: str | None = None, *, return_dropped: bool = False,
):
    """Parsea la hoja y devuelve la lista de pendientes de banco.

    Args:
        source: path (str) o bytes del .xlsx.
        sheet:  nombre de la pestaña (ej. "FEB2023"). Si None, auto-detecta.
        return_dropped: si True, devuelve ``(rows, dropped)`` donde `dropped`
            son las filas con concepto+VALOR que NO se pudieron parsear (antes
            se descartaban en silencio → un mov "no cargaba como las demás" y
            nadie se enteraba). Default False = comportamiento histórico.

    Returns:
        list[{fecha, concepto, documento, monto (>0), tipo: 'C'|'D', fila}]
        o (rows, dropped) si return_dropped.
    """
    import openpyxl

    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = openpyxl.load_workbook(source, data_only=True)

    ws = _elegir_hoja(wb, sheet)
    header_row, cols = _localizar_columnas(ws)
    c_fecha = cols["fecha"]
    c_concepto = cols["concepto"]
    c_doc = cols["documento"]
    c_valor = cols["valor"]

    out: list[dict] = []
    dropped: list[dict] = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        def _cell(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        concepto_raw = _cell(c_concepto)
        concepto = str(concepto_raw).strip() if concepto_raw is not None else ""
        # Fila de cierre (TOTAL / SALDO / DIFERENCIA) → no es pendiente.
        if concepto.upper() in _FILAS_CIERRE:
            continue

        valor_raw = _cell(c_valor)
        valor = _to_float(valor_raw)
        if valor is None or valor == 0:
            # SOLO es "drop sospechoso" si la celda VALOR tenía contenido que
            # no pudimos parsear (texto raro, formato roto) Y hay concepto.
            # Las filas en blanco (sin concepto ni valor) se ignoran como antes.
            tiene_valor = valor_raw is not None and str(valor_raw).strip() != ""
            if concepto and tiene_valor and valor is None:
                _LOG.warning(
                    "hoja fila %d: no pude parsear VALOR %r (concepto=%r)",
                    i, valor_raw, concepto[:60],
                )
                dropped.append({
                    "fila": i, "concepto": concepto[:120],
                    "valor_raw": str(valor_raw)[:40],
                    "documento": str(_cell(c_doc) or "").strip()[:40],
                })
            continue

        doc_raw = _cell(c_doc)
        documento = str(doc_raw).strip() if doc_raw is not None else ""
        # openpyxl a veces trae el código como float (ej. 56379469.0)
        if documento.endswith(".0"):
            documento = documento[:-2]

        tipo = "D" if valor < 0 else "C"
        monto = round(abs(valor), 2)
        fecha = _parse_fecha(_cell(c_fecha))

        out.append({
            "fecha": fecha,
            "concepto": concepto[:120],
            "documento": documento[:40],
            "monto": monto,
            "tipo": tipo,
            "fila": i,
        })
    if return_dropped:
        return out, dropped
    return out

    return out


def resumen(rows: list[dict]) -> dict:
    """Totales de control para mostrar antes de confirmar la carga."""
    cred = sum(r["monto"] for r in rows if r["tipo"] == "C")
    deb = sum(r["monto"] for r in rows if r["tipo"] == "D")
    return {
        "n": len(rows),
        "n_cred": sum(1 for r in rows if r["tipo"] == "C"),
        "n_deb": sum(1 for r in rows if r["tipo"] == "D"),
        "creditos": round(cred, 2),
        "debitos": round(deb, 2),
        "neto": round(cred - deb, 2),
    }


def reemplazar_historicos_desde_hoja(
    no_banco: int, rows: list[dict], *, fuente: str = "hoja", usuario: str = "web",
    conn=None,
) -> dict:
    """DELETE todos los históricos del banco + INSERT los de la hoja.

    Idempotente (como restaurar_173): correrlo dos veces deja el mismo
    estado. NO toca los ya conciliados de otras tablas; solo reescribe
    banco_historicos_pendientes para `no_banco`.

    TMT 2026-06-04 dueña: "si subimos hoja se toma de la hoja". La hoja es
    la verdad de los pendientes de banco.
    """
    import db as _db

    def _run(c):
        n_del = _db.execute(
            "DELETE FROM scintela.banco_historicos_pendientes WHERE no_banco = %s",
            (no_banco,), conn=c,
        ) or 0
        n_ins = 0
        for r in rows:
            _db.execute(
                """
                INSERT INTO scintela.banco_historicos_pendientes
                    (no_banco, fecha, concepto, documento, monto, tipo,
                     fuente, creado_por)
                VALUES (%s, %s, %s, %s, %s::numeric, %s, %s, %s)
                """,
                (
                    no_banco, r.get("fecha"), (r.get("concepto") or "")[:120],
                    (r.get("documento") or "")[:40], str(r.get("monto") or 0),
                    (r.get("tipo") or "C")[:1],
                    f"{fuente}:fila{r.get('fila','')}"[:60], usuario[:50],
                ),
                conn=c,
            )
            n_ins += 1
        return n_del, n_ins

    if conn is not None:
        n_del, n_ins = _run(conn)
    else:
        with _db.tx() as c:
            n_del, n_ins = _run(c)

    res = resumen(rows)
    res.update({"borradas": n_del, "insertadas": n_ins})
    return res
