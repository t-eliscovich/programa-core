"""/inventario-rotativo — lo que se vende siempre y hay que tener.

Dos cortes del mismo dato: por COLOR (default, dueña 2026-08-18) y por tela.
El filtrado —estado, familia, búsqueda— se hace en el navegador: la consulta
a Asinfo tarda unos segundos y recargar la página entera para tildar un filtro
haría la pantalla incómoda justo en el uso normal.

Rutas:
  GET /inventario-rotativo         (stock.ver)
  GET /inventario-rotativo/excel   (stock.ver)

La hoja impresa sale de la MISMA pantalla (`?imprimir=1`): si fuera otra
plantilla, las dos se irían separando y lo que se lleva a planta dejaría de
ser lo que se mira. Mismo criterio que la hoja del portal de vendedores.

Y respeta el filtro que haya en pantalla (dueña 2026-08-18): el botón le
cuelga `est`/`fam`/`q` a la URL y ACÁ se filtran las filas, no en el
navegador. Sin eso, pedir "las 62 que hay que teñir" imprimía las 289 —
trece páginas para tirar. El filtro se aplica en el servidor y no por JS
porque la hoja tiene que salir igual aunque el JS no haya corrido todavía
cuando el navegador arma el papel.

Va en el menú "Producción y stocks", arriba de Inventario.
"""
from __future__ import annotations

import io

from flask import Blueprint, Response, flash, redirect, render_template, request, url_for

from auth import requiere_login, requiere_permiso
from filters import today_ec

from . import service

inventario_rotativo_bp = Blueprint(
    "inventario_rotativo", __name__, template_folder="templates"
)

#: El corte por COLOR primero: es como la dueña mira el inventario.
CORTES = ("color", "tela")
CORTES_LBL = {"color": "Color", "tela": "Tela"}


def _filtros() -> dict:
    """Los filtros de la URL, normalizados. Lo que no se entiende no filtra."""
    est = (request.args.get("est") or "").strip().lower()
    return {
        "est": est if est in ("rojo", "ambar", "ok", "sobra") else "",
        "fam": (request.args.get("fam") or "").strip(),
        "q": (request.args.get("q") or "").strip().lower(),
    }


def _aplicar(filas: list[dict], f: dict) -> list[dict]:
    """Las mismas tres condiciones que el JS de la pantalla.

    Si esto y el JS se separan, la hoja deja de ser lo que se ve — por eso son
    tres líneas y no una máquina de filtros.
    """
    out = filas
    if f["est"]:
        out = [x for x in out if x["estado"] == f["est"]]
    if f["fam"]:
        out = [x for x in out if x["familia"] == f["fam"]]
    if f["q"]:
        out = [x for x in out
               if f["q"] in f"{x['color']} {x['tela']}".lower()]
    return out


@inventario_rotativo_bp.route("/inventario-rotativo")
@requiere_login
@requiere_permiso("stock.ver")
def lista():
    filas, disponible = service.rotativo()

    corte = (request.args.get("ver") or "").strip()
    if corte not in CORTES:
        corte = CORTES[0]

    imprimir = request.args.get("imprimir") == "1"
    filtros = _filtros()
    # En pantalla filtra el JS (instantáneo, sin recargar). En el papel filtra
    # el servidor: no hay dónde clickear después.
    if imprimir:
        filas = _aplicar(filas, filtros)

    return render_template(
        "inventario_rotativo/lista.html",
        semanas_corta=service.SEMANAS_CORTA,
        disponible=disponible,
        corte=corte,
        imprimir=imprimir,
        filtros=filtros,
        hoy=today_ec(),
        cortes=CORTES,
        cortes_lbl=CORTES_LBL,
        bloques=service.agrupar(filas, corte) if filas else [],
        resumen=service.resumen(filas),
        sem_rojo=service.SEM_ROJO,
        lead_semanas=service.LEAD_SEMANAS,
        semanas_minimas=service.SEMANAS_MINIMAS,
    )


#: Las columnas del Excel, en el mismo orden que la pantalla.
#: Una sola columna de faltante, en la unidad de la fila (dueña 2026-08-18:
#: "falta tiene que ser en rollos o en kg pero no dos columnas, cuando hay
#: unidades ya está"). La de kilos era para poder sumar la planilla entera,
#: pero un total que mezcla telas con cuellos no se usa para nada.
COLUMNAS = ("Color", "Tela", "Familia", "Acabado", "Un.", "Por semana",
            "En bodega", "Pedido (informativo)", "En proceso", "Alcanza (sem)",
            "Falta")

#: Ancho de cada columna, en el orden de COLUMNAS.
_ANCHOS = (10, 26, 13, 8, 6, 11, 11, 18, 11, 13, 9)


def _valores(f: dict) -> list:
    """Una fila del service → la fila del Excel, en la unidad que se lee."""
    return [
        f["color"], f["tela"], f["familia"], f.get("acabado") or "", f["unidad"],
        f["sem"], f["stock"], f["pedido"], f["proceso"],
        # "no sé" (sin venta reciente) va vacío y no como -1: un número
        # negativo en una columna de semanas se lee como un dato, no como
        # un hueco.
        None if f["alcanza"] < 0 else f["alcanza"],
        f["falta"],
    ]


@inventario_rotativo_bp.route("/inventario-rotativo/excel")
@requiere_login
@requiere_permiso("stock.ver")
def excel():
    """La misma tabla, en un xlsx.

    Sale EN EL MISMO ORDEN QUE LA PANTALLA (dueña 2026-08-18: "cuando bajo el
    excel aparece en otro orden") y con el mismo filtro: se agrupa por el corte
    que se esté mirando, los bloques van por faltante y dentro de cada uno las
    filas por urgencia. Una planilla ordenada distinto obliga a buscar de nuevo
    lo que ya se tenía en pantalla.

    Plana y con una sola hoja: se filtra y se ordena en Excel, que es lo que se
    hace con ella. La columna de grupo va primero para que agrupar sea ordenar
    por la columna A.
    """
    filas, disponible = service.rotativo()
    if not disponible:
        flash("No pude consultar Asinfo, así que no armé el Excel. "
              "Probá de nuevo en un minuto.", "error")
        return redirect(url_for("inventario_rotativo.lista"))

    corte = (request.args.get("ver") or "").strip()
    if corte not in CORTES:
        corte = CORTES[0]
    filtros = _filtros()
    filas = _aplicar(filas, filtros)
    # El mismo agrupado de la pantalla, aplanado: bloque por bloque, en orden.
    ordenadas = [f for b in service.agrupar(filas, corte) for f in b["filas"]]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("openpyxl no instalado en el server — pedí al admin que lo instale.",
              "error")
        return redirect(url_for("inventario_rotativo.lista"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Rotación del inventario"

    encabezado = Font(bold=True, color="FFFFFF", size=10)
    relleno = PatternFill("solid", fgColor="0F172A")
    rojo = Font(bold=True, color="B91C1C")
    ambar = Font(bold=True, color="B45309")
    gris = Font(color="94A3B8")
    linea = Border(bottom=Side(style="thin", color="E2E8F0"))

    for i, h in enumerate(COLUMNAS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = encabezado
        c.fill = relleno
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26

    # Las tres primeras columnas son texto; de la 6 en adelante, números.
    for fila, f in enumerate(ordenadas, 2):
        pinta = {"rojo": rojo, "ambar": ambar}.get(f["estado"], gris)
        for i, v in enumerate(_valores(f), 1):
            c = ws.cell(row=fila, column=i, value=v)
            c.border = linea
            if i >= 6:
                c.number_format = "#,##0" if i != 10 else "#,##0.0"
                c.alignment = Alignment(horizontal="right")
            # El semáforo de la pantalla, en las dos columnas que lo llevan.
            if i in (10, 11):
                c.font = pinta

    for i, ancho in enumerate(_ANCHOS, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{max(len(ordenadas) + 1, 2)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"inventario_rotativo_{today_ec():%Y_%m_%d}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )
