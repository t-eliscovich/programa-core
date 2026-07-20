"""Informes gerenciales — read-only v1."""

import csv
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from flask import (
    Blueprint,
    Response,
    abort,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    session,
    url_for,
)

import db
from auth import requiere_login, requiere_permiso, tiene_permiso
from error_messages import flash_exc
from exports import csv_response
from filters import today_ec

from . import queries

informes_bp = Blueprint(
    "informes",
    __name__,
    template_folder="templates",
)


def _safe(fn, default):
    """Run a query; on error return (default, error_message)."""
    try:
        return fn(), None
    except Exception as e:
        return default, str(e)


# ── Cache de la pantalla /flujo-produccion ──────────────────────────────────
# Federico 2026-07-17: armar esta pantalla dispara ~12 consultas externas
# (dBase + Asinfo/Metabase + formulas). Las de Asinfo ya están cacheadas en su
# service (5–10 min), pero las de dBase/formulas (movimientos del mes, tintorería
# mensual del AÑO entero, gastos, químicos) se re-ejecutaban en CADA carga → entrar
# y salir de la pantalla tardaba lo mismo siempre. Cacheamos el contexto ya armado
# por (anio, mes) unos minutos: la primera carga sigue costando, pero las repetidas
# (mismo mes, cualquier usuario — los datos NO dependen del usuario) salen
# instantáneas. TTL corto (2 min) para no servir números viejos; solo se cachean
# cargas exitosas (con data), nunca un Asinfo caído o un mes vacío.
import logging as _logging_fp
import time as _time_fp

_LOG_FP = _logging_fp.getLogger(__name__)
_FLUJO_PROD_CACHE: dict = {}
_FLUJO_PROD_TTL_SECS = 300  # 5 min (antes 10 — duena 2026-07-18); con el warmup
# de Asinfo corriendo (modules/_lib/warmup.py) la recompute es barata y la
# pantalla queda como maximo 5 min desactualizada.


def reset_flujo_produccion_cache() -> None:
    """Vaciar el cache de la pantalla /flujo-produccion (tests / tras deploy)."""
    _FLUJO_PROD_CACHE.clear()


def _build_mov_asinfo(data, inv_inic, inv_act, anio=None, mes=None,
                      proy_quimico=None) -> dict | None:
    """Tabla 'movimientos del mes' con la MISMA lógica del dBase/TINT.BAT pero
    con los datos viniendo de Asinfo (no de scintela.iniciales/historia).

    El dBase arma una cadena de balance de masa (hilado→crudo→terminado) donde
    el egreso de una etapa = el ingreso de la siguiente. Acá los SALDOS (inicial
    as-of el 1° del mes + actual live) salen de Asinfo; el único flujo que Asinfo
    no da por mes es "lo tejido" (W=ktej), que se toma del dBase como ANCLA, y el
    resto (lo tinturado, compras de hilo, ventas) se DERIVA para cerrar el balance
    contra los saldos de Asinfo. Ver el bloque de cálculo abajo.

    $/kg: se mantienen los del dBase (Asinfo no tiene dólares confiables), aplicados
    a las nuevas bases de kg. COLORANTES no tiene stock en Asinfo → valor PC.

    Mapeo etapa → categoría: Hilado←hilo_total, Crudo/Tejido←cruda_total,
    Terminado←terminada (misma combinación que la tabla 'Inventario Asinfo live').

    Devuelve None si no hay `header` o si Asinfo (inicial O actual) no está
    disponible → la vista muestra un aviso apagado.
    """
    if not isinstance(data, dict):
        return None
    header = data.get("header") or {}
    if not header:
        return None
    if not (isinstance(inv_inic, dict) and inv_inic.get("disponible")):
        return None
    if not (isinstance(inv_act, dict) and inv_act.get("disponible")):
        return None

    def _f(d, k):
        try:
            return float(d.get(k) or 0)
        except (TypeError, ValueError):
            return 0.0

    hl = dict(header.get("hilado", {}))
    tj = dict(header.get("tejido", {}))
    te = dict(header.get("terminado", {}))
    co = dict(header.get("colorantes", {}))

    # ── Saldos de Asinfo: inicial (as-of 1° del mes) y ACTUAL (live) ──────
    # TMT 2026-07-09 (dueña "solo bodegas 51/52/53, sin en-proceso"): usamos los
    # saldos de bodega puros (hilo=51, tela_cruda=52, terminada=53), NO los
    # *_total que suman el WIP (en_proceso). El inicial as-of NO puede
    # reconstruir el WIP → si el actual lo incluyera, la cadena mostraría un
    # salto fantasma. Con bodegas puras, inicial y actual quedan consistentes.
    hi0 = _f(inv_inic, "hilo")
    tc0 = _f(inv_inic, "tela_cruda")
    pf0 = _f(inv_inic, "terminada")
    hi1 = _f(inv_act, "hilo")
    tc1 = _f(inv_act, "tela_cruda")
    pf1 = _f(inv_act, "terminada")

    # ── EN MÁQUINAS (WIP live) — fila "En máquinas" (dueña 2026-07-09): lo que
    # está en proceso HOY, para que bodega (Stock act.) + máquinas coincida con
    # el "INVENTARIO ASINFO (LIVE)" de arriba (hilo_total / cruda_total).
    #   · Hilado: en_proceso_tc = hilo en telares (hilo → tela cruda).
    #   · Crudo:  en_proceso_pt = cruda en tintorería (tela cruda → terminada).
    #   · Terminado: no tiene WIP después. El inicial as-of no reconstruye WIP.
    maq_hilado = _f(inv_act, "en_proceso_tc")
    maq_crudo = _f(inv_act, "en_proceso_pt")

    # ── Cadena de balance de masa (misma lógica que el dBase/TINT.BAT, pero
    # con datos de Asinfo). El dBase hace:
    #   hilado_act = hilado_ini + compras − ktej   (ktej = lo TEJIDO del mes)
    #   crudo_act  = crudo_ini  + ktej   − ktin     (ktin = lo TINTURADO)
    #   term_act   = term_ini   + ktin   − kvent    (kvent = lo VENDIDO)
    # y el EGRESO de cada etapa = el INGRESO de la siguiente
    # (hilado egreso = ktej = crudo ingreso; crudo egreso = ktin = term ingreso).
    #
    # Acá los SALDOS (inicial y actual) vienen de Asinfo. El único flujo que no
    # da Asinfo por mes es "lo tejido" (W) → lo tomamos del dBase como ANCLA, y
    # derivamos el resto para que TODO cierre contra los saldos de Asinfo:
    #   W (lo tejido)      = ktej del dBase              [ancla]
    #   D (lo tinturado)   = crudo_ini + W − crudo_act   [balance crudo]
    #   compras (hilo)     = hilado_act − hilado_ini + W [balance hilado]
    #   ventas (terminado) = term_ini + D − term_act     [balance terminado]
    # TMT 2026-07-08 (dueña: "el egreso de hilado tiene que ser lo que se fue a
    # tejer" + "los datos vienen de Asinfo, no se cargan").
    # ── TODO de Asinfo — nada del dBase. Asinfo NO tiene foto del "en proceso"
    # (WIP) al 1° del mes, así que los flujos del medio se DERIVAN de los saldos
    # de Asinfo (la "cuenta para que dé"), en vez de tomarse del dBase. Dueña
    # 2026-07-09: "todo de asinfo o de formulas; si en proceso no hay 1 del mes,
    # hay que hacer una cuenta para que dé". El dBase queda SOLO para comparar
    # (la tabla de abajo). Antes W ("lo tejido") venía del dBase → por eso
    # coincidían los 84.067 en las dos tablas; ese era el dato que no cambiaba.
    #
    # HILADO ingreso = importaciones RECIBIDAS en Asinfo (por fecha de recepción).
    compras = 0.0
    compras_us = 0.0
    if anio and mes:
        try:
            from modules.asinfo import service as _asvc
            compras = float(_asvc.hilado_recibido_mes(int(anio), int(mes)) or 0.0)
        except Exception:  # noqa: BLE001 -- fail-soft
            compras = 0.0
        # $ del hilado que entró = costo-a-la-fecha de NUESTRA base (anticipos +
        # compras de la importación, atribuido por año). El dBase no da $/kg de
        # ingreso (por eso la columna estaba en "—"). Dueña 2026-07-10.
        try:
            from modules.importaciones import service as _impsvc
            compras_us = float(
                (_impsvc.costo_hilado_recibido_mes(int(anio), int(mes)) or {}).get("us")
                or 0.0
            )
        except Exception:  # noqa: BLE001 -- fail-soft
            compras_us = 0.0
    # FLUJO DE FABRICACIÓN REAL de Asinfo (dueña 2026-07-09). Por bodega, de las
    # órdenes CERRADAS en el mes (fecha_cierre): material consumido (issued) y
    # producto fabricado (fab). El desperdicio (merma) = issued − fab.
    #   · Tejeduría (b52): hilo consumido − tela cruda producida.
    #   · Tintura   (b53): crudo consumido − producto terminado producido.
    b52 = {"issued": 0.0, "fab": 0.0}
    b53 = {"issued": 0.0, "fab": 0.0}
    if anio and mes:
        try:
            from modules.asinfo import service as _asvc2
            b52 = _asvc2.fabricacion_flujo_mes(52, int(anio), int(mes)) or b52
            b53 = _asvc2.fabricacion_flujo_mes(53, int(anio), int(mes)) or b53
        except Exception:  # noqa: BLE001 -- fail-soft
            pass
    hilo_consumido = float(b52.get("issued") or 0.0)   # HILADO egreso (a tejer)
    ci_crudo = float(b52.get("fab") or 0.0)            # CRUDO ingreso (producido)
    crudo_consumido = float(b53.get("issued") or 0.0)  # CRUDO egreso (a tinturar)
    ci_term = float(b53.get("fab") or 0.0)             # TERM ingreso (producido)
    ventas = max(pf0 + ci_term - pf1, 0.0)             # venta DERIVADA (comparación)
    # TERM ingreso/egreso = MOVIMIENTO REAL de la bodega 53 (deltas del saldo por
    # lote DESDE el 1° del mes exclusivo), la MISMA fuente que el stock. Cierra
    # por construcción: pf0 + ingreso − egreso = pf1 (identidad telescópica).
    # Dueña 2026-07-10 "no podemos tener residuo de 12k": el residuo venía de
    # mezclar fuentes (inicial de foto vieja + ingreso=producción que subcuenta
    # + egreso=despacho que corre ~7% arriba del saldo por timing). Con todo del
    # mismo saldo cierra exacto. El facturado/despacho (venta) se muestran al
    # lado como referencia; corren un poco arriba del egreso de bodega (desfase).
    # MOVIMIENTO REAL de bodega por etapa (deltas del saldo por lote DESDE el 1°
    # del mes, MISMA fuente/corte que el stock) → hilado (51), crudo (52),
    # terminado (53) cierran TODOS por telescopía: inicial + ingreso − egreso =
    # saldo final. Dueña 2026-07-10: el hilado/crudo mostraban "lo que se fue a
    # tejer/producción" (subcuenta el movimiento real de bodega → no cerraba);
    # verificado en Asinfo que el egreso real de bodega 51 fue 88.178 vs 46.540
    # de "a tejer". "Importaciones" y "a tejer/producción" quedan de referencia.
    mov51 = {"ingreso": 0.0, "egreso": 0.0}
    mov52 = {"ingreso": 0.0, "egreso": 0.0}
    mov53 = {"ingreso": 0.0, "egreso": 0.0}
    despacho = 0.0
    if anio and mes:
        try:
            from datetime import date as _date_mov

            from modules.asinfo import service as _asvcd
            _corte = _date_mov(int(anio), int(mes), 1)  # mismo corte que el inicial
            mov51 = _asvcd.movimiento_bodega_mes(51, _corte) or mov51
            mov52 = _asvcd.movimiento_bodega_mes(52, _corte) or mov52
            mov53 = _asvcd.movimiento_bodega_mes(53, _corte) or mov53
            despacho = float(_asvcd.despacho_fisico_mes(int(anio), int(mes)) or 0.0)
        except Exception:  # noqa: BLE001 -- fail-soft
            pass
    ing_hilo = float(mov51.get("ingreso") or 0.0)   # HILADO ingreso (real bodega 51)
    egr_hilo = float(mov51.get("egreso") or 0.0)    # HILADO egreso (real bodega 51)
    ing_crudo = float(mov52.get("ingreso") or 0.0)  # CRUDO ingreso (real bodega 52)
    egr_crudo = float(mov52.get("egreso") or 0.0)   # CRUDO egreso (real bodega 52)
    ing_term = float(mov53.get("ingreso") or 0.0)   # TERM ingreso (real bodega)
    egr_term = float(mov53.get("egreso") or 0.0)    # TERM egreso (real bodega)
    # DESPERDICIO = egreso del proceso anterior − ingreso al siguiente (= merma).
    desp_crudo = hilo_consumido - ci_crudo   # tejeduría: hilo − cruda
    desp_term = crudo_consumido - ci_term    # tintura: crudo − PT

    # HILADO — ingreso=compras (importaciones), egreso=hilo consumido a tejer.
    # COSTEO POR PROMEDIO PONDERADO (dueña 2026-07-10), TODO de Programa Core:
    #   · Apertura: $/kg = promedio de las compras/importaciones de PC (Σ costo ÷
    #     Σ kg), NO el $/kg del dBase. Fail-soft al $/kg histórico si Asinfo cae.
    #   · Ingreso: al costo REAL de la importación (anticipos+compras de PC).
    #   · Egreso y stock: al PROMEDIO = (apertura $ + ingreso $) / (kg apert + kg ing).
    #   · En máquinas (WIP): valuado al $/kg de apertura.
    #   · Stock act $ = apertura $ + ingreso $ − egreso $  → la columna CIERRA, y el
    #     ingreso barato BAJA el $/kg (antes el egreso/act tomaban el $/kg del dBase
    #     y por eso no daba la cuenta).
    # TMT 2026-07-12 (dueña): valuar el inicial al COSTO DEL STOCK (≈ dBase), no
    # al promedio de importaciones. Ese promedio corría ALTO (2,986 vs 2,951 del
    # dBase) porque valuaba TODO el stock viejo al precio de las importaciones
    # recientes → inflaba ~119k. Con el costo del stock que ya carga el dBase/
    # cierre (stock_act_ukg ≈ 2,951) + las compras del mes al promedio, la tarifa
    # queda ≈ 2,954, pegada al dBase (decisión dueña: "dejá el 2954 estamos ok").
    # El promedio de importaciones queda de FALLBACK si no hay costo de stock.
    _open_ukg = _f(hl, "stock_act_ukg") or _f(hl, "stock_inic_ukg")
    if not _open_ukg:
        try:
            from modules.importaciones import service as _impsvc2
            _open_ukg = _impsvc2.promedio_hilado_usd_kg()
        except Exception:  # noqa: BLE001 -- fail-soft
            _open_ukg = None
    # KG: ingreso/egreso = MOVIMIENTO REAL de bodega 51 → cierra por telescopía
    #   inicial + ingreso + en máquinas − egreso = stock actual (kg).
    # El ingreso real de bodega (ing_hilo) = importaciones (compras) + reingresos
    # de lote; el egreso real (egr_hilo) = a tejer + otras salidas (ajustes).
    # $: promedio ponderado — la apertura y las IMPORTACIONES son los únicos
    # eventos de costo externo; los reingresos entran al promedio (neutros) y el
    # egreso/stock se valúan al promedio → la compra barata BAJA el $/kg y la
    # columna $ TAMBIÉN cierra: inic + ingreso − egreso + máquinas = stock act.
    # $/kg del hilado = MISMA variable que usa el balance (asinfo_service.
    # mov_hilado_valuacion) → el 2,954 es único para flujo y balance. Fallback a
    # la cuenta inline si Asinfo no responde. Dueña 2026-07-13.
    _hv_hil = {}
    try:
        from modules.asinfo import service as _asvc_hv
        if anio and mes:
            _hv_hil = _asvc_hv.mov_hilado_valuacion(int(anio), int(mes), _open_ukg) or {}
    except Exception:  # noqa: BLE001 -- fail-soft
        _hv_hil = {}
    _hv_ok = bool(_hv_hil.get("disponible"))  # solo usar la func si Asinfo respondió
    _avg_ukg = float(_hv_hil["avg_ukg"]) if _hv_ok else (
        (hi0 * _open_ukg + compras_us) / (hi0 + compras)
        if (hi0 + compras) else _open_ukg
    )
    hl["stock_inic_kg"] = hi0
    hl["stock_inic_ukg"] = _open_ukg
    hl["stock_inic_us"] = hi0 * _open_ukg
    # Dueña 2026-07-17: "Ingresos" de la banda = IMPORTACIONES RECIBIDAS — el
    # MISMO número que COMPRAS HILADO y que el chequeo ("que los 3 sean
    # iguales"). Los reingresos de lote (hilo que vuelve de tejeduría,
    # correcciones = ing_hilo − compras) NO son ingreso: se NETEAN contra el
    # egreso (consumo real = salidas − reingresos). La telescopía kg y $ se
    # preserva (se resta lo mismo del ingreso y del egreso, al promedio).
    # La fórmula del egreso vive en asinfo_service.hilado_egresos_mes — LA
    # MISMA que usa el balance para Materia Prima (dueña: "el usuario tiene
    # que poder ver de dónde viene"). Acá le pasamos los datos ya consultados.
    from modules.asinfo import service as _asvc_egr
    _hegr = _asvc_egr.hilado_egresos_mes(
        anio or 0, mes or 0,
        mov={"ingreso": ing_hilo, "egreso": egr_hilo},
        importaciones_kg=compras,
    )
    _reingresos_kg = float(_hegr.get("reingresos_kg") or 0.0)
    hl["ingresos_kg"] = compras
    hl["ingresos_us"] = compras_us
    hl["ingresos_ukg"] = (compras_us / compras) if compras else 0.0
    # Referencias (trazabilidad; se muestran al lado / en el chequeo):
    hl["ref_import_kg"] = compras          # importaciones recibidas del mes
    hl["ref_import_us"] = compras_us       # lo que pagamos por ellas
    hl["ref_import_ukg"] = (compras_us / compras) if compras else 0.0
    hl["ref_bodega_ing_kg"] = ing_hilo     # ingreso bruto real de bodega 51
    hl["ref_reingresos_kg"] = _reingresos_kg
    hl["ref_tejer_kg"] = hilo_consumido    # lo que se fue a tejer (órdenes)
    hl["egresos_kg"] = float(_hegr.get("egresos_kg") or 0.0) if _hegr.get("disponible") \
        else max(egr_hilo - _reingresos_kg, 0.0)
    hl["egresos_ukg"] = _avg_ukg
    hl["egresos_us"] = hl["egresos_kg"] * _avg_ukg
    # En máquinas (WIP) al $/kg de apertura; suma al stock actual (es stock nuestro).
    _maq_us = maq_hilado * _open_ukg
    hl["stock_act_kg"] = float(_hv_hil["stock_act_kg"]) if _hv_ok else (hi1 + maq_hilado)
    hl["stock_act_us"] = float(_hv_hil["stock_act_us"]) if _hv_ok else (hi1 * _avg_ukg + _maq_us)
    hl["stock_act_ukg"] = (
        hl["stock_act_us"] / hl["stock_act_kg"] if hl["stock_act_kg"] else _avg_ukg
    )

    # CRUDO — ingreso = cruda producida (real), egreso = crudo consumido a
    # tintura (real). El % = rendimiento: cruda producida / hilo consumido.
    # CRUDO — ingreso/egreso = MOVIMIENTO REAL de bodega 52 (deltas del saldo) →
    # cierra por telescopía igual que hilado/terminado: inicial + ingreso + en
    # máquinas − egreso = stock actual. "Cruda producida" (ci_crudo) y "a tinturar"
    # (crudo_consumido) quedan de referencia; el % de merma sigue con la producción.
    tj["stock_inic_kg"] = tc0
    tj["ingresos_kg"] = ing_crudo
    tj["egresos_kg"] = egr_crudo
    tj["ref_prod_kg"] = ci_crudo            # cruda producida (órdenes)
    tj["ref_tinturar_kg"] = crudo_consumido  # crudo consumido a tintura
    tj["ingresos_pct"] = (desp_crudo / hilo_consumido * 100.0) if hilo_consumido else 0.0
    tj["stock_act_kg"] = tc1 + maq_crudo

    # TERMINADO — ingreso/egreso = MOVIMIENTO REAL de la bodega (mismo saldo que
    # el stock) → la columna CIERRA exacto (pf0 + ingreso − egreso = pf1). El %
    # sigue siendo la merma de tintura (crudo consumido − PT producido).
    te["stock_inic_kg"] = pf0
    te["ingresos_kg"] = ing_term
    te["egresos_kg"] = egr_term
    te["ingresos_pct"] = (desp_term / crudo_consumido * 100.0) if crudo_consumido else 0.0
    te["stock_act_kg"] = pf1
    # Dueña 2026-07-09: mostrar AMBAS ventas — la derivada (arriba) y lo
    # FACTURADO directo de Asinfo — y la diferencia. El facturado no cierra la
    # columna; el gap es el desfase stock↔facturación.
    ventas_facturado = 0.0
    if anio and mes:
        try:
            from modules.asinfo import service as _asvc3
            ventas_facturado = float(_asvc3.ventas_facturado_kg(int(anio), int(mes)) or 0.0)
        except Exception:  # noqa: BLE001 -- fail-soft
            ventas_facturado = 0.0
    te["facturado_kg"] = ventas_facturado
    # Δ = egreso de bodega (movimiento real) − facturado. El facturado corre un
    # poco arriba del egreso de bodega por el desfase de facturación (se factura
    # algo que salió antes / NTEN). El despacho físico (documento) va aparte.
    te["facturado_diff"] = egr_term - ventas_facturado
    te["despacho_kg"] = despacho

    # COLORANTES — sin stock en Asinfo: se deja tal cual (valor PC).

    # ── Comparación "mostrar ambos por ahora así elegimos" (dueña 2026-07-09):
    #   · VENTAS: la DERIVADA (que cierra el terminado) vs la REAL de
    #     scintela.factura (kg físicos del mes).
    #   · COLORANTES: el valor del programa vs el stock de químicos de
    #     formulas_app (Σ stock_kg × precio_us).
    ventas_real = 0.0
    if anio and mes:
        try:
            import db as _db
            _vr = _db.fetch_one(
                """
                SELECT COALESCE(SUM(kg), 0) AS kg
                  FROM scintela.factura
                 WHERE EXTRACT(YEAR FROM fecha)  = %s
                   AND EXTRACT(MONTH FROM fecha) = %s
                   AND (stat IS NULL OR stat <> 'X')
                """,
                (int(anio), int(mes)),
            )
            ventas_real = float((_vr or {}).get("kg") or 0)
        except Exception:  # noqa: BLE001 -- fail-soft
            ventas_real = 0.0
    # DESCUBRIMIENTO: valor de stock de químicos de formulas_app POR FAMILIA,
    # para ver cuál es "colorante" (los auxiliares casi no se mueven — dueña).
    # OJO: esto es el stock ACTUAL (foto), no el movimiento inic/consumo/act del
    # mes; el número final de colorantes se arma con la familia correcta + el
    # inicial/consumo (siguiente iteración). Dueña 2026-07-09.
    # 2026-07-18: el loop (stock_quimicos + factor_iva por producto) vive en
    # quimicos_flujo.color_familias_valuadas — caché 240s + warmup (era parte
    # de los 7,4s de mov_asinfo_quimicos). Mismo cálculo, mismo resultado.
    color_formulas = None
    color_familias: dict = {}
    try:
        from modules.informes.quimicos_flujo import (
            color_familias_valuadas as _q_familias,
            color_movimiento_mes as _q_color_mov,
        )
        _fams = _q_familias()
        color_familias = dict(_fams) if _fams else {}
        color_formulas = sum(color_familias.values()) if color_familias else None
    except Exception:  # noqa: BLE001 -- fail-soft
        color_formulas = None
    _fam_top = sorted(color_familias.items(), key=lambda kv: kv[1], reverse=True)[:10]

    # MOVIMIENTO de colorantes desde formulas_app, con criterio (dueña "un poco
    # creativo", "fijate compras y cuántas órdenes"): el inventario tiene lecturas
    # muy espaciadas → no sirve para el movimiento del mes. Lo REAL está en:
    #   · CONSUMO del mes = colorante (familias POLI+ALG) usado en las órdenes de
    #     tintura TERMINADAS del mes (orden_lineas × precio). = egreso.
    #   · COMPRAS del mes = colorante comprado en el mes (compras × precio). = ingreso.
    #   · STOCK actual = valor de colorantes hoy (última lectura POLI+ALG).
    #   · STOCK inicial = actual + consumo − compras (para que cierre, como el dBase).
    # Todo de formulas_app vía formulas_db (fail-soft).
    _COLOR_FAMS = ("POLI", "ALG")   # dueña 2026-07-13: colorantes SIN auxiliares (338, no 393)
    # 2026-07-18: las 2 queries viven en quimicos_flujo.color_movimiento_mes
    # (caché 240s + warmup). Misma SQL, mismo resultado, mismo fail-soft.
    color_consumo = color_compras = color_stock = None
    color_ordenes = None
    if anio and mes:
        try:
            _cmov = _q_color_mov(int(anio), int(mes))
            if _cmov is not None:
                color_consumo = float(_cmov.get("consumo_us") or 0)
                color_ordenes = int(_cmov.get("n_ordenes") or 0)
                color_compras = float(_cmov.get("compras_us") or 0)
        except Exception:  # noqa: BLE001 -- fail-soft
            color_consumo = color_compras = None
    # Stock actual de colorantes (POLI+ALG) = suma del stock foto de hoy.
    try:
        color_stock = sum(v for k, v in color_familias.items() if k in _COLOR_FAMS)
    except Exception:  # noqa: BLE001
        color_stock = None
    # COLORANTES 100% de formulas_app "como viene" (dueña 2026-07-09, confirmado
    # por el equipo: "se usa formulas app como viene, está ok"). Colorante puro
    # POLI+ALG, sin auxiliares, sin dBase:
    #   egreso  = consumo de colorante en las órdenes de tintura del mes.
    #   compras = colorante comprado en el mes.
    #   stock   = inventario físico de colorante hoy.
    #   inicial = stock + consumo − compras (deriva para que cierre).
    color_inic_der = None
    if color_stock is not None and color_consumo is not None and color_compras is not None:
        color_inic_der = color_stock + color_consumo - color_compras

    cmp = {
        "ventas_derivada": ventas,
        "ventas_real": ventas_real,
        "en_proceso": ventas_real - ventas,   # "el restante" si se usan ventas reales
        "color_consumo": color_consumo,
        "color_familias": _fam_top,
        "color_compras": color_compras,
        "color_stock": color_stock,
        "color_inic_der": color_inic_der,
        "color_ordenes": color_ordenes,
    }

    # ── MODELO DE STOCK DE QUÍMICOS (programa vs formulas) ──────────────────
    # TMT 2026-07-09 (dueña): "necesito mi inicial de químicos (programa) +
    # compras químicos programa − egresos formulas = final programa, y luego
    # ajuste contra final formulas". El ajuste NO se postea, se MUESTRA.
    #   · inicial (programa) = iniciales.vq del mes anterior (semilla dBase).
    #     Es el mismo VQ0 que usa el balance (VQX = VQ0 + VQQ − ITIN).
    #   · compras (programa) = compras tipo 'Q' del mes (scintela.compra),
    #     las que carga el botón "Cargar químicos". Mismo filtro que VQQ.
    #   · egresos (formulas) = consumo de químicos de las órdenes del mes
    #     (color_consumo, POLI+ALG+AUX de orden_lineas).
    #   · final (programa) = inicial + compras − egresos (computado).
    #   · final (formulas) = stock físico vivo hoy (color_stock, foto).
    #   · ajuste = final formulas − final programa (la varianza, se muestra).
    # TMT 2026-07-12 (dueña): la banda descuadraba −66.100 por DOS datos mal
    # tomados que se tapaban a medias (verif live formulas_app db3):
    #   1. INICIAL salía de iniciales.vq del dBase (144.637), un libro HEREDADO
    #      que PC arrastra sin recalcular (write-back lee su propio vqx=vq) →
    #      quedó viejo. El físico REAL de químicos (formulas, POLI+ALG+AUX) al
    #      cierre de junio era 417.943, no 144.637.
    #   2. COMPRAS metía las 2 facturas de importación tipo Q (Colourtex 241.330,
    #      09/07), PERO esa tintura ya había entrado al físico de formulas el
    #      18/06 (229.189 en 18 líneas) → contarla como compra de julio la
    #      DUPLICA sobre un inicial que ya la incluye.
    # Fix: inicial = físico al-día de formulas al cierre del mes anterior;
    # compras = lo que REALMENTE entró a bodega este mes (compras de formulas,
    # no la factura); físico = al-día de hoy (no la foto cruda vieja). Cierra
    # al ~1% (el ajuste queda como varianza real de inventario). La factura de
    # importación se muestra como memo informativo (facturado_prog).
    quimicos_modelo = None   # COLUMNA QUÍM.$ (modelo A / programa)
    quimicos_banda = None    # BANDA (solo formulas / físico colorante)
    if anio and mes:
        try:
            import calendar as _cal3
            from datetime import date as _date3
            from datetime import timedelta as _td3

            from filters import today_ec as _today_ec3
            from modules.tintura import service as _tsvc3

            # FÍSICO de colorante = MISMA variable que el balance (POLI+ALG, sin
            # AUX) → tintura_service.stock_colorante_fisico. NO recalcular acá
            # (dueña 2026-07-13: "stock quimicos idem que hilado, la variable del
            # flujo"). El balance (vqx) llama a la misma función → 338 único.
            # 2026-07-18: vía quimicos_flujo (caché 240s + warmup) — era parte
            # de los 7,4s de mov_asinfo_quimicos.
            from modules.informes.quimicos_flujo import (
                consumo_quimico_desglose as _q_desglose,
                fisico_colorante_al_dia as _q_fisico,
            )

            def _fisico_quimicos_aldia(_corte):
                _v = _q_fisico(_corte)
                if _v is None:  # fallback directo si el caché/módulo falló
                    _v = _tsvc3.stock_colorante_fisico(_corte)
                return float(_v or 0)

            _corte_ini = _date3(int(anio), int(mes), 1) - _td3(days=1)
            _last = _date3(int(anio), int(mes),
                           _cal3.monthrange(int(anio), int(mes))[1])
            _corte_fin = min(_last, _today_ec3())

            # MODELO A (dueña 2026-07-16): LIBRO CONTABLE del programa
            # reconciliado al físico. El inicial (VQ0) y las compras (tipo Q =
            # TODO el químico facturado) las pone el PROGRAMA; el consumo sale
            # del TINTURADO DIARIO de formulas — TODO el químico (POLI+ALG+AUX)
            # por fecha de tinturado, que coincide con el ITIN del dBase (~114k),
            # NO el color_consumo de arriba (solo colorante por fecha_terminado,
            # subcuenta ~40k). El ajuste = físico − libro es la revaluación real
            # (~+20k). "Copiamos lo que tiene el programa y el ajuste nos ajusta
            # al físico."
            _vq0 = float((header.get("colorantes") or {}).get("stock_inic_us") or 0)

            _qc = db.fetch_one(
                """
                SELECT COALESCE(SUM(importe), 0) AS importe,
                       COUNT(*)                  AS n
                  FROM scintela.compra
                 WHERE EXTRACT(YEAR  FROM fecha) = %s
                   AND EXTRACT(MONTH FROM fecha) = %s
                   AND UPPER(COALESCE(tipo, '')) = 'Q'
                   AND COALESCE(stat, '') NOT IN ('X', 'Y')
                   AND COALESCE(usuario_crea, '') <> 'asinfo-backfill'
                """,
                (int(anio), int(mes)),
            ) or {}
            _compras_prog = float(_qc.get("importe") or 0)

            # Consumo del mes = tinturado diario de formulas, TODO el químico
            # (POLI+ALG+AUX) por fecha de tinturado (ordenes.fecha 'DD/MM/YYYY').
            # DESGLOSE para reconciliar con la tintorería de abajo — que solo
            # costea el teñido con kg de tela cargado y sin lavados:
            #   costeado = teñido con kg de tela (= Total de COSTOS DE TINTORERÍA)
            #   proceso  = teñido sin cerrar la tela (sin kg) → trabajo en proceso
            #   lavado   = órdenes de lavado
            # total = costeado + proceso + lavado = egreso de químico del mes.
            # 2026-07-18: el desglose vive en quimicos_flujo (caché 240s +
            # warmup) — la query sobre orden_lineas (TO_DATE sobre texto) era
            # el grueso de los 7,4s de mov_asinfo_quimicos. Misma SQL, mismo
            # resultado; fail-soft idéntico (None → respaldo).
            _consumo_prog = None
            _egr_cost = _egr_proc = _egr_lav = 0.0
            _desg = _q_desglose(int(anio), int(mes))
            if _desg is not None:
                _egr_cost = float(_desg.get("costeado") or 0)
                _egr_proc = float(_desg.get("proceso") or 0)
                _egr_lav = float(_desg.get("lavado") or 0)
                _consumo_prog = _egr_cost + _egr_proc + _egr_lav

            q_inicial = _vq0
            q_compras = _compras_prog
            # Consumo del mes = PROYECTADO de tintura (mismo número que la tabla
            # de abajo). Dueña 2026-07-16: un solo número de químico en toda la
            # pantalla, sin lavados ni desglose. Si formulas no dio proyectado,
            # cae al tinturado total del programa (_consumo_prog) de respaldo.
            q_egresos = (float(proy_quimico) if proy_quimico is not None
                         else _consumo_prog)
            # En máquinas (dueña 2026-07-17): químico ya dosificado en órdenes
            # SIN cerrar la tela (_egr_proc del desglose de arriba). Con el
            # consumo = costeado (tela cerrada), este monto ya salió del
            # estante pero el consumo no lo descuenta → se resta en la cuenta,
            # a la vista, y el ajuste queda midiendo solo arranque + merma.
            # En meses cerrados tiende a 0 solo (la tela se va cerrando).
            q_en_maquinas = float(_egr_proc or 0) if _consumo_prog is not None else 0.0
            q_final_prog = (q_inicial + q_compras - float(q_egresos or 0)
                            - q_en_maquinas)
            q_final_form = _fisico_quimicos_aldia(_corte_fin)   # físico formulas
            q_ajuste = q_final_form - q_final_prog               # de arranque (cutover)

            quimicos_modelo = {
                "inicial": q_inicial,
                "compras": q_compras,
                "compras_n": int(_qc.get("n") or 0),
                "egresos": q_egresos,
                "en_maquinas": q_en_maquinas,
                "final_prog": q_final_prog,
                "final_form": q_final_form,
                "ajuste": q_ajuste,
                "facturado_prog": _compras_prog,
                "facturado_n": int(_qc.get("n") or 0),
            }

            # BANDA "stock de químicos" = MISMO número que la columna QUÍM.$.
            # Dueña 2026-07-16: unificar — un solo consumo (proyectado) y un
            # solo ajuste (de arranque, cutover dBase→programa) arriba y abajo.
            quimicos_banda = dict(quimicos_modelo)
        except Exception:  # noqa: BLE001 -- fail-soft, no rompe la vista
            quimicos_modelo = None
            quimicos_banda = None

    # COLUMNA QUÍM.$ de la tabla de movimientos: inicial VQ0 + compras tipo Q −
    # consumo (= PROYECTADO de tintura, el mismo número que la tabla de abajo),
    # ajuste = físico − libro. Dueña 2026-07-16: un solo número de químico y un
    # solo ajuste (de arranque) en toda la pantalla; la banda de abajo lo repite.
    if quimicos_modelo and quimicos_modelo.get("final_form") is not None:
        co["stock_inic_us"] = round(float(quimicos_modelo.get("inicial") or 0), 0)
        co["ingresos_us"] = round(float(quimicos_modelo.get("compras") or 0), 0)
        co["egresos_us"] = round(float(quimicos_modelo.get("egresos") or 0), 0)
        co["ajuste_us"] = round(float(quimicos_modelo.get("ajuste") or 0), 0)   # físico − libro (de arranque)
        co["maquinas_us"] = round(float(quimicos_modelo.get("en_maquinas") or 0), 0)
        co["stock_act_us"] = round(float(quimicos_modelo["final_form"] or 0), 0)

    return {
        "hilado": hl, "tejido": tj, "terminado": te, "colorantes": co,
        "cmp": cmp, "quimicos_modelo": quimicos_modelo,
        "quimicos_banda": quimicos_banda,
        "maquinas": {
            "hilado": maq_hilado, "crudo": maq_crudo,
            "hilado_ukg": hl.get("stock_inic_ukg") or 0,          # WIP al $/kg de apertura
            "hilado_us": maq_hilado * (hl.get("stock_inic_ukg") or 0),
        },
        "desperdicio": {"crudo": desp_crudo, "term": desp_term},
    }


def _asof_dia_overrides(comp: dict, as_of) -> None:
    """Afina TOTC/TOTF de balance_components_as_of() para fotos DÍA a día.

    balance_components_as_of() fue pensada para cierres de mes y su TOTC usa
    COALESCE(fecha_recibido, fecha) — pero `fecha` en cheques es la fecha de
    COBRO (posfechada, futura) → para "ayer" excluía casi toda la cartera
    (TMT 2026-06-12: daba 33k contra 2,18M reales). Acá:

      · TOTC = cheques que EXISTÍAN al as_of (fecha_crea/fecha_recibido) y
        que o siguen vivos hoy (Z,1,2,3,P,D), o fueron depositados DESPUÉS
        (fechaing > as_of), o salieron de cartera DESPUÉS por cobro en
        efectivo / endoso / terminal (fechaout > as_of). Anulados X/Y no
        se resucitan a propósito: son typos corregidos, el pasado que
        queremos mostrar es el corregido.
      · TOTF = fórmula canónica de totf() (saldo NETO, sin filtro de signo,
        sin asinfo-backfill) + fecha <= as_of + creada antes del as_of.
        Aproximación: usa el saldo ACTUAL (no rebobina abonos).

    Recalcula cart/subt/totl/patr/utilidad con los valores afinados.
    NO toca queries.balance_components_as_of — la usan los snapshots de
    historia y /fuentes-y-usos (coordinar con Federico antes de cambiarla).
    """
    totc_row = db.fetch_one(
        """
        SELECT COALESCE(SUM(importe), 0) AS total
          FROM scintela.cheque
         WHERE (fecha_crea IS NULL OR fecha_crea::date <= %s)
           AND (fecha_recibido IS NULL OR fecha_recibido <= %s)
           AND ( stat IN ('Z','1','2','3','P','D')
                 OR (stat IN ('B','A') AND fechaing IS NOT NULL AND fechaing > %s)
                 OR (stat IN ('C','9','E','T') AND fechaout IS NOT NULL AND fechaout > %s) )
        """,
        (as_of, as_of, as_of, as_of),
    ) or {}
    # TOTF con rewind de abonos PC: cada aplicación hecha POR EL PROGRAMA
    # queda fechada en scintela.chequesxfact (fechaing) — se re-suman las
    # posteriores al as_of, y eso también resucita facturas que pasaron a
    # T (canceladas) DESPUÉS de la fecha. Los abonos que llegaron por sync
    # del dBase no traen historia → para esos vale el saldo actual. Con
    # operación 100% en PC la foto es exacta hacia adelante.
    totf_row = db.fetch_one(
        """
        SELECT COALESCE(SUM(
                 CASE WHEN (f.stat IS NULL OR f.stat IN ('Z','A','',' '))
                      THEN COALESCE(f.saldo, 0) ELSE 0 END
                 + COALESCE(ab.post, 0)
               ), 0) AS total
          FROM scintela.factura f
          LEFT JOIN (
                SELECT id_fact, SUM(COALESCE(importe, 0)) AS post
                  FROM scintela.chequesxfact
                 WHERE fechaing > %s
                 GROUP BY id_fact
               ) ab ON ab.id_fact = f.id_factura
         WHERE COALESCE(f.stat, '') NOT IN ('X','Y')
           AND COALESCE(f.usuario_crea, '') <> 'asinfo-backfill'
           AND f.fecha <= %s
           AND (f.fecha_crea IS NULL OR f.fecha_crea::date <= %s)
        """,
        (as_of, as_of, as_of),
    ) or {}
    comp["totc"] = float(totc_row.get("total") or 0)
    comp["totf"] = float(totf_row.get("total") or 0)
    comp["cart"] = comp["totc"] + comp["totf"]
    # banco (= bancos + caja) ya viene all-in de balance_components_as_of.
    comp["subt"] = comp["banco"] + comp["cart"]
    comp["totl"] = (comp["subt"] + comp["vsto"] + comp["vqx"]
                    + comp["umaq"] + comp["uact"] + comp["antic"])
    comp["patr"] = comp["patrimonio"] = comp["totl"] - comp["totp"]
    comp["utilidad"] = comp["usuti"] = (comp["patr"] - comp["patant"]) + comp["usret"]


@informes_bp.route("/balance")
@requiere_login
@requiere_permiso("informes.ver")
def balance():
    # ?as_of=AAAA-MM-DD → foto histórica SOLO LECTURA del balance (pedido
    # dueña 2026-06-12: "ver el balance como estaba ayer"). Usa
    # balance_components_as_of() — la misma cuenta de los snapshots y de
    # /fuentes-y-usos, con sus aproximaciones (saldo de factura ACTUAL,
    # stock/activos del último cierre mensual anterior). Importante:
    # mirar el pasado NO escribe nada — se saltea provisiones diarias,
    # persist YY y auto-cierre, que sí corren en el balance live.
    as_of_raw = (request.args.get("as_of") or "").strip()
    if as_of_raw:
        from datetime import date as _date

        try:
            as_of = _date.fromisoformat(as_of_raw)
        except ValueError:
            flash(f"Fecha inválida: {as_of_raw} — uso AAAA-MM-DD.", "warning")
            return redirect(url_for("informes.balance"))
        if as_of >= today_ec():
            # Hoy o futuro → balance live normal.
            return redirect(url_for("informes.balance"))
        comp, error = _safe(lambda: queries.balance_components_as_of(as_of), {})
        if not error:
            _asof_dia_overrides(comp, as_of)
        return render_template(
            "informes/balance_as_of.html",
            c=comp,
            as_of=as_of,
            hoy=today_ec(),
            error=error,
        )

    # Provisiones diarias automáticas (replica MENU.PRG L282-333).
    # Idempotente — sólo aplica si HOY > última fecha guardada y no es
    # domingo. Si falla, no rompe el balance — la migración de la tabla
    # sistema_meta puede no haber corrido todavía (decorador defensivo).
    #
    # Persistir acumulación YY/RT en el importe guardado (dBase REPLACE DAILY).
    # Idempotente (baseline=hoy → no-op). Sin esto el Pasivos YY queda congelado
    # bajo el dBase. TMT 2026-06-05.
    try:
        from modules.posdat.queries import persistir_acumulacion_yy
        persistir_acumulacion_yy()
    except Exception:  # noqa: BLE001
        # NO silenciar del todo: el except:pass mudo escondió 5 días que el
        # persist no corría (columna fantasma) y los Pasivos drifteaban
        # 32k/día. Si esto loguea, hay que mirarlo YA. TMT 2026-06-10.
        import logging
        logging.getLogger("programa_core.posdat").exception(
            "persistir_acumulacion_yy FALLÓ — Pasivos YY van a driftear vs dBase"
        )

    # Provisiones diarias: SOLO automáticas (catch-up de días hábiles pendientes).
    # El bypass GET ?forzar_provisiones=1 se removió — un GET no debe mutar estado
    # financiero (un refresh/prefetch/favorito lo disparaba). El forzado manual
    # sigue disponible para scripts vía correr_provisiones_diarias(forzar=True).
    # TMT 2026-07-11.
    try:
        prov_result = queries.correr_provisiones_diarias()
    except Exception as e:  # noqa: BLE001
        prov_result = {"aplicado": False, "error": str(e)}

    # ITEM #5 — Auto-cierre de stock mensual (replica MENU.PRG L246-263).
    # Idempotente. Si ya se cerró el mes destino, no hace nada. Si falla,
    # no rompe el balance — la tabla scintela.sistema_meta puede no estar
    # inicializada todavía. Decorador defensivo.
    try:
        from modules.iniciales.views import auto_cerrar_mes_si_corresponde

        auto_cerrar_mes_si_corresponde()
    except Exception as e:  # noqa: BLE001
        {"aplicado": False, "error": str(e)}

    data, error = _safe(queries.informe_balance, {})
    return render_template(
        "informes/balance.html",
        b=data,
        error=error,
        provisiones=prov_result,
    )


# Feature A — tab Compras en /informes/balance (TMT 2026-05-19 v6).
@informes_bp.route("/balance/compras")
@requiere_login
@requiere_permiso("informes.ver")
def balance_compras():
    """Drill-down de compras del período. Reuse de /informes/balance."""

    hoy = today_ec()
    try:
        anio = int(request.args.get("anio") or hoy.year)
    except (TypeError, ValueError):
        anio = hoy.year
    try:
        mes = int(request.args.get("mes") or hoy.month)
    except (TypeError, ValueError):
        mes = hoy.month
    mes = max(1, min(mes, 12))
    prov = (request.args.get("prov") or "").strip().upper() or None
    try:
        num_v = int(request.args.get("v") or 0) or None
    except (TypeError, ValueError):
        num_v = None
    try:
        data = queries.compras_del_periodo(
            anio=anio,
            mes=mes,
            prov=prov,
            num_v=num_v,
        )
        error = None
    except Exception as e:  # noqa: BLE001
        data, error = (
            {
                "filas": [],
                "total_importe": 0,
                "total_kg": 0,
                "n_filas": 0,
                "prov_options": [],
                "anio": anio,
                "mes": mes,
                "prov_actual": prov,
                "num_v_actual": num_v,
            },
            str(e),
        )
    return render_template(
        "informes/balance_compras.html",
        data=data,
        anio=anio,
        mes=mes,
        prov=prov,
        num_v=num_v,
        error=error,
    )


# Feature B — matriz histórica TINT.BAT (TMT 2026-05-19 v7).
# TMT 2026-05-20 — refactor: ahora la vista DEFAULT es la matriz fija 5+1
# (5 meses pasados + mes actual con múltiples snapshots para comparar).
# Toma snapshot del mes actual al entrar (throttle 1h) y permite validar
# o borrar cada snapshot del mes actual.
@informes_bp.route("/historico-12m")
@requiere_login
@requiere_permiso("informes.ver")
def historico_12m():
    """Matriz fija 5 meses pasados + mes actual (con N snapshots).

    Pedido dueña 2026-05-20: la pantalla siempre muestra los últimos 5
    meses cerrados + el mes actual. Al entrar, toma un snapshot nuevo
    del mes actual (sin pisar el anterior, throttle 1h) para que la
    dueña pueda comparar. Cada snapshot del mes actual puede ser
    "validado" (deja éste, borra el resto) o "borrado" (solo éste).

    Query params (todos opcionales):
      modo    "matriz" (default), "mom" (mes vs mes — modo viejo).
      a_a/m_a/a_b/m_b: para modo "mom".
    """
    modo = (request.args.get("modo") or "matriz").strip().lower()
    if modo not in ("matriz", "mom"):
        modo = "matriz"

    error = None
    data: dict = {}
    mom: dict = {}
    snap_info: dict = {}
    meses_disponibles: list[tuple[int, int]] = []

    if modo == "mom":
        # Defaults: comparar mes actual (b) vs mes anterior (a).

        hoy = today_ec()

        def _parse_par(prefix: str, default_anio: int, default_mes: int):
            try:
                a_ = int(request.args.get(f"a_{prefix}") or default_anio)
                m_ = int(request.args.get(f"m_{prefix}") or default_mes)
            except (TypeError, ValueError):
                return default_anio, default_mes
            return a_, max(1, min(12, m_))

        mes_actual_a, mes_actual_m = hoy.year, hoy.month
        prev_a, prev_m = mes_actual_a, mes_actual_m - 1
        if prev_m < 1:
            prev_m = 12
            prev_a -= 1
        a_a, m_a = _parse_par("a", prev_a, prev_m)
        a_b, m_b = _parse_par("b", mes_actual_a, mes_actual_m)
        try:
            mom = queries.historico_mom(a_a, m_a, a_b, m_b)
            meses_disponibles = queries.historico_meses_disponibles()
        except Exception as e:  # noqa: BLE001
            mom = {"par_a": (a_a, m_a), "par_b": (a_b, m_b), "lineas": [], "meses_sin_snap": []}
            error = str(e)
    else:
        # Federico 2026-05-21 -- foto automatica al entrar (reactivada).
        # El bug que la habia desactivado (snapshot con ktej/ktin en 0)
        # quedo resuelto con el carry-forward en insertar_snapshot.
        # Flujo: (1) tomar a lo sumo UNA foto por dia (throttle 24h) para que
        # las columnas sean "ayer vs hoy" y no "hace 3 minutos"; (2) consolidar
        # dejando las 2 columnas mas recientes (la previa + la de hoy).
        # TMT 2026-06-04 (Bug #3): el throttle era 180s -> se creaba una columna
        # nueva por cada visita >3min y las 2 columnas salian casi identicas
        # (delta ~0). Con 24h la comparacion intra-mes tiene sentido. Para
        # forzar una foto fuera de hora esta el boton "Snapshot ahora".
        try:
            snap_info = queries.tomar_snapshot_mes_actual(
                usuario=(g.user or {}).get("username", "web"),
                throttle_segundos=86400,
            )
        except Exception as e:  # noqa: BLE001
            snap_info = {"accion": "error", "error": str(e)}
        try:
            queries.consolidar_snapshots_mes_actual(conservar=2)
        except Exception:  # noqa: BLE001
            pass
        try:
            data = queries.historico_5m_con_actual(max_actual=3)
        except Exception as e:  # noqa: BLE001
            data = {"columnas": [], "lineas": [], "meses_sin_snap": [], "n_actual": 0, "hoy": None}
            error = str(e)

    return render_template(
        "informes/historico_12m.html",
        data=data,
        mom=mom,
        modo=modo,
        meses_disponibles=meses_disponibles,
        error=error,
        snap_info=snap_info,
    )


@informes_bp.route("/historico-12m/_api/<int:id_historia>/validar", methods=["POST"])
@requiere_login
@requiere_permiso("informes.editar")  # TMT 2026-06-03 audit: era .ver — borra los OTROS snapshots del mes
def historico_validar(id_historia: int):
    """Marca un snapshot como canónico — borra los OTROS del mismo mes."""
    try:
        r = queries.validar_snapshot(
            id_historia,
            usuario=(g.user or {}).get("username", "web"),
        )
        return jsonify({"ok": True, **r})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:  # noqa: BLE001
        return jsonify({"ok": False, "error": f"No pude validar: {e}"}), 500


@informes_bp.route("/historico-12m/_api/<int:id_historia>/borrar", methods=["POST"])
@requiere_login
@requiere_permiso("informes.editar")  # TMT 2026-06-03 audit: era .ver — DELETE destructivo
def historico_borrar(id_historia: int):
    """Borra UN snapshot específico de scintela.historia."""
    try:
        n = queries.borrar_snapshot(id_historia)
        return jsonify({"ok": True, "n_borrados": int(n or 0)})
    except Exception as e:  # noqa: BLE001
        return jsonify({"ok": False, "error": f"No pude borrar: {e}"}), 500


@informes_bp.route("/historico-12m/_api/snapshot-ahora", methods=["POST"])
@requiere_login
@requiere_permiso("informes.editar")  # TMT 2026-06-03 audit: era .ver — INSERT en historia
def historico_snapshot_ahora():
    """Fuerza un snapshot del mes actual ignorando el throttle de 24h.

    TMT 2026-05-20 — pedido dueña: cuando los KPIs muestran 0 porque
    el snapshot viejo se tomó con la lógica vieja, queremos un botón
    que rehace el snapshot con la lógica nueva sin esperar 24h.
    """
    try:
        usuario = (g.user or {}).get("username", "web")
        # Pasamos throttle_segundos=0 para que se inserte sí o sí.
        r = queries.tomar_snapshot_mes_actual(usuario=usuario, throttle_segundos=0)
        return jsonify({"ok": True, **r})
    except Exception as e:  # noqa: BLE001
        return jsonify({"ok": False, "error": f"No pude crear snapshot: {e}"}), 500


@informes_bp.route("/historico-12m/_api/eliminar-ultima", methods=["POST"])
@requiere_login
@requiere_permiso("informes.editar")  # TMT 2026-06-03 audit: era .ver — DELETE destructivo
def historico_eliminar_ultima():
    """Borra la columna mas reciente del mes actual (boton Eliminar ultima)."""
    try:
        r = queries.eliminar_ultima_columna_mes_actual()
        return jsonify({"ok": True, **r})
    except Exception as e:  # noqa: BLE001
        return jsonify({"ok": False, "error": f"No pude eliminar: {e}"}), 500


@informes_bp.route("/balance/utilidad-debug")
@requiere_login
@requiere_permiso("informes.ver")
def utilidad_debug():
    """Diagnóstico para identificar por qué la UTILIDAD del balance no
    coincide con el dBase. Muestra:
      - Fila de scintela.historia usada como PATANT (todos los campos)
      - Componentes de PATR (subt, vsto, vqx, umaq, uact, uret, antic, totp)
      - 4 fórmulas alternativas con su resultado, para que el gerente
        identifique cuál da el número correcto.
    """
    import db

    data, error = _safe(queries.informe_balance, {})

    # Levantar la fila de historia que se usa como PATANT
    hist_row = (
        db.fetch_one(
            """
        SELECT *
        FROM scintela.historia
        WHERE fecha < date_trunc('month', (CURRENT_TIMESTAMP - INTERVAL '5 hours')::date)::date
        ORDER BY fecha DESC
        LIMIT 1
        """
        )
        or {}
    )

    # Levantar TODAS las filas de historia (últimas 24) para auditar qué
    # está cargado y comparar con el dBase. Usuario reportó 2026-05-06:
    # April=20,115,887, Marzo=176,556,980 (probable typo, debería ser
    # ~17,655,698 dado que April es 20M).
    hist_all = (
        db.fetch_all(
            """
        SELECT fecha, patrimonio, ustock, uqui, usret, usuti, kvent, uvent
        FROM scintela.historia
        ORDER BY fecha DESC
        LIMIT 24
        """
        )
        or []
    )

    # Componentes de PATR
    componentes = {
        "subt": float(data.get("subt") or 0),
        "vsto_display": float(data.get("vsto") or 0),  # post-override
        "vqx": float(data.get("vqx") or 0),
        "umaq": float(data.get("umaq") or 0),
        "uact": float(data.get("uact") or 0),
        "uret": float(data.get("uret") or 0),
        "antic": float(data.get("antic") or 0),
        "totp": float(data.get("totp") or 0),
    }
    vsto_orig = float(hist_row.get("ustock") or 0)
    componentes["vsto_orig"] = vsto_orig

    # PATR alternativos
    patr_post = (
        componentes["subt"]
        + componentes["vsto_display"]
        + componentes["vqx"]
        + componentes["umaq"]
        + componentes["uact"]
        + componentes["uret"]
        + componentes["antic"]
        - componentes["totp"]
    )
    patr_pre = (
        componentes["subt"]
        + componentes["vsto_orig"]
        + componentes["vqx"]
        + componentes["umaq"]
        + componentes["uact"]
        + componentes["uret"]
        + componentes["antic"]
        - componentes["totp"]
    )

    patrimonio_hist = float(hist_row.get("patrimonio") or 0)
    usret_hist = float(hist_row.get("usret") or 0)
    usuti_hist = float(hist_row.get("usuti") or 0)

    # Fórmulas alternativas
    formulas = [
        {
            "label": "A) patr_pre_override − patrimonio_hist (= lo actual)",
            "patr": patr_pre,
            "patant": patrimonio_hist,
            "result": patr_pre - patrimonio_hist,
        },
        {
            "label": "B) patr_post_override − patrimonio_hist",
            "patr": patr_post,
            "patant": patrimonio_hist,
            "result": patr_post - patrimonio_hist,
        },
        {
            "label": "C) patr_pre_override − (patrimonio_hist − usret_hist)",
            "patr": patr_pre,
            "patant": patrimonio_hist - usret_hist,
            "result": patr_pre - (patrimonio_hist - usret_hist),
        },
        {
            "label": "D) patr_post_override − (patrimonio_hist − usret_hist)",
            "patr": patr_post,
            "patant": patrimonio_hist - usret_hist,
            "result": patr_post - (patrimonio_hist - usret_hist),
        },
        {
            "label": "E) usuti_hist (= utilidad guardada en cierre anterior, sin recalcular)",
            "patr": 0,
            "patant": 0,
            "result": usuti_hist,
        },
        {
            "label": "F) patr_pre_override − patrimonio_hist + usret_hist (= delta + retiros del cierre)",
            "patr": patr_pre,
            "patant": patrimonio_hist - usret_hist,
            "result": patr_pre - patrimonio_hist + usret_hist,
        },
        {
            "label": "G) patr_post_override − patrimonio_hist + usret_hist",
            "patr": patr_post,
            "patant": patrimonio_hist - usret_hist,
            "result": patr_post - patrimonio_hist + usret_hist,
        },
    ]

    return render_template(
        "informes/utilidad_debug.html",
        hist_row=hist_row,
        hist_all=hist_all,
        componentes=componentes,
        patr_pre=patr_pre,
        patr_post=patr_post,
        patrimonio_hist=patrimonio_hist,
        usret_hist=usret_hist,
        usuti_hist=usuti_hist,
        formulas=formulas,
        error=error,
    )


@informes_bp.route("/cartera")
@requiere_login
@requiere_permiso("informes.ver")
def cartera():
    filas, error = _safe(queries.cartera_por_cliente, [])
    if request.args.get("export") == "csv":
        return csv_response(
            filas,
            columnas=[
                ("codigo_cli", "Código"),
                ("nombre", "Cliente"),
                ("n_facturas", "# facturas"),
                ("saldo_total", "Saldo"),
                ("factura_mas_vieja", "Fact. más vieja"),
                ("vence_mas_viejo", "Vence más vieja"),
            ],
            filename="cartera_clientes.csv",
        )
    total = sum(float(r["saldo_total"] or 0) for r in filas)
    return render_template("informes/cartera.html", filas=filas, total=total, error=error)


@informes_bp.route("/check-totales")
@requiere_login
@requiere_permiso("informes.ver")
def check_totales():
    """Diagnóstico de consistencia entre pantallas.

    TMT 2026-05-20 — pedido dueña: "necesito hacer un check de los totales,
    no me gusta llegar a una pantalla y ver otra cosa". Compara cada
    total cruzado entre vistas y marca diffs con badge rojo.

    Cruces verificados:
      A. Cartera bruta:    /cartera.total  vs  Resultados (TOTC + TOTF)
      B. Pasivos:          /deudas.total   vs  Resultados.TOTP
      C. Posdat (deuda):   /posdat default vs  /deudas.total
      D. TOTC (sin 'A'):   informes.totc() vs  cheques live Z+1+2+3+P+D
      E. TOTF:             informes.totf() vs  facturas Z+A live
    """
    import db as _db

    error = None
    checks: list[dict] = []

    def _diff_check(
        label: str, a_label: str, a_val: float, b_label: str, b_val: float, ok_tol: float = 0.5
    ) -> dict:
        a_f = float(a_val or 0)
        b_f = float(b_val or 0)
        diff = a_f - b_f
        ok = abs(diff) <= ok_tol
        return {
            "label": label,
            "a_label": a_label,
            "a_val": a_f,
            "b_label": b_label,
            "b_val": b_f,
            "diff": diff,
            "ok": ok,
        }

    try:
        # ─── Building blocks (queries canónicas) ─────────────────────
        totc = queries.totc()
        totf = queries.totf()
        totp = queries.posdat_totales()["totp"]

        # /cartera total (bruto)
        from modules.cartera import queries as _cq

        cartera_tot = _cq.aging_totales()
        # TMT 2026-05-20 v4 Federico — usar el saldo NETO (incluye sobrepagos)
        # para que el check matchee TOTF de Resultados. /cartera muestra
        # saldo_facturas (positivos) pero el check compara contra TOTF que
        # netea los sobrepagos. Sin esto, hay drift = SUM(saldo<0 stat=Z|A).
        cartera_facturas = float(cartera_tot.get("saldo_facturas_net") or 0)
        cartera_sobrepagos = float(cartera_tot.get("sobrepagos") or 0)
        cartera_cheques = float(cartera_tot.get("cheques_en_cartera") or 0)

        # /deudas total
        deudas_filas = queries.deudas_por_proveedor()
        deudas_total = sum(float(r.get("saldo_total") or 0) for r in deudas_filas)

        # /posdat default (= banc=0 + no anulada)
        from modules.posdat import queries as _pq

        posdat_resumen = _pq.resumen(solo_abiertas=True, tab="posdatados")
        posdat_resumen_yy = _pq.resumen(solo_abiertas=True, tab="yy")
        posdat_total_no_yy = float(posdat_resumen.get("total_abierto") or 0)
        posdat_total_yy = float(posdat_resumen_yy.get("total_abierto") or 0)

        # Cheques live (mismo filtro que totc — sanity).
        chq_live = _db.fetch_one(
            "SELECT COALESCE(SUM(importe), 0) AS t "
            "FROM scintela.cheque "
            "WHERE stat IN ('Z','1','2','3','P','D')"
        )
        cheques_live = float((chq_live or {}).get("t") or 0)

        fact_live = _db.fetch_one(
            "SELECT COALESCE(SUM(saldo), 0) AS t "
            "FROM scintela.factura "
            "WHERE stat IS NULL OR stat IN ('Z','A','',' ')"
        )
        facturas_live = float((fact_live or {}).get("t") or 0)

        # ─── Construir los checks ────────────────────────────────────
        # TMT 2026-05-20 v3 — labels SIEMPRE referencian "Resultados → X".
        # IMPORTANTE: NO comparamos "Subtotal Cartera" vs "/cartera total"
        # porque son números semánticamente DISTINTOS:
        #   - Resultados.Subtotal Cartera = cheques + facturas (BRUTO,
        #     activos comerciales).
        #   - /cartera total = facturas − cheques (NETO, lo que me deben).
        # En cambio, comparamos los SUMANDOS individuales (cheques y
        # facturas separados) que SÍ deben coincidir entre las 2 vistas.
        checks = [
            _diff_check(
                "Cheques en cartera — Resultados vs /cartera",
                "Resultados → Cheques",
                totc,
                "/cartera → Cheques en cartera",
                cartera_cheques,
            ),
            _diff_check(
                "Facturas vivas — Resultados vs /cartera (netas)",
                "Resultados → Facturas",
                totf,
                f"/cartera → Saldo facturas + sobrepagos ({cartera_sobrepagos:,.2f})",
                cartera_facturas,
            ),
            _diff_check(
                "Pasivos — Resultados vs /deudas",
                "Resultados → Pasivos",
                totp,
                "/deudas → Total deudas",
                deudas_total,
            ),
            _diff_check(
                "Posdatas — Resultados vs /posdat",
                "Resultados → ↳ Posdatas (total)",
                totp,
                "/posdat → tab Posdatados + tab YY",
                posdat_total_no_yy + posdat_total_yy,
            ),
            # Sanity adicionales — chequea queries internas vs live SQL.
            _diff_check(
                "Sanity TOTC: queries vs live SQL",
                "Resultados → Cheques (totc())",
                totc,
                "SELECT SUM live (Z+1+2+3+P+D)",
                cheques_live,
            ),
            _diff_check(
                "Sanity TOTF: queries vs live SQL",
                "Resultados → Facturas (totf())",
                totf,
                "SELECT SUM live (Z+A)",
                facturas_live,
            ),
        ]
    except Exception as e:  # noqa: BLE001
        error = str(e)
        import traceback

        traceback.print_exc()

    return render_template(
        "informes/check_totales.html",
        checks=checks,
        error=error,
    )


# TMT 2026-07-01 (dueña, review accesos Alex): el informe de Deudas a
# proveedores lo puede ver quien tenga `deudas.ver` (Alex, Compras, Gerente,
# Contabilidad) además de informes.ver — antes exigía informes.ver y dejaba
# afuera a Alex, que tiene deudas.ver pero no el módulo Informes completo.
@informes_bp.route("/deudas")
@requiere_login
def deudas():
    if not (tiene_permiso("informes.ver") or tiene_permiso("deudas.ver")):
        from flask import abort
        abort(404)
    filas, error = _safe(queries.deudas_por_proveedor, [])
    if request.args.get("export") == "csv":
        return csv_response(
            filas,
            columnas=[
                ("codigo_prov", "Código"),
                ("nombre", "Proveedor"),
                ("tipo", "Tipo"),
                ("n_posdats", "# posdatados"),
                ("saldo_total", "Saldo"),
                ("posdat_mas_vieja", "Posdat más vieja"),
                ("vence_mas_viejo", "Vence más vieja"),
            ],
            filename="deudas_proveedores.csv",
        )
    total = sum(float(r["saldo_total"] or 0) for r in filas)

    # TMT 2026-05-20 — Agrupar por categoría según proveedor.tipo (pedido
    # dueña: "subtotales de mat.prima, maquinaria, bancos, etc. con %").
    # Mapeo de tipos a categorías canónicas:
    #   H, Q  → Mat. Prima
    #   U     → Maquinaria
    #   B     → Bancos
    #   Y, '' → Otros / Servicios
    cats_orden = [
        (1, "Mat. Prima", {"H", "Q"}),
        (2, "Maquinaria", {"U"}),
        (3, "Bancos", {"B"}),
        (4, "Otros", {"Y", ""}),
    ]

    def _categoria_de(tipo: str) -> tuple[int, str]:
        t = (tipo or "").strip().upper()
        for orden, label, codes in cats_orden:
            if t in codes:
                return (orden, label)
        return (4, "Otros")

    # Anotar cada fila con categoria + categoria_orden + pct.
    filas_anotadas: list[dict] = []
    for r in filas:
        cat_orden, cat_label = _categoria_de(r.get("tipo") or "")
        saldo = float(r.get("saldo_total") or 0)
        filas_anotadas.append(
            {
                **dict(r),
                "categoria": cat_label,
                "categoria_orden": cat_orden,
                "pct": round(100.0 * saldo / total, 1) if total > 0 else 0.0,
            }
        )
    # Sort por categoría ASC + dentro por saldo DESC.
    filas_anotadas.sort(
        key=lambda r: (r["categoria_orden"], -float(r.get("saldo_total") or 0)),
    )

    # Subtotales por categoría (con %).
    subtotales: dict[int, dict] = {}
    for r in filas_anotadas:
        cat = r["categoria_orden"]
        s = subtotales.setdefault(
            cat,
            {
                "orden": cat,
                "label": r["categoria"],
                "n": 0,
                "total": 0.0,
            },
        )
        s["n"] += 1
        s["total"] += float(r.get("saldo_total") or 0)
    for s in subtotales.values():
        s["pct"] = round(100.0 * s["total"] / total, 1) if total > 0 else 0.0

    return render_template(
        "informes/deudas.html",
        filas=filas_anotadas,
        total=total,
        error=error,
        subtotales=subtotales,
    )


@informes_bp.route("/_diag/stock")
@requiere_login
@requiere_permiso("informes.ver")
def diag_stock():
    """TMT 2026-05-18 — diagnóstico del flujo de stock terminado.

    Muestra las queries crudas que alimentan /stock para entender por qué
    Terminado=0. Pensado para que la dueña abra la URL una vez y mande
    screenshot — más eficiente que pelear con SSM PowerShell quoting.
    """

    import db

    y = today_ec().year

    def _safe_q(sql, params=()):
        try:
            return db.fetch_all(sql, params) or []
        except Exception as e:
            return [{"error": str(e)}]

    tinto = _safe_q(
        """
        SELECT EXTRACT(MONTH FROM fecha)::int AS mes,
               COUNT(*) AS n,
               SUM(COALESCE(kg, 0))::int AS kg_col,
               SUM(COALESCE(kgn, 0))::int AS kgn_col,
               SUM(COALESCE(toper,0)+COALESCE(jersey,0)+COALESCE(pique,0)
                 + COALESCE(messi,0)+COALESCE(james,0)+COALESCE(franela,0)
                 + COALESCE(otros,0)+COALESCE(j3,0)+COALESCE(jlyc,0)
                 + COALESCE(flyc,0)+COALESCE(falso,0)+COALESCE(kiana,0))::int AS suma_indiv
          FROM scintela.tinto
         WHERE EXTRACT(YEAR FROM fecha) = %s
         GROUP BY 1 ORDER BY 1
    """,
        (y,),
    )

    iniciales = _safe_q(
        """
        SELECT yy, mesnum, hilado, tejido, terminado, vq,
               um, uk, uf, uq
          FROM scintela.iniciales
         WHERE yy = %s
         ORDER BY mesnum
    """,
        (y,),
    )

    facturas_mes = _safe_q(
        """
        SELECT EXTRACT(MONTH FROM fecha)::int AS mes,
               COUNT(*) AS n,
               SUM(COALESCE(kg, 0))::int AS kg
          FROM scintela.factura
         WHERE EXTRACT(YEAR FROM fecha) = %s
           AND COALESCE(stat, '') <> 'X'
         GROUP BY 1 ORDER BY 1
    """,
        (y,),
    )

    compras_tipo = _safe_q(
        """
        SELECT UPPER(TRIM(COALESCE(tipo, ''))) AS tipo,
               COUNT(*) AS n,
               SUM(COALESCE(kg, 0))::int AS kg,
               SUM(COALESCE(importe, 0))::int AS importe
          FROM scintela.compra
         WHERE EXTRACT(YEAR FROM fecha) = %s
           AND COALESCE(stat, '') != 'Y'
         GROUP BY 1 ORDER BY 1
    """,
        (y,),
    )

    return render_template(
        "informes/diag_stock.html",
        anio=y,
        tinto=tinto,
        iniciales=iniciales,
        facturas_mes=facturas_mes,
        compras_tipo=compras_tipo,
    )


@informes_bp.route("/snapshot-mes", methods=["POST"])
@requiere_login
@requiere_permiso("informes.editar")  # TMT 2026-06-03 audit: era .ver — escribe historia
def snapshot_mes():
    """Cierra snapshot mensual en scintela.historia para el mes indicado.

    POST con form (anio, mes). Idempotente.
    """

    try:
        anio = int(request.form.get("anio") or today_ec().year)
        mes = int(request.form.get("mes") or today_ec().month)
    except (TypeError, ValueError):
        flash("Parámetros inválidos.", "error")
        return redirect(url_for("informes.fuentes_y_usos"))
    mes = max(1, min(mes, 12))
    try:
        usuario = (g.user or {}).get("username", "web")
        r = queries.crear_snapshot_historia(anio, mes, usuario=usuario)
        if r.get("aplicado"):
            flash(f"Snapshot {mes:02d}/{anio} creado.", "ok")
        else:
            flash(r.get("razon", "Nada que hacer."), "info")
    except Exception as e:
        flash_exc("Snapshot falló", e)
    return redirect(url_for("informes.fuentes_y_usos", anio=anio, mes=mes))


@informes_bp.route("/snapshot-backfill", methods=["POST"])
@requiere_login
@requiere_permiso("informes.editar")  # TMT 2026-06-03 audit: era .ver — escribe historia
def snapshot_backfill():
    """Backfill: crea snapshots para los últimos N meses (default 3).

    POST con form (meses=N). Idempotente.
    """

    try:
        n = int(request.form.get("meses") or 3)
    except (TypeError, ValueError):
        n = 3
    # TMT 2026-05-19 v6 — Feature B permite hasta 12 meses (antes 12 cap).
    n = max(1, min(n, 12))
    hoy = today_ec()
    aplicados, saltados = [], []
    usuario = (g.user or {}).get("username", "web")
    for i in range(1, n + 1):
        # Mes pasado i: retrocedemos i meses desde el primero del mes actual
        m = hoy.month - i
        a = hoy.year
        while m < 1:
            m += 12
            a -= 1
        try:
            r = queries.crear_snapshot_historia(a, m, usuario=usuario)
            (aplicados if r.get("aplicado") else saltados).append(f"{m:02d}/{a}")
        except Exception as e:
            saltados.append(f"{m:02d}/{a} (error: {e})")
    if aplicados:
        flash(f"Backfilled: {', '.join(aplicados)}.", "ok")
    if saltados:
        flash(f"Salteados (ya existían o error): {', '.join(saltados)}.", "info")
    return redirect(url_for("informes.fuentes_y_usos"))


@informes_bp.route("/fuentes-y-usos")
@requiere_login
@requiere_permiso("informes.ver")
def fuentes_y_usos():
    """Cuadro de Fuentes y Usos en un rango DESDE-HASTA (mensual).

    Pedido dueña 2026-05-19 (docx "Para Claude 2", item 14): seleccionar
    DESDE-HASTA y mostrar 2 columnas con totales iguales (réplica de
    INFORMES.PRG::PROCEDURE FUENTES L1654-1727). Granularidad: mensual,
    porque la data viene de scintela.historia (un snapshot por mes).
    """

    hoy = today_ec()

    def _p(k, default):
        try:
            return int(request.args.get(k) or default)
        except (TypeError, ValueError):
            return default

    # Default: ventana de 1 mes terminando en mes actual (compatible con
    # comportamiento anterior cuando solo había un picker).
    hasta_anio = _p("hasta_anio", _p("anio", hoy.year))
    hasta_mes = _p("hasta_mes", _p("mes", hoy.month))
    desde_anio = _p("desde_anio", hasta_anio if hasta_mes > 1 else hasta_anio - 1)
    desde_mes = _p("desde_mes", hasta_mes - 1 if hasta_mes > 1 else 12)
    hasta_mes = max(1, min(hasta_mes, 12))
    desde_mes = max(1, min(desde_mes, 12))

    try:
        data = queries.fuentes_y_usos(
            desde_anio=desde_anio,
            desde_mes=desde_mes,
            hasta_anio=hasta_anio,
            hasta_mes=hasta_mes,
        )
    except Exception as e:
        data = {
            "anio_ini": desde_anio,
            "mes_ini": desde_mes,
            "anio": hasta_anio,
            "mes": hasta_mes,
            "fuentes": [],
            "usos": [],
            "total_fuentes": 0,
            "total_usos": 0,
            "delta_liquido": 0,
            "delta_banco": 0,
            "h_ini": {},
            "h_fin": {},
            "error": str(e),
        }
    return render_template(
        "informes/fuentes_usos.html",
        data=data,
        # Para back-compat con el template (siguen existiendo `anio`/`mes`
        # como los del HASTA, además de los explícitos `desde_*`/`hasta_*`).
        anio=hasta_anio,
        mes=hasta_mes,
        desde_anio=desde_anio,
        desde_mes=desde_mes,
        hasta_anio=hasta_anio,
        hasta_mes=hasta_mes,
    )


@informes_bp.route("/flujo")
@requiere_login
@requiere_permiso("informes.ver")
def flujo():
    dias = request.args.get("dias", default=30, type=int)
    filas, error = _safe(lambda: queries.flujo_ultimos_dias(dias), [])
    if request.args.get("export") == "csv":
        return csv_response(
            filas,
            columnas=[
                ("fecha", "Fecha"),
                ("cheques", "Cheques"),
                ("facturas", "Facturas"),
                ("pichincha", "Pichincha"),
                ("inter", "Internacional"),
                ("posdat1", "Pos.dat 1"),
                ("posdat2", "Pos.dat 2"),
                ("mprima", "M. prima"),
                ("gastos", "Gastos"),
                ("saldo", "Saldo"),
                ("pagos", "Pagos"),
                ("dolares", "Dólares"),
                ("usaldo", "USD saldo"),
            ],
            filename=f"flujo_{dias}d.csv",
        )
    return render_template("informes/flujo.html", filas=filas, dias=dias, error=error)


@informes_bp.route("/flujo/grafico")
@requiere_login
@requiere_permiso("informes.ver")
def flujo_grafico():
    """Gráfico de flujo de caja — la vista del gerente, con proyección.

    Equivalente moderno del GRAFICO del viejo dBase: muestra historia
    reciente + proyección a 365 días (postdatados, provisiones, pagos
    programados ya cargados en scintela.flujo).
    """
    # Default 70d para matchear el rango del chart dBase (May 11 → Jul 20).
    # El MIN del flujo cae a los 68-70 días.
    ventana = request.args.get("ventana", default=90, type=int)
    ventana = max(7, min(ventana, 365))  # clamp

    # Fuente del flujo:
    #   1. `flujo_calculado()` — proyección en vivo desde cheques+posdat+saldos.
    #      Es la fuente primaria desde 2026-04-29 (batch 19) porque la tabla
    #      legacy scintela.flujo nunca se carga en producción.
    #   2. `flujo_proyeccion()` — lee scintela.flujo (la tabla legacy alimentada
    #      por el dBase). Sólo si tenemos filas ahí Y `?fuente=tabla` explícito.
    #
    # Esto resuelve el bug histórico "el gráfico nunca muestra nada": ahora
    # arranca con un saldo bancario real y proyecta los cheques pendientes
    # de cobro y los posdat pendientes de pago.
    fuente = (request.args.get("fuente") or "calculado").lower()
    # Modo "peor caso": ignorar cheques en cartera (asumir que ninguno se
    # cobra). Útil cuando se sospecha que la cartera Z está stale.
    ignorar_cheques = request.args.get("ignorar_cheques") in ("1", "true", "yes", "on")
    if fuente == "tabla":
        filas, error = _safe(
            lambda: queries.flujo_proyeccion(dias_atras=14, dias_adelante=365),
            [],
        )
    else:
        filas, error = _safe(
            lambda: queries.flujo_calculado(
                dias_atras=14,
                dias_adelante=365,
                ignorar_cheques=ignorar_cheques,
            ),
            [],
        )

    # Pass dates as ISO strings — the JS parses them deterministically
    # instead of relying on the browser's Date(string) forgiveness.
    datos = [
        {
            "fecha": r["fecha"].isoformat() if hasattr(r["fecha"], "isoformat") else r["fecha"],
            "saldo": float(r["saldo"] or 0),
            "cheques": float(r["cheques"] or 0),
            "facturas": float(r["facturas"] or 0),
            "posdat1": float(r["posdat1"] or 0),
            "posdat2": float(r["posdat2"] or 0),
            "pichincha": float(r["pichincha"] or 0),
            "inter": float(r["inter"] or 0),
            "mprima": float(r["mprima"] or 0),
            "gastos": float(r["gastos"] or 0),
            "pagos": float(r["pagos"] or 0),
            "dolares": float(r["dolares"] or 0),
        }
        for r in filas
    ]

    # Lista de posdat egresos para mostrar al lado del gráfico — ayuda al
    # gerente a saber QUÉ se está restando, no sólo el total agregado.
    posdat_egresos, _ = _safe(
        lambda: queries.posdat_egresos_proximos(dias_adelante=365),
        [],
    )
    egresos_lista = [
        {
            "id_posdat": int(r["id_posdat"]) if r.get("id_posdat") else None,
            "fecha_efectiva": r["fecha_efectiva"].isoformat()
            if hasattr(r["fecha_efectiva"], "isoformat")
            else r["fecha_efectiva"],
            "fechad": r["fechad"].isoformat()
            if r.get("fechad") and hasattr(r["fechad"], "isoformat")
            else None,
            "prov": r.get("prov") or "",
            "concepto": r.get("concepto") or "",
            "importe": float(r.get("importe") or 0),
            "banc": int(r.get("banc") or 0),
            "vencido": bool(r.get("fechad") and r["fechad"] < today_ec()),
        }
        for r in posdat_egresos
    ]

    # Posdatados (banc=0) ordenados por fecha de vencimiento para la caja
    # inferior del gráfico "Posdatados contabilizados" (dueña 2026-07-09).
    posdat_filas = sorted(
        (e for e in egresos_lista if e.get("banc") == 0),
        key=lambda e: (e.get("fechad") or e.get("fecha_efectiva") or ""),
    )

    # Plazos PLAZ.COBR / PLAZ.DEUDA — calculados server-side con la fórmula
    # de dBase (plazo otorgado ponderado por importe). El JS antes los
    # calculaba sobre la ventana del gráfico con `fecha-hoy`, lo cual no
    # representa el plazo real otorgado y daba números muy bajos (23/25 vs
    # 32.9/96.7 de dBase).
    plazos, _ = _safe(lambda: queries.plazos_dbase(), {"cobro": 0, "deuda": 0})

    return render_template(
        "informes/flujo_grafico.html",
        datos=datos,
        egresos_lista=egresos_lista,
        posdat_filas=posdat_filas,
        hoy=today_ec().isoformat(),
        ventana_dias=ventana,
        ignorar_cheques=ignorar_cheques,
        plazos=plazos,
        error=error,
    )


@informes_bp.route("/flujo/grafico/export.xlsx")
@requiere_login
@requiere_permiso("informes.ver")
def flujo_grafico_export():
    """Excel del flujo día por día — qué SUMA (ingresos) y qué RESTA (egresos)
    cada día, con el saldo acumulado. Pedido dueña 2026-07-08 ("un excel del
    flujo que muestre cada día que suma y que resta"). Misma fuente que el
    gráfico (flujo_calculado): cheques cobrados = ingreso (+); posdat P1/P2,
    materia prima y gastos = egresos (−). Fechas en DD/MM/AAAA."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ignorar_cheques = request.args.get("ignorar_cheques") in ("1", "true", "yes", "on")
    filas, _ = _safe(
        lambda: queries.flujo_calculado(
            dias_atras=0, dias_adelante=365, ignorar_cheques=ignorar_cheques
        ),
        [],
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Flujo"
    headers = [
        "Fecha", "Ingresos (cheques)", "Posdat P1", "Posdat P2",
        "Materia prima", "Gastos", "Neto del día", "Saldo acumulado",
    ]
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="0F172A")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    money_fmt = "#,##0.00"
    rownum = 2
    for r in filas:
        ing = float(r.get("cheques") or 0)   # +
        p1 = float(r.get("posdat1") or 0)    # ya negativo (egreso)
        p2 = float(r.get("posdat2") or 0)
        mp = float(r.get("mprima") or 0)
        g = float(r.get("gastos") or 0)
        # Sólo días con movimiento (el resto es la línea plana del saldo).
        if ing == 0 and p1 == 0 and p2 == 0 and mp == 0 and g == 0:
            continue
        neto = round(ing + p1 + p2 + mp + g, 2)
        ws.cell(row=rownum, column=1, value=r.get("fecha")).number_format = "DD/MM/YYYY"
        for i, v in enumerate(
            [ing, p1, p2, mp, g, neto, round(float(r.get("saldo") or 0), 2)], start=2
        ):
            ws.cell(row=rownum, column=i, value=v).number_format = money_fmt
        rownum += 1

    for i, w in enumerate([13, 18, 13, 13, 14, 13, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="flujo_{today_ec().isoformat()}.xlsx"'
        },
    )


# ---------------------------------------------------------------------------
# Flujo — carga manual / CSV  (v1: la forma más rápida de poblar scintela.flujo
# sin necesidad de importar desde el dBase viejo ni correr scripts a mano).
# ---------------------------------------------------------------------------

_FLUJO_HEADERS = [
    "fecha",
    "saldo",
    "cheques",
    "facturas",
    "posdat1",
    "posdat2",
    "pichincha",
    "inter",
    "mprima",
    "gastos",
    "pagos",
    "dolares",
    "usaldo",
]


def _parse_fecha(value: str):
    """Accept 2026-04-16, 16/04/2026, 16-04-2026."""
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_monto(value):
    """Accept 1234.56, 1.234,56 (es-EC), empty → None."""
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    # es-EC: punto miles, coma decimal. Si hay coma, asumí ese formato.
    if "," in s and s.count(",") == 1:
        s = s.replace(".", "").replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


@informes_bp.route("/flujo/cargar", methods=["GET", "POST"])
@requiere_login
@requiere_permiso("informes.editar")  # TMT 2026-06-03 audit: era .ver — INSERT en scintela.flujo
def flujo_cargar():
    """Carga manual o por CSV de la tabla scintela.flujo.

    - GET ?plantilla=1 → descarga un CSV template vacío.
    - GET              → muestra el formulario (manual + upload).
    - POST             → procesa CSV o una sola fila manual.
    """
    # --- CSV template download --------------------------------------------
    if request.method == "GET" and request.args.get("plantilla"):
        buf = io.StringIO()
        buf.write("\ufeff")
        w = csv.writer(buf, delimiter=";")
        w.writerow(_FLUJO_HEADERS)
        # Una fila de ejemplo para que el usuario vea el formato.
        w.writerow(
            [
                today_ec().isoformat(),
                "0",
                "0",
                "0",
                "0",
                "0",
                "0",
                "0",
                "0",
                "0",
                "0",
                "0",
                "0",
            ]
        )
        return Response(
            buf.getvalue().encode("utf-8"),
            mimetype="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": 'attachment; filename="flujo_plantilla.csv"',
            },
        )

    resultado = None
    errores: list[str] = []

    if request.method == "POST":
        usuario = (g.user or {}).get("username", "web")
        rows: list[dict] = []

        # Rama 1: upload de CSV.
        f = request.files.get("archivo")
        if f and f.filename:
            try:
                raw = f.stream.read().decode("utf-8-sig", errors="replace")
            except Exception as e:  # pragma: no cover — defensivo
                errores.append(f"No pude leer el archivo: {e}")
                raw = ""
            if raw:
                # Detectá ; o , como separador mirando la primera línea.
                first = raw.split("\n", 1)[0]
                delim = ";" if first.count(";") >= first.count(",") else ","
                reader = csv.DictReader(io.StringIO(raw), delimiter=delim)
                for i, r in enumerate(reader, start=2):  # línea 1 = header
                    # Normalizá keys (lowercase, strip, quitar BOM residual).
                    rn = {(k or "").strip().lower().lstrip("\ufeff"): v for k, v in r.items()}
                    fecha = _parse_fecha(rn.get("fecha"))
                    if fecha is None:
                        errores.append(f"Línea {i}: fecha inválida ({rn.get('fecha')!r})")
                        continue
                    row: dict = {"fecha": fecha}
                    for col in queries.FLUJO_COLS:
                        if col in rn and rn[col] not in (None, ""):
                            monto = _parse_monto(rn[col])
                            if monto is None:
                                errores.append(f"Línea {i} col {col}: monto inválido ({rn[col]!r})")
                            else:
                                row[col] = monto
                    rows.append(row)

        # Rama 2: una sola fila manual.
        else:
            fecha = _parse_fecha(request.form.get("fecha"))
            if fecha is None:
                errores.append("Fecha requerida (formato AAAA-MM-DD o DD/MM/AAAA).")
            else:
                row = {"fecha": fecha}
                for col in queries.FLUJO_COLS:
                    val = request.form.get(col)
                    if val not in (None, ""):
                        monto = _parse_monto(val)
                        if monto is None:
                            errores.append(f"Monto inválido en {col}: {val!r}")
                        else:
                            row[col] = monto
                if fecha is not None:
                    rows.append(row)

        if rows and not errores:
            try:
                resultado = queries.upsert_flujo_rows(rows, usuario)
                flash(
                    f"Flujo: {resultado['inserted']} insertadas, {resultado['updated']} actualizadas.",
                    "ok",
                )
                return redirect(url_for("informes.flujo_grafico"))
            except Exception as e:
                errores.append(f"Error guardando: {e}")

    return render_template(
        "informes/flujo_cargar.html",
        headers=_FLUJO_HEADERS,
        cols=queries.FLUJO_COLS,
        errores=errores,
        resultado=resultado,
        hoy=today_ec().isoformat(),
    )


@informes_bp.route("/ventas/multianual")
@requiere_login
@requiere_permiso("informes.ver")
def ventas_multianual():
    """Matriz ventas mes × año — replica MODIFICA.PRG PROCEDURE VENTAS L144-217.

    Default 4 años para alinear con la captura legacy (2020-21-22-23).
    """
    anios = request.args.get("anios", default=4, type=int)
    data, error = _safe(lambda: queries.ventas_multianual(anios), {})
    return render_template(
        "informes/ventas_multianual.html",
        data=data,
        anios=anios,
        error=error,
    )


@informes_bp.route("/ventas")
@requiere_login
@requiere_permiso("informes.ver")
def ventas():
    # TMT 2026-05-19 v8 — dueña: al clickear "Ventas" del balance quiere ver
    # la pantalla TINT.BAT del dBase (ranking clientes del mes). Por default
    # ahora redirigimos al ranking del mes; el listado multi-mes vive en
    # ventas_multianual (link sigue disponible desde ahí).

    hoy = today_ec()
    try:
        anio = int(request.args.get("anio") or hoy.year)
    except (TypeError, ValueError):
        anio = hoy.year
    try:
        mes = int(request.args.get("mes") or hoy.month)
    except (TypeError, ValueError):
        mes = hoy.month
    mes = max(1, min(mes, 12))
    data, error = _safe(
        lambda: queries.ventas_clientes_del_mes(anio=anio, mes=mes),
        {},
    )
    return render_template(
        "informes/ventas_mes.html",
        data=data,
        anio=anio,
        mes=mes,
        error=error,
    )


# TMT 2026-05-19 v8 — pantalla multi-mes eliminada (pedido dueña).
# La query `ventas_mensuales` y el template `informes/ventas.html`
# quedan en el repo por si se necesitan más adelante, pero ya no hay
# ruta que los exponga.


@informes_bp.route("/ventas-anio")
@requiere_login
@requiere_permiso("ventas.ver")
def ventas_anio():
    """Ventas del año en curso — mes a mes con acumulado.

    TMT 2026-05-20 — pedido dueña: pantalla simple desde
    /informes/balance al click 'Ventas del año'. Columnas:
    mes · kg · precio (U$/kg) · importe · acum.
    """
    filas, error = _safe(queries.ventas_mes_a_mes_anio_actual, [])
    total_kg = sum(float(r.get("kg") or 0) for r in filas)
    total_importe = sum(float(r.get("importe") or 0) for r in filas)
    precio_prom = (total_importe / total_kg) if total_kg > 0 else 0.0
    if request.args.get("export") == "csv":
        return csv_response(
            filas,
            columnas=[
                ("mes_nombre", "Mes"),
                ("kg", "Kg"),
                ("precio", "Precio U$/kg"),
                ("importe", "Importe"),
                ("acum", "Acumulado"),
            ],
            filename="ventas_anio.csv",
        )

    return render_template(
        "informes/ventas_anio.html",
        filas=filas,
        total_kg=total_kg,
        total_importe=total_importe,
        precio_prom=precio_prom,
        anio=today_ec().year,
        error=error,
    )


def _chequeo_coherencia(data, mov_asinfo, prod_tej_asinfo, tol_pct=1.0):
    """Chequeos de coherencia entre bandas del flujo (dueña 2026-07-15): detectar
    descuadres automáticamente en vez de a ojo. Cada chequeo compara dos números
    que TIENEN que coincidir; marca 'warn' si el desvío supera tol_pct. Los de
    tipo 'ajuste' (químicos: físico vs libro) son informativos, nunca 'warn'.
    Fail-soft: si falta un dato queda 'nd' (no disponible)."""
    data = data or {}
    mov = mov_asinfo or {}

    def _g(d, *ks):
        for k in ks:
            if not isinstance(d, dict):
                return None
            d = d.get(k)
        return d

    checks = []

    def add(clave, etiqueta, a, a_lbl, b, b_lbl, unidad, tipo="cuadre"):
        try:
            a = None if a is None else float(a)
            b = None if b is None else float(b)
        except (TypeError, ValueError):
            a = b = None
        if a is None or b is None:
            estado, delta, pct = "nd", None, None
        else:
            delta = a - b
            pct = (delta / b * 100.0) if b else None
            if tipo == "ajuste":
                estado = "info"
            elif pct is None:
                estado = "warn" if delta else "ok"
            else:
                estado = "ok" if abs(pct) <= tol_pct else "warn"
        checks.append({
            "clave": clave, "etiqueta": etiqueta, "unidad": unidad, "tipo": tipo,
            "a": a, "a_lbl": a_lbl, "b": b, "b_lbl": b_lbl,
            "delta": delta, "pct": pct, "estado": estado,
        })

    # Hilo (dueña 2026-07-17: "que los 3 sean iguales"): Compras hilado =
    # Ingresos de la banda MOVIMIENTOS = Importaciones recibidas. La banda ya
    # muestra las importaciones recibidas como Ingresos (los reingresos de
    # lote van neteados en el egreso), así que acá se chequean los 3 números
    # de a pares: si cualquiera difiere, algún par marca ⚠.
    add("hilo", "Hilo comprado = ingresado",
        _g(data, "compras_hilado_total", "kg"), "Compras hilado",
        _g(mov, "hilado", "ingresos_kg"), "Ingresos hilado", "kg")
    add("hilo_import", "Hilo ingresado = importaciones recibidas",
        _g(mov, "hilado", "ingresos_kg"), "Ingresos hilado",
        _g(mov, "hilado", "ref_import_kg"), "Importaciones recibidas", "kg")
    # (La línea "reingresos de lote / ingreso bodega 75k" se SACÓ — dueña
    # 2026-07-17: "el check sigue diciendo 75k", confundía. Los reingresos ya
    # van neteados en el egreso; la trazabilidad queda en
    # mov.hilado.ref_bodega_ing_kg / ref_reingresos_kg si hace falta auditar.)
    add("tejido", "Tejido producido = crudo ingresado",
        _g(data, "produc_tejido_total", "kg"), "Producción tejido",
        _g(mov, "tejido", "ingresos_kg"), "Ingresos crudo", "kg")
    # Químicos: físico vs libro. El consumo = proyectado de tintura y el ajuste
    # este mes es de arranque (cutover dBase→programa); tipo="ajuste" → informa,
    # no marca descuadre. La banda y la columna QUÍM.$ muestran lo mismo.
    qm = _g(mov, "quimicos_banda") or _g(mov, "quimicos_modelo")
    if qm:
        add("quimicos", "Químicos: físico vs libro",
            _g(qm, "final_form"), "Físico",
            _g(qm, "final_prog"), "Libro", "US$", tipo="ajuste")

    return checks


@informes_bp.route("/flujo-produccion")
@requiere_login
@requiere_permiso("informes.ver")
def flujo_produccion():
    """Pantalla TINT.BAT replica — flujo de producción + costos unitarios.

    TMT 2026-05-19 v8 — pedido dueña: linkeable desde "Stock" de
    /informes/balance, muestra MOVIMIENTOS MES (hilado/crudo/term/col),
    COMPRAS HILADO, PRODUC.TEJIDO, TINTORERIA y CS.COLORANTES/PRODUCCION.
    """

    hoy = today_ec()
    try:
        anio = int(request.args.get("anio") or hoy.year)
    except (TypeError, ValueError):
        anio = hoy.year
    try:
        mes = int(request.args.get("mes") or hoy.month)
    except (TypeError, ValueError):
        mes = hoy.month
    mes = max(1, min(mes, 12))

    # Cache-hit: contexto ya armado para este (anio, mes) dentro del TTL. Los
    # datos no dependen del usuario, así que se comparte entre usuarios.
    _ck = (anio, mes)
    _now_fp = _time_fp.time()
    _t0_fp = _time_fp.perf_counter()
    _cached_fp = _FLUJO_PROD_CACHE.get(_ck)
    if _cached_fp and (_now_fp - _cached_fp[0]) < _FLUJO_PROD_TTL_SECS:
        _ctx = _cached_fp[1]
        # el context-processor de comparativa_tintoreria lee g._tint_mensual
        g._tint_mensual = _ctx.get("_tint_mensual")
        _LOG_FP.info("flujo_produccion %s/%s cache HIT", mes, anio)
        return render_template("informes/flujo_produccion.html", **_ctx["render"])

    # Federico 2026-07-17 — instrumentación temporal para ver qué consulta domina
    # la primera carga (se expone en un comentario HTML al final de la página).
    _timings = {}

    # TMT 2026-07-08 (dueña): tabla live de Asinfo por etapa (kg) al tope, y
    # un segundo cuadro "MOVIMIENTOS DEL MES (inicial Asinfo)" — clon del de
    # abajo pero con el Stock inicial tomado del snapshot de Asinfo al arranque
    # del mes. Ambas fail-soft: si Asinfo está caído la vista igual renderiza.
    from modules.asinfo import service as asinfo_service

    fecha_corte = date(anio, mes, 1)  # arranque del mes = corte del as-of

    # Federico 2026-07-17 — PERF: estas 3 consultas son INDEPENDIENTES entre sí y
    # sus backends son thread-safe (dBase = ThreadedConnectionPool; Asinfo = HTTP
    # con requests). En serie sumaban ~9s; en paralelo bajan a ~la más lenta (~4s).
    # NO se paraleliza formulas_db (SimpleConnectionPool, NO thread-safe) ni la
    # lógica de químicos. Cada tarea va envuelta en _safe → fail-soft por separado.
    from concurrent.futures import ThreadPoolExecutor
    _s = _time_fp.perf_counter()
    with ThreadPoolExecutor(max_workers=3) as _ex:
        _f_data = _ex.submit(
            _safe, lambda: queries.movimientos_mes_dbase(anio=anio, mes=mes), {})
        _f_inv = _ex.submit(_safe, asinfo_service.inventario_por_etapa, {})
        _f_inv_asof = _ex.submit(
            _safe, lambda: asinfo_service.inventario_por_etapa_a_fecha(fecha_corte), {})
        data, error = _f_data.result()
        inv_asinfo, _e_inv = _f_inv.result()
        inv_asinfo_inic, _e_inic = _f_inv_asof.result()
    _timings["fetch_paralelo_3"] = round(_time_fp.perf_counter() - _s, 3)

    if not isinstance(inv_asinfo, dict):
        inv_asinfo = {}
    if not isinstance(inv_asinfo_inic, dict):
        inv_asinfo_inic = {}

    # Proyectado de tintura (mismo número que la tabla de abajo) = consumo de
    # químico de la columna QUÍM.$ y de la banda. Se calcula UNA vez acá y se
    # deja en g para que el context-processor de comparativa_tintoreria lo
    # reuse (no recalcular). Dueña 2026-07-16: un solo número en la pantalla.
    _tint_mensual = None
    _proy_quimico = None
    try:
        from modules.comparativa_tintoreria.views import _build_tintoreria_mensual
        _s = _time_fp.perf_counter()
        _tint_mensual, _e_tint = _safe(
            lambda: _build_tintoreria_mensual(anio, mes), None)
        _timings["tintoreria_mensual"] = round(_time_fp.perf_counter() - _s, 3)
        g._tint_mensual = _tint_mensual
        _filas_t = (_tint_mensual or {}).get("filas") or []
        if _filas_t:
            # Dueña 2026-07-17: usar el TOTAL ACTUAL de la tabla (lo costeado,
            # tela cerrada) — NO el proyectado. Así el consumo de la banda es
            # exactamente el mismo 90k de COSTOS DE TINTORERÍA, y no respira
            # con las órdenes en proceso (que no descuentan de ningún lado
            # hasta cerrarse).
            _proy_quimico = (_filas_t[0] or {}).get("t_imp")
    except Exception:  # noqa: BLE001 -- fail-soft: si falla, cae al respaldo
        _tint_mensual = None
        _proy_quimico = None

    _s = _time_fp.perf_counter()
    mov_asinfo = _build_mov_asinfo(data, inv_asinfo_inic, inv_asinfo,
                                   anio=anio, mes=mes, proy_quimico=_proy_quimico)
    _timings["mov_asinfo_quimicos"] = round(_time_fp.perf_counter() - _s, 3)

    # TMT 2026-07-14 (dueña): "tengo que tener EXACTAMENTE la misma tabla, copiá
    # la segunda igual a la primera". La tabla PRODUCCIÓN TEJIDO de Asinfo tiene
    # el MISMO formato que la del dBase (Prov | Kg | $/Kg | $), no un comparativo.
    # Kg = ingreso real a bodega 52 (= "Ingresos crudo", cierra con el stock); $ =
    # costo (tercerizados = compra cargada; INTELA autoprod = kg × (hilo + 0,5)).
    # Los números salen de tejeduria_asinfo.resumen_mes (misma fuente que la tab).
    prod_tej_asinfo = None
    _s = _time_fp.perf_counter()
    try:
        from modules.tejeduria_asinfo import service as _tej_svc
        _res = _tej_svc.resumen_mes(anio, mes)
        if _res and _res.get("disponible"):
            # $ INTELA (autoprod) = GASTO DE TEJEDURÍA (V1+V2+V3 + DTJ) prorrateado,
            # NO hilo+0,5: el hilo ya está en Compras hilado, valuarlo a hilo+0,5
            # lo contaría dos veces (por eso saltaba de ~55k a ~410k). Dueña
            # 2026-07-15. Tercerizados = lo facturado (su maquila ya es el costo).
            try:
                _gxg = queries.gastos_xgast_v1_a_v9_mes()
                _amort = queries.amortizaciones_mensuales()
                _gs_tej = (float((_gxg or {}).get("gtej_sin_dtj") or 0)
                           + float((_amort or {}).get("dtj") or 0))
            except Exception:  # noqa: BLE001 -- fail-soft
                _gs_tej = 0.0
            _kg_intela = sum(float(_t.get("kg") or 0)
                             for _t in _res.get("tejedores", []) if _t.get("es_intela"))
            _rows = []
            for _t in _res.get("tejedores", []):
                _es_intela = bool(_t.get("es_intela"))
                _prov = "INTELA" if _es_intela else (_t.get("cod") or _t.get("label"))
                _kg = float(_t.get("kg") or 0)
                if _es_intela:
                    _imp = _gs_tej * (_kg / _kg_intela) if _kg_intela else 0.0
                else:
                    _imp = float(_t.get("costo") or 0)
                _rows.append({
                    "prov": _prov,
                    "kg": _kg,
                    "importe": _imp,
                    "ukg": (_imp / _kg if _kg else 0.0),
                })
            _tot_kg = float(_res.get("total_kg") or 0)
            _tot_imp = sum(r["importe"] for r in _rows)
            prod_tej_asinfo = {
                "filas": _rows,
                "total": {
                    "kg": _tot_kg,
                    "importe": _tot_imp,
                    "ukg": (_tot_imp / _tot_kg if _tot_kg else 0),
                },
            }
    except Exception:  # noqa: BLE001 -- best-effort, la vista no rompe
        prod_tej_asinfo = None
    _timings["prod_tejido_asinfo"] = round(_time_fp.perf_counter() - _s, 3)

    # dueña 2026-07-15: "Asinfo manda". "Producción tejido" = FÍSICO de bodega 52
    # (resumen_mes), NO las compras tipo K — que estaban viejas del match del 13/07
    # y ni aparecen en /compras. Unifica la variable con "Ingresos crudo" y con la
    # pantalla Tejeduría Asinfo. Fail-soft: si Asinfo no está, queda el tipo K.
    if isinstance(data, dict) and prod_tej_asinfo and prod_tej_asinfo.get("filas"):
        data["produc_tejido"] = prod_tej_asinfo["filas"]
        data["produc_tejido_total"] = prod_tej_asinfo["total"]

    _s = _time_fp.perf_counter()
    coherencia, _e_coh = _safe(
        lambda: _chequeo_coherencia(data, mov_asinfo, prod_tej_asinfo),
        [],
    )
    _timings["coherencia"] = round(_time_fp.perf_counter() - _s, 3)
    _timings["TOTAL"] = round(_time_fp.perf_counter() - _t0_fp, 3)

    _render_kw = dict(
        data=data,
        anio=anio,
        mes=mes,
        error=error,
        inv_asinfo=inv_asinfo,
        mov_asinfo=mov_asinfo,
        prod_tej_asinfo=prod_tej_asinfo,
        coherencia=coherencia,
    )
    # Guardar en cache SOLO cargas buenas: sin error y con data del mes. Así un
    # Asinfo caído o un mes vacío no queda "pegado" el TTL entero.
    if not error and isinstance(data, dict) and data.get("header"):
        _FLUJO_PROD_CACHE[_ck] = (
            _now_fp, {"_tint_mensual": _tint_mensual, "render": dict(_render_kw)})
    _LOG_FP.info("flujo_produccion %s/%s cache MISS %ss %s",
                 mes, anio, _timings.get("TOTAL"), _timings)
    # _perf va SOLO en la respuesta viva (no en el cache) → comentario HTML.
    return render_template("informes/flujo_produccion.html", _perf=_timings, **_render_kw)


@informes_bp.route("/gastos")
@requiere_login
@requiere_permiso("gastos.ver")
def gastos():
    filas, error = _safe(queries.gastos_mes_corriente, [])
    if request.args.get("export") == "csv":
        return csv_response(
            filas,
            columnas=[
                ("fecha", "Fecha"),
                ("documento", "Doc"),
                ("concepto", "Concepto"),
                ("proveedor", "Proveedor"),
                ("banco", "Banco"),
                ("importe", "Importe"),
            ],
            filename="gastos_mes.csv",
        )
    total = sum(float(r["importe"] or 0) for r in filas)
    # Matriz V1-V9 (xgast por NUM) + amortizaciones por rubro.
    # Layout 3x3: filas = personal/servicios/otros, cols = tej/tinto/admin.
    # Coincide con la convención del PRG INFORMES.PRG líneas 211-217:
    #   GTEJ = V1+V2+V3 + DTJ
    #   GTIN = V4+V5+V6 + DCC
    #   GGF  = V7+V8+V9 + DEPRCAR
    v, _e = _safe(queries.gastos_xgast_v1_a_v9_mes, {})
    a, _e = _safe(queries.amortizaciones_mensuales, {})

    def gv(k):
        return float((v or {}).get(k) or 0)

    def ga(k):
        return float((a or {}).get(k) or 0)

    matriz = {
        "personal": {
            "tej": gv("v1"),
            "tin": gv("v4"),
            "adm": gv("v7"),
        },
        "servicios": {
            "tej": gv("v2"),
            "tin": gv("v5"),
            "adm": gv("v8"),
        },
        "otros": {
            "tej": gv("v3"),
            "tin": gv("v6"),
            "adm": gv("v9"),
        },
    }
    # Totales por columna (V1+V2+V3 etc.) y los GTEJ/GTIN/GGF con amort.
    col_v = {
        "tej": gv("v1") + gv("v2") + gv("v3"),
        "tin": gv("v4") + gv("v5") + gv("v6"),
        "adm": gv("v7") + gv("v8") + gv("v9"),
    }
    col_amort = {"tej": ga("dtj"), "tin": ga("dcc"), "adm": ga("deprcar")}
    col_total = {k: col_v[k] + col_amort[k] for k in col_v}
    # Totales por fila (personal/servicios/otros — sin amort, son sólo V1-V9).
    fil_total = {
        "personal": matriz["personal"]["tej"] + matriz["personal"]["tin"] + matriz["personal"]["adm"],
        "servicios": matriz["servicios"]["tej"] + matriz["servicios"]["tin"] + matriz["servicios"]["adm"],
        "otros": matriz["otros"]["tej"] + matriz["otros"]["tin"] + matriz["otros"]["adm"],
    }
    suma_v_total = sum(col_v.values())
    suma_amort_total = sum(col_amort.values())
    suma_grand = sum(col_total.values())
    # TMT 2026-05-19 v5 — pedido dueña: banner "Sin clasificar" con link
    # al wizard. xgast.num NULL → no aparece en V1..V9 → invisible al ojo.
    # Mostrar al pie cuánta plata hay en ese limbo.
    sin_num_resumen = {"n": 0, "total": 0.0, "n_conceptos_unicos": 0}
    try:
        from modules.gastos import queries as _gq

        sin_num_resumen = _gq.xgast_sin_num_resumen()
    except Exception:
        pass

    return render_template(
        "informes/gastos.html",
        filas=filas,
        total=total,
        error=error,
        matriz=matriz,
        col_v=col_v,
        col_amort=col_amort,
        col_total=col_total,
        fil_total=fil_total,
        suma_v_total=suma_v_total,
        suma_amort_total=suma_amort_total,
        suma_grand=suma_grand,
        sin_num_resumen=sin_num_resumen,
    )


@informes_bp.route("/gastos/detalle/<int:num>")
@requiere_login
@requiere_permiso("gastos.ver")
def gastos_detalle(num):
    """Drill-down de una categoría V1..V12 — DETALGAST del PRG.

    Lista las filas de `scintela.xgast` para esa categoría (mes en curso)
    agrupadas por concepto (EEQ/CMB/EMAAP/etc).
    """
    # TMT 2026-05-15: decisión #3 — antes era `abort(404)` para num fuera de
    # rango. La dueña pidió un 400 explícito con el rango válido y el valor
    # recibido, para que se entienda qué pasó al tipear una URL inválida.
    if num < 1 or num > 12:
        abort(400, description=f"categoría debe estar entre 1 y 12, recibido {num}")
    data, error = _safe(lambda: queries.gastos_detalle_categoria(num), {})
    return render_template(
        "informes/gastos_detalle.html",
        data=data,
        num=num,
        error=error,
    )


@informes_bp.route("/retiros")
@requiere_login
@requiere_permiso("informes.ver")
def retiros():
    """Dividendos — 2 tabs (mes/año) con KPIs combinados arriba.

    TMT 2026-05-20 v2 — pedido dueña: unificar pantallas de retiros.
    Reemplaza la antigua /capital + /retiros con un solo destino. Cada
    tab muestra los retiros del periodo, pero los KPIs (mes + año)
    aparecen siempre en ambas.
    """
    tab = (request.args.get("tab") or "mes").strip().lower()
    if tab not in ("mes", "anio"):
        tab = "mes"

    if tab == "anio":
        filas, error = _safe(queries.retiros_del_anio_actual, [])
    else:
        filas, error = _safe(queries.retiros_del_mes_actual, [])

    if request.args.get("export") == "csv":
        return csv_response(
            filas,
            columnas=[
                ("fecha", "Fecha"),
                ("concepto", "Concepto"),
                ("ret", "Importe"),
            ],
            filename=f"dividendos_{tab}.csv",
        )

    # KPIs — siempre mes + año (visibles en ambas tabs).
    total_mes, _ = _safe(queries.retiros_total_mes_actual, 0.0)
    total_anual, _ = _safe(queries.retiros_total_anual, 0.0)

    # Conteos para los badges del switcher de tabs (best-effort).
    try:
        n_mes = len(queries.retiros_del_mes_actual())
        n_anio = len(queries.retiros_del_anio_actual())
    except Exception:  # noqa: BLE001
        n_mes, n_anio = 0, 0

    return render_template(
        "informes/retiros.html",
        filas=filas,
        tab=tab,
        total_mes=total_mes,
        total_anual=total_anual,
        n_mes=n_mes,
        n_anio=n_anio,
        error=error,
    )


@informes_bp.route("/activos")
@requiere_login
@requiere_permiso("informes.ver")
def activos():
    filas, error = _safe(queries.activos_lista, [])
    if request.args.get("export") == "csv":
        return csv_response(
            filas,
            columnas=[
                ("fecha", "Fecha"),
                ("concepto", "Concepto"),
                ("tipo", "Tipo"),
                ("proveedor", "Proveedor"),
                ("inicial", "Inicial"),
                ("amortizac", "Amort. acum."),
                ("amortimes", "Amort. mes"),
                ("valor", "Valor neto"),
                ("cuota", "Cuota"),
                ("vida_util", "Vida útil"),
                ("ult_mes_amortizado", "Últ. mes amort."),
            ],
            filename="activos_fijos.csv",
        )
    return render_template("informes/activos.html", filas=filas, error=error)


@informes_bp.route("/historia/multianual")
@requiere_login
@requiere_permiso("informes.ver")
def historia_multianual():
    """Vista cruzada mes × año — replica INFORMES.PRG L1336-1550 modo '1/2/3'.

    Muestra los últimos N meses (default 12) con las métricas principales
    (patrimonio, ventas U$, utilidad U$, kg vendidos, stock MP+PT, etc.)
    desplegadas por año (corriente + 2 anteriores) y con la variación %
    año contra año. Útil para detectar tendencias estacionales.
    """
    meses = request.args.get("meses", default=12, type=int)
    data, error = _safe(lambda: queries.historia_multianual(meses), {})
    return render_template(
        "informes/historia_multianual.html",
        data=data,
        meses=meses,
        error=error,
    )


@informes_bp.route("/historia")
@requiere_login
@requiere_permiso("informes.ver")
def historia():
    filas, error = _safe(queries.historia_lista, [])
    if request.args.get("export") == "csv":
        return csv_response(
            filas,
            columnas=[
                ("fecha", "Mes"),
                ("stock", "Stock"),
                ("kcom", "Kg compra"),
                ("ktej", "Kg tejido"),
                ("ktin", "Kg tinto"),
                ("ustock", "U stock"),
                ("uqui", "U químicos"),
                ("kvent", "Kg venta"),
                ("uvent", "U venta"),
                ("costo", "Costo"),
                ("ucom", "U compra"),
                ("utej", "U tejido"),
                ("utin", "U tinto"),
                ("gasto", "Gasto mes"),
                ("gstotal", "Gasto total"),
                ("banco", "Banco"),
                ("cart", "Cartera"),
                ("deuda", "Deuda"),
                ("retiro", "Retiro"),
                ("patrimonio", "Patrimonio"),
                ("anticipos", "Anticipos"),
                ("dolar", "Dólar"),
                ("maquinaria", "Maquinaria"),
                ("realty", "Inmueble"),
                ("usret", "USD retiro"),
                ("usuti", "USD utilidad"),
            ],
            filename="historia_mensual.csv",
        )
    return render_template("informes/historia.html", filas=filas, error=error)


@informes_bp.route("/iniciales")
@requiere_login
@requiere_permiso("informes.ver")
def iniciales():
    anio = request.args.get("anio", type=int)
    filas, error = _safe(lambda: queries.iniciales_lista(anio), [])
    if request.args.get("export") == "csv":
        return csv_response(
            filas,
            columnas=[
                ("yy", "Año"),
                ("mesnum", "#"),
                ("mesnom", "Mes"),
                ("hilado", "Hilado"),
                ("tejido", "Tejido"),
                ("terminado", "Terminado"),
                ("vq", "VQ"),
                ("um", "UM"),
                ("uk", "UK"),
                ("uf", "UF"),
                ("uq", "UQ"),
                ("pre", "Precio"),
                ("kprog", "Kg prog."),
                ("gprog", "Gasto prog."),
                ("numnot", "# notas"),
                ("dificil", "Dificultad"),
                ("pretej", "Precio tej."),
                ("pretin", "Precio tin."),
                ("preadm", "Precio adm."),
                ("pretot", "Precio tot."),
            ],
            filename=f"iniciales_{anio or 'todos'}.csv",
        )
    return render_template("informes/iniciales.html", filas=filas, anio=anio, error=error)


# ---------------------------------------------------------------------------
# Stock inicial mensual (foto de Asinfo persistida) — TMT 2026-07-08
# ---------------------------------------------------------------------------
# Asinfo es live-only: para tener un "stock inicial" por mes se toma una foto
# del inventario live (bodegas 51/52/53 + WIP) y se persiste en
# scintela.stock_inicial_mes. Ver modules/informes/stock_inicial.py.

@informes_bp.route("/stock-inicial")
@requiere_login
@requiere_permiso("informes.ver")
def stock_inicial():
    """Muestra las fotos de stock inicial guardadas por mes + form de captura."""
    from . import stock_inicial as si_q

    hoy = today_ec()
    meses, error = _safe(lambda: si_q.meses_capturados(24), [])
    if not isinstance(meses, list):
        meses = []
    return render_template(
        "informes/stock_inicial.html",
        meses=meses,
        anio=hoy.year,
        mes=hoy.month,
        etapa_label=si_q.ETAPA_LABEL,
        error=error,
    )


@informes_bp.route("/stock-inicial/capturar", methods=["POST"])
@requiere_login
@requiere_permiso("informes.editar")  # escribe una foto — mismo gate que snapshot-mes
def stock_inicial_capturar():
    """Toma la foto live de Asinfo y la persiste para (anio, mes). Fail-soft."""
    from . import stock_inicial as si_q

    try:
        anio = int(request.form.get("anio") or today_ec().year)
        mes = int(request.form.get("mes") or today_ec().month)
    except (TypeError, ValueError):
        flash("Parámetros inválidos.", "error")
        return redirect(url_for("informes.stock_inicial"))
    mes = max(1, min(mes, 12))
    try:
        usuario = (g.user or {}).get("username", "web")
        r = si_q.capturar(anio, mes, usuario=usuario)
        if r.get("aplicado"):
            flash(r.get("razon", f"Foto {mes:02d}/{anio} guardada."), "ok")
        else:
            flash(r.get("razon", "Asinfo no disponible — nada que guardar."), "info")
    except Exception as e:
        flash_exc("Captura de stock inicial falló", e)
    return redirect(url_for("informes.stock_inicial"))


@informes_bp.route("/estado-cuenta", methods=["GET"])
@requiere_login
# TMT 2026-07-09 (dueña "todos los usuarios tienen acceso a estado de cuenta"):
# sin gate de permiso — cualquier usuario logueado puede ver estados de cuenta.
def estado_cuenta_landing():
    """Landing/lookup page para estado de cuenta de cliente.

    Muestra top deudores (los candidatos más probables a mirar) + un buscador
    por código o nombre. Si el usuario envía ?codigo=XYZ lo redirige al
    estado de cuenta. Si busca por nombre, lista los matches.
    """
    codigo = (request.args.get("codigo") or "").strip().upper()
    if codigo:
        return redirect(url_for("informes.estado_cuenta", codigo_cli=codigo))

    busqueda = (request.args.get("q") or "").strip()
    matches: list[dict] = []
    if busqueda:
        matches, _ = _safe(lambda: queries.buscar_clientes(busqueda), [])

    top, error = _safe(queries.cartera_por_cliente, [])
    # top 10 deudores como atajos
    top = top[:10] if top else []
    return render_template(
        "informes/estado_cuenta_landing.html",
        top=top,
        matches=matches,
        q=busqueda,
        error=error,
    )


# Provincias del Ecuador (canónicas, UPPER) para normalizar el texto libre de
# scintela.cliente.provincia (viene truncado a ~10 chars y con typos).
_PROVINCIAS_EC = (
    "AZUAY", "BOLIVAR", "CAÑAR", "CARCHI", "CHIMBORAZO", "COTOPAXI", "EL ORO",
    "ESMERALDAS", "GALAPAGOS", "GUAYAS", "IMBABURA", "LOJA", "LOS RIOS",
    "MANABI", "MORONA SANTIAGO", "NAPO", "ORELLANA", "PASTAZA", "PICHINCHA",
    "SANTA ELENA", "SANTO DOMINGO", "SUCUMBIOS", "TUNGURAHUA", "ZAMORA CHINCHIPE",
)
# Typos / ciudades que no matchean por prefijo → canónica.
_PROV_ALIAS = {
    "TUNGURAGUA": "TUNGURAHUA", "TUNGUAHUA": "TUNGURAHUA", "AMBATO": "TUNGURAHUA",
    "PICHICHA": "PICHINCHA", "PICHIHCHA": "PICHINCHA",
    "STO DOMING": "SANTO DOMINGO", "STO DMGO": "SANTO DOMINGO",
    "STO DOMIGO": "SANTO DOMINGO", "STO DMG": "SANTO DOMINGO",
    "GUAYAQUIL": "GUAYAS", "STA ELENA": "SANTA ELENA", "TENA": "NAPO",
}


def _normalizar_provincia(raw) -> str:
    """Texto libre de provincia → provincia canónica del Ecuador.

    Colapsa mayúsc/espacios, truncados (VARCHAR 10 → 'TUNGURAHU', 'STO DOMING'),
    typos ('TUNGURAGUA', 'PICHICHA') y algunas ciudades ('AMBATO'→Tungurahua).
    Basura numérica o irreconocible → '(sin provincia)'.
    """
    s = (raw or "").strip().upper()
    if not s or s == "(SIN PROVINCIA)":
        return "(sin provincia)"
    if s in _PROV_ALIAS:
        return _PROV_ALIAS[s]
    if s in _PROVINCIAS_EC:
        return s
    # Truncados / con sufijo: prefijo en cualquier dirección (STO..., GUAYAS 15D).
    for p in _PROVINCIAS_EC:
        if p.startswith(s) or s.startswith(p):
            return p
    if not any(ch.isalpha() for ch in s):
        return "(sin provincia)"
    return s


def _ec_group_key(por, r):
    """(clave, etiqueta) de agrupación de un cliente según la dimensión.

    Compartida por el listado agrupado y por la impresión en lote.
    """
    if por == "vendedor":
        if r.get("vendedor_activo") and (r.get("vend") or "").strip():
            return (r.get("vend"), r.get("vendedor_nombre") or r.get("vend"))
        return ("~", "(sin vendedor)")
    if por == "provincia":
        prov = _normalizar_provincia(r.get("provincia"))
        return (prov, prov)
    return (r.get("grupo_codigo") or "~", r.get("grupo_nombre") or "(sin grupo)")


@informes_bp.route("/estado-cuenta/grupos", methods=["GET"])
@requiere_login
# TMT 2026-07-09 (dueña): estado de cuenta abierto a todos los usuarios logueados.
def estado_cuenta_grupos():
    """Estado de cuenta de TODOS los clientes con saldo, AGRUPADO para imprimir.

    Dimensión elegible por ?por= : vendedor (quién atiende), grupo (grupo de
    clientes) o provincia. Pedido dueña 2026-07-09 ("imprimir por grupos: por
    vendedor, grupo clientes y por provincia"). Reemplaza la PROCEDURE GRUPOS.
    """
    por = (request.args.get("por") or "vendedor").strip().lower()
    if por not in ("vendedor", "grupo", "provincia"):
        por = "vendedor"
    sel = (request.args.get("sel") or "").strip()
    filas, error = _safe(queries.estado_cuenta_clientes_saldos, [])
    filas = filas or []

    def _keylabel(r):
        return _ec_group_key(por, r)

    ctx = {"por": por, "error": error, "sel": sel}

    # GRUPOS: mostrar SOLO grupos reales (2+ clientes); los que están solos no
    # se muestran. Todos los grupos juntos, uno debajo del otro.
    if por == "grupo":
        gmap: dict = {}
        for r in filas:
            k, label = _keylabel(r)
            grp = gmap.setdefault(k, {"label": label, "clientes": [], "saldo": 0.0})
            grp["clientes"].append(r)
            grp["saldo"] += float(r.get("saldo") or 0)
        grupos = [g for g in gmap.values() if len(g["clientes"]) >= 2]
        grupos.sort(key=lambda x: x["saldo"], reverse=True)
        for grp in grupos:
            grp["clientes"].sort(key=lambda r: float(r.get("saldo") or 0), reverse=True)
        ctx.update(
            mode="grupo",
            grupos=grupos,
            total=sum(g["saldo"] for g in grupos),
            n_clientes=sum(len(g["clientes"]) for g in grupos),
        )
        return render_template("informes/estado_cuenta_grupos.html", **ctx)

    # VENDEDOR / PROVINCIA: elegir uno (sel) e imprimir SUS clientes. Sin sel,
    # mostrar la lista de opciones para elegir.
    if sel:
        clientes = []
        sel_label = sel
        for r in filas:
            k, label = _keylabel(r)
            if str(k) == sel:
                clientes.append(r)
                sel_label = label
        clientes.sort(key=lambda r: float(r.get("saldo") or 0), reverse=True)
        ctx.update(
            mode="list",
            sel_label=sel_label,
            clientes=clientes,
            total=sum(float(r.get("saldo") or 0) for r in clientes),
        )
        return render_template("informes/estado_cuenta_grupos.html", **ctx)

    opt_map: dict = {}
    for r in filas:
        k, label = _keylabel(r)
        o = opt_map.setdefault(k, {"sel": k, "label": label, "n": 0, "saldo": 0.0})
        o["n"] += 1
        o["saldo"] += float(r.get("saldo") or 0)
    options = sorted(opt_map.values(), key=lambda x: x["saldo"], reverse=True)
    ctx.update(
        mode="picker",
        options=options,
        total=sum(o["saldo"] for o in options),
        n_clientes=len(filas),
    )
    return render_template("informes/estado_cuenta_grupos.html", **ctx)


@informes_bp.route("/estado-cuenta/imprimir", methods=["GET"])
@requiere_login
def estado_cuenta_lote_imprimir():
    """Imprime el estado de cuenta COMPLETO de cada cliente de la selección,
    uno tras otro (no el resumen). Dueña 2026-07-09: "cuando pongo imprimir
    sean todos los estados de cuenta completos, no el total. uno por uno".
    """
    por = (request.args.get("por") or "vendedor").strip().lower()
    if por not in ("vendedor", "grupo", "provincia"):
        por = "vendedor"
    sel = (request.args.get("sel") or "").strip()
    filas, _err = _safe(queries.estado_cuenta_clientes_saldos, [])
    filas = filas or []

    # Códigos de la selección, en orden de impresión.
    codes: list[str] = []
    titulo = ""
    if por == "grupo":
        gmap: dict = {}
        for r in filas:
            k, label = _ec_group_key("grupo", r)
            g = gmap.setdefault(k, {"label": label, "clientes": [], "saldo": 0.0})
            g["clientes"].append(r)
            g["saldo"] += float(r.get("saldo") or 0)
        grupos = [g for g in gmap.values() if len(g["clientes"]) >= 2]
        grupos.sort(key=lambda x: x["saldo"], reverse=True)
        for g in grupos:
            g["clientes"].sort(key=lambda r: float(r.get("saldo") or 0), reverse=True)
            codes += [c["codigo_cli"] for c in g["clientes"]]
        titulo = "Grupos de clientes"
    else:
        sub = []
        for r in filas:
            k, label = _ec_group_key(por, r)
            if str(k) == sel:
                sub.append(r)
                titulo = ("Vendedor: " if por == "vendedor" else "Provincia: ") + label
        sub.sort(key=lambda r: float(r.get("saldo") or 0), reverse=True)
        codes = [r["codigo_cli"] for r in sub]

    # Estado de cuenta completo por cliente (facturas + totales).
    clientes = []
    for code in codes:
        d, _e = _safe(lambda c=code: queries.estado_cuenta_cliente(c), {})
        if d and d.get("cliente"):
            clientes.append(d)

    return render_template(
        "informes/estado_cuenta_lote_print.html",
        clientes=clientes,
        titulo=titulo,
        por=por,
        n=len(clientes),
    )


@informes_bp.route("/estado-cuenta/<codigo_cli>")
@requiere_login
# TMT 2026-07-09 (dueña): estado de cuenta abierto a todos los usuarios logueados.
def estado_cuenta(codigo_cli):
    codigo_up = codigo_cli.upper()
    data, error = _safe(lambda: queries.estado_cuenta_cliente(codigo_up), {})
    if not data or not data.get("cliente"):
        abort(404)
    try:
        from modules.recientes import queries as rec

        cli = data.get("cliente") or {}
        rec.registrar(
            "cliente",
            codigo_up,
            etiqueta=f"{codigo_up} — {cli.get('nombre') or ''}",
        )
    except Exception:
        pass
    # Lista de clientes para el cargador "Nuevo estado de cuenta" (autocomplete).
    try:
        from modules.autocomplete.queries import clientes_para_datalist
        _clientes_dl = clientes_para_datalist()
    except Exception:  # noqa: BLE001
        _clientes_dl = []
    return render_template(
        "informes/estado_cuenta.html",
        data=data,
        error=error,
        clientes_datalist=_clientes_dl,
        # TMT 2026-07-09 (dueña): facturas totalizadas (T), para poder
        # REABRIRLAS desde el panel A↔T. Solo en la vista individual.
        facturas_totalizadas=(_safe(
            lambda: queries.facturas_totalizadas_cliente(codigo_up), [])[0]),
        # TMT 2026-07-06 (dueña): banner de éxito del TOTALIZAR (one-shot,
        # patrón cobranza_ok) + botón imprimir el resultado.
        totalizar_ok=session.pop("totalizar_ok", None),
        neteo_ok=session.pop("neteo_ok", None),
    )


@informes_bp.route("/estado-cuenta/<codigo_cli>/factura/<int:id_factura>/toggle-stat",
                   methods=["POST"])
@requiere_login
@requiere_permiso("clientes.ver")
def estado_cuenta_factura_toggle(codigo_cli, id_factura):
    """Cerrar (A/Z→T) o reabrir (T→A) UNA factura desde el estado de cuenta.

    TMT 2026-07-09 (dueña): "poder pasar facturas de A→T y T→A". Mismo permiso
    que TOTALIZAR (clientes.ver, decisión dueña #4) — lo usan todos los que
    gestionan la cuenta.
    """
    codigo_up = codigo_cli.upper()
    usuario = (g.user or {}).get("username", "web") if hasattr(g, "user") else "web"
    try:
        res = queries.factura_cambiar_stat_a_t(id_factura, codigo_up, usuario=usuario)
        if res.get("accion") == "cerrada":
            flash(f"Factura {res.get('numf')} cerrada (→T).", "ok")
        else:
            flash(
                f"Factura {res.get('numf')} reabierta (→A, saldo "
                f"{res.get('saldo_nuevo', 0):,.2f}).", "ok")
    except ValueError as e:
        flash(str(e), "warn")
    except Exception as e:
        flash_exc("No pude cambiar el estado de la factura", e)
    return redirect(url_for("informes.estado_cuenta", codigo_cli=codigo_up))


@informes_bp.route("/estado-cuenta/<codigo_cli>/netear", methods=["POST"])
@requiere_login
@requiere_permiso("cheques.anular")
def estado_cuenta_netear(codigo_cli):
    """Netear (anular) cheque(s) contra anticipo(s) del cliente.

    TMT 2026-07-09 (dueña): "cancelar cheques y anticipos (netearlos)". Los dos
    lados se cancelan entre sí (stat='X') si suman igual. Gate cheques.anular.
    """
    from modules.cheques import queries as chq
    codigo_up = codigo_cli.upper()
    usuario = (g.user or {}).get("username", "web") if hasattr(g, "user") else "web"
    ids_cheques = [int(x) for x in request.form.getlist("ch") if x.strip().isdigit()]
    ids_anticipos = [int(x) for x in request.form.getlist("ant") if x.strip().isdigit()]
    try:
        res = chq.netear_cheques_con_anticipos(
            codigo_cli=codigo_up, ids_cheques=ids_cheques,
            ids_anticipos=ids_anticipos, usuario=usuario,
        )
        session["neteo_ok"] = res
        _reab = res.get("facturas_reabiertas") or []
        _msg = (
            f"Neteado: {res['n_cheques']} cheque(s) y {res['n_anticipos']} "
            f"anticipo(s) por {res['total']:,.2f} anulados entre sí."
        )
        if _reab:
            _nums = ", ".join(
                (f"#{r['numf']}" if r.get("numf") else f"id {r['id_factura']}")
                for r in _reab
            )
            _msg += (
                f" Se reabrió(eron) {len(_reab)} factura(s) que el/los "
                f"cheque(s) tenían aplicada(s): {_nums} (quedan con saldo "
                "pendiente; reversible desde el Historial)."
            )
        flash(_msg, "ok")
    except ValueError as e:
        flash(str(e), "warn")
    except Exception as e:
        flash_exc("No pude netear", e)
    return redirect(url_for("informes.estado_cuenta", codigo_cli=codigo_up))


@informes_bp.route("/estado-cuenta/<codigo_cli>/totalizar", methods=["GET", "POST"])
@requiere_login
@requiere_permiso("clientes.ver")
def estado_cuenta_totalizar(codigo_cli):
    """TOTALIZAR estado de cuenta — re-liquidación FIFO (CUENTA.PRG rama 'Y').

    TMT 2026-07-06 (dueña): junta todos los abonos de las facturas vivas del
    cliente y los redistribuye de la más vieja a la más nueva. GET = pantalla
    de confirmación (preview de lo que va a quedar, estilo dBase, imprimible);
    POST = ejecuta (IRREVERSIBLE — se pierden los vínculos cheque↔factura) y
    vuelve al estado de cuenta con banner de éxito.

    Mismo permiso que el estado de cuenta (decisión dueña #4): lo usan todos
    los que ven la cuenta.
    """
    codigo_up = codigo_cli.upper()
    # TMT 2026-07-06 (dueña): "¿cómo elegís hasta dónde totalizar?" — corte
    # opcional por fecha (inclusive). Vacío = todas las vivas.
    _hasta_raw = (request.values.get("hasta") or "").strip()
    _hasta = None
    if _hasta_raw:
        try:
            from datetime import date as _date
            _hasta = _date.fromisoformat(_hasta_raw)
        except ValueError:
            flash("Fecha 'hasta' inválida — se ignora el corte.", "warn")
            _hasta = None
    if request.method == "POST":
        usuario = (g.user or {}).get("username", "web") if hasattr(g, "user") else "web"
        try:
            res = queries.totalizar_estado_cuenta_ejecutar(
                codigo_up, usuario=usuario, hasta=_hasta)
        except ValueError as e:
            flash(str(e), "warn")
            return redirect(url_for("informes.estado_cuenta", codigo_cli=codigo_up))
        except Exception as e:
            flash_exc("No pude totalizar la cuenta", e)
            return redirect(url_for("informes.estado_cuenta_totalizar", codigo_cli=codigo_up))
        session["totalizar_ok"] = {
            "codigo_cli": res["codigo_cli"],
            "n_facturas": res["n_facturas"],
            "pool": res["pool"],
            "n_T": res["n_T"],
            "n_A": res["n_A"],
            "n_Z": res["n_Z"],
            "n_links_borrados": res["n_links_borrados"],
        }
        return redirect(url_for("informes.estado_cuenta", codigo_cli=codigo_up))

    data, error = _safe(
        lambda: queries.totalizar_estado_cuenta_preview(codigo_up, hasta=_hasta), {})
    if not data or not data.get("cliente"):
        abort(404)
    return render_template(
        "informes/totalizar_preview.html", data=data, error=error, hoy=today_ec(),
        hasta=_hasta_raw,
    )


# ---------------------------------------------------------------------------
# Gastos forzados — endpoints JSON. Migración localStorage → DB.
# Pedido dueña 2026-05-19 v8: cargás en Chrome, abrís en Safari, no
# aparecía nada. Ahora la fuente de verdad es scintela.gasto_forzado.
# El JS del flujo_grafico.html llama a estos endpoints (en lugar de
# tocar localStorage) — ver bloque `gfLoad/gfSave` del template.
# ---------------------------------------------------------------------------


def _parse_fecha_iso(s: str):
    """YYYY-MM-DD → date, o None si no parsea."""
    if not s:
        return None
    try:
        return datetime.strptime(str(s).strip(), "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def _parse_importe_payload(raw) -> float | None:
    """Acepta float/int/str con '.' o ','. Devuelve None si no parsea."""
    if raw is None or raw == "":
        return None
    try:
        return float(str(raw).replace(",", "."))
    except (TypeError, ValueError):
        return None


@informes_bp.route("/flujo/gastos-forzados", methods=["GET"])
@requiere_login
@requiere_permiso("informes.ver")
def gastos_forzados_listar():
    try:
        items = queries.gastos_forzados_listar()
        return jsonify({"ok": True, "items": items})
    except Exception as e:  # noqa: BLE001
        return jsonify({"ok": False, "error": str(e)}), 500


@informes_bp.route("/flujo/gastos-forzados", methods=["POST"])
@requiere_login
@requiere_permiso("informes.editar")
def gastos_forzados_crear():
    payload = request.get_json(silent=True) or {}
    fecha = _parse_fecha_iso(payload.get("fecha"))
    importe = _parse_importe_payload(payload.get("importe"))
    concepto = (payload.get("concepto") or "").strip()[:80]
    prov = (payload.get("prov") or "").strip().upper()[:5]
    if not fecha or importe is None or importe <= 0:
        return jsonify(
            {
                "ok": False,
                "error": "Datos inválidos: fecha (YYYY-MM-DD) y importe > 0 requeridos.",
            }
        ), 400
    usuario = (g.user or {}).get("username", "web")
    try:
        item = queries.gasto_forzado_crear(
            fecha=fecha,
            importe=importe,
            concepto=concepto,
            prov=prov,
            usuario=usuario,
        )
        return jsonify({"ok": True, "item": item}), 201
    except Exception as e:  # noqa: BLE001
        return jsonify({"ok": False, "error": str(e)}), 500


@informes_bp.route(
    "/flujo/gastos-forzados/<int:id_gasto>",
    methods=["PUT", "PATCH"],
)
@requiere_login
@requiere_permiso("informes.editar")
def gastos_forzados_actualizar(id_gasto: int):
    payload = request.get_json(silent=True) or {}
    expected = payload.get("expected_version")
    if expected is None:
        return jsonify({"ok": False, "error": "expected_version requerido"}), 400
    try:
        expected_v = int(expected)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "expected_version inválido"}), 400
    fecha = _parse_fecha_iso(payload.get("fecha")) if "fecha" in payload else None
    importe = _parse_importe_payload(payload.get("importe")) if "importe" in payload else None
    concepto = (payload.get("concepto") or "").strip()[:80] if "concepto" in payload else None
    prov = (payload.get("prov") or "").strip().upper()[:5] if "prov" in payload else None
    usuario = (g.user or {}).get("username", "web")
    try:
        r = queries.gasto_forzado_actualizar(
            id_gasto_forzado=id_gasto,
            expected_version=expected_v,
            fecha=fecha,
            importe=importe,
            concepto=concepto,
            prov=prov,
            usuario=usuario,
        )
    except Exception as e:  # noqa: BLE001
        return jsonify({"ok": False, "error": str(e)}), 500
    if not r.get("ok"):
        status = 409 if r.get("reason", "").startswith("version_conflict") else 404
        return jsonify(r), status
    return jsonify(r)


@informes_bp.route(
    "/flujo/gastos-forzados/<int:id_gasto>",
    methods=["DELETE"],
)
@requiere_login
@requiere_permiso("informes.editar")
def gastos_forzados_eliminar(id_gasto: int):
    try:
        ok = queries.gasto_forzado_eliminar(id_gasto)
        if not ok:
            return jsonify({"ok": False, "error": "not_found"}), 404
        return jsonify({"ok": True})
    except Exception as e:  # noqa: BLE001
        return jsonify({"ok": False, "error": str(e)}), 500


@informes_bp.route(
    "/informes/flujo/gastos-forzados/importar",
    methods=["POST"],
)
@requiere_login
@requiere_permiso("informes.editar")
def gastos_forzados_importar():
    """One-time migration: el cliente envía el contenido de localStorage
    `flujo_gastos_forzados_v1` y los inserta en DB (dedup por
    fecha+importe+concepto)."""
    payload = request.get_json(silent=True) or {}
    items = payload.get("items") or []
    if not isinstance(items, list):
        return jsonify({"ok": False, "error": "items debe ser lista"}), 400
    usuario = (g.user or {}).get("username", "web")
    try:
        r = queries.gastos_forzados_importar_bulk(items, usuario=usuario)
        return jsonify({"ok": True, **r})
    except Exception as e:  # noqa: BLE001
        return jsonify({"ok": False, "error": str(e)}), 500


@informes_bp.route("/flujo-fondos")
@requiere_login
@requiere_permiso("informes.ver")
def flujo_fondos():
    """Flujo de Fondos DIARIO — réplica de la opción 6 del dBase (MENU.PRG
    PROCEDURE FLUJO L560-720 + FLUJO.DBF).

    TMT 2026-07-06 (dueña) "flujo de fondos no está como el dBase": antes
    esta pantalla agrupaba por SEMANA con columnas propias. El dBase muestra
    UNA FILA POR DÍA con vencimientos, columnas FECHA / PICH / INTER / CHEQ /
    MAT.PR / GASTS / SALDO, primera fila = arranque (hoy: Pichincha + Inter +
    Caja) y un corte por semana con el acumulado de ingresos/egresos
    (ACUMI/ACUME). Reescrito para calcar eso 1:1 — el cálculo vive en
    queries.flujo_fondos_diario() (mismos helpers que usa el gráfico).

    Igual que antes: por default las facturas NO suman (dBase L701 hardcodea
    FA=0 — flujo conservador, solo cheques en mano); ?incluir_facturas=1
    proyecta también los cobros de cartera a VENCIM+50 (regla dBase L647-649).
    """
    incluir_fact = request.args.get("incluir_facturas") == "1"
    data, error = _safe(
        lambda: queries.flujo_fondos_diario(incluir_facturas=incluir_fact),
        None,
    )
    if not data:
        data = {
            "arranque": {"s1": 0.0, "s2": 0.0, "p1": 0.0, "p2": 0.0,
                         "caja": 0.0, "total": 0.0},
            "filas": [], "saldo_final": 0.0, "saldo_min": 0.0,
            "fecha_min": today_ec(),
        }

    # TMT 2026-07-08 (dueña "estos -87k no sé de dónde salen"): enriquecemos cada
    # fila-día con el DETALLE de los posdat que componen GASTOS y MAT.PRIMA, para
    # mostrarlo en un tooltip. Best-effort: si falla, la tabla sigue igual.
    # TMT 2026-07-08 (dueña "no bajamos tanto de la nada"): DESGLOSE del saldo
    # final en sus componentes, para que se vea que la caída NO viene de la
    # deuda actual sino de las COMPRAS FUTURAS de hilado (banc=9) que el flujo
    # proyecta sin las ventas futuras que las pagan (FA=0). Reusa el detalle ya
    # traído. Best-effort: si falla, la página sigue igual sin la caja.
    desglose = None
    try:
        _det = queries.flujo_egresos_detalle()
        _by_day: dict = {}
        for _it in _det:
            _by_day.setdefault((_it["fecha"], _it["tipo"]), []).append(_it)
        for _f in data.get("filas", []):
            if _f.get("tipo") == "dia":
                _f["gastos_det"] = _by_day.get((_f.get("fecha"), "gasto"), [])
                _f["mprima_det"] = _by_day.get((_f.get("fecha"), "mprima"), [])
        # Buckets por banc: 0 = deuda actual (pasivos); 9 = importaciones de
        # hilado a futuro; 1/2 = cheques emitidos sin debitar (van al arranque).
        _egr_deuda = sum(i["importe"] for i in _det if i["banc"] == 0)
        _egr_import = sum(i["importe"] for i in _det if i["banc"] == 9)
        _ing_cheques = sum(
            f.get("cheques", 0.0) for f in data.get("filas", [])
            if f.get("tipo") == "dia"
        )
        _arr = data["arranque"]["total"]
        desglose = {
            "arranque": round(_arr, 2),
            "cheques": round(_ing_cheques, 2),
            "deuda": round(_egr_deuda, 2),
            "importaciones": round(_egr_import, 2),
            # Saldo "operativo": lo que quedaría si NO proyectáramos las compras
            # futuras de hilado (que se pagan con ventas futuras no contadas).
            "sin_importaciones": round(_arr + _ing_cheques - _egr_deuda, 2),
        }
    except Exception:  # noqa: BLE001
        pass

    return render_template(
        "informes/flujo_fondos.html",
        arranque=data["arranque"],
        filas=data["filas"],
        saldo_final=data["saldo_final"],
        saldo_min=data["saldo_min"],
        fecha_min=data["fecha_min"],
        hoy=today_ec(),
        incluir_fact=incluir_fact,
        desglose=desglose,
        error=error,
    )
