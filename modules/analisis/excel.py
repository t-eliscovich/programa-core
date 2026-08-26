"""Los dos Excel de la sección Análisis, con formato de verdad.

⭐ Dueña 26/08/2026: *"bajar a excel se baja horrible: bajalo a algo con
formato"*, y enseguida *"y lo mismo con los vendidos"*.

Lo que bajaba era un CSV armado en el navegador: punto y coma, todo texto,
columnas de un caracter de ancho y los kilos como cadenas —Excel los abría
alineados a la izquierda y no se podían sumar—. Ahora es un `.xlsx` real:
encabezado fijo, filtro automático, anchos pensados y los números como números.

⚠ Baja SIEMPRE todo, sin los filtros de la pantalla. No es un olvido: los
filtros son de JavaScript y replicarlos acá sería escribir la misma regla dos
veces en dos lenguajes — el día que una cambie, el archivo diría algo distinto
de la pantalla sin que nadie se entere. Y el motivo por el que uno baja algo a
Excel es justamente filtrarlo y pivotearlo ahí: por eso van GRUPO y TELA como
columnas propias. El botón lo dice.
"""
from __future__ import annotations

import io
from datetime import date

from flask import Response

#: Ancho de columna, en caracteres de Excel. Sin esto, "Jersey Forro Spun"
#: entra en una columna de 8 y se lee "Jersey F…".
_ENCABEZADO_BG = "0F172A"


def libro(hojas: list[dict]) -> bytes:
    """Arma el .xlsx. `hojas` = [{titulo, columnas:[(rotulo, ancho, formato)],
    filas:[[valor, ...]]}]."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    negrita = Font(bold=True, color="FFFFFF", size=10)
    relleno = PatternFill("solid", fgColor=_ENCABEZADO_BG)
    linea = Border(bottom=Side(style="thin", color="E2E8F0"))

    for hoja in hojas:
        ws = wb.create_sheet(hoja["titulo"][:31])
        cols = hoja["columnas"]
        for i, (rotulo, ancho, _fmt) in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=rotulo)
            c.font = negrita
            c.fill = relleno
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = ancho
        ws.row_dimensions[1].height = 24

        for f, valores in enumerate(hoja["filas"], 2):
            for i, ((_r, _a, fmt), v) in enumerate(zip(cols, valores, strict=False), 1):
                c = ws.cell(row=f, column=i, value=v)
                c.border = linea
                if fmt:
                    c.number_format = fmt
                    if not isinstance(v, date):
                        c.alignment = Alignment(horizontal="right")

        # ⚠ El encabezado FIJO y el filtro automático: son 700 renglones, y sin
        # esto a la fila 200 ya no se sabe qué columna se está mirando.
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(len(hoja['filas']) + 1, 2)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def respuesta(nombre: str, datos: bytes) -> Response:
    return Response(
        datos,
        mimetype=("application/vnd.openxmlformats-officedocument"
                  ".spreadsheetml.sheet"),
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )
