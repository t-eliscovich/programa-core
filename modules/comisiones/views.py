"""Comisiones de vendedores — list + inline edit + detalle mensual."""

import io

from flask import (
    Blueprint,
    Response,
    abort,
    flash,
    g,
    redirect,
    render_template,
    request,
    url_for,
)

from auth import requiere_login, requiere_permiso
from error_messages import flash_exc
from exports import csv_response
from filters import today_ec
from parsers import parse_monto

from . import queries

comisiones_bp = Blueprint("comisiones", __name__, template_folder="templates")


def _yy_mm() -> tuple[int, int]:
    """Lee ?anio= y ?mes= del query string. Default: mes en curso."""
    hoy = today_ec()
    try:
        yy = int(request.args.get("anio") or hoy.year)
    except (TypeError, ValueError):
        yy = hoy.year
    try:
        mm = int(request.args.get("mes") or hoy.month)
    except (TypeError, ValueError):
        mm = hoy.month
    mm = max(1, min(mm, 12))
    return yy, mm


@comisiones_bp.route("/comisiones")
@requiere_login
@requiere_permiso("comisiones.ver")
def lista():
    yy, mm = _yy_mm()
    try:
        filas = queries.lista(anio=yy, mes=mm)
        error = None
    except Exception as e:
        filas = []
        msg = str(e)
        # TMT 2026-05-18 — mensaje legible si la migración 0032 no corrió.
        if "scintela.vendedor" in msg and "does not exist" in msg:
            error = (
                "La tabla scintela.vendedor todavía no existe. Aplicá la migración: python scripts/migrate.py"
            )
        else:
            error = msg

    if request.args.get("export") == "csv":
        return csv_response(
            filas,
            columnas=[
                ("codigo", "Código"),
                ("nombre", "Nombre"),
                ("pct_comision", "% comisión"),
                ("n_clientes", "N° clientes"),
                ("ventas_mes", "Ventas mes"),
                ("cobranzas_mes", "Cobranzas mes"),
                ("comision_mes", "Comisión mes"),
            ],
            filename=f"comisiones_{yy}_{mm:02d}.csv",
        )

    totales = {
        "cobranzas": sum(float(f.get("cobranzas_mes") or 0) for f in filas),
        "ventas": sum(float(f.get("ventas_mes") or 0) for f in filas),
        "comision": sum(float(f.get("comision_mes") or 0) for f in filas),
    }

    return render_template(
        "comisiones/lista.html",
        filas=filas,
        totales=totales,
        anio=yy,
        mes=mm,
        error=error,
    )


@comisiones_bp.route("/comisiones/debug")
@requiere_login
@requiere_permiso("comisiones.ver")
def debug():
    """TMT 2026-05-27 dueña: '/comisiones?mes=4 sigue sin funcionar'.
    Endpoint para diagnosticar paso a paso qué encuentra la query.
    Devuelve JSON con counts por stat / vend / orphan / cobros / fecha.
    """
    from flask import jsonify

    import db as _db

    yy, mm = _yy_mm()

    # 1) Cheques en el mes (fechad), distribuidos por stat
    by_stat = _db.fetch_all(
        """
        SELECT COALESCE(ch.stat, '(null)') AS stat,
               COUNT(*) AS n,
               COALESCE(SUM(ch.importe), 0) AS importe_total
          FROM scintela.cheque ch
         WHERE EXTRACT(YEAR FROM ch.fechad)  = %(yy)s
           AND EXTRACT(MONTH FROM ch.fechad) = %(mm)s
         GROUP BY COALESCE(ch.stat, '(null)')
         ORDER BY n DESC
        """,
        {"yy": yy, "mm": mm},
    ) or []

    # 2) Cheques depositados del mes, agrupados por vend del cliente
    by_vend = _db.fetch_all(
        """
        SELECT COALESCE(UPPER(TRIM(c.vend)), '(null)') AS vend,
               COUNT(*) AS n,
               COALESCE(SUM(ch.importe), 0) AS importe_total
          FROM scintela.cheque ch
          LEFT JOIN scintela.cliente c ON c.codigo_cli = ch.codigo_cli
         WHERE EXTRACT(YEAR FROM ch.fechad)  = %(yy)s
           AND EXTRACT(MONTH FROM ch.fechad) = %(mm)s
           AND ch.stat IN ('B','V','W','I','J','K','A')
         GROUP BY COALESCE(UPPER(TRIM(c.vend)), '(null)')
         ORDER BY n DESC
        """,
        {"yy": yy, "mm": mm},
    ) or []

    # 3) Cheques con codigo_cli huérfano
    huerfanos = _db.fetch_one(
        """
        SELECT COUNT(*) AS n,
               COALESCE(SUM(ch.importe), 0) AS importe_total
          FROM scintela.cheque ch
          LEFT JOIN scintela.cliente c ON c.codigo_cli = ch.codigo_cli
         WHERE EXTRACT(YEAR FROM ch.fechad)  = %(yy)s
           AND EXTRACT(MONTH FROM ch.fechad) = %(mm)s
           AND ch.stat IN ('B','V','W','I','J','K','A')
           AND c.codigo_cli IS NULL
        """,
        {"yy": yy, "mm": mm},
    ) or {"n": 0, "importe_total": 0}

    # 4) Cobros no-cheque del mes
    cobros_no_cheque = _db.fetch_all(
        """
        SELECT COALESCE(UPPER(co.tipo_doc), '(null)') AS tipo,
               COUNT(*) AS n,
               COALESCE(SUM(co.valor), 0) AS valor_total
          FROM scintela.cobro co
         WHERE EXTRACT(YEAR FROM co.fecha)  = %(yy)s
           AND EXTRACT(MONTH FROM co.fecha) = %(mm)s
         GROUP BY COALESCE(UPPER(co.tipo_doc), '(null)')
         ORDER BY n DESC
        """,
        {"yy": yy, "mm": mm},
    ) or []

    # 5) Comparativa: cuántos cheques caen en mes por fechad vs fecha vs fechaing
    by_fecha = _db.fetch_one(
        """
        SELECT
            COUNT(*) FILTER (WHERE EXTRACT(YEAR FROM ch.fechad)=%(yy)s AND EXTRACT(MONTH FROM ch.fechad)=%(mm)s) AS n_fechad,
            COUNT(*) FILTER (WHERE EXTRACT(YEAR FROM ch.fecha)=%(yy)s AND EXTRACT(MONTH FROM ch.fecha)=%(mm)s) AS n_fecha,
            COUNT(*) FILTER (WHERE EXTRACT(YEAR FROM ch.fechaing)=%(yy)s AND EXTRACT(MONTH FROM ch.fechaing)=%(mm)s) AS n_fechaing
          FROM scintela.cheque ch
         WHERE ch.stat IN ('B','V','W','I','J','K','A')
        """,
        {"yy": yy, "mm": mm},
    ) or {}

    return jsonify({
        "anio": yy, "mes": mm,
        "by_stat_fechad_en_mes": [dict(r) for r in by_stat],
        "by_vend_depositados_fechad_en_mes": [dict(r) for r in by_vend],
        "huerfanos_sin_cliente": dict(huerfanos),
        "cobros_no_cheque_fecha_en_mes": [dict(r) for r in cobros_no_cheque],
        "comparativa_fechas": dict(by_fecha),
    })


@comisiones_bp.route("/comisiones/<codigo>")
@requiere_login
@requiere_permiso("comisiones.ver")
def detalle(codigo: str):
    yy, mm = _yy_mm()
    v = queries.por_codigo(codigo)
    if not v:
        abort(404)
    try:
        cobranzas = queries.cobranzas_detalle(codigo, anio=yy, mes=mm)
        ventas = queries.ventas_detalle(codigo, anio=yy, mes=mm)
        error = None
    except Exception as e:
        cobranzas, ventas, error = [], [], str(e)

    total_cobr = sum(float(r.get("importe") or 0) for r in cobranzas)
    total_vent = sum(float(r.get("importe") or 0) for r in ventas)
    pct = float(v.get("pct_comision") or 0)
    comision = round(total_cobr * pct / 100.0, 2)

    return render_template(
        "comisiones/detalle.html",
        vendedor=v,
        cobranzas=cobranzas,
        ventas=ventas,
        total_cobr=total_cobr,
        total_vent=total_vent,
        pct=pct,
        comision=comision,
        anio=yy,
        mes=mm,
        error=error,
    )


@comisiones_bp.route("/comisiones/<codigo>/pct", methods=["POST"])
@requiere_login
@requiere_permiso("comisiones.ver")
def actualizar_pct(codigo: str):
    """Inline edit del % desde la lista."""
    v = queries.por_codigo(codigo)
    if not v:
        abort(404)
    pct = parse_monto(request.form.get("pct_comision"))
    if pct is None or pct < 0 or pct > 100:
        flash("% inválido (debe ser 0-100).", "error")
        return redirect(
            url_for("comisiones.lista", anio=request.form.get("anio"), mes=request.form.get("mes"))
        )
    try:
        usuario = (g.user or {}).get("username", "web")
        queries.actualizar_pct(codigo, pct, usuario=usuario)
        flash(f"% de {codigo} actualizado a {pct}%.", "ok")
    except Exception as e:
        flash_exc("No pude actualizar", e)
    return redirect(url_for("comisiones.lista", anio=request.form.get("anio"), mes=request.form.get("mes")))


@comisiones_bp.route("/comisiones/<codigo>/excel")
@requiere_login
@requiere_permiso("comisiones.ver")
def exportar_excel(codigo: str):
    """Descarga XLSX con cobranzas + ventas del vendedor en el mes.

    TMT 2026-05-19 v8 — pedido dueña: "dejame descargar por vendedor un
    excel asi lo vemos live con ellos". 2 hojas: Cobranzas + Ventas.
    """
    yy, mm = _yy_mm()
    v = queries.por_codigo(codigo)
    if not v:
        abort(404)

    try:
        cobranzas = queries.cobranzas_detalle(codigo, anio=yy, mes=mm)
        ventas = queries.ventas_detalle(codigo, anio=yy, mes=mm)
    except Exception as e:
        flash_exc("No pude armar el Excel", e)
        return redirect(url_for("comisiones.detalle", codigo=codigo, anio=yy, mes=mm))

    total_cobr = sum(float(r.get("importe") or 0) for r in cobranzas)
    total_vent = sum(float(r.get("importe") or 0) for r in ventas)
    pct = float(v.get("pct_comision") or 0)
    comision = round(total_cobr * pct / 100.0, 2)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        flash("openpyxl no instalado en el server — pedí al admin que lo instale.", "error")
        return redirect(url_for("comisiones.detalle", codigo=codigo, anio=yy, mes=mm))

    wb = Workbook()

    # Hoja 1 — Resumen / KPIs
    ws_res = wb.active
    ws_res.title = "Resumen"
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="E2E8F0")
    for cell in ("A1",):
        ws_res[cell].font = bold
    ws_res["A1"] = f"Comisión {codigo} — {mm:02d}/{yy}"
    ws_res["A1"].font = Font(bold=True, size=14)
    ws_res["A3"] = "Vendedor"
    ws_res["B3"] = codigo
    ws_res["A4"] = "% Comisión"
    ws_res["B4"] = pct
    ws_res["A5"] = "Total cobranzas USD"
    ws_res["B5"] = total_cobr
    ws_res["A6"] = "Total ventas USD"
    ws_res["B6"] = total_vent
    ws_res["A7"] = "Comisión a pagar USD"
    ws_res["B7"] = comision
    ws_res["A7"].font = bold
    ws_res["B7"].font = bold
    ws_res["B5"].number_format = "#,##0.00"
    ws_res["B6"].number_format = "#,##0.00"
    ws_res["B7"].number_format = "#,##0.00"
    ws_res.column_dimensions["A"].width = 25
    ws_res.column_dimensions["B"].width = 18

    # Hoja 2 — Cobranzas (separadas por origen)
    ws_cob = wb.create_sheet("Cobranzas")
    headers_cob = ["Tipo", "Fecha", "Cliente código", "Cliente nombre", "N°/Doc", "Banco", "Importe USD"]
    for i, h in enumerate(headers_cob, 1):
        c = ws_cob.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="left")
    cheques = [r for r in cobranzas if (r.get("origen") or "") == "CHE"]
    otros = [r for r in cobranzas if (r.get("origen") or "") != "CHE"]
    row_idx = 2
    for r in cheques + otros:
        ws_cob.cell(row=row_idx, column=1, value=r.get("origen") or "—")
        ws_cob.cell(row=row_idx, column=2, value=r.get("fecha"))
        ws_cob.cell(row=row_idx, column=3, value=r.get("codigo_cli") or "")
        ws_cob.cell(row=row_idx, column=4, value=r.get("cliente") or "")
        ws_cob.cell(row=row_idx, column=5, value=r.get("doc") or "")
        ws_cob.cell(row=row_idx, column=6, value=r.get("banco") or "")
        imp = ws_cob.cell(row=row_idx, column=7, value=float(r.get("importe") or 0))
        imp.number_format = "#,##0.00"
        row_idx += 1
    # Total
    if cobranzas:
        total_row = row_idx + 1
        c = ws_cob.cell(row=total_row, column=1, value="TOTAL")
        c.font = bold
        c.fill = hdr_fill
        t = ws_cob.cell(row=total_row, column=7, value=total_cobr)
        t.font = bold
        t.number_format = "#,##0.00"
        t.fill = hdr_fill
    for col, w in [("A", 8), ("B", 12), ("C", 14), ("D", 30), ("E", 14), ("F", 18), ("G", 16)]:
        ws_cob.column_dimensions[col].width = w

    # Hoja 3 — Ventas (facturas emitidas)
    ws_vta = wb.create_sheet("Ventas")
    headers_vta = ["Fecha", "Cliente código", "Cliente nombre", "N° factura", "Importe USD", "Saldo USD"]
    for i, h in enumerate(headers_vta, 1):
        c = ws_vta.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="left")
    for i, r in enumerate(ventas, 2):
        ws_vta.cell(row=i, column=1, value=r.get("fecha"))
        ws_vta.cell(row=i, column=2, value=r.get("codigo_cli") or "")
        ws_vta.cell(row=i, column=3, value=r.get("cliente") or "")
        ws_vta.cell(row=i, column=4, value=r.get("numf_completo") or r.get("numf") or "")
        imp = ws_vta.cell(row=i, column=5, value=float(r.get("importe") or 0))
        imp.number_format = "#,##0.00"
        sal = ws_vta.cell(
            row=i, column=6, value=float(r.get("saldo") or 0) if r.get("saldo") is not None else None
        )
        sal.number_format = "#,##0.00"
    if ventas:
        total_row = len(ventas) + 3
        c = ws_vta.cell(row=total_row, column=1, value="TOTAL")
        c.font = bold
        c.fill = hdr_fill
        t = ws_vta.cell(row=total_row, column=5, value=total_vent)
        t.font = bold
        t.number_format = "#,##0.00"
        t.fill = hdr_fill
    for col, w in [("A", 12), ("B", 14), ("C", 30), ("D", 16), ("E", 16), ("F", 16)]:
        ws_vta.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"comisiones_{codigo}_{yy}_{mm:02d}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@comisiones_bp.route("/comisiones/<codigo>/nombre", methods=["POST"])
@requiere_login
@requiere_permiso("comisiones.ver")
def actualizar_nombre(codigo: str):
    """Inline edit del nombre desde la lista."""
    v = queries.por_codigo(codigo)
    if not v:
        abort(404)
    nombre = (request.form.get("nombre") or "").strip()
    if not nombre:
        flash("Nombre vacío.", "error")
        return redirect(url_for("comisiones.lista"))
    try:
        usuario = (g.user or {}).get("username", "web")
        queries.actualizar_nombre(codigo, nombre, usuario=usuario)
        flash(f"Nombre de {codigo} actualizado.", "ok")
    except Exception as e:
        flash_exc("No pude actualizar", e)
    return redirect(url_for("comisiones.lista", anio=request.form.get("anio"), mes=request.form.get("mes")))
