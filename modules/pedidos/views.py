"""/pedidos — qué pidieron los clientes y todavía no se despachó.

Cinco cortes del mismo dato, en el orden del Excel "Detalle de Pedidos" de la
dueña: Color · Tipo de tela · Acabado · Cliente · Categoría. El corte por
acabado no está todavía (el acabado es atributo de LOTE en Asinfo y el pedido
no tiene lote — no hay fuente limpia a nivel producto).

  Color      → una fila por color (código grande, nombre chico), con pedido,
               stock, producción y faltante; se abre a sus telas y clientes.
  Tipo tela  → una pestaña por familia, un bloque por tela, una fila por color.
  Cliente    → total pedido por cliente (mezcla unidades, como el Excel).
  Categoría  → total pedido por familia de tela.

Todo se lee en la unidad en que se pide: rollos ENTEROS para las telas,
unidades para cuellos y puños.

Rutas:
  GET /pedidos                  (facturas.ver)
  GET /pedidos/color/<codigo>   (facturas.ver)

Va en el menú debajo de Factura Proforma (dueña 2026-08-17).
"""
from __future__ import annotations

import io

from flask import (
    Blueprint,
    Response,
    flash,
    redirect,
    render_template,
    request,
    url_for,
)

from auth import requiere_login, requiere_permiso
from modules._lib import formulas_memos

from . import service

pedidos_bp = Blueprint("pedidos", __name__, template_folder="templates")

#: Los cortes disponibles, EN EL ORDEN del Excel de la dueña. El primero es el
#: default: la pantalla abre en Color.
CORTES = ("color", "tela", "cliente", "pedido", "categoria")
CORTES_LBL = {
    "color": "Color",
    "tela": "Tipo de tela",
    "cliente": "Cliente",
    "pedido": "Pedido",
    "categoria": "Categoría",
}


@pedidos_bp.route("/pedidos")
@requiere_login
@requiere_permiso("facturas.ver")
def lista():
    filas, disponible = service.pendientes()
    categorias = service.por_categoria(filas)

    corte = (request.args.get("corte") or "").strip()
    if corte not in CORTES:
        corte = CORTES[0]

    ctx = dict(
        disponible=disponible,
        categorias=categorias,
        corte=corte,
        cortes=CORTES,
        cortes_lbl=CORTES_LBL,
        dias_pedido_max=service.DIAS_PEDIDO_MAX,
        dias_produccion_viva=service.DIAS_PRODUCCION_VIVA,
        kg_por_rollo=service.KG_POR_ROLLO,
        # se completan según el corte
        colores=[], telas=[], activa="", resumen=None,
        clientes=[], solo_faltan=False, cubiertas=[],
        pedidos_por_color={}, pedidos_lista=[], etapas={},
    )

    if not disponible or not categorias:
        return render_template("pedidos/lista.html", **ctx)

    if corte in ("color", "tela"):
        service.marcar_acabado(filas)   # el punto TUB/ABI; fail-soft

    if corte == "color":
        ctx["colores"] = service.por_color(filas)
        ctx["pedidos_por_color"] = service.pedidos_por_color()
    elif corte == "cliente":
        ctx["clientes"] = service.por_cliente()
    elif corte == "pedido":
        pedidos_lista, disp_ped = service.por_pedido()
        ctx["pedidos_lista"] = pedidos_lista
        ctx["disponible"] = disponible and disp_ped
        # Qué pedidos ya tienen memo en la fábrica (fail-soft: {} si el
        # bridge no está — los botones se muestran y el UNIQUE de formulas
        # frena un doble envío igual).
        numeros = [p["numero"] for p in pedidos_lista]
        memo_estados = formulas_memos.estados(numeros)
        # Un memo cancelado no cuenta como "tiene memo" para la pantalla: el
        # pedido tiene que volver a mostrar "Enviar memo" como si nunca se
        # hubiera mandado.
        activos = {n: v for n, v in memo_estados.items() if v.get("estado") != "cancelado"}
        ctx["memo_estados"] = activos
        # La ETAPA de cada pedido (enviado → en tintura → terminado), armada
        # con las órdenes de formulas y la OFT en Asinfo. Sin porcentajes
        # (dueña 27/08).
        ctx["etapas"] = service.etapas_por_pedido(pedidos_lista, activos)
    elif corte == "tela":
        # La pestaña por defecto es la que MÁS falta, no la primera alfabética.
        pedida = (request.args.get("cat") or "").strip()
        nombres = [c["categoria"] for c in categorias]
        activa = pedida if pedida in nombres else (nombres[0] if nombres else "")
        solo_faltan = request.args.get("falta") == "1"
        telas = service.por_tela(filas, activa, solo_faltan=solo_faltan) if activa else []
        for t in telas:
            service.en_unidad(t["filas"], "alt")
            t["faltan_d"], t["dec"], t["u"] = service.total_en_unidad(
                t["faltan_kg"], "alt",
                any(f["en_unidades"] for f in t["filas"]),
                next((f["un_por_kg"] for f in t["filas"] if f["un_por_kg"]), 0.0))
        ctx.update(
            activa=activa, telas=telas, solo_faltan=solo_faltan,
            cubiertas=service.telas_sin_faltante(filas, activa) if solo_faltan else [],
            resumen=next((c for c in categorias if c["categoria"] == activa), None),
            pedidos_por_color=service.pedidos_por_color(),
        )
    # corte == "categoria" usa `categorias`, ya calculado.

    return render_template("pedidos/lista.html", **ctx)


@pedidos_bp.route("/pedidos/excel")
@requiere_login
@requiere_permiso("facturas.ver")
def excel():
    """Todo lo que se ve en pantalla, en un xlsx con una hoja por corte.

    Color (a nivel color + tela), Tipo de tela, Cliente y Categoría — el mismo
    dato de las cuatro pestañas, para ordenar y filtrar en Excel.
    """
    filas, disponible = service.pendientes()
    if not disponible:
        flash("No pude consultar Asinfo, así que no armé el Excel. "
              "Probá de nuevo en un minuto.", "error")
        return redirect(url_for("pedidos.lista"))
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        flash("openpyxl no instalado en el server — pedí al admin que lo instale.", "error")
        return redirect(url_for("pedidos.lista"))

    from filters import today_ec

    service.marcar_acabado(filas)
    colores = service.por_color(filas)          # variants ya traen pedido_d, etc.
    categorias = service.por_categoria(filas)
    clientes = service.por_cliente()

    wb = Workbook()
    bold = Font(bold=True)
    relleno = PatternFill("solid", fgColor="E2E8F0")

    def encabezado(ws, cols):
        for i, h in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = bold
            c.fill = relleno
            c.alignment = Alignment(horizontal="left")

    def cerrar(ws, anchos, nfilas):
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho
        ws.freeze_panes = "A2"
        ncol = ws.cell(row=1, column=len(anchos)).column_letter
        ws.auto_filter.ref = f"A1:{ncol}{max(nfilas + 1, 2)}"

    def falta(v):
        return v if v > 0 else 0

    # ── Hoja Color (color + tela) ──
    ws = wb.active
    ws.title = "Color"
    encabezado(ws, ["Código", "Color", "Tela", "Acabado", "Unidad",
                    "Pedido", "Stock", "En produc.", "Falta"])
    r = 1
    for g in colores:
        nombre = f"{g['codigo']} {g['nombre']}".strip()
        for f in g["variants"]:
            r += 1
            for i, v in enumerate([f["codigo"], nombre, f["tela"], f.get("acabado", ""),
                                   f["u"], f["pedido_d"], f["inventario_d"],
                                   f["produccion_d"], falta(f["faltan_d"])], 1):
                ws.cell(row=r, column=i, value=v)
    cerrar(ws, (12, 24, 22, 8, 8, 10, 10, 12, 10), r)

    # ── Hoja Tipo de tela (familia · tela · color) ──
    ws = wb.create_sheet("Tipo de tela")
    encabezado(ws, ["Familia", "Tela", "Color", "Código", "Acabado", "Unidad",
                    "Pedido", "Stock", "En produc.", "Falta"])
    r = 1
    for f in sorted(filas, key=lambda x: (x["categoria"], x["tela"], x["color"])):
        r += 1
        for i, v in enumerate([f["categoria"], f["tela"], f["color"], f["codigo"],
                               f.get("acabado", ""), f["u"], f["pedido_d"],
                               f["inventario_d"], f["produccion_d"], falta(f["faltan_d"])], 1):
            ws.cell(row=r, column=i, value=v)
    cerrar(ws, (12, 22, 10, 12, 8, 8, 10, 10, 12, 10), r)

    # ── Hoja Cliente ──
    ws = wb.create_sheet("Cliente")
    encabezado(ws, ["Cliente", "Pedido"])
    for r, cl in enumerate(clientes, 2):
        ws.cell(row=r, column=1, value=cl["cliente"])
        ws.cell(row=r, column=2, value=cl["pedido"])
    cerrar(ws, (14, 12), len(clientes))

    # ── Hoja Categoría ──
    ws = wb.create_sheet("Categoría")
    encabezado(ws, ["Familia", "Pedido", "Unidad"])
    cats = sorted(categorias, key=lambda c: -c["pedido_d"])
    for r, c in enumerate(cats, 2):
        ws.cell(row=r, column=1, value=c["categoria"])
        ws.cell(row=r, column=2, value=c["pedido_d"])
        ws.cell(row=r, column=3, value=c["u"])
    cerrar(ws, (16, 12, 8), len(cats))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"pedidos_pendientes_{today_ec():%Y_%m_%d}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@pedidos_bp.route("/pedidos/color/<codigo>")
@requiere_login
@requiere_permiso("facturas.ver")
def color(codigo: str):
    """Detalle de un color: quién lo pidió y qué se está tinturando.

    No devuelve 404 cuando el código no existe: si Asinfo no contesta no
    podemos distinguir "no existe" de "no pude preguntar", y un 404 ahí
    mentiría. Muestra la ficha vacía con el aviso.
    """
    ficha, pedidos, ordenes, disponible = service.detalle_color(codigo)
    return render_template(
        "pedidos/color.html",
        codigo=codigo.strip().upper(),
        ficha=ficha,
        pedidos=pedidos,
        ordenes=ordenes,
        repetidos=service.repetidos(pedidos),
        disponible=disponible,
        kg_por_rollo=service.KG_POR_ROLLO,
    )


@pedidos_bp.route("/pedidos/enviar-memo", methods=["POST"])
@requiere_login
@requiere_permiso("pedidos.enviar_memo")
def enviar_memo():
    """Manda UN pedido como memo al tab Memos de formulas_app.

    El memo es la foto del pedido al momento de enviar. Si ya estaba enviado
    (por quien sea), avisa y no duplica — lo garantiza el UNIQUE de
    `pedido_numero` del lado de formulas_app, no esta vista.
    """
    from flask import g

    numero = (request.form.get("numero") or "").strip()
    detalle = service.armar_memo(numero) if numero else None
    if detalle is None:
        flash("No encontré ese pedido entre los pendientes, así que no mandé "
              "el memo. Recargá la pantalla y probá de nuevo.", "error")
        return redirect(url_for("pedidos.lista", corte="pedido"))

    usuario = g.user["username"] if getattr(g, "user", None) else ""
    ok, motivo = formulas_memos.enviar(
        numero=numero,
        cliente=detalle["cliente"]["nombre"],
        vendedor=service.etiqueta_dueno(detalle["vendedor"]),
        enviado_por=usuario,
        detalle=detalle,
    )
    if ok:
        flash(f"Memo del pedido {numero} enviado a la fábrica.", "success")
    elif motivo == "ya_enviado":
        flash(f"El pedido {numero} ya tenía memo enviado — no se mandó otro.",
              "warning")
    else:
        flash("No pude hablar con formulas_app — el memo NO se envió. "
              "Probá de nuevo en un rato.", "error")
    return redirect(url_for("pedidos.lista", corte="pedido"))


@pedidos_bp.route("/pedidos/cancelar-memo", methods=["POST"])
@requiere_login
@requiere_permiso("pedidos.enviar_memo")
def cancelar_memo():
    """Cancela un memo PENDIENTE para poder mandar uno nuevo (el pedido
    cambió). Si la fábrica ya lo tomó, no hace nada — avisa y no rompe."""
    from flask import g

    numero = (request.form.get("numero") or "").strip()
    if not numero:
        flash("Falta el número de pedido.", "error")
        return redirect(url_for("pedidos.lista", corte="pedido"))

    usuario = g.user["username"] if getattr(g, "user", None) else ""
    ok, motivo = formulas_memos.cancelar(numero, usuario)
    if ok:
        flash(f"Memo del pedido {numero} cancelado. Ya lo podés mandar de nuevo.",
              "success")
    elif motivo == "no_pendiente":
        flash(f"El memo del pedido {numero} ya no está pendiente — no se "
              "pudo cancelar.", "warning")
    else:
        flash("No pude hablar con formulas_app — el memo NO se canceló. "
              "Probá de nuevo en un rato.", "error")
    return redirect(url_for("pedidos.lista", corte="pedido"))
