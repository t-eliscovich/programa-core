"""Bancos — lista, movimientos y emisión de cheques propios (chequera)."""

from flask import Blueprint, abort, flash, g, redirect, render_template, request, url_for

from auth import requiere_login, requiere_permiso
from error_messages import flash_exc
from exports import csv_response
from filters import today_ec
from parsers import parse_date, parse_int, parse_monto

from . import queries

bancos_bp = Blueprint("bancos", __name__, template_folder="templates")


@bancos_bp.route("/bancos")
@requiere_login
@requiere_permiso("bancos.ver")
def lista():
    """Vista hub de Bancos — TMT 2026-05-19 v2 (pedido dueña).

    Default: muestra Pichincha con su saldo y los 4 botones de acción
    (Emitir cheque, Depositar, NC, ND) pre-llenados con `?no_banco=`.
    Botón "Cambiar banco" permite cambiar a Internacional u otro.
    El "banco actual" se elige con ?banco=<no_banco> o `?banco=PICHINCHA`/
    `?banco=INTERNACIONAL` (resuelto por NOMBRE — el no_banco hardcoded
    no nos sirve porque varía por instalación).
    """
    mostrar_todos = request.args.get("todos") == "1"
    try:
        filas_all = queries.lista_bancos()
        error = None
    except Exception as e:
        filas_all, error = [], str(e)

    # TMT 2026-06-25 (dueña): la lista de CUENTAS no debe incluir los
    # bancos-concepto/espejo (DEP.PICH 90, DEP.INTER 91, CANCELA ANTICIPO 95,
    # ANTICIPO 97, UKN 98, EFECTIVO 99) — son rubros contables internos de la
    # cobranza, no cuentas bancarias. Los reales son no_banco < 90. Antes
    # DEP.PICH se colaba al resumen con un residuo (ej. -455,89). El filtro
    # _es_operativo ya los excluía, pero solo se usaba para "Cambiar banco".
    def _es_cuenta_real(b):
        try:
            return int(b.get("no_banco") or 0) < 90
        except (TypeError, ValueError):
            return True

    filas_reales = [b for b in filas_all if _es_cuenta_real(b)]
    if mostrar_todos:
        filas = filas_reales
    else:
        filas = [
            b
            for b in filas_reales
            if round(float(b["saldo_stored"] or 0), 2) != 0.0
            or round(float(b["saldo_derivado"] or 0), 2) != 0.0
        ]
    ocultos = len(filas_reales) - len(filas)

    # --- Resolver el banco "actual" del hub ----------------------------
    # Prioridad: ?banco=<no_banco_int o nombre> → fallback PICHINCHA.
    banco_param = (request.args.get("banco") or "").strip().upper()

    def _busca_por_nombre(needle):
        for b in filas_all:
            if needle in (b.get("nombre") or "").upper():
                return b
        return None

    banco_actual = None
    if banco_param.isdigit():
        no = int(banco_param)
        banco_actual = next((b for b in filas_all if int(b["no_banco"]) == no), None)
    elif banco_param in ("INTERNACIONAL", "INTER", "INTERNAC"):
        banco_actual = _busca_por_nombre("INTER")
    elif banco_param == "PICHINCHA" or banco_param == "PICH":
        banco_actual = _busca_por_nombre("PICHINC")
    if banco_actual is None:
        # Default: Pichincha por nombre.
        banco_actual = _busca_por_nombre("PICHINC")
    if banco_actual is None and filas_all:
        # Fallback: primer banco con saldo > 0, o el primero.
        banco_actual = filas[0] if filas else filas_all[0]

    # Lista de otros bancos para el selector "Cambiar banco" (excluye el actual).
    # TMT 2026-05-19 v7 — dueña: "solo podemos cambiar al internacional".
    # TMT 2026-05-19 v8 — dueña: "BORRAR DEP INTER DE BANCOS". El nombre
    # "INTER" matcheaba tanto INTERNACI (operativo) como DEP. INTER.
    # (banco depósito-en-tránsito, no operativo). Cambio a "INTERNAC" para
    # excluir el de DEP.
    def _es_operativo(b):
        n = (b.get("nombre") or "").upper().strip()
        # Excluir explícitamente cualquier "DEP. ..." (depósitos en tránsito).
        if n.startswith("DEP"):
            return False
        return "PICHINC" in n or "INTERNAC" in n

    otros_bancos = [
        b
        for b in filas_all
        if _es_operativo(b) and (not banco_actual or int(b["no_banco"]) != int(banco_actual["no_banco"]))
    ]

    return render_template(
        "bancos/lista.html",
        filas=filas,
        error=error,
        mostrar_todos=mostrar_todos,
        ocultos=ocultos,
        banco_actual=banco_actual,
        otros_bancos=otros_bancos,
    )


@bancos_bp.route("/bancos/_api/preview-concepto")
@requiere_login
@requiere_permiso("bancos.ver")
def preview_concepto():
    """JSON: previsualiza el tipo de cheque desde el concepto.

    Para el form emitir_cheque (chequera): mientras la usuaria tipea el
    concepto, el JS llama acá para sugerir el radio adecuado (proveedor /
    retiro / caja / etc). Auto-detect estilo dBase legacy.

    Mapeo:
      - PR <prov_valido>  → proveedor (con beneficiario sugerido)
      - RR <socio>        → retiro (con de_socio sugerido)
      - INHB / caja / efectivo → caja (banco→caja física)
      - PICH / INTER (transfer entre bancos) → caja (genérico) o proveedor
      - cualquier otro    → otro (sólo movimiento banco)
    """
    import concepto_parser
    import db as _db

    concepto = (request.args.get("concepto") or "").strip()
    if not concepto:
        return {"tipo_sugerido": None, "descripcion": "", "extras": {}}

    provs_validos = {
        (r.get("codigo_prov") or "").strip().upper()
        for r in (_db.fetch_all("SELECT codigo_prov FROM scintela.proveedor") or [])
    }
    bancos_map: dict = {}
    for b in _db.fetch_all("SELECT no_banco, COALESCE(nombre, '') AS nombre FROM scintela.banco") or []:
        n = (b.get("nombre") or "").upper().strip()
        if "PICHINC" in n:
            bancos_map.setdefault("PICHINCHA", int(b["no_banco"]))
        if "INTER" in n:
            bancos_map.setdefault("INTERNACIONAL", int(b["no_banco"]))

    parsed = concepto_parser.parse_concepto(
        concepto,
        {"provs_validos": provs_validos, "bancos": bancos_map},
    )
    parsed_tipo = parsed.get("tipo")
    # Mapeo parser → form radio.
    if parsed_tipo == "compra_proveedor":
        tipo_sugerido = "proveedor"
        extras = {"beneficiario": parsed.get("prov")}
    elif parsed_tipo == "retiro_socio":
        tipo_sugerido = "retiro"
        extras = {"de_socio": parsed.get("socio")}
    elif parsed_tipo == "caja_inhb":
        tipo_sugerido = "caja"
        extras = {}
    elif parsed_tipo == "transfer_banco":
        # Transferencia entre bancos propios — no aplica al form de cheques.
        # El form de cheques es para egresos a terceros.
        tipo_sugerido = "otro"
        extras = {}
    elif parsed_tipo == "dolares":
        # TMT 2026-05-17: antes mapeaba a "otro" (no creaba el anticipo).
        # Ahora va a la card dedicada `anticipo_usd` y pasa la cuenta de
        # 2 letras como beneficiario para el INSERT en scintela.dolares.
        tipo_sugerido = "anticipo_usd"
        extras = {"beneficiario": parsed.get("cuenta")}
    else:
        tipo_sugerido = "gasto"  # default razonable: gasto pagado con cheque
        extras = {}

    return {
        "tipo_parser": parsed_tipo,
        "tipo_sugerido": tipo_sugerido,
        "descripcion": concepto_parser.descripcion_humana(parsed),
        "extras": extras,
    }


@bancos_bp.route("/bancos/transferir", methods=["GET", "POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def transferir():
    """Mueve plata de un banco al otro. Atómico: insert CH origen + DE destino.

    Flujo en 2 pasos (TMT 2026-05-14 #19 audit):
      1. GET form → user llena.
      2. POST sin `confirmado=1` → muestra wizard con saldo_preview.
      3. POST con `confirmado=1` → ejecuta la transferencia.
    """
    # Sólo bancos operativos (Pichincha + Internacional) — la tabla scintela.banco
    # tiene 30 filas pero la mayoría son rubros contables del PRG legacy. TMT 2026-05-12.
    bancos = queries.bancos_operativos() or []
    if request.method == "POST":
        no_origen = parse_int(request.form.get("no_banco_origen"))
        no_destino = parse_int(request.form.get("no_banco_destino"))
        importe = parse_monto(request.form.get("importe"))
        fecha = parse_date(request.form.get("fecha")) or today_ec()
        concepto = (request.form.get("concepto") or "").strip()
        confirmado = request.form.get("confirmado") in ("1", "true", "yes")

        # Validaciones básicas comunes (al paso 2 y al paso 3).
        errores: list[str] = []
        if not no_origen:
            errores.append("Elegí el banco de origen.")
        if not no_destino:
            errores.append("Elegí el banco de destino.")
        if no_origen and no_destino and no_origen == no_destino:
            errores.append("Origen y destino no pueden ser el mismo banco.")
        if importe is None or float(importe or 0) <= 0:
            errores.append("Importe debe ser mayor a cero.")
        if errores:
            for e in errores:
                flash(e, "warn")
            return render_template(
                "bancos/transferir.html",
                bancos=bancos,
                hoy=today_ec().isoformat(),
                form={
                    "no_banco_origen": no_origen,
                    "no_banco_destino": no_destino,
                    "importe": request.form.get("importe"),
                    "fecha": request.form.get("fecha"),
                    "concepto": concepto,
                },
            )

        # Paso 2: si NO confirmó, mostrar wizard con saldo_preview.
        if not confirmado:
            try:
                # Buscar nombres + saldos actuales de los dos bancos.
                bo = next((b for b in bancos if int(b.get("no_banco")) == int(no_origen)), None)
                bd = next((b for b in bancos if int(b.get("no_banco")) == int(no_destino)), None)
                imp_f = float(importe or 0)
                saldo_preview = [
                    {
                        "label": f"{(bo or {}).get('nombre') or no_origen} (origen)",
                        "antes": float((bo or {}).get("saldo") or 0),
                        "despues": float((bo or {}).get("saldo") or 0) - imp_f,
                    },
                    {
                        "label": f"{(bd or {}).get('nombre') or no_destino} (destino)",
                        "antes": float((bd or {}).get("saldo") or 0),
                        "despues": float((bd or {}).get("saldo") or 0) + imp_f,
                    },
                ]
            except Exception:
                saldo_preview = None
            # Bug C fix (TMT 2026-05-16): el template _confirmar_accion.html
            # espera `extras_hidden` (list of {name,value}) y `accion_url`,
            # no `hidden_fields` + `form_action`. Antes los hidden inputs no
            # se renderizaban → el POST de confirm llegaba sin datos y volvía
            # al form vacío con errores "Elegí banco origen", etc.
            return render_template(
                "_confirmar_accion.html",
                titulo="Confirmar transferencia entre bancos",
                resumen=(
                    f"Vas a transferir <strong>$ {float(importe or 0):,.2f}</strong> "
                    f"de {(bo or {}).get('nombre') or 'origen'} a "
                    f"{(bd or {}).get('nombre') or 'destino'} "
                    f"el {fecha.strftime('%d/%m/%Y')}."
                ),
                concepto=concepto,
                saldo_preview=saldo_preview,
                accion_url=url_for("bancos.transferir"),
                cancel_url=url_for("bancos.transferir"),
                extras_hidden=[
                    {"name": "no_banco_origen", "value": no_origen},
                    {"name": "no_banco_destino", "value": no_destino},
                    {"name": "importe", "value": request.form.get("importe") or ""},
                    {"name": "fecha", "value": fecha.isoformat()},
                    {"name": "concepto", "value": concepto},
                    {"name": "confirmado", "value": "1"},
                ],
                motivo_obligatorio=False,
                motivo_requerido=False,
                confirm_label="Sí, transferir",
            )

        # Paso 3: confirmado → ejecutar.
        try:
            usuario = (g.user or {}).get("username", "web")
            r = queries.transferir_entre_bancos(
                no_banco_origen=no_origen,
                no_banco_destino=no_destino,
                importe=importe,
                fecha=fecha,
                concepto=concepto,
                usuario=usuario,
            )
            flash(
                f"Transferencia OK: {r['origen']['nombre']} → {r['destino']['nombre']} "
                f"por $ {r['importe']:.2f}. Saldos: origen ${r['origen']['saldo_nuevo']:.2f} · "
                f"destino ${r['destino']['saldo_nuevo']:.2f}.",
                "ok",
            )
            return redirect(url_for("bancos.lista"))
        except ValueError as e:
            flash(str(e), "warn")
        except Exception as e:  # noqa: BLE001
            flash_exc("No pude completar la transferencia", e)
    return render_template(
        "bancos/transferir.html",
        bancos=bancos,
        hoy=today_ec().isoformat(),
    )


@bancos_bp.route("/bancos/emitir-cheque", methods=["GET", "POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def emitir_cheque():
    """Wizard para emitir un cheque propio (chequera).

    El legacy `BANCOS.PRG::CHEQUERA` infería el tipo de movimiento del
    proveedor + concepto. Este wizard pide el tipo EXPLÍCITO (proveedor /
    retiro / caja / gasto / otro) y aplica el side-effect correspondiente
    en una sola transacción.
    """
    bancos = queries.bancos_operativos() or []

    if request.method == "POST":
        tipo = (request.form.get("tipo") or "").strip().lower()
        no_banco = parse_int(request.form.get("no_banco"))
        importe = parse_monto(request.form.get("importe"))
        fecha = parse_date(request.form.get("fecha")) or today_ec()
        no_cheque = (request.form.get("no_cheque") or "").strip()
        beneficiario = (request.form.get("beneficiario") or "").strip().upper()
        concepto = (request.form.get("concepto") or "").strip()
        id_posdat = parse_int(request.form.get("id_posdat"))
        # TMT 2026-05-27 dueña: 'cuando emito cheques, puedes dejarme
        # seleccionar multiples proveedores'. Multi-select via getlist;
        # parsea ints, ignora vacíos / inválidos. Si vienen los dos
        # (single legacy + multi), gana multi en queries.emitir_cheque.
        id_posdats_raw = request.form.getlist("id_posdats[]")
        id_posdats: list[int] = []
        for _x in id_posdats_raw:
            _v = parse_int(_x)
            if _v:
                id_posdats.append(_v)
        de_socio = (request.form.get("de_socio") or "").strip().upper()
        es_postdatado = bool(request.form.get("es_postdatado"))
        fechad = parse_date(request.form.get("fechad"))
        # TMT 2026-05-19 v4 audit — xgast_num para clasificar V1..V9 cuando
        # tipo='gasto'. Sin esto el xgast quedaba con num=NULL → invisible
        # en /informes/gastos.
        xgast_num_raw = (request.form.get("xgast_num") or "").strip()
        xgast_num_val: int | None = None
        if xgast_num_raw:
            try:
                v = int(xgast_num_raw)
                if 1 <= v <= 9:
                    xgast_num_val = v
            except (TypeError, ValueError):
                xgast_num_val = None

        try:
            usuario = (g.user or {}).get("username", "web")
            r = queries.emitir_cheque(
                tipo=tipo,
                no_banco=no_banco,
                importe=importe,
                fecha=fecha,
                no_cheque=no_cheque,
                beneficiario=beneficiario,
                concepto=concepto,
                id_posdat=id_posdat,
                id_posdats=id_posdats or None,
                de_socio=de_socio,
                es_postdatado=es_postdatado,
                fechad=fechad,
                usuario=usuario,
                xgast_num=xgast_num_val,
            )
            flash(
                f"Cheque emitido OK desde {r['banco_nombre']} por $ {r['importe']:.2f}. "
                f"Side-effect: {r['side_effect']}.",
                "ok",
            )
            return redirect(url_for("bancos.movimientos", no_banco=r["no_banco"]))
        except ValueError as e:
            flash(str(e), "warn")
        except Exception as e:
            flash_exc("No pude emitir el cheque", e)

    # GET o POST con error
    import contextlib

    posdats = []
    prov_filter = (request.args.get("prov") or "").strip().upper() or None
    with contextlib.suppress(Exception):
        posdats = queries.posdat_abiertas_de(prov_filter)

    # Autocomplete: conceptos históricos + proveedores activos.
    import contextlib as _ctx

    conceptos = []
    proveedores = []
    with _ctx.suppress(Exception):
        conceptos = queries.conceptos_frecuentes_egresos(limite=50)
    with _ctx.suppress(Exception):
        proveedores = queries.proveedores_activos(limite=500)

    # TMT 2026-05-17: si la URL trae ?tipo=anticipo_usd (acceso directo
    # desde sidebar), el template pre-selecciona esa card. Aceptamos los
    # 6 tipos válidos; cualquier otro valor se ignora.
    tipo_inicial = (request.args.get("tipo") or "").strip().lower()
    if tipo_inicial not in ("proveedor", "retiro", "caja", "gasto", "anticipo_usd", "otro"):
        tipo_inicial = ""

    # TMT 2026-05-19 — item 19 (pedido dueña): si vienen con ?id_posdat=N
    # desde el botón "Pagar" de /posdat, levantamos esa posdat y la pasamos
    # al template para mostrar un banner de contexto + pre-seleccionar la
    # fila + pre-llenar importe/concepto. Antes el wizard llegaba "vacío"
    # y la dueña tenía que recordar todos los datos.
    import db as _db

    posdat_target = None
    id_posdat_param = (request.args.get("id_posdat") or "").strip()
    if id_posdat_param:
        try:
            posdat_target = _db.fetch_one(
                """
                SELECT pd.id_posdat, pd.num, pd.prov, pd.fechad,
                       pd.importe, pd.concepto, pd.banc,
                       COALESCE(p.nombre, '') AS proveedor_nombre
                  FROM scintela.posdat pd
                  LEFT JOIN scintela.proveedor p ON p.codigo_prov = pd.prov
                 WHERE pd.id_posdat = %s
                   AND (pd.anulada IS NOT TRUE OR pd.anulada IS NULL)
                """,
                (int(id_posdat_param),),
            )
            if posdat_target:
                # Default tipo = "proveedor" cuando viene desde el botón Pagar.
                if not tipo_inicial:
                    tipo_inicial = "proveedor"
                # Si tenemos prov, filtramos las posdats al mismo proveedor
                # para que la lista no abrume.
                if posdat_target.get("prov") and not prov_filter:
                    prov_filter = (posdat_target["prov"] or "").strip().upper() or None
                    with contextlib.suppress(Exception):
                        posdats = queries.posdat_abiertas_de(prov_filter)
        except (ValueError, TypeError):
            posdat_target = None

    # TMT 2026-05-19 v2 — banco pre-seleccionado vía ?no_banco= cuando se
    # entra desde la action bar de /bancos. La dueña trabaja siempre desde
    # Pichincha por defecto, esto evita el "elegir banco" extra.
    # TMT 2026-05-19 v7 — si no viene ?no_banco=, default = Pichincha.
    # Pedido literal dueña: "siempre en default deja banco pichincha cuando
    # estemos completando algo que haya que elegir cheque (emitir/pagar)".
    no_banco_inicial = parse_int(request.args.get("no_banco"))
    if not no_banco_inicial:
        for b in bancos or []:
            if "PICHINC" in (b.get("nombre") or "").upper():
                no_banco_inicial = int(b["no_banco"])
                break

    # TMT 2026-05-19 v7 — pedido dueña: "cuando aprieto en una fila pagar,
    # tiene que venir todo pre cargado. (...) numero de cheque agarrar el
    # que viene despues del ultimo que tenemos cargado". Calculamos
    # MAX(numreferencia)+1 del banco pre-seleccionado.
    no_cheque_sugerido = None
    if no_banco_inicial:
        with contextlib.suppress(Exception):
            row = _db.fetch_one(
                """
                SELECT MAX(CASE WHEN COALESCE(NULLIF(TRIM(numreferencia_manual),''),
                                             numreferencia::text) ~ '^[0-9]+$'
                                THEN COALESCE(NULLIF(TRIM(numreferencia_manual),''),
                                              numreferencia::text)::bigint END) AS ultimo
                  FROM scintela.transacciones_bancarias
                 WHERE no_banco = %s
                   AND documento = 'CH'
                """,
                (int(no_banco_inicial),),
            )
            if row and row.get("ultimo"):
                no_cheque_sugerido = int(row["ultimo"]) + 1

    return render_template(
        "bancos/emitir_cheque.html",
        bancos=bancos,
        posdats=posdats,
        prov_filter=prov_filter,
        tipo_inicial=tipo_inicial,
        no_banco_inicial=no_banco_inicial,
        no_cheque_sugerido=no_cheque_sugerido,
        hoy=today_ec().isoformat(),
        conceptos=conceptos,
        proveedores=proveedores,
        posdat_target=posdat_target,
    )


@bancos_bp.route("/bancos/cheque-emitido/<int:id_transaccion>/reversar", methods=["GET", "POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def reversar_cheque_emitido(id_transaccion: int):
    """Wizard de 2 pasos para reversar un cheque emitido.

    GET: muestra detalles + input de motivo.
    POST: ejecuta queries.reversar_cheque_emitido (atómico, registra mov_doble).
    """
    import db as _db

    tx = _db.fetch_one(
        """
        SELECT t.id_transaccion, t.fecha, t.no_banco, t.documento,
               t.importe, t.concepto, t.prov, t.numreferencia,
               COALESCE(b.nombre, '') AS banco_nombre,
               COALESCE(p.nombre, '') AS prov_nombre
          FROM scintela.transacciones_bancarias t
          LEFT JOIN scintela.banco b ON b.no_banco = t.no_banco
          LEFT JOIN scintela.proveedor p ON p.codigo_prov = t.prov
         WHERE t.id_transaccion = %s
        """,
        (id_transaccion,),
    )
    if not tx:
        abort(404)
    if (tx.get("documento") or "").strip().upper() != "CH":
        flash(
            f"La transacción #{id_transaccion} no es un cheque emitido (documento={tx.get('documento')!r}).",
            "warn",
        )
        return redirect(url_for("bancos.movimientos", no_banco=tx["no_banco"]))

    if request.method == "POST":
        motivo = (request.form.get("motivo") or "").strip()
        if False:  # TMT 2026-05-21 dueña: motivo opcional sin minlen
            flash("Motivo requerido (mín. 5 caracteres).", "warn")
            return render_template(
                "bancos/reversar_cheque_emitido.html",
                tx=tx,
                motivo=motivo,
            ), 400
        try:
            usuario = (g.user or {}).get("username", "web")
            r = queries.reversar_cheque_emitido(
                id_transaccion=id_transaccion,
                motivo=motivo,
                usuario=usuario,
            )
            msg = (
                f"Cheque emitido reversado. Compensación bancaria "
                f"#{r['id_transaccion_compensacion']} (ND $ {r['importe']:.2f})."
            )
            if r.get("side_effect_revertido"):
                se = r["side_effect_revertido"]
                msg += f" Side effect revertido: {se.get('tipo')}."
            flash(msg, "ok")
            return redirect(url_for("bancos.movimientos", no_banco=tx["no_banco"]))
        except ValueError as e:
            flash(str(e), "warn")
            return render_template(
                "bancos/reversar_cheque_emitido.html",
                tx=tx,
                motivo=motivo,
            ), 400
        except Exception as e:
            flash_exc("No pude reversar el cheque emitido", e)
            return redirect(url_for("bancos.movimientos", no_banco=tx["no_banco"]))

    return render_template(
        "bancos/reversar_cheque_emitido.html",
        tx=tx,
        motivo="",
    )


@bancos_bp.route("/bancos/transferencia/<int:id_mov_doble>/reversar", methods=["GET", "POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def reversar_transferencia(id_mov_doble: int):
    """Wizard de 2 pasos para reversar una transferencia banco↔banco.

    GET: muestra detalles + input de motivo (opcional).
    POST: ejecuta queries.reversar_transferencia. NC en origen + CH en
    destino, atómico, registra mov_doble del reverso linkeado.
    TMT 2026-05-13.
    """
    import db as _db

    md = _db.fetch_one(
        """
        SELECT m.id_mov_doble, m.tipo, m.fecha_operacion, m.importe,
               m.concepto, m.estado,
               m.origen_id, m.destino_id, m.metadata,
               tx_o.no_banco AS no_banco_origen,
               tx_d.no_banco AS no_banco_destino,
               COALESCE(b_o.nombre, '') AS banco_origen,
               COALESCE(b_d.nombre, '') AS banco_destino
          FROM scintela.mov_doble m
          LEFT JOIN scintela.transacciones_bancarias tx_o
                 ON tx_o.id_transaccion = m.origen_id
          LEFT JOIN scintela.transacciones_bancarias tx_d
                 ON tx_d.id_transaccion = m.destino_id
          LEFT JOIN scintela.banco b_o ON b_o.no_banco = tx_o.no_banco
          LEFT JOIN scintela.banco b_d ON b_d.no_banco = tx_d.no_banco
         WHERE m.id_mov_doble = %s
        """,
        (id_mov_doble,),
    )
    if not md:
        abort(404)
    if md.get("tipo") != "transfer_banco_banco":
        flash(
            f"mov_doble #{id_mov_doble} no es una transferencia banco↔banco.",
            "warn",
        )
        return redirect(url_for("historial.lista"))
    if md.get("estado") != "activo":
        flash(
            f"La transferencia ya está en estado {md.get('estado')!r} — no se puede reversar otra vez.",
            "warn",
        )
        return redirect(url_for("historial.lista"))

    if request.method == "POST":
        motivo = (request.form.get("motivo") or "").strip()
        try:
            usuario = (g.user or {}).get("username", "web")
            r = queries.reversar_transferencia(
                id_mov_doble=id_mov_doble,
                motivo=motivo,
                usuario=usuario,
            )
            flash(
                f"Transferencia reversada. NC en banco origen "
                f"(tx #{r['compensacion_origen']}) + CH en destino "
                f"(tx #{r['compensacion_destino']}). Importe $ {r['importe']:.2f}.",
                "ok",
            )
            return redirect(url_for("historial.lista"))
        except ValueError as e:
            flash(str(e), "warn")
        except Exception as e:
            flash_exc("No pude reversar la transferencia", e)
        return redirect(url_for("historial.lista"))

    detalle = {
        "Importe": f"$ {float(md.get('importe') or 0):,.2f}",
        "Fecha": md.get("fecha_operacion"),
        "Banco origen": f"{md.get('banco_origen') or ''} (#{md.get('no_banco_origen') or '?'})",
        "Banco destino": f"{md.get('banco_destino') or ''} (#{md.get('no_banco_destino') or '?'})",
        "Concepto": md.get("concepto") or "—",
    }
    return render_template(
        "_confirmar_accion.html",
        titulo=f"Reversar transferencia bancaria #{id_mov_doble}",
        mensaje=(
            "Vas a reversar la transferencia banco↔banco. Se va a insertar "
            "una NC (ingreso) en el banco origen y una CH (egreso) en el "
            "destino — ambas compensan los movimientos originales. Atómico."
        ),
        detalle_registro=detalle,
        accion_url=url_for("bancos.reversar_transferencia", id_mov_doble=id_mov_doble),
        volver_url=url_for("historial.lista"),
        motivo_requerido=True,
        motivo_obligatorio=False,
        confirm_label="Confirmar reverso de transferencia",
    )


@bancos_bp.route("/bancos/<int:no_banco>")
@requiere_login
@requiere_permiso("bancos.ver")
def movimientos(no_banco):
    desde = request.args.get("desde") or None
    hasta = request.args.get("hasta") or None
    # TMT 2026-05-26 dueña: filtro por estado de conciliación bancaria.
    # Valores: '' (todos) | 'si' (solo conciliados) | 'no' (solo sin conciliar).
    conciliado_filtro = (request.args.get("conciliado") or "").strip().lower()
    # TMT 2026-05-28 dueña: filtros adicionales para cheques de cliente.
    cliente_filtro = (request.args.get("cliente") or "").strip()
    doc_num_filtro = (request.args.get("doc_num") or "").strip()
    monto_raw = (request.args.get("monto") or "").strip()
    monto_filtro = parse_monto(monto_raw) if monto_raw else None
    try:
        banco = queries.banco_info(no_banco)
        if not banco:
            abort(404)
        filas = queries.movimientos(
            no_banco,
            desde,
            hasta,
            cliente=cliente_filtro or None,
            monto=float(monto_filtro) if monto_filtro is not None else None,
            doc_num=doc_num_filtro or None,
        )
        # Aplicar filtro post-fetch (las filas vienen con r["conciliacion_id"]
        # gracias al enrichment en queries.movimientos).
        if conciliado_filtro == "si":
            filas = [r for r in filas if r.get("conciliacion_id")]
        elif conciliado_filtro == "no":
            filas = [r for r in filas if not r.get("conciliacion_id")]
        error = None
    except Exception as e:
        filas, banco, error = [], None, str(e)

    if request.args.get("export") == "csv" and filas:
        return csv_response(
            filas,
            columnas=[
                ("fecha", "Fecha"),
                ("tipo", "Tipo"),
                ("referencia", "Referencia"),
                ("descripcion", "Descripción"),
                ("debito", "Débito"),
                ("credito", "Crédito"),
                ("saldo", "Saldo"),
            ],
            filename=f"banco_{no_banco}_movimientos.csv",
        )

    # TMT 2026-05-20 — Excel export de movimientos. Pedido dueña: "del
    # banco dejame descargar un excel". Mismo patrón que comisiones —
    # openpyxl con bold + número format. Una sola hoja con todos los
    # movimientos del rango filtrado + total al final.
    if request.args.get("export") == "xlsx" and filas:
        import io

        from flask import Response

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            flash("openpyxl no instalado en el server.", "error")
            return redirect(url_for("bancos.movimientos", no_banco=no_banco))

        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"
        bold = Font(bold=True)
        hdr_fill = PatternFill("solid", fgColor="E2E8F0")
        # Cabecera del banco arriba.
        ws["A1"] = f"Movimientos · {banco.get('nombre') or ('Banco ' + str(no_banco))}"
        ws["A1"].font = Font(bold=True, size=14)
        if desde or hasta:
            ws["A2"] = f"Período: {desde or 'inicio'} → {hasta or 'hoy'}"
        ws["A3"] = f"Generado: {today_ec().isoformat()}"

        # Headers de la tabla en la fila 5.
        headers = ["Fecha", "Doc", "Concepto", "F. depósito", "Importe", "Saldo", "Estado"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=5, column=i, value=h)
            c.font = bold
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="left")

        # Filas — orden ORIGINAL (DESC) tal como las ve en pantalla.
        row_idx = 6
        for m in filas:
            doc = (m.get("documento") or "").strip()
            es_egreso = doc in ("CH", "ND", "DB")
            imp_raw = m.get("importe") or 0
            imp_abs = imp_raw if imp_raw is None or imp_raw >= 0 else -imp_raw
            imp_excel = float(imp_abs) * (-1 if es_egreso else 1)
            estado = m.get("mov_estado") or ""
            ws.cell(row=row_idx, column=1, value=m.get("fecha"))
            ws.cell(row=row_idx, column=2, value=doc)
            ws.cell(row=row_idx, column=3, value=m.get("concepto") or "")
            ws.cell(row=row_idx, column=4, value=m.get("fechad"))
            c_imp = ws.cell(row=row_idx, column=5, value=imp_excel)
            c_imp.number_format = "#,##0.00"
            c_sal = ws.cell(
                row=row_idx,
                column=6,
                value=(float(m.get("saldo")) if m.get("saldo") is not None else None),
            )
            c_sal.number_format = "#,##0.00"
            ws.cell(row=row_idx, column=7, value=estado)
            row_idx += 1

        # Total firmado al final (saldo neto del período visible).
        total_row = row_idx + 1
        c = ws.cell(row=total_row, column=1, value="TOTAL")
        c.font = bold
        c.fill = hdr_fill
        total_imp = 0.0
        for m in filas:
            doc = (m.get("documento") or "").strip()
            es_egreso = doc in ("CH", "ND", "DB")
            imp_raw = m.get("importe") or 0
            imp_abs = imp_raw if imp_raw is None or imp_raw >= 0 else -imp_raw
            total_imp += float(imp_abs) * (-1 if es_egreso else 1)
        t = ws.cell(row=total_row, column=5, value=total_imp)
        t.font = bold
        t.number_format = "#,##0.00"
        t.fill = hdr_fill

        # Anchos de columnas.
        for col, w in [("A", 12), ("B", 8), ("C", 50), ("D", 12), ("E", 16), ("F", 16), ("G", 14)]:
            ws.column_dimensions[col].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"banco_{no_banco}_{desde or 'inicio'}_{hasta or 'hoy'}.xlsx"
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # TMT 2026-05-28 dueña v3: 'los cheques en el banco haz que aparezcan
    # linea por linea, no agrupados'. Revertimos el agrupamiento por lote
    # (v1 2026-05-20, v2 2026-05-27 ventana fecha_crea). Cada DE va como
    # row individual. Si más adelante quisiera volver al lote, está en git.
    items: list[dict] = [{"_kind": "row", **f} for f in filas]

    # TMT 2026-05-27 dueña: el total tiene que ser EL MISMO en /bancos y
    # /conciliacion/banco. Antes sumaba solo filas visibles (LIMIT 500) y
    # daba un numero distinto que el saldo pendiente de conciliacion.
    # Ahora hacemos una query AGGREGATE separada sobre TODAS las txs que
    # matchean los filtros + 'saldo antes/despues' como en dBase.
    _DOCS_CR = ('DE','AC','NC','TR')
    _DOCS_DB = ('CH','ND','DB','GS','PA')
    total_filtrado = 0.0
    total_abs_filtrado = 0.0
    n_filtrado = 0
    saldo_pre_concil = 0.0  # suma signada de los movs que faltan conciliar
    saldo_banco = 0.0       # running balance actual del banco
    try:
        # Saldo actual del banco (último running)
        row_saldo = queries.banco_info(no_banco) or {}
        saldo_banco = float(row_saldo.get("saldo_stored") or 0)
        # Suma signada sobre todas las txs del filtro (sin LIMIT)
        import db as _db
        where_filt = []
        params: dict = {"no_banco": no_banco}
        if desde:
            where_filt.append("t.fecha >= %(desde)s::date")
            params["desde"] = desde
        if hasta:
            where_filt.append("t.fecha <= %(hasta)s::date")
            params["hasta"] = hasta
        if monto_filtro is not None:
            where_filt.append("t.importe = %(monto)s::numeric")
            params["monto"] = float(monto_filtro)
        if doc_num_filtro:
            where_filt.append(
                "UPPER(COALESCE(t.numreferencia::text,'')) LIKE %(doc_like)s"
            )
            params["doc_like"] = f"%{doc_num_filtro.upper()}%"
        if cliente_filtro:
            where_filt.append(
                "EXISTS (SELECT 1 FROM scintela.chequextransaccion cxt "
                "JOIN scintela.cheque c ON c.id_cheque = cxt.id_cheque "
                "LEFT JOIN scintela.cliente cli ON cli.codigo_cli = c.codigo_cli "
                "WHERE cxt.id_transaccion = t.id_transaccion "
                "AND (UPPER(COALESCE(cli.nombre,'')) LIKE %(cli_like)s "
                "OR UPPER(COALESCE(c.codigo_cli,'')) LIKE %(cli_like)s))"
            )
            params["cli_like"] = f"%{cliente_filtro.upper()}%"
        if conciliado_filtro == "no":
            # Excluir conciliados PC + conciliados dBase (stat='*')
            where_filt.append("TRIM(COALESCE(t.stat,'')) <> '*'")
            where_filt.append("""NOT EXISTS (
                SELECT 1 FROM scintela.banco_conciliacion_match m
                 WHERE m.id_transaccion = t.id_transaccion
                   AND m.deshecho_en IS NULL)""")
        elif conciliado_filtro == "si":
            where_filt.append("""(TRIM(COALESCE(t.stat,'')) = '*' OR EXISTS (
                SELECT 1 FROM scintela.banco_conciliacion_match m
                 WHERE m.id_transaccion = t.id_transaccion
                   AND m.deshecho_en IS NULL))""")
        sql_agg = (
            "SELECT COUNT(*) AS n, "
            "COALESCE(SUM(CASE WHEN t.documento IN ('CH','ND','DB','GS','PA') "
            "                  THEN -t.importe ELSE t.importe END), 0) AS signed, "
            "COALESCE(SUM(t.importe), 0) AS absoluta "
            "FROM scintela.transacciones_bancarias t "
            "WHERE t.no_banco = %(no_banco)s "
        )
        if where_filt:
            sql_agg += "AND " + " AND ".join(where_filt)
        row_agg = _db.fetch_one(sql_agg, params) or {}
        n_filtrado = int(row_agg.get("n") or 0)
        total_filtrado = round(float(row_agg.get("signed") or 0), 2)
        total_abs_filtrado = round(float(row_agg.get("absoluta") or 0), 2)
        if conciliado_filtro == "no":
            saldo_pre_concil = total_filtrado
    except Exception as _e:
        import logging
        logging.getLogger("programa_core.bancos").exception(
            "total_filtrado AGG falló: %s", _e
        )
    # Saldo despues de conciliar todo lo pendiente = saldo actual − total_pendiente
    # (las tx pendientes ya impactaron el saldo_stored porque existen; al
    # marcarlas conciliadas el saldo no cambia. Es solo display).
    saldo_post_concil = round(saldo_banco - saldo_pre_concil, 2) if conciliado_filtro == "no" else saldo_banco

    # TMT 2026-05-28 dueña: 'pongamos debajo del saldo del banco Saldo
    # conciliado que tiene que sumar todo lo que tiene conciliado'.
    # v1 (ANTES): SUM(signed) WHERE stat='*' → daba el delta acumulado
    #             ($349K para Pichincha) — NO es lo que la dueña espera.
    # v2 (AHORA): saldo_banco − sum_pendientes_signed → da el "saldo si
    #             concilio todo" ($2.557K para Pichincha). Cuadra con el
    #             "SALDO SISTEMA" del archivo de conciliación. Equivale a
    #             saldo_banco descontando los movs NO conciliados.
    saldo_conciliado = 0.0
    n_conciliado = 0
    try:
        import db as _db
        # Suma signada de los movs PENDIENTES (no conciliados PC ni dBase).
        row_pend = _db.fetch_one(
            "SELECT COUNT(*) AS n, "
            "COALESCE(SUM(CASE WHEN t.documento IN ('CH','ND','DB','GS','PA') "
            "                  THEN -t.importe ELSE t.importe END), 0) AS signed "
            "FROM scintela.transacciones_bancarias t "
            "WHERE t.no_banco = %(no_banco)s "
            "  AND TRIM(COALESCE(t.stat, '')) <> '*' "
            "  AND NOT EXISTS ("
            "      SELECT 1 FROM scintela.banco_conciliacion_match m "
            "       WHERE m.id_transaccion = t.id_transaccion "
            "         AND m.deshecho_en IS NULL"
            "  )",
            {"no_banco": no_banco},
        ) or {}
        n_conciliado = int(row_pend.get("n") or 0)
        pendientes_signed = round(float(row_pend.get("signed") or 0), 2)
        saldo_conciliado = round(saldo_banco - pendientes_signed, 2)
    except Exception as _e:
        import logging
        logging.getLogger("programa_core.bancos").exception(
            "saldo_conciliado AGG falló: %s", _e
        )

    hay_filtro = bool(
        desde or hasta or conciliado_filtro
        or cliente_filtro or doc_num_filtro or monto_filtro is not None
    )

    return render_template(
        "bancos/movimientos.html",
        banco=banco,
        filas=filas,
        items=items,
        desde=desde,
        hasta=hasta,
        error=error,
        total_filtrado=total_filtrado,
        total_abs_filtrado=total_abs_filtrado,
        n_filtrado=n_filtrado,
        hay_filtro=hay_filtro,
        conciliado_filtro=conciliado_filtro,
        cliente_filtro=cliente_filtro,
        doc_num_filtro=doc_num_filtro,
        monto_filtro_raw=monto_raw,
        saldo_banco=saldo_banco,
        saldo_pre_concil=saldo_pre_concil,
        saldo_post_concil=saldo_post_concil,
        saldo_conciliado=saldo_conciliado,
        n_conciliado=n_conciliado,
    )


@bancos_bp.route("/bancos/recompute-saldos", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.editar")
def recompute_saldos():
    """Recalcula el running `saldo` de TODAS las filas de
    `transacciones_bancarias` para todos los bancos. Necesario una sola vez
    para repara los depósitos que se hicieron con código viejo (saldo=NULL).

    El walk-forward es seguro de correr: si todo está bien, devuelve los
    mismos números. Si hay filas con saldo=NULL o desfasadas, las arregla.

    Reportado por TMT 2026-05-11.
    """
    import bank_helpers
    import db as _db

    try:
        bancos = (
            _db.fetch_all(
                "SELECT no_banco, COALESCE(nombre,'') AS nombre FROM scintela.banco ORDER BY no_banco"
            )
            or []
        )
        total_filas = 0
        bancos_sin_filas = 0
        # TX SEPARADA POR BANCO (TMT 2026-05-14 #20 audit):
        # Antes había una sola tx envolvente para todos los bancos. Si un
        # banco tenía 10k+ filas el lock bloqueaba INSERTs concurrentes en
        # transacciones_bancarias por minutos. Una tx por banco da pasos
        # cortos y permite a otros usuarios escribir entre medio.
        for b in bancos:
            no_banco = int(b["no_banco"])
            # ANCLA = SEGUNDA fila más vieja (NOT primera). bug #R1 audit
            # 2026-05-14: usar MIN(id_transaccion) es equivalente a
            # ancla=None porque NO hay filas con id < MIN(id), entonces
            # `saldo_previo` queda en 0 y el walk arranca de cero —
            # reintroduce el bug histórico que destruyó Pichincha el
            # 2026-05-12. El ancla debe respetar el saldo de la primera
            # fila como opening implícito y walkear desde la segunda.
            anc = _db.fetch_one(
                """
                SELECT id_transaccion AS ancla
                  FROM scintela.transacciones_bancarias
                 WHERE no_banco = %s
                 ORDER BY fecha ASC, id_transaccion ASC
                 OFFSET 1 LIMIT 1
                """,
                (no_banco,),
            )
            ancla_id = anc.get("ancla") if anc else None
            if not ancla_id:
                # Banco con 0 o 1 fila — nada que recomputar (sin segunda
                # fila no hay walk; la única fila es ya el opening).
                bancos_sin_filas += 1
                continue
            with _db.tx() as conn:
                n = bank_helpers.recompute_saldos_desde(
                    conn,
                    no_banco=no_banco,
                    no_cta=None,
                    ancla_id=int(ancla_id),
                )
            total_filas += int(n or 0)
        msg = (
            f"Saldos recalculados: {total_filas} fila(s) tocadas en "
            f"{len(bancos) - bancos_sin_filas} banco(s)."
        )
        if bancos_sin_filas:
            msg += f" {bancos_sin_filas} banco(s) sin transacciones."
        flash(msg, "ok")
    except Exception as e:
        flash_exc("No pude recalcular saldos bancarios", e)
    return redirect(url_for("bancos.lista"))


# Usuarios "del sync" — sus filas vienen del dBase y NUNCA se borran acá.
_USUARIOS_SYNC = ("dbf-import", "asinfo-carga", "asinfo-backfill")


@bancos_bp.route("/bancos/<int:no_banco>/tx/<int:id_transaccion>/eliminar",
                 methods=["GET", "POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def eliminar_movimiento_pc(no_banco: int, id_transaccion: int):
    """Elimina un movimiento de banco CARGADO EN PC que no tiene reverso
    automático (sin mov_doble) ni cheques enganchados.

    TMT 2026-06-26 (dueña): un depósito que alex cargó a mano en INTERNACI
    ($500, "1 ch.TOM") que el dBase no tiene, y la pantalla no tenía cómo
    sacarlo. Guards (todos deben cumplirse): solo movs PC (usuario_crea no es
    del sync), sin conciliar, sin mov_doble activo y sin cheques enganchados.
    El dBase no se toca; si el mov viene del dBase se corrige allá y se
    re-sincroniza.
    """
    import bank_helpers
    import db as _db

    tx = _db.fetch_one(
        """
        SELECT t.id_transaccion, t.no_banco, t.fecha, t.documento, t.importe,
               t.concepto, t.usuario_crea,
               COALESCE(b.nombre, '') AS banco_nombre
          FROM scintela.transacciones_bancarias t
          LEFT JOIN scintela.banco b ON b.no_banco = t.no_banco
         WHERE t.id_transaccion = %s AND t.no_banco = %s
        """,
        (id_transaccion, no_banco),
    )
    if not tx:
        abort(404)

    usuario_crea = (tx.get("usuario_crea") or "").strip().lower()
    if (not usuario_crea) or usuario_crea in _USUARIOS_SYNC:
        flash("Este movimiento viene del dBase (sync) — no se borra desde acá. "
              "Corregilo en el dBase y re-sincronizá.", "warn")
        return redirect(url_for("bancos.movimientos", no_banco=no_banco))

    conc = _db.fetch_one(
        "SELECT 1 FROM scintela.banco_conciliacion_match "
        "WHERE id_transaccion = %s AND deshecho_en IS NULL LIMIT 1",
        (id_transaccion,),
    )
    if conc:
        flash("Este movimiento está conciliado — desconciliá primero.", "warn")
        return redirect(url_for("bancos.movimientos", no_banco=no_banco))

    md = _db.fetch_one(
        "SELECT id_mov_doble FROM scintela.mov_doble "
        "WHERE destino_table = 'transacciones_bancarias' AND destino_id = %s "
        "  AND estado = 'activo' LIMIT 1",
        (id_transaccion,),
    )
    if md:
        flash("Este movimiento tiene reverso propio (mov_doble) — reversalo "
              "desde /historial, no se borra a mano.", "warn")
        return redirect(url_for("bancos.movimientos", no_banco=no_banco))

    # Cheques enganchados: NO abortamos. Hacemos la reversa COMPLETA del
    # depósito (TMT 2026-06-26 dueña: alex depositó un cheque TOM en INTERNACI
    # por error). Un cheque "vivo" todavía depositado (stat B/I) vuelve a
    # cartera (Z); si el enlace quedó HUÉRFANO (el sync de CHEQUES.DBF recargó
    # los cheques con ids nuevos), solo se limpia el enlace. En ambos casos se
    # borra el mov de banco y se recalcula el saldo.
    links = _db.fetch_all(
        "SELECT id_cheque FROM scintela.chequextransaccion WHERE id_transaccion = %s",
        (id_transaccion,),
    ) or []
    cheques_vivos = []
    for lk in links:
        ch = _db.fetch_one(
            "SELECT id_cheque, no_cheque, stat FROM scintela.cheque WHERE id_cheque = %s",
            (lk.get("id_cheque"),),
        )
        if ch and (ch.get("stat") or "").upper() in ("B", "I"):
            cheques_vivos.append(ch)

    if request.method == "POST":
        usuario = (g.user or {}).get("username", "web")
        try:
            with _db.tx() as conn:
                # 1) Cheques vivos depositados → vuelven a cartera (Z).
                for ch in cheques_vivos:
                    _db.execute(
                        "UPDATE scintela.cheque "
                        "   SET stat = 'Z', fechaing = NULL, "
                        "       usuario_modifica = %s, fecha_modifica = CURRENT_TIMESTAMP "
                        " WHERE id_cheque = %s",
                        (usuario, ch.get("id_cheque")), conn=conn,
                    )
                # 2) Borrar los enlaces cheque↔transacción (vivos y huérfanos).
                _db.execute(
                    "DELETE FROM scintela.chequextransaccion WHERE id_transaccion = %s",
                    (id_transaccion,), conn=conn,
                )
                # 3) Borrar el mov de banco.
                _db.execute(
                    "DELETE FROM scintela.transacciones_bancarias "
                    "WHERE id_transaccion = %s AND no_banco = %s",
                    (id_transaccion, no_banco), conn=conn,
                )
                # 4) Recompute del running saldo (ancla = 2da fila más vieja).
                anc = _db.fetch_one(
                    "SELECT id_transaccion AS ancla "
                    "  FROM scintela.transacciones_bancarias "
                    " WHERE no_banco = %s "
                    " ORDER BY fecha ASC, id_transaccion ASC OFFSET 1 LIMIT 1",
                    (no_banco,), conn=conn,
                )
                if anc and anc.get("ancla"):
                    bank_helpers.recompute_saldos_desde(
                        conn, no_banco=no_banco, no_cta=None,
                        ancla_id=int(anc["ancla"]),
                    )
            extra = (f" {len(cheques_vivos)} cheque(s) devuelto(s) a cartera."
                     if cheques_vivos else "")
            flash(
                f"Movimiento #{id_transaccion} eliminado "
                f"({tx.get('concepto') or ''} {float(tx.get('importe') or 0):,.2f})."
                + extra,
                "ok",
            )
        except Exception as e:
            flash_exc("No pude eliminar el movimiento", e)
        return redirect(url_for("bancos.movimientos", no_banco=no_banco))

    return render_template(
        "bancos/eliminar_mov.html",
        tx=tx,
        cheques_vivos=cheques_vivos,
        n_links=len(links),
    )


# ---------------------------------------------------------------------------
# Nuevo movimiento simple (DE / NC / ND) — TMT 2026-05-19 (pedido dueña).
# Una vista genérica con un parámetro `doc` que decide el tipo. Reutiliza
# el mismo template.
# ---------------------------------------------------------------------------
_LABELS_DOC = {
    "DE": ("Depósito", "Suma al saldo del banco"),
    "NC": ("Nota de crédito", "Suma al saldo (devolución, intereses, reverso de cargo)"),
    "ND": ("Nota de débito", "Resta del saldo (cargo del banco, comisión, ISI)"),
}


@bancos_bp.route("/bancos/nuevo-movimiento", methods=["GET", "POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def nuevo_movimiento():
    """Form genérico para crear DE / NC / ND.

    El tipo de documento llega vía `?doc=DE|NC|ND` (GET) o el campo hidden
    `documento` (POST). Comparte template `bancos/nuevo_movimiento.html`.
    """
    doc = (request.args.get("doc") or request.form.get("documento") or "").upper().strip()
    if doc not in _LABELS_DOC:
        flash("Documento inválido (debe ser DE, NC o ND).", "warn")
        return redirect(url_for("bancos.lista"))
    label, ayuda = _LABELS_DOC[doc]

    try:
        bancos = queries.lista_bancos()
    except Exception as e:
        bancos, _err = [], str(e)
        flash_exc("No pude listar bancos", e)

    pregunta_activa = None  # cuenta USD pendiente de ACTIVA? (paridad dBase)

    if request.method == "POST":
        no_banco = parse_int(request.form.get("no_banco"))
        importe = parse_monto(request.form.get("importe"))
        fecha = parse_date(request.form.get("fecha")) or today_ec()
        concepto = (request.form.get("concepto") or "").strip()
        prov = (request.form.get("beneficiario") or "").strip().upper() or None
        usuario = (g.user or {}).get("username", "web")
        # TMT 2026-06-09: override del guard anti-duplicado (repetidos reales).
        permitir_dup = (request.form.get("permitir_duplicado") or "") == "1"
        # ACTIVA? — réplica del prompt del dBase (BANCOS.PRG ~216). 'S'/'N'
        # llega del segundo submit; sin respuesta = None → queries levanta
        # ActivaRequerida y acá se re-pregunta.
        activa_raw = (request.form.get("activa") or "").strip().upper()
        activa = True if activa_raw == "S" else (False if activa_raw == "N" else None)
        # TMT 2026-06-15: anticipo de proveedor EXPLÍCITO (select del form, solo ND).
        anticipo_prov = (request.form.get("anticipo_prov") or "").strip().upper() or None
        # TMT 2026-06-15: destino especial EXPLÍCITO (retiro/caja/posdato) via
        # select del form, en vez de tipear el concepto en formato mágico. Si se
        # elige uno, reescribimos el concepto al formato que _routear_mov_simple
        # ya maneja (RR/CAJA/INOP) y limpiamos beneficiario para que no caiga en
        # COMPRAS. Retrocompat: sin selección, el tipeo viejo sigue andando.
        mov_destino = (request.form.get("mov_destino") or "").strip().lower()
        mov_param = (request.form.get("mov_destino_param") or "").strip().upper()[:2]
        if doc == "ND" and not anticipo_prov and mov_destino in ("retiro", "caja", "posdat"):
            _desc = concepto
            # El código de 2 letras va en posición FIJA (de=c[3:5] para RR,
            # prov=c[5:7] para INOP); ljust(2) lo mantiene aunque venga corto.
            _p2 = (mov_param or "").ljust(2)
            if mov_destino == "retiro":
                concepto = ("RR " + _p2 + " " + _desc).strip()
            elif mov_destino == "caja":
                concepto = ("CAJA " + _desc).strip()
            elif mov_destino == "posdat":
                concepto = ("INOP " + _p2 + " " + _desc).strip()
            prov = None
        try:
            r = queries.crear_movimiento_simple(
                no_banco=no_banco,
                documento=doc,
                importe=importe,
                fecha=fecha,
                concepto=concepto,
                prov=prov,
                usuario=usuario,
                permitir_duplicado=permitir_dup,
                activa=activa,
                anticipo_prov=anticipo_prov,
            )
            if r.get("dedupe"):
                # TMT 2026-06-09: dedupe silencioso — el mov ya estaba.
                msg = (
                    f"{label} por $ {r['importe']:.2f} ya estaba cargada "
                    f"(mov #{r['id_transaccion']}) — no se duplicó."
                )
            else:
                msg = f"{label} registrada por $ {r['importe']:.2f}. Nuevo saldo: $ {r['saldo_nuevo']:.2f}."
            if r.get("side_effect"):
                # TMT 2026-06-09 paridad dBase: contraparte creada a la vez
                # (anticipo USD / retiro / caja).
                msg += f" {r['side_effect']}."
            flash(msg, "ok")
            return redirect(url_for("bancos.movimientos", no_banco=r["no_banco"]))
        except queries.ActivaRequerida as e:
            # Réplica del prompt dBase: nada se grabó — re-render con la
            # pregunta y los valores tal como vinieron.
            pregunta_activa = e.cta
        except ValueError as e:
            flash(str(e), "warn")
        except Exception as e:
            flash_exc(f"No pude registrar la {label.lower()}", e)

    # GET o POST con error
    import contextlib as _ctx

    proveedores = []
    with _ctx.suppress(Exception):
        proveedores = queries.proveedores_activos(limite=500)

    # Andres 2026-06-18: cuentas OP (anticipo por proveedor) con saldo, para el
    # datalist del destino 'Posdato (INOP)'. Solo aplica a ND.
    cuentas_op = []
    if doc == "ND":
        with _ctx.suppress(Exception):
            cuentas_op = queries.proveedores_op_saldos(limite=500)

    # TMT 2026-05-19 v2 — banco pre-seleccionado vía ?no_banco= (default
    # Pichincha cuando se entra desde /bancos).
    # TMT 2026-05-19 v7 — si no viene ?no_banco=, default = Pichincha.
    no_banco_inicial = parse_int(request.args.get("no_banco"))
    if not no_banco_inicial:
        for b in bancos or []:
            if "PICHINC" in (b.get("nombre") or "").upper():
                no_banco_inicial = int(b["no_banco"])
                break

    # Si venimos de un POST (ACTIVA? o error), preservar lo tipeado.
    form_vals = {}
    if request.method == "POST":
        form_vals = {
            "no_banco": request.form.get("no_banco") or "",
            "importe": request.form.get("importe") or "",
            "fecha": request.form.get("fecha") or "",
            "beneficiario": request.form.get("beneficiario") or "",
            "concepto": request.form.get("concepto") or "",
            "anticipo_prov": request.form.get("anticipo_prov") or "",
            "mov_destino": request.form.get("mov_destino") or "",
            "mov_destino_param": request.form.get("mov_destino_param") or "",
        }
        if form_vals["no_banco"]:
            no_banco_inicial = parse_int(form_vals["no_banco"]) or no_banco_inicial

    return render_template(
        "bancos/nuevo_movimiento.html",
        doc=doc,
        label=label,
        ayuda=ayuda,
        bancos=bancos,
        no_banco_inicial=no_banco_inicial,
        proveedores=proveedores,
        cuentas_op=cuentas_op,
        hoy=today_ec().isoformat(),
        pregunta_activa=pregunta_activa,
        form_vals=form_vals,
    )


@bancos_bp.route("/bancos/mov-simple/<int:id_mov_doble>/reversar", methods=["GET", "POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def confirmar_reverso_movimiento_simple(id_mov_doble: int):
    """Wizard de reverso para DE / NC / ND creados vía /bancos/nuevo-movimiento.

    GET: muestra detalles + input de motivo.
    POST: ejecuta queries.reversar_movimiento_simple (atómico).
    """
    import db as _db

    md = _db.fetch_one(
        """
        SELECT id_mov_doble, tipo, origen_id, importe, fecha, concepto, estado
          FROM scintela.mov_doble
         WHERE id_mov_doble = %s
        """,
        (id_mov_doble,),
    )
    if not md:
        flash(f"mov_doble #{id_mov_doble} no existe.", "warn")
        return redirect(url_for("historial.lista"))
    if (md.get("estado") or "") != "activo":
        flash(f"Este mov_doble ya está {md.get('estado')}.", "warn")
        return redirect(url_for("historial.lista"))

    # Info del movimiento original para mostrar.
    tx = _db.fetch_one(
        """
        SELECT t.id_transaccion, t.no_banco, t.documento, t.importe, t.fecha,
               t.saldo, COALESCE(b.nombre, '') AS banco_nombre
          FROM scintela.transacciones_bancarias t
          LEFT JOIN scintela.banco b ON b.no_banco = t.no_banco
         WHERE t.id_transaccion = %s
        """,
        (md.get("origen_id"),),
    )

    if request.method == "POST":
        motivo = (request.form.get("motivo") or "").strip()
        usuario = (g.user or {}).get("username", "web")
        try:
            r = queries.reversar_movimiento_simple(
                id_mov_doble=id_mov_doble,
                motivo=motivo,
                usuario=usuario,
            )
            flash(
                f"Reverso OK: {r['doc_orig']} compensado con {r['doc_reverso']}. "
                f"Nuevo saldo: $ {r['saldo_nuevo']:.2f}.",
                "ok",
            )
            return redirect(url_for("historial.lista"))
        except ValueError as e:
            flash(str(e), "warn")
        except Exception as e:
            flash_exc("No pude reversar el movimiento", e)

    # Render wizard de confirmación.
    tipo_legible = (md.get("tipo") or "").replace("_", " ")
    return render_template(
        "_confirmar_accion.html",
        titulo=f"Reversar {tipo_legible}",
        mensaje=(
            "Esta acción compensa el movimiento original con un documento de "
            "signo opuesto (DE/NC → CH; ND → NC). El saldo del banco queda como "
            "estaba antes del alta."
        ),
        detalle_registro={
            "Tipo": tipo_legible,
            "Banco": tx.get("banco_nombre", "") if tx else "",
            "Importe": f"$ {md.get('importe', 0):.2f}",
            "Concepto": md.get("concepto") or "(sin concepto)",
            "Fecha": md.get("fecha"),
        },
        accion_url=url_for("bancos.confirmar_reverso_movimiento_simple", id_mov_doble=id_mov_doble),
        volver_url=url_for("historial.lista"),
        # TMT 2026-07-08 dueña: el motivo de reverso NO es obligatorio ("se
        # hace largo"). El handler ya lo acepta vacío.
        motivo_obligatorio=False,
        confirm_label="Reversar",
    )


# ---------------------------------------------------------------------------
# Toggle conciliado manual — TMT 2026-05-28 (dueña).
# Pedido literal: "o dejame editar el conciliado si no con una advertencia".
# Click en la celda Conc. → JS pregunta confirmación → POST acá.
# Marca/desmarca scintela.transacciones_bancarias.stat ('*' ↔ NULL) y, si la
# fila tiene un match PC activo, lo deshace en la misma operación. Esto
# cubre los casos "fantasma-conciliado" que sobreviven al deshacer (bug del
# dual-write previo al fix de matcher_banco.romper_match).
# ---------------------------------------------------------------------------
@bancos_bp.route("/bancos/<int:no_banco>/tx/<int:id_transaccion>/set-numreferencia",
                 methods=["POST"], endpoint="set_numreferencia")
@requiere_login
@requiere_permiso("bancos.conciliar")
def set_numreferencia(no_banco: int, id_transaccion: int):
    """TMT 2026-06-03 dueña: 'aca quiero ver documento y poder editar, asi
    les agrego numero de documento para hacer la conciliacion por num de
    documento'. Update inline del campo numreferencia_manual (sobrevive
    al sync dBase, mig 0074)."""
    import db as _db
    raw = (request.form.get("numreferencia") or "").strip()
    valor = raw[:30] if raw else None
    try:
        n = _db.execute(
            """
            UPDATE scintela.transacciones_bancarias
               SET numreferencia_manual = %s
             WHERE id_transaccion = %s AND no_banco = %s
            """,
            (valor, id_transaccion, no_banco),
        )
        if n:
            flash(f"N° doc actualizado: {valor or '(vacío)'}", "ok")
        else:
            flash("No encontré la transacción.", "warn")
    except Exception as e:
        flash(f"Error al actualizar: {e}", "error")
    return redirect(request.referrer or url_for("bancos.movimientos", no_banco=no_banco))


@bancos_bp.route("/bancos/<int:no_banco>/tx/<int:id_transaccion>/toggle-conciliado",
                 methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def toggle_conciliado(no_banco: int, id_transaccion: int):
    import db as _db

    row = _db.fetch_one(
        """
        SELECT t.id_transaccion, t.no_banco, t.stat,
               (SELECT m.id FROM scintela.banco_conciliacion_match m
                 WHERE m.id_transaccion = t.id_transaccion
                   AND m.deshecho_en IS NULL
                 LIMIT 1) AS match_activo_id
          FROM scintela.transacciones_bancarias t
         WHERE t.id_transaccion = %s AND t.no_banco = %s
        """,
        (id_transaccion, no_banco),
    )
    if not row:
        abort(404)
    estaba_conciliado = (
        (row.get("stat") or "").strip() == "*"
        or row.get("match_activo_id") is not None
    )
    usuario = (g.user or {}).get("username", "web")
    try:
        if estaba_conciliado:
            # Desconciliar: limpiar stat y deshacer el match activo si existe.
            _db.execute(
                "UPDATE scintela.transacciones_bancarias SET stat = NULL "
                "WHERE id_transaccion = %s",
                (id_transaccion,),
            )
            if row.get("match_activo_id"):
                from modules.conciliacion.matcher_banco import romper_match
                romper_match(
                    match_id=int(row["match_activo_id"]),
                    usuario=usuario,
                )
            flash(
                f"Mov #{id_transaccion}: marcado como NO conciliado. "
                f"Si era un error, volvé a clickear para conciliar.",
                "ok",
            )
        else:
            # Marcar conciliado manualmente (sin extracto). Solo stat='*'.
            _db.execute(
                "UPDATE scintela.transacciones_bancarias SET stat = '*' "
                "WHERE id_transaccion = %s",
                (id_transaccion,),
            )
            flash(
                f"Mov #{id_transaccion}: marcado como conciliado (manual, "
                f"sin extracto).",
                "ok",
            )
    except Exception as e:
        flash_exc("No pude cambiar el estado de conciliación", e)
    # Vuelve a la lista de movimientos del banco preservando filtros.
    return redirect(
        url_for(
            "bancos.movimientos",
            no_banco=no_banco,
            desde=request.args.get("desde") or request.form.get("desde") or None,
            hasta=request.args.get("hasta") or request.form.get("hasta") or None,
            conciliado=request.args.get("conciliado") or request.form.get("conciliado") or None,
            cliente=request.args.get("cliente") or request.form.get("cliente") or None,
            monto=request.args.get("monto") or request.form.get("monto") or None,
            doc_num=request.args.get("doc_num") or request.form.get("doc_num") or None,
        )
    )


# ---------------------------------------------------------------------------
# Carga multi-línea (bulk) — TMT 2026-06-13 (pedido dueña: "dejame cargar
# varios movimientos del banco de una, no uno por uno"). Réplica del patrón
# de /tinto-carga: N filas en un solo POST (inputs repetidos vía getlist),
# all-or-nothing en el backend (si UNA fila con datos falla, no se carga
# NADA). Solo Pichincha (no_banco=10 hidden, sin selector de banco). Cada
# fila reusa queries.crear_movimiento_simple() — el mismo helper que el alta
# de a uno — así el saldo y el mov_doble quedan exactos (no se reinventa la
# lógica de saldo).
# ---------------------------------------------------------------------------
_BANCO_PICHINCHA = 10  # Intela solo opera Pichincha (convención "solo Pichincha").
_DOCS_CARGA = ("DE", "ND")  # DE = depósito (entra), ND = nota de débito (sale).


@bancos_bp.route("/bancos/cargar")
@requiere_login
@requiere_permiso("bancos.conciliar")
def cargar():
    """Pantalla de carga multi-línea de movimientos de Pichincha."""
    return render_template(
        "bancos/cargar.html",
        hoy=today_ec().isoformat(),
        no_banco=_BANCO_PICHINCHA,
    )


@bancos_bp.route("/bancos/cargar/agregar", methods=["POST"])
@requiere_login
@requiere_permiso("bancos.conciliar")
def cargar_agregar():
    """Alta de N movimientos de Pichincha en un solo POST.

    Inputs repetidos fecha/tipo/importe/concepto/doc_banco vía getlist.
    Filas totalmente vacías se ignoran. Si CUALQUIER fila con datos tiene
    error, no se carga NADA (all-or-nothing, igual que /tinto-carga). Cada
    línea pasa por queries.crear_movimiento_simple() (mismo helper que el
    alta de a uno) → saldo + mov_doble correctos.
    """
    fechas = request.form.getlist("fecha")
    tipos = request.form.getlist("tipo")
    importes = request.form.getlist("importe")
    conceptos = request.form.getlist("concepto")
    docs = request.form.getlist("doc_banco")
    n = max(len(fechas), len(tipos), len(importes), len(conceptos), len(docs))

    def _at(lst, i):
        return lst[i] if i < len(lst) else ""

    filas: list[dict] = []
    errores: list[str] = []
    for i in range(n):
        f_raw = str(_at(fechas, i) or "").strip()
        tipo = (_at(tipos, i) or "").upper().strip()
        imp_raw = str(_at(importes, i) or "").strip()
        concepto = (_at(conceptos, i) or "").strip()
        doc_banco = (_at(docs, i) or "").strip()
        if not imp_raw and not concepto and not doc_banco:
            continue  # línea vacía — ignorar
        rotulo = f"línea {len(filas) + len(errores) + 1}"
        fecha = parse_date(f_raw) or today_ec()
        if tipo not in _DOCS_CARGA:
            errores.append(f"{rotulo}: tipo inválido (debe ser DE o ND)")
            continue
        importe = parse_monto(imp_raw)
        if importe is None or float(importe or 0) <= 0:
            errores.append(f"{rotulo}: importe inválido")
            continue
        filas.append({
            "fecha": fecha,
            "tipo": tipo,
            "importe": float(importe),
            "concepto": concepto,
            "doc_banco": doc_banco or None,
        })

    if not filas and not errores:
        flash("No hay líneas para cargar.", "warn")
        return redirect(url_for("bancos.cargar"))
    if errores:
        flash("No se cargó nada. " + " | ".join(errores), "warn")
        return redirect(url_for("bancos.cargar"))

    usuario = (g.user or {}).get("username", "web")
    creados = 0
    tot_entra = 0.0
    tot_sale = 0.0
    try:
        for f in filas:
            # concepto: si el usuario puso doc_banco pero no concepto, el
            # doc viaja como concepto (es el comprobante del banco). Si puso
            # ambos, se concatena para no perder ninguno.
            concepto = f["concepto"]
            if f["doc_banco"]:
                concepto = (f"{concepto} {f['doc_banco']}").strip() if concepto else f["doc_banco"]
            # activa=False: un ND que parezca anticipo USD NO dispara el
            # prompt ACTIVA? (queda solo en el banco). En carga bulk no hay
            # modo de contestar S/N por fila, y el flujo de anticipos va por
            # su pantalla; acá es carga llana de movimientos del banco.
            r = queries.crear_movimiento_simple(
                no_banco=_BANCO_PICHINCHA,
                documento=f["tipo"],
                importe=f["importe"],
                fecha=f["fecha"],
                concepto=concepto,
                usuario=usuario,
                activa=False,
            )
            if not r.get("dedupe"):
                creados += 1
                if f["tipo"] == "DE":
                    tot_entra += f["importe"]
                else:
                    tot_sale += f["importe"]
    except queries.ActivaRequerida as e:
        # No debería ocurrir (activa=False), pero por las dudas: cortar y
        # avisar en vez de dejar una carga a medias.
        flash(
            f"Una línea es candidata a anticipo USD (cuenta {e.cta}). "
            "Cargá los anticipos por su pantalla y dejá acá solo movimientos "
            "simples del banco.",
            "warn",
        )
        return redirect(url_for("bancos.cargar"))
    except ValueError as e:
        flash(f"No se cargó todo: {e}", "warn")
        return redirect(url_for("bancos.cargar"))
    except Exception as e:
        flash_exc("No pude cargar los movimientos", e)
        return redirect(url_for("bancos.cargar"))

    partes = [f"Cargados {creados} movimiento{'s' if creados != 1 else ''}"]
    if tot_entra:
        partes.append(f"entran $ {tot_entra:,.2f}")
    if tot_sale:
        partes.append(f"salen $ {tot_sale:,.2f}")
    flash(" · ".join(partes) + ".", "ok")
    return redirect(url_for("bancos.cargar"))


@bancos_bp.route("/bancos/cheque-imprimible")
@requiere_login
@requiere_permiso("bancos.ver")
def cheque_imprimible():
    """Impresión del cheque con monto EN LETRAS — réplica de IMPCHEQ (MODIFICA.PRG).

    El dBase imprime sobre el cheque físico pre-impreso: nombre del
    beneficiario + monto en números, el monto en letras con guiones de
    relleno, y "QUITO, <fecha>". Esta pantalla reproduce ese layout en una
    página imprimible; las posiciones son calibrables con `?offx=`/`?offy=`
    (milímetros) para alinear con el cheque de cada banco.

    Fuentes de datos (en orden):
      - `?id_transaccion=N`  → levanta un cheque emitido (documento='CH').
      - params sueltos: `?importe=&beneficiario=&fecha=&no_cheque=&ciudad=`

    Solo lectura: no toca la base.
    """
    from datetime import datetime as _dt

    from .letras import numero_a_letras

    beneficiario = (request.args.get("beneficiario") or "").strip().upper()
    no_cheque = (request.args.get("no_cheque") or "").strip()
    ciudad = (request.args.get("ciudad") or "QUITO").strip().upper()
    importe = parse_monto(request.args.get("importe"))
    fecha_str = (request.args.get("fecha") or "").strip()
    try:
        fecha = _dt.strptime(fecha_str, "%Y-%m-%d").date() if fecha_str else today_ec()
    except ValueError:
        fecha = today_ec()

    # Si viene una transacción, la usamos como fuente autoritativa.
    id_transaccion = parse_int(request.args.get("id_transaccion"))
    if id_transaccion:
        import contextlib as _ctx

        import db as _db

        with _ctx.suppress(Exception):
            row = _db.fetch_one(
                """
                SELECT t.importe, t.fecha, t.numreferencia, t.concepto, t.prov,
                       COALESCE(p.nombre, '') AS proveedor_nombre
                  FROM scintela.transacciones_bancarias t
                  LEFT JOIN scintela.proveedor p ON p.codigo_prov = t.prov
                 WHERE t.id_transaccion = %s
                """,
                (id_transaccion,),
            )
            if row:
                importe = abs(float(row.get("importe") or 0))
                fecha = row.get("fecha") or fecha
                if row.get("numreferencia"):
                    no_cheque = str(row["numreferencia"])
                if not beneficiario:
                    beneficiario = (row.get("proveedor_nombre") or row.get("concepto") or "").strip().upper()

    # Calibración de posición (mm) para el cheque físico.
    offx = parse_int(request.args.get("offx")) or 0
    offy = parse_int(request.args.get("offy")) or 0

    letras = numero_a_letras(importe or 0)

    return render_template(
        "bancos/cheque_imprimible.html",
        beneficiario=beneficiario,
        no_cheque=no_cheque,
        ciudad=ciudad,
        importe=float(importe or 0),
        fecha=fecha,
        letras=letras,
        offx=offx,
        offy=offy,
    )
